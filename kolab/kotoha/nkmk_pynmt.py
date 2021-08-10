import sys
def add(a, b):
return a + b
if __name__ == '__main__':
print(add(float(sys.argv[1]), float(sys.argv[2])))
import argparse
import add_module
parser = argparse.ArgumentParser()
parser.add_argument('a', type=float)
parser.add_argument('b', type=float)
args = parser.parse_args()
print(add_module.add(args.a, args.b))
import sys
import add_module
print(add_module.add(float(sys.argv[1]), float(sys.argv[2])))
import add_module
print(add_module.add(100, 200))
l = [0, 1, 2, 3, 4]
print([i > 2 for i in l])
print(all([i > 2 for i in l]))
print(any([i > 2 for i in l]))
print(type([i > 2 for i in l]))
print(type((i > 2 for i in l)))
print(type(i > 2 for i in l))
print(all(i > 2 for i in l))
print(any(i > 2 for i in l))
print(sum(i > 2 for i in l))
print(sum(not (i > 2) for i in l))
l = list(range(100000))
print(l[:5])
print(l[-5:])
print(len(l))
all([i < 0 for i in l])
std. dev. of
all(i < 0 for i in l)
std. dev. of
all([i >= 0 for i in l])
std. dev. of
all(i >= 0 for i in l)
std. dev. of
all(i < 50000 for i in l)
std. dev. of
any([i >= 0 for i in l])
std. dev. of
any(i >= 0 for i in l)
std. dev. of
any([i < 0 for i in l])
std. dev. of
any(i < 0 for i in l)
std. dev. of
any(i > 50000 for i in l)
std. dev. of
print(all([True, True, True]))
print(all([True, False, True]))
print(all((True, True, True)))
print(all({True, True, True}))
print(all([]))
print(all([100, [0, 1, 2], 'abc']))
print(all([100, [0, 1, 2], 'abc', {}]))
x = -9
print(x)
print(bin(x))
-9
-0b1001
print(bin(x & 0xff))
print(format(x & 0xffff, 'x'))
x = 9
y = 10
print(x & y)
print(bin(x & y))
print(x | y)
print(bin(x | y))
print(x ^ y)
print(bin(x ^ y))
x = 9
-10
-0b1010
x = 9
print(x << 1)
print(bin(x << 1))
print(x >> 1)
print(bin(x >> 1))
x = -9
print(bin(x))
print(bin(x & 0xff))
-0b1001
print(x << 1)
print(bin(x << 1))
print(bin((x << 1) & 0xff))
-18
-0b10010
print(x >> 1)
print(bin(x >> 1))
print(bin((x >> 1) & 0xff))
-5
-0b101
print(True and True)
print(True and False)
print(False and True)
print(False and False)
a = 10
print(0 < a)
print(a < 100)
print(0 < a and a < 100)
print(0 < a < 100)
print(True or True)
print(True or False)
print(False or True)
print(False or False)
print(not True)
print(not False)
print(True or True and False)
print(True or (True and False))
print((True or True) and False)
print(bool(10))
print(bool(0))
print(bool(''))
print(bool('0'))
print(bool('False'))
print(bool([]))
print(bool([False]))
x = 10
y = 0
print(x and y)
print(x or y)
print(not x)
x = 10
y = 100
print(x and y)
print(y and x)
print(x or y)
print(y or x)
x = 0
y = 0.0
print(x and y)
print(y and x)
print(x or y)
print(y or x)
print(bool(x and y))
print(any([True, False, False]))
print(any([False, False, False]))
print(any((True, False, False)))
print(any({True, False, False}))
print(any([]))
print(not any([False, False, False]))
print(not any([True, False, False]))
import argparse
parser = argparse.ArgumentParser()
parser.add_argument('--en', action='store_true')
args = parser.parse_args()
print(args.en)
print(type(args.en))
import argparse
parser = argparse.ArgumentParser()
parser.add_argument('arg_bool', type=bool)
args = parser.parse_args()
print(args.arg_bool)
print(type(args.arg_bool))
import argparse
parser = argparse.ArgumentParser()
parser.add_argument('arg_int', type=int)
args = parser.parse_args()
print(args.arg_int)
print(type(args.arg_int))
import argparse
from distutils.util import strtobool
parser = argparse.ArgumentParser()
parser.add_argument('arg_bool', type=strtobool)
args = parser.parse_args()
print(args.arg_bool)
print(type(args.arg_bool))
def my_sum(*args):
return sum(args)
print(my_sum(1, 2, 3, 4))
print(my_sum(1, 2, 3, 4, 5, 6, 7, 8))
def my_sum2(*args):
print('args: ', args)
print('type: ', type(args))
print('sum : ', sum(args))
my_sum2(1, 2, 3, 4)
def func_args(arg1, arg2, *args):
print('arg1: ', arg1)
print('arg2: ', arg2)
print('args: ', args)
func_args(0, 1, 2, 3, 4)
func_args(0, 1)
args:  ()
def func_args2(arg1, *args, arg2):
print('arg1: ', arg1)
print('arg2: ', arg2)
print('args: ', args)
func_args2(0, 1, 2, 3, 4)
TypeError: func_args2
func_args2(0, 1, 2, 3, arg2=4)
print('arg1: ', arg1)
print('arg2: ', arg2)
func_args_kw_only(100, 200)
TypeError: func_args_kw_only
func_args_kw_only(100, arg2=200)
def func_default(arg1, arg2='default_x', arg3='default_y'):
print(arg1)
print(arg2)
print(arg3)
func_default('a')
func_default('a', 'b')
func_default('a', arg3='c')
def func_default_error(arg2='default_a', arg3='default_b', arg1):
print(arg1)
print(arg2)
SyntaxError: non
-default
def func_default_list(l=[0, 1, 2], v=3):
l.append(v)
print(l)
func_default_list([0, 0, 0], 100)
func_default_list()
func_default_list()
func_default_list()
def func_default_dict(d={'default': 0}, k='new', v=100):
d[k] = v
print(d)
func_default_dict()
func_default_dict(k='new2', v=200)
def func_default_list_none(l=None, v=3):
if l is None:
l = [0, 1, 2]
l.append(v)
print(l)
func_default_list_none()
func_default_list_none()
def func(arg1, arg2, arg3):
print(arg1)
print(arg2)
print(arg3)
d = {'arg1': 'one', 'arg2': 'two', 'arg3': 'three'}
func(**d)
func(**{'arg1': 'one', 'arg2': 'two', 'arg3': 'three'})
func(**{'arg1': 'one', 'arg2': 'two'})
TypeError: func
func(**{'arg1': 'one', 'arg2': 'two', 'arg3': 'three', 'arg4': 'four'})
TypeError: func
def func_default(arg1=1, arg2=2, arg3=3):
print(arg1)
print(arg2)
print(arg3)
func_default(**{'arg1': 'one'})
func_default(**{'arg2': 'two', 'arg3': 'three'})
func_default(**{'arg1': 'one', 'arg4': 'four'})
TypeError: func_default
def func_kwargs(arg1, **kwargs):
print('arg1', arg1)
for k, v in kwargs.items():
print(k, v)
func_kwargs(**{'arg1': 'one', 'arg2': 'two', 'arg3': 'three'})
func_kwargs(**{'arg1': 'one', 'arg2': 'two', 'arg3': 'three', 'arg4': 'four'})
func_kwargs(**{'arg1': 'one', 'arg3': 'three'})
def func(arg1, arg2, arg3):
print(arg1)
print(arg2)
print(arg3)
l = ['one', 'two', 'three']
func(*l)
func(*['one', 'two', 'three'])
t = ('one', 'two', 'three')
func(*t)
func(*('one', 'two', 'three'))
func(*['one', 'two'])
TypeError: func
func(*['one', 'two', 'three', 'four'])
TypeError: func
def func_default(arg1=1, arg2=2, arg3=3):
print(arg1)
print(arg2)
print(arg3)
func_default(*['one', 'two'])
func_default(*['one'])
func_default(*['one', 'two', 'three', 'four'])
TypeError: func_default
def func_args(arg1, *args):
print(arg1)
for arg in args:
print(arg)
func_args(*['one', 'two'])
func_args(*['one', 'two', 'three'])
func_args(*['one', 'two', 'three', 'four'])
a_l = [0, 1, 2]
b_l = [10, 20, 30]
a_t = (0, 1, 2)
b_t = (10, 20, 30)
a_s = 'abc'
b_s = 'xyz'
print(a_l + b_l)
print(a_t + b_t)
print(a_s + b_s)
print(a_l + 3)
TypeError: can
list (not "int") 
print(a_l + [3])
print(a_t + (3))
TypeError: can
tuple (not "int") 
print(a_t + (3, ))
a_l += b_l
print(a_l)
a_t += b_t
print(a_t)
a_s += b_s
print(a_s)
print(b_l * 3)
print(3 * b_l)
print(b_t * 3)
print(3 * b_t)
print(b_s * 3)
print(3 * b_s)
print(b_l * 0.5)
TypeError: can
print(b_l * -1)
b_l *= 3
print(b_l)
b_t *= 3
print(b_t)
b_s *= 3
print(b_s)
a_l = [0, 1, 2]
b_l = [10, 20, 30]
c_l = a_l + b_l * 3
print(c_l)
print(10 + 3)
print(10 - 3)
print(10 * 3)
print(10 / 3)
print(10 // 3)
print(10 % 3)
print(10 ** 3)
print(2 ** 0.5)
print(10 ** -2)
print(0 ** 0)
print(10 / 0)
ZeroDivisionError: integer
division or modulo
print(10 // 0)
ZeroDivisionError: integer
division or modulo
print(10 % 0)
ZeroDivisionError: integer
division or modulo
print(0 ** -1)
raised
a = 10
b = 3
c = a + b
print('a:', a)
print('b:', b)
print('c:', c)
a = 10
b = 3
a += b
print('a:', a)
print('b:', b)
a = 10
b = 3
a %= b
print('a:', a)
print('b:', b)
a = 10
b = 3
a **= b
print('a:', a)
print('b:', b)
print(2 + 3.0)
print(type(2 + 3.0))
print(10 / 2)
print(type(10 / 2))
print(2 ** 3)
print(type(2 ** 3))
print(2.0 ** 3)
print(type(2.0 ** 3))
print(25 ** 0.5)
print(type(25 ** 0.5))
print(0.01 ** -2)
print(type(0.01 ** -2))
print(100 / 10 ** 2 + 2 * 3 - 5)
print(100 / (10 ** 2) + (2 * 3) - 5)
print((100 / 10) ** 2 + 2 * (3 - 5))
import array
arr_int = array.array('i', [0, 1, 2])
print(arr_int)
array('i', [0, 1, 2])
arr_float = array.array('f', [0.0, 0.1, 0.2])
print(arr_float)
array('f', [0.0, 0.10000000149011612, 0.20000000298023224])
arr_int = array.array('i', [0, 0.1, 2])
TypeError: integer
print(arr_int[1])
print(sum(arr_int))
import pprint
import arxiv
import pandas as pd
l = arxiv.query(query='au:"Grisha Perelman"')
print(type(l))
print(len(l))
print(type(l[0]))
feedparser.FeedParserDict
pprint.pprint(l[0], width=200)
math.DG
time.struct_time(tm_year=2002, tm_mon=11, tm_mday=11, tm_hour=16, tm_min=11, tm_sec=49, tm_wday=0, tm_yday=315, tm_isdst=0)
ensemble. Several
given. In
diffeomorphism and scaling
not quickly 
away. We
assertions
search_query=au%3
start=0&max_results=1000&sortBy=relevance&sortOrder=descending
ensemble. Several
given. In
diffeomorphism and scaling
not quickly 
away. We
assertions
math.DG
flow and its
search_query=au%3
start=0&max_results=1000&sortBy=relevance&sortOrder=descending
flow and its
time.struct_time(tm_year=2002, tm_mon=11, tm_mday=11, tm_hour=16, tm_min=11, tm_sec=49, tm_wday=0, tm_yday=315, tm_isdst=0)
print(l[0]['author'])
print(l[0]['title'])
flow and its
print(l[0]['arxiv_url'])
print(l[0]['pdf_url'])
print(l[0]['summary'])
ensemble. Several
given. In
diffeomorphism and scaling
not quickly 
away. We
assertions
pprint.pprint([a['id'] for a in l])
pprint.pprint([[a['id'], a['published']] for a in l])
df = pd.io.json.json_normalize(l)
print(df.shape)
print(df[['title', 'published']])
flow and its
l = arxiv.query(query='cat:cs.AI', max_results=10, sort_by='submittedDate')
pprint.pprint([[a['id'], a['published']] for a in l])
l = arxiv.query
query='cat:cs.AI'
max_results=10
sort_by='submittedDate'
sort_order='ascending'
pprint.pprint([[a['id'], a['published']] for a in l])
l = arxiv.query
query='cat:cs.AI AND submittedDate:[20190101 TO 20190131235959]'
sort_by='submittedDate'
sort_order='ascending'
df = pd.io.json.json_normalize(l)
print(df.shape)
print(df.head()[['id', 'published']])
print(df.tail()[['id', 'published']])
l = arxiv.query
query='cat:cs.AI AND ti:"deep learning" AND submittedDate:[20190101 TO 20190131235959]'
sort_by='submittedDate'
sort_order='ascending'
df = pd.io.json.json_normalize(l)
print(df[['title', 'published']])
networks for the interp
Recognition: In
Area and Volume
DLocRL: A
l = arxiv.query(id_list=['1902.00358v2', '1902.00358', 'math/0211159v1'])
for a in l:
print(a['arxiv_url'])
import arxiv
import time
l = arxiv.query(query='au:"Grisha Perelman"')
arxiv.download(l[0], 'data/temp/')
v1.The_entropy_formula_for_the_Ricci_flow_and_its_geometric_applications.pdf
arxiv.download(l[0], 'data/temp/', lambda x: x.get('id').split('/')[-1])
v1.pdf
for a in l:
arxiv.download(a, 'data/temp/')
time.sleep(10)
import pprint
import feedparser
url = 'http://arxiv.org/rss/cs.CV'
d = feedparser.parse(url)
pprint.pprint(d, depth=1)
time.struct_time(tm_year=2019, tm_mon=7, tm_mday=26, tm_hour=0, tm_min=30, tm_sec=0, tm_wday=4, tm_yday=207, tm_isdst=0)
print(type(d['entries']))
print(len(d['entries']))
print(type(d['entries'][0]))
feedparser.FeedParserDict
pprint.pprint(d['entries'][0], width=100)
href="http://arxiv.org/find/cs/1/au:+Kurmi_V/0/1/0/all/0/1">Vinod
href="http://arxiv.org/find/cs/1/au:+Bajaj_V/0/1/0/all/0/1">Vipul
href="http://arxiv.org/find/cs/1/au:+Subramanian_V/0/1/0/all/0/1">Venkatesh
href="http://arxiv.org/find/cs/1/au:+Namboodiri_V/0/1/0/all/0/1">Vinay
href="http://arxiv.org/find/cs/1/au:+Kurmi_V/0/1/0/all/0/1">Vinod
href="http://arxiv.org/find/cs/1/au:+Bajaj_V/0/1/0/all/0/1">Vipul
href="http://arxiv.org/find/cs/1/au:+Subramanian_V/0/1/0/all/0/1">Venkatesh
href="http://arxiv.org/find/cs/1/au:+Namboodiri_V/0/1/0/all/0/1">Vinay
href="http://arxiv.org/find/cs/1/au:+Kurmi_V/0/1/0/all/0/1">Vinod
href="http://arxiv.org/find/cs/1/au:+Bajaj_V/0/1/0/all/0/1">Vipul
href="http://arxiv.org/find/cs/1/au:+Subramanian_V/0/1/0/all/0/1">Venkatesh
href="http://arxiv.org/find/cs/1/au:+Namboodiri_V/0/1/0/all/0/1">Vinay
datasets. Adversarial
source and target
close. However
gap. This
classifiers or using
discriminator. Specifically
distribution and the
source and target
representations. The
results and thorough
datasets. Adversarial
source and target
close. However
classifiers or using
distribution and the
source and target
representations. The
results and thorough
cs.LG
v1 [cs.LG]
print(d['entries'][0]['link'])
print(d['entries'][0]['title'])
v1 [cs.LG]
import ast
s = '["a", "b", "c"]'
l = ast.literal_eval(s)
print(l)
print(type(l))
s = '["x", 1, True]'
l = ast.literal_eval(s)
print(l)
print(type(l[0]))
print(type(l[1]))
print(type(l[2]))
s = '{"key1": 1, "key2": 2}'
d = ast.literal_eval(s)
print(d)
print(type(d))
s = '{1, 2, 3}'
se = ast.literal_eval(s)
print(se)
print(type(se))
s = '["x", 1 + 10]'
print(eval(s))
print(ast.literal_eval(s))
ValueError: malformed
node or string
a = 100
print(eval('[1, a]'))
a = 100
print(ast.literal_eval('[1, a]'))
ValueError: malformed
node or string
import json
s = '{"key1": [1, 2, 3], "key2": "abc"}'
print(json.loads(s))
print(ast.literal_eval(s))
s = '[True, False, None]'
print(json.loads(s))
JSONDecodeError: Expecting
print(ast.literal_eval(s))
s = '[true, false, null]'
print(json.loads(s))
print(ast.literal_eval(s))
ValueError: malformed
node or string
s = "{'key1': 'abc', 'key2': '''xyz'''}"
print(json.loads(s))
JSONDecodeError: Expecting
print(ast.literal_eval(s))
s = 'a, b, c'
l = s.split(', ')
print(l)
print(type(l))
s = '1-2-3'
l = s.split('-')
print(l)
print(type(l[0]))
l = [int(c) for c in s.split('-')]
print(l)
print(type(l[0]))
import platform
import numpy
import scipy
print('Python:       ', platform.python_version())
print('NumPy:        ', numpy.__version__)
print('SciPy:        ', scipy.__version__)
import sklearn
import numba
import networkx
print('scikit-learn: ', sklearn.__version__)
print('Numba:        ', numba.__version__)
print('NetworkX:     ', networkx.__version__)
except:
pass
import urllib.request
from bs4 import BeautifulSoup
url = 'http://www.yahoo.co.jp/'
ua = 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_3) '
req = urllib.request.Request(url, headers={'User-Agent': ua})
html = urllib.request.urlopen(req)
soup = BeautifulSoup(html, "html.parser")
li_list = soup.find(id='yahooservice').find_all('li')
li_list = soup.find('div', attrs={'id': 'yahooservice'}).find_all('li')
li_list = soup.select('div#yahooservice > ul > li')
for li in li_list:
print(li.find('a').text)
import os
import time
import urllib.error
import urllib.request
from bs4 import BeautifulSoup
url = 'https://news.yahoo.co.jp/list/'
ua = 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_3) '
req = urllib.request.Request(url, headers={'User-Agent': ua})
html = urllib.request.urlopen(req)
soup = BeautifulSoup(html, "html.parser")
url_list = [img.get('data-src') for img in soup.find(class_='list').find_all('img')]
def download_file(url, dst_path):
urllib.request.urlopen(url) 
open(dst_path, 'wb') 
local_file.write(web_file.read())
urllib.error.URLError
print(e)
def download_file_to_dir(url, dst_dir):
download_file(url, os.path.join(dst_dir, os.path.basename(url)))
download_dir = 'data/temp'
sleep_time_sec = 1
for url in url_list:
print(url)
download_file_to_dir(url, download_dir)
time.sleep(sleep_time_sec)
print(type(True))
print(type(False))
print(issubclass(bool, int))
print(True == 1)
print(False == 0)
print(True + True)
print(True * 10)
print(sum([True, False, True]))
l = [0, 1, 2, 3, 4]
print([i > 2 for i in l])
print(sum(i > 2 for i in l))
if 'abc':
print('True!')
print(int(True))
print(int(False))
print(float(True))
print(float(False))
print(complex(True))
print(complex(False))
print(str(True))
print(str(False))
print(type(str(True)))
print(type(str(False)))
print(bool(str(False)))
print(list(True))
print(bool('True'))
print(bool('False'))
print(bool('abc'))
print(bool(''))
print(bool(1))
print(bool(2))
print(bool(1.23))
print(bool(-1))
print(bool(0))
print(bool(0.0))
print(bool([1, 2, 3]))
print(bool((1, 2, 3)))
print(bool({1, 2, 3}))
print(bool({'k1': 1, 'k2':2, 'k3': 3}))
print(bool([]))
print(bool({}))
print(bool(None))
l1 = [1, 2, 3]
l2 = [10, 20, 30]
for i in l1:
for j in l2:
print(i, j)
for i in l1:
for j in l2:
print(i, j)
if i == 2 and j == 20 :
print('BREAK')
break
for i in l1:
for j in l2:
print(i, j)
if i == 2 and j == 20:
print('BREAK')
break
continue
break
for i in l1:
print('Start outer loop')
for j in l2:
print('--', i, j)
if i == 2 and j == 20:
print('-- BREAK inner loop')
break
print('-- Finish inner loop without BREAK')
continue
print('BREAK outer loop')
break
-- 1
-- 1
-- 1
-- Finish
-- 2
-- 2
-- BREAK
l1 = [1, 2, 3]
l2 = [10, 20, 30]
l3 = [100, 200, 300]
for i in l1:
for j in l2:
for k in l3:
print(i, j, k)
if i == 2 and j == 20 and k == 200:
print('BREAK')
break
continue
break
continue
break
l1 = [1, 2, 3]
l2 = [10, 20, 30]
flag = False
for i in l1:
for j in l2:
print(i, j)
if i == 2 and j == 20:
flag = True
print('BREAK')
break
if flag:
break
l1 = [1, 2, 3]
l2 = [10, 20, 30]
l3 = [100, 200, 300]
flag = False
for i in l1:
for j in l2:
for k in l3:
print(i, j, k)
if i == 2 and j == 20 and k == 200:
flag = True
print('BREAK')
break
if flag:
break
if flag:
break
import itertools
l1 = [1, 2, 3]
l2 = [10, 20, 30]
for i, j in itertools.product(l1, l2):
print(i, j)
for i, j in itertools.product(l1, l2):
print(i, j)
if i == 2 and j == 20:
print('BREAK')
break
l1 = [1, 2, 3]
l2 = [10, 20, 30]
l3 = [100, 200, 300]
for i, j, k in itertools.product(l1, l2, l3):
print(i, j, k)
if i == 2 and j == 20 and k == 200:
print('BREAK')
break
for i, j in itertools.product(l1, l2):
x = i * 2 + j * 3
print(i, j, x)
for i in l1:
temp = i * 2
for j in l2:
x = temp + j * 3
print(i, j, x)
import itertools
n = 100
l1 = range(n)
l2 = range(n)
l3 = range(n)
x = n - 1
for i in l1:
for j in l2:
for k in l3:
if i == x and j == x and k == x:
break
continue
break
continue
break
std. dev. of
flag = False
for i in l1:
for j in l2:
for k in l3:
if i == x and j == x and k == x:
flag = True
break
if flag:
break
if flag:
break
std. dev. of
for i, j, k in itertools.product(l1, l2, l3):
if i == x and j == x and k == x:
break
std. dev. of
print(len([0, 1, 2]))
print(abs(-10))
s = 'abc'
print(s)
print(type(s))
print(s.upper())
import builtins
print(len('abc'))
print(builtins.len('abc'))
print(len)
print(builtins.len)
print(builtins.len is len)
print(__builtins__.len('abc'))
print(__builtins__.len is len)
print(__builtins__ is builtins)
import calendar
import datetime
dt = datetime.datetime(2019, 1, 1, 10, 10, 10)
print(dt)
print(calendar.isleap(dt.year))
d = datetime.date(2020, 1, 1)
print(d)
print(calendar.isleap(d.year))
def isleap_datetime(dt):
return calendar.isleap(dt.year)
print(dt)
print(isleap_datetime(dt))
print(d)
print(isleap_datetime(d))
def isleap_datetime2(dt):
return dt.year % 4 == 0 and (dt.year % 100 != 0 or dt.year % 400 == 0)
print(dt)
print(isleap_datetime2(dt))
print(d)
print(isleap_datetime2(d))
import calendar
import datetime
print(calendar.month(2019, 1))
def get_nth_week(day):
return (day - 1) // 7 + 1
def get_nth_dow(year, month, day):
return get_nth_week(day), calendar.weekday(year, month, day)
print(get_nth_dow(2019, 1, 7))
print(get_nth_dow(2019, 1, 20))
def get_nth_dow_datetime(year, month, day):
return get_nth_week(day), datetime.date(year, month, day).weekday()
print(get_nth_dow_datetime(2019, 1, 7))
print(get_nth_dow_datetime(2019, 1, 20))
def get_nth_dow_datetime_dt(dt):
return get_nth_week(dt.day), dt.weekday()
dt = datetime.datetime(2019, 1, 20)
print(get_nth_dow_datetime_dt(dt))
d = datetime.date(2019, 1, 20)
print(get_nth_dow_datetime_dt(d))
print(datetime.date.today())
print(get_nth_dow_datetime_dt(datetime.date.today()))
print(calendar.month(2019, 1))
calendar.setfirstweekday(6)
print(calendar.month(2019, 1))
def get_nth_week2(year, month, day, firstweekday=0):
first_dow = calendar.monthrange(year, month)[0]
offset = (first_dow - firstweekday) % 7
return (day + offset - 1) // 7 + 1
print(get_nth_week2(2019, 1, 20))
print(get_nth_week2(2019, 1, 20, 6))
def get_nth_week2_datetime(year, month, day, firstweekday=0):
first_dow = datetime.date(year, month, 1).weekday()
offset = (first_dow - firstweekday) % 7
return (day + offset - 1) // 7 + 1
print(get_nth_week2_datetime(2019, 1, 20))
print(get_nth_week2_datetime(2019, 1, 20, 6))
def get_nth_week2_datetime_dt(dt, firstweekday=0):
first_dow = dt.replace(day=1).weekday()
offset = (first_dow - firstweekday) % 7
return (dt.day + offset - 1) // 7 + 1
dt = datetime.datetime(2019, 1, 20)
print(get_nth_week2_datetime_dt(dt))
print(get_nth_week2_datetime_dt(dt, 6))
d = datetime.date(2019, 1, 20)
print(get_nth_week2_datetime_dt(d))
print(get_nth_week2_datetime_dt(d, 6))
print(datetime.date.today())
print(get_nth_week2_datetime_dt(datetime.date.today()))
print(get_nth_week2_datetime_dt(datetime.date.today(), 6))
import calendar
import datetime
def get_day_of_last_week(year, month, dow):
dow: Monday
- Sunday(6)
n = calendar.monthrange(year, month)[1]
l = range(n - 6, n + 1)
w = calendar.weekday(year, month, l[0])
w_l = [i % 7 for i in range(w, w + 7)]
return l[w_l.index(dow)]
print(calendar.month(2019, 1))
print(get_day_of_last_week(2019, 1, 0))
print(get_day_of_last_week(2019, 1, 3))
print(get_day_of_last_week(2019, 1, 4))
year = 2019
month = 1
dow = 0
n = calendar.monthrange(year, month)[1]
print(n)
l = range(n - 6, n + 1)
print(list(l))
w = calendar.weekday(year, month, l[0])
print(w)
w_l = [i % 7 for i in range(w, w + 7)]
print(w_l)
print(l[w_l.index(dow)])
def get_date_of_last_week(dt, dow):
dow: Monday
- Sunday(6)
return dt.replace(day=get_day_of_last_week(dt.year, dt.month, dow))
dt = datetime.datetime(2019, 1, 10, 10, 10, 10)
print(dt)
print(get_date_of_last_week(dt, 0))
print(get_date_of_last_week(dt, 3))
d = datetime.date(2019, 1, 10)
print(d)
print(get_date_of_last_week(d, 0))
print(get_date_of_last_week(d, 3))
def get_date_of_last_week2(year, month, dow):
dow: Monday
- Sunday(6)
return datetime.date(year, month, get_day_of_last_week(year, month, dow))
print(get_date_of_last_week2(2019, 1, 0))
import calendar
import datetime
def get_day_of_nth_dow(year, month, nth, dow):
dow: Monday
- Sunday(6)
if nth < 1 or dow < 0 or dow > 6:
return None
first_dow, n = calendar.monthrange(year, month)
day = 7 * (nth - 1) + (dow - first_dow) % 7 + 1
return day if day <= n else None
print(calendar.month(2019, 1))
print(get_day_of_nth_dow(2019, 1, 1, 1))  
Tuesday(1)
print(get_day_of_nth_dow(2019, 1, 2, 0))  
Monday(0)
print(get_day_of_nth_dow(2019, 1, 3, 6))  
Sunday(6)
print(get_day_of_nth_dow(2019, 1, 5, 3))  
Thursday(3)
print(get_day_of_nth_dow(2019, 1, 5, 4))
print(get_day_of_nth_dow(2019, 1, 0, 4))
print(get_day_of_nth_dow(2019, 1, 1, 10))
print(get_day_of_nth_dow(2019, 1, 2, 1.8))
def get_date_of_nth_dow(year, month, nth, dow):
day = get_day_of_nth_dow(year, month, nth, dow)
return datetime.date(year, month, day) if day else None
print(get_date_of_nth_dow(2019, 1, 1, 1))
print(get_date_of_nth_dow(2019, 1, 1, 10))
print(get_date_of_nth_dow(2019, 1, 2, 1.8))
TypeError: integer
print([(m, get_day_of_nth_dow(2019, m, 2, 0)) for m in range(1, 13)])
for y in range(2020, 2030):
print(get_date_of_nth_dow(y, 1, 2, 0))
import calendar
print(calendar.monthrange(2019, 1))
print(calendar.monthrange(2019, 1)[1])
print(calendar.monthrange(2019, 2)[1])
print(calendar.monthrange(2020, 2)[1])
print(calendar.monthlen(2019, 1))
print(calendar.__all__)
print('monthlen' in calendar.__all__)
print('monthrange' in calendar.__all__)
import calendar
hc = calendar.HTMLCalendar()
print(hc.formatmonth(2019, 1, withyear=False))
border="0"
cellpadding="0"
cellspacing="0"
class="month"
colspan="7"
class="month">January
class="mon">Mon
class="tue">Tue
class="wed">Wed
class="thu">Thu
class="fri">Fri
class="sat">Sat
class="sun">Sun
class="noday"
class="tue">1
class="wed">2
class="thu">3
class="fri">4
class="sat">5
class="sun">6
class="mon">7
class="tue">8
class="wed">9
class="thu">10
class="fri">11
class="sat">12
class="sun">13
class="mon">14
class="tue">15
class="wed">16
class="thu">17
class="fri">18
class="sat">19
class="sun">20
class="mon">21
class="tue">22
class="wed">23
class="thu">24
class="fri">25
class="sat">26
class="sun">27
class="mon">28
class="tue">29
class="wed">30
class="thu">31
class="noday"
class="noday"
class="noday"
print(type(hc.formatmonth(2019, 1)))
print(hc.formatyear(2019, width=4))
border="0"
cellpadding="0"
cellspacing="0"
class="year"
colspan="4"
class="year">2019
border="0"
cellpadding="0"
cellspacing="0"
class="month"
colspan="7"
class="month">January
class="mon">Mon
class="tue">Tue
class="wed">Wed
class="thu">Thu
class="fri">Fri
class="sat">Sat
class="sun">Sun
class="noday"
class="tue">1
class="wed">2
class="thu">3
class="fri">4
class="sat">5
class="sun">6
class="mon">7
class="tue">8
class="wed">9
class="thu">10
class="fri">11
class="sat">12
class="sun">13
class="mon">14
class="tue">15
class="wed">16
class="thu">17
class="fri">18
class="sat">19
class="sun">20
class="mon">21
class="tue">22
class="wed">23
class="thu">24
class="fri">25
class="sat">26
class="sun">27
class="mon">28
class="tue">29
class="wed">30
class="thu">31
class="noday"
class="noday"
class="noday"
border="0"
cellpadding="0"
cellspacing="0"
class="month"
colspan="7"
class="month">February
class="mon">Mon
class="tue">Tue
class="wed">Wed
class="thu">Thu
class="fri">Fri
class="sat">Sat
class="sun">Sun
class="noday"
class="noday"
class="noday"
class="noday"
class="fri">1
class="sat">2
class="sun">3
class="mon">4
class="tue">5
class="wed">6
class="thu">7
class="fri">8
class="sat">9
class="sun">10
class="mon">11
class="tue">12
class="wed">13
class="thu">14
class="fri">15
class="sat">16
class="sun">17
class="mon">18
class="tue">19
class="wed">20
class="thu">21
class="fri">22
class="sat">23
class="sun">24
class="mon">25
class="tue">26
class="wed">27
class="thu">28
class="noday"
class="noday"
class="noday"
border="0"
cellpadding="0"
cellspacing="0"
class="month"
colspan="7"
class="month">March
class="mon">Mon
class="tue">Tue
class="wed">Wed
class="thu">Thu
class="fri">Fri
class="sat">Sat
class="sun">Sun
class="noday"
class="noday"
class="noday"
class="noday"
class="fri">1
class="sat">2
class="sun">3
class="mon">4
class="tue">5
class="wed">6
class="thu">7
class="fri">8
class="sat">9
class="sun">10
class="mon">11
class="tue">12
class="wed">13
class="thu">14
class="fri">15
class="sat">16
class="sun">17
class="mon">18
class="tue">19
class="wed">20
class="thu">21
class="fri">22
class="sat">23
class="sun">24
class="mon">25
class="tue">26
class="wed">27
class="thu">28
class="fri">29
class="sat">30
class="sun">31
border="0"
cellpadding="0"
cellspacing="0"
class="month"
colspan="7"
class="month">April
class="mon">Mon
class="tue">Tue
class="wed">Wed
class="thu">Thu
class="fri">Fri
class="sat">Sat
class="sun">Sun
class="mon">1
class="tue">2
class="wed">3
class="thu">4
class="fri">5
class="sat">6
class="sun">7
class="mon">8
class="tue">9
class="wed">10
class="thu">11
class="fri">12
class="sat">13
class="sun">14
class="mon">15
class="tue">16
class="wed">17
class="thu">18
class="fri">19
class="sat">20
class="sun">21
class="mon">22
class="tue">23
class="wed">24
class="thu">25
class="fri">26
class="sat">27
class="sun">28
class="mon">29
class="tue">30
class="noday"
class="noday"
class="noday"
class="noday"
class="noday"
border="0"
cellpadding="0"
cellspacing="0"
class="month"
colspan="7"
class="month">May
class="mon">Mon
class="tue">Tue
class="wed">Wed
class="thu">Thu
class="fri">Fri
class="sat">Sat
class="sun">Sun
class="noday"
class="noday"
class="wed">1
class="thu">2
class="fri">3
class="sat">4
class="sun">5
class="mon">6
class="tue">7
class="wed">8
class="thu">9
class="fri">10
class="sat">11
class="sun">12
class="mon">13
class="tue">14
class="wed">15
class="thu">16
class="fri">17
class="sat">18
class="sun">19
class="mon">20
class="tue">21
class="wed">22
class="thu">23
class="fri">24
class="sat">25
class="sun">26
class="mon">27
class="tue">28
class="wed">29
class="thu">30
class="fri">31
class="noday"
class="noday"
border="0"
cellpadding="0"
cellspacing="0"
class="month"
colspan="7"
class="month">June
class="mon">Mon
class="tue">Tue
class="wed">Wed
class="thu">Thu
class="fri">Fri
class="sat">Sat
class="sun">Sun
class="noday"
class="noday"
class="noday"
class="noday"
class="noday"
class="sat">1
class="sun">2
class="mon">3
class="tue">4
class="wed">5
class="thu">6
class="fri">7
class="sat">8
class="sun">9
class="mon">10
class="tue">11
class="wed">12
class="thu">13
class="fri">14
class="sat">15
class="sun">16
class="mon">17
class="tue">18
class="wed">19
class="thu">20
class="fri">21
class="sat">22
class="sun">23
class="mon">24
class="tue">25
class="wed">26
class="thu">27
class="fri">28
class="sat">29
class="sun">30
border="0"
cellpadding="0"
cellspacing="0"
class="month"
colspan="7"
class="month">July
class="mon">Mon
class="tue">Tue
class="wed">Wed
class="thu">Thu
class="fri">Fri
class="sat">Sat
class="sun">Sun
class="mon">1
class="tue">2
class="wed">3
class="thu">4
class="fri">5
class="sat">6
class="sun">7
class="mon">8
class="tue">9
class="wed">10
class="thu">11
class="fri">12
class="sat">13
class="sun">14
class="mon">15
class="tue">16
class="wed">17
class="thu">18
class="fri">19
class="sat">20
class="sun">21
class="mon">22
class="tue">23
class="wed">24
class="thu">25
class="fri">26
class="sat">27
class="sun">28
class="mon">29
class="tue">30
class="wed">31
class="noday"
class="noday"
class="noday"
class="noday"
border="0"
cellpadding="0"
cellspacing="0"
class="month"
colspan="7"
class="month">August
class="mon">Mon
class="tue">Tue
class="wed">Wed
class="thu">Thu
class="fri">Fri
class="sat">Sat
class="sun">Sun
class="noday"
class="noday"
class="noday"
class="thu">1
class="fri">2
class="sat">3
class="sun">4
class="mon">5
class="tue">6
class="wed">7
class="thu">8
class="fri">9
class="sat">10
class="sun">11
class="mon">12
class="tue">13
class="wed">14
class="thu">15
class="fri">16
class="sat">17
class="sun">18
class="mon">19
class="tue">20
class="wed">21
class="thu">22
class="fri">23
class="sat">24
class="sun">25
class="mon">26
class="tue">27
class="wed">28
class="thu">29
class="fri">30
class="sat">31
class="noday"
border="0"
cellpadding="0"
cellspacing="0"
class="month"
colspan="7"
class="month">September
class="mon">Mon
class="tue">Tue
class="wed">Wed
class="thu">Thu
class="fri">Fri
class="sat">Sat
class="sun">Sun
class="noday"
class="noday"
class="noday"
class="noday"
class="noday"
class="noday"
class="sun">1
class="mon">2
class="tue">3
class="wed">4
class="thu">5
class="fri">6
class="sat">7
class="sun">8
class="mon">9
class="tue">10
class="wed">11
class="thu">12
class="fri">13
class="sat">14
class="sun">15
class="mon">16
class="tue">17
class="wed">18
class="thu">19
class="fri">20
class="sat">21
class="sun">22
class="mon">23
class="tue">24
class="wed">25
class="thu">26
class="fri">27
class="sat">28
class="sun">29
class="mon">30
class="noday"
class="noday"
class="noday"
class="noday"
class="noday"
class="noday"
border="0"
cellpadding="0"
cellspacing="0"
class="month"
colspan="7"
class="month">October
class="mon">Mon
class="tue">Tue
class="wed">Wed
class="thu">Thu
class="fri">Fri
class="sat">Sat
class="sun">Sun
class="noday"
class="tue">1
class="wed">2
class="thu">3
class="fri">4
class="sat">5
class="sun">6
class="mon">7
class="tue">8
class="wed">9
class="thu">10
class="fri">11
class="sat">12
class="sun">13
class="mon">14
class="tue">15
class="wed">16
class="thu">17
class="fri">18
class="sat">19
class="sun">20
class="mon">21
class="tue">22
class="wed">23
class="thu">24
class="fri">25
class="sat">26
class="sun">27
class="mon">28
class="tue">29
class="wed">30
class="thu">31
class="noday"
class="noday"
class="noday"
border="0"
cellpadding="0"
cellspacing="0"
class="month"
colspan="7"
class="month">November
class="mon">Mon
class="tue">Tue
class="wed">Wed
class="thu">Thu
class="fri">Fri
class="sat">Sat
class="sun">Sun
class="noday"
class="noday"
class="noday"
class="noday"
class="fri">1
class="sat">2
class="sun">3
class="mon">4
class="tue">5
class="wed">6
class="thu">7
class="fri">8
class="sat">9
class="sun">10
class="mon">11
class="tue">12
class="wed">13
class="thu">14
class="fri">15
class="sat">16
class="sun">17
class="mon">18
class="tue">19
class="wed">20
class="thu">21
class="fri">22
class="sat">23
class="sun">24
class="mon">25
class="tue">26
class="wed">27
class="thu">28
class="fri">29
class="sat">30
class="noday"
border="0"
cellpadding="0"
cellspacing="0"
class="month"
colspan="7"
class="month">December
class="mon">Mon
class="tue">Tue
class="wed">Wed
class="thu">Thu
class="fri">Fri
class="sat">Sat
class="sun">Sun
class="noday"
class="noday"
class="noday"
class="noday"
class="noday"
class="noday"
class="sun">1
class="mon">2
class="tue">3
class="wed">4
class="thu">5
class="fri">6
class="sat">7
class="sun">8
class="mon">9
class="tue">10
class="wed">11
class="thu">12
class="fri">13
class="sat">14
class="sun">15
class="mon">16
class="tue">17
class="wed">18
class="thu">19
class="fri">20
class="sat">21
class="sun">22
class="mon">23
class="tue">24
class="wed">25
class="thu">26
class="fri">27
class="sat">28
class="sun">29
class="mon">30
class="tue">31
class="noday"
class="noday"
class="noday"
class="noday"
class="noday"
print(hc.cssclasses)
hc.cssclasses = ['mon', 'tue', 'wed', 'thu', 'fri', 'sat blue', 'sun red']
print(hc.cssclass_month)
print(hc.cssclass_year)
print(hc.cssclass_noday)
hc_sun = calendar.HTMLCalendar(firstweekday=6)
print(hc_sun.formatmonth(2019, 1))
border="0"
cellpadding="0"
cellspacing="0"
class="month"
colspan="7"
class="month">January
class="sun">Sun
class="mon">Mon
class="tue">Tue
class="wed">Wed
class="thu">Thu
class="fri">Fri
class="sat">Sat
class="noday"
class="noday"
class="tue">1
class="wed">2
class="thu">3
class="fri">4
class="sat">5
class="sun">6
class="mon">7
class="tue">8
class="wed">9
class="thu">10
class="fri">11
class="sat">12
class="sun">13
class="mon">14
class="tue">15
class="wed">16
class="thu">17
class="fri">18
class="sat">19
class="sun">20
class="mon">21
class="tue">22
class="wed">23
class="thu">24
class="fri">25
class="sat">26
class="sun">27
class="mon">28
class="tue">29
class="wed">30
class="thu">31
class="noday"
class="noday"
lhc = calendar.LocaleHTMLCalendar(firstweekday=6, locale='ja_jp')
print(lhc.formatmonth(2019, 1))
border="0"
cellpadding="0"
cellspacing="0"
class="month"
colspan="7"
class="month">1
class="sun"
class="mon"
class="tue"
class="wed"
class="thu"
class="fri"
class="sat"
class="noday"
class="noday"
class="tue">1
class="wed">2
class="thu">3
class="fri">4
class="sat">5
class="sun">6
class="mon">7
class="tue">8
class="wed">9
class="thu">10
class="fri">11
class="sat">12
class="sun">13
class="mon">14
class="tue">15
class="wed">16
class="thu">17
class="fri">18
class="sat">19
class="sun">20
class="mon">21
class="tue">22
class="wed">23
class="thu">24
class="fri">25
class="sat">26
class="sun">27
class="mon">28
class="tue">29
class="wed">30
class="thu">31
class="noday"
class="noday"
import calendar
print(calendar.isleap(2019))
print(calendar.isleap(2020))
print(calendar.isleap(1900))
print(calendar.isleap(2000))
print(calendar.leapdays(2019, 2030))
print(calendar.leapdays(2019, 2020))
print([y for y in range(2019, 2030) if calendar.isleap(y)])
print([y for y in range(2000, 2020) if calendar.isleap(y)])
import calendar
import pprint
pprint.pprint(calendar.monthcalendar(2019, 1))
calendar.setfirstweekday(6)
pprint.pprint(calendar.monthcalendar(2019, 1))
c = calendar.Calendar(firstweekday=0)
pprint.pprint(c.monthdayscalendar(2019, 1))
pprint.pprint(c.yeardayscalendar(2019), depth=2)
pprint.pprint(c.yeardayscalendar(2019, width=4), depth=2)
pprint.pprint(c.monthdays2calendar(2019, 1))
pprint.pprint(c.monthdatescalendar(2019, 1))
datetime.date(2018, 12, 31)
datetime.date(2019, 1, 1)
datetime.date(2019, 1, 2)
datetime.date(2019, 1, 3)
datetime.date(2019, 1, 4)
datetime.date(2019, 1, 5)
datetime.date(2019, 1, 6)
datetime.date(2019, 1, 7)
datetime.date(2019, 1, 8)
datetime.date(2019, 1, 9)
datetime.date(2019, 1, 10)
datetime.date(2019, 1, 11)
datetime.date(2019, 1, 12)
datetime.date(2019, 1, 13)
datetime.date(2019, 1, 14)
datetime.date(2019, 1, 15)
datetime.date(2019, 1, 16)
datetime.date(2019, 1, 17)
datetime.date(2019, 1, 18)
datetime.date(2019, 1, 19)
datetime.date(2019, 1, 20)
datetime.date(2019, 1, 21)
datetime.date(2019, 1, 22)
datetime.date(2019, 1, 23)
datetime.date(2019, 1, 24)
datetime.date(2019, 1, 25)
datetime.date(2019, 1, 26)
datetime.date(2019, 1, 27)
datetime.date(2019, 1, 28)
datetime.date(2019, 1, 29)
datetime.date(2019, 1, 30)
datetime.date(2019, 1, 31)
datetime.date(2019, 2, 1)
datetime.date(2019, 2, 2)
datetime.date(2019, 2, 3)
import calendar
print(calendar.calendar(2017))
print(calendar.month(2100, 1))
l = calendar.monthcalendar(2100, 1)
print(type(l))
print(l)
print(calendar.weekday(2001, 9, 11)) 
import calendar
print(calendar.month(2019, 1))
print(type(calendar.month(2019, 1)))
print(calendar.month(2019, 1, w=3, l=2))
calendar.prmonth(2019, 1)
print(calendar.calendar(2019))
print(type(calendar.calendar(2019)))
print(calendar.calendar(2019, c=3, m=4))
calendar.prcal(2019)
calendar.setfirstweekday(calendar.SUNDAY)
print(calendar.month(2019, 1))
print(calendar.MONDAY)
print(calendar.SUNDAY)
calendar.setfirstweekday(0)
print(calendar.month(2019, 1))
print(calendar.firstweekday())
ltc_de = calendar.LocaleTextCalendar(locale='de_de')
print(ltc_de.formatmonth(2019, 1))
ltc_ja = calendar.LocaleTextCalendar(locale='ja_jp')
print(ltc_ja.formatmonth(2019, 1))
import calendar
import pprint
print(calendar.month(2019, 3))
pprint.pprint(calendar.monthcalendar(2019, 3))
print(len(calendar.monthcalendar(2019, 3)))
calendar.setfirstweekday(calendar.SUNDAY)
print(calendar.month(2019, 3))
pprint.pprint(calendar.monthcalendar(2019, 3))
print(len(calendar.monthcalendar(2019, 3)))
print(calendar.MONDAY)
print(calendar.SUNDAY)
calendar.setfirstweekday(0)
print(calendar.firstweekday())
x = 15
print(10 < x < 20)
print(10 < x and x < 20)
x = 0
print(10 < x < 20)
print(10 < x and x < 20)
x = 15
y = 25
print(10 < x < 20 < y < 30)
print(10 < x and x < 20 and 20 < y and y < 30)
x = 15
y = 40
print(10 < x < 20 < y < 30)
print(10 < x and x < 20 and 20 < y and y < 30)
def test(x):
print('function is called')
return(x)
print(test(15))
print(10 < test(15) < 20)
print(10 < test(15) and test(15) < 20)
print(10 < test(0) < 20)
print(10 < test(0) and test(0) < 20)
x = 15
if 10 < x < 20:
print('result: 10 < x < 20')
print('result: x <= 10 or 20 <= x')
x = 30
if 10 < x < 20:
print('result: 10 < x < 20')
print('result: x <= 10 or 20 <= x')
result: x
10 or 20 <= x
a = 10
b = 10
c = 10
if a == b == c:
print('all equal')
print('not all equal')
a = 10
b = 1
c = 10
if a == b == c:
print('all equal')
print('not all equal')
not all 
a = 10
b = 1
c = 100
print(a != b != c)
a = 10
b = 10
c = 1
print(a != b != c)
a = 10
b = 1
c = 10
print(a != b != c)
a = 100
l = [0, 10, 100, 1000]
print(50 < a in l)
print(50 < a and a in l)
i = 100
f = 1.23
print(type(i))
print(type(f))
print(isinstance(i, int))
print(isinstance(i, float))
print(isinstance(f, int))
print(isinstance(f, float))
f_i = 100.0
print(type(f_i))
print(isinstance(f_i, int))
print(isinstance(f_i, float))
f = 1.23
print(f.is_integer())
f_i = 100.0
print(f_i.is_integer())
def is_integer_num(n):
if isinstance(n, int):
return True
if isinstance(n, float):
return n.is_integer()
return False
print(is_integer_num(100))
print(is_integer_num(1.23))
print(is_integer_num(100.0))
print(is_integer_num('100'))
def is_integer(n):
float(n)
except ValueError:
return False
return float(n).is_integer()
print(is_integer(100))
print(is_integer(100.0))
print(is_integer(1.23))
print(is_integer('100'))
print(is_integer('100.0'))
print(is_integer('1.23'))
print(is_integer('string'))
i = ord('A')
print(i)
print(type(i))
ord('abc')
TypeError: ord
s = hex(i)
print(s)
print(type(s))
print(format(i, '04x'))
print(format(i, '#06x'))
print(format(ord('X'), '#08x'))
print(format(ord('💯'), '#08x'))
ord('🇯🇵')
TypeError: ord
print(len('🇯🇵'))
print(chr(65))
print(type(chr(65)))
print(65 == 0x41)
print(chr(0x41))
print(chr(0x000041))
s = '0x0041'
print(int(s, 16))
print(chr(int(s, 16)))
s = 'U+0041'
print(s[2:])
print(chr(int(s[2:], 16)))
print('\x41')
print('\u0041')
print('\U00000041')
print('\U0001f4af')
print('\U0000041')
print('\u0041\u0042\u0043')
print(len('\u0041\u0042\u0043'))
import collections
l = ['a', 'a', 'a', 'a', 'b', 'c', 'c']
c = collections.Counter(l)
print(c)
Counter({'a': 4, 'c': 2, 'b': 1})
print(type(c))
collections.Counter
print(issubclass(type(c), dict))
print(c['a'])
print(c['b'])
print(c['c'])
print(c['d'])
print(c.keys())
dict_keys(['a', 'b', 'c'])
print(c.values())
dict_values([4, 1, 2])
print(c.items())
dict_items([('a', 4), ('b', 1), ('c', 2)])
print(c.most_common())
print(c.most_common()[0])
print(c.most_common()[-1])
print(c.most_common()[0][0])
print(c.most_common()[0][1])
print(c.most_common()[::-1])
print(c.most_common(2))
values, counts = zip(*c.most_common())
print(values)
print(counts)
l = ['a', 'a', 'a', 'a', 'b', 'c', 'c']
c = collections.Counter(l)
print(len(c))
print(set(l))
print(len(set(l)))
l = ['a', 'a', 'a', 'a', 'b', 'c', 'c']
c = collections.Counter(l)
print(c.items())
dict_items([('a', 4), ('b', 1), ('c', 2)])
print([i for i in l if c[i] >= 2])
print([i[1] for i in c.items() if i[1] >= 2])
print(sum(i[1] for i in c.items() if i[1] >= 2))
print([i[0] for i in c.items() if i[1] >= 2])
print([i[1] >= 2 for i in c.items()])
print(sum(i[1] >= 2 for i in c.items()))
s = 'government of the people, by the people, for the people.'
s_remove = s.replace(',', '').replace('.', '')
print(s_remove)
word_list = s_remove.split()
print(word_list)
print(word_list.count('people'))
print(len(set(word_list)))
c = collections.Counter(word_list)
print(c)
Counter({'the': 3, 'people': 3, 'government': 1, 'of': 1, 'by': 1, 'for': 1})
print(c.most_common()[0][0])
s = 'supercalifragilisticexpialidocious'
print(s.count('p'))
c = collections.Counter(s)
print(c)
Counter({'i': 7, 's': 3, 'c': 3, 'a': 3, 'l': 3, 'u': 2, 'p': 2, 'e': 2, 'r': 2, 'o': 2, 'f': 1, 'g': 1, 't': 1, 'x': 1, 'd': 1})
print(c.most_common(5))
values, counts = zip(*c.most_common(5))
print(values)
import collections
l_100 = list(range(100))
collections.Counter(l_100)
std. dev. of
l_10000 = list(range(10000))
collections.Counter(l_10000)
std. dev. of
l_10000_2 = list(range(100)) * 100
print(len(l_10000_2))
collections.Counter(l_10000_2)
std. dev. of
from collections import deque
d = deque()
print(d)
deque([])
print(type(d))
collections.deque
d = deque(['m', 'n'])
print(d)
deque(['m', 'n'])
d.append('o')
print(d)
deque(['m', 'n', 'o'])
d.appendleft('l')
print(d)
deque(['l', 'm', 'n', 'o'])
d.extend(['p', 'q'])
print(d)
deque(['l', 'm', 'n', 'o', 'p', 'q'])
d.extendleft(['k', 'j'])
print(d)
deque(['j', 'k', 'l', 'm', 'n', 'o', 'p', 'q'])
d.insert(3, 'XXX')
print(d)
deque(['j', 'k', 'l', 'XXX', 'm', 'n', 'o', 'p', 'q'])
d.insert(-1, 'YYY')
print(d)
deque(['j', 'k', 'l', 'XXX', 'm', 'n', 'o', 'p', 'YYY', 'q'])
d.insert(100, 'ZZZ')
print(d)
deque(['j', 'k', 'l', 'XXX', 'm', 'n', 'o', 'p', 'YYY', 'q', 'ZZZ'])
d.insert(-100, 'XYZ')
print(d)
deque(['XYZ', 'j', 'k', 'l', 'XXX', 'm', 'n', 'o', 'p', 'YYY', 'q', 'ZZZ'])
d = deque(['a', 'b', 'c', 'b', 'd'])
print(d)
deque(['a', 'b', 'c', 'b', 'd'])
print(d.pop())
print(d)
deque(['a', 'b', 'c', 'b'])
print(d.popleft())
print(d)
deque(['b', 'c', 'b'])
d.remove('b')
print(d)
deque(['c', 'b'])
d.remove('X')
ValueError: deque
remove(x)
d.clear()
print(d)
deque([])
d.pop()
IndexError: pop
d.popleft()
IndexError: pop
d.clear()
print(d)
deque([])
d = deque(['a', 'b', 'c', 'd', 'e'])
print(d)
deque(['a', 'b', 'c', 'd', 'e'])
d.rotate()
print(d)
deque(['e', 'a', 'b', 'c', 'd'])
d = deque(['a', 'b', 'c', 'd', 'e'])
d.rotate(2)
print(d)
deque(['d', 'e', 'a', 'b', 'c'])
d = deque(['a', 'b', 'c', 'd', 'e'])
d.rotate(-1)
print(d)
deque(['b', 'c', 'd', 'e', 'a'])
d = deque(['a', 'b', 'c', 'd', 'e'])
d.rotate(6)
print(d)
deque(['e', 'a', 'b', 'c', 'd'])
d = deque(['a', 'b', 'c', 'c', 'd'])
print(d[0])
print(d[-1])
d[2] = 'X'
print(d)
deque(['a', 'b', 'X', 'c', 'd'])
print(d[2:4])
TypeError: sequence
not 'slice'
print(d.index('c'))
print(d.index('x'))
d = deque(['a', 'a', 'b', 'c'])
print(len(d))
print(d.count('a'))
print(d.count('x'))
print('b' in d)
print('x' in d)
d = deque(['a', 'b', 'c', 'd', 'e'])
d.reverse()
print(d)
deque(['e', 'd', 'c', 'b', 'a'])
d = deque(['a', 'b', 'c', 'd', 'e'])
print(deque(reversed(d)))
deque(['e', 'd', 'c', 'b', 'a'])
d = deque(['a', 'b', 'c'])
for v in d:
print(v)
d = deque(['a', 'b', 'c'])
l = list(d)
print(l)
print(type(l))
from collections import deque
d = deque(['l', 'm', 'n'], 3)
print(d)
deque(['l', 'm', 'n'], maxlen=3)
d.append('o')
print(d)
deque(['m', 'n', 'o'], maxlen=3)
d.appendleft('l')
print(d)
deque(['l', 'm', 'n'], maxlen=3)
d.extend(['o', 'p'])
print(d)
deque(['n', 'o', 'p'], maxlen=3)
d.extendleft(['m', 'l'])
print(d)
deque(['l', 'm', 'n'], maxlen=3)
d.insert(0, 'XXX')
IndexError: deque
print(d.pop())
print(d)
deque(['l', 'm'], maxlen=3)
d.insert(1, 'XXX')
print(d)
deque(['l', 'XXX', 'm'], maxlen=3)
print(d.maxlen)
d.maxlen = 5
AttributeError: attribute
collections.deque
from collections import deque
d = deque(['a', 'b', 'c'])
print(d)
deque(['a', 'b', 'c'])
d.append('d')
print(d)
deque(['a', 'b', 'c', 'd'])
print(d.popleft())
print(d)
deque(['b', 'c', 'd'])
from collections import deque
d = deque(['a', 'b', 'c'])
print(d)
deque(['a', 'b', 'c'])
d.append('d')
print(d)
deque(['a', 'b', 'c', 'd'])
print(d.pop())
print(d)
deque(['a', 'b', 'c'])
from operator import mul
from functools import reduce
def combinations_count(n, r):
r = min(r, n - r)
numer = reduce(mul, range(n, n - r, -1), 1)
denom = reduce(mul, range(1, r + 1), 1)
return numer // denom
print(combinations_count(4, 2))
print(combinations_count(4, 0))
a = 1
a = 1
b = 2
a = 1
a = 1
b = 2
c = 3
d = 4
e = 5
a = 1
a = 1
a = 1
xyz = 100
a = 1
a = 1
a = 1
def test(a, b):
print(a)
print(b)
a = 1
b = 2
c = 3
d = 4
e = 5
def test(a, b):
print(a)
print(b)
def test(a, b):
print(a)
print(b)
IndentationError: unexpected
def test(a, b):
print(a)
print(b)
def test(a, b):
print(a)
print(b)
c = 3 + 4
print(c)
print(type(c))
c = 3 + j
NameError: name
not defined
c = 3 + 1
print(c)
c = 3
print(c)
c = 3 + 0
print(c)
c = 1.2e3 + 3
print(c)
c = complex(3, 4)
print(c)
print(type(c))
c = 3 + 4
print(c.real)
print(type(c.real))
print(c.imag)
print(type(c.imag))
c.real = 5.5
AttributeError: readonly
c = 3 + 4
print(c.conjugate())
c = 3 + 4
print(abs(c))
c = 1 + 1
print(abs(c))
c1 = 3 + 4
c2 = 2 - 1
print(c1 + c2)
print(c1 - c2)
print(c1 * c2)
print(c1 / c2)
print(c1 ** 3)
print(c1 + 3)
print(c1 * 0.5)
import cmath
import math
c = 1 + 1
print(math.atan2(c.imag, c.real))
print(cmath.phase(c))
print(cmath.phase(c) == math.atan2(c.imag, c.real))
print(math.degrees(cmath.phase(c)))
c = 1 + 1
print(cmath.polar(c))
print(type(cmath.polar(c)))
print(cmath.polar(c)[0] == abs(c))
print(cmath.polar(c)[1] == cmath.phase(c))
print(cmath.rect(1, 1))
print(cmath.rect(1, 0))
print(cmath.rect(cmath.polar(c)[0], cmath.polar(c)[1]))
r = 2
ph = math.pi
print(cmath.rect(r, ph).real == r * math.cos(ph))
print(cmath.rect(r, ph).imag == r * math.sin(ph))
print((-1) ** 0.5)
cmath.sqrt
print(cmath.sqrt(-1))
a = 1
result = 'even' if a % 2 == 0 else 'odd'
print(result)
a = 2
result = 'even' if a % 2 == 0 else 'odd'
print(result)
a = 1
result = a * 2 if a % 2 == 0 else a * 3
print(result)
a = 2
result = a * 2 if a % 2 == 0 else a * 3
print(result)
a = 1
print('even') if a % 2 == 0 else print('odd')
a = 1
if a % 2 == 0:
print('even')
print('odd')
a = -2
result = 'negative and even' if a < 0 and a % 2 == 0 else 'positive or odd'
print(result)
negative and even
a = -1
result = 'negative and even' if a < 0 and a % 2 == 0 else 'positive or odd'
print(result)
positive or odd
a = 2
result = 'negative' if a < 0 else 'positive' if a > 0 else 'zero'
print(result)
a = 0
result = 'negative' if a < 0 else 'positive' if a > 0 else 'zero'
print(result)
a = -2
result = 'negative' if a < 0 else 'positive' if a > 0 else 'zero'
print(result)
result = 'negative' if a < 0 else ('positive' if a > 0 else 'zero')
print(result)
result = ('negative' if a < 0 else 'positive') if a > 0 else 'zero'
print(result)
l = ['even' if i % 2 == 0 else i for i in range(10)]
print(l)
l = [i * 10 if i % 2 == 0 else i for i in range(10)]
print(l)
get_odd_even = lambda x: 'even' if x % 2 == 0 else 'odd'
print(get_odd_even(1))
print(get_odd_even(2))
import csv
with open('data/src/sample.csv', 'r') as f:
reader = csv.reader(f)
for row in reader:
print(row)
with open('data/src/sample_space.csv', 'r') as f:
reader = csv.reader(f)
for row in reader:
print(row)
with open('data/src/sample_space.csv', 'r') as f:
reader = csv.reader(f, skipinitialspace=True)
for row in reader:
print(row)
with open('data/src/sample_double_quotation.csv', 'r') as f:
reader = csv.reader(f)
for row in reader:
print(row)
with open('data/src/sample_double_quotation.csv', 'r') as f:
reader = csv.reader(f, skipinitialspace=True)
for row in reader:
print(row)
import csv
import pprint
with open('data/src/sample_header.csv') as f:
print(f.read())
with open('data/src/sample_header.csv') as f:
reader = csv.DictReader(f)
l = [row for row in reader]
pprint.pprint(l)
OrderedDict([('a', '11'), ('b', '12'), ('c', '13'), ('d', '14')])
OrderedDict([('a', '21'), ('b', '22'), ('c', '23'), ('d', '24')])
OrderedDict([('a', '31'), ('b', '32'), ('c', '33'), ('d', '34')])
print(l[1])
OrderedDict([('a', '21'), ('b', '22'), ('c', '23'), ('d', '24')])
print(l[1]['c'])
with open('data/src/sample.csv') as f:
print(f.read())
with open('data/src/sample.csv') as f:
reader = csv.DictReader
f, fieldnames=['a', 'b', 'c', 'd']
for row in reader:
print(row)
OrderedDict([('a', '11'), ('b', '12'), ('c', '13'), ('d', '14')])
OrderedDict([('a', '21'), ('b', '22'), ('c', '23'), ('d', '24')])
OrderedDict([('a', '31'), ('b', '32'), ('c', '33'), ('d', '34')])
with open('data/src/sample_header_index.csv') as f:
print(f.read())
with open('data/src/sample_header_index.csv') as f:
reader = csv.DictReader(f)
l = [row for row in reader]
pprint.pprint(l, width=100)
OrderedDict([('', 'ONE'), ('a', '11'), ('b', '12'), ('c', '13'), ('d', '14')])
OrderedDict([('', 'TWO'), ('a', '21'), ('b', '22'), ('c', '23'), ('d', '24')])
OrderedDict([('', 'THREE'), ('a', '31'), ('b', '32'), ('c', '33'), ('d', '34')])
print([od.pop('') for od in l])
pprint.pprint(l)
OrderedDict([('a', '11'), ('b', '12'), ('c', '13'), ('d', '14')])
OrderedDict([('a', '21'), ('b', '22'), ('c', '23'), ('d', '24')])
OrderedDict([('a', '31'), ('b', '32'), ('c', '33'), ('d', '34')])
import csv
d1 = {'a': 1, 'b': 2, 'c': 3}
d2 = {'a': 10, 'c': 30}
with open('data/temp/sample_dictwriter.csv', 'w') as f:
writer = csv.DictWriter(f, ['a', 'b', 'c'])
writer.writeheader()
writer.writerow(d1)
writer.writerow(d2)
with open('data/temp/sample_dictwriter.csv') as f:
print(f.read())
with open('data/temp/sample_dictwriter_list.csv', 'w') as f:
writer = csv.DictWriter(f, ['a', 'b', 'c'])
writer.writeheader()
writer.writerows([d1, d2])
with open('data/temp/sample_dictwriter_list.csv') as f:
print(f.read())
with open('data/temp/sample_dictwriter_ignore.csv', 'w') as f:
writer = csv.DictWriter(f, ['a', 'c'], extrasaction='ignore')
writer.writeheader()
writer.writerows([d1, d2])
with open('data/temp/sample_dictwriter_ignore.csv') as f:
print(f.read())
import numpy as np
with open('data/src/sample.csv') as f:
print(f.read())
a = np.loadtxt('data/src/sample.csv', delimiter=',')
print(type(a))
numpy.ndarray
print(a)
print(a[1:, :2])
print(a.mean())
print(a.sum(axis=0))
import pandas as pd
with open('data/src/sample_pandas_normal.csv') as f:
print(f.read())
df = pd.read_csv('data/src/sample_pandas_normal.csv', index_col=0)
print(type(df))
pandas.core.frame.DataFrame
print(df)
print(df['age'])
Name: age
dtype: int64
print(df.at['Bob', 'age'])
print(df.query('state == "NY"'))
print(df.query('age > 30'))
print(df.describe())
import csv
import pprint
with open('data/src/sample.csv') as f:
print(f.read())
with open('data/src/sample.csv') as f:
reader = csv.reader(f)
for row in reader:
print(row)
with open('data/src/sample.csv') as f:
reader = csv.reader(f)
l = [row for row in reader]
print(l)
print(l[1])
print(l[1][1])
l_T = [list(x) for x in zip(*l)]
print(l_T)
print(l_T[1])
print(l[0][0])
print(type(l[0][0]))
r = l[0]
print(r)
print([int(v) for v in r])
print([[int(v) for v in row] for row in l])
with open('data/src/sample.csv') as f:
reader = csv.reader(f, quoting=csv.QUOTE_NONNUMERIC)
l_f = [row for row in reader]
print(l_f)
print(l_f[0][0])
print(type(l_f[0][0]))
with open('data/src/sample.txt') as f:
print(f.read())
with open('data/src/sample.txt') as f:
reader = csv.reader(f, delimiter=' ')
l = [row for row in reader]
print(l)
with open('data/src/sample_quote.csv') as f:
print(f.read())
with open('data/src/sample_quote.csv') as f:
reader = csv.reader(f)
for row in reader:
print(row)
with open('data/src/sample_quote.csv') as f:
reader = csv.reader(f, quoting=csv.QUOTE_NONE)
for row in reader:
print(row)
with open('data/src/sample_linebreak.csv') as f:
print(f.read())
with open('data/src/sample_linebreak.csv', newline='') as f:
reader = csv.reader(f)
for row in reader:
print(row)
with open('data/src/sample_header_index.csv') as f:
print(f.read())
with open('data/src/sample_header_index.csv') as f:
reader = csv.reader(f)
l = [row for row in reader]
pprint.pprint(l)
print([row[1:] for row in l[1:]])
print([[int(v) for v in row[1:]] for row in l[1:]])
import csv
with open('data/temp/sample_writer_row.csv', 'w') as f:
writer = csv.writer(f)
writer.writerow([0, 1, 2])
writer.writerow(['a', 'b', 'c'])
with open('data/temp/sample_writer_row.csv') as f:
print(f.read())
l = [[11, 12, 13, 14], [21, 22, 23, 24], [31, 32, 33, 34]]
print(l)
with open('data/temp/sample_writer.csv', 'w') as f:
writer = csv.writer(f)
writer.writerows(l)
with open('data/temp/sample_writer.csv') as f:
print(f.read())
with open('data/temp/sample_writer_row.csv') as f:
print(f.read())
with open('data/temp/sample_writer_row.csv', 'a') as f:
writer = csv.writer(f)
writer.writerow(['X', 'Y', 'Z'])
with open('data/temp/sample_writer_row.csv') as f:
print(f.read())
with open('data/temp/sample_writer.tsv', 'w') as f:
writer = csv.writer(f, delimiter='\t')
writer.writerows(l)
with open('data/temp/sample_writer.tsv') as f:
print(f.read())
l = [[0, 1, 2], ['a,b,c', 'x', 'y']]
with open('data/temp/sample_writer_quote.csv', 'w') as f:
writer = csv.writer(f)
writer.writerows(l)
with open('data/temp/sample_writer_quote.csv') as f:
print(f.read())
with open('data/temp/sample_writer_quote_all.csv', 'w') as f:
writer = csv.writer(f, quoting=csv.QUOTE_ALL)
writer.writerows(l)
with open('data/temp/sample_writer_quote_all.csv') as f:
print(f.read())
with open('data/temp/sample_writer_quote_nonnumeric.csv', 'w') as f:
writer = csv.writer(f, quoting=csv.QUOTE_NONNUMERIC)
writer.writerows(l)
with open('data/temp/sample_writer_quote_nonnumeric.csv') as f:
print(f.read())
with open('data/temp/sample_writer_quote_none.csv', 'w') as f:
writer = csv.writer(f, quoting=csv.QUOTE_NONE, escapechar='\\')
writer.writerows(l)
with open('data/temp/sample_writer_quote_none.csv') as f:
print(f.read())
with open('data/temp/sample_writer_quote_char.csv', 'w') as f:
writer = csv.writer(f, quotechar="'")
writer.writerows(l)
with open('data/temp/sample_writer_quote_char.csv') as f:
print(f.read())
l = [[0, 1, 2], ['a\nb', 'x', 'y']]
with open('data/temp/sample_writer_linebreak.csv', 'w', newline='') as f:
writer = csv.writer(f)
writer.writerows(l)
with open('data/temp/sample_writer_linebreak.csv') as f:
print(f.read())
l = [[11, 12, 13, 14], [21, 22, 23, 24], [31, 32, 33, 34]]
print(l)
header = ['', 'a', 'b', 'c', 'd']
index = ['ONE', 'TWO', 'THREE']
with open('data/temp/sample_writer_header_index.csv', 'w') as f:
writer = csv.writer(f)
writer.writerow(header)
for i, row in zip(index, l):
writer.writerow([i] + row)
with open('data/temp/sample_writer_header_index.csv') as f:
print(f.read())
import datetime
dt_now = datetime.datetime.now()
print(dt_now)
print(type(dt_now))
datetime.datetime
print(dt_now.year)
print(dt_now.hour)
dt = datetime.datetime(2018, 2, 1, 12, 15, 30, 2000)
print(dt)
print(dt.minute)
print(dt.microsecond)
dt = datetime.datetime(2018, 2, 1)
print(dt)
print(dt.minute)
print(dt_now)
print(type(dt_now))
datetime.datetime
print(dt_now.date())
print(type(dt_now.date()))
datetime.date
d_today = datetime.date.today()
print(d_today)
print(type(d_today))
datetime.date
print(d_today.year)
d = datetime.date(2018, 2, 1)
print(d)
print(d.month)
t = datetime.time(12, 15, 30, 2000)
print(t)
print(type(t))
datetime.time
print(t.hour)
t = datetime.time()
print(t)
td = dt_now - dt
print(td)
print(type(td))
datetime.timedelta
print(td.days)
print(td.seconds)
print(td.microseconds)
print(td.total_seconds())
td_1w = datetime.timedelta(weeks=1)
print(td_1w)
print(td_1w.days)
d_1w = d_today - td_1w
print(d_1w)
td_10d = datetime.timedelta(days=10)
print(td_10d)
dt_10d = dt_now + td_10d
print(dt_10d)
td_50m = datetime.timedelta(minutes=50)
print(td_50m)
print(td_50m.seconds)
dt_50m = dt_now + td_50m
print(dt_50m)
d_target = datetime.date(2020, 7, 24)
td = d_target - d_today
print(td)
print(td.days)
print(dt_now.strftime('%Y-%m-%d %H:%M:%S'))
print(d_today.strftime('%y%m%d'))
print(d_today.strftime('%A, %B %d, %Y'))
print(d_today.strftime('%Y年%m月%d日'))
print('日番号（1年の何日目か / 正月が001）:', d_today.strftime('%j'))
print('週番号（週の始まりは日曜日 / 正月が00）:', d_today.strftime('%U'))
print('週番号（週の始まりは月曜日 / 正月が00）:', d_today.strftime('%W'))
week_num_mon = int(d_today.strftime('%W'))
print(week_num_mon)
print(type(week_num_mon))
d = datetime.date(2018, 2, 1)
td = datetime.timedelta(weeks=2)
n = 8
f = '%Y年%m月%d日'
l = []
for i in range(n):
l.append((d + i * td).strftime(f))
print(l)
print('\n'.join(l))
l = [(d + i * td).strftime(f) for i in range(n)]
print(l)
date_str = '2018/2/1 12:30'
date_dt = datetime.datetime.strptime(date_str, '%Y/%m/%d %H:%M')
print(date_dt)
print(type(date_dt))
datetime.datetime
print(date_dt.strftime('%Y年%m月%d日 %H時%M分'))
date_str = '2018年2月1日'
date_format = '%Y年%m月%d日'
td_10_d = datetime.timedelta(days=10)
date_dt = datetime.datetime.strptime(date_str, date_format)
date_dt_new = date_dt - td_10_d
date_str_new = date_dt_new.strftime(date_format)
print(date_str_new)
import datetime
dt = datetime.datetime(2018, 1, 1)
print(dt)
print(dt.weekday())
print(type(dt.weekday()))
print(dt.isoweekday())
print(type(dt.isoweekday()))
w_list = ['月曜日', '火曜日', '水曜日', '木曜日', '金曜日', '土曜日', '日曜日']
print(w_list[dt.weekday()])
def get_day_of_week_jp(dt):
w_list = ['月曜日', '火曜日', '水曜日', '木曜日', '金曜日', '土曜日', '日曜日']
return(w_list[dt.weekday()])
print(get_day_of_week_jp(dt))
dt2 = datetime.datetime(2018, 1, 2)
print(dt2)
print(get_day_of_week_jp(dt2))
s = '2018年1月10日'
print(get_day_of_week_jp(datetime.datetime.strptime(s, '%Y年%m月%d日')))
def get_day_of_week_jp_s(s):
return get_day_of_week_jp(datetime.datetime.strptime(s, '%Y年%m月%d日'))
print(get_day_of_week_jp_s(s))
def get_month_jp(dt):
m_list = [None, '睦月', '如月', '弥生', '卯月', '皐月', '水無月', '文月', '葉月', '長月', '神無月', '霜月', '師走']
return(m_list[dt.month])
print(get_month_jp(dt2))
import datetime
import locale
dt = datetime.datetime(2018, 1, 1)
print(dt)
print(dt.strftime('%A, %a, %B, %b'))
print(locale.getlocale(locale.LC_TIME))
locale.setlocale(locale.LC_TIME, 'ja_JP.UTF-8')
print(locale.getlocale(locale.LC_TIME))
print(dt.strftime('%A, %a, %B, %b'))
locale.setlocale(locale.LC_TIME, 'en_US.UTF-8')
print(dt.strftime('%A, %a, %B, %b'))
locale.setlocale(locale.LC_TIME, 'de_DE.UTF-8')
print(dt.strftime('%A, %a, %B, %b'))
locale.setlocale(locale.LC_TIME, 'ja_JP.UTF-8')
s = '2018-01-01'
s_dow = datetime.datetime.strptime(s, '%Y-%m-%d').strftime('%A')
print(s_dow)
import datetime
def get_first_date(dt):
return dt.replace(day=1)
dt = datetime.datetime(2019, 1, 10, 10, 10, 10)
print(dt)
print(get_first_date(dt))
d = datetime.date(2019, 1, 10)
print(d)
print(get_first_date(d))
print(datetime.date.today())
print(get_first_date(datetime.date.today()))
def get_first_date2(year, month):
return datetime.date(year, month, 1)
print(get_first_date2(2019, 1))
import datetime
s = '2018-12-31'
d = datetime.date.fromisoformat(s)
print(d)
print(type(d))
datetime.date
print(datetime.date.fromisoformat('2018-12'))
ValueError: Invalid
print(datetime.date.fromisoformat('2018-01-01'))
print(datetime.date.fromisoformat('2018-1-1'))
ValueError: Invalid
s = '05:00:30.001000'
t = datetime.time.fromisoformat(s)
print(t)
print(type(t))
datetime.time
print(datetime.time.fromisoformat('05'))
print(datetime.time.fromisoformat('5:00:30'))
ValueError: Invalid
s = '2018-12-31T05:00:30.001000'
dt = datetime.datetime.fromisoformat(s)
print(dt)
print(type(dt))
datetime.datetime
print(datetime.datetime.fromisoformat('2018-12-31x05:00:30.001000'))
print(datetime.datetime.fromisoformat('2018-12-31xx05:00:30.001000'))
ValueError: Invalid
print(datetime.datetime.fromisoformat('2018-12-31T05'))
print(datetime.datetime.fromisoformat('2018-12-31'))
print(datetime.datetime.fromisoformat('2018-12-31T5:00'))
ValueError: Invalid
s = '2018-12-31T05:00:30.001000'
print(datetime.date.fromisoformat(s))
ValueError: Invalid
print(datetime.time.fromisoformat(s))
ValueError: Invalid
d = datetime.datetime.fromisoformat(s).date()
print(d)
print(type(d))
datetime.date
t = datetime.datetime.fromisoformat(s).time()
print(t)
print(type(t))
datetime.time
s = '2018-12-31T05:00:30'
s_basic = s.replace('-', '').replace(':', '')
print(s_basic)
s = '2018-12-31T05:00:30.001000'
s_basic = s.split('.')[0].replace('-', '').replace(':', '')
print(s_basic)
s_ex = datetime.datetime.strptime(s_basic, '%Y%m%dT%H%M%S').isoformat()
print(s_ex)
import datetime
d = datetime.date(2018, 12, 31)
print(d)
print(type(d))
datetime.date
print(d.isoformat())
print(type(d.isoformat()))
t = datetime.time(5, 0, 30, 1000)
print(t)
print(type(t))
datetime.time
print(t.isoformat())
print(type(t.isoformat()))
print(t.isoformat('hours'))
print(t.isoformat('minutes'))
print(t.isoformat('seconds'))
print(t.isoformat('milliseconds'))
print(t.isoformat('microseconds'))
dt = datetime.datetime(2018, 12, 31, 5, 0, 30, 1000)
print(dt)
print(type(dt))
datetime.datetime
print(dt.isoformat())
print(type(dt.isoformat()))
print(dt.isoformat(' '))
print(dt.isoformat('x'))
print(dt.isoformat('xx'))
TypeError: isoformat
not str
print(dt.isoformat(timespec='hours'))
print(dt.isoformat(timespec='minutes'))
print(dt.isoformat(timespec='seconds'))
print(dt.isoformat(timespec='milliseconds'))
print(dt.isoformat(timespec='microseconds'))
print(dt.isoformat(' ', 'seconds'))
print(dt.date())
print(type(dt.date()))
datetime.date
print(dt.date().isoformat())
print(type(dt.date().isoformat()))
print(dt.time())
print(type(dt.time()))
datetime.time
print(dt.time().isoformat())
print(type(dt.time().isoformat()))
print(dt.time().isoformat('seconds'))
import datetime
s = '2018-12-31T05:00:30.001000'
dt = datetime.datetime.fromisoformat(s)
print(dt)
print(dt.tzinfo)
s = '2018-12-31T05:00:30.001000+09:00'
dt = datetime.datetime.fromisoformat(s)
print(dt)
print(dt.tzinfo)
print(dt.isoformat())
s = '2018-12-31T05:00:30.001000Z'
dt = datetime.datetime.fromisoformat(s)
ValueError: Invalid
print(s.replace('Z', '+00:00'))
dt_utc = datetime.datetime.fromisoformat(s.replace('Z', '+00:00'))
print(dt_utc)
print(dt_utc.tzinfo)
print(s.replace('Z', ''))
dt_none = datetime.datetime.fromisoformat(s.replace('Z', ''))
print(dt_none)
print(dt_none.tzinfo)
import datetime
import calendar
print(calendar.monthrange(2019, 1))
print(calendar.monthrange(2019, 1)[1])
def get_last_date(dt):
return dt.replace(day=calendar.monthrange(dt.year, dt.month)[1])
dt = datetime.datetime(2019, 1, 10, 10, 10, 10)
print(dt)
print(get_last_date(dt))
d = datetime.date(2019, 1, 10)
print(d)
print(get_last_date(d))
print(datetime.date.today())
print(get_last_date(datetime.date.today()))
def get_last_date2(year, month):
return datetime.date(year, month, calendar.monthrange(year, month)[1])
print(get_last_date2(2019, 1))
import datetime
d1 = datetime.date(2010, 10, 1)
print(d1)
print(type(d1))
print(d1.year, d1.month, d1.day)
datetime.date
d1_tt = d1.timetuple()
print(d1_tt)
print(type(d1_tt))
print(d1_tt.tm_year)
print(d1_tt[0])
time.struct_time(tm_year=2010, tm_mon=10, tm_mday=1, tm_hour=0, tm_min=0, tm_sec=0, tm_wday=4, tm_yday=274, tm_isdst=-1)
time.struct_time
d2 = datetime.date.today()
print(d2)
t = datetime.time(10, 5, 45)
print(t)
print(type(t))
print(t.hour, t.minute, t.second)
datetime.time
dt1 = datetime.datetime(2001, 5, 10, 14, 10, 5)
print(dt1)
print(type(dt1))
print(dt1.year, dt1.month, dt1.day, dt1.hour, dt1.minute, dt1.second)
print(dt1.date())
print(dt1.time())
datetime.datetime
dt2 = datetime.datetime.now()
print(dt2)
print(dt2.microsecond)
td1 = dt2 - dt1
print(td1)
print(type(td1))
print(td1.days, td1.seconds, td1.microseconds)
print(td1.total_seconds())
print(td1.days * 24 * 60 * 60 + td1.seconds + td1.microseconds / 1000000)
datetime.timedelta
dt = datetime.datetime(2015, 3, 20, 20, 0, 10)
td = datetime.timedelta(days=100)
dt_a = dt + td
print(dt)
print(dt_a)
dt = datetime.datetime(2015, 3, 20, 20, 0, 10)
print(dt)
print(dt.strftime('%Y/%m/%d(%a) %H:%M:%S'))
print(dt.strftime('%d %b. %Y'))
print(dt.strftime('%x %X'))
s = '2010/4/1'
dts = datetime.datetime.strptime(s, '%Y/%m/%d')
print(dts)
print(dts.year, dts.hour)
print(type(dts))
datetime.datetime
import datetime
dt_now = datetime.datetime.now()
print(dt_now)
print(type(dt_now))
datetime.datetime
print(dt_now.tzinfo)
print(dt_now.strftime('%Y年%m月%d日 %H:%M:%S'))
print(type(dt_now.strftime('%Y年%m月%d日 %H:%M:%S')))
print(dt_now.isoformat())
print(type(dt_now.isoformat()))
print(dt_now.year)
print(dt_now.month)
print(dt_now.day)
print(dt_now.hour)
print(dt_now.minute)
print(dt_now.second)
print(dt_now.microsecond)
print(type(dt_now.year))
dt_now_utc_aware = datetime.datetime.now(datetime.timezone.utc)
print(dt_now_utc_aware)
print(dt_now_utc_aware.tzinfo)
dt_now_jst_aware = datetime.datetime.now
datetime.timezone(datetime.timedelta(hours=9))
print(dt_now_jst_aware)
print(dt_now_jst_aware.tzinfo)
dt_now_utc_naive = datetime.datetime.utcnow()
print(dt_now_utc_naive)
print(dt_now_utc_naive.tzinfo)
d_today = datetime.date.today()
print(d_today)
print(type(d_today))
datetime.date
d_today_utc = datetime.datetime.utcnow().date()
print(d_today_utc)
print(type(d_today_utc))
datetime.date
t_now = datetime.datetime.now().time()
print(t_now)
print(type(t_now))
datetime.time
print(t_now.tzinfo)
dt_now_utc_aware = datetime.datetime.now(datetime.timezone.utc)
print(dt_now_utc_aware)
print(dt_now_utc_aware.tzinfo)
print(dt_now_utc_aware.time())
print(dt_now_utc_aware.time().tzinfo)
print(dt_now_utc_aware.timetz())
print(dt_now_utc_aware.timetz().tzinfo)
import datetime
import pytz
jst = pytz.timezone('Asia/Tokyo')
print(jst)
eastern = pytz.timezone('US/Eastern')
print(eastern)
print(type(jst))
print(issubclass(type(jst), datetime.tzinfo))
dt_aware = datetime.datetime
tzinfo=datetime.timezone.utc
print(dt_aware)
print(dt_aware.astimezone(jst))
print(dt_aware.astimezone(eastern))
print(dt_aware.replace(tzinfo=jst))
print(dt_aware.replace(tzinfo=eastern))
print(jst.localize(dt_aware.replace(tzinfo=None)))
print(eastern.localize(dt_aware.replace(tzinfo=None)))
dt_naive = datetime.datetime(2018, 12, 31, 5, 0, 30, 1000)
print(dt_naive)
print(dt_naive.replace(tzinfo=jst))
print(dt_naive.replace(tzinfo=eastern))
print(jst.localize(dt_naive))
print(eastern.localize(dt_naive))
datetime.datetime
tzinfo=jst
print(jst.localize(datetime.datetime(2018, 12, 31, 5, 0, 30, 1000)))
print(datetime.datetime.now(jst))
dt_dst = datetime.datetime(2018, 3, 11, 2, 0, 0, 0)
print(eastern.localize(dt_dst))
print(eastern.normalize(eastern.localize(dt_dst)))
import datetime
dt1 = datetime.datetime(year=2017, month=10, day=10, hour=15)
print(dt1)
print(type(dt1))
datetime.datetime
dt2 = datetime.datetime(year=2019, month=1, day=1, hour=12)
print(dt2)
td = dt2 - dt1
print(td)
print(type(td))
datetime.timedelta
print(td.days)
print(td.seconds)
print(td.microseconds)
print(td.total_seconds())
td_m = dt1 - dt2
print(td_m)
-448
print(td_m.days)
-448
print(td_m.seconds)
print(td_m.total_seconds())
-38696400.0
td_abs = abs(dt1 - dt2)
print(td_abs)
print(td_abs.days)
print(td_abs.seconds)
print(td_abs.total_seconds())
d1 = datetime.date(year=2017, month=10, day=10)
print(d1)
print(type(d1))
datetime.date
d2 = datetime.date(year=2019, month=1, day=1)
print(d2)
td = abs(d1 - d2)
print(td)
print(type(td))
datetime.timedelta
print(dt2 - d1)
TypeError: unsupported
type(s) 
datetime.datetime
datetime.date
print(dt2.date())
print(type(dt2.date()))
datetime.date
print(dt2.date() - d1)
print(datetime.datetime.combine(d1, datetime.time()))
print(type(datetime.datetime.combine(d1, datetime.time())))
datetime.datetime
print(dt2 - datetime.datetime.combine(d1, datetime.time()))
print(datetime.datetime.now() - dt2)
print(datetime.date.today() - d2)
td = datetime.timedelta(weeks=1, hours=20)
print(td)
print(type(td))
datetime.timedelta
print(dt1)
print(dt1 + td)
print(dt1 - td)
print(d1)
print(d1 + td)
print(d1 - td)
import datetime
td = datetime.timedelta
weeks=1
days=1
hours=1
minutes=1
seconds=1
milliseconds=1
microseconds=1
print(td)
print(type(td))
datetime.timedelta
print(datetime.timedelta(days=0.5, hours=-6, minutes=120))
td = datetime.timedelta
weeks=1
days=1
hours=1
minutes=1
seconds=1
milliseconds=1
microseconds=1
print(td)
print(type(td))
datetime.timedelta
print(td.days)
print(type(td.days))
print(td.seconds)
print(type(td.seconds))
print(td.microseconds)
print(type(td.microseconds))
print(td.total_seconds())
print(type(td.total_seconds()))
print(td.days * 24 * 60 * 60 + td.seconds + td.microseconds / 1000000)
s = str(td)
print(s)
print(type(s))
td = datetime.timedelta(days=1, hours=1, milliseconds=100)
sec = td.total_seconds()
print(sec)
print(type(sec))
sec = datetime.timedelta(days=1, hours=1, milliseconds=100).total_seconds()
print(sec)
td = datetime.timedelta(seconds=123456.789)
print(td)
s = str(td)
print(s)
print(type(s))
print(td.days)
print(td.seconds)
print(td.microseconds)
def get_h_m_s(td):
m, s = divmod(td.seconds, 60)
h, m = divmod(m, 60)
return h, m, s
print(get_h_m_s(td))
print(type(get_h_m_s(td)))
h, m, s = get_h_m_s(td)
print(h)
print(m)
print(s)
def get_d_h_m_s_us(sec):
td = datetime.timedelta(seconds=sec)
m, s = divmod(td.seconds, 60)
h, m = divmod(m, 60)
return td.days, h, m, s, td.microseconds
print(get_d_h_m_s_us(123456.789))
print(get_d_h_m_s_us(123.456789))
td_m = datetime.timedelta(seconds=-123456.789)
print(td_m)
-2
print(td_m.days)
-2
print(td_m.seconds)
print(td_m.microseconds)
print(td_m.total_seconds())
-123456.789
print(get_h_m_s(td_m))
print(get_d_h_m_s_us(-123456.789))
-2
print(abs(td_m))
print(get_h_m_s(abs(td_m)))
print(get_d_h_m_s_us(abs(-123456.789)))
import datetime
dt = datetime.datetime.fromtimestamp(0)
print(dt)
print(type(dt))
datetime.datetime
print(dt.tzinfo)
dt_utc_aware = datetime.datetime.fromtimestamp(0, datetime.timezone.utc)
print(dt_utc_aware)
print(dt_utc_aware.tzinfo)
dt_jst_aware = datetime.datetime.fromtimestamp(0, datetime.timezone(datetime.timedelta(hours=9)))
print(dt_jst_aware)
print(dt_jst_aware.tzinfo)
dt_utc_naive = datetime.datetime.utcfromtimestamp(0)
print(dt_utc_naive)
print(dt_utc_naive.tzinfo)
print(dt)
print(dt.timestamp())
print(type(dt.timestamp()))
print(dt_utc_aware)
print(dt_utc_aware.timestamp())
print(dt_jst_aware)
print(dt_jst_aware.timestamp())
print(dt_utc_naive)
print(dt_utc_naive.timestamp())
-32400.0
import datetime
dt_naive = datetime.datetime(2018, 12, 31, 5, 0, 30, 1000)
print(dt_naive)
print(dt_naive.tzinfo)
dt_aware = datetime.datetime
tzinfo=datetime.timezone.utc
print(dt_aware)
print(dt_aware.tzinfo)
print(type(dt_aware.tzinfo))
datetime.timezone
print(datetime.timezone.utc)
print(type(datetime.timezone.utc))
datetime.timezone
print(issubclass(type(datetime.timezone.utc), datetime.tzinfo))
tz_jst = datetime.timezone(datetime.timedelta(hours=9))
print(tz_jst)
print(type(tz_jst))
datetime.timezone
tz_jst_name = datetime.timezone(datetime.timedelta(hours=9), name='JST')
print(tz_jst_name)
dt_utc = datetime.datetime
tzinfo=datetime.timezone.utc
print(dt_utc)
print(dt_utc.tzinfo)
dt_jst = datetime.datetime
tzinfo=datetime.timezone(datetime.timedelta(hours=9))
print(dt_jst)
print(dt_jst.tzinfo)
import datetime
dt_jst = datetime.datetime
tzinfo=datetime.timezone(datetime.timedelta(hours=9))
print(dt_jst)
print(dt_jst.tzinfo)
dt_jst_to_utc = dt_jst.astimezone(datetime.timezone.utc)
print(dt_jst_to_utc)
print(dt_jst_to_utc.tzinfo)
dt_jst_to_m5h = dt_jst.astimezone(datetime.timezone(datetime.timedelta(hours=-5)))
print(dt_jst_to_m5h)
print(dt_jst_to_m5h.tzinfo)
dt_jst_to_utc_replace = dt_jst.replace(tzinfo=datetime.timezone.utc)
print(dt_jst_to_utc_replace)
print(dt_jst_to_utc_replace.tzinfo)
dt_jst_to_naive = dt_jst.replace(tzinfo=None)
print(dt_jst_to_naive)
print(dt_jst_to_naive.tzinfo)
dt_naive = datetime.datetime(2018, 12, 31, 5, 0, 30, 1000)
print(dt_naive)
print(dt_naive.tzinfo)
dt_naive_to_utc_replace = dt_naive.replace(tzinfo=datetime.timezone.utc)
print(dt_naive_to_utc_replace)
print(dt_naive_to_utc_replace.tzinfo)
dt_naive_to_jst_replace = dt_naive.replace(tzinfo=datetime.timezone(datetime.timedelta(hours=9)))
print(dt_naive_to_jst_replace)
print(dt_naive_to_jst_replace.tzinfo)
dt_naive_to_utc = dt_naive.astimezone(datetime.timezone.utc)
print(dt_naive_to_utc)
print(dt_naive_to_utc.tzinfo)
import datetime
dt_now = datetime.datetime.now()
print(dt_now)
print(dt_now.tzinfo)
dt_now_utc = datetime.datetime.now(datetime.timezone.utc)
print(dt_now_utc)
print(dt_now_utc.tzinfo)
dt_now_jst = datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=9)))
print(dt_now_jst)
print(dt_now_jst.tzinfo)
dt_utcnow = datetime.datetime.utcnow()
print(dt_utcnow)
print(dt_utcnow.tzinfo)
import datetime
s = '2018/12/31 05:00:30+0900'
dt = datetime.datetime.strptime(s, '%Y/%m/%d %H:%M:%S%z')
print(dt)
print(dt.tzinfo)
s = '2018-12-31T05:00:30+09:00'
dt = datetime.datetime.fromisoformat(s)
print(dt)
print(dt.tzinfo)
import datetime
dt_utc = datetime.datetime
tzinfo=datetime.timezone.utc
print(dt_utc.tzinfo)
print(type(dt_utc.tzinfo))
datetime.timezone
print(dt_utc.utcoffset())
print(type(dt_utc.utcoffset()))
datetime.timedelta
dt_jst = datetime.datetime
tzinfo=datetime.timezone(datetime.timedelta(hours=9))
print(dt_jst.tzinfo)
print(type(dt_jst.tzinfo))
datetime.timezone
print(dt_jst.utcoffset())
print(type(dt_jst.utcoffset()))
datetime.timedelta
from decimal import Decimal, ROUND_HALF_UP, ROUND_HALF_EVEN
print(Decimal(0.05))
print(type(Decimal(0.05)))
decimal.Decimal
print(Decimal(0.5))
print(Decimal('0.05'))
f = 123.456
print(Decimal(str(f)).quantize(Decimal('0'), rounding=ROUND_HALF_UP))
print(Decimal(str(f)).quantize(Decimal('0.1'), rounding=ROUND_HALF_UP))
print(Decimal(str(f)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
print('0.4 =>', Decimal(str(0.4)).quantize(Decimal('0'), rounding=ROUND_HALF_UP))
print('0.5 =>', Decimal(str(0.5)).quantize(Decimal('0'), rounding=ROUND_HALF_UP))
print('0.6 =>', Decimal(str(0.6)).quantize(Decimal('0'), rounding=ROUND_HALF_UP))
print('0.05 =>', round(0.05, 1))
print('0.15 =>', round(0.15, 1))
print('0.25 =>', round(0.25, 1))
print('0.35 =>', round(0.35, 1))
print('0.45 =>', round(0.45, 1))
print('0.05 =>', Decimal(0.05).quantize(Decimal('0.1'), rounding=ROUND_HALF_EVEN))
print('0.15 =>', Decimal(0.15).quantize(Decimal('0.1'), rounding=ROUND_HALF_EVEN))
print('0.25 =>', Decimal(0.25).quantize(Decimal('0.1'), rounding=ROUND_HALF_EVEN))
print('0.35 =>', Decimal(0.35).quantize(Decimal('0.1'), rounding=ROUND_HALF_EVEN))
print('0.45 =>', Decimal(0.45).quantize(Decimal('0.1'), rounding=ROUND_HALF_EVEN))
print('0.05 =>', Decimal(str(0.05)).quantize(Decimal('0.1'), rounding=ROUND_HALF_EVEN))
print('0.15 =>', Decimal(str(0.15)).quantize(Decimal('0.1'), rounding=ROUND_HALF_EVEN))
print('0.25 =>', Decimal(str(0.25)).quantize(Decimal('0.1'), rounding=ROUND_HALF_EVEN))
print('0.35 =>', Decimal(str(0.35)).quantize(Decimal('0.1'), rounding=ROUND_HALF_EVEN))
print('0.45 =>', Decimal(str(0.45)).quantize(Decimal('0.1'), rounding=ROUND_HALF_EVEN))
print(Decimal(2.675))
print(Decimal(2.675).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
print(Decimal(str(2.675)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
print(Decimal(2.675).quantize(Decimal('0.01'), rounding=ROUND_HALF_EVEN))
print(Decimal(str(2.675)).quantize(Decimal('0.01'), rounding=ROUND_HALF_EVEN))
d = Decimal('123.456').quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
print(d)
print(type(d))
decimal.Decimal
print(1.2 + d)
TypeError: unsupported
type(s) 
decimal.Decimal
print(1.2 + float(d))
i = 99518
print(Decimal(i).quantize(Decimal('10'), rounding=ROUND_HALF_UP))
print(Decimal('10').as_tuple())
DecimalTuple(sign=0, digits=(1, 0), exponent=0)
print(Decimal('1E1').as_tuple())
DecimalTuple(sign=0, digits=(1,), exponent=1)
print(Decimal(i).quantize(Decimal('1E1'), rounding=ROUND_HALF_UP))
print(int(Decimal(i).quantize(Decimal('1E1'), rounding=ROUND_HALF_UP)))
print(int(Decimal(i).quantize(Decimal('1E2'), rounding=ROUND_HALF_UP)))
print(int(Decimal(i).quantize(Decimal('1E3'), rounding=ROUND_HALF_UP)))
print('4 =>', int(Decimal(4).quantize(Decimal('1E1'), rounding=ROUND_HALF_UP)))
print('5 =>', int(Decimal(5).quantize(Decimal('1E1'), rounding=ROUND_HALF_UP)))
print('6 =>', int(Decimal(6).quantize(Decimal('1E1'), rounding=ROUND_HALF_UP)))
d = {'k1': 1, 'k2': 2}
d['k3'] = 3
print(d)
d['k1'] = 100
print(d)
d1 = {'k1': 1, 'k2': 2}
d2 = {'k1': 100, 'k3': 3, 'k4': 4}
d1.update(d2)
print(d1)
d1 = {'k1': 1, 'k2': 2}
d2 = {'k3': 3, 'k4': 4}
d3 = {'k5': 5, 'k6': 6}
d1.update(d2, d3)
TypeError: update
d1.update(**d2, **d3)
print(d1)
d = {'k1': 1, 'k2': 2}
d.update(k1=100, k3=3, k4=4)
print(d)
d = {'k1': 1, 'k2': 2}
d.update([('k1', 100), ('k3', 3), ('k4', 4)])
print(d)
d = {'k1': 1, 'k2': 2}
keys = ['k1', 'k3', 'k4']
values = [100, 3, 4]
d.update(zip(keys, values))
print(d)
d = {'k1': 1, 'k2': 2, 'k3': 3}
d['k10'] = d['k1']
del d['k1'
print(d)
d = {'k1': 1, 'k2': 2, 'k3': 3}
print(d.pop('k1'))
print(d)
d = {'k1': 1, 'k2': 2, 'k3': 3}
d['k10'] = d.pop('k1')
print(d)
d = {'k1': 1, 'k2': 2, 'k3': 3}
print(d.pop('k10'))
print(d.pop('k10', None))
print(d)
def change_dict_key(d, old_key, new_key, default_value=None):
d[new_key] = d.pop(old_key, default_value)
d = {'k1': 1, 'k2': 2, 'k3': 3}
change_dict_key(d, 'k1', 'k10')
print(d)
d = {'k1': 1, 'k2': 2, 'k3': 3}
change_dict_key(d, 'k10', 'k100')
print(d)
d = {'k1': 1, 'k2': 2, 'k3': 3}
change_dict_key(d, 'k10', 'k100', 100)
print(d)
d = {'k1': 1, 'k2': 2, 'k3': 3}
change_dict_key(d, 'k1', 'k2')
print(d)
def change_dict_key_setdefault(d, old_key, new_key, default_value=None):
d.setdefault(new_key, d.pop(old_key, default_value))
d = {'k1': 1, 'k2': 2, 'k3': 3}
change_dict_key_setdefault(d, 'k1', 'k2')
print(d)
d = {'k1': 1, 'k2': 2, 'k3': 3}
change_dict_key_setdefault(d, 'k1', 'k10')
print(d)
d = {'k1': 1, 'k2': 2, 'k3': 3}
change_dict_key_setdefault(d, 'k10', 'k100')
print(d)
def change_dict_key_exist(d, old_key, new_key):
if old_key in d:
d[new_key] = d.pop(old_key)
d = {'k1': 1, 'k2': 2, 'k3': 3}
change_dict_key_exist(d, 'k1', 'k10')
print(d)
d = {'k1': 1, 'k2': 2, 'k3': 3}
change_dict_key_exist(d, 'k10', 'k100')
print(d)
d = {'k1': 1, 'k2': 2, 'k3': 3}
change_dict_key_exist(d, 'k1', 'k2')
print(d)
def change_dict_key_exist_setdefault(d, old_key, new_key):
if old_key in d:
d.setdefault(new_key, d.pop(old_key))
d = {'k1': 1, 'k2': 2, 'k3': 3}
change_dict_key_exist_setdefault(d, 'k1', 'k2')
print(d)
d = {'k1': 1, 'k2': 2, 'k3': 3}
d.clear()
print(d)
d = {'k1': 1, 'k2': 2, 'k3': 3}
removed_value = d.pop('k1')
print(d)
print(removed_value)
d = {'k1': 1, 'k2': 2, 'k3': 3}
removed_value = d.pop('k4')
print(d)
d = {'k1': 1, 'k2': 2, 'k3': 3}
removed_value = d.pop('k4', None)
print(d)
print(removed_value)
d = {'k1': 1, 'k2': 2}
k, v = d.popitem()
print(k)
print(v)
print(d)
k, v = d.popitem()
print(k)
print(v)
print(d)
k, v = d.popitem()
popitem()
d = {'k1': 1, 'k2': 2, 'k3': 3}
del d['k2'
print(d)
d = {'k1': 1, 'k2': 2, 'k3': 3}
del d['k1'
d['k3']
print(d)
d = {'k1': 1, 'k2': 2, 'k3': 3}
del d['k4'
print(d)
d = {'k1': 1, 'k2': 2, 'k3': 3}
print(d)
d = {'k1': 1, 'k2': 2, 'k3': 3, 'k3': 300}
print(d)
d = dict(k1=1, k2=2, k3=3)
print(d)
d = dict([('k1', 1), ('k2', 2), ('k3', 4)])
print(d)
d = dict((['k1', 1], ['k2', 2], ['k3', 4]))
print(d)
d = dict([{'k1', 1}, {'k2', 2}, {'k3', 4}])
print(d)
keys = ['k1', 'k2', 'k3']
values = [1, 2, 3]
d = dict(zip(keys, values))
print(d)
d_other = {'k10': 10, 'k100': 100}
d = dict(d_other)
print(d)
print(d == d_other)
print(d is d_other)
l = ['Alice', 'Bob', 'Charlie']
d = {s: len(s) for s in l}
print(d)
keys = ['k1', 'k2', 'k3']
values = [1, 2, 3]
d = {k: v for k, v in zip(keys, values)}
print(d)
d = {k: v for k, v in zip(keys, values) if v % 2 == 1}
print(d)
d1 = {'k1': 1, 'k2': 2}
d2 = {'k3': 3, 'k4': 4}
print(d)
print(d1)
print(d2)
d3 = {'k5': 5, 'k6': 6}
d4 = {'k1': 100, 'k3': 300}
d1 = {'k1': 1, 'k2': 2}
d2 = {'k3': 3, 'k4': 4}
d = dict(d1, d2)
TypeError: dict
d = dict(**d1, **d2)
print(d)
d = {'key1': 'val1', 'key2': 'val2', 'key3': 'val3'}
print(d['key1'])
print(d['key4'])
d['key4'] = 'val4'
print(d)
print(d.get('key1'))
print(d.get('key5'))
print(d.get('key5', 'NO KEY'))
print(d.get('key5', 100))
print(d)
d = {'key1': 'aaa', 'key2': 'aaa', 'key3': 'bbb'}
value = d['key1']
print(value)
keys = [k for k, v in d.items() if v == 'aaa']
print(keys)
keys = [k for k, v in d.items() if v == 'bbb']
print(keys)
keys = [k for k, v in d.items() if v == 'xxx']
print(keys)
key = [k for k, v in d.items() if v == 'aaa'][0]
print(key)
key = [k for k, v in d.items() if v == 'bbb'][0]
print(key)
key = [k for k, v in d.items() if v == 'xxx'][0]
print(key)
IndexError: list
def get_keys_from_value(d, val):
return [k for k, v in d.items() if v == val]
keys = get_keys_from_value(d, 'aaa')
print(keys)
def get_key_from_value(d, val):
keys = [k for k, v in d.items() if v == val]
if keys:
return keys[0]
return None
key = get_key_from_value(d, 'aaa')
print(key)
key = get_key_from_value(d, 'bbb')
print(key)
key = get_key_from_value(d, 'xxx')
print(key)
d_num = {'key1': 1, 'key2': 2, 'key3': 3}
keys = [k for k, v in d_num.items() if v >= 2]
print(keys)
keys = [k for k, v in d_num.items() if v % 2 == 1]
print(keys)
d_str = {'key1': 'aaa@xxx.com', 'key2': 'bbb@yyy.net', 'key3': 'ccc@zzz.com'}
keys = [k for k, v in d_str.items() if v.endswith('com')]
print(keys)
d = {'key1': 'val1', 'key2': 'val2', 'key3': 'val3'}
print('key1' in d)
print('val1' in d)
print('key4' not in d)
print(d['key1'])
print(d['key4'])
print(d.get('key4'))
print('val1' in d.values())
print('val4' not in d.values())
print(('key1', 'val1') in d.items())
print(('key1', 'val2') in d.items())
print(('key1', 'val2') not in d.items())
d1 = {'a': 1, 'b': 2, 'c': 3}
d2 = {'b': 2, 'c': 4, 'd': 5}
print(list(d1.keys()))
print(type(d1.keys()))
print(list(d1.items()))
print(type(d1.items()))
intersection_keys = d1.keys() & d2.keys()
print(intersection_keys)
print(type(intersection_keys))
intersection_items = d1.items() & d2.items()
print(intersection_items)
intersection_dict = dict(d1.items() & d2.items())
print(intersection_dict)
print(type(intersection_dict))
union_keys = d1.keys() | d2.keys()
print(union_keys)
union_items = d1.items() | d2.items()
print(union_items)
union_dict = dict(d1.items() | d2.items())
print(union_dict)
symmetric_difference_keys = d1.keys() ^ d2.keys()
print(symmetric_difference_keys)
symmetric_difference_items = d1.items() ^ d2.items()
print(symmetric_difference_items)
symmetric_difference_dict = dict(d1.items() ^ d2.items())
print(symmetric_difference_dict)
difference_keys = d1.keys() - d2.keys()
print(difference_keys)
difference_items = d1.items() - d2.items()
print(difference_items)
difference_dict = dict(d1.items() - d2.items())
print(difference_dict)
d = {'key1': 1, 'key2': 2, 'key3': 3}
for k in d:
print(k)
for k in d.keys():
print(k)
keys = d.keys()
print(keys)
print(type(keys))
dict_keys(['key1', 'key2', 'key3'])
k_list = list(d.keys())
print(k_list)
print(type(k_list))
for v in d.values():
print(v)
values = d.values()
print(values)
print(type(values))
dict_values([1, 2, 3])
v_list = list(d.values())
print(v_list)
print(type(v_list))
for k, v in d.items():
print(k, v)
for t in d.items():
print(t)
print(type(t))
print(t[0])
print(t[1])
print('---')
items = d.items()
print(items)
print(type(items))
dict_items([('key1', 1), ('key2', 2), ('key3', 3)])
i_list = list(d.items())
print(i_list)
print(type(i_list))
print(i_list[0])
print(type(i_list[0]))
import pprint
pprint.pprint(l)
sorted(l)
not supported 
pprint.pprint
l, key=lambda x: x['Age']
pprint.pprint
l, key=lambda x: x['Name']
pprint.pprint
l, key=lambda x: x['Age'], reverse=True
l, key=lambda x: x['Point']
l, key=lambda x: x.get('Point')
not supported 
pprint.pprint
l, key=lambda x: x.get('Point', 75)
pprint.pprint
l, key=lambda x: x.get('Point', float('inf'))
pprint.pprint
l, key=lambda x: x.get('Point', -float('inf'))
import operator
pprint.pprint(sorted(l, key=operator.itemgetter('Age')))
pprint.pprint(sorted(l, key=operator.itemgetter('Name')))
sorted(l, key=operator.itemgetter('Point'))
pprint.pprint(l_dup)
pprint.pprint(sorted(l_dup, key=operator.itemgetter('State')))
pprint.pprint(sorted(l_dup, key=operator.itemgetter('State', 'Age')))
pprint.pprint(sorted(l_dup, key=operator.itemgetter('Age', 'State')))
pprint.pprint
l_dup, key=lambda x: (x['State'], x['Age'])
max(l)
not supported 
l, key=lambda x: x['Age']
l, key=lambda x: x['Age']
l, key=lambda x: x['Age']
print(max(l, key=operator.itemgetter('Age')))
l_name = [d.get('Name') for d in l]
print(l_name)
l_age = [d.get('Age') for d in l]
print(l_age)
l_point = [d.get('Point') for d in l]
print(l_point)
l_name = [d['Name'] for d in l]
print(l_name)
l_point = [d['Point'] for d in l]
l_point_default = [d.get('Point', 0) for d in l]
print(l_point_default)
l_point_ignore = [d.get('Point') for d in l if d.get('Point')]
print(l_point_ignore)
d1 = {'k1': 1, 'k2': 2}
d2 = {'k1': 100, 'k3': 3, 'k4': 4}
print(d1 | d2)
print(d2 | d1)
d1 = {'k1': 1, 'k2': 2}
d2 = {'k3': 3, 'k4': 4}
d3 = {'k5': 5, 'k6': 6}
print(d1 | d2 | d3)
d1 = {'k1': 1, 'k2': 2}
d2 = {'k1': 100, 'k3': 3, 'k4': 4}
d1 |= d2
print(d1)
d = {'k1': 1, 'k2': 2}
d |= [('k1', 100), ('k3', 3), ('k4', 4)]
print(d)
print(d | [('k1', 100), ('k3', 3), ('k4', 4)])
TypeError: unsupported
type(s) 
d = {'k1': 1, 'k2': 2}
d.setdefault('k3', 3)
print(d)
d.setdefault('k4')
print(d)
d.setdefault('k1', 100)
print(d)
print(d.setdefault('k5', 5))
print(d)
print(d.setdefault('k6'))
print(d)
print(d.setdefault('k1', 100))
print(d.setdefault('k1', -100))
print(d.setdefault('k1'))
print(d)
d = {'key1': 'val1', 'key2': 'val2', 'key3': 'val3'}
d_swap = {v: k for k, v in d.items()}
print(d_swap)
def get_swap_dict(d):
return {v: k for k, v in d.items()}
d_swap = get_swap_dict(d)
print(d_swap)
d_duplicate = {'key1': 'val1', 'key2': 'val1', 'key3': 'val3'}
d_duplicate_swap = get_swap_dict(d_duplicate)
print(d_duplicate_swap)
d = {'a': 100, 'b': 20, 'c': 50, 'd': 100, 'e': 80}
max_d = max(d)
print(max_d)
min_d = min(d)
print(min_d)
max_v = max(d.values())
print(max_v)
min_v = min(d.values())
print(min_v)
max_k = max(d, key=d.get)
print(max_k)
min_k = min(d, key=d.get)
print(min_k)
max_kv = max
d.items()
key=lambda x: x[1]
print(max_kv)
print(type(max_kv))
max_k, max_v = max
d.items()
key=lambda x: x[1]
print(max_k)
print(max_v)
min_kv = min
d.items()
key=lambda x: x[1]
print(min_kv)
max_kv_list = [kv for kv in d.items() if kv[1] == max(d.values())]
print(max_kv_list)
max_k_list = [kv[0] for kv in d.items() if kv[1] == max(d.values())]
print(max_k_list)
min_kv_list = [kv for kv in d.items() if kv[1] == min(d.values())]
print(min_kv_list)
import pandas as pd
d = {'a': 100, 'b': 20, 'c': 50, 'd': 100, 'e': 80}
s = pd.Series(d)
print(s)
dtype: int64
print(s.index)
Index(['a', 'b', 'c', 'd', 'e'], dtype='object')
print(s.values)
print(s.max())
print(s.min())
print(max(s.index))
print(min(s.index))
print(s.idxmax())
print(s.idxmin())
print(s[s == s.max()])
dtype: int64
print(s[s == s.max()].index)
Index(['a', 'd'], dtype='object')
print(s[s == s.max()].index.tolist())
print(list(s[s == s.max()].index))
print(s[s == s.min()])
dtype: int64
print(s[s == s.min()].index)
Index(['b'], dtype='object')
print(s[s == s.min()].index.tolist())
print(list(s[s == s.min()].index))
print(s.sort_values())
dtype: int64
print(s[s > 60])
dtype: int64
d = {'one': 1, 'two': 2, 'three': 3}
print(d)
print(d.keys())
print(d.values())
print(d.items())
dict_keys(['one', 'two', 'three'])
dict_values([1, 2, 3])
dict_items([('one', 1), ('two', 2), ('three', 3)])
print(d.get('one'))
print(d.get('four'))
print(d.get('five', 'default_value'))
print(d)
print(d.setdefault('one'))
print(d.setdefault('four'))
print(d)
print(d.setdefault('five', 'default_value'))
print(d)
f = 987.6543
print(f)
print(type(f))
s = str(f)
print(s)
print(type(s))
print(s.split('.'))
print(type(s.split('.')))
print(type(s.split('.')[0]))
s_i, s_d = s.split('.')
print(s_i)
print(s_d)
print(len(s_i))
print(len(s_d))
print(len(str(f).split('.')[0]))
print(len(str(f).split('.')[1]))
print(str(0.123).split('.'))
print(len(str(0.123).split('.')[0]))
print(str(0.123).strip('0').split('.'))
print(len(str(0.123).strip('0').split('.')[0]))
print(s_i[-1])
print(s_i[-3])
print(s_d[0])
print(s_d[3])
print(type(s_d[3]))
print(int(s_d[3]))
print(type(int(s_d[3])))
print(str(f).split('.')[0][-3])
print(int(str(f).split('.')[0][-3]))
print(str(f).split('.')[1][3])
print(int(str(f).split('.')[1][3]))
print(str(0.0001))
print(str(0.00001))
s_format = format(0.00001, '.8f')
print(s_format)
print(type(s_format))
print('{:.8f}'.format(0.00001))
print(f'{0.00001:.8f}')
s_rstrip = s_format.rstrip('0')
print(s_rstrip)
print(format(0.1, '.8f').rstrip('0'))
print(format(0.0001, '.8f').rstrip('0'))
print(format(0.00000001, '.8f').rstrip('0'))
print(format(0.000000001, '.8f').rstrip('0'))
print(format(0.00001, '.8f').strip('0'))
s_i, s_d = format(0.00001, '.8f').strip('0').split('.')
print(s_i)
print(type(s_i))
print(len(s_i))
print(s_d)
s_i, s_d = format(1.00001, '.8f').strip('0').split('.')
print(s_i)
print(s_d)
i = 9876
print(i)
print(type(i))
s = str(i)
print(s)
print(type(s))
print(len(s))
print(type(len(s)))
print(len(str(i)))
i_minus = -9876
s_minus = str(i_minus)
print(s_minus)
-9876
print(len(s_minus))
print(abs(i_minus))
print(len(str(abs(i_minus))))
print(s[-1])
print(s[-3])
print(s[-10])
IndexError: string
print(type(s[-1]))
print(int(s[-1]))
print(str(i)[-1])
print(int(str(i)[-1]))
s = '9,675'
print(s.replace(',', ''))
print(type(s.replace(',', '')))
print(len(s.replace(',', '')))
print(s.replace(',', '')[-3])
import sys
print(dir(sys))
print(dir(int))
print(dir(str))
import pprint
print(type(dir(__builtins__)))
print(len(dir(__builtins__)))
pprint.pprint(dir(__builtins__), compact=True)
break
delattr
globals
print(dir(__builtins__)[0])
print(type(dir(__builtins__)[0]))
pprint.pprint([s for s in dir(__builtins__) if s.islower() and not s.startswith('_')], compact=True)
break
delattr
globals
pprint.pprint([s for s in dir(__builtins__) if s.endswith('Error')], compact=True)
pprint.pprint([s for s in dir(__builtins__) if s.endswith('Warning')], compact=True)
print('len' in dir(__builtins__))
q = 10 // 3
mod = 10 % 3
print(q, mod)
q, mod = divmod(10, 3)
print(q, mod)
answer = divmod(10, 3)
print(answer)
print(answer[0], answer[1])
def my_func():
print(my_func.__doc__)
print(type(my_func.__doc__))
help(my_func)
my_func()
def my_func2():
print(my_func2.__doc__)
def my_func_error():
a = 100
print(my_func_error.__doc__)
class MyClass
print(MyClass.__doc__)
def func_rest(param1, param2):
param1: Description
param1: int
param2: Description
param2: str
returns
return value
rtype: bool
return True
def func_numpy(param1, param2):
param1 : int
param2 : str
return value
return True
def func_google(param1, param2):
param1 (int)
param2 (str)
bool: Description
return value
return True
def add(a, b):
add(1, 2)
add(5, 10)
return a + b
if __name__ == '__main__':
import doctest
doctest.testmod()
def add(a, b):
add(1, 2)
add(5, 10)
return a * b
if __name__ == '__main__':
import doctest
doctest.testmod()
import doctest
doctest.testfile('doctest_text.txt')
def add(a, b):
add(1, 2)
add(5, 10)
return a + b
if __name__ == '__main__':
import doctest
doctest.testmod(verbose=True)
def add(a, b):
add(1, 2)
add(5, 10)
return a + b
if __name__ == '__main__':
import sys
result = add(int(sys.argv[1]), int(sys.argv[2]))
print(result)
print(Ellipsis)
print(type(Ellipsis))
import os
import shutil
os.makedirs('temp/dir/', exist_ok=True)
with open('file.txt', 'w') as f:
f.write('')
print(os.listdir('temp/'))
test.txt
target_dir = 'temp'
shutil.rmtree(target_dir)
os.mkdir(target_dir)
print(os.listdir('temp/'))
l = ['Alice', 'Bob', 'Charlie']
for name in l:
print(name)
for i, name in enumerate(l):
print(i, name)
for i, name in enumerate(l, 1):
print(i, name)
for i, name in enumerate(l, 42):
print(i, name)
for i, name in enumerate(l, 1):
print('{:03}_{}'.format(i, name))
step = 3
for i, name in enumerate(l):
print(i * step, name)
SyntaxError: unexpected
print('hello')
print(i)
SyntaxError: invalid
for i in range(3):
print(i)
n = 100
if n == 100:
print('n is 100')
print('n is not 100')
IndentationError: unindent
not match 
n = 100
if n == 100:
print('n is 100')
print('n is not 100')
import mathematics
ModuleNotFoundError: No
import math
print(math.pi)
import math.pi
ModuleNotFoundError: No
math.pi
not a 
from math import pi
print(pi)
from math import COS
ImportError: cannot
import name
from math import cos
print(cos(0))
import math
print(math.PI)
AttributeError: module
print(math.pi)
l = 100
l.append(200)
l = [100]
l.append(200)
print(l)
n = '100'
print(n + 200)
TypeError: can
str (not "int") 
n = 100
print(100 + 200)
print(float(['1.23E-3']))
TypeError: float
string or a
not 'list'
print(float('1.23E-3'))
print(float('float number'))
ValueError: could
not convert 
print(float('1.23E-3'))
n = 100
zero = 0
print(n / zero)
ZeroDivisionError: division
print(n // zero)
ZeroDivisionError: integer
division or modulo
print(n % zero)
ZeroDivisionError: integer
division or modulo
my_number = 100
print(myNumber)
NameError: name
not defined
print(my_number)
l = [0, 1, 2]
print(l[100])
IndexError: list
print(len(l))
print(l[1])
d = {'a': 1, 'b': 2, 'c': 3}
print(d['x'])
print(d['a'])
print(d)
print(list(d.keys()))
print(d.get('x'))
with open('not_exist_file.txt') as f:
print(f.read())
file or directory
not_exist_file.txt
import os
os.mkdir('data')
print(1 / 0)
ZeroDivisionError: division
print(1 / 0)
except ZeroDivisionError:
print('Error')
print(1 / 0)
except ZeroDivisionError as e:
print(e)
print(type(e))
print(issubclass(ZeroDivisionError, ArithmeticError))
print(1 / 0)
except ArithmeticError as e:
print(e)
print(type(e))
for i in [-2, -1, 0, 1, 2]:
print(1 / i)
except ZeroDivisionError as e:
print(e)
-0.5
-1.0
def divide(a, b):
print(a / b)
except ZeroDivisionError as e:
print('catch ZeroDivisionError:', e)
divide(1, 0)
ZeroDivisionError: division
divide('a', 'b')
TypeError: unsupported
type(s) 
def divide_each(a, b):
print(a / b)
except ZeroDivisionError as e:
print('catch ZeroDivisionError:', e)
except TypeError as e:
print('catch TypeError:', e)
divide_each(1, 0)
ZeroDivisionError: division
divide_each('a', 'b')
TypeError: unsupported
type(s) 
def divide_same(a, b):
print(a / b)
except (ZeroDivisionError, TypeError) 
print(e)
divide_same(1, 0)
divide_same('a', 'b')
type(s) 
def divide_wildcard(a, b):
print(a / b)
except:
print('Error')
divide_wildcard(1, 0)
divide_wildcard('a', 'b')
def divide_exception(a, b):
print(a / b)
except Exception as e:
print(e)
divide_exception(1, 0)
divide_exception('a', 'b')
type(s) 
def divide_else(a, b):
print(a / b)
except ZeroDivisionError as e:
print('catch ZeroDivisionError:', e)
print('finish (no error)')
divide_else(1, 2)
divide_else(1, 0)
ZeroDivisionError: division
def divide_finally(a, b):
print(a / b)
except ZeroDivisionError as e:
print('catch ZeroDivisionError:', e)
print('all finish')
divide_finally(1, 2)
divide_finally(1, 0)
ZeroDivisionError: division
def divide_else_finally(a, b):
print(a / b)
except ZeroDivisionError as e:
print('catch ZeroDivisionError:', e)
print('finish (no error)')
print('all finish')
divide_else_finally(1, 2)
divide_else_finally(1, 0)
ZeroDivisionError: division
def divide_pass(a, b):
print(a / b)
except ZeroDivisionError:
pass
divide_pass(1, 0)
a = 123
b = 'abc'
print('{} and {}'.format(a, b))
123 and abc
print('{first} and {second}'.format(first=a, second=b))
123 and abc
print(f'{a} and {b}')
123 and abc
print(F'{a} and {b}')
123 and abc
print(f"{a} and {b}")
123 and abc
print(f'''{a} and {b}''')
123 and abc
print(f"""{a} and {b}""")
123 and abc
s = 'abc'
print(f'right : {s:_>8}')
print(f'center: {s:_^8}')
print(f'left  : {s:_<8}')
right : _____abc
center: __abc___
left  : abc_____
i = 1234
print(f'zero padding: {i:08}')
print(f'comma: {i:,}')
print(f'bin: {i:b}')
print(f'oct: {i:o}')
print(f'hex: {i:x}')
print(f'bin: {i:#b}')
print(f'oct: {i:#o}')
print(f'hex: {i:#x}')
f = 12.3456
print(f'digit(decimal): {f:.3f}')
print(f'digit(all)    : {f:.3g}')
digit(decimal)
digit(all)    
print(f'exponent: {f:.3e}')
f = 0.123
print(f'percent: {f:.2%}')
n = 123
print(f'{{}}-{n}-{{{n}}}')
n = 123
i = 8
print('{n:0{i}}'.format(n=n, i=i))
print(f'{n:0{i}}')
f = 1.2345
for i in range(5):
print(f'{f:.{i}f}')
print('x\ty')
x = 'XXX'
y = 'YYY'
print(f'{x}\t{y}')
a = 3
b = 4
print('{a} + {b} = {a + b}'.format(a=a, b=b))
print(f'{a} + {b} = {a + b}')
print(f'{a} * {b} = {a * b}')
print(f'{a} / {b} = {a / b:.2e}')
d = {'key1': 3, 'key2': 4}
print('{0[key1]}, {0[key2]}'.format(d))
print('{0["key1"]}, {0["key2"]}'.format(d))
print(f'{d["key1"]}, {d["key2"]}')
print(f'{d[key1]}, {d[key2]}')
NameError: name
not defined
print(f'{d['key1']}, {d['key2']}')
SyntaxError: invalid
print(f"{d['key1']}, {d['key2']}")
SyntaxError: f
-string
i = 123
i=123
i = 123
i  =   123
i = 0b1111011
l = [0, 1, 2]
l[0] 
l = [0, 1, 2]
l[0] = 0
d = {'key1': 3, 'key2': 4}
d["key1"] 
d = {'key1': 3, 'key2': 4}
d["key1"] = 3
import feedparser
import pprint
import time
print(feedparser.__version__)
d_atom = feedparser.parse('http://gihyo.jp/feed/atom')
print(type(d_atom))
feedparser.FeedParserDict
pprint.pprint(d_atom, depth=1)
time.struct_time(tm_year=2018, tm_mon=6, tm_mday=30, tm_hour=7, tm_min=22, tm_sec=1, tm_wday=5, tm_yday=181, tm_isdst=0)
print(d_atom['encoding'])
print(d_atom.get('encoding'))
print(list(d_atom.keys()))
d_rss1 = feedparser.parse('http://gihyo.jp/feed/rss1')
print(type(d_rss1))
feedparser.FeedParserDict
pprint.pprint(d_rss1, depth=1)
time.struct_time(tm_year=2018, tm_mon=6, tm_mday=30, tm_hour=7, tm_min=22, tm_sec=1, tm_wday=5, tm_yday=181, tm_isdst=0)
d_rss2 = feedparser.parse('http://gihyo.jp/feed/rss2')
print(type(d_rss2))
feedparser.FeedParserDict
pprint.pprint(d_rss2, depth=1)
time.struct_time(tm_year=2018, tm_mon=6, tm_mday=30, tm_hour=7, tm_min=22, tm_sec=1, tm_wday=5, tm_yday=181, tm_isdst=0)
feed = feedparser.parse('http://gihyo.jp/feed/atom')['feed']
print(type(feed))
feedparser.FeedParserDict
pprint.pprint(feed)
gihyo.jp
gihyo.jp
gihyo.jp
gihyo.jp
gihyo.jp
gihyo.jp
gihyo.jp
time.struct_time(tm_year=2018, tm_mon=6, tm_mday=30, tm_hour=7, tm_min=22, tm_sec=1, tm_wday=5, tm_yday=181, tm_isdst=0)
print(feed['updated'])
print(type(feed['updated']))
t = feed['updated_parsed']
print(t)
time.struct_time(tm_year=2018, tm_mon=6, tm_mday=30, tm_hour=7, tm_min=22, tm_sec=1, tm_wday=5, tm_yday=181, tm_isdst=0)
print(type(t))
time.struct_time
print(t.tm_year)
print(t.tm_mon)
print(t.tm_mday)
print(time.strftime('%Y-%m-%d %H:%M:%S', t))
entries = feedparser.parse('http://gihyo.jp/feed/atom')['entries']
print(type(entries))
print(len(entries))
entry = entries[0]
print(type(entry))
feedparser.FeedParserDict
pprint.pprint(entry)
time.struct_time(tm_year=2018, tm_mon=6, tm_mday=29, tm_hour=6, tm_min=46, tm_sec=0, tm_wday=4, tm_yday=180, tm_isdst=0)
time.struct_time(tm_year=2018, tm_mon=6, tm_mday=29, tm_hour=6, tm_min=46, tm_sec=0, tm_wday=4, tm_yday=180, tm_isdst=0)
d = feedparser.parse('http://gihyo.jp/feed/atom')
urls = [entry['link'] for entry in d['entries']]
pprint.pprint(urls)
titles = [entry['title'] for entry in d['entries']]
pprint.pprint(titles)
dicts = [{'url': e['link'], 'title': e['title']} for e in d['entries']]
pprint.pprint(dicts)
print(dicts[0]['url'])
print(dicts[0]['title'])
print('\u3000' == '　')
title = d['entries'][0]['title']
print(repr(title))
print(title)
print(title.replace('\u3000', ' '))
titles_space = [entry['title'].replace('\u3000', ' ') for entry in d['entries']]
pprint.pprint(titles_space)
path = 'data/src/test.txt'
f = open(path)
print(type(f))
_io.TextIOWrapper
f.close()
with open(path) as f:
print(type(f))
_io.TextIOWrapper
with open('data/src/test_error.txt') as f:
print(type(f))
file or directory
with open(path) as f:
s = f.read()
print(type(s))
print(s)
with open(path) as f:
s = f.read()
print(s)
with open(path) as f:
l = f.readlines()
print(type(l))
print(l)
with open(path) as f:
l_strip = [s.strip() for s in f.readlines()]
print(l_strip)
with open(path) as f:
l = f.readlines()
print(l[1])
with open(path) as f:
for s_line in f:
print(s_line)
with open(path) as f:
s_line = f.readline()
print(s_line)
with open(path) as f:
while True:
s_line = f.readline()
print(s_line)
if not s_line:
break
path_w = 'data/src/test_w.txt'
s = 'New file'
with open(path_w, mode='w') as f:
f.write(s)
with open(path_w) as f:
print(f.read())
with open('data/src/new_dir/test_w.txt', mode='w') as f:
f.write(s)
file or directory
s = 'New line 1\nNew line 2\nNew line 3'
with open(path_w, mode='w') as f:
f.write(s)
with open(path_w) as f:
print(f.read())
l = ['One', 'Two', 'Three']
with open(path_w, mode='w') as f:
f.writelines(l)
with open(path_w) as f:
print(f.read())
with open(path_w, mode='w') as f:
f.write('\n'.join(l))
with open(path_w) as f:
print(f.read())
with open(path_w, mode='x') as f:
f.write(s)
with open(path_w, mode='x') as f:
f.write(s)
except FileExistsError:
pass
import os
if not os.path.isfile(path_w):
with open(path_w, mode='w') as f:
f.write(s)
with open(path_w, mode='a') as f:
f.write('Four')
with open(path_w) as f:
print(f.read())
with open(path_w, mode='a') as f:
f.write('\nFour')
with open(path_w) as f:
print(f.read())
with open(path_w, mode='r+') as f:
f.write('123456')
with open(path_w) as f:
print(f.read())
with open(path_w, mode='r+') as f:
f.seek(3)
f.write('-')
with open(path_w) as f:
print(f.read())
with open(path_w) as f:
l = f.readlines()
l.insert(0, 'FIRST\n')
with open(path_w, mode='w') as f:
f.writelines(l)
with open(path_w) as f:
print(f.read())
l = [-2, -1, 0, 1, 2]
print(filter(lambda x: x % 2 == 0, l))
print(type(filter(lambda x: x % 2 == 0, l)))
for i in filter(lambda x: x % 2 == 0, l):
print(i)
-2
print(list(filter(lambda x: x % 2 == 0, l)))
-2
print(list(filter(lambda x: x % 2 != 0, l)))
-1
l_s = ['apple', 'orange', 'strawberry']
print(list(filter(lambda x: x.endswith('e'), l_s)))
print(list(filter(lambda x: not x.endswith('e'), l_s)))
def is_even(x):
return x % 2 == 0
l = [-2, -1, 0, 1, 2]
print(list(filter(is_even, l)))
-2
l = [-2, -1, 0, 1, 2]
print(list(filter(lambda x: x % 2 == 0 and x > 0, l)))
print(list(filter(lambda x: x % 2 == 0 or x > 0, l)))
-2
l_b = [True, False]
print(list(filter(None, l_b)))
l = [-2, -1, 0, 1, 2]
print(list(filter(None, l)))
-2
-1
l_2d = [[0, 1, 2], [], [3, 4, 5]]
print(list(filter(None, l_2d)))
l_s = ['apple', '', 'orange', 'strawberry']
print(list(filter(None, l_s)))
import itertools
l = [-2, -1, 0, 1, 2]
print(list(itertools.filterfalse(lambda x: x % 2 == 0, l)))
-1
print(list(itertools.filterfalse(lambda x: x % 2 != 0, l)))
-2
l_s = ['apple', 'orange', 'strawberry']
print(list(itertools.filterfalse(lambda x: x.endswith('e'), l_s)))
l = [-2, -1, 0, 1, 2]
print(list(itertools.filterfalse(None, l)))
l = [-2, -1, 0, 1, 2]
print([x for x in l if x % 2 == 0])
-2
print([x for x in l if x % 2 != 0])
-1
l_s = ['apple', 'orange', 'strawberry']
print([x for x in l_s if x.endswith('e')])
print([x for x in l_s if not x.endswith('e')])
l = [-2, -1, 0, 1, 2]
print([x for x in l if x])
-2
-1
l_2d = [[0, 1, 2], [], [3, 4, 5]]
print([x for x in l_2d if x])
f = 256.0
print(f.hex())
print(type(f.hex()))
print(256.0.hex())
print(0.5.hex())
print(42.195.hex())
i = 256
print(i.hex())
s = '0x1.0000000000000p+8'
print(float.fromhex(s))
print(type(float.fromhex(s)))
print(float.fromhex('0x1p+8'))
print(float.fromhex('1p+8'))
print(float.fromhex('0x100'))
print(float.fromhex('100'))
print(float.fromhex('0xf2.f8p-10'))
print((15 * 16**1 + 2 * 16**0 + 15 * 16**-1 + 8 * 16**-2) * 2**-10)
import sys
f_max = sys.float_info.max
print(f_max)
print(f_max.hex())
print(float.fromhex('0x1.fffffffffffffp+1023'))
print(float.fromhex('0x1.0000000000000p+1024'))
OverflowError: hexadecimal
print(float.fromhex('0x2.0000000000000p+1023'))
OverflowError: hexadecimal
f_min = sys.float_info.min
print(f_min)
print(f_min.hex())
print(float.fromhex('0x1.0000000000000p-1022'))
print(float.fromhex('0x0.0000000000001p-1022'))
print(format(float.fromhex('0x0.0000000000001p-1022'), '.17'))
print(float.fromhex('0x0.0000000000001p-1023'))
import struct
import sys
f_max = sys.float_info.max
print(f_max)
print(struct.pack('>d', f_max))
print(type(struct.pack('>d', f_max)))
print(struct.pack('<d', f_max))
print(struct.unpack('>Q', struct.pack('>d', f_max)))
print(type(struct.unpack('>Q', struct.pack('>d', f_max))))
print(struct.unpack('>Q', struct.pack('>d', f_max))[0])
print(type(struct.unpack('>Q', struct.pack('>d', f_max))[0]))
print(struct.unpack('>d', struct.pack('>d', f_max))[0])
print(hex(struct.unpack('>Q', struct.pack('>d', f_max))[0]))
print(type(hex(struct.unpack('>Q', struct.pack('>d', f_max))[0])))
def double_to_hex(f):
return hex(struct.unpack('>Q', struct.pack('>d', f))[0])
print(double_to_hex(f_max))
print(double_to_hex(42.195))
print(double_to_hex(1e500))
print(double_to_hex(1e-500))
print(int(double_to_hex(f_max), 16))
print(bin(int(double_to_hex(f_max), 16)))
print(oct(int(double_to_hex(f_max), 16)))
def double_to_bin(f):
return bin(struct.unpack('>Q', struct.pack('>d', f))[0])
def double_to_oct(f):
return oct(struct.unpack('>Q', struct.pack('>d', f))[0])
print(double_to_bin(f_max))
print(double_to_oct(f_max))
def float_to_hex(f):
return hex(struct.unpack('>I', struct.pack('>f', f))[0])
print(float_to_hex(42.195))
names = ['Alice', 'Bob', 'Charlie']
ages = [24, 50, 18]
for i, (name, age) in enumerate(zip(names, ages)):
print(i, name, age)
for i, t in enumerate(zip(names, ages)):
print(i, t)
0 ('Alice', 24)
1 ('Bob', 50)
2 ('Charlie', 18)
for i, t in enumerate(zip(names, ages)):
print(i, t[0], t[1])
for i in range(3):
print(i)
print(range(3))
print(type(range(3)))
range(0, 3)
print(list(range(3)))
print(list(range(6)))
print(list(range(10, 13)))
print(list(range(0, 10, 3)))
print(list(range(10, 0, -3)))
for i in range(10, 0, -3):
print(i)
l = ['Alice', 'Bob', 'Charlie']
for name in reversed(l):
print(name)
for i in reversed(range(3)):
print(i)
for i in range(2, -1, -1):
print(i)
for i, name in reversed(enumerate(l)):
print(i, name)
for i, name in reversed(list(enumerate(l))):
print(i, name)
for i, name in enumerate(reversed(l)):
print(i, name)
l2 = [24, 50, 18]
for name, age in reversed(zip(l, l2)):
print(name, age)
for name, age in reversed(list(zip(l, l2))):
print(name, age)
l = ['Alice', 'Bob', 'Charlie']
for name in l:
print(name)
l = ['Alice', 'Bob', 'Charlie']
for name in l:
if name == 'Bob':
print('!!BREAK!!')
break
print(name)
l = ['Alice', 'Bob', 'Charlie']
for name in l:
if name == 'Bob':
print('!!SKIP!!')
continue
print(name)
l = ['Alice', 'Bob', 'Charlie']
for name in l:
print(name)
print('!!FINISH!!')
l = ['Alice', 'Bob', 'Charlie']
for name in l:
if name == 'Bob':
print('!!BREAK!!')
break
print(name)
print('!!FINISH!!')
l = ['Alice', 'Bob', 'Charlie']
for name in l:
if name == 'Bob':
print('!!SKIP!!')
continue
print(name)
print('!!FINISH!!')
l = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
for c in l[2:5]:
print(c)
for c in l[::2]:
print(c)
for c in l[1::2]:
print(c)
s = format(255, '04x')
print(s)
print(type(s))
print(format('center', '*^16'))
s = '{:04x}'.format(255)
print(s)
print(type(s))
print('{:*^16}'.format('center'))
print('{}-{}-{}'.format('100', '二百', 300))
-300
print('{0}-{1}-{0}'.format('foo', 'bar'))
print('{year}年{month}月{day}日'.format(year=2018, month=1, day=11))
l = ['one', 'two', 'three']
print('{0[0]}-{0[1]}-{0[2]}'.format(l))
d1 = {'name': 'Alice', 'age': 20}
d2 = {'name': 'Bob', 'age': 30}
print('{0[name]} is {0[age]} years old.\n{1[name]} is {1[age]} years old.'.format(d1, d2))
l = ['one', 'two', 'three']
print('{}-{}-{}'.format(*l))
d = {'name': 'Alice', 'age': 20}
print('{name} is {age} years old.'.format(**d))
print('{{}}-{num}-{{{num}}}'.format(num=100))
print('{num:x}'.format(num=255))
print('{year}年{month:02}月{day:02}日'.format(year=2018, month=1, day=11))
print('left  : {:<10}'.format(100))
print('center: {:^10}'.format(100))
print('right : {:>10}'.format(100))
print('left  : {:*<10}'.format(100))
print('center: {:a^10}'.format(100))
print('right : {:鬼>10}'.format(100))
center: aaa100aaaa
print('sign: {:0>10}'.format(-100))
print('sign: {:0=10}'.format(-100))
print('sign: {:0=+10}'.format(100))
-000000100
+000000100
print('sign: {:0=10}'.format('-100'))
not allowed in string 
print('sign: {:0=10}'.format(int('-100')))
-000000100
print('left  : {:*<10}'.format(1.23))
print('center: {:a^10}'.format(1.23))
print('right : {:鬼>10}'.format(1.23))
center: aaa1
print('sign: {:0>10}'.format(-1.23))
print('sign: {:0=10}'.format(-1.23))
print('sign: {:0=+10}'.format(1.23))
-000001.23
+000001.23
print('zero padding: {:0=10}'.format(100))
print('zero padding: {:010}'.format(100))
print('zero padding: {:0=10}'.format(-100))
print('zero padding: {:010}'.format(-100))
-000000100
-000000100
print('zero padding: {:010}'.format('-100'))
not allowed in string 
print('sign: {}'.format(100))
print('sign: {}'.format(-100))
-100
print('sign: {:+}'.format(100))
print('sign: {:+}'.format(-100))
+100
-100
print('sign: {: }'.format(100))
print('sign: {: }'.format(-100))
-100
print('sign: {:06}'.format(100))
print('sign: {:06}'.format(-100))
-00100
print('sign: {:+06}'.format(100))
print('sign: {:+06}'.format(-100))
+00100
-00100
print('sign: {: 06}'.format(100))
print('sign: {: 06}'.format(-100))
-00100
print('sign: {:_>6}'.format(100))
print('sign: {:_>6}'.format(-100))
sign: ___100
sign: __
-100
print('sign: {:_>+6}'.format(100))
print('sign: {:_>+6}'.format(-100))
sign: __
+100
sign: __
-100
print('sign: {:_> 6}'.format(100))
print('sign: {:_> 6}'.format(-100))
sign: __
sign: __
-100
print('{:,}'.format(100000000))
print('{:_}'.format(100000000))
print('{:,}'.format(1234.56789))
print('bin: {:b}'.format(255))
print('oct: {:o}'.format(255))
print('dec: {:d}'.format(255))
print('hex: {:x}'.format(255))
print('HEX: {:X}'.format(255))
hex: ff
HEX: FF
print('bin: {:08b}'.format(255))
print('oct: {:08o}'.format(255))
print('dec: {:08d}'.format(255))
print('hex: {:08x}'.format(255))
print('HEX: {:08X}'.format(255))
print('bin: {:#010b}'.format(255))
print('oct: {:#010o}'.format(255))
print('dec: {:#010d}'.format(255))
print('hex: {:#010x}'.format(255))
print('HEX: {:#010X}'.format(255))
print('hex: {:08x}'.format(255))
print('hex: {:09_x}'.format(255))
print('hex: {:#011_x}'.format(255))
print('hex: {:08x}'.format('255'))
ValueError: Unknown
print('hex: {:08x}'.format(int('255')))
print('{:.2f}'.format(123.456))
print('{:.5f}'.format(123.456))
print('{:.3f}'.format(0.0001234))
print('{:>12.5f}'.format(123.456))
print('{:012.5f}'.format(123.456))
print('{:06.5f}'.format(123.456))
print('{:.0f}'.format(0.4))
print('{:.0f}'.format(0.5))
print('{:.0f}'.format(0.6))
print('{}'.format(0.0001234))
print('{}'.format(0.00001234))
print('{}'.format(1234000000000000.0))
print('{}'.format(12340000000000000.0))
print('{}'.format(12340000000000000000000000))
print('{:e}'.format(0.0001234))
print('{:E}'.format(0.0001234))
print('{:.5e}'.format(0.0001234))
print('{:.2E}'.format(0.0001234))
print('{:.5e}'.format(987.65))
print('{:.2E}'.format(987.65))
print('{:>12.5e}'.format(987.65))
print('{:012.2E}'.format(987.65))
print('{:.2g}'.format(123.456))
print('{:.3g}'.format(123.456))
print('{:.8g}'.format(123.456))
print('{:.3g}'.format(0.0001234))
print('{:.2}'.format(123.456))
print('{:.3}'.format(123.456))
print('{:.8}'.format(123.456))
print('{:.3}'.format(0.0001234))
print('{:.3f}'.format(123.456))
print('{:.3e}'.format(123.456))
print('{:.3g}'.format(123.456))
print('{:.3}'.format(123.456))
print('{:.8f}'.format(123.456))
print('{:.8e}'.format(123.456))
print('{:.8g}'.format(123.456))
print('{:.8}'.format(123.456))
print('{:.4e}'.format(123.456))
print('{:.4e}'.format(0.000012345))
print('{:.4e}'.format(12))
print('{:%}'.format(0.12345))
print('{:.2%}'.format(0.12345))
print('{:%}'.format(10))
print('{:.2%}'.format(10))
print('{:>7.2%}'.format(0.12345))
print('{:07.2%}'.format(0.12345))
l = [0, 1]
print(type(l))
print('{:*^16}'.format(l))
TypeError: unsupported
pass
list.__format__
print(type(str(l)))
print('{:*^16}'.format(str(l)))
from fractions import Fraction
print(Fraction(1, 3))
print(Fraction(2, 6))
print(Fraction(3))
print(Fraction(0.25))
print(Fraction(0.33))
print(Fraction('2/5'))
print(Fraction('16/48'))
a = Fraction(1, 3)
print(a)
print(a.numerator)
print(type(a.numerator))
print(a.denominator)
print(type(a.denominator))
a.numerator = 7
AttributeError: can
result = Fraction(1, 6) ** 2 + Fraction(1, 3) / Fraction(1, 2)
print(result)
print(type(result))
fractions.Fraction
print(Fraction(7, 13) > Fraction(8, 15))
a_f = float(a)
print(a_f)
print(type(a_f))
b = a + 0.1
print(b)
print(type(b))
a_s = str(a)
print(a_s)
print(type(a_s))
pi = Fraction(3.14159265359)
print(pi)
print(pi.limit_denominator(10))
print(pi.limit_denominator(100))
print(pi.limit_denominator(1000))
e = Fraction(2.71828182846)
print(e)
print(e.limit_denominator(10))
print(e.limit_denominator(100))
print(e.limit_denominator(1000))
a = Fraction(0.565656565656)
print(a)
print(a.limit_denominator())
a = Fraction(0.3333)
print(a)
print(a.limit_denominator())
print(a.limit_denominator(100))
def func(x, y):
return x * y
print(func('abc', 3))
print(func(4, 3))
return x * y
print(func_annotations('abc', 3))
print(func_annotations(4, 3))
return x * y
print(func_annotations_default('abc'))
print(func_annotations_default(4))
def func_annotations_type(x: str, y: int) -> str:
return x * y
print(func_annotations_type('abc', 3))
print(func_annotations_type(4, 3))
return x * y
print(type(func_annotations.__annotations__))
print(func_annotations.__annotations__)
return': '
print(func_annotations.__annotations__['x'])
def hello():
print('Hello')
hello()
def add(a, b):
x = a + b
return x
x = add(3, 4)
print(x)
def func(a, b, c):
print(f'a={a}, b={b}, c={c}')
func(1, 10, 100)
a=1
b=10
c=100
func(1)
TypeError: func
func(1, 10, 100, 1000)
TypeError: func
func(b=10, c=100, a=1)
a=1
b=10
c=100
func(1, c=100, b=10)
a=1
b=10
c=100
func(a=1, 10, 100)
SyntaxError: positional
print(f'a={a}, b={b}, c={c}')
func_pos_only(a=1, b=10, c=100)
TypeError: func_pos_only
pass
func_pos_only(1, 10, 100)
a=1
b=10
c=100
func_pos_only(1, 10, c=100)
a=1
b=10
c=100
print(f'a={a}, b={b}, c={c}')
func_kw_only(1, 10, 100)
TypeError: func_kw_only
func_kw_only(1, 10, c=100)
a=1
b=10
c=100
func_kw_only(1, c=100, b=10)
a=1
b=10
c=100
print(f'a={a}, b={b}, c={c}')
func_pos_kw_only(1, 10, 100)
TypeError: func_pos_kw_only
func_pos_kw_only(a=1, b=10, c=100)
TypeError: func_pos_kw_only
pass
func_pos_kw_only(1, 10, c=100)
a=1
b=10
c=100
func_pos_kw_only(1, c=100, b=10)
a=1
b=10
c=100
print(f'a={a}, b={b}, c={c}')
SyntaxError: invalid
def func_default(a, b, c=100):
print(f'a={a}, b={b}, c={c}')
func_default(1, 10)
a=1
b=10
c=100
func_default(1, 10, 200)
a=1
b=10
c=200
def func_default(a=1, b, c=100):
print(f'a={a}, b={b}, c={c}')
SyntaxError: non
-default
def func_args(*args):
print(args)
func_args(1, 10)
func_args(1, 10, 100, 1000)
def func_kwargs(**kwargs):
print(kwargs)
func_kwargs(a=1, b=10)
func_kwargs(c=1, b=10, d=1000, a=100)
def func(a, b, c):
print(f'a={a}, b={b}, c={c}')
l = [1, 10, 100]
func(*l)
a=1
b=10
c=100
l = [1, 10]
func(*l)
TypeError: func
d = {'a': 1, 'b': 10, 'c': 100}
func(**d)
a=1
b=10
c=100
d = {'a': 1, 'b': 10, 'x': 100}
func(**d)
TypeError: func
def func_return(a, b):
return a + b
x = func_return(3, 4)
print(x)
print(type(x))
x = func_return(0.3, 0.4)
print(x)
print(type(x))
def func_none():
pass
x = func_none()
print(x)
def func_none2():
return
x = func_none2()
print(x)
def func_none3():
return None
x = func_none3()
print(x)
def func_return_multi(a, b):
return a + b, a * b, a / b
x = func_return_multi(3, 4)
print(x)
print(type(x))
x, y, z = func_return_multi(3, 4)
print(x)
print(y)
print(z)
import math
print(math.gcd(6, 4))
print(math.lcm(6, 4))
def my_lcm(x, y):
return (x * y) // math.gcd(x, y)
print(my_lcm(6, 4))
import math
print(math.gcd(27, 18, 9))
print(math.gcd(27, 18, 9, 3))
print(math.lcm(27, 9, 3))
print(math.lcm(27, 18, 9, 3))
l = [27, 18, 9, 3]
print(math.gcd(*l))
print(math.lcm(*l))
from functools import reduce
def my_gcd(*numbers):
return reduce(math.gcd, numbers)
print(my_gcd(27, 18, 9))
print(my_gcd(27, 18, 9, 3))
l = [27, 18, 9, 3]
print(my_gcd(*l))
def my_lcm_base(x, y):
return (x * y) // math.gcd(x, y)
def my_lcm(*numbers):
return reduce(my_lcm_base, numbers, 1)
print(my_lcm(27, 9, 3))
print(my_lcm(27, 18, 9, 3))
l = [27, 18, 9, 3]
print(my_lcm(*l))
l = [i**2 for i in range(5)]
print(l)
print(type(l))
g = (i**2 for i in range(5))
print(g)
print(type(g))
for i in g:
print(i)
col == row
print(type(g_cells))
for i in g_cells:
print(i)
print(sum([i**2 for i in range(5)]))
print(sum((i**2 for i in range(5))))
print(sum(i**2 for i in range(5)))
t = tuple(i**2 for i in range(5))
print(t)
print(type(t))
import glob
print(glob.glob('temp/*.txt'))
print(glob.glob('temp/**', recursive=True))
print(glob.glob('temp/**/*.txt', recursive=True))
print(glob.glob('temp/**/*.text', recursive=True))
print(glob.glob('temp/**/???.text', recursive=True))
print(glob.glob('temp/**/[0-9][0-9][0-9].txt', recursive=True))
import util_make_files
util_make_files.glob_example_detail()
import glob
import re
import os
l = glob.glob('temp/*.txt')
print(l)
print(type(l))
print(glob.glob('temp/*'))
print(glob.glob('temp/*.txt'))
print(glob.glob('temp/dir/*/*.text'))
print(glob.glob('temp/???.*'))
print(glob.glob('temp/[0-9].*'))
print(glob.glob('temp/[0-9][0-9].*'))
print(glob.glob('temp/[a-z][a-z][a-z].*'))
print(glob.glob('temp/[[]*'))
print(glob.glob('temp/dir/*/*.text'))
print(glob.glob('temp/**/*.text', recursive=True))
print(glob.glob('temp/**', recursive=True))
p for p in glob.glob('temp/**', recursive=True)
os.path.isfile(p)
os.path.basename(p) for p in glob.glob('temp/**', recursive=True)
os.path.isfile(p)
aaa.text
ccc.text
bbb.txt
ddd.text
print(glob.glob('temp/**/', recursive=True))
print(os.path.join('temp', '**' + os.sep))
print(glob.glob(os.path.join('temp', '**' + os.sep), recursive=True))
os.path.basename(p.rstrip(os.sep)) 
glob.glob(os.path.join('temp', '**' + os.sep), recursive=True)
p for p in glob.glob('temp/**', recursive=True)
re.search
p for p in glob.glob('temp/**', recursive=True)
re.search
print(type(glob.iglob('temp/*.txt')))
for p in glob.iglob('temp/*.txt'):
print(p)
import shutil
shutil.rmtree('temp')
os.mkdir('temp')
import os
os.makedirs('temp/dir', exist_ok=True)
with open('temp/012.txt', 'w') as f:
pass
with open('temp/abc.txt', 'w') as f:
pass
with open('temp/file.text', 'w') as f:
pass
with open('temp/dir/789.txt', 'w') as f:
pass
with open('temp/dir/xyz.text', 'w') as f:
pass
import glob
import re
print(glob.glob('temp/**', recursive=True))
p for p in glob.glob('temp/**', recursive=True) if re.search
p for p in glob.glob('temp/**', recursive=True) if re.search
path = 'data/src/sample_for_grep.txt'
with open(path) as f:
print(f.read())
with open(path) as f:
lines = f.readlines()
print(lines)
print(type(lines))
lines_strip = [line.strip() for line in lines]
print(lines_strip)
l_XXX = [line for line in lines_strip if 'XXX' in line]
print(l_XXX)
for line in l_XXX:
print(line)
print(l_XXX[0])
print(l_XXX[-1])
l_XXX_start = [line for line in lines_strip if line.startswith('XXX')]
print(l_XXX_start)
l_XXX_ZZZ_and = [line for line in lines_strip if ('XXX' in line) and ('ZZZ' in line)]
print(l_XXX_ZZZ_and)
l_XXX_xxx = [line for line in lines_strip if 'xxx' in line.lower()]
print(l_XXX_xxx)
l_XXX_i = [i for i, line in enumerate(lines_strip) if 'XXX' in line]
print(l_XXX_i)
l_XXX_both = [(i, line) for i, line in enumerate(lines_strip) if 'XXX' in line]
print(l_XXX_both)
l_i, l_str = list(zip(*l_XXX_both))
print(l_i)
print(l_str)
with open(path) as f:
for i, line in enumerate(f):
if 'aaa' in line:
break
print(i)
print(line)
with open(path) as f:
for i, line in enumerate(f):
if line == 'ZZZ XXX\n':
break
print(i)
print(line)
import requests
url = 'https://note.nkmk.me'
hb_count = 'http://api.b.st-hatena.com/entry.count'
r = requests.get(hb_count, params={'url': url})
print(r.url)
url=https%3
Fnote.nkmk.me
print(r.text)
print(type(r.text))
print(int(r.text))
print(type(int(r.text)))
hb_counts = 'http://api.b.st-hatena.com/entry.counts'
r = requests.get(hb_counts, params={'url': ['https://www.google.co.jp', 'https://www.yahoo.co.jp']})
print(r.url)
url=https%3
Fwww.yahoo.co.jp
j = r.json()
print(j)
www.google.co.jp
www.yahoo.co.jp
print(type(j))
hb_total_count = 'http://api.b.st-hatena.com/entry.total_count'
r = requests.get(hb_total_count, params={'url': url})
print(r.url)
url=https%3
Fnote.nkmk.me
j = r.json()
print(j)
note.nkmk.me
print(j['total_bookmarks'])
hb_entry = 'http://b.hatena.ne.jp/entry/jsonlite/'
r = requests.get(hb_entry, params={'url': url})
print(r.url)
url=https%3
Fnote.nkmk.me
j = r.json()
import pprint
pprint.pprint(j)
note.nkmk.me
print(type(j['bookmarks']))
print(type(j['bookmarks'][0]))
for b in j['bookmarks']:
print(b['timestamp'])
import heapq
l = [3, 6, 7, -1, 23, -10, 18]
print(heapq.nlargest(3, l))
print(heapq.nsmallest(3, l))
-10
-1
print(l)
-1
-10
def func():
print('Hello!')
print('__name__ is', __name__)
print('Hello!')
def func():
print('Hello!')
print('__name__ is', __name__)
if __name__ == '__main__':
func()
def main():
print('Hello!')
if __name__ == '__main__':
main()
import struct
import binascii
f_max_s = '7fefffffffffffff'
print(binascii.unhexlify(f_max_s))
print(type(binascii.unhexlify(f_max_s)))
print(struct.unpack('>d', binascii.unhexlify(f_max_s)))
print(struct.unpack('>d', binascii.unhexlify(f_max_s))[0])
print(type(struct.unpack('>d', binascii.unhexlify(f_max_s))[0]))
def hex_to_double(s):
if s.startswith('0x'):
s = s[2:]
s = s.replace(' ', '')
return struct.unpack('>d', binascii.unhexlify(s))[0]
print(hex_to_double('7fefffffffffffff'))
print(hex_to_double('0x7fefffffffffffff'))
print(hex_to_double('0x7fef ffff ffff ffff'))
print(hex_to_double('0x4045 18f5 c28f 5c29'))
print(hex_to_double('7ff0000000000000'))
print(hex_to_double('7ff0000000000001'))
print(hex_to_double('0000000000000001'))
print(hex_to_double('ffff ffff ffff ffff ff'))
error: unpack
print(hex_to_double('ffff ffff ffff ff'))
error: unpack
def hex_to_float(s):
if s.startswith('0x'):
s = s[2:]
s = s.replace(' ', '')
return struct.unpack('>f', binascii.unhexlify(s))[0]
print(hex_to_float('0x4228c7ae'))
AbcDef_123 = 100
print(AbcDef_123)
SyntaxError: can
abc = 100
SyntaxError: invalid
_abc = 100
print(_abc)
SyntaxError: invalid
SyntaxError: invalid
ABC = -100
-100
-100
print('AbcDef_123'.isidentifier())
print('AbcDef-123'.isidentifier())
print('変数その１'.isidentifier())
print('☺'.isidentifier())
print('None'.isidentifier())
None = 100
SyntaxError: can
print(len)
print(len('abc'))
len = 100
print(len)
print(len('abc'))
def if_test(num):
if num > 100:
print('100 < num')
elif num > 50:
print('50 < num <= 100')
elif num > 0:
print('0 < num <= 50')
elif num == 0:
print('num == 0')
print('num < 0')
if_test(1000)
if_test(70)
if_test(0)
num == 0
if_test(-100)
def if_test2(num):
if 50 < num < 100:
print('50 < num < 100')
print('num <= 50 or num >= 100')
if_test2(70)
if_test2(0)
num <= 50 or num >= 100
def if_test_in(s):
if 'a' in s:
print('a is in string')
print('a is NOT in string')
if_test_in('apple')
if_test_in('melon')
if 10:
print('True')
if [0, 1, 2]:
print('True')
print(bool(10))
print(bool(0.0))
print(bool([]))
print(bool('False'))
def if_test_list(l):
if l:
print('list is NOT empty')
print('list is empty')
if_test_list([0, 1, 2])
if_test_list([])
def if_test_and_not(num):
if num >= 0 and not num % 2 == 0:
print('num is positive odd')
print('num is NOT positive odd')
if_test_and_not(5)
if_test_and_not(10)
if_test_and_not(-10)
def if_test_and_not_or(num):
if num >= 0 and not num % 2 == 0 or num == -10:
print('num is positive odd or -10')
print('num is NOT positive odd or -10')
if_test_and_not_or(5)
odd or -10
if_test_and_not_or(10)
odd or -10
if_test_and_not_or(-10)
odd or -10
def if_test_and_backslash(num):
not num % 2 == 0
print('num is positive odd')
print('num is NOT positive odd')
if_test_and_backslash(5)
def if_test_and_brackets(num):
not num % 2 == 0
print('num is positive odd')
print('num is NOT positive odd')
if_test_and_brackets(5)
import math
print(type(math))
print(math)
print(math.radians(180))
print(type(math.radians))
print(math.pi)
print(type(math.pi))
import math as m
print(m.pi)
print(math.pi)
NameError: name
not defined
from math import pi 
print(PI)
print(pi)
NameError: name
not defined
from math import pi
print(pi)
print(math.radians(180))
NameError: name
not defined
from math import pi, radians
print(pi)
print(radians(180))
print(e)
print(exp(1))
from math import *
print(pi)
print(cos(0))
print(sin(0))
import collections
print(collections)
print(collections.Counter)
collections.Counter
import collections.Counter
ModuleNotFoundError: No
collections.Counter
from collections import Counter
print(Counter)
collections.Counter
import urllib
print(type(urllib))
print(urllib)
print(urllib.error)
AttributeError: module
import urllib.error
print(urllib.error)
urllib.error
print(urllib.error.HTTPError)
urllib.error.HTTPError
from urllib import error
print(error)
urllib.error
print(error.HTTPError)
urllib.error.HTTPError
from urllib.error import HTTPError
print(HTTPError)
urllib.error.HTTPError
import os
import os
import sys
import math
import os
import sys
import Requests
import my_package1
import my_package2
from my_package import mod2
from my_package.sub_package2 import sub_mod2
mod2.func_same()
mod2.func_sub()
sub_mod2.func_parent()
sub_mod2.func_parent_sub()
from my_package.mod2 import func_same, func_sub
from my_package.sub_package2.sub_mod2 import func_parent, func_parent_sub
func_same()
func_sub()
func_parent()
func_parent_sub()
print(1 in [0, 1, 2])
print(100 in [0, 1, 2])
print(1 in (0, 1, 2))
print(1 in {0, 1, 2})
print(1 in range(3))
l = [0, 1, 2]
i = 0
if i in l:
print('{} is a member of {}.'.format(i, l))
print('{} is not a member of {}.'.format(i, l))
l = [0, 1, 2]
i = 100
if i in l:
print('{} is a member of {}.'.format(i, l))
print('{} is not a member of {}.'.format(i, l))
l = [0, 1, 2]
if l:
print('not empty')
print('empty')
not empty
l = []
if l:
print('not empty')
print('empty')
d = {'key1': 'value1', 'key2': 'value2', 'key3': 'value3'}
print('key1' in d)
print('value1' in d)
print('value1' in d.values())
print(('key1', 'value1') in d.items())
print(('key1', 'value2') in d.items())
print('a' in 'abc')
print('x' in 'abc')
print('ab' in 'abc')
print('ac' in 'abc')
print(10 in [1, 2, 3])
print(10 not in [1, 2, 3])
print(not 10 in [1, 2, 3])
print(not (10 in [1, 2, 3]))
print((not 10) in [1, 2, 3])
print(not 10)
print(False in [1, 2, 3])
print([0, 1] in [0, 1, 2])
print([0, 1] in [[0, 1], [1, 0]])
l = [0, 1, 2]
v1 = 0
v2 = 100
print(v1 in l and v2 in l)
print(v1 in l or v2 in l)
print((v1 in l) or (v2 in l))
l1 = [0, 1, 2, 3, 4]
l2 = [0, 1, 2]
l3 = [0, 1, 5]
l4 = [5, 6, 7]
print(set(l2) <= set(l1))
print(set(l3) <= set(l1))
print(set(l1).isdisjoint(set(l4)))
print(not set(l1).isdisjoint(set(l3)))
l = [0, 1, 2]
for i in l:
print(i)
print([i * 10 for i in l])
n_small = 10
n_large = 10000
l_small = list(range(n_small))
l_large = list(range(n_large))
std. dev. of
std. dev. of
std. dev. of
std. dev. of
std. dev. of
s_small = set(l_small)
s_large = set(l_large)
std. dev. of
std. dev. of
std. dev. of
std. dev. of
std. dev. of
for i in range(n_large):
std. dev. of
s_large_ = set(l_large)
for i in range(n_large):
std. dev. of
d = dict(zip(l_large, l_large))
print(len(d))
print(d[0])
print(d[9999])
for i in range(n_large):
std. dev. of
dv = d.values()
for i in range(n_large):
std. dev. of
di = d.items()
for i in range(n_large):
std. dev. of
print(float('inf') + 100)
print(float('inf') + float('inf'))
print(float('inf') - 100)
print(float('inf') - float('inf'))
print(type(float('inf') - float('inf')))
print(float('inf') * 2)
print(float('inf') * float('inf'))
print(float('inf') * 0)
print(float('inf') / 2)
print(float('inf') / float('inf'))
print(0 / float('inf'))
print(float('inf') / 0)
ZeroDivisionError: float
print(float('inf') ** 2)
print(float('inf') ** float('inf'))
print(float('inf') ** 0)
print(2 ** float('inf'))
print(1 ** float('inf'))
print(0 ** float('inf'))
import math
import numpy as np
print(1e1000)
print(1e100)
print(1e1000 == float('inf'))
print(1e100 == float('inf'))
print(float('inf') == math.inf == np.inf)
print(1e1000 == math.inf)
print(1e100 == math.inf)
print(float('inf') == float('inf') * 100)
print(math.isinf(1e1000))
print(math.isinf(1e100))
print(math.isinf(-1e1000))
print(np.isinf(1e1000))
print(np.isinf(1e100))
print(np.isinf(-1e1000))
a_inf = np.array([1, np.inf, 3, -np.inf])
print(a_inf)
print(type(a_inf))
numpy.ndarray
print(np.isinf(a_inf))
a_inf[np.isinf(a_inf)] = 0
print(a_inf)
import sys
print(sys.float_info.max)
print(float('inf') > sys.float_info.max)
print(-float('inf') < -sys.float_info.max)
print(float('nan'))
print(type(float('nan')))
print(float('inf') > float('nan'))
print(float('inf') < float('nan'))
print(float('inf') == float('nan'))
print(float('inf') > 100)
large_int = int(sys.float_info.max) * 10
print(large_int)
print(type(large_int))
print(large_int > sys.float_info.max)
print(float('inf') > large_int)
print(float(10**308))
print(float(10**309))
OverflowError: int
f_inf = float('inf')
print(f_inf)
print(type(f_inf))
f_inf_minus = -float('inf')
print(f_inf_minus)
-inf
print(type(f_inf_minus))
print(int(f_inf))
OverflowError: cannot
print(str(f_inf))
print(type(str(f_inf)))
print(float('inf'))
print(float('infinity'))
print(float('INF'))
print(float('INFinity'))
import sys
f_inf_num = sys.float_info.max * 2
print(f_inf_num)
import math
print(math.inf)
print(type(math.inf))
print(float('inf') == math.inf)
import numpy as np
print(np.inf)
print(type(np.inf))
print(float('inf') == np.inf)
val = input()
print(val)
print(type(val))
val = input('Enter your name: ')
name: Alice
print(val)
print(type(val))
val = input('Enter number: ')
print(val)
print(type(val))
i = int(val)
print(i)
print(type(i))
f = float(val)
print(f)
print(type(f))
val = input('Enter number: ')
number: abc
print(val)
i = int(val)
ValueError: invalid
int() 
i = int(val)
except ValueError:
i = 0
print(i)
val_1 = input('Enter 1st value: ')
val_2 = input('Enter 2nd value: ')
val_3 = input('Enter 3rd value: ')
value: x
value: y
value: z
print(val_1)
print(val_2)
print(val_3)
l = []
print('Enter "over" then finish')
while True:
val = input('Enter value: ')
if val == 'over':
print('FINISH')
break
l.append(val)
value: x
value: y
value: z
value: over
print(l)
l = list(iter(input, 'over'))
print(l)
l = list(iter(lambda: input('Enter value: '), 'over'))
value: x
value: y
value: z
value: over
print(l)
s = '\n'.join(iter(input, ''))
print(s)
print(type(s))
val = input('Enter values separated by comma: ')
comma: x
print(val)
l = val.split(',')
print(l)
print(type(l))
val = input('Enter your name: ')
print('You are', val)
print(int(10.123))
print(int(10.987))
print(int(10))
print(type(int(10.123)))
print(int(-10.123))
-10
print(int(-10.987))
-10
print(int('10'))
print(int('10.123'))
ValueError: invalid
int() 
print(int('FF', 16))
i = 10
print(type(i))
f = 10.0
print(type(f))
print(i == f)
print(i is f)
print(isinstance('string', str))
print(isinstance(100, str))
print(isinstance(100, (int, str)))
def is_str(v):
return isinstance(v, str)
print(is_str('string'))
print(is_str(100))
print(is_str([0, 1, 2]))
def is_str_or_int(v):
return isinstance(v, (int, str))
print(is_str_or_int('string'))
print(is_str_or_int(100))
print(is_str_or_int([0, 1, 2]))
def type_condition(v):
if isinstance(v, str):
print('type is str')
elif isinstance(v, int):
print('type is int')
print('type is not str or int')
type_condition('string')
type_condition(100)
type_condition([0, 1, 2])
type is not str or int
class Base
pass
class Derive(Base
pass
base = Base()
print(type(base))
__main__.Base
derive = Derive()
print(type(derive))
__main__.Derive
print(type(derive) is Derive)
print(type(derive) is Base)
print(isinstance(derive, Derive))
print(isinstance(derive, Base))
print(type(True))
print(type(True) is bool)
print(type(True) is int)
print(isinstance(True, bool))
print(isinstance(True, int))
print(issubclass(bool, int))
print(issubclass(bool, float))
print(issubclass(bool, bool))
print(issubclass(bool, (int, float)))
print(issubclass(bool, (str, float)))
print(issubclass(ZeroDivisionError, ArithmeticError))
print(issubclass(ZeroDivisionError, Exception))
print(issubclass(ZeroDivisionError, BaseException))
print(issubclass(True, int))
TypeError: issubclass
print(issubclass(type(True), int))
print(isinstance(True, bool))
print(isinstance(True, int))
print(isinstance(True, float))
print(isinstance(True, (int, float)))
import itertools
import operator
l = [1, 2, 3, 4, 5, 6]
print(itertools.accumulate(l))
itertools.accumulate
print(type(itertools.accumulate(l)))
itertools.accumulate
for i in itertools.accumulate(l):
print(i)
print(list(itertools.accumulate(l)))
print([0] + list(itertools.accumulate(l)))
print(list(itertools.accumulate(reversed(l))))
print(operator.mul(2, 3))
print(list(itertools.accumulate(l, func=operator.mul)))
print(list(itertools.accumulate(l, func=operator.sub)))
-1
-4
-8
-13
-19
print(list(itertools.accumulate(l, func=operator.truediv)))
print(list(itertools.accumulate([1, 3, 2, 6, 5, 4], func=max)))
itertools.accumulate
l, func=lambda x, y: x * y
itertools.accumulate
l, func=lambda x, y: int(str(x) + str(y))
print(list(itertools.accumulate(l)))
print(list(itertools.accumulate(l, initial=0)))
print(list(itertools.accumulate(l, initial=100)))
import itertools
for i in itertools.count():
print(i)
if i > 3:
break
for i in itertools.count(2):
print(i)
if i > 3:
break
for i in itertools.count(step=3):
print(i)
if i > 8:
break
for i in itertools.count(2, 3):
print(i)
if i > 8:
break
for i in itertools.count(10, -1):
print(i)
if i < 8:
break
for i in itertools.count(0.1, 1.5):
print(i)
if i > 3:
break
for i in itertools.count():
ii = 0.1 + 1.5 * i
print(ii)
if ii > 3:
break
l1 = ['a', 'b', 'c']
l2 = ['x', 'y', 'z']
print(list(zip(itertools.count(), l1, l2)))
print(list(enumerate(zip(l1, l2))))
import itertools
l = [1, 10, 100]
sum_value = 0
for i in itertools.cycle(l):
print(i)
sum_value += i
if sum_value > 300:
break
sum_value = 0
for i in itertools.cycle(range(3)):
print(i)
sum_value += i
if sum_value > 5:
break
l1 = [1, 10, 100]
l2 = [0, 1, 2, 3, 4, 5, 6]
print(list(zip(itertools.cycle(l1), l2)))
import itertools
l = ['a', 'b', 'c', 'd']
p = itertools.permutations(l, 2)
print(type(p))
itertools.permutations
for v in itertools.permutations(l, 2):
print(v)
p_list = list(itertools.permutations(l, 2))
print(p_list)
print(len(p_list))
for v in itertools.permutations(l):
print(v)
print(len(list(itertools.permutations(l))))
l = ['a', 'a']
for v in itertools.permutations(l, 2):
print(v)
l = ['a', 'b', 'c', 'd']
c = itertools.combinations(l, 2)
print(type(c))
itertools.combinations
for v in itertools.combinations(l, 2):
print(v)
c_list = list(itertools.combinations(l, 2))
print(c_list)
print(len(c_list))
h = itertools.combinations_with_replacement(l, 2)
print(type(h))
itertools.combinations_with_replacement
for v in itertools.combinations_with_replacement(l, 2):
print(v)
h_list = list(itertools.combinations_with_replacement(l, 2))
print(h_list)
print(len(h_list))
s = 'arc'
for v in itertools.permutations(s):
print(v)
anagram_list = [''.join(v) for v in itertools.permutations(s)]
print(anagram_list)
import itertools
import pprint
l1 = ['a', 'b', 'c']
l2 = ['X', 'Y', 'Z']
p = itertools.product(l1, l2)
print(p)
itertools.product
print(type(p))
itertools.product
for v in p:
print(v)
for v in p:
print(v)
for v1, v2 in itertools.product(l1, l2):
print(v1, v2)
for v1 in l1:
for v2 in l2:
print(v1, v2)
l_p = list(itertools.product(l1, l2))
pprint.pprint(l_p)
print(type(l_p))
print(type(l_p[0]))
t = ('one', 'two')
d = {'key1': 'value1', 'key2': 'value2'}
r = range(2)
l_p = list(itertools.product(t, d, r))
pprint.pprint(l_p)
l1 = ['a', 'b']
pprint.pprint(list(itertools.product(l1, repeat=3)))
pprint.pprint(list(itertools.product(l1, l1, l1)))
l1 = ['a', 'b']
l2 = ['X', 'Y']
pprint.pprint(list(itertools.product(l1, l2, repeat=2)))
pprint.pprint(list(itertools.product(l1, l2, l1, l2)))
import itertools
A = range(1000)
for x in itertools.product(A, A):
pass
std. dev. of
for a1, a2 in itertools.product(A, A):
pass
std. dev. of
for a1 in A:
for a2 in A:
pass
std. dev. of
for x in ((a1, a2) for a1 in A for a2 in A):
pass
std. dev. of
for a1, a2 in ((a1, a2) for a1 in A for a2 in A):
pass
std. dev. of
v = 0
for a1, a2 in itertools.product(A, A):
v += a1 * a2
std. dev. of
v = 0
for a1 in A:
for a2 in A:
v += a1 * a2
std. dev. of
v = sum(a1 * a2 for a1, a2 in itertools.product(A, A))
std. dev. of
v = sum(a1 * a2 for a1 in A for a2 in A)
std. dev. of
B = range(100)
for x in itertools.product(B, B, B):
pass
std. dev. of
for b1, b2, b3 in itertools.product(B, B, B):
pass
std. dev. of
for b1 in B:
for b2 in B:
for b3 in B:
pass
std. dev. of
for x in ((b1, b2, b3) for b1 in B for b2 in B for b3 in B):
pass
std. dev. of
for b1, b2, b3 in ((b1, b2, b3) for b1 in B for b2 in B for b3 in B):
pass
std. dev. of
import itertools
sum_value = 0
for i in itertools.repeat(10):
print(i)
sum_value += i
if sum_value > 40:
break
for i in itertools.repeat(10, 3):
print(i)
for l in itertools.repeat([0, 1, 2], 3):
print(l)
for func in itertools.repeat(len, 3):
print(func('abc'))
l = [0, 1, 2, 3]
print(list(zip(itertools.repeat(10), l)))
from janome.tokenizer import Tokenizer
from janome.analyzer import Analyzer
from janome.charfilter import *
from janome.tokenfilter import *
t = Tokenizer()
s = '<div>PythonとＰＹＴＨＯＮとパイソンとﾊﾟｲｿﾝ</div>'
for token in t.tokenize(s):
print(token)
UnicodeNormalizeCharFilter()
RegexReplaceCharFilter('<.*?>', '')
POSKeepFilter(['名詞'])
LowerCaseFilter()
ExtractAttributeFilter('surface')
a = Analyzer(char_filters=char_filters, token_filters=token_filters)
for token in a.analyze(s):
print(token)
s = '自然言語処理による日本国憲法の形態素解析'
for token in t.tokenize(s):
print(token)
a = Analyzer(token_filters=[CompoundNounFilter()])
for token in a.analyze(s):
print(token)
s = '人民の人民による人民のための政治'
a = Analyzer(token_filters=[POSKeepFilter(['名詞']), TokenCountFilter()])
g_count = a.analyze(s)
print(type(g_count))
for i in g_count:
print(i)
l_count = list(a.analyze(s))
print(type(l_count))
print(l_count)
d_count = dict(a.analyze(s))
print(type(d_count))
print(d_count)
print(d_count['人民'])
print(d_count['国民'])
print(d_count.get('国民', 0))
s = '走れと言われたので走ると言った'
a = Analyzer(token_filters=[TokenCountFilter()])
print(list(a.analyze(s)))
a = Analyzer(token_filters=[TokenCountFilter(att='base_form')])
print(list(a.analyze(s)))
a = Analyzer(token_filters=[TokenCountFilter(att='part_of_speech')])
print(list(a.analyze(s)))
s = '吾輩は猫である'
a = Analyzer(token_filters=[POSKeepFilter('助動詞')])
for token in a.analyze(s):
print(token)
a = Analyzer(token_filters=[POSKeepFilter(['助動詞'])])
for token in a.analyze(s):
print(token)
from janome.tokenizer import Tokenizer
import collections
t = Tokenizer()
s = '人民の人民による人民のための政治'
for token in t.tokenize(s):
print(token)
c = collections.Counter(t.tokenize(s, wakati=True))
print(type(c))
collections.Counter
print(c)
Counter({'人民': 3, 'の': 3, 'による': 1, 'ため': 1, '政治': 1})
print(c['人民'])
print(c['国民'])
mc = c.most_common()
print(mc)
print(mc[0][0])
print(mc[0][1])
words, counts = zip(*c.most_common())
print(words)
print(counts)
s = '走れと言われたので走ると言った'
print(collections.Counter(t.tokenize(s, wakati=True)))
Counter({'と': 2, 'た': 2, '走れ': 1, '言わ': 1, 'れ': 1, 'ので': 1, '走る': 1, '言っ': 1})
print(collections.Counter(token.base_form for token in t.tokenize(s)))
Counter({'走る': 2, 'と': 2, '言う': 2, 'た': 2, 'れる': 1, 'ので': 1})
print(type(token.base_form for token in t.tokenize(s)))
collections.Counter
token.base_form for token in t.tokenize(s)
token.part_of_speech.startswith('動詞,自立')
Counter({'走る': 2, '言う': 2})
print(collections.Counter(token.part_of_speech.split(',')[0] for token in t.tokenize(s)))
Counter({'動詞': 5, '助詞': 3, '助動詞': 2})
from janome.tokenizer import Tokenizer
t = Tokenizer()
s = 'すもももももももものうち'
print(type(t.tokenize(s)))
print(type(t.tokenize(s).__next__()))
janome.tokenizer.Token
for token in t.tokenize(s):
print(token)
print(type(list(t.tokenize(s))))
print(type(list(t.tokenize(s))[0]))
janome.tokenizer.Token
token = t.tokenize('走れ').__next__()
print(type(token))
janome.tokenizer.Token
print(token)
print(token.surface)
print(token.part_of_speech)
print(token.part_of_speech.split(','))
print(token.part_of_speech.split(',')[0])
print(token.infl_type)
print(token.infl_form)
print(token.base_form)
print(token.reading)
print(token.phonetic)
s = '走れと言われたので走ると言った'
for token in t.tokenize(s):
print(token)
print(type(t.tokenize(s, wakati=True)))
print(type(t.tokenize(s, wakati=True).__next__()))
print(list(t.tokenize(s, wakati=True)))
t_wakati = Tokenizer(wakati=True)
print(list(t_wakati.tokenize(s)))
print([token.surface for token in t.tokenize(s)])
print([token.base_form for token in t.tokenize(s)])
print([token.part_of_speech.split(',')[0] for token in t.tokenize(s)])
token.surface for token in t.tokenize(s)
token.part_of_speech.startswith('動詞')
token.surface for token in t.tokenize(s)
not token.part_of_speech.startswith('動詞')
token.surface for token in t.tokenize(s)
token.part_of_speech.startswith('動詞,自立')
token.surface for token in t.tokenize(s)
import os
import pathlib
import pprint
import joblib
src_dir = 'data/temp/joblib/src_file'
os.makedirs(src_dir, exist_ok=True)
for i in range(10):
pathlib.Path(src_dir).joinpath(f'file{i:05}.txt').write_text(f'This is file{i:05}')
def func_write(i):
pathlib.Path(src_dir).joinpath(f'file{i:05}.txt').write_text(f'This is file{i:05}')
_ = joblib.Parallel(n_jobs=-1)(joblib.delayed(func_write)(i) for i in range(10))
first_lines = []
for p in pathlib.Path(src_dir).glob('*.txt'):
with p.open() as f:
first_lines.append((p.name, f.readline()))
pprint.pprint(sorted(first_lines))
file00000.txt
file00001.txt
file00002.txt
file00003.txt
file00004.txt
file00005.txt
file00006.txt
file00007.txt
file00008.txt
file00009.txt
def func_read_first_line(p):
with p.open() as f:
return p.name, f.readline()
first_lines = joblib.Parallel(n_jobs=-1)
joblib.delayed(func_read_first_line)(p) for p in pathlib.Path(src_dir).glob('*.txt')
pprint.pprint(sorted(first_lines))
file00000.txt
file00001.txt
file00002.txt
file00003.txt
file00004.txt
file00005.txt
file00006.txt
file00007.txt
file00008.txt
file00009.txt
import os
import glob
from PIL import Image
import joblib
dst_dir = 'data/temp/joblib/dst_img'
os.makedirs(dst_dir, exist_ok=True)
files = glob.glob('data/temp/joblib/src_img/*')
for f in files:
img = Image.open(f)
img_resize = img.resize((img.width // 2, img.height // 2))
root, ext = os.path.splitext(f)
basename = os.path.basename(root)
img_resize.save(os.path.join(dst_dir, basename + '_half' + ext))
except OSError as e:
pass
def func(f):
img = Image.open(f)
img_resize = img.resize((img.width // 2, img.height // 2))
root, ext = os.path.splitext(f)
basename = os.path.basename(root)
img_resize.save(os.path.join(dst_dir, basename + '_half' + ext))
except OSError as e:
pass
_ = joblib.Parallel(n_jobs=-1)(joblib.delayed(func)(f) for f in files)
import joblib
print(joblib.__version__)
def func(i):
return i
result = joblib.Parallel(n_jobs=-1)(joblib.delayed(func)(i) for i in range(5))
print(result)
result = joblib.Parallel(n_jobs=-1, verbose=1)(joblib.delayed(func)(i) for i in range(5))
Parallel(n_jobs=-1)
Parallel(n_jobs=-1)
result = joblib.Parallel(n_jobs=-1, verbose=11)(joblib.delayed(func)(i) for i in range(5))
Parallel(n_jobs=-1)
Parallel(n_jobs=-1)
Parallel(n_jobs=-1)
batch_size=2.
Parallel(n_jobs=-1)
Parallel(n_jobs=-1)
Parallel(n_jobs=-1)
Parallel(n_jobs=-1)
result = joblib.Parallel(n_jobs=2, verbose=1)(joblib.delayed(func)(i) for i in range(5))
Parallel(n_jobs=2)
Parallel(n_jobs=2)
result = joblib.Parallel(n_jobs=-2, verbose=1)(joblib.delayed(func)(i) for i in range(5))
Parallel(n_jobs=-2)
Parallel(n_jobs=-2)
def func_multi(i):
return i, i**2, i**3
results = joblib.Parallel(n_jobs=-1)(joblib.delayed(func_multi)(i) for i in range(5))
print(results)
a, b, c = zip(*results)
print(a)
print(b)
print(c)
def func_multi2(i, j, k):
return i, j, k
results = joblib.Parallel(n_jobs=-1)
joblib.delayed(func_multi2)(i, j, k) for i, j, k in zip(range(5), range(5, 10), range(10, 15))
print(results)
def func_none():
pass
results = joblib.Parallel(n_jobs=-1)(joblib.delayed(func_none)() for i in range(5))
print(results)
joblib.Parallel(n_jobs=-1)(joblib.delayed(func_none)() for i in range(5))
_ = joblib.Parallel(n_jobs=-1)(joblib.delayed(func_none)() for i in range(5))
import joblib
def func(i):
return i
for i in range(100):
func(i)
std. dev. of
joblib.Parallel(n_jobs=-1)(joblib.delayed(func)(i) for i in range(100))
std. dev. of
import time
def func_sleep(i):
time.sleep(0.1)
return i
for i in range(8):
func_sleep(i)
std. dev. of
joblib.Parallel(n_jobs=-1)(joblib.delayed(func_sleep)(i) for i in range(8))
std. dev. of
joblib.Parallel(n_jobs=2)(joblib.delayed(func_sleep)(i) for i in range(8))
std. dev. of
import json
from collections import OrderedDict
import pprint
s = r
print(s)
d = json.loads(s)
pprint.pprint(d, width=40)
print(type(d))
od = json.loads(s, object_pairs_hook=OrderedDict)
pprint.pprint(od)
OrderedDict([('i', 1), ('j', 2)])
OrderedDict([('X', 1), ('Y', 10)])
OrderedDict([('X', 2), ('Y', 20)])
b = s.encode()
print(b)
print(type(b))
db = json.loads(b)
pprint.pprint(db, width=40)
print(type(db))
sb = b.decode()
print(sb)
print(type(sb))
dsb = json.loads(sb)
pprint.pprint(dsb, width=40)
print(type(dsb))
sb_u = b.decode('unicode-escape')
print(sb_u)
print(type(sb_u))
dsb_u = json.loads(sb_u)
pprint.pprint(dsb_u, width=40)
print(type(dsb_u))
with open('data/src/test.json') as f:
print(f.read())
with open('data/src/test.json') as f:
df = json.load(f)
pprint.pprint(df, width=40)
print(type(df))
pprint.pprint(d, width=40)
print(d['A'])
print(d['A']['i'])
print(d['B'])
print(d['B'][0])
print(d['B'][0]['X'])
value = d['B'][1]['Y']
print(value)
print(d['D'])
print(d.get('D'))
d['C'] = 'ん'
pprint.pprint(d, width=40)
d.pop('C')
pprint.pprint(d, width=40)
d['C'] = 'あ'
pprint.pprint(d, width=40)
sd = json.dumps(d)
print(sd)
print(type(sd))
pprint.pprint(od)
OrderedDict([('i', 1), ('j', 2)])
OrderedDict([('X', 1), ('Y', 10)])
OrderedDict([('X', 2), ('Y', 20)])
sod = json.dumps(od)
print(sod)
print(type(sod))
print(json.dumps(d, separators=(',', ':')))
print(json.dumps(d, separators=(' / ', '-> ')))
print(json.dumps(d, indent=4))
print(json.dumps(od))
print(json.dumps(od, sort_keys=True))
print(json.dumps(od, ensure_ascii=False))
with open('data/dst/test2.json', 'w') as f:
json.dump(d, f, indent=4)
with open('data/dst/test2.json') as f:
print(f.read())
d_new = {'A': 100, 'B': 'abc', 'C': 'あいうえお'}
with open('data/dst/test_new.json', 'w') as f:
json.dump(d_new, f, indent=2, ensure_ascii=False)
with open('data/dst/test_new.json') as f:
print(f.read())
with open('data/dst/test_new.json') as f:
d_update = json.load(f, object_pairs_hook=OrderedDict)
print(d_update)
OrderedDict([('A', 100), ('B', 'abc'), ('C', 'あいうえお')])
d_update['A'] = 200
d_update.pop('B')
d_update['D'] = 'new value'
print(d_update)
OrderedDict([('A', 200), ('C', 'あいうえお'), ('D', 'new value')])
with open('data/dst/test_new_update.json', 'w') as f:
json.dump(d_update, f, indent=2, ensure_ascii=False)
with open('data/dst/test_new_update.json') as f:
print(f.read())
with open('data/src/test.json') as f:
print(f.read())
with open('data/src/test.json', encoding='unicode-escape') as f:
print(f.read())
print(b)
print(b.decode())
print(b.decode(encoding='unicode-escape'))
d = {"A": 100, "B": 'abc', "C": 'あいうえお'}
print(str(d))
print(json.loads(str(d)))
JSONDecodeError: Expecting
quotes: line
print(json.dumps(d))
print(json.loads(json.dumps(d)))
print(type(json.loads(json.dumps(d))))
f = 123.456789
print(f)
import numpy as np
a = np.array([0.123456789, 0.987654321])
array([0.12345679, 0.98765432])
array([0.123, 0.988])
print(a)
print(np.get_printoptions()['precision'])
np.set_printoptions(precision=5)
array([0.12346, 0.98765])
print(a)
print(a[0])
array([0.12345679, 0.98765432])
print(np.get_printoptions()['precision'])
array([0.12345679, 0.98765432])
np.set_printoptions(formatter={'float': '{:.2e}'.format})
array([1.23e-01, 9.88e-01])
import pandas as pd
df = pd.DataFrame({'a': [0.123456789], 'b': [0.987654321]})
print(pd.options.display.precision)
print(pd.options.display.precision)
pd.options.display.precision = 3
print(df)
array_example.ipynb
array_example.py
array_example.ipynb
array_example.py
output = !date
print(output)
print(type(output))
IPython.utils.text.SList
print(isinstance(output, list))
print(len(output))
print(output[0])
print(type(output[0]))
output = !ls -l 
print(output)
array_example.ipynb
array_example.py
print(len(output))
print(output[0])
array_example.ipynb
print(output[1])
array_example.py
print(output.n)
array_example.ipynb
array_example.py
print(type(output.n))
print('\n'.join(output))
array_example.ipynb
array_example.py
array_example.ipynb
array_example.py
output = !ls 
print(output)
array_example.ipynb
array_example.py
print(len(output))
array_example.ipynb
array_example.py
array_example.ipynb
array_example.py
array_example.ipynb
array_example.py
output = !ls -C 
print(output)
array_example.ipynb
tarray_example.py
print(len(output))
print(output[0])
array_example.ipynb
array_example.py
import os
print(os.getcwd())
print(os.getcwd())
print(os.getcwd())
print(os.getcwd())
os.chdir('data')
print(os.getcwd())
import IPython.display
IPython.display.YouTubeVideo('6XvmhE1J9PY')
IPython.lib.display.YouTubeVideo
IPython.display.YouTubeVideo('6XvmhE1J9PY', width=480, height=270)
IPython.lib.display.YouTubeVideo
IPython.display.YouTubeVideo('6XvmhE1J9PY', start=30)
IPython.lib.display.YouTubeVideo
IPython.display.VimeoVideo('289502328')
IPython.lib.display.VimeoVideo
import numpy as np
from sklearn import datasets, model_selection
from keras.models import Sequential
from keras.layers import Dense
from keras import optimizers
mnist = datasets.fetch_mldata('MNIST original', data_home='data/src/download')
X = mnist.data / 255
y = mnist.target
Y = np.identity(10)[y.astype(int)]
train_size = 60000
test_size = 10000
X_train, X_test, Y_train, Y_test = model_selection.train_test_split
X, Y, test_size=test_size, train_size=train_size
model = Sequential()
model.add(Dense(units=10, input_dim=784, activation='softmax'))
sgd = optimizers.SGD(lr=0.5)
model.compile
optimizer=sgd
loss='categorical_crossentropy'
metrics=['accuracy']
batch_size = 100
epochs = 20
model.fit(X_train, Y_train, epochs=epochs, batch_size=batch_size)
- 1
- 1
- 1
- 1
- 1
- 1
- 1
- 1
- 1
- 1
- 1
- 1
- 1
- 1
- 1
- 1
- 1
- 1
- 1
- 1
keras.callbacks.History
score = model.evaluate(X_test, Y_test)
print(score)
- 0
l = [1, -3, 2]
print(sorted(l))
-3
print(sorted(l, key=abs))
-3
l_abs = [abs(i) for i in l]
print(l_abs)
print(sorted(l_abs))
l.sort(key=abs)
print(l)
-3
l = [1, -3, 2]
print(max(l))
print(max(l, key=abs))
-3
print(min(l))
-3
print(min(l, key=abs))
l_str = ['bbb', 'c', 'aa']
print(sorted(l_str))
print(sorted(l_str, key=len))
l_2d = [[2, 10], [1, -30], [-3, 20]]
print(sorted(l_2d))
-3
-30
print(sorted(l_2d, key=max))
-30
-3
l_2d, key=lambda x: max([abs(i) for i in x])
-3
-30
l_2d, key=lambda x: max(abs(i) for i in x)
-3
-30
def max_abs(x):
return max(abs(i) for i in x)
print(sorted(l_2d, key=max_abs))
-3
-30
import operator
l_2d = [[2, 10], [1, -30], [-3, 20]]
print(sorted(l_2d, key=operator.itemgetter(1)))
-30
-3
f = operator.itemgetter(1)
print(f([2, 10]))
print(operator.itemgetter(1)([2, 10]))
l_2d, key=lambda x: x[1]
-30
-3
l_dict = [{'k1': 2, 'k2': 10}, {'k1': 1}, {'k1': 3}]
print(sorted(l_dict))
not supported 
print(sorted(l_dict, key=operator.itemgetter('k1')))
print(sorted(l_dict, key=operator.itemgetter('k2')))
l_dict, key=lambda x: x['k1']
l_dict = [{'k1': 2, 'k2': 'ccc'}, {'k1': 1, 'k2': 'ccc'}, {'k1': 2, 'k2': 'aaa'}]
print(operator.itemgetter('k1', 'k2')(l_dict[0]))
print(sorted(l_dict, key=operator.itemgetter('k1')))
print(sorted(l_dict, key=operator.itemgetter('k1', 'k2')))
print(sorted(l_dict, key=operator.itemgetter('k2', 'k1')))
l_dict, key=lambda x: (x['k1'], x['k2'])
import datetime
l_dt = [datetime.date(2003, 2, 10), datetime.date(2001, 3, 20), datetime.date(2002, 1, 30)]
print(l_dt[0])
print(l_dt[0].day)
f = operator.attrgetter('day')
print(f(l_dt[0]))
print(sorted(l_dt))
datetime.date(2001, 3, 20)
datetime.date(2002, 1, 30)
datetime.date(2003, 2, 10)
print(sorted(l_dt, key=operator.attrgetter('day')))
datetime.date(2003, 2, 10)
datetime.date(2001, 3, 20)
datetime.date(2002, 1, 30)
l_dt, key=lambda x: x.day
datetime.date(2003, 2, 10)
datetime.date(2001, 3, 20)
datetime.date(2002, 1, 30)
l_str = ['0_xxxxA', '1_Axxxx', '2_xxAxx']
print(l_str[0])
print(l_str[0].find('A'))
f = operator.methodcaller('find', 'A')
print(f(l_str[0]))
print(sorted(l_str))
print(sorted(l_str, key=operator.methodcaller('find', 'A')))
l_str, key=lambda x: x.find('A')
import keyword
import pprint
print(type(keyword.kwlist))
print(len(keyword.kwlist))
pprint.pprint(keyword.kwlist, compact=True)
assert', '
break
continue
pass
raise', '
return', '
yield
print(keyword.kwlist[0])
print(type(keyword.kwlist[0]))
True = 100
SyntaxError: can
print(keyword.iskeyword('None'))
print(keyword.iskeyword('none'))
def func_kwargs(**kwargs):
print('kwargs: ', kwargs)
print('type: ', type(kwargs))
func_kwargs(key1=1, key2=2, key3=3)
def func_kwargs_positional(arg1, arg2, **kwargs):
print('arg1: ', arg1)
print('arg2: ', arg2)
print('kwargs: ', kwargs)
func_kwargs_positional(0, 1, key1=1)
d = {'key1': 1, 'key2': 2, 'arg1': 100, 'arg2': 200}
func_kwargs_positional(**d)
def func_kwargs_error(**kwargs, arg):
print(kwargs)
SyntaxError: invalid
def add_def(a, b=1):
return a + b
add_lambda = lambda a, b=1: a + b
print(add_def(3, 4))
print(add_def(3))
print(add_lambda(3, 4))
print(add_lambda(3))
get_odd_even = lambda x: 'even' if x % 2 == 0 else 'odd'
print(get_odd_even(3))
print(get_odd_even(4))
l = ['Charle', 'Bob', 'Alice']
l_sorted = sorted(l)
print(l_sorted)
print(len('Alice'))
l_sorted_len = sorted(l, key=len)
print(l_sorted_len)
print((lambda x: x[1])('Alice'))
l_sorted_second = sorted
l, key=lambda x: x[1]
print(l_sorted_second)
l = [0, 1, 2, 3]
map_square = map(lambda x: x**2, l)
print(map_square)
print(list(map_square))
l_square = [x**2 for x in l]
print(l_square)
g_square = (x**2 for x in l)
print(g_square)
print(list(g_square))
filter_even = filter(lambda x: x % 2 == 0, l)
print(list(filter_even))
l_even = [x for x in l if x % 2 == 0]
print(l_even)
l = [0, 1, 2]
print(len(l))
t = (0, 1, 2)
print(len(t))
s = {0, 1, 2}
print(len(s))
d = {'key0': 0, 'key1': 1, 'key2': 2}
print(len(d))
s = 'abcde'
print(len(s))
s = 'あいうえお'
print(len(s))
import numpy as np
a_1d = np.arange(3)
print(a_1d)
print(len(a_1d))
a_2d = np.arange(12).reshape((3, 4))
print(a_2d)
print(len(a_2d))
a_3d = np.arange(24).reshape((2, 3, 4))
print(a_3d)
print(len(a_3d))
import pandas as pd
df = pd.DataFrame({'A': [0, 1, 2], 'B': [3, 4, 5]}, index=['a', 'b', 'c'])
print(df)
print(len(df))
s = pd.Series([0, 1, 2], index=['a', 'b', 'c'])
print(s)
dtype: int64
print(len(s))
print(len(100))
TypeError: object
len()
print(len(0.1))
TypeError: object
len()
print(len(True))
TypeError: object
len()
def convert_1d_to_2d(l, cols):
return [l[i:i + cols] for i in range(0, len(l), cols)]
l = [0, 1, 2, 3, 4, 5]
print(convert_1d_to_2d(l, 2))
print(convert_1d_to_2d(l, 3))
print(convert_1d_to_2d(l, 4))
def convert_1d_to_2d_rows(l, rows):
return convert_1d_to_2d(l, len(l) // rows)
print(convert_1d_to_2d_rows(l, 2))
print(convert_1d_to_2d_rows(l, 3))
print(convert_1d_to_2d_rows(l, 4))
import pprint
print([100] > [-100])
print([1, 2, 100] > [1, 2, -100])
print([1, 2, 100] > [1, 100])
l_2d = [[20, 3, 100], [1, 200, 30], [300, 10, 2]]
pprint.pprint(l_2d, width=20)
pprint.pprint(sorted(l_2d), width=20)
pprint.pprint(l_2d, width=20)
pprint.pprint([sorted(l) for l in l_2d], width=20)
pprint.pprint
list(x) for x in zip(*[sorted(l) for l in zip(*l_2d)])
width=20
import numpy as np
print(np.sort(l_2d))
print(np.sort(l_2d, axis=0))
print(type(np.sort(l_2d)))
numpy.ndarray
print(np.sort(l_2d).tolist())
print(type(np.sort(l_2d).tolist()))
l_2d_error = [[1, 2], [3, 4, 5]]
print(np.sort(l_2d_error))
list([1, 2]) 
list([3, 4, 5])
VisibleDeprecationWarning: Creating
lengths or shapes
deprecated. If
dtype=object
return array(a, dtype, copy=False, order=order, subok=True)
pprint.pprint
l_2d, key=lambda x: x[1]
width=20
pprint.pprint
l_2d, key=lambda x: x[2]
width=20
import operator
pprint.pprint(sorted(l_2d, key=operator.itemgetter(1)), width=20)
l_2d_dup = [[20, 3, 100], [1, 200, 30], [1, 10, 2]]
pprint.pprint(l_2d_dup, width=20)
pprint.pprint(sorted(l_2d_dup), width=20)
pprint.pprint(sorted(l_2d_dup, key=operator.itemgetter(0)), width=20)
pprint.pprint(sorted(l_2d_dup, key=operator.itemgetter(0, 1)), width=20)
pprint.pprint
l_2d_dup, key=lambda x: (x[0], x[1])
width=20
import pandas as pd
df = pd.DataFrame(l_2d_dup, columns=['A', 'B', 'C'], index=['X', 'Y', 'Z'])
print(df)
print(df.sort_values('A'))
print(df.sort_values('X', axis=1))
print(df.sort_values(['A', 'B']))
l = list(range(3))
print(l)
l.append(100)
print(l)
l.append('new')
print(l)
l.append([3, 4, 5])
print(l)
l = list(range(3))
print(l)
l.extend([100, 101, 102])
print(l)
l.extend((-1, -2, -3))
print(l)
-1
-2
-3
l.extend('new')
print(l)
-1
-2
-3
l2 = l + [5, 6, 7]
print(l2)
-1
-2
-3
l += [5, 6, 7]
print(l)
-1
-2
-3
l = list(range(3))
print(l)
l.insert(0, 100)
print(l)
l.insert(-1, 200)
print(l)
l.insert(0, [-1, -2, -3])
print(l)
-1
-2
-3
l = list(range(3))
print(l)
l[1:1] = [100, 200, 300]
print(l)
l = list(range(3))
print(l)
l[1:2] = [100, 200, 300]
print(l)
l1 = ['a', 'b', 'c']
l2 = ['b', 'c', 'd']
l3 = ['c', 'd', 'e']
l1_l2_and = set(l1) & set(l2)
print(l1_l2_and)
print(type(l1_l2_and))
l1_l2_and_list = list(l1_l2_and)
print(l1_l2_and_list)
print(type(l1_l2_and_list))
print(len(l1_l2_and))
l1_l2_l3_and = set(l1) & set(l2) & set(l3)
print(l1_l2_l3_and)
l1_l2_sym_diff = set(l1) ^ set(l2)
print(l1_l2_sym_diff)
print(list(l1_l2_sym_diff))
print(len(l1_l2_sym_diff))
l1_l2_l3_sym_diff = set(l1) ^ set(l2) ^ set(l3)
print(l1_l2_l3_sym_diff)
l_all = l1 + l2 + l3
print(l_all)
print(set(l_all))
l_all_only = [x for x in set(l_all) if l_all.count(x) == 1]
print(l_all_only)
l1_duplicate = ['a', 'a', 'b', 'c'] 
l_duplicate_all = l1_duplicate + l2 + l3
l_duplicate_all_only = [x for x in set(l_duplicate_all) if l_duplicate_all.count(x) == 1]
print(l_duplicate_all_only)
l_unique_all = list(set(l1_duplicate)) + list(set(l2)) + list(set(l3))
print(l_unique_all)
l_uniaues_all_only = [x for x in set(l_unique_all) if l_unique_all.count(x) == 1]
print(l_uniaues_all_only)
l1_l2_or = set(l1 + l2)
print(l1_l2_or)
print(list(l1_l2_or))
print(len(l1_l2_or))
l1_l2_l3_or = set(l1 + l2 + l3)
print(l1_l2_l3_or)
squares = [i**2 for i in range(5)]
print(squares)
squares = []
for i in range(5):
squares.append(i**2)
print(squares)
odds = [i for i in range(10) if i % 2 == 1]
print(odds)
odds = []
for i in range(10):
if i % 2 == 1:
odds.append(i)
print(odds)
odd_even = ['odd' if i % 2 == 1 else 'even' for i in range(10)]
print(odd_even)
odd_even = []
for i in range(10):
if i % 2 == 1:
odd_even.append('odd')
odd_even.append('even')
print(odd_even)
odd10 = [i * 10 if i % 2 == 1 else i for i in range(10)]
print(odd10)
matrix = [[1, 2, 3], [4, 5, 6], [7, 8, 9]]
flat = [x for row in matrix for x in row]
print(flat)
flat = []
for row in matrix:
for x in row:
flat.append(x)
print(flat)
cells = [(row, col) for row in range(3) for col in range(2)]
print(cells)
col == row
print(cells)
print(cells)
l_str1 = ['a', 'b', 'c']
l_str2 = ['x', 'y', 'z']
l_zip = [(s1, s2) for s1, s2 in zip(l_str1, l_str2)]
print(l_zip)
l_zip = []
for s1, s2 in zip(l_str1, l_str2):
l_zip.append((s1, s2))
print(l_zip)
l_enu = [(i, s) for i, s in enumerate(l_str1)]
print(l_enu)
l_enu = []
for i, s in enumerate(l_str1):
l_enu.append((i, s))
print(l_enu)
l_zip_if = [(s1, s2) for s1, s2 in zip(l_str1, l_str2) if s1 != 'b']
print(l_zip_if)
l_int1 = [1, 2, 3]
l_int2 = [10, 20, 30]
l_sub = [i2 - i1 for i1, i2 in zip(l_int1, l_int2)]
print(l_sub)
def has_duplicates(seq):
return len(seq) != len(set(seq))
l = [0, 1, 2]
print(has_duplicates(l))
l = [0, 1, 1, 2]
print(has_duplicates(l))
l_2d = [[0, 1], [1, 1], [0, 1], [1, 0]]
print(has_duplicates(l_2d))
TypeError: unhashable
def has_duplicates2(seq):
seen = []
unique_list = [x for x in seq if x not in seen and not seen.append(x)]
return len(seq) != len(unique_list)
l_2d = [[0, 0], [0, 1], [1, 1], [1, 0]]
print(has_duplicates2(l_2d))
l_2d = [[0, 0], [0, 1], [1, 1], [1, 1]]
print(has_duplicates2(l_2d))
l = [0, 1, 2]
print(has_duplicates2(l))
l = [0, 1, 1, 2]
print(has_duplicates2(l))
l_2d = [[0, 1], [2, 3]]
print(sum(l_2d, []))
print(has_duplicates(sum(l_2d, [])))
l_2d = [[0, 1], [2, 0]]
print(has_duplicates(sum(l_2d, [])))
l = ['apple', 100, 0.123]
print(l)
l_2d = [[0, 1, 2], [3, 4, 5], [6, 7, 8]]
print(l_2d)
print(l[1])
print(l_2d[1])
print(l_2d[1][1])
print(l[:2])
l_num = [0, 10, 20, 30]
print(min(l_num))
print(max(l_num))
print(sum(l_num))
print(sum(l_num) / len(l_num))
l_str = ['apple', 'orange', 'banana']
for s in l_str:
print(s)
import itertools
l_2d = [[0, 1], [2, 3]]
print(list(itertools.chain.from_iterable(l_2d)))
t_2d = ((0, 1), (2, 3))
print(tuple(itertools.chain.from_iterable(t_2d)))
l_3d = [[[0, 1], [2, 3]], [[4, 5], [6, 7]]]
print(list(itertools.chain.from_iterable(l_3d)))
l_mix = [[0, 1], [2, 3], 4]
print(list(itertools.chain.from_iterable(l_mix)))
print(sum(l_2d, []))
print(sum(l_2d))
TypeError: unsupported
type(s) 
print(sum(l_3d, []))
print(sum(l_mix, []))
TypeError: can
list (not "int") 
import collections
def flatten(l):
for el in l:
if isinstance(el, collections.abc.Iterable) and not isinstance(el, (str, bytes)):
yield from
flatten(el)
yield el
print(list(flatten(l_2d)))
print(list(flatten(l_3d)))
print(list(flatten(l_mix)))
l_t_r_mix = [[0, 1], (2, 3), 4, range(5, 8)]
print(list(flatten(l_t_r_mix)))
def flatten_list(l):
for el in l:
if isinstance(el, list):
yield from
flatten_list(el)
yield el
print(list(flatten_list(l_2d)))
print(list(flatten_list(l_3d)))
print(list(flatten_list(l_mix)))
print(list(flatten_list(l_t_r_mix)))
range(5, 8)
def flatten_list_tuple_range(l):
for el in l:
if isinstance(el, (list, tuple, range)):
yield from
flatten_list_tuple_range(el)
yield el
print(list(flatten_list_tuple_range(l_t_r_mix)))
import itertools
l_2d_5 = [[0, 1, 2]] * 5
print(l_2d_5)
list(itertools.chain.from_iterable(l_2d_5))
std. dev. of
sum(l_2d_5, [])
std. dev. of
l_2d_100 = [[0, 1, 2]] * 100
list(itertools.chain.from_iterable(l_2d_100))
std. dev. of
sum(l_2d_100, [])
std. dev. of
l_2d_10000 = [[0, 1, 2]] * 10000
list(itertools.chain.from_iterable(l_2d_10000))
std. dev. of
sum(l_2d_10000, [])
std. dev. of
l = list('abcde')
print(l)
print(l.index('a'))
print(l.index('c'))
print(l.index('x'))
def my_index(l, x, default=False):
if x in l:
return l.index(x)
return default
print(my_index(l, 'd'))
print(my_index(l, 'x'))
print(my_index(l, 'x', -1))
-1
l_dup = list('abcba')
print(l_dup)
print(l_dup.index('a'))
print(l_dup.index('b'))
print([i for i, x in enumerate(l_dup) if x == 'a'])
print([i for i, x in enumerate(l_dup) if x == 'b'])
print([i for i, x in enumerate(l_dup) if x == 'c'])
print([i for i, x in enumerate(l_dup) if x == 'x'])
def my_index_multi(l, x):
return [i for i, _x in enumerate(l) if _x == x]
print(my_index_multi(l_dup, 'a'))
print(my_index_multi(l_dup, 'c'))
print(my_index_multi(l_dup, 'x'))
t = tuple('abcde')
print(t)
print(t.index('a'))
print(t.index('x'))
ValueError: tuple
index(x)
print(my_index(t, 'c'))
print(my_index(t, 'x'))
t_dup = tuple('abcba')
print(t_dup)
print(my_index_multi(t_dup, 'a'))
def my_index2(l, x, default=False):
return l.index(x) if x in l else default
print(my_index2(l, 'd'))
print(my_index2(l, 'x'))
print(my_index2(l, 'x', -1))
-1
l = [30, 40, 20, 0, 10]
print(l.index(max(l)))
print(l.index(min(l)))
l_dup = [0, 40, 20, 0, 40]
print(l_dup.index(max(l_dup)))
print(l_dup.index(min(l_dup)))
print(my_index_multi(l_dup, max(l_dup)))
print(my_index_multi(l_dup, min(l_dup)))
l_empty = []
print(l_empty)
print(len(l_empty))
l_empty.append(100)
l_empty.append(200)
print(l_empty)
l_empty.remove(100)
print(l_empty)
l = [0] * 10
print(l)
print(len(l))
print([0, 1, 2] * 3)
l_2d_ng = [[0] * 4] * 3
print(l_2d_ng)
l_2d_ng[0][0] = 5
print(l_2d_ng)
l_2d_ng[0].append(100)
print(l_2d_ng)
print(id(l_2d_ng[0]) == id(l_2d_ng[1]) == id(l_2d_ng[2]))
l_2d_ok = [[0] * 4 for i in range(3)]
print(l_2d_ok)
l_2d_ok[0][0] = 100
print(l_2d_ok)
print(id(l_2d_ok[0]) == id(l_2d_ok[1]) == id(l_2d_ok[2]))
l_2d_ok_2 = [[0] * 4 for i in [1] * 3]
print(l_2d_ok_2)
l_2d_ok_2[0][0] = 100
print(l_2d_ok_2)
print(id(l_2d_ok_2[0]) == id(l_2d_ok_2[1]) == id(l_2d_ok_2[2]))
l_3d = [[[0] * 2 for i in range(3)] for j in range(4)]
print(l_3d)
l_3d[0][0][0] = 100
print(l_3d)
t = (0,) * 5
print(t)
import array
a = array.array('i', [0] * 5)
print(a)
array('i', [0, 0, 0, 0, 0])
l = [0, 1, 2, 3]
print(len(l))
l_length = len(l)
print(l_length)
print(type(l_length))
l_2d = [[0, 1, 2], [3, 4, 5]]
print(len(l_2d))
print([len(v) for v in l_2d])
print(sum(len(v) for v in l_2d))
import numpy as np
l_2d = [[0, 1, 2], [3, 4, 5]]
arr_2d = np.array(l_2d)
print(arr_2d)
print(arr_2d.size)
print(arr_2d.shape)
l_multi = [[0, 1, 2, [10, 20, 30]], [3, 4, 5], 100]
print(len(l_multi))
arr_multi = np.array(l_multi)
print(arr_multi)
list([0, 1, 2, [10, 20, 30]]) 
list([3, 4, 5]) 
print(arr_multi.size)
print(arr_multi.shape)
def my_len(l):
count = 0
if isinstance(l, list):
for v in l:
count += my_len(v)
return count
return 1
l_multi = [[0, 1, 2, [10, 20, 30]], [3, 4, 5], 100]
print(my_len(l_multi))
l_2d = [[0, 1, 2], [3, 4, 5]]
print(my_len(l_2d))
l = [0, 1, 2, 3]
print(my_len(l))
l = ['a', 'a', 'a', 'a', 'b', 'c', 'c']
print(len(l))
l = ['a', 'a', 'a', 'a', 'b', 'c', 'c']
print(l.count('a'))
print(l.count('b'))
print(l.count('c'))
print(l.count('d'))
l = list(range(-5, 6))
print(l)
-5
-4
-3
-2
-1
print([i < 0 for i in l])
print(sum([i < 0 for i in l]))
print(sum((i < 0 for i in l)))
print(sum(i < 0 for i in l))
print([not (i < 0) for i in l])
print(sum(not (i < 0) for i in l))
print(sum(i >= 0 for i in l))
print([i % 2 == 1 for i in l])
print(sum(i % 2 == 1 for i in l))
l = ['apple', 'orange', 'banana']
print([s.endswith('e') for s in l])
print(sum(s.endswith('e') for s in l))
import statistics
import pprint
l = [0, 1, 2, 3, 4]
print(l)
def min_max(l):
l_min = min(l)
l_max = max(l)
return [(i - l_min) / (l_max - l_min) for i in l]
print(min_max(l))
def standardization(l):
l_mean = statistics.mean(l)
l_stdev = statistics.stdev(l)
return [(i - l_mean) / l_stdev for i in l]
pprint.pprint(standardization(l))
-1.2649110640673518
-0.6324555320336759
def standardization_p(l):
l_mean = statistics.mean(l)
l_pstdev = statistics.pstdev(l)
return [(i - l_mean) / l_pstdev for i in l]
pprint.pprint(standardization_p(l))
-1.414213562373095
-0.7071067811865475
l_2d = [[0, 1, 2], [3, 4, 5], [6, 7, 8]]
print(l_2d)
pprint.pprint([min_max(l_1d) for l_1d in l_2d], width=40)
pprint.pprint([standardization(l_1d) for l_1d in l_2d], width=40)
-1.0
-1.0
-1.0
pprint.pprint([standardization_p(l_1d) for l_1d in l_2d])
-1.224744871391589
-1.224744871391589
-1.224744871391589
l_2d_min_max_col = list(zip(*[min_max(l_1d) for l_1d in list(zip(*l_2d))]))
pprint.pprint(l_2d_min_max_col, width=40)
l_2d_standardization_col = list(zip(*[standardization(l_1d) for l_1d in list(zip(*l_2d))]))
pprint.pprint(l_2d_standardization_col, width=40)
-1.0
-1.0
-1.0
l_2d_standardization_p_col = list(zip(*[standardization_p(l_1d) for l_1d in list(zip(*l_2d))]))
pprint.pprint(l_2d_standardization_p_col)
-1.2247448713915892
-1.2247448713915892
-1.2247448713915892
def min_max_2d_all(l_2d):
l_flatten = sum(l_2d, [])
l_2d_min = min(l_flatten)
l_2d_max = max(l_flatten)
return 
pprint.pprint(min_max_2d_all(l_2d), width=40)
def standardization_2d_all(l):
l_flatten = sum(l_2d, [])
l_2d_mean = statistics.mean(l_flatten)
l_2d_stdev = statistics.stdev(l_flatten)
return 
l_2d_stdev for i in l_1d
pprint.pprint(standardization_2d_all(l_2d))
-1.4605934866804429
-1.0954451150103321
-0.7302967433402214
-0.3651483716701107
def standardization_p_2d_all(l):
l_flatten = sum(l_2d, [])
l_2d_mean = statistics.mean(l_flatten)
l_2d_pstdev = statistics.pstdev(l_flatten)
return 
l_2d_pstdev for i in l_1d
pprint.pprint(standardization_p_2d_all(l_2d))
-1.5491933384829668
-1.161895003862225
-0.7745966692414834
-0.3872983346207417
l = list(range(10))
print(l)
l.clear()
print(l)
l = list(range(10))
print(l)
print(l.pop(0))
print(l)
print(l.pop(3))
print(l)
print(l.pop(-2))
print(l)
print(l.pop())
print(l)
print(l.pop(100))
IndexError: pop
l = ['Alice', 'Bob', 'Charlie', 'Bob', 'Dave']
print(l)
l.remove('Alice')
print(l)
l.remove('Bob')
print(l)
l.remove('xxx')
ValueError: list
remove(x)
l = list(range(10))
print(l)
print([i for i in l if i % 2 == 0])
print([i for i in l if i % 2 != 0])
print(l)
l = ['Alice', 'Bob', 'Charlie', 'Bob', 'David']
print(l)
print([s for s in l if s != 'Bob'])
print([s for s in l if s.endswith('e')])
print(list(set(l)))
l = list(range(10))
print(l)
del l[0
print(l)
del l[-1
print(l)
del l[6
print(l)
l = list(range(10))
print(l)
del l[2:5]
print(l)
l = list(range(10))
del l[:3]
print(l)
l = list(range(10))
del l[4:]
print(l)
l = list(range(10))
del l[-3:]
print(l)
l = list(range(10))
del l[:]
print(l)
l = list(range(10))
del l[2:8:2]
print(l)
l = list(range(10))
del l[::3]
print(l)
l = list(range(-5, 6))
print(l)
-5
-4
-3
-2
-1
l_square = [i**2 for i in l]
print(l_square)
l_str = [str(i) for i in l]
print(l_str)
-5
-4
-3
-2
-1
l_even = [i for i in l if i % 2 == 0]
print(l_even)
-4
-2
l_minus = [i for i in l if i < 0]
print(l_minus)
-5
-4
-3
-2
-1
l_odd = [i for i in l if not i % 2 == 0]
print(l_odd)
-5
-3
-1
l_plus = [i for i in l if not i < 0]
print(l_plus)
l_odd = [i for i in l if i % 2 != 0]
print(l_odd)
-5
-3
-1
l_plus = [i for i in l if i >= 0]
print(l_plus)
l_minus_or_even = [i for i in l if (i < 0) or (i % 2 == 0)]
print(l_minus_or_even)
-5
-4
-3
-2
-1
l_minus_and_odd = [i for i in l if (i < 0) and not (i % 2 == 0)]
print(l_minus_and_odd)
-5
-3
-1
a = 80
x = 100 if a > 50 else 0
print(x)
b = 30
y = 100 if b > 50 else 0
print(y)
l_replace = [100 if i > 0 else i for i in l]
print(l_replace)
-5
-4
-3
-2
-1
l_replace2 = [100 if i > 0 else 0 for i in l]
print(l_replace2)
l_convert = [i * 10 if i % 2 == 0 else i for i in l]
print(l_convert)
-5
-40
-3
-20
-1
l_convert2 = [i * 10 if i % 2 == 0 else i / 10 for i in l]
print(l_convert2)
-0.5
-40
-0.3
-20
-0.1
l_n = [-0.5, 0, 1.0, 100, 1.2e-2, 0xff, 0b11]
l_n_str = [str(n) for n in l_n]
print(l_n_str)
-0.5
l_i = [0, 64, 128, 192, 256]
l_i_hex1 = [hex(i) for i in l_i]
print(l_i_hex1)
l_i_hex2 = [format(i, '04x') for i in l_i]
print(l_i_hex2)
l_i_hex3 = [format(i, '#06x') for i in l_i]
print(l_i_hex3)
l_f = [0.0001, 123.456, 123400000]
l_f_e1 = [format(f, 'e') for f in l_f]
print(l_f_e1)
l_f_e2 = [format(f, '.3E') for f in l_f]
print(l_f_e2)
l_si = ['-10', '0', '100']
l_si_i = [int(s) for s in l_si]
print(l_si_i)
-10
l_sf = ['.123', '1.23', '123']
l_sf_f = [float(s) for s in l_sf]
print(l_sf_f)
l_sb = ['0011', '0101', '1111']
l_sb_i = [int(s, 2) for s in l_sb]
print(l_sb_i)
l_sbox = ['100', '0b100', '0o77', '0xff']
l_sbox_i = [int(s, 0) for s in l_sbox]
print(l_sbox_i)
l_se = ['1.23e3', '0.123e-1', '123']
l_se_f = [float(s) for s in l_se]
print(l_se_f)
def is_int(s):
int(s)
except ValueError:
return False
return True
def is_float(s):
float(s)
except ValueError:
return False
return True
l_multi = ['-100', '100', '1.23', '1.23e2', 'one']
l_multi_i = [int(s) for s in l_multi if is_int(s)]
print(l_multi_i)
-100
l_multi_f = [float(s) for s in l_multi if is_float(s)]
print(l_multi_f)
-100.0
import re
l = ['oneXXXaaa', 'twoXXXbbb', 'three999aaa', '000111222']
l_re_match = [s for s in l if re.match('.*XXX.*', s)]
print(l_re_match)
re.sub
print(l_re_sub_all)
re.sub
re.match('.*XXX.*', s)
print(l_re_sub)
l = ['oneXXXaaa', 'twoXXXbbb', 'three999aaa', '000111222']
l_in = [s for s in l if 'XXX' in s]
print(l_in)
l_in_not = [s for s in l if 'XXX' not in s]
print(l_in_not)
l_replace = [s.replace('XXX', 'ZZZ') for s in l]
print(l_replace)
l_replace_all = ['ZZZ' if 'XXX' in s else s for s in l]
print(l_replace_all)
l_start = [s for s in l if s.startswith('t')]
print(l_start)
l_start_not = [s for s in l if not s.startswith('t')]
print(l_start_not)
l_end = [s for s in l if s.endswith('aaa')]
print(l_end)
l_end_not = [s for s in l if not s.endswith('aaa')]
print(l_end_not)
l_lower = [s for s in l if s.islower()]
print(l_lower)
l_upper_all = [s.upper() for s in l]
print(l_upper_all)
l_lower_to_upper = [s.upper() if s.islower() else s for s in l]
print(l_lower_to_upper)
l_isalpha = [s for s in l if s.isalpha()]
print(l_isalpha)
l_isnumeric = [s for s in l if s.isnumeric()]
print(l_isnumeric)
l_multi = [s for s in l if s.isalpha() and not s.startswith('t')]
print(l_multi)
l_multi_or = [s for s in l if (s.isalpha() and not s.startswith('t')) or ('bbb' in s)]
print(l_multi_or)
import numpy as np
import pandas as pd
l_2d = [[0, 1, 2], [3, 4, 5]]
arr_t = np.array(l_2d).T
print(arr_t)
print(type(arr_t))
numpy.ndarray
l_2d_t = np.array(l_2d).T.tolist()
print(l_2d_t)
print(type(l_2d_t))
df_t = pd.DataFrame(l_2d).T
print(df_t)
print(type(df_t))
pandas.core.frame.DataFrame
l_2d_t = pd.DataFrame(l_2d).T.values.tolist()
print(l_2d_t)
print(type(l_2d_t))
l_2d_t_tuple = list(zip(*l_2d))
print(l_2d_t_tuple)
print(type(l_2d_t_tuple))
print(l_2d_t_tuple[0])
print(type(l_2d_t_tuple[0]))
l_2d_t = [list(x) for x in zip(*l_2d)]
print(l_2d_t)
print(type(l_2d_t))
print(l_2d_t[0])
print(type(l_2d_t[0]))
print(*l_2d)
print(list(zip([0, 1, 2], [3, 4, 5])))
print([list(x) for x in [(0, 3), (1, 4), (2, 5)]])
l = [0, 1, 2]
print(l)
print(type(l))
t = ('one', 'two', 'three')
print(t)
print(type(t))
r = range(10)
print(r)
print(type(r))
range(0, 10)
tl = list(t)
print(tl)
print(type(tl))
rl = list(r)
print(rl)
print(type(rl))
lt = tuple(l)
print(lt)
print(type(lt))
rt = tuple(r)
print(rt)
print(type(rt))
t[0] = 'ONE'
not support 
tl = list(t)
tl[0] = 'ONE'
t_new = tuple(tl)
print(t_new)
print(type(t_new))
t2 = t + ('four', 'five')
print(t)
print(t2)
t2 = t + ('four')
TypeError: can
tuple (not "str") 
t2 = t + ('four', )
print(t)
print(t2)
tl = list(t)
tl.insert(2, 'XXX')
t_new = tuple(tl)
print(t_new)
print(type(t_new))
l = [3, 3, 2, 1, 5, 1, 4, 2, 3]
print(set(l))
print(list(set(l)))
print(dict.fromkeys(l))
print(list(dict.fromkeys(l)))
print(sorted(set(l), key=l.index))
l_2d = [[1, 1], [0, 1], [0, 1], [0, 0], [1, 0], [1, 1], [1, 1]]
l_2d_unique = list(set(l_2d))
TypeError: unhashable
l_2d_unique_order = dict.fromkeys(l_2d)
TypeError: unhashable
def get_unique_list(seq):
seen = []
return [x for x in seq if x not in seen and not seen.append(x)]
print(get_unique_list(l_2d))
print(get_unique_list(l))
import collections
l = [3, 3, 2, 1, 5, 1, 4, 2, 3]
print(collections.Counter(l))
Counter({3: 3, 2: 2, 1: 2, 5: 1, 4: 1})
print([k for k, v in collections.Counter(l).items() if v > 1])
print(sorted([k for k, v in collections.Counter(l).items() if v > 1], key=l.index))
cc = collections.Counter(l)
print([x for x in l if cc[x] > 1])
def get_duplicate_list(seq):
seen = []
return [x for x in seq if not seen.append(x) and seen.count(x) == 2]
def get_duplicate_list_order(seq):
seen = []
return [x for x in seq if seq.count(x) > 1 and not seen.append(x) and seen.count(x) == 1]
print(get_duplicate_list(l_2d))
print(get_duplicate_list_order(l_2d))
print(get_duplicate_list(l))
print(get_duplicate_list_order(l))
print([x for x in l_2d if l_2d.count(x) > 1])
print(collections.Counter(l_2d))
TypeError: unhashable
import locale
print(locale.getpreferredencoding())
n = 1 + 2
+ 3
print(n)
s = 'aaa'
print(s)
s = 'https://ja.wikipedia.org/wiki/'
print(s)
s_var = 'xxx'
s = 'aaa'
SyntaxError: invalid
s = 'aaa' + s_var + 'bbb'
print(s)
s = 'aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa'
+ s_var
+ 'bbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbb'
print(s)
print(s)
+ s_var
+ 'bbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbb'
print(s)
l = [-2, -1, 0]
print(map(abs, l))
print(type(map(abs, l)))
for i in map(abs, l):
print(i)
for i in l:
print(abs(i))
print(list(map(abs, l)))
l_s = ['apple', 'orange', 'strawberry']
print(list(map(len, l_s)))
print(list(map(abs, range(-2, 1))))
l = [-2, -1, 0]
print(list(map(lambda x: x**2, l)))
def square(x):
return x**2
print(list(map(square, l)))
l_1 = [1, 2, 3]
l_2 = [10, 20, 30]
print(list(map(lambda x, y: x * y, l_1, l_2)))
print(list(map(abs, l_1, l_2)))
TypeError: abs
l_3 = [100, 200, 300, 400]
print(list(map(lambda x, y, z: x * y * z, l_1, l_2, l_3)))
l = [-2, -1, 0]
print([abs(x) for x in l])
print([x**2 for x in l])
l_1 = [1, 2, 3]
l_2 = [10, 20, 30]
print([x * y for x, y in zip(l_1, l_2)])
import numpy as np
a = np.array([-2, -1, 0])
print(np.abs(a))
print(a**2)
a_1 = np.array([1, 2, 3])
a_2 = np.array([10, 20, 30])
print(a_1 * a_2)
import csv
import glob
import os
import pprint
from bs4 import BeautifulSoup
import markdown2
import pandas as pd
def get_links_from_md(file_path, markdowner=markdown2.Markdown()):
with open(file_path) as f:
md = f.read()
html = markdowner.convert(md)
soup = BeautifulSoup(html, 'html.parser')
l = [[file_path, a.text, a.attrs.get('href')] for a in soup.find_all('a')]
return l
def get_links_from_md_in_list(file_path_list, markdowner=markdown2.Markdown()):
l = []
for path in file_path_list:
l.extend(get_links_from_md(path, markdowner))
return l
def get_links_from_md_in_dir(dir_path, markdowner=markdown2.Markdown()):
return get_links_from_md_in_list
glob.glob(os.path.join(dir_path, '**', '*.md'), recursive=True)
with open('data/src/md/test1.md') as f:
print(f.read())
www.instagram.com
and [Twitter]
twitter.com
Python.org
www.python.org
pprint.pprint(get_links_from_md('data/src/md/test1.md'))
www.instagram.com
twitter.com
Python.org
www.python.org
with open('data/src/md/test1.md') as f:
md = f.read()
markdowner = markdown2.Markdown()
html = markdowner.convert(md)
print(html)
href="https://www.instagram.com/">Instagram
href="https://twitter.com">Twitter
href="https://www.python.org/">[Py] 
Python.org
href="../test/">relative
l = BeautifulSoup(html, 'html.parser').find_all('a')
pprint.pprint(l)
href="https://www.instagram.com/">Instagram
href="https://twitter.com">Twitter
href="https://www.python.org/">[Py] 
Python.org
href="../test/">relative
a = l[0]
print(type(a))
bs4.element.Tag
print(a.attrs)
www.instagram.com
print(a.attrs.get('href'))
www.instagram.com
print(a.text)
html_en = markdowner.convert('abcde')
print(html_en)
print(BeautifulSoup(html_en, 'html.parser'))
print(BeautifulSoup(html_en, 'lxml'))
html_jp = markdowner.convert('あいうえお')
print(html_jp)
print(BeautifulSoup(html_jp, 'html.parser'))
print(BeautifulSoup(html_jp, 'lxml'))
print(BeautifulSoup(html_jp.encode(), 'lxml'))
print(BeautifulSoup('<p>abcdeあいうえお</p>', 'lxml'))
print(BeautifulSoup('<html>' + html_jp + '</html>', 'lxml'))
print(BeautifulSoup('abcde' + html_jp, 'lxml'))
pprint.pprint(get_links_from_md_in_list(glob.glob('data/src/md/*.md')))
Python.org
www.python.org
www.instagram.com
twitter.com
Python.org
www.python.org
pprint.pprint(get_links_from_md_in_dir('data/src/md/'))
Python.org
www.python.org
www.instagram.com
twitter.com
Python.org
www.python.org
www.instagram.com
twitter.com
l = get_links_from_md('data/src/md/test1.md')
l.insert(0, ['file', 'anchor text', 'URL'])
with open('data/temp/md_links_csv.csv', 'w') as f:
writer = csv.writer(f)
writer.writerows(l)
l = get_links_from_md('data/src/md/test1.md')
df = pd.DataFrame(l, columns=['file', 'anchor text', 'URL'])
print(df)
www.instagram.com
twitter.com
Python.org
www.python.org
df.to_csv('data/temp/md_links_df.csv', index=False)
import pprint
import re
file_path, p=re.compile
l = []
with open(file_path) as f:
for i, line in enumerate(f):
for result in p.findall(line):
l.append([file_path, i + 1, result[0], result[1]])
return l
pprint.pprint(get_links_from_md_regex('data/src/md/test1.md'))
www.instagram.com
twitter.com
Python.org
www.python.org
s = '[text](URL_with())'
p1 = re.compile
print(p1.findall(s))
p2 = re.compile
print(p2.findall(s))
URL_with()
s_inline = '[text](URL_with()) and [text2](URL2)'
print(p2.findall(s_inline))
URL_with()
and [text2]
p3 = re.compile
print(p3.findall(s))
URL_with()
print(p3.findall(s_inline))
URL_with()
s_jp = '[text](日本語URL)'
print(p1.findall(s_jp))
print(p2.findall(s_jp))
print(p3.findall(s_jp))
p4 = re.compile
print(p4.findall(s_jp))
s_jp_inline = '[text](日本語URL)と括弧(xxx)。'
print(p1.findall(s_jp_inline))
print(p2.findall(s_jp_inline))
print(p3.findall(s_jp_inline))
print(p4.findall(s_jp_inline))
p_title = re.compile
print(p_title.findall('[text](URL "title")'))
print(p_title.findall('[text](URL)'))
import math
print(math.ceil(10.123))
print(math.ceil(10.987))
print(type(math.ceil(10.123)))
print(math.ceil(10))
print(math.ceil('10'))
TypeError: must
not str
print(hasattr(10, '__ceil__'))
print(hasattr('10', '__ceil__'))
print(math.ceil(-10.123))
-10
print(math.ceil(-10.987))
-10
import math
print(math.copysign(123, -100))
-123.0
print(math.copysign(123.0, -100.0))
-123.0
def my_sign_with_copysign(x):
return int(math.copysign(1, x))
print(my_sign_with_copysign(100))
print(my_sign_with_copysign(-100))
-1
print(type(my_sign_with_copysign(100)))
print(my_sign_with_copysign(1.23))
print(my_sign_with_copysign(-1.23))
-1
print(type(my_sign_with_copysign(1.23)))
print(my_sign_with_copysign(0))
print(my_sign_with_copysign(0.0))
print(my_sign_with_copysign(-0.0))
-1
print(my_sign_with_copysign(float('nan')))
print(my_sign_with_copysign(float('-nan')))
-1
math.copysign
TypeError: can
TypeError: can
import math
print(math.e)
print(2**4)
print(pow(2, 4))
print(math.pow(2, 4))
math.pow
TypeError: can
print(pow(2, 4, 5))
print(2**0.5)
print(math.sqrt(2))
print(2**0.5 == math.sqrt(2))
math.sqrt
TypeError: can
print((-1)**0.5)
print(math.sqrt(-1))
ValueError: math
import cmath
cmath.sqrt
print(cmath.sqrt(-1))
print(math.exp(2))
print(math.exp(2) == math.e**2)
print(math.log(25, 5))
print(math.log(math.e))
print(math.log10(100000))
print(math.log2(1024))
import math
print(math.factorial(5))
print(math.factorial(0))
print(math.factorial(1.5))
ValueError: factorial
print(math.factorial(-1))
ValueError: factorial
not defined 
def permutations_count(n, r):
return math.factorial(n) // math.factorial(n - r)
print(permutations_count(4, 2))
print(permutations_count(4, 4))
def combinations_count(n, r):
return math.factorial(n) // (math.factorial(n - r) * math.factorial(r))
print(combinations_count(4, 2))
def combinations_with_replacement_count(n, r):
return combinations_count(n + r - 1, r)
print(combinations_with_replacement_count(4, 2))
import math
print(math.floor(10.123))
print(math.floor(10.987))
print(type(math.floor(10.123)))
print(math.floor(10))
print(math.floor('10'))
TypeError: must
not str
print(hasattr(10, '__floor__'))
print(hasattr('10', '__floor__'))
print(math.floor(-10.123))
-11
print(math.floor(-10.987))
-11
print(0.1)
print(format(0.1, '.20f'))
print(0.1 + 0.1 + 0.1)
print(0.1 + 0.1 + 0.1 == 0.3)
print((19 / 155) * (155 / 19))
print((19 / 155) * (155 / 19) == 1)
print(round(0.1 + 0.1 + 0.1, 10) == round(0.3, 10))
print(abs((0.1 + 0.1 + 0.1) - 0.3) < 1e-10)
print(1e5)
print(1e-3)
import math
print(math.isclose(0.1 + 0.1 + 0.1, 0.3))
print(math.isclose((19 / 155) * (155 / 19), 1))
print(math.isclose(1, 1.001))
print(math.isclose(1, 1.001, rel_tol=0.01))
print(math.isclose(0, 0.001))
print(math.isclose(0, 0.001, rel_tol=0.01))
print(math.isclose(0, 0.001, abs_tol=0.01))
print(math.sin(math.pi))
print(math.sin(math.pi) == 0)
print(math.isclose(math.sin(math.pi), 0))
print(math.isclose(math.sin(math.pi), 0, abs_tol=1e-10))
print(round(math.sin(math.pi), 10) == 0)
print(abs(math.sin(math.pi)) < 1e-10)
import math
print(math.modf(1.5))
print(type(math.modf(1.5)))
f, i = math.modf(1.5)
print(i)
print(f)
print(type(i))
print(type(f))
f, i = math.modf(-1.5)
print(i)
print(f)
-1.0
-0.5
f, i = math.modf(100)
print(i)
print(f)
a = 1.5
i = int(a)
f = a - int(a)
print(i)
print(f)
print(type(i))
print(type(f))
import math
print(math.pi)
print(math.degrees(math.pi))
print(math.radians(180))
sin30 = math.sin(math.radians(30))
print(sin30)
print(round(sin30, 3))
print(type(round(sin30, 3)))
print('{:.3}'.format(sin30))
print(type('{:.3}'.format(sin30)))
print(format(sin30, '.3'))
print(type(format(sin30, '.3')))
print(math.isclose(sin30, 0.5))
asin05 = math.degrees(math.asin(0.5))
print(asin05)
print(round(asin05, 3))
print(math.cos(math.radians(60)))
print(math.degrees(math.acos(0.5)))
print(math.tan(math.radians(45)))
print(math.degrees(math.atan(1)))
print(math.degrees(math.atan(0)))
print(math.degrees(math.atan(1)))
print(math.degrees(math.atan(-1)))
-45.0
print(math.degrees(math.atan(math.inf)))
print(math.degrees(math.atan(-math.inf)))
-90.0
print(math.degrees(math.atan2(0, 1)))
print(math.degrees(math.atan2(1, 1)))
print(math.degrees(math.atan2(1, 0)))
print(math.degrees(math.atan2(1, -1)))
print(math.degrees(math.atan2(0, -1)))
print(math.degrees(math.atan2(-1, -1)))
-135.0
print(math.degrees(math.atan2(-1, 0)))
-90.0
print(math.degrees(math.atan2(-1, 1)))
-45.0
print(math.degrees(math.atan2(-0.0, -1)))
-180.0
print(-1 / math.inf)
-0.0
print(-1.0 * 0.0)
-0.0
print(-0.0)
-0.0
print(-0)
print(math.degrees(math.atan2(0.0, 0.0)))
print(math.degrees(math.atan2(-0.0, 0.0)))
-0.0
print(math.degrees(math.atan2(-0.0, -0.0)))
-180.0
print(math.degrees(math.atan2(0.0, -0.0)))
print(math.sin(0.0))
print(math.sin(-0.0))
-0.0
print(math.asin(0.0))
print(math.asin(-0.0))
-0.0
print(math.tan(0.0))
print(math.tan(-0.0))
-0.0
print(math.atan(0.0))
print(math.atan(-0.0))
-0.0
print(math.atan2(0.0, 1.0))
print(math.atan2(-0.0, 1.0))
-0.0
import matplotlib as mpl
import matplotlib.pyplot as plt
import numpy as np
x = np.arange(0, 2 * np.pi, 0.1)
y = np.sin(x)
plt.plot(x, y)
plt.savefig('data/dst/matplotlib_style_default.png')
plt.show()
matplotlib.figure.Figure
plt.style.context('classic')
plt.plot(x, y)
plt.savefig('data/dst/matplotlib_style_classic.png')
plt.show()
matplotlib.figure.Figure
mpl.rcParams['axes.autolimit_mode'] = 'round_numbers'
mpl.rcParams['axes.xmargin'] = 0
mpl.rcParams['axes.ymargin'] = 0
plt.plot(x, y)
plt.savefig('data/dst/matplotlib_style_change_axes_margin.png')
plt.show()
matplotlib.figure.Figure
import numpy as np
import matplotlib.pyplot as plt
plt.subplots(nrows=2, ncols=3, sharey=True)
x = np.arange(4)
ax00.plot(x, x, 'ro--')
ax01.plot(x, x**1.5, 'g^-.')
ax02.plot(x, x**2, 'bs:')
ax10.bar(x, x + 1, width=0.5, align='center', color='r')
ax11.bar(x, x**1.5 + 1, width=0.5, align='center', color='g')
ax12.bar(x, x**2 + 1, width=0.5, align='center', color='b')
plt.rcParams['font.size'] = 10
plt.tight_layout()
plt.savefig("data/dst/matplotlib_example_multi.png")
import numpy as np
import matplotlib.pyplot as plt
x = np.linspace(0, 2.0, 20)
plt.plot(x, x, 'b:')
plt.plot(x, x**1.5, 'rs-')
plt.plot(x, x**2, color='#30F050', marker='^', markerfacecolor='blue', markeredgecolor='blue')
plt.tight_layout()
plt.savefig("data/dst/matplotlib_example_single.png")
from PIL import Image
import numpy as np
import matplotlib.pyplot as plt
im = Image.open('data/src/lena_square.png')
r = np.array(im)
flatten()
g = np.array(im)
flatten()
b = np.array(im)
flatten()
bins_range = range(0, 257, 8)
xtics_range = range(0, 257, 32)
plt.subplots(nrows=3, ncols=1, sharex=True, sharey=True)
ax0.hist(r, bins=bins_range, color='r')
ax1.hist(g, bins=bins_range, color='g')
ax2.hist(b, bins=bins_range, color='b')
plt.setp((ax0, ax1, ax2), xticks=xtics_range, xlim=(0, 256))
ax0.grid(True)
ax1.grid(True)
ax2.grid(True)
plt.savefig("data/dst/matplotlib_histogram_multi.png")
from PIL import Image
import numpy as np
import matplotlib.pyplot as plt
im = Image.open('data/src/lena_square.png')
r = np.array(im)
flatten()
g = np.array(im)
flatten()
b = np.array(im)
flatten()
bins_range = range(0, 257, 8)
xtics_range = range(0, 257, 32)
plt.hist
bins=bins_range
color=['r', 'g', 'b']
label=['Red', 'Green', 'Blue']
plt.legend(loc=2)
plt.grid(True)
plt.axis()
plt.axis([0, 256, 0, ymax])
plt.xticks(xtics_range)
plt.savefig("data/dst/matplotlib_histogram_single.png")
import matplotlib as mpl
import matplotlib.pyplot as plt
import numpy as np
print(mpl.get_configdir())
print(mpl.matplotlib_fname())
print(plt.style.available)
x = np.arange(0, 2 * np.pi, 0.1)
y = np.sin(x)
plt.plot(x, y)
plt.savefig('data/dst/matplotlib_style_default.png')
plt.show()
matplotlib.figure.Figure
plt.style.context('dark_background')
plt.plot(x, y)
plt.savefig('data/dst/matplotlib_style_dark_background.png')
plt.show()
matplotlib.figure.Figure
plt.style.context(['ggplot', 'dark_background'])
plt.plot(x, y)
plt.savefig('data/dst/matplotlib_style_ggplot_dark_background.png')
plt.show()
matplotlib.figure.Figure
plt.style.context('data/src/test.mplstyle')
plt.plot(x, y)
plt.savefig('data/dst/matplotlib_style_test.png')
plt.show()
matplotlib.figure.Figure
plt.style.context(['ggplot', 'data/src/test.mplstyle'])
plt.plot(x, y)
plt.savefig('data/dst/matplotlib_style_test_ggplot.png')
plt.show()
matplotlib.figure.Figure
plt.style.use('ggplot')
plt.plot(x, y)
plt.savefig('data/dst/matplotlib_style_ggplot.png')
plt.show()
matplotlib.figure.Figure
import numpy as np
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import axes3d
from matplotlib import cm
fig = plt.figure()
ax = fig.add_subplot(111, projection='3d')
x = y = np.arange(-15, 15, 0.5)
X, Y = np.meshgrid(x, y)
sigma = 4
Z = np.exp(-(X**2 + Y**2)/(2*sigma**2)) / (2*np.pi*sigma**2)
ax.plot_surface(X, Y, Z, rstride=1, cstride=1, cmap=cm.coolwarm)
plt.savefig("data/dst/matplotlib_mplot3d_surface.png")
ax.clear()
ax.plot_wireframe(X, Y, Z, rstride=2, cstride=2)
plt.savefig("data/dst/matplotlib_mplot3d_wireframe.png")
ax.clear()
ax.scatter(X, Y, Z, s=1)
plt.savefig("data/dst/matplotlib_mplot3d_scatter.png")
import matplotlib.pyplot as plt
import matplotlib.patches as patches
fig = plt.figure()
ax = plt.axes()
fc = face
color, ec = edge
c = patches.Circle(xy=(0, 0), radius=0.5, fc='g', ec='r')
e = patches.Ellipse(xy=(-0.25, 0), width=0.5, height=0.25, fc='b', ec='y')
r = patches.Rectangle(xy=(0, 0), width=0.25, height=0.5, ec='#000000', fill=False)
ax.add_patch(c)
ax.add_patch(e)
ax.add_patch(r)
plt.axis('scaled')
ax.set_aspect('equal')
plt.savefig('data/dst/matplotlib_patches.png')
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
sns.set()
sns.set_style("whitegrid", {'grid.linestyle': '--'})
sns.set_context("paper", 1.5, {"lines.linewidth": 4})
sns.set_palette("winter_r", 8, 1)
sns.set
font_scale=1.5
rc={"lines.linewidth": 2, 'grid.linestyle': '--'}
x = np.arange(0, 2.1, 0.1)
plt.plot(x, x)
plt.plot(x, x**1.5)
plt.plot(x, x**2)
plt.savefig('data/dst/matplotlib_seaborn_set_all.png')
l = [3, 6, 7, -1, 23, -10, 18]
print(max(l))
print(min(l))
-10
ld = sorted(l, reverse=True)
print(ld)
-1
-10
print(ld[:3])
la = sorted(l)
print(la)
-10
-1
print(la[:3])
-10
-1
print(sorted(l, reverse=True)[:3])
print(sorted(l)[:3])
-10
-1
print(l)
-1
-10
l.sort(reverse=True)
print(l[:3])
print(l)
-1
-10
l.sort()
print(l[:3])
-10
-1
print(l)
-10
-1
import mimetypes
print(mimetypes.guess_type('test.jpg'))
print(type(mimetypes.guess_type('test.jpg')))
print(mimetypes.guess_type('test.jpg')[0])
print(type(mimetypes.guess_type('test.jpg')[0]))
print(mimetypes.guess_type('test.tar.gz'))
print(mimetypes.guess_type('dir/test.txt'))
print(mimetypes.guess_type('https://xxx.com/test.html'))
print(mimetypes.guess_type('test.JPG'))
print(mimetypes.guess_all_extensions('image/jpeg'))
print(mimetypes.guess_extension('image/jpeg'))
print(mimetypes.guess_all_extensions('text/plain'))
print(mimetypes.guess_extension('text/plain'))
print(mimetypes.guess_all_extensions('image/jpg'))
print(mimetypes.guess_extension('image/jpg'))
if not mimetypes.inited:
mimetypes.init()
print(type(mimetypes.types_map))
print(len(mimetypes.types_map))
print(mimetypes.types_map['.jpg'])
import pprint
pprint.pprint(mimetypes.types_map)
gpp2.tcap
print(bool.__mro__)
print(type(True).__mro__)
print(type(bool.__mro__))
print(type(bool.__mro__[0]))
print([c.__name__ for c in bool.__mro__])
import pprint
pprint.pprint(ZeroDivisionError.__mro__)
BaseA()
pass
BaseB()
pass
Sub(BaseA, BaseB)
pass
pprint.pprint(Sub.__mro__)
__main__.Sub
__main__.BaseA
__main__.BaseB
print(bool.__bases__)
print(type(bool.__bases__))
print(type(bool.__bases__[0]))
print(object.__bases__)
print(Sub.__bases__)
__main__.BaseA
__main__.BaseB
print(bool.__base__)
print(Sub.__base__)
__main__.BaseA
a = 100
b = 200
print(a)
print(b)
a, b = 100, 200
print(a)
print(b)
a, b, c = 0.1, 100, 'string'
print(a)
print(b)
print(c)
a = 100
print(a)
print(type(a))
a, b = 100, 200, 300
ValueError: too
a, b, c = 100, 200
ValueError: not
b = 100
print(a)
print(type(a))
print(b)
print(type(b))
a, b = 100, 200, 300
print(a)
print(type(a))
print(b)
print(type(b))
a = b = 100
print(a)
print(b)
a = 200
print(a)
print(b)
a = b = c = 'string'
print(a)
print(b)
print(c)
a = b = [0, 1, 2]
print(a is b)
a[0] = 100
print(a)
print(b)
b = [0, 1, 2]
a = b
print(a is b)
a[0] = 100
print(a)
print(b)
a = [0, 1, 2]
b = [0, 1, 2]
print(a is b)
a[0] = 100
print(a)
print(b)
def my_round(val, digit=0):
p = 10 ** digit
return (val * p * 2 + 1) // 2 / p
my_round_int = lambda x: int((x * 2 + 1) // 2)
f = 123.456
print(int(my_round(f)))
print(my_round_int(f))
print(my_round(f, 1))
print(my_round(f, 2))
print(int(my_round(0.4)))
print(int(my_round(0.5)))
print(int(my_round(0.6)))
i = 99518
print(int(my_round(i, -1)))
print(int(my_round(i, -2)))
print(int(my_round(i, -3)))
print(int(my_round(4, -1)))
print(int(my_round(5, -1)))
print(int(my_round(6, -1)))
print(int(my_round(-0.4)))
print(int(my_round(-0.5)))
print(int(my_round(-0.6)))
-1
import math
def my_round2(val, digit=0):
p = 10 ** digit
s = math.copysign(1, val)
return (s * val * p * 2 + 1) // 2 / p * s
print(int(my_round2(-0.4)))
print(int(my_round2(-0.5)))
print(int(my_round2(-0.6)))
-1
-1
import numpy as np
print(np.sign(100))
print(np.sign(-100))
-1
print(type(np.sign(100)))
numpy.int64
print(np.sign(1.23))
print(np.sign(-1.23))
-1.0
print(type(np.sign(1.23)))
numpy.float64
print(np.sign(0))
print(np.sign(0.0))
print(np.sign(-0.0))
print(np.sign(float('nan')))
print(np.sign(float('-nan')))
print(100 > 0)
print(True - False)
print(False - True)
-1
print(False - False)
def my_sign(x):
return (x > 0) - (x < 0)
print(my_sign(100))
print(my_sign(-100))
-1
print(type(my_sign(100)))
print(my_sign(1.23))
print(my_sign(-1.23))
-1
print(type(my_sign(-1.23)))
print(my_sign(0))
print(my_sign(0.0))
print(my_sign(-0.0))
print(my_sign(float('nan')))
print(my_sign(float('-nan')))
not supported 
def my_sign_with_abs(x):
return 0.0 if abs(x) == 0 else x / abs(x)
print(my_sign_with_abs(100))
print(my_sign_with_abs(-100))
-1.0
print(type(my_sign_with_abs(100)))
print(my_sign_with_abs(1.23))
print(my_sign_with_abs(-1.23))
-1.0
print(type(my_sign_with_abs(1.23)))
print(my_sign_with_abs(0))
print(my_sign_with_abs(0.0))
print(my_sign_with_abs(-0.0))
print(my_sign_with_abs(float('nan')))
print(my_sign_with_abs(float('-nan')))
import math
import numpy as np
print(math.__name__)
print(np.__name__)
import hello
print(hello.__name__)
hello.func()
collections.namedtuple
from collections import namedtuple
Point = namedtuple('Point', ['x', 'y'])
print(Point)
__main__.Point
p1 = Point(1, 2)
print(p1)
print(type(p1))
Point(x=1, y=2)
__main__.Point
print(p1.x, p1.y)
print(p1[0], p1[1])
p_new = p1._replace(x=10)
print(p1)
print(p_new)
Point(x=1, y=2)
Point(x=10, y=2)
p_d = p1._asdict()
print(p_d)
print(type(p_d))
OrderedDict([('x', 1), ('y', 2)])
collections.OrderedDict
print(Point._fields)
Point_3D = namedtuple('Point_3d', Point._fields + ('z', ))
print(Point_3D)
__main__.Point_3d
p_3d = Point_3D(1, 2, 3)
print(p_3d)
print(type(p_3d))
Point_3d(x=1, y=2, z=3)
__main__.Point_3d
p2 = Point(3, 4)
print(p2)
Point(x=3, y=4)
p = p1 + p2
print(p)
print(type(p))
Point_a(namedtuple('Point_a', ['x', 'y']))
def __add__(self, other):
return Point_a(self.x + other.x, self.y + other.y)
p1_a = Point_a(1, 2)
p2_a = Point_a(3, 4)
print(p1_a)
print(p2_a)
Point_a(x=1, y=2)
Point_a(x=3, y=4)
p_a = p1_a + p2_a
print(p_a)
Point_a(x=4, y=6)
bin_num = 0b10
oct_num = 0
hex_num = 0x10
print(bin_num)
print(oct_num)
print(hex_num)
Bin_num = 0B10
Oct_num = 0
Hex_num = 0X10
print(Bin_num)
print(Oct_num)
print(Hex_num)
print(type(bin_num))
print(type(oct_num))
print(type(hex_num))
print(type(Bin_num))
print(type(Oct_num))
print(type(Hex_num))
result = 0b10 * 0
print(result)
print(0b111111111111 == 0b1_1_1_1_1_1_1_1_1_1_1_1)
bin_num = 0b1111_1111_1111
print(bin_num)
i = 255
print(bin(i))
print(oct(i))
print(hex(i))
print(type(bin(i)))
print(type(oct(i)))
print(type(hex(i)))
print(bin(i)[2:])
print(oct(i)[2:])
print(hex(i)[2:])
print(str(i))
print(type(str(i)))
print(format(i, 'b'))
print(format(i, 'o'))
print(format(i, 'x'))
print(type(format(i, 'b')))
print(type(format(i, 'o')))
print(type(format(i, 'x')))
print(format(i, '#b'))
print(format(i, '#o'))
print(format(i, '#x'))
print(format(i, '08b'))
print(format(i, '08o'))
print(format(i, '08x'))
print(format(i, '#010b'))
print(format(i, '#010o'))
print(format(i, '#010x'))
print('{:08b}'.format(i))
print('{:08o}'.format(i))
print('{:08x}'.format(i))
print(f'{i:08b}')
print(f'{i:08o}')
print(f'{i:08x}')
print(int('10'))
print(int('10', 2))
print(int('10', 8))
print(int('10', 16))
print(type(int('10')))
print(type(int('10', 2)))
print(type(int('10', 8)))
print(type(int('10', 16)))
print(int('0b10', 0))
print(int('0o10', 0))
print(int('0x10', 0))
print(int('0B10', 0))
print(int('0O10', 0))
print(int('0X10', 0))
print(int('10', 0))
print(int('010', 0))
ValueError: invalid
int() 
print(int('010'))
print(int('00ff', 16))
print(int('0x00ff', 0))
print(int('ff', 2))
ValueError: invalid
int() 
print(int('0a10', 0))
ValueError: invalid
int() 
print(int('0bff', 0))
ValueError: invalid
int() 
a = '0b1001'
b = '0b0011'
c = int(a, 0) + int(b, 0)
print(c)
print(bin(c))
a_0b = '0b1110001010011'
print(format(int(a, 0), '#010x'))
print(format(int(a, 0), '#010o'))
import numpy as np
a = np.arange(6)
print(a)
print(a.reshape(2, 3))
print(a.reshape(-1, 3))
print(a.reshape(2, -1))
print(a.reshape(3, 4))
ValueError: cannot
shape (3,4)
print(a.reshape(-1, 4))
ValueError: cannot
shape (4)
l = [0, 1, 2, 3, 4, 5]
print(np.array(l).reshape(-1, 3).tolist())
print(np.array(l).reshape(3, -1).tolist())
import numpy as np
a = np.array([0.3, 0.1 + 0.1 + 0.1])
print(a)
b = np.array([0.3, 0.3])
print(b)
c = np.array([0.2, 0.3])
print(c)
print(np.allclose(a, b))
print(np.allclose(a, c))
a_100 = np.array([99, 100, 101])
print(a_100)
print(np.allclose(a_100, 100))
print(np.allclose(a_100, 100, rtol=0, atol=1))
a_nan = np.array([np.nan, 1, 2])
print(a_nan)
b_nan = np.array([np.nan, 1, 2])
print(b_nan)
print(np.allclose(a_nan, b_nan))
print(np.allclose(a_nan, b_nan, equal_nan=True))
import numpy as np
a = np.arange(3)
print(a)
a_append = np.append(a, 3)
print(a_append)
print(a)
print(np.append(a, [3, 4, 5]))
print(np.append(a, np.arange(3, 6)))
print(np.append(-1, a))
-1
print(np.append([-3, -2, -1], a))
print(np.append(np.arange(-3, 0), a))
a_2d = np.arange(6).reshape(2, 3)
print(a_2d)
print(np.append(a_2d, 10))
a_2d_ex = np.arange(6).reshape(2, 3) * 10
print(a_2d_ex)
print(np.append(a_2d, a_2d_ex))
print(np.append(a_2d, a_2d_ex, axis=0))
print(np.append(a_2d, a_2d_ex, axis=1))
print(np.append(a_2d, a_2d_ex, axis=2))
AxisError: axis
a_2d_ex2 = np.arange(2).reshape(1, 2) * 10
print(a_2d_ex2)
print(np.append(a_2d, a_2d_ex2, axis=0))
ValueError: all
print(np.append(a_2d, a_2d_ex2, axis=1))
ValueError: all
print(np.append(a_2d_ex, a_2d, axis=0))
print(np.append(a_2d_ex, a_2d, axis=1))
print(np.append(a_2d, [[0, 10, 20], [30, 40, 50]], axis=0))
a_row_1d = np.arange(3) * 10
print(a_row_1d)
print(np.append(a_2d, a_row_1d, axis=0))
ValueError: all
a_row_2d = np.arange(3).reshape(1, 3) * 100
print(a_row_2d)
print(np.append(a_2d, a_row_2d, axis=0))
print(np.append(a_2d, a_row_1d.reshape(1, 3), axis=0))
a_col_1d = np.arange(2) * 10
print(a_col_1d)
print(np.append(a_2d, a_col_1d, axis=1))
ValueError: all
a_col_2d = np.arange(2).reshape(2, 1) * 100
print(a_col_2d)
print(np.append(a_2d, a_col_2d, axis=1))
print(np.append(a_2d, a_col_1d.reshape(2, 1), axis=1))
print(np.vstack([a_2d, a_row_1d]))
print(np.vstack([a_row_2d, a_2d, a_row_1d, [[0, -1, -2], [-3, -4, -5]]]))
print(np.hstack([a_2d, a_col_1d]))
ValueError: all
print(np.hstack([a_col_2d, a_2d, a_col_1d.reshape(2, 1), [[0, -1], [-2, -3]]]))
a_3d = np.arange(12).reshape(2, 3, 2)
print(a_3d)
print(np.append(a_3d, 100))
a_3d_ex = np.arange(12).reshape(2, 3, 2) * 10
print(a_3d_ex)
print(a_3d_ex.shape)
print(np.append(a_3d, a_3d_ex, axis=0))
print(np.append(a_3d, a_3d_ex, axis=0).shape)
print(np.append(a_3d, a_3d_ex, axis=1))
print(np.append(a_3d, a_3d_ex, axis=1).shape)
print(np.append(a_3d, a_3d_ex, axis=2))
print(np.append(a_3d, a_3d_ex, axis=2).shape)
print(np.concatenate([a_3d_ex, a_3d, a_3d_ex], axis=2))
import numpy as np
print(np.arange(3))
print(np.arange(3, 10))
print(np.arange(3, 10, 2))
print(np.arange(0.3, 1.0, 0.2))
print(np.arange(-3, 3))
print(np.arange(10, 3))
print(np.arange(10, 3, -2))
print(np.arange(3, dtype=float))
print(np.arange(3, 10, dtype=float))
print(np.arange(3, 10, 2, dtype=float))
print(np.arange(3, 10, 2))
print(np.arange(9, 2, -2))
print(np.arange(3, 10, 2)[::-1])
print(np.flip(np.arange(3, 10, 2)))
print(np.arange(12).reshape(3, 4))
print(np.arange(24).reshape(2, 3, 4))
import numpy as np
a = np.array([1, 100, 10])
print(a)
print(np.argmax(a))
a = np.array([1, 10, 10])
print(a)
print(np.argmax(a))
a_2d = np.array([[20, 50, 30], [60, 40, 10]])
print(a_2d)
print(np.argmax(a_2d))
print(a_2d.flatten())
print(np.argmax(a_2d.flatten()))
print(np.argmax(a_2d, axis=0))
print(np.max(a_2d, axis=0))
print(np.argmax(a_2d, axis=1))
print(np.max(a_2d, axis=1))
idx = np.unravel_index(np.argmax(a_2d), a_2d.shape)
print(idx)
print(a_2d[idx])
print(np.max(a_2d))
a = np.array([1, 100, 10])
print(a)
print(a.argmax())
print(a.max())
a_2d = np.array([[20, 50, 30], [60, 40, 10]])
print(a_2d)
print(a_2d.argmax())
print(a_2d.argmax(axis=0))
print(a_2d.max(axis=0))
print(a_2d.argmax(axis=1))
print(a_2d.max(axis=1))
print(np.unravel_index(a_2d.argmax(), a_2d.shape))
import numpy as np
a = np.array([1, 100, 10])
print(a)
print(np.argmin(a))
print(a.argmin())
a_2d = np.array([[20, 50, 30], [60, 40, 10]])
print(a_2d)
print(np.argmin(a_2d))
print(a_2d.argmin())
print(np.argmin(a_2d, axis=0))
print(a_2d.argmin(axis=0))
print(np.min(a_2d, axis=0))
print(a_2d.min(axis=0))
print(np.argmin(a_2d, axis=1))
print(a_2d.argmin(axis=1))
print(np.min(a_2d, axis=1))
print(a_2d.min(axis=1))
idx = np.unravel_index(np.argmin(a_2d), a_2d.shape)
print(idx)
print(a_2d[idx])
print(np.min(a_2d))
print(np.unravel_index(a_2d.argmin(), a_2d.shape))
import numpy as np
a = np.arange(3)
print(a)
b = np.arange(3)
print(b)
c = np.arange(1, 4)
print(c)
print(np.all(a == b))
print(np.all(a == c))
print(np.array_equal(a, b))
print(np.array_equal(a, c))
print(np.array_equiv(a, b))
print(np.array_equiv(a, c))
b_f = np.arange(3, dtype=float)
print(b_f)
print(np.array_equal(a, b_f))
print(np.array_equiv(a, b_f))
ones = np.ones(3)
print(ones)
print(np.array_equal(ones, 1))
print(np.array_equiv(ones, 1))
a_2d = np.array([[0, 1, 2], [0, 1, 2], [0, 1, 2]])
print(a_2d)
print(np.array_equal(a_2d, b))
print(np.array_equiv(a_2d, b))
a_nan = np.array([np.nan, 1, 2])
print(a_nan)
b_nan = np.array([np.nan, 1, 2])
print(b_nan)
print(np.array_equal(a_nan, b_nan))
print(np.array_equiv(a_nan, b_nan))
print(np.all(a_nan == b_nan))
import numpy as np
a = np.arange(15).reshape(3, 5)
print(a)
np.split(a, 2, 0)
ValueError: array
not result in an 
a0, a1 = np.array_split(a, 2, 0)
print(a0)
print(a1)
a0, a1, a2 = np.array_split(a, 3, 1)
print(a0)
print(a1)
print(a2)
import numpy as np
a = np.array([1, 2, 3])
print(a)
print(a.dtype)
a_float = a.astype(np.float32)
print(a_float)
print(a_float.dtype)
print(a)
print(a.dtype)
a_float = a.astype(float)
print(a_float)
print(a_float.dtype)
a_str = a.astype('str')
print(a_str)
print(a_str.dtype)
a_int = a.astype('int32')
print(a_int)
print(a_int.dtype)
a = np.arange(50).reshape((5, 10)) / 10 - 2
print(a)
print(a.dtype)
a_int = a.astype('int64')
print(a_int)
print(a_int.dtype)
-1
print(np.round(a).astype(int))
my_round_int = lambda x: np.round((x * 2 + 1) // 2)
print(my_round_int(a).astype(int))
def my_round(x, digit=0):
p = 10 ** digit
s = np.copysign(1, x)
return (s * x * p * 2 + 1) // 2 / p * s
print(my_round(a).astype(int))
import numpy as np
print(np.__version__)
a_int = np.array([0, 1, 3])  
b_int = np.array([1, 0, 2])  
print(a_int.dtype)
print(b_int.dtype)
print(a_int & b_int)
print(a_int | b_int)
print(a_int ^ b_int)
print(a_int << b_int)
print(a_int >> b_int)
print(np.bitwise_and(a_int, b_int))
print(np.bitwise_or(a_int, b_int))
print(np.bitwise_xor(a_int, b_int))
print(np.bitwise_not(a_int))
print(np.invert(a_int))
print(-(a_int + 1))
a_uint8 = np.array([0, 1, 3], dtype=np.uint8)
a_uint16 = np.array([0, 1, 3], dtype=np.uint16)
c_int_2d = np.arange(6).reshape(2, 3)
print(c_int_2d)
print(c_int_2d & a_int)
print(np.bitwise_and(c_int_2d, a_int))
print(c_int_2d & 2)
print(np.bitwise_and(c_int_2d, 2))
d_float = np.array([0, 1, 3], dtype=float)
print(a_int & d_float)
TypeError: ufunc
not supported for the input
not be 
TypeError: ufunc
not supported for the input
not be 
print(np.bitwise_and(a_int, d_float))
TypeError: ufunc
not supported for the input
not be 
print(np.bitwise_not(d_float))
TypeError: ufunc
not supported for the input
not be 
e_bool = np.array([True, False, True])
print(a_int & e_bool)
print((a_int & e_bool).dtype)
print(np.bitwise_and(a_int, e_bool))
print(np.bitwise_and(a_int, e_bool).dtype)
print(np.logical_not(e_bool))
print(np.bitwise_not(e_bool))
print(np.invert(e_bool))
import numpy as np
a1 = np.ones((2, 3), int)
print(a1)
a2 = np.full((2, 3), 2)
print(a2)
print(np.block([a1, a2]))
print(np.block([[a1], [a2]]))
print(np.block([[a1, a2], [a2, a1]]))
print(np.block([[[a1]], [[a2]]]))
print(np.block([[[a1]], [[a2]]]).shape)
a3 = np.full(6, 3)
print(a3)
print(np.block([[a1, a2], [a3]]))
print(np.block([[a1, a2], a3]))
ValueError: List
mismatched. First
1 (arrays[1])
print(np.block([[a1, a2, a3]]))
ValueError: all
import numpy as np
a = np.arange(3)
print(a)
print(a.shape)
b = np.arange(3).reshape(3, 1)
print(b)
print(b.shape)
arrays = np.broadcast_arrays(a, b)
print(type(arrays))
print(len(arrays))
print(arrays[0])
print(arrays[1])
print(type(arrays[0]))
numpy.ndarray
c = np.zeros((2, 2))
print(c)
print(c.shape)
arrays = np.broadcast_arrays(a, c)
ValueError: shape
mismatch: objects
import numpy as np
a = np.arange(3)
print(a)
print(a.shape)
print(np.broadcast_to(a, (3, 3)))
print(type(np.broadcast_to(a, (3, 3))))
numpy.ndarray
print(np.broadcast_to(a, (2, 2)))
ValueError: operands
not be 
shape (2,2)
import numpy as np
a = np.zeros((3, 3), np.int)
print(a)
print(a.shape)
b = np.arange(3)
print(b)
print(b.shape)
print(a + b)
b_1_3 = b.reshape(1, 3)
print(b_1_3)
print(b_1_3.shape)
print(np.tile(b_1_3, (3, 1)))
b_3_1 = b.reshape(3, 1)
print(b_3_1)
print(b_3_1.shape)
print(a + b_3_1)
print(np.tile(b_3_1, (1, 3)))
print(b_1_3)
print(b_1_3.shape)
print(b_3_1)
print(b_3_1.shape)
print(b_1_3 + b_3_1)
print(np.tile(b_1_3, (3, 1)))
print(np.tile(b_3_1, (1, 3)))
print(np.tile(b_1_3, (3, 1)) + np.tile(b_3_1, (1, 3)))
c = np.arange(4)
print(c)
print(c.shape)
print(b_3_1)
print(b_3_1.shape)
print(c + b_3_1)
print(np.tile(c.reshape(1, 4), (3, 1)))
print(np.tile(b_3_1, (1, 4)))
print(np.tile(c.reshape(1, 4), (3, 1)) + np.tile(b_3_1, (1, 4)))
import numpy as np
a = np.zeros((2, 3, 4), dtype=np.int)
print(a)
print(a.shape)
b = np.arange(4)
print(b)
print(b.shape)
print(a + b)
b_1_1_4 = b.reshape(1, 1, 4)
print(b_1_1_4)
print(np.tile(b_1_1_4, (2, 3, 1)))
import numpy as np
a = np.zeros((4, 3), dtype=np.int)
print(a)
print(a.shape)
b = np.arange(6).reshape(2, 3)
print(b)
print(b.shape)
print(a + b)
ValueError: operands
not be 
shapes (4,3) (2,3) 
a = np.zeros((2, 3, 4), dtype=np.int)
print(a)
print(a.shape)
b = np.arange(3)
print(b)
print(b.shape)
print(a + b)
ValueError: operands
not be 
shapes (2,3,4) 
b_3_1 = b.reshape(3, 1)
print(b_3_1)
print(b_3_1.shape)
print(a + b_3_1)
import numpy as np
a = b = np.arange(3)
print(a)
print(a.shape)
print(a * b)
print(np.multiply(a, b))
a_1_3 = a.reshape(1, 3)
print(a_1_3)
print(a_1_3.shape)
b_3_1 = b.reshape(3, 1)
print(b_3_1)
print(b_3_1.shape)
print(a_1_3 * b_3_1)
print(np.multiply(a_1_3, b_3_1))
print(a * b_3_1)
print(np.multiply(a, b_3_1))
print(np.matmul(a_1_3, b_3_1))
print(np.dot(a_1_3, b_3_1))
numpy.ndarray
print(np.matmul(a_1_3, b))
print(np.dot(a_1_3, b))
numpy.ndarray
print(np.matmul(a, b))
print(np.dot(a, b))
numpy.int64
a = np.arange(6).reshape(2, 3)
print(a)
b = np.arange(2).reshape(1, 2)
print(b)
ValueError: shapes
and (1,2) 
not aligned
print(np.tile(b, (3, 1)))
np.tile(b, (3, 1))
import numpy as np
a = np.arange(10)
print(a)
print(np.clip(a, 2, 7))
print(np.clip(a, None, 7))
print(np.clip(a, 2, None))
print(np.clip(a, 2))
TypeError: clip
a_clip = np.clip(a, 2, 7)
print(a_clip)
print(a)
print(a.clip(2, 7))
print(a.clip(None, 7))
print(a.clip(2, None))
print(a.clip(2))
print(a.clip(min=2))
print(a.clip(max=7))
a_clip = a.clip(2, 7)
print(a_clip)
print(a)
import numpy as np
a = np.arange(12).reshape(3, 4)
print(a)
b = np.arange(12).reshape(4, 3).T
print(b)
a_compare = a < b
print(a_compare)
print(type(a_compare))
numpy.ndarray
print(a_compare.dtype)
print(a > b)
print(a <= b)
print(a >= b)
print(a == b)
print(a != b)
b_float = b.astype(float)
print(b_float)
print(b_float.dtype)
print(a == b_float)
b_1d = np.arange(4, 8)
print(b_1d)
print(a < b_1d)
print(a < 6)
print(a % 2)
print(a % 2 == 0)
print(np.count_nonzero(a < 6))
print(np.all(a < 6))
print(np.all(a < 6, axis=1))
print(np.any(a < 6))
print(np.any(a < 6, axis=1))
a_nan = np.array([0, 1, np.nan])
print(a_nan)
0.  1. nan
print(a_nan == np.nan)
print(np.isnan(a_nan))
print(a_nan > 0)
RuntimeWarning: invalid
a_nan_only = np.array([np.nan])
print(a_nan_only)
print(a_nan_only > 0)
print(np.nan > 0)
x = 6
print(4 < x < 8)
print(4 < a < 8)
ValueError: The
a.any() or a.all()
print((a > 4) & (a < 8))
print((a > 4) and (a < 8))
ValueError: The
a.any() or a.all()
print(a > 4 & a < 8)
ValueError: The
a.any() or a.all()
x = 6
y = 6
z = 6
print(x == y == z)
c = np.zeros((3, 4), int)
print(c)
print(a == b == c)
ValueError: The
a.any() or a.all()
print((a == b) & (b == c))
import numpy as np
a1 = np.ones((2, 3), int)
print(a1)
a2 = np.full((2, 3), 2)
print(a2)
print(np.concatenate([a1, a2]))
a3 = np.full((2, 3), 3)
print(a3)
print(np.concatenate([a1, a2, a3]))
print(np.concatenate([a1, a2], 0))
print(np.concatenate([a1, a2], 1))
print(np.concatenate([a1, a2], 2))
AxisError: axis
a2_ = np.full((3, 3), 2)
print(a2_)
print(np.concatenate([a1, a2_], 0))
print(np.concatenate([a1, a2_], 1))
ValueError: all
a1 = np.ones(3, int)
print(a1)
a2 = np.full(3, 2)
print(a2)
print(np.concatenate([a1, a2], 0))
print(np.concatenate([a1, a2], 1))
AxisError: axis
a1 = np.ones((2, 3), int)
print(a1)
a2 = np.full(3, 2)
print(a2)
print(np.concatenate([a1, a2], 0))
ValueError: all
import numpy as np
a = np.arange(12).reshape((3, 4))
print(a)
print(a < 5)
print(a[a < 5])
print(a < 10)
print(a[a < 10])
b = a[a < 10]
print(b)
print(a)
print(a[a < 5].sum())
print(a[a < 5].mean())
print(a[a < 5].max())
print(a[a < 10].min())
print(a[a < 10].std())
print(a < 5)
print(np.all(a < 5))
print(np.all(a < 5, axis=0))
print(np.all(a < 5, axis=1))
print(a < 10)
print(np.all(a < 10, axis=0))
print(np.all(a < 10, axis=1))
print(a[:, np.all(a < 10, axis=0)])
print(a[np.all(a < 10, axis=1), :])
print(a[np.all(a < 10, axis=1)])
print(a[:, np.all(a < 5, axis=0)])
print(a[np.all(a < 5, axis=1)])
print(a[np.all(a < 5, axis=1)].ndim)
print(a[np.all(a < 5, axis=1)].shape)
print(a < 5)
print(np.any(a < 5))
print(np.any(a < 5, axis=0))
print(np.any(a < 5, axis=1))
print(a[:, np.any(a < 5, axis=0)])
print(a[np.any(a < 5, axis=1)])
print(a[:, np.all(a < 10, axis=0)])
np.all(a < 10, axis=0)
print(a[np.any(a < 5, axis=1)])
np.any(a < 5, axis=1)
print(a)
print(np.delete(a, [0, 2], axis=0))
print(np.delete(a, [0, 2], axis=1))
print(a < 2)
print(np.where(a < 2))
array([0, 0])
array([0, 1])
print(np.where(a < 2)[0])
print(np.where(a < 2)[1])
print(np.delete(a, np.where(a < 2)[0], axis=0))
print(np.delete(a, np.where(a < 2)[1], axis=1))
print(a == 6)
print(np.where(a == 6))
array([1])
array([2])
print(np.delete(a, np.where(a == 6)))
print(np.delete(a, np.where(a == 6)[0], axis=0))
print(np.delete(a, np.where(a == 6)[1], axis=1))
print(a[(a < 10) & (a % 2 == 1)])
print(a[np.any((a == 2) | (a == 10), axis=1)])
np.any((a == 2) | (a == 10), axis=0)
import numpy as np
a = np.arange(12).reshape(3, 4)
print(a)
b = np.arange(-5, 7).reshape(3, 4)
print(b)
-1
a_copysign = np.copysign(a, b)
print(a_copysign)
-4.  5
print(a_copysign.dtype)
print(np.copysign(10, -5))
-10.0
print(type(np.copysign(10, -5)))
numpy.float64
print(a)
b_small = np.array([-100, -100, 100, 100])
print(b_small)
print(a + b_small)
print(np.copysign(a, b_small))
b_mismatch = np.array([-100, -100, 100])
print(b_mismatch)
print(np.copysign(a, b_mismatch))
ValueError: operands
not be 
shapes (3,4) 
print(np.copysign(b, -10))
print(np.abs(b) * -1)
-1
print(np.abs(b) * -1.0)
a_special = np.array([0.0, -0.0, np.inf, -np.inf, np.nan])
print(a_special)
print(np.copysign(a_special, 1))
0.  0. inf
print(np.copysign(a_special, -1))
print(np.copysign([10, 10, 10, 10, 10], a_special))
-10.  10
print(np.copysign([-10, -10, -10, -10, -10], a_special))
-10.  10
print(np.copysign(10, 0))
print(np.copysign(0, 10))
print(np.copysign(0, -10))
-0.0
a_complex = np.array
print(a_complex)
print(np.copysign(a_complex, 1))
TypeError: ufunc
not supported for the input
not be 
print(np.copysign([1, 1], a_complex))
TypeError: ufunc
not supported for the input
not be 
import numpy as np
a = np.arange(12).reshape((3, 4))
print(a)
print(a < 4)
print(a % 2 == 1)
print(np.count_nonzero(a < 4))
print(np.count_nonzero(a % 2 == 1))
print(np.sum(a < 4))
print(np.sum(a % 2 == 1))
print(np.count_nonzero(a < 4, axis=0))
print(np.count_nonzero(a < 4, axis=1))
print(np.count_nonzero(a % 2 == 1, axis=0))
print(np.count_nonzero(a % 2 == 1, axis=1))
print(np.any(a < 4))
print(np.any(a > 100))
print(np.any(a < 4, axis=0))
print(np.any(a < 4, axis=1))
print(np.all(a < 4))
print(np.all(a < 100))
print(np.all(a < 4, axis=0))
print(np.all(a < 4, axis=1))
print((a < 4) | (a % 2 == 1))
print(np.count_nonzero((a < 4) | (a % 2 == 1)))
print(np.count_nonzero((a < 4) | (a % 2 == 1), axis=0))
print(np.count_nonzero((a < 4) | (a % 2 == 1), axis=1))
import numpy as np
a_inf = np.array([-np.inf, 0, np.inf])
print(a_inf)
-inf
print(np.isinf(a_inf))
print(a_inf == np.inf)
print(a_inf == -np.inf)
import numpy as np
a_nan = np.genfromtxt('data/src/sample_nan.csv', delimiter=',')
print(a_nan)
11. 12. nan
print(np.nan == np.nan)
print(a_nan == np.nan)
print(np.isnan(a_nan))
print(np.count_nonzero(np.isnan(a_nan)))
print(np.count_nonzero(np.isnan(a_nan), axis=0))
print(np.count_nonzero(np.isnan(a_nan), axis=1))
np.isnan(a_nan)
import numpy as np
print(np.__version__)
a_1d = np.array([1, 2, 3, 4, 5, 6])
print(a_1d)
print(np.cumsum(a_1d))
print(np.cumsum(a_1d, dtype=float))
print(a_1d.cumsum())
print(a_1d.cumsum(dtype=float))
l = [1, 2, 3, 4, 5, 6]
print(np.cumsum(l))
print(type(np.cumsum(l)))
numpy.ndarray
a_2d = a_1d.reshape(2, 3)
print(a_2d)
print(np.cumsum(a_2d))
print(np.cumsum(a_2d, axis=0))
print(np.cumsum(a_2d, axis=1))
print(a_2d.cumsum())
print(a_2d.cumsum(axis=0))
print(a_2d.cumsum(axis=1))
l_2d = [[1, 2, 3], [4, 5, 6]]
print(np.cumsum(l_2d))
print(np.cumsum(l_2d, axis=0))
print(np.cumsum(l_2d, axis=1))
l_2d_error = [[1, 2, 3], [4, 5]]
print(np.cumsum(l_2d_error))
list([1, 2, 3]) 
list([1, 2, 3, 4, 5])
VisibleDeprecationWarning: Creating
lengths or shapes
deprecated. If
dtype=object
return array(a, dtype, copy=False, order=order)
print(np.cumprod(a_1d))
print(np.cumprod(a_1d, dtype=float))
print(a_1d.cumprod())
print(a_1d.cumprod(dtype=float))
print(np.cumprod(a_2d))
print(np.cumprod(a_2d, axis=0))
print(np.cumprod(a_2d, axis=1))
print(a_2d.cumprod())
print(a_2d.cumprod(axis=0))
print(a_2d.cumprod(axis=1))
print(np.cumprod(l))
print(np.cumprod(l_2d))
print(np.cumprod(l_2d, axis=0))
print(np.cumprod(l_2d, axis=1))
import numpy as np
a = np.arange(12).reshape(3, 4)
print(a)
a_del = np.delete(a, 1, 0)
print(a_del)
print(a)
print(np.delete(a, 0, 0))
print(np.delete(a, 2, 0))
print(np.delete(a, 3, 0))
IndexError: index
print(np.delete(a, 1, 0))
print(np.delete(a, 1, 1))
print(np.delete(a, 1, 2))
AxisError: axis
print(np.delete(a, 1, None))
print(np.delete(a, 1))
print(np.delete(a, [0, 3], 1))
print(np.delete(a, [0, 1, 3], 1))
print(np.delete(a, slice(2), 1))
print(np.delete(a, slice(1, 3), 1))
print(np.delete(a, slice(None, None, 2), 1))
print(np.delete(a, np.s_[:2], 1))
print(np.delete(a, np.s_[1:3], 1))
print(np.delete(a, np.s_[::2], 1))
print(np.delete(np.delete(a, 1, 0), 1, 1))
a_3d = np.arange(24).reshape(2, 3, 4)
print(a_3d)
print(a_3d.shape)
print(np.delete(a_3d, 1, 0))
print(np.delete(a_3d, 1, 1))
print(np.delete(a_3d, 1, 2))
print(np.delete(a_3d, [0, 3], 2))
print(np.delete(a_3d, np.s_[::2], 2))
import numpy as np
a = np.array([10, 20, 30])
print(a)
print(np.diag(a))
print(np.diag([100, 200, 300]))
print(np.diag(a, k=1))
print(np.diag(a, k=-2))
print(np.diag([1, 1, 1]))
print(np.identity(3))
print(np.identity(3, int))
import numpy as np
a = np.arange(9).reshape((3, 3))
print(a)
print(np.diag(a))
print(np.diag(a, k=1))
print(np.diag(a, k=2))
print(np.diag(a, k=3))
print(np.diag(a, k=-1))
print(np.diag(a, k=-2))
print(np.diag(a, k=-3))
a = np.arange(12).reshape((3, 4))
print(a)
print(np.diag(a))
print(np.diag(a, k=1))
print(np.diag(a, k=-1))
a = np.arange(27).reshape((3, 3, 3))
print(a)
print(np.diag(a))
ValueError: Input
a = np.arange(9).reshape((3, 3))
print(a)
a_diag = np.diag(a)
print(a_diag)
a_diag[0] = 100
ValueError: assignment
a_diag.flags.writeable = True
a_diag[0] = 100
print(a_diag)
print(a)
a_diag_copy = np.diag(a).copy()
print(a_diag_copy)
a_diag_copy[1] = 100
print(a_diag_copy)
print(a)
import numpy as np
a = np.arange(9).reshape((3, 3))
print(a)
print(a.diagonal())
print(a.diagonal(offset=1))
print(a.diagonal(offset=3))
print(a.diagonal(offset=-2))
a = np.arange(12).reshape((3, 4))
print(a)
print(a.diagonal())
print(a.diagonal(offset=1))
a = np.arange(3)
print(a)
a.diagonal()
ValueError: diag
import numpy as np
a_3d = np.arange(24).reshape(2, 3, 4)
print(a_3d)
print(a_3d.shape)
a0, a1 = np.dsplit(a_3d, 2)
print(a0)
print(a0.shape)
print(a1)
print(a1.shape)
a0, a1 = np.dsplit(a_3d, [1])
print(a0)
print(a1)
a = np.arange(16).reshape(4, 4)
print(a)
np.dsplit(a, 2)
ValueError: dsplit
3 or more
import numpy as np
a1 = np.ones((3, 4), int)
print(a1)
a2 = np.full((3, 4), 2)
print(a2)
print(np.dstack([a1, a2]))
print(np.dstack([a1, a2]).shape)
np.dstack([a1, a2])
np.dstack([a1, a2])
print(np.concatenate([a1.reshape(3, 4, 1), a2.reshape(3, 4, 1)], 2))
a1 = np.ones(3, int)
print(a1)
a2 = np.full(3, 2)
print(a2)
print(np.dstack([a1, a2]))
print(np.dstack([a1, a2]).shape)
np.dstack([a1, a2])
np.dstack([a1, a2])
import numpy as np
a = np.array([1, 2, 3], dtype=np.int64)
print(a.dtype)
a = np.array([1, 2, 3], dtype='int64')
print(a.dtype)
a = np.array([1, 2, 3], dtype='i8')
print(a.dtype)
print(int is np.int)
a = np.array([1, 2, 3], dtype=int)
print(a.dtype)
a = np.array([1, 2, 3], dtype='int')
print(a.dtype)
a_str = np.array([1, 2, 3], dtype=str)
print(a_str)
print(a_str.dtype)
a_str[0] = 'abcde'
print(a_str)
a_str10 = np.array([1, 2, 3], dtype='U10')
print(a_str10.dtype)
a_str10[0] = 'abcde'
print(a_str10)
a_object = np.array([1, 0.1, 'one'], dtype=object)
print(a_object)
print(a_object.dtype)
print(type(a_object[0]))
print(type(a_object[1]))
print(type(a_object[2]))
a_object[2] = 'oneONE'
print(a_object)
l = [1, 0.1, 'oneONE']
print(type(l[0]))
print(type(l[1]))
print(type(l[2]))
print(a_object * 2)
print(l * 2)
import numpy as np
a = np.arange(120).reshape(2, 3, 4, 5)
print(a.shape)
print(a[Ellipsis, 0])
IndexError: an
ellipsis ('...')
print(a[0, 0])
import numpy as np
print(np.empty(3))
print(np.empty((2, 3)))
-3.10503618e+231
print(np.empty(3).dtype)
print(np.empty(3, dtype=np.int))
print(np.empty(3, dtype=np.int).dtype)
a_int = np.arange(6).reshape((2,3))
print(a_int)
print(np.empty_like(a_int))
a_float = np.arange(6).reshape((2,3)) / 10
print(a_float)
print(np.empty_like(a_float))
print(np.empty_like(a_float, dtype=np.int))
import numpy as np
print(np.__version__)
a_bool = np.array([True, True, True])
b_bool = np.array([True, False, False])
if a_bool:
pass
ValueError: The
a.any() or a.all()
a_bool and b_bool
ValueError: The
a.any() or a.all()
a_bool or b_bool
ValueError: The
a.any() or a.all()
not b_bool
ValueError: The
a.any() or a.all()
print(bool([0, 1, 2]))
print(bool([]))
print(not [0, 1, 2])
print(not [])
bool(a_bool)
ValueError: The
a.any() or a.all()
print(a_bool.all())
print(a_bool.any())
print(b_bool.all())
print(b_bool.any())
a_bool_2d = np.array([[True, True, True], [True, False, False]])
print(a_bool_2d)
print(a_bool_2d.all())
print(a_bool_2d.all(axis=0))
print(a_bool_2d.all(axis=1))
print(type(a_bool_2d.all(axis=0)))
numpy.ndarray
print(a_bool.size)
print(a_bool.size == 0)
print(a_bool & b_bool)
print(a_bool | b_bool)
print(a_bool ^ b_bool)
a_int = np.array([0, 1, 3])  
b_int = np.array([1, 0, 2])  
print(a_int & b_int)
print(a_int | b_int)
print(a_int ^ b_int)
a = np.arange(12).reshape(3, 4)
print(a)
print(a > 3)
print(a % 2 == 0)
print(a > 3 & a % 2 == 0)
ValueError: The
a.any() or a.all()
print(a > (3 & (a % 2)) == 0)
ValueError: The
a.any() or a.all()
print((a > 3) & (a % 2 == 0))
x = 10
print(x > 3)
print(x % 2 == 1)
print(x > 3 or x % 2 == 1)
print((x > 3) or (x % 2 == 1))
a_single = np.array([0])
b_single = np.array([1])
c_single = np.array([2])
print(bool(a_single))
print(bool(b_single))
print(bool(c_single))
print(b_single and c_single)
print(c_single and b_single)
print(b_single or c_single)
print(c_single or b_single)
print(b_single & c_single)
print(b_single | c_single)
print(not a_single)
print(not b_single)
print(not c_single)
-1
-2
-3
a_empty = np.array([])
print(a_empty)
print(bool(a_empty))
DeprecationWarning: The
error. Use
import numpy as np
a = np.arange(6).reshape(2, 3)
print(a)
print(np.expand_dims(a, 0))
print(np.expand_dims(a, 0).shape)
print(np.expand_dims(a, 1).shape)
print(np.expand_dims(a, 2).shape)
print(np.expand_dims(a, -1).shape)
print(np.expand_dims(a, -2).shape)
print(np.expand_dims(a, -3).shape)
print(np.expand_dims(a, 3).shape)
DeprecationWarning: Both
axis > a.ndim and axis < -a.ndim - 1
deprecated and will
raise an
print(np.expand_dims(a, -4).shape)
DeprecationWarning: Both
axis > a.ndim and axis < -a.ndim - 1
deprecated and will
raise an
print(np.expand_dims(a, (0, 1)).shape)
not supported 
a_expand_dims = np.expand_dims(a, 0)
print(np.shares_memory(a, a_expand_dims))
import numpy as np
e = np.eye(4)
print(type(e))
print(e)
print(e.dtype)
numpy.ndarray
e = np.eye(4, M=3, k=1, dtype=np.int8)
print(e)
print(e.dtype)
i = np.identity(4)
print(i)
print(i.dtype)
i = np.identity(4, dtype=np.uint8)
print(i)
print(i.dtype)
a = [3, 0, 8, 1, 9]
a_one_hot = np.identity(10)[a]
print(a)
print(a_one_hot)
a = [2, 2, 0, 1, 0]
a_one_hot = np.identity(3)[a]
print(a)
print(a_one_hot)
import numpy as np
a = np.arange(10) * 10
print(a)
print(a[5])
print(a[8])
print(a[[5, 8]])
print(a[[5, 4, 8, 0]])
print(a[[5, 5, 5, 5]])
idx = np.array([[5, 4], [8, 0]])
print(idx)
print(a[idx])
print(a[[[5, 4], [8, 0]]])
IndexError: too
print(a[[[[5, 4], [8, 0]]]])
a_2d = np.arange(12).reshape((3, 4))
print(a_2d)
print(a_2d[0])
print(a_2d[2])
print(a_2d[[2, 0]])
print(a_2d[[2, 2, 2]])
print(a_2d[:, 1])
print(a_2d[:, 3])
print(a_2d[:, 1:2])
print(a_2d[:, [3, 1]])
print(a_2d[:, [3, 3, 3]])
print(a_2d[0, 1])
print(a_2d[2, 3])
print(a_2d[[0, 2], [1, 3]])
print(a_2d[[0, 2, 1], [1, 3]])
IndexError: shape
mismatch: indexing
not be 
print(a_2d[[[0, 0], [2, 2]], [[1, 3], [1, 3]]])
print(a_2d[[[0], [2]], [1, 3]])
idxs = np.ix_([0, 2], [1, 3])
print(idxs)
array([[1, 3]])
print(type(idxs))
print(type(idxs[0]))
numpy.ndarray
print(idxs[0])
print(idxs[1])
print(a_2d[np.ix_([0, 2], [1, 3])])
print(a_2d[np.ix_([2, 0], [3, 3, 3])])
print(a_2d[[0, 2]][:, [1, 3]])
a_2d = np.arange(12).reshape((3, 4))
print(a_2d)
a_2d[np.ix_([0, 2], [1, 3])] = 100
print(a_2d)
a_2d[np.ix_([0, 2], [1, 3])] = [100, 200]
print(a_2d)
a_2d[np.ix_([0, 2], [1, 3])] = [[100, 200], [300, 400]]
print(a_2d)
print(a_2d[[0, 2]][:, [1, 3]])
a_2d[[0, 2]][:, [1, 3]] = 0
print(a_2d)
a_2d = np.arange(12).reshape((3, 4))
print(a_2d)
a_2d[[2, 0]] = [[100, 200, 300, 400], [500, 600, 700, 800]]
print(a_2d)
a_2d[[2, 2]] = [[-1, -2, -3, -4], [-5, -6, -7, -8]]
print(a_2d)
a_2d = np.arange(12).reshape((3, 4))
print(a_2d)
a_fancy = a_2d[np.ix_([0, 2], [1, 3])]
print(a_fancy)
a_fancy[0, 0] = 100
print(a_fancy)
print(a_2d)
a_2d = np.arange(12).reshape((3, 4))
print(a_2d)
print(a_2d[[2, 0], ::-1])
print(a_2d[::2, [3, 0, 1]])
import numpy as np
a = np.arange(3)
print(a)
print(a.flags)
C_CONTIGUOUS : True
F_CONTIGUOUS : True
OWNDATA : True
WRITEABLE : True
ALIGNED : True
WRITEBACKIFCOPY : False
UPDATEIFCOPY : False
print(type(a.flags))
numpy.flagsobj
print(a.flags.writeable)
print(a.flags['WRITEABLE'])
a[0] = 100
print(a)
a.flags.writeable = False
a[0] = 0
ValueError: assignment
a.flags['WRITEABLE'] = False
a.setflags(write=False)
a = np.arange(3)
print(a)
a.flags.writeable = False
a_view = a[1:]
print(a_view)
print(a_view.flags.writeable)
a_view[0] = 100
ValueError: assignment
a_view.flags.writeable = True
ValueError: cannot
a.flags.writeable = True
print(a_view.flags.writeable)
a_view.flags.writeable = True
a_view[0] = 100
print(a_view)
print(a)
a_view.flags.writeable = False
a_view[1] = 1
ValueError: assignment
print(a.flags.writeable)
a[1] = 1
print(a)
print(a_view)
a.flags.writeable = False
a_copy = a[1:].copy()
print(a_copy)
print(a_copy.flags.writeable)
a_copy[0] = 100
print(a_copy)
print(a)
import numpy as np
a_2d = np.arange(6).reshape(2, 3)
print(a_2d)
a_2d_flip = np.flip(a_2d)
print(a_2d_flip)
print(np.shares_memory(a_2d, a_2d_flip))
print(a_2d[::-1, ::-1])
a_1d = np.arange(3)
print(a_1d)
print(np.flip(a_1d))
print(a_1d[::-1])
a_3d = np.arange(24).reshape(2, 3, 4)
print(a_3d)
print(np.flip(a_3d))
-1
-1
-1
print(np.flip(a_3d, 0))
print(a_3d[::-1])
print(np.flipud(a_3d))
print(np.flip(a_3d, 1))
print(a_3d[:, ::-1])
print(np.fliplr(a_3d))
print(np.flip(a_3d, 2))
-1
print(np.flip(a_3d, (0, 1)))
print(a_3d[::-1, ::-1])
print(np.flip(a_3d, (1, 2)))
-1
-1
print(np.flip(a_3d, (0, 2)))
-1
-1
import numpy as np
from PIL import Image
img = np.array(Image.open('data/src/lena.jpg'))
print(type(img))
numpy.ndarray
print(img.shape)
Image.fromarray(np.flipud(img)).save('data/dst/lena_np_flipud.jpg')
Image.fromarray(np.fliplr(img)).save('data/dst/lena_np_fliplr.jpg')
Image.fromarray(np.flip(img, (0, 1))).save('data/dst/lena_np_flip_ud_lr.jpg')
import numpy as np
a_2d = np.arange(6).reshape(2, 3)
print(a_2d)
a_2d_fliplr = np.fliplr(a_2d)
print(a_2d_fliplr)
print(np.shares_memory(a_2d, a_2d_fliplr))
print(a_2d[:, ::-1])
a_1d = np.arange(3)
print(a_1d)
print(np.fliplr(a_1d))
ValueError: Input
a_3d = np.arange(24).reshape(2, 3, 4)
print(a_3d)
print(np.fliplr(a_3d))
print(a_3d[:, ::-1])
import numpy as np
a_2d = np.arange(6).reshape(2, 3)
print(a_2d)
a_2d_flipud = np.flipud(a_2d)
print(a_2d_flipud)
print(np.shares_memory(a_2d, a_2d_flipud))
a_2d_flipud_copy = np.flipud(a_2d).copy()
print(a_2d_flipud_copy)
print(np.shares_memory(a_2d, a_2d_flipud_copy))
print(a_2d[::-1])
a_1d = np.arange(3)
print(a_1d)
print(np.flipud(a_1d))
print(a_1d[::-1])
a_3d = np.arange(24).reshape(2, 3, 4)
print(a_3d)
print(np.flipud(a_3d))
print(a_3d[::-1])
import numpy as np
print(np.__version__)
a = np.array([[10.0, 10.1, 10.9], [-10.0, -10.1, -10.9]])
print(a)
print(np.floor(a))
print(np.floor(a).dtype)
print(np.floor(a).astype(int))
print(np.floor(10.1))
print(np.trunc(a))
print(np.fix(a))
print(a.astype(int))
print(np.ceil(a))
print(np.copysign(np.ceil(np.abs(a)), a))
import numpy as np
a = np.array([0, 2, 3, 6])
b = np.array([3, 4, 5, 15])
print(np.gcd(a, b))
print(type(np.gcd(a, b)))
numpy.ndarray
l_a = [0, 2, 3, 6]
l_b = [3, 4, 5, 14]
print(np.gcd(l_a, l_b))
print(type(np.gcd(l_a, l_b)))
numpy.ndarray
print(np.gcd(6, 15))
print(type(np.gcd(6, 15)))
numpy.int64
a_2d = np.array([[0, 2, 3, 6], [0, 2, 3, 6]])
print(a_2d)
print(b)
print(a_2d + b)
print(np.gcd(a_2d, b))
a_mismatch = np.array([0, 1, 2])
print(np.gcd(a_mismatch, b))
ValueError: operands
not be 
print(np.gcd(a, 15))
print(np.gcd(15, a))
import numpy as np
a = np.array([0, 2, 3, 6])
b = np.array([3, 4, 5, 15])
c = np.array([6, 8, 9, 9])
print(np.gcd.reduce([a, b, c]))
print(np.gcd(np.gcd(a, b), c))
print(np.lcm.reduce([a, b, c]))
print(np.gcd.reduce([6, 9, 12, 15]))
print(np.gcd.reduce((a, b, c)))
print(np.gcd.reduce(a, b, c))
TypeError: data
not understood
a_2d = np.array([[0, 2, 3, 6], [0, 2, 3, 6]])
print(a_2d)
print(np.gcd.reduce([a, b, a_2d]))
ValueError: The
a.any() or a.all()
print(np.gcd(np.gcd(a, b), a_2d))
print(np.gcd.reduce([a, b, 9]))
ValueError: The
a.any() or a.all()
print(np.gcd(np.gcd(a, b), 9))
a_1d = np.array([4, 6, 12])
print(np.gcd.reduce(a_1d))
print(np.lcm.reduce(a_1d))
a_2d = np.array([[4, 6, 12], [2, 12, 16]])
print(a_2d)
print(np.gcd.reduce(a_2d))
print(np.gcd(a_2d[0], a_2d[1]))
print(a_2d.ravel())
print(np.gcd.reduce(a_2d.ravel()))
print(np.lcm.reduce(a_2d.ravel()))
import numpy as np
def get_gradient_2d(start, stop, width, height, is_horizontal):
if is_horizontal:
return np.tile(np.linspace(start, stop, width), (height, 1))
return np.tile(np.linspace(start, stop, height), (width, 1)).T
def get_gradient_3d(width, height, start_list, stop_list, is_horizontal_list):
result = np.zeros((height, width, len(start_list)), dtype=np.float)
for i, (start, stop, is_horizontal) in enumerate(zip(start_list, stop_list, is_horizontal_list)):
get_gradient_2d(start, stop, width, height, is_horizontal)
return result
from PIL import Image
array = get_gradient_3d(512, 256, (0, 0, 0), (255, 255, 255), (True, True, True))
Image.fromarray(np.uint8(array)).save('data/dst/gray_gradient_h.jpg', quality=95)
array = get_gradient_3d(512, 256, (0, 0, 0), (255, 255, 255), (False, False, False))
Image.fromarray(np.uint8(array)).save('data/dst/gray_gradient_v.jpg', quality=95)
array = get_gradient_3d(512, 256, (0, 0, 192), (255, 255, 64), (True, False, False))
Image.fromarray(np.uint8(array)).save('data/dst/color_gradient.jpg', quality=95)
import numpy as np
with open('data/src/sample_nan.csv') as f:
print(f.read())
a = np.loadtxt('data/src/sample_nan.csv', delimiter=',')
ValueError: could
not convert 
a = np.genfromtxt('data/src/sample_nan.csv', delimiter=',')
print(a)
11. 12. nan
print(a[0, 2])
print(type(a[0, 2]))
numpy.float64
with open('data/src/sample_pandas_normal.csv') as f:
print(f.read())
a = np.loadtxt
delimiter
skiprows=1
print(type(a))
numpy.ndarray
print(a)
print(a.dtype)
a = np.genfromtxt
delimiter
names=True
dtype=None
encoding='utf-8'
print(type(a))
numpy.ndarray
print(a)
print(a.dtype)
import numpy as np
a = np.arange(16).reshape(4, 4)
print(a)
a0, a1 = np.hsplit(a, 2)
print(a0)
print(a1)
a0, a1 = np.hsplit(a, [1])
print(a0)
print(a1)
a_1d = np.arange(6)
print(a_1d)
np.split(a_1d, 2, 1)
IndexError: tuple
a0, a1 = np.hsplit(a_1d, 2)
print(a0)
print(a1)
import numpy as np
a1 = np.ones((2, 3), int)
print(a1)
a2 = np.full((2, 3), 2)
print(a2)
print(np.hstack([a1, a2]))
print(np.concatenate([a1, a2], 1))
a1 = np.ones(3, int)
print(a1)
a2 = np.full(3, 2)
print(a2)
print(np.hstack([a1, a2]))
print(np.concatenate([a1, a2], 0))
a1 = np.ones((2, 3), int)
print(a1)
a2 = np.full(2, 2)
print(a2)
print(np.hstack([a1, a2]))
ValueError: all
import numpy as np
ii64 = np.iinfo(np.int64)
print(type(ii64))
numpy.iinfo
print(ii64)
min = -9223372036854775808
max = 9223372036854775807
print(ii64.max)
print(type(ii64.max))
print(ii64.min)
-9223372036854775808
print(ii64.bits)
print(np.iinfo('int16'))
min = -32768
max = 32767
print(np.iinfo('i4'))
min = -2147483648
max = 2147483647
print(np.iinfo(int))
min = -9223372036854775808
max = 9223372036854775807
print(np.iinfo('uint64'))
min = 0
max = 18446744073709551615
i = 100
print(type(i))
print(np.iinfo(i))
min = -9223372036854775808
max = 9223372036854775807
ui = np.uint8(100)
print(type(ui))
numpy.uint8
print(np.iinfo(ui))
min = 0
max = 255
a = np.array([1, 2, 3], dtype=np.int8)
print(type(a))
numpy.ndarray
print(np.iinfo(a))
ValueError: Invalid
print(np.iinfo(a.dtype))
min = -128
max = 127
print(np.iinfo(a[0]))
min = -128
max = 127
print(np.iinfo(np.float64))
ValueError: Invalid
fi64 = np.finfo(np.float64)
print(type(fi64))
numpy.finfo
print(fi64)
precision =  15
resolution = 1.0000000000000001e-15
machep =    -52
eps =        2.2204460492503131e-16
negep =     -53
epsneg =     1.1102230246251565e-16
minexp =  -1022
tiny =       2.2250738585072014e-308
maxexp =   1024
max =        1.7976931348623157e+308
nexp =       11
min =        -max
print(fi64.max)
print(type(fi64.max))
numpy.float64
print(fi64.min)
-1.7976931348623157e+308
print(fi64.eps)
print(fi64.bits)
print(fi64.iexp)
print(fi64.nmant)
import numpy as np
from PIL import Image
src1 = np.array(Image.open('data/src/lena.jpg'))
src2 = np.array(Image.open('data/src/rocket.jpg').resize(src1.shape[1::-1], Image.BILINEAR))
print(src1.dtype)
dst = src1 * 0.5 + src2 * 0.5
print(dst.dtype)
Image.fromarray(dst.astype(np.uint8)).save('data/dst/numpy_image_alpha_blend.jpg')
dst = src1 * 0.5 + src2 * 0.2 + (96, 128, 160)
print(dst.max())
dst = dst.clip(0, 255)
print(dst.max())
Image.fromarray(dst.astype(np.uint8)).save('data/dst/numpy_image_alpha_blend_gamma.jpg')
import numpy as np
from PIL import Image
src1 = np.array(Image.open('data/src/lena.jpg'))
src2 = np.array(Image.open('data/src/rocket.jpg').resize(src1.shape[1::-1], Image.BILINEAR))
mask1 = np.array(Image.open('data/src/gradation_h.jpg').resize(src1.shape[1::-1], Image.BILINEAR))
mask1 = mask1 / 255
dst = src1 * mask1 + src2 * (1 - mask1)
Image.fromarray(dst.astype(np.uint8)).save('data/dst/numpy_image_ab_grad.jpg')
mask2 = np.array(Image.open('data/src/horse_r.png').resize(src1.shape[1::-1], Image.BILINEAR))
mask2 = mask2 / 255
dst = (src1 * mask1 + src2 * (1 - mask1)) * mask2
Image.fromarray(dst.astype(np.uint8)).save('data/dst/numpy_image_ab_mask_grad.jpg')
import numpy as np
from PIL import Image
im_gray = np.array(Image.open('data/src/lena_square_half.png').convert('L'))
print(type(im_gray))
numpy.ndarray
thresh = 128
im_bool = im_gray > thresh
print(im_bool)
maxval = 255
im_bin = (im_gray > thresh) * maxval
print(im_bin)
Image.fromarray(np.uint8(im_bin)).save('data/dst/numpy_binarization.png')
im_bin_keep = (im_gray > thresh) * im_gray
print(im_bin_keep)
Image.fromarray(np.uint8(im_bin_keep)).save('data/dst/numpy_binarization_keep.png')
im_bool = im_gray > 128
im_dst = np.empty
im_gray.shape
r, g, b = 255, 128, 32
Image.fromarray(np.uint8(im_dst)).save('data/dst/numpy_binarization_color.png')
im_bool = im_gray > 128
im_dst = np.empty
im_gray.shape
r, g, b = 128, 160, 192
Image.fromarray(np.uint8(im_dst)).save('data/dst/numpy_binarization_color2.png')
im = np.array(Image.open('data/src/lena_square_half.png'))
im_th = np.empty_like(im)
thresh = 128
maxval = 255
for i in range(3):
Image.fromarray(np.uint8(im_th)).save('data/dst/numpy_binarization_from_color.png')
l_thresh = [64, 128, 192]
l_maxval = [64, 128, 192]
for i, thresh, maxval in zip(range(3), l_thresh, l_maxval):
Image.fromarray(np.uint8(im_th)).save('data/dst/numpy_binarization_from_color2.png')
from PIL import Image
import numpy as np
im = np.array(Image.open('data/src/lena_square.png'))
im_R = im.copy()
im_G = im.copy()
im_B = im.copy()
im_RGB = np.concatenate((im_R, im_G, im_B), axis=1)
im_RGB = np.hstack((im_R, im_G, im_B))
im_RGB = np.c_
pil_img_RGB = Image.fromarray(im_RGB)
pil_img_RGB.save('data/dst/lena_numpy_split_color.jpg')
im_gray = 0.299 * im
print(im.dtype)
print(im_gray.dtype)
print(im.shape)
print(im_gray.shape)
pil_img_gray = Image.fromarray(np.uint8(im_gray))
pil_img_gray.save('data/dst/lena_numpy_gray.jpg')
im_swap = im.copy()
pil_img_swap = Image.fromarray(im_swap)
pil_img_swap.save('data/dst/lena_numpy_swap_color.jpg')
import numpy as np
from PIL import Image
im = np.array(Image.open('data/src/lena_square.png').resize((256, 256)))
im_32 = im // 32 * 32
im_128 = im // 128 * 128
im_dec = np.concatenate((im, im_32, im_128), axis=1)
Image.fromarray(im_dec).save('data/dst/lena_numpy_dec_color.png')
import cv2
import numpy as np
im = cv2.imread('data/src/lena.jpg')
print(im.shape)
print(im.dtype)
cv2.imwrite('data/temp/lena.png', im)
im_png = cv2.imread('data/temp/lena.png')
print(np.array_equal(im, im_png))
cv2.imwrite('data/temp/lena.bmp', im)
im_bmp = cv2.imread('data/temp/lena.bmp')
print(np.array_equal(im, im_bmp))
print(np.array_equal(im_png, im_bmp))
cv2.imwrite('data/dst/lena_q25.jpg', im, [cv2.IMWRITE_JPEG_QUALITY, 25])
im_q25 = cv2.imread('data/dst/lena_q25.jpg')
print(np.array_equal(im, im_q25))
print(im.shape == im_q25.shape)
print(im.shape == (250, 400, 3))
im_diff = im.astype(int) - im_q25.astype(int)
print(im_diff.max())
print(im_diff.min())
-101
im_diff_abs = np.abs(im_diff)
print(im_diff_abs.max())
print(im_diff_abs.min())
cv2.imwrite('data/dst/lena_diff_abs.png', im_diff_abs)
im_diff_abs_norm = im_diff_abs / im_diff_abs.max() * 255
print(im_diff_abs_norm.max())
print(im_diff_abs_norm.min())
cv2.imwrite('data/dst/lena_diff_abs_norm.png', im_diff_abs_norm)
im_diff_center = np.floor_divide(im_diff, 2) + 128
print(im_diff_center.max())
print(im_diff_center.min())
cv2.imwrite('data/dst/lena_diff_center.png', im_diff_center)
im_diff_center_norm = im_diff / np.abs(im_diff).max() * 127.5 + 127.5
print(im_diff_center_norm.max())
print(im_diff_center_norm.min())
cv2.imwrite('data/dst/lena_diff_center_norm.png', im_diff_center_norm)
im_diff_bin = (im_diff_abs > 32) * 255
cv2.imwrite('data/dst/lena_diff_bin.png', im_diff_bin)
print(list(zip(*np.where(im_diff_abs == np.max(im_diff_abs)))))
print(list(zip(*np.where(im_diff_abs > 100))))
from PIL import Image
import numpy as np
im = np.array(Image.open('data/src/lena_square.png'))
im_1_22 = 255.0 * (im / 255.0)**(1 / 2.2)
im_22 = 255.0 * (im / 255.0)**2.2
im_gamma = np.concatenate((im_1_22, im, im_22), axis=1)
pil_img = Image.fromarray(np.uint8(im_gamma))
pil_img.save('data/dst/lena_numpy_gamma.jpg')
import numpy as np
from PIL import Image
im = np.array(Image.open('data/src/lena_square.png').resize((256, 256)))
im_i = 255 - im
Image.fromarray(im_i).save('data/dst/lena_numpy_inverse.jpg')
import numpy as np
from PIL import Image
src = np.array(Image.open('data/src/lena.jpg'))
mask = np.array(Image.open('data/src/horse_r.png').resize(src.shape[1::-1], Image.BILINEAR))
print(mask.dtype, mask.min(), mask.max())
mask = mask / 255
print(mask.dtype, mask.min(), mask.max())
dst = src * mask
Image.fromarray(dst.astype(np.uint8)).save('data/dst/numpy_image_mask.jpg')
mask = np.array(Image.open('data/src/horse_r.png').convert('L').resize(src.shape[1::-1], Image.BILINEAR))
print(mask.shape)
mask = mask / 255
dst = src * mask
ValueError: operands
not be 
shapes (225,400,3) (225,400) 
mask = mask
np.newaxis
mask = mask.reshape(*mask.shape, 1)
print(mask.shape)
dst = src * mask
Image.fromarray(dst.astype(np.uint8)).save('data/dst/numpy_image_mask_l.jpg')
import numpy as np
from PIL import Image
src = np.array(Image.open('data/src/lena_square.png').resize((128, 128)))
dst = np.array(Image.open('data/src/lena_square.png').resize((256, 256))) // 4
dst_copy = dst.copy()
dst_copy[64:128, 128:192] = src[32:96, 32:96]
Image.fromarray(dst_copy).save('data/dst/lena_numpy_paste.jpg')
dst_copy = dst.copy()
dst_copy[64:192, 64:192] = src
Image.fromarray(dst_copy).save('data/dst/lena_numpy_paste_all.jpg')
from PIL import Image
import numpy as np
im = np.array(Image.open('data/src/lena.jpg'))
print(im.shape)
print(im[100, 150])
print(type(im[100, 150]))
numpy.ndarray
R, G, B = im[100, 150]
print(R)
print(G)
print(B)
im[100, 150] = (0, 50, 100)
print(im[100, 150])
print(im[100, 150])
from PIL import Image
import numpy as np
im = np.array(Image.open('data/src/lena_square.png').resize((256, 256)))
print(im.shape)
im_0 = im[:, :100]
im_1 = im[:, 100:]
print(im_0.shape)
print(im_1.shape)
Image.fromarray(im_0).save('data/dst/lena_numpy_split_0.jpg')
Image.fromarray(im_1).save('data/dst/lena_numpy_split_1.jpg')
im_0, im_1 = np.hsplit(im, 2)
print(im_0.shape)
print(im_1.shape)
im_0, im_1, im_2 = np.hsplit(im, [100, 150])
print(im_0.shape)
print(im_1.shape)
print(im_2.shape)
im_0, im_1, im_2 = np.hsplit(im, 3)
ValueError: array
not result in an 
im_0, im_1, im_2 = np.array_split(im, 3, axis=1)
print(im_0.shape)
print(im_1.shape)
print(im_2.shape)
from PIL import Image
import numpy as np
im = np.array(Image.open('data/src/lena_square.png'))
im_R = im.copy()
im_G = im.copy()
im_B = im.copy()
im_RGB = np.concatenate((im_R, im_G, im_B), axis=1)
im_RGB = np.hstack((im_R, im_G, im_B))
im_RGB = np.c_
pil_img = Image.fromarray(im_RGB)
pil_img.save('data/dst/lena_numpy_split_color.jpg')
from PIL import Image
import numpy as np
im = np.array(Image.open('data/src/lena_square.png'))
print(im.shape)
im_trim1 = im[128:384, 128:384]
print(im_trim1.shape)
Image.fromarray(im_trim1).save('data/dst/lena_numpy_trim.jpg')
def trim(array, x, y, width, height):
return array[y:y + height, x:x+width]
im_trim2 = trim(im, 128, 192, 256, 128)
print(im_trim2.shape)
Image.fromarray(im_trim2).save('data/dst/lena_numpy_trim2.jpg')
im_trim3 = trim(im, 128, 192, 512, 128)
print(im_trim3.shape)
Image.fromarray(im_trim3).save('data/dst/lena_numpy_trim3.jpg')
import numpy as np
a = np.array([1, 2, 3])
print(a)
print(a.dtype)
print((a / 1).dtype)
print((a / 1.0).dtype)
print((a + 1).dtype)
print((a + 1.0).dtype)
print((a - 1).dtype)
print((a - 1.0).dtype)
print((a * 1).dtype)
print((a * 1.0).dtype)
print((a // 1).dtype)
print((a // 1.0).dtype)
print((a ** 1).dtype)
print((a ** 1.0).dtype)
ones_int16 = np.ones(3, np.int16)
print(ones_int16)
ones_int32 = np.ones(3, np.int32)
print(ones_int32)
print((ones_int16 + ones_int32).dtype)
ones_float16 = np.ones(3, np.float16)
print(ones_float16)
print((ones_int16 + ones_float16).dtype)
ones_int16[0] = 10.9
print(ones_int16)
print(ones_int16.dtype)
import numpy as np
a = np.arange(4)
print(a)
print(np.insert(a, 2, 100))
print(np.insert(a, 1, [100, 101, 102]))
print(np.insert(a, [0, 2, 4], [100, 101, 102]))
_a = a.copy()
_a[1] = 100
print(_a)
_a = a.copy()
_a[1:3] = [100, 101]
print(_a)
_a = a.copy()
_a[1] = [100, 101, 102]
print(_a)
ValueError: setting
_a = np.insert(a, 1, [100, 101, 102])
_a = np.delete(_a, 4)
print(_a)
a = np.arange(12).reshape((3, 4))
print(a)
print(np.insert(a, 2, 100))
print(np.insert(a, 2, 100, axis=0))
b1 = np.arange(100, 104)
print(b1)
print(np.insert(a, 1, b1, axis=0))
print(np.insert(a, 3, b1, axis=0))
print(np.insert(a, [0, 2], b1, axis=0))
b2 = np.arange(100, 112).reshape((3, 4))
print(b2)
print(np.insert(a, 2, b2, axis=0))
print(np.insert(a, 2, b2[2], axis=0))
print(np.insert(a, [0, 2, 3], b2, axis=0))
print(np.insert(a, range(3), b2, axis=0))
print(np.vstack((a, b1)))
print(np.vstack((b2, a)))
_a = a.copy()
_a[2] = b1
print(_a)
_a = a.copy()
_a[1] = b2[1]
print(_a)
_a = a.copy()
_a[1:] = b2[[0, 2]]
print(_a)
print(a)
print(np.insert(a, 1, 100, axis=1))
c1 = np.arange(100, 103)
print(c1)
print(np.insert(a, 1, c1, axis=1))
print(np.insert(a, 3, c1, axis=1))
c2 = np.arange(100, 106).reshape((3, 2))
print(c2)
print(np.insert(a, 1, c2, axis=1))
ValueError: could
not broadcast 
shape (2,3) 
shape (3,3)
print(np.insert(a, [1], c2, axis=1))
print(np.insert(a, [0, 2], c2, axis=1))
print(c1)
print(np.insert(a, 1, c1, axis=1))
print(np.insert(a, [1], c1, axis=1))
print(np.insert(a, [1, 3, 4], c1, axis=1))
_c1 = c1.reshape((3, 1))
print(_c1)
print(np.insert(a, 1, _c1, axis=1))
print(np.insert(a, [1], _c1, axis=1))
print(np.insert(a, [1, 3, 4], _c1, axis=1))
print(np.hstack((a, c1)))
ValueError: all
print(_c1)
print(np.hstack((a, _c1)))
print(np.hstack((_c1, a)))
print(np.hstack((a, c2)))
print(np.hstack((c2, a)))
_a = a.copy()
_a[:, 1] = c1
print(_a)
_a = a.copy()
_a[:, :2] = c2
print(_a)
_a = a.copy()
_a[:, [0, 3]] = c2
print(_a)
import numpy as np
print(0.1 + 0.1 + 0.1)
a = np.array([0.3, 0.1 + 0.1 + 0.1])
print(a)
b = np.array([0.3, 0.3])
print(b)
print(a == b)
np.set_printoptions(precision=18)
print(a)
print(np.isclose(a, b))
print(np.isclose(a, 0.3))
print(np.isclose(0.1 + 0.1 + 0.1, 0.3))
print(np.isclose(100, 101))
print(np.isclose(100, 101, rtol=0, atol=1))
print(np.isclose(np.nan, np.nan))
print(np.isclose(np.nan, np.nan, equal_nan=True))
print(np.isclose(np.nan, 100, equal_nan=True))
a_nan = np.array([np.nan, 1, 2])
print(a_nan)
b_nan = np.array([np.nan, np.nan, 2])
print(b_nan)
print(np.isclose(a_nan, b_nan))
print(np.isclose(a_nan, b_nan, equal_nan=True))
import numpy as np
a = np.array([0, 2, 3, 6])
b = np.array([3, 4, 5, 15])
print(np.lcm(a, b))
print(type(np.lcm(a, b)))
numpy.ndarray
print(np.lcm(6, 15))
a_2d = np.array([[0, 2, 3, 6], [0, 2, 3, 6]])
print(a_2d)
print(np.lcm(a_2d, b))
print(np.lcm(a, 15))
import numpy as np
print(np.linspace(0, 10, 3))
print(np.linspace(0, 10, 4))
print(np.linspace(0, 10, 5))
print(np.linspace(10, 0, 5))
print(np.linspace(0, 10, 3, dtype=int))
print(np.linspace(0, 10, 5))
print(np.linspace(0, 10, 5, endpoint=False))
result = np.linspace(0, 10, 5, retstep=True)
print(result)
array([ 0. ,  2.5,  5. ,  7.5, 10. ])
print(type(result))
print(result[0])
print(result[1])
print(np.linspace(0, 10, 5, retstep=True)[1])
print(np.linspace(0, 10, 5, retstep=True, endpoint=False)[1])
print(np.linspace(0, 10, 5))
print(np.linspace(10, 0, 5))
print(np.linspace(0, 10, 5)[::-1])
print(np.flip(np.linspace(0, 10, 5)))
print(np.linspace(0, 10, 12).reshape(3, 4))
print(np.linspace(0, 10, 24).reshape(2, 3, 4))
import numpy as np
l = [[0, 0, 0], [0, 0, 0]]
arr = np.array(l)
print(arr)
print(arr.dtype, arr.shape, arr.ndim)
int64 (2, 3) 
l2 = [[0, 0, 0], [0, 0]]
arr2 = np.array(l2)
print(arr2)
print(arr2.dtype, arr2.shape, arr2.ndim)
import numpy as np
with open('data/src/sample.txt') as f:
print(f.read())
a = np.loadtxt('data/src/sample.txt')
print(type(a))
numpy.ndarray
print(a)
print(a.dtype)
with open('data/src/sample.csv') as f:
print(f.read())
print(np.loadtxt('data/src/sample.csv'))
ValueError: could
not convert 
print(np.loadtxt('data/src/sample.csv', delimiter=','))
a = np.loadtxt('data/src/sample.csv', delimiter=',', dtype='int64')
print(a)
print(a.dtype)
with open('data/src/sample_header_index.csv') as f:
print(f.read())
a = np.loadtxt
delimiter
dtype='int64'
skiprows=1
usecols=[1, 2, 3, 4]
print(a)
import numpy as np
print(np.__version__)
a_bool = np.array([True, True, False, False])
b_bool = np.array([True, False, True, False])
print(a_bool.dtype)
print(b_bool.dtype)
print(a_bool & b_bool)
print(a_bool | b_bool)
print(a_bool ^ b_bool)
print(type(a_bool & b_bool))
numpy.ndarray
print((a_bool & b_bool).dtype)
print(a_bool and b_bool)
ValueError: The
a.any() or a.all()
print(np.logical_and(a_bool, b_bool))
print(np.logical_or(a_bool, b_bool))
print(np.logical_xor(a_bool, b_bool))
print(np.logical_not(a_bool))
c_int = np.arange(4)
print(c_int)
print(np.logical_not(c_int))
d_int = c_int + 4
print(d_int)
print(np.logical_not(d_int))
print(np.logical_and(c_int, d_int))
print(c_int & d_int)
a_bool_2d = np.array([[True, True, False, False], [False, False, True, True]])
print(a_bool_2d)
print(a_bool_2d & b_bool)
print(np.logical_and(a_bool_2d, a_bool))
print(a_bool & True)
print(np.logical_and(a_bool, True))
print(c_int)
print(c_int < 2)
print(c_int % 2 == 0)
print((c_int < 2) & (c_int % 2 == 0))
print(c_int < 2 & c_int % 2 == 0)
ValueError: The
a.any() or a.all()
print(c_int < (2 & (c_int % 2)) == 0)
ValueError: The
a.any() or a.all()
print(np.logical_and(c_int < 2, c_int % 2 == 0))
import numpy as np
arr = np.array([[0, 1], [2, 3]])
det = np.linalg.det(arr)
print(det)
-2.0
import numpy as np
arr = np.array([[8, 1], [4, 5]])
w, v = np.linalg.eig(arr)
print(w)
print(v)
print('value: ', w[0])
print('vector: ', v[:, 0] / v[0, 0])
print(w[np.argmax(w)])
print(v[:, np.argmax(w)])
def get_eigenpairs(arr):
w, v = np.linalg.eig(arr)
eigenpairs = []
for i, val in enumerate(w):
vec = v[:, i] / np.min(np.abs(v[:, i][v[:, i] != 0]))
eigenpairs.append((val, vec))
return eigenpairs
eigenpairs = get_eigenpairs(arr)
for val, vec in eigenpairs:
print('value: {}, vector: {}'.format(val, vec))
-1.  4
arr = np.array([[1, 1, 2], [0, 2, -1], [0, 0, 3]])
eigenpairs = get_eigenpairs(arr)
for val, vec in eigenpairs:
print('value: {}, vector: {}'.format(val, vec))
arr = np.array([[3, 2], [-2, 3]])
eigenpairs = get_eigenpairs(arr)
for val, vec in eigenpairs:
print('value: {}, vector: {}'.format(val, vec))
import numpy as np
arr = np.array([[2, 5], [1, 3]])
arr_inv = np.linalg.inv(arr)
print(arr_inv)
-1.  2
mat = np.matrix([[2, 5], [1, 3]])
mat_inv = np.linalg.inv(mat)
print(mat_inv)
-1.  2
mat_inv = mat**-1
print(mat_inv)
-1.  2
mat_inv = mat.I
print(mat_inv)
-1.  2
result = mat * mat.I
print(result)
print(arr.I)
numpy.ndarray
arr_s = np.array([[0, 0], [1, 3]])
print(np.linalg.inv(arr_s))
LinAlgError: Singular
arr_pinv = np.linalg.pinv(arr_s)
print(arr_pinv)
print(np.linalg.pinv(arr_pinv))
print(np.linalg.inv(arr))
-1.  2
print(np.linalg.pinv(arr))
-1.  2
mat_s = np.mat([[0, 0], [1, 3]])
print(np.linalg.inv(mat_s))
LinAlgError: Singular
print(mat_s**-1)
LinAlgError: Singular
print(mat_s.I)
LinAlgError: Singular
print(np.linalg.pinv(mat_s))
import numpy as np
l_2d = [[0, 1, 2], [3, 4, 5]]
print(l_2d)
print(type(l_2d))
arr = np.array([[0, 1, 2], [3, 4, 5]])
print(arr)
print(type(arr))
numpy.ndarray
arr = np.arange(6)
print(arr)
arr = np.arange(6).reshape((2, 3))
print(arr)
mat = np.matrix([[0, 1, 2], [3, 4, 5]])
print(mat)
print(type(mat))
numpy.matrix
mat = np.matrix(arr)
print(mat)
print(type(mat))
numpy.matrix
mat_1d = np.matrix([0, 1, 2])
print(mat_1d)
print(type(mat_1d))
numpy.matrix
print(mat_1d.shape)
mat_3d = np.matrix([[[0, 1, 2]]])
ValueError: matrix
print(l_2d)
print(l_2d[0][1])
l_2d[0][1] = 100
print(l_2d)
print(arr)
print(arr[0, 1])
arr[0, 1] = 100
print(arr)
l_2d = [[0, 1, 2], [3, 4, 5]]
print(l_2d)
print([list(x) for x in list(zip(*l_2d))])
arr = np.arange(6).reshape((2, 3))
print(arr)
print(arr.T)
l_2d_1 = [[0, 1, 2], [3, 4, 5]]
l_2d_2 = [[0, 2, 4], [6, 8, 10]]
print(l_2d_1 + l_2d_2)
print(l_2d_1 - l_2d_2)
TypeError: unsupported
type(s) 
arr1 = np.arange(6).reshape((2, 3))
print(arr1)
arr2 = np.arange(0, 12, 2).reshape((2, 3))
print(arr2)
print(arr1 + arr2)
print(arr1 - arr2)
mat1 = np.matrix(arr1)
mat2 = np.matrix(arr2)
print(mat1 + mat2)
print(mat1 - mat2)
print(l_2d_1 * 2)
print(l_2d_1 * l_2d_2)
TypeError: can
print(arr1 * 2)
print(mat1 * 2)
print(np.multiply(arr1, arr2))
print(np.multiply(mat1, mat2))
print(arr1 * arr2)
arr1 = np.arange(4).reshape((2, 2))
print(arr1)
arr2 = np.arange(6).reshape((2, 3))
print(arr2)
print(np.dot(arr1, arr2))
print(arr1.dot(arr2))
print(np.matmul(arr1, arr2))
mat1 = np.matrix(arr1)
mat2 = np.matrix(arr2)
print(np.dot(mat1, mat2))
print(mat1.dot(mat2))
print(np.matmul(mat1, mat2))
print(mat1 * mat2)
arr = np.arange(1, 5).reshape(2, 2)
print(arr)
print(arr**2)
mat = np.matrix(arr)
print(mat)
print(mat**2)
print(mat**2 == mat * mat)
print(mat**3 == mat * mat * mat)
print(arr**-1)
ValueError: Integers
not allowed
arr_f = np.array(arr, dtype=float)
print(arr_f**-1)
print(mat**-1)
-2.   1
print(mat**-2)
-3.75
print(mat**-2 == mat**-1 * mat**-1)
print(mat**-3 == mat**-1 * mat**-1 * mat**-1)
import numpy as np
a = np.arange(6).reshape(2, 3)
print(a)
print(a.max())
print(a.max(axis=0))
print(a.max(axis=1))
import numpy as np
a = np.array([0, 1, 2])
b = np.array([2, 0, 6])
print(np.maximum(a, b))
print(np.fmax(a, b))
l_a = [0, 1, 2]
l_b = [2, 0, 6]
print(np.maximum(l_a, l_b))
print(np.fmax(l_a, l_b))
print(np.maximum(0, 2))
print(np.fmax(0, 2))
print(max(0, 2))
a_2d = np.arange(6).reshape(2, 3)
print(a_2d)
print(b)
print(a_2d + b)
print(np.maximum(a_2d, b))
print(np.fmax(a_2d, b))
a_mismatch = np.array([0, 1, 2, 3])
print(np.maximum(a_mismatch, b))
ValueError: operands
not be 
print(np.fmax(a_mismatch, b))
ValueError: operands
not be 
print(np.maximum(a_2d, 2))
print(np.fmax(a_2d, 2))
print(np.maximum(2, a_2d))
print(np.fmax(2, a_2d))
print(np.maximum([np.nan, np.nan], [np.inf, 0]))
print(np.fmax([np.nan, np.nan], [np.inf, 0]))
import numpy as np
a = np.array([0, 1, 2])
b = np.array([3, 0, 6])
c = np.array([1, 2, 3])
print(np.maximum.reduce([a, b, c]))
print(np.maximum(np.maximum(a, b), c))
print(np.fmax.reduce([a, b, c]))
print(np.minimum.reduce([a, b, c]))
print(np.fmin.reduce([a, b, c]))
print(np.maximum.reduce((a, b, c)))
print(np.maximum.reduce(a, b, c))
TypeError: data
not understood
a_2d = np.arange(6).reshape(2, 3)
print(a_2d)
print(np.maximum.reduce([a_2d, b, c]))
ValueError: The
a.any() or a.all()
print(np.maximum(np.maximum(a_2d, b), c))
print(np.maximum.reduce([4, b, c]))
ValueError: The
a.any() or a.all()
print(np.maximum(np.maximum(4, b), c))
print(a)
print(np.maximum.reduce(a))
print(a.max())
print(max(a))
a_2d = np.array([[0, 1, 2], [3, 0, 6], [1, 2, 3]])
print(a_2d)
print(np.maximum.reduce(a_2d))
print(np.maximum(np.maximum(a_2d[0], a_2d[1]), a_2d[2]))
print(a_2d.max(axis=0))
import numpy as np
a = np.arange(12)
a = a.reshape(3, 4)
a = a.clip(2, 9)
print(a)
a_mc = np.arange(12).reshape(3, 4).clip(2, 9)
print(a_mc)
np.arange(12)
reshape(3, 4)
clip(2, 9)
print(a_mc_break_parens)
import numpy as np
a = np.array([0, 1, 2])
b = np.array([2, 0, 6])
print(np.minimum(a, b))
print(np.fmin(a, b))
a_2d = np.arange(6).reshape(2, 3)
print(a_2d)
print(np.minimum(a_2d, b))
print(np.fmin(a_2d, b))
print(np.minimum(a_2d, 2))
print(np.fmin(a_2d, 2))
print(np.minimum([np.nan, np.nan], [np.inf, 0]))
print(np.fmin([np.nan, np.nan], [np.inf, 0]))
import numpy as np
a = np.genfromtxt('data/src/sample_nan.csv', delimiter=',')
print(a)
11. 12. nan
print(np.isnan(a))
np.isnan(a)
np.isnan(a)
print(np.isnan(a).any(axis=1))
np.isnan(a).any(axis=1)
np.isnan(a).any(axis=1)
np.isnan(a).any(axis=1)
np.isnan(a).any(axis=0)
np.isnan(a).any(axis=0)
a = np.genfromtxt('data/src/sample_nan.csv', delimiter=',')
a[2, 2] = np.nan
print(a)
11. 12. nan
31. 32. nan
np.isnan(a).any(axis=0)
np.isnan(a).all(axis=0)
import numpy as np
a = np.genfromtxt('data/src/sample_nan.csv', delimiter=',')
print(a)
11. 12. nan
a_nan = np.array([0, 1, np.nan, float('nan')])
print(a_nan)
0.  1. nan
print(np.nan == np.nan)
print(np.isnan(np.nan))
print(a_nan == np.nan)
print(np.isnan(a_nan))
a_fill = np.genfromtxt('data/src/sample_nan.csv', delimiter=',', filling_values=0)
print(a_fill)
a = np.genfromtxt('data/src/sample_nan.csv', delimiter=',')
print(np.nan_to_num(a))
print(a)
11. 12. nan
print(np.nan_to_num(a, copy=False))
print(a)
a = np.genfromtxt('data/src/sample_nan.csv', delimiter=',')
print(np.nan_to_num(a, nan=-1))
-1. 14
print(np.nanmean(a))
print(np.nan_to_num(a, nan=np.nanmean(a)))
a = np.genfromtxt('data/src/sample_nan.csv', delimiter=',')
print(np.isnan(a))
a[np.isnan(a)] = 0
print(a)
a = np.genfromtxt('data/src/sample_nan.csv', delimiter=',')
a[np.isnan(a)] = np.nanmean(a)
print(a)
import numpy as np
arr = np.genfromtxt('data/src/sample_nan.csv', delimiter=',')
print(arr)
11.  12.  nan
print(arr.sum())
print(np.sum(arr))
print(np.nansum(arr))
print(np.nansum(arr, axis=0))
print(np.nansum(arr, axis=1))
print(np.nanmean(arr))
print(np.nanmax(arr))
print(np.nanmin(arr))
print(np.nanstd(arr))
print(np.nanvar(arr))
import numpy as np
a = np.arange(10)
print(a)
a_0 = a[:6]
print(a_0)
a_1 = a_0.reshape(2, 3)
print(a_1)
print(a_0.base)
print(a_1.base)
a_copy = a.copy()
print(a_copy)
print(a_copy.base)
print(a.base)
print(a_0.base is None)
print(a_copy.base is None)
print(a.base is None)
print(a_0.base is a)
print(a_0.base is a_1.base)
import numpy as np
arr = np.array([0, 1, 2])
print(arr)
arr_2d = np.array([[0, 1, 2], [3, 4, 5]])
print(arr_2d)
print(arr[1])
print(arr_2d[1, 1])
print(arr_2d[0, 1:])
print(np.sqrt(arr_2d))
arr_1 = np.array([[1, 2], [3, 4]])
arr_2 = np.array([[1, 2, 3], [4, 5, 6]])
print(np.dot(arr_1, arr_2))
import numpy as np
l_1d = [0, 1, 2]
arr_1d = np.array(l_1d)
print(arr_1d)
print(arr_1d.dtype)
arr_1d_f = np.array(l_1d, dtype=float)
print(arr_1d_f)
print(arr_1d_f.dtype)
l_2d = [[0, 1, 2], [3, 4, 5]]
arr_2d = np.array(l_2d)
print(arr_2d)
l_2d_error = [[0, 1, 2], [3, 4]]
arr_2d_error = np.array(l_2d_error)
print(arr_2d_error)
list([0, 1, 2]) 
list([3, 4])
print(arr_2d_error.dtype)
print(arr_2d_error.shape)
arr_1d = np.arange(3)
print(arr_1d)
l_1d = arr_1d.tolist()
print(l_1d)
arr_2d = np.arange(6).reshape((2, 3))
print(arr_2d)
l_2d = arr_2d.tolist()
print(l_2d)
arr_3d = np.arange(24).reshape((2, 3, 4))
print(arr_3d)
l_3d = arr_3d.tolist()
print(l_3d)
print(l_3d[0])
print(l_3d[0][0])
print(l_3d[0][0][0])
import numpy as np
a_1d = np.arange(3)
print(a_1d)
a_2d = np.arange(12).reshape((3, 4))
print(a_2d)
a_3d = np.arange(24).reshape((2, 3, 4))
print(a_3d)
print(a_1d.ndim)
print(type(a_1d.ndim))
print(a_2d.ndim)
print(a_3d.ndim)
print(a_1d.shape)
print(type(a_1d.shape))
print(a_2d.shape)
print(a_3d.shape)
print(a_2d.shape[0])
print(a_2d.shape[1])
row, col = a_2d.shape
print(row)
print(col)
print(a_1d.size)
print(type(a_1d.size))
print(a_2d.size)
print(a_3d.size)
print(len(a_1d))
print(a_1d.shape[0])
print(a_1d.size)
print(len(a_2d))
print(a_2d.shape[0])
print(len(a_3d))
print(a_3d.shape[0])
import numpy as np
print(np.newaxis is None)
a = np.arange(6).reshape(2, 3)
print(a)
print(a.shape)
np.newaxis
np.newaxis
np.newaxis
np.newaxis
np.newaxis
np.newaxis
print(a[:, np.newaxis])
print(a[:, np.newaxis].shape)
print(a[np.newaxis])
print(a[np.newaxis].shape)
np.newaxis
np.newaxis
np.newaxis
np.newaxis
np.newaxis
np.newaxis
np.newaxis
np.newaxis
a_newaxis = a
np.newaxis
print(np.shares_memory(a, a_newaxis))
a = np.zeros(27, dtype=np.int).reshape(3, 3, 3)
print(a)
print(a.shape)
b = np.arange(9).reshape(3, 3)
print(b)
print(b.shape)
print(a + b)
np.newaxis
np.newaxis
np.newaxis
np.newaxis
np.newaxis
np.newaxis
a = np.arange(6).reshape(2, 3)
print(a)
print(a.shape)
print(a[np.newaxis])
print(a[np.newaxis].shape)
print(np.expand_dims(a, 0))
print(np.expand_dims(a, 0).shape)
print(a.reshape(1, 2, 3))
print(a.reshape(1, 2, 3).shape)
print(a.reshape(1, *a.shape))
print(a.reshape(1, *a.shape).shape)
import numpy as np
import scipy.stats
from sklearn import preprocessing
a = np.array([0, 1, 2, 3, 4])
print(a)
print((a - a.min()) / (a.max() - a.min()))
print((a - a.mean()) / a.std())
print((a - a.mean()) / a.std(ddof=1))
a_2d = np.array([[0, 1, 2], [3, 4, 5], [6, 7, 8]])
print(a_2d)
def min_max(x, axis=None):
x_min = x.min(axis=axis, keepdims=True)
x_max = x.max(axis=axis, keepdims=True)
return (x - x_min) / (x_max - x_min)
print(min_max(a_2d))
print(min_max(a_2d, axis=0))
print(min_max(a_2d, axis=1))
print(min_max(a))
def standardization(x, axis=None, ddof=0):
x_mean = x.mean(axis=axis, keepdims=True)
x_std = x.std(axis=axis, keepdims=True, ddof=ddof)
return (x - x_mean) / x_std
print(standardization(a_2d))
-0.38729833
print(standardization(a_2d, ddof=1))
-0.36514837
print(standardization(a_2d, axis=0))
print(standardization(a_2d, axis=0, ddof=1))
print(standardization(a_2d, axis=1))
-1.22474487
-1.22474487
-1.22474487
print(standardization(a_2d, axis=1, ddof=1))
-1.  0
-1.  0
-1.  0
print(standardization(a))
print(standardization(a, ddof=1))
print(scipy.stats.zscore(a))
print(scipy.stats.zscore(a_2d))
print(scipy.stats.zscore(a_2d, axis=None, ddof=1))
-0.36514837
mm = preprocessing.MinMaxScaler()
print(mm.fit_transform(a_2d.astype(float)))
print(preprocessing.minmax_scale(a.astype(float)))
print(preprocessing.minmax_scale(a_2d.astype(float), axis=1))
ss = preprocessing.StandardScaler()
print(ss.fit_transform(a_2d.astype(float)))
print(preprocessing.scale(a.astype(float)))
print(preprocessing.scale(a_2d.astype(float), axis=1))
-1.22474487
-1.22474487
-1.22474487
import numpy as np
from PIL import Image
src_path = "data/src/lena.jpg"
img = Image.open(src_path).convert('RGB')
print(type(img))
PIL.Image.Image
print(img.size)  
arr = np.array(img)
print(type(arr))
numpy.ndarray
print(arr.shape)  
img2 = Image.fromarray(np.uint8(arr))
print(type(img2))
PIL.Image.Image
import numpy as np
import cv2
def psnr(img_1, img_2, data_range=255):
mse = np.mean((img_1.astype(float) - img_2.astype(float)) ** 2)
return 10 * np.log10((data_range ** 2) / mse)
img_org = cv2.imread('data/src/lena.jpg')
img_q95 = cv2.imread('data/src/lena_q95.jpg')
img_q50 = cv2.imread('data/src/lena_q50.jpg')
print(psnr(img_org, img_q95))
print(psnr(img_org, img_q50))
print(psnr(img_org, img_org))
RuntimeWarning: divide
imports
import numpy as np
a = np.arange(12).reshape(3, 4)
print(a)
print(np.ravel(a))
print(type(np.ravel(a)))
numpy.ndarray
print(np.ravel([[0, 1, 2, 3], [4, 5, 6, 7], [8, 9, 10, 11]]))
print(type(np.ravel([[0, 1, 2, 3], [4, 5, 6, 7], [8, 9, 10, 11]])))
numpy.ndarray
print(np.ravel([[0, 1, 2, 3], [4, 5, 6, 7], [8, 9]]))
list([0, 1, 2, 3]) 
list([4, 5, 6, 7]) 
list([8, 9])
print(type(np.ravel([[0, 1, 2, 3], [4, 5, 6, 7], [8, 9]])))
numpy.ndarray
print(a.ravel())
print(a.flatten())
print(a.reshape(-1))
print(np.reshape(a, -1))
print(np.reshape([[0, 1, 2, 3], [4, 5, 6, 7], [8, 9, 10, 11]], -1))
print(np.reshape([[0, 1, 2, 3], [4, 5, 6, 7], [8, 9]], -1))
list([0, 1, 2, 3]) 
list([4, 5, 6, 7]) 
list([8, 9])
print(np.shares_memory(a, a.ravel()))
print(np.shares_memory(a, np.ravel(a)))
print(np.shares_memory(a, a.flatten()))
print(np.shares_memory(a, a.reshape(-1)))
print(np.shares_memory(a, np.reshape(a, -1)))
a_ravel = a.ravel()
print(a_ravel)
a_ravel[0] = 100
print(a_ravel)
print(a)
a = np.arange(12).reshape(3, 4)
print(a)
print(a[:, ::3])
print(np.shares_memory(a[:, ::3], np.ravel(a[:, ::3])))
print(np.shares_memory(a[:, ::3], np.reshape(a[:, ::3], -1)))
print(a[:, ::2])
print(np.shares_memory(a[:, ::2], np.ravel(a[:, ::2])))
print(np.shares_memory(a[:, ::2], np.reshape(a[:, ::2], -1)))
a_3d = np.arange(24).reshape(2, 3, 4)
print(a_3d)
print(a_3d.ravel())
print(a_3d.ravel('F'))
import numpy as np
a = np.arange(12).reshape(3, 4)
print(a)
print(a.ravel())
print(a.ravel('F'))
print(np.ravel(a, 'F'))
print(a.flatten('F'))
print(a.reshape(-1, order='F'))
print(np.reshape(a, -1, order='F'))
np.info(a)
class:  ndarray
aligned:  True
contiguous:  True
fortran:  False
byteorder:  little
byteswap:  False
type: int64
print(a.ravel('C'))
print(a.ravel('F'))
print(a.ravel('A'))
print(a.ravel('K'))
print(a.T)
np.info(a.T)
class:  ndarray
aligned:  True
contiguous:  False
fortran:  True
byteorder:  little
byteswap:  False
type: int64
print(a.T.ravel('C'))
print(a.T.ravel('F'))
print(a.T.ravel('A'))
print(a.T.ravel('K'))
print(a.T[::-1])
np.info(a.T[::-1])
class:  ndarray
-8
aligned:  True
contiguous:  False
fortran:  False
byteorder:  little
byteswap:  False
type: int64
print(a.T[::-1].ravel('C'))
print(a.T[::-1].ravel('F'))
print(a.T[::-1].ravel('A'))
print(a.T[::-1].ravel('K'))
import numpy as np
a = np.arange(12).reshape(3, 4)
print(a)
a.ravel()
std. dev. of
a.flatten()
std. dev. of
a.reshape(-1)
std. dev. of
a_large = np.arange(1000000).reshape(100, 100, 100)
a_large.ravel()
std. dev. of
a_large.flatten()
std. dev. of
a_large.reshape(-1)
std. dev. of
import numpy as np
a = np.arange(24)
print(a)
print(a.shape)
print(a.ndim)
a_4_6 = a.reshape([4, 6])
print(a_4_6)
print(a_4_6.shape)
print(a_4_6.ndim)
a_2_3_4 = a.reshape([2, 3, 4])
print(a_2_3_4)
print(a_2_3_4.shape)
print(a_2_3_4.ndim)
a_5_6 = a.reshape([5, 6])
ValueError: cannot
shape (5,6)
print(a.reshape(4, 6))
print(a.reshape(2, 3, 4))
print(np.reshape(a, [4, 6]))
print(np.reshape(a, [2, 3, 4]))
print(np.reshape(a, [5, 6]))
ValueError: cannot
shape (5,6)
print(a.reshape(4, 6))
print(np.reshape(a, 4, 6))
ValueError: cannot
print(a.reshape([4, 6], order='C'))
print(a.reshape([4, 6], order='F'))
print(a.reshape([2, 3, 4], order='C'))
print(a.reshape([2, 3, 4], order='F'))
print(np.reshape(a, [4, 6], order='F'))
print(a.reshape([4, 6], 'F'))
print(np.reshape(a, [4, 6], 'F'))
print(a.reshape([4, -1]))
print(a.reshape([2, -1, 4]))
print(a.reshape([2, -1, -1]))
ValueError: can
print(a.reshape([2, -1, 5]))
ValueError: cannot
shape (2,newaxis,5)
a = np.arange(8)
print(a)
a_2_4 = a.reshape([2, 4])
print(a_2_4)
print(np.shares_memory(a, a_2_4))
a[0] = 100
print(a)
print(a_2_4)
a_2_4[0, 0] = 0
print(a_2_4)
print(a)
a_2_4_copy = a.reshape([2, 4]).copy()
print(a_2_4_copy)
print(np.shares_memory(a, a_2_4_copy))
a[0] = 100
print(a)
print(a_2_4_copy)
a_2_4_copy[0, 0] = 200
print(a_2_4_copy)
print(a)
a = np.arange(6).reshape(2, 3)
print(a)
a_step = a[:, ::2]
print(a_step)
print(a_step.reshape(-1))
print(np.shares_memory(a_step, a_step.reshape(-1)))
np.info(a)
class:  ndarray
aligned:  True
contiguous:  True
fortran:  False
byteorder:  little
byteswap:  False
type: int64
np.info(a_step)
class:  ndarray
aligned:  True
contiguous:  False
fortran:  False
byteorder:  little
byteswap:  False
type: int64
np.info(a_step.reshape(-1))
class:  ndarray
aligned:  True
contiguous:  True
fortran:  True
byteorder:  little
byteswap:  False
type: int64
a = np.arange(8).reshape(2, 4)
print(a)
a_step = a[:, ::2]
print(a_step)
print(a_step.reshape(-1))
print(np.shares_memory(a_step, a_step.reshape(-1)))
np.info(a)
class:  ndarray
aligned:  True
contiguous:  True
fortran:  False
byteorder:  little
byteswap:  False
type: int64
np.info(a_step)
class:  ndarray
aligned:  True
contiguous:  False
fortran:  False
byteorder:  little
byteswap:  False
type: int64
np.info(a_step.reshape(-1))
class:  ndarray
aligned:  True
contiguous:  False
fortran:  False
byteorder:  little
byteswap:  False
type: int64
import numpy as np
a = np.arange(10)
print(a)
a_roll = np.roll(a, 3)
print(a_roll)
print(a)
a_roll[0] = 100
print(a_roll)
print(a)
print(np.roll(a, -3))
print(np.roll(a, 12))
print(np.roll(a, 0.5))
TypeError: slice
integers or None or have
a_2d = np.arange(12).reshape(3, 4)
print(a_2d)
print(np.roll(a_2d, 2))
print(np.roll(a_2d, 5))
print(np.roll(a_2d, 1, axis=0))
print(np.roll(a_2d, 2, axis=1))
print(np.roll(a_2d, (1, 2), axis=(0, 1)))
a_3d = np.arange(24).reshape(2, 3, 4)
print(a_3d)
print(np.roll(a_3d, 3))
print(np.roll(a_3d, 2, axis=2))
import numpy as np
from PIL import Image
img = np.array(Image.open('data/src/lena.jpg'))
print(img.shape)
img_scroll = np.roll(img, (50, 100), axis=(0, 1))
Image.fromarray(img_scroll).save('data/dst/lena_numpy_roll.jpg')
import numpy as np
a_2d = np.arange(6).reshape(2, 3)
print(a_2d)
a_2d_rot = np.rot90(a_2d)
print(a_2d_rot)
print(np.shares_memory(a_2d, a_2d_rot))
a_2d_rot[0, 0] = 100
print(a_2d_rot)
print(a_2d)
a_2d[0, 2] = 2
print(a_2d)
print(a_2d_rot)
a_2d_rot_copy = np.rot90(a_2d).copy()
print(a_2d_rot_copy)
print(np.shares_memory(a_2d, a_2d_rot_copy))
print(np.rot90(a_2d, 2))
print(np.rot90(a_2d, 3))
print(np.rot90(a_2d, 4))
print(np.rot90(a_2d, 100))
print(np.rot90(a_2d, -1))
print(np.rot90(a_2d, -2))
a_1d = np.arange(3)
print(a_1d)
print(np.rot90(a_1d))
ValueError: Axes
a_2d_row = np.arange(3).reshape(1, 3)
print(a_2d_row)
print(np.rot90(a_2d_row))
a_3d = np.arange(24).reshape(2, 3, 4)
print(a_3d)
print(a_3d.shape)
print(np.rot90(a_3d))
print(np.rot90(a_3d).shape)
np.rot90(a_3d)
print(np.rot90(a_3d, axes=(0, 1)))
print(np.rot90(a_3d, axes=(1, 2)))
print(np.rot90(a_3d, axes=(1, 2)).shape)
print(np.rot90(a_3d, axes=(2, 1)))
print(np.rot90(a_3d, axes=(2, 1)).shape)
print(np.rot90(a_3d, 2, axes=(1, 2)))
print(np.rot90(a_3d, -1, axes=(1, 2)))
import numpy as np
from PIL import Image
img = np.array(Image.open('data/src/lena.jpg'))
print(type(img))
numpy.ndarray
print(img.shape)
Image.fromarray(np.rot90(img)).save('data/dst/lena_np_rot90.jpg')
Image.fromarray(np.rot90(img, 2)).save('data/dst/lena_np_rot90_180.jpg')
Image.fromarray(np.rot90(img, 3)).save('data/dst/lena_np_rot90_270.jpg')
import numpy as np
a = np.arange(5)
print(a)
np.save('data/temp/np_save', a)
print(type(np.load('data/temp/np_save.npy')))
numpy.ndarray
print(np.load('data/temp/np_save.npy'))
np.save('data/temp/np_save', a.astype('float32'))
b = np.load('data/temp/np_save.npy')
print(b)
print(b.dtype)
import numpy as np
a = np.arange(6).reshape(2, 3)
print(a)
np.savetxt('data/temp/np_savetxt.txt', a)
with open('data/temp/np_savetxt.txt') as f:
print(f.read())
np.savetxt('data/temp/np_savetxt_5e.txt', a, fmt='%.5e')
with open('data/temp/np_savetxt_5e.txt') as f:
print(f.read())
print(np.loadtxt('data/temp/np_savetxt.txt'))
np.savetxt('data/temp/np_savetxt_5f.txt', a, fmt='%.5f')
with open('data/temp/np_savetxt_5f.txt') as f:
print(f.read())
np.savetxt('data/temp/np_savetxt_d.txt', a, fmt='%d')
with open('data/temp/np_savetxt_d.txt') as f:
print(f.read())
print(a * 10)
np.savetxt('data/temp/np_savetxt_x.txt', a * 10, fmt='%04x')
with open('data/temp/np_savetxt_x.txt') as f:
print(f.read())
np.savetxt('data/temp/np_savetxt.csv', a, delimiter=',', fmt='%d')
with open('data/temp/np_savetxt.csv') as f:
print(f.read())
np.savetxt('data/temp/np_savetxt.tsv', a, delimiter='\t', fmt='%d')
with open('data/temp/np_savetxt.tsv') as f:
print(f.read())
a_3d = np.arange(24).reshape(2, 3, 4)
print(a_3d)
np.savetxt('data/temp/np_savetxt_3d.txt', a_3d)
ValueError: Expected
D or 2
import numpy as np
a1 = np.arange(5)
print(a1)
a2 = np.arange(5, 10)
print(a2)
np.savez('data/temp/np_savez', a1, a2)
print(type(np.load('data/temp/np_savez.npz')))
numpy.lib.npyio.NpzFile
npz = np.load('data/temp/np_savez.npz')
print(npz.files)
print(npz['arr_0'])
print(npz['arr_1'])
np.savez('data/temp/np_savez_kw', x=a1, y=a2)
npz_kw = np.load('data/temp/np_savez_kw.npz')
print(npz_kw.files)
print(npz_kw['x'])
print(npz_kw['y'])
np.savez('data/temp/np_savez_kw2', a1, y=a2)
npz_kw2 = np.load('data/temp/np_savez_kw2.npz')
print(npz_kw2.files)
print(npz_kw2['arr_0'])
print(npz_kw2['y'])
import numpy as np
a1 = np.arange(5)
print(a1)
a2 = np.arange(5, 10)
print(a2)
np.savez_compressed('data/temp/np_savez_comp', a1, a2)
print(type(np.load('data/temp/np_savez_comp.npz')))
numpy.lib.npyio.NpzFile
npz_comp = np.load('data/temp/np_savez_comp.npz')
print(npz_comp.files)
print(npz_comp['arr_0'])
print(npz_comp['arr_1'])
np.savez_compressed('data/temp/np_savez_comp_kw', x=a1, y=a2)
npz_comp_kw = np.load('data/temp/np_savez_comp_kw.npz')
print(npz_comp_kw.files)
print(npz_comp_kw['x'])
print(npz_comp_kw['y'])
import numpy as np
a_2d = np.arange(12).reshape(3, 4)
print(a_2d)
a_2d[0, 0] = 100
print(a_2d)
a_2d[0] = 100
print(a_2d)
a_2d[np.ix_([False, True, True], [1, 3])] = 200
print(a_2d)
a_2d = np.arange(12).reshape(3, 4)
print(a_2d)
print(a_2d[::2, :3])
print(np.arange(6).reshape(2, 3) * 100)
a_2d[::2, :3] = np.arange(6).reshape(2, 3) * 100
print(a_2d)
a_2d = np.arange(12).reshape(3, 4)
print(a_2d)
print(a_2d[::2, :3])
print(np.arange(3) * 100)
a_2d[::2, :3] = np.arange(3) * 100
print(a_2d)
a_2d = np.arange(12).reshape(3, 4)
print(a_2d)
print(a_2d[::2, :3])
print(np.arange(2) * 100)
a_2d[::2, :3] = np.arange(2) * 100
ValueError: could
not broadcast 
shape (2) 
shape (2,3)
import numpy as np
a_1d = np.arange(10)
print(a_1d)
print(a_1d[3])
a_2d = np.arange(12).reshape(3, 4)
print(a_2d)
print(a_2d[0, 2])
a_3d = np.arange(24).reshape(2, 3, 4)
print(a_3d)
import numpy as np
a_1d = np.arange(4)
print(a_1d)
print(a_1d[[True, False, True, False]])
print(a_1d[np.array([True, False, True, False])])
print(a_1d[[True, False]])
IndexError: boolean
not match 
a_2d = np.arange(12).reshape(3, 4)
print(a_2d)
print(a_2d[[True, False, True]])
print(a_2d[:, [True, False, True, False]])
print(a_2d[[True, False, True], [True, False, True, False]])
print(a_2d[np.ix_([True, False, True], [True, False, True, False])])
print(a_2d[:, [True, False, False, False]])
print(a_2d > 5)
print(type(a_2d > 5))
numpy.ndarray
print(a_2d[a_2d > 5])
print((a_2d > 5) & (a_2d < 10))
print(a_2d[(a_2d > 5) & (a_2d < 10)])
import numpy as np
a_1d = np.arange(4)
print(a_1d)
print(a_1d[[0, 2]])
print(a_1d[[0, 3, 2, 1, 2, -1, -2]])
print(a_1d[np.array([0, 3, 2, 1, 2, -1, -2])])
a_2d = np.arange(12).reshape(3, 4)
print(a_2d)
print(a_2d[[0, 2]])
print(a_2d[:, [0, 2]])
print(a_2d[[0, 2], [0, 2]])
print(a_2d[np.ix_([0, 2], [0, 2])])
print(a_2d[np.ix_([0, 2, 1, 1, -1, -1], [0, 2, 1, 3])])
print(a_2d[:, [1]])
print(a_2d[:, [1]].shape)
print(a_2d[:, 1])
print(a_2d[:, 1].shape)
import numpy as np
a_1d = np.arange(10)
print(a_1d)
print(a_1d[0])
print(a_1d[4])
print(a_1d[-1])
print(a_1d[-4])
print(a_1d[100])
IndexError: index
print(a_1d[-100])
IndexError: index
a_2d = np.arange(12).reshape(3, 4)
print(a_2d)
print(a_2d[0, -1])
print(a_2d[0])
print(a_2d[0, :])
print(a_2d[:, 0])
print(type(a_2d[0, -1]))
numpy.int64
print(type(a_2d[0]))
numpy.ndarray
import numpy as np
a_2d = np.arange(12).reshape(3, 4)
print(a_2d)
print(a_2d[1, 1:3])
print(a_2d[1, [False, True, True, False]])
print(a_2d[1, [1, 2]])
print(a_2d[1:3, 1:3])
print(a_2d[1:3, [False, True, True, False]])
print(a_2d[1:3, [1, 2]])
print(a_2d[np.ix_([1, 2], [False, True, True, False])])
a_3d = np.arange(24).reshape(2, 3, 4)
print(a_3d)
np.ix_
SyntaxError: invalid
print(a_3d[np.ix_(range(2), [0, 2], [0, 2])])
print(a_3d[np.ix_(range(a_3d.shape[0]), [0, 2], [0, 2])])
print(a_3d[np.ix_(0, [0, 2], [0, 2])])
ValueError: Cross
print(a_3d[np.ix_([0], [0, 2], [0, 2])])
import numpy as np
a_1d = np.arange(10)
print(a_1d)
print(a_1d[2:7])
print(a_1d[:7])
print(a_1d[2:])
print(a_1d[2:7:2])
a_2d = np.arange(12).reshape(3, 4)
print(a_2d)
print(a_2d[:2, 1:3])
print(a_2d[:2, :])
print(a_2d[:2])
print(a_2d[1:2])
print(a_2d[1:2].shape)
print(a_2d[1])
print(a_2d[1].shape)
print(a_2d[:, 2:3])
print(a_2d[:, 2:3].shape)
print(a_2d[:, 2])
print(a_2d[:, 2].shape)
print(a_2d[1:2, 2:3])
print(a_2d[1:2, 2:3].shape)
print(a_2d[1, 2])
print(type(a_2d[1, 2]))
numpy.int64
a_3d = np.arange(24).reshape(2, 3, 4)
print(a_3d)
print(a_3d[0, 0])
print(a_3d[0])
print(a_3d[0:1, 0:1])
print(a_3d[0:1])
import numpy as np
a_2d = np.arange(12).reshape(3, 4)
print(a_2d)
a_slice = a_2d[:2, :2]
print(a_slice)
print(np.shares_memory(a_2d, a_slice))
a_slice[0, 0] = 100
print(a_slice)
print(a_2d)
a_2d[0, 0] = 0
print(a_2d)
print(a_slice)
a_fancy_index = a_2d[[0, 1]]
print(a_fancy_index)
print(np.shares_memory(a_2d, a_fancy_index))
a_fancy_index[0, 0] = 100
print(a_fancy_index)
print(a_2d)
a_slice_copy = a_2d[:2, :2].copy()
print(a_slice_copy)
print(np.shares_memory(a_2d, a_slice_copy))
a_slice_copy[0, 0] = 100
print(a_slice_copy)
print(a_2d)
a_fancy_index_slice = a_2d[[0, 1], :3]
print(a_fancy_index_slice)
print(np.shares_memory(a_2d, a_fancy_index_slice))
a_scalar_slice = a_2d[1, :3]
print(a_scalar_slice)
print(np.shares_memory(a_2d, a_scalar_slice))
import numpy as np
a = np.arange(6).reshape(2, 3)
print(a)
b = np.zeros_like(a)
print(b)
print(a)
a[:, :] = 0
print(a)
a[:] = 1
print(a)
a[:] = 0.1
print(a)
a = a.astype(np.float)
a[:] = 0.1
print(a)
import numpy as np
print(np.get_printoptions())
-', '
a = np.array([12.3456, 0.123456789])
print(a)
np.set_printoptions(precision=3)
print(a)
np.set_printoptions(precision=10)
print(a)
np.set_printoptions(precision=4, floatmode='maxprec')
print(a)
np.set_printoptions(precision=10, floatmode='maxprec')
print(a)
np.set_printoptions(precision=4, floatmode='fixed')
print(a)
np.set_printoptions(precision=10, floatmode='fixed')
print(a)
np.set_printoptions(precision=4, floatmode='maxprec_equal')
print(a)
np.set_printoptions(precision=10, floatmode='maxprec_equal')
print(a)
np.set_printoptions(precision=4, floatmode='unique')
print(a)
np.set_printoptions(precision=10, floatmode='unique')
print(a)
np.set_printoptions(precision=8, floatmode='maxprec')
b = np.round(a, 2)
print(b)
b = np.round(a, -1)
print(b)
b = np.round([1234.56, 123456.789], -2)
print(b)
a = np.array([0.123456, 0.123456])
print(a)
a = np.array([0.123456, 0.0000123456])
print(a)
a = np.array([123.456, 0.0123456])
print(a)
np.set_printoptions(suppress=True)
print(a)
np.set_printoptions(suppress=True, precision=2)
print(a)
np.set_printoptions(suppress=False, precision=2)
print(a)
np.set_printoptions(suppress=False, precision=8, floatmode='fixed')
print(a)
np.set_printoptions(precision=8, floatmode='maxprec', suppress=False)
a = np.array([123.456, 0.0123456])
print(a)
np.set_printoptions(formatter={'float': '{:.2f}'.format})
print(a)
np.set_printoptions(formatter={'float': '{:.8f}'.format})
print(a)
np.set_printoptions(formatter={'float': '{:.2e}'.format})
print(a)
np.set_printoptions(formatter={'float': '{:.8e}'.format})
print(a)
a = np.array([12, 1234])
print(a)
np.set_printoptions(formatter={'int': '{:08d}'.format})
print(a)
np.set_printoptions(formatter={'int': '{:b}'.format})
print(a)
np.set_printoptions(formatter={'int': '{:o}'.format})
print(a)
np.set_printoptions(formatter={'int': '{:x}'.format})
print(a)
a = np.array(['One', 'Two'])
print(a)
np.set_printoptions(formatter={'numpystr': str.upper})
print(a)
np.set_printoptions(formatter={'numpystr': lambda x: '***' + x + '***'})
print(a)
np.set_printoptions(formatter={'float': '{:0.8e}'.format, 'int': '{:08d}'.format})
a = np.array([12, 12.34])
print(a.dtype)
print(a)
a = np.array([12, 123])
print(a.dtype)
print(a)
import numpy as np
a = np.arange(10)
print(a)
np.set_printoptions(threshold=10)
print(a)
np.set_printoptions(threshold=9)
print(a)
a = np.arange(100).reshape((10, 10))
np.set_printoptions(threshold=100)
print(a)
np.set_printoptions(threshold=99)
print(a)
print(np.get_printoptions())
-', '
print(np.get_printoptions()['threshold'])
np.set_printoptions(threshold=0)
print(np.arange(100))
print(np.arange(7))
print(np.arange(6))
np.set_printoptions(edgeitems=1)
print(np.arange(6))
np.set_printoptions(edgeitems=0)
print(np.arange(6))
np.set_printoptions(edgeitems=3)
print(np.arange(60).reshape((6, 10)))
np.set_printoptions(edgeitems=2)
print(np.arange(60).reshape((6, 10)))
np.set_printoptions(threshold=np.inf)
print(np.arange(1001))
np.set_printoptions(linewidth=100)
print(np.arange(1001))
import numpy as np
a = np.arange(6)
print(a)
a_reshape = a.reshape(2, 3)
print(a_reshape)
print(np.shares_memory(a, a_reshape))
a_slice = a[2:5]
print(a_slice)
print(np.shares_memory(a_reshape, a_slice))
a_reshape_copy = a.reshape(2, 3).copy()
print(a_reshape_copy)
print(np.shares_memory(a, a_reshape_copy))
import numpy as np
a = np.arange(10)
print(a)
a_0 = a[::2]
print(a_0)
a_1 = a[1::2]
print(a_1)
print(np.shares_memory(a_0, a_1))
print(np.may_share_memory(a_0, a_1))
a_2 = a[:5]
print(a_2)
a_3 = a[5:]
print(a_3)
print(np.shares_memory(a_2, a_3))
print(np.may_share_memory(a_2, a_3))
np.shares_memory(a_0, a_1)
std. dev. of
np.may_share_memory(a_0, a_1)
std. dev. of
import numpy as np
a = np.array([-100, -10, 0, 10, 100])
print(a)
print(np.sign(a))
print(type(np.sign(a)))
numpy.ndarray
print(np.sign(a).dtype)
a_float = np.array([-1.23, 0.0, 1.23])
print(a_float)
-1.23
print(np.sign(a_float))
-1.  0
print(np.sign(a_float).dtype)
print(np.sign(100))
print(type(np.sign(100)))
numpy.int64
print(np.sign(-1.23))
-1.0
print(type(np.sign(-1.23)))
numpy.float64
a_special = np.array([0.0, -0.0, np.inf, -np.inf, np.nan])
print(a_special)
print(np.sign(a_special))
print(np.sign(a_special).dtype)
a_complex = np.array
-10
-10
np.nan
print(a_complex)
print(np.sign(a_complex))
print(a_complex.real)
print(np.sign(a_complex.real))
print(a_complex.imag)
print(np.sign(a_complex.imag))
import numpy as np
a = np.array([-100, -10, 0, 10, 100])
print(a)
print(np.signbit(a))
print(type(np.signbit(a)))
numpy.ndarray
print(np.signbit(a).dtype)
print(np.signbit(-100))
print(a == 0)
print(a > 0)
print(a >= 0)
print(a < 0)
print(a <= 0)
print(np.count_nonzero(np.signbit(a)))
np.signbit(a)
np.count_nonzero
np.signbit(a)
print(np.count_nonzero(a == 0))
print(np.count_nonzero(a < 0))
print(np.count_nonzero(a > 0))
a_special = np.array([0.0, -0.0, np.inf, -np.inf, np.nan])
print(a_special)
print(np.signbit(a_special))
print(a_special == 0)
print(a_special < 0)
RuntimeWarning: invalid
print(a_special > 0)
RuntimeWarning: invalid
a_complex = np.array
print(a_complex)
print(np.signbit(a_complex))
TypeError: ufunc
not supported for the input
not be 
print(np.abs(a_complex))
print(a_complex.real)
print(a_complex.imag)
print(np.signbit(a_complex.real))
print(a_complex.real < 0)
import numpy as np
print(np.__version__)
print(np.pi)
print(np.radians(180))
print(type(np.radians(180)))
numpy.float64
a = np.array([0, 90, 180])
print(type(a))
numpy.ndarray
print(np.radians(a))
print(type(np.radians(a)))
numpy.ndarray
l = [0, 90, 180]
print(type(l))
print(np.radians(l))
print(type(np.radians(l)))
numpy.ndarray
print(np.radians(a))
print(np.round(np.radians(a)))
print(np.round(np.radians(a), 2))
print(np.sin(np.radians(30)))
print(np.round(np.sin(np.radians(30)), 1))
print(np.sin(np.radians([0, 30, 90])))
print(np.sin(np.radians([0, 30, 90]))[1])
np.set_printoptions(precision=20)
print(np.sin(np.radians([0, 30, 90])))
np.set_printoptions(precision=8)  
print(np.radians([0, 90, 180]))
print(np.deg2rad([0, 90, 180]))
print(np.degrees([0, np.pi / 2, np.pi]))
print(np.rad2deg([0, np.pi / 2, np.pi]))
print(np.sin(np.radians([0, 30, 90])))
print(np.degrees(np.arcsin([0, 0.5, 1])))
print(np.cos(np.radians([0, 60, 90])))
print(np.round(np.cos(np.radians([0, 60, 90])), 1))
print(np.degrees(np.arccos([0, 0.5, 1])))
print(np.tan(np.radians([0, 45, 90])))
print(np.degrees(np.arctan([0, 1, np.inf])))
print(np.degrees(np.arctan([-np.inf, -1, 0, 1, np.inf])))
print(np.degrees(np.arctan2(-1, 1)))
-45.0
print(np.degrees(np.arctan2(1, -1)))
np.degrees
np.arctan2
-1
-1
np.degrees
np.arctan2
-1
-1
-1
-1
-1
print(np.degrees(np.arctan2(0, -1)))
print(np.degrees(np.arctan2(-0.0, -1.0)))
-180.0
print(np.degrees(np.arctan2(-0, -1)))
np.degrees
np.arctan2
-0.0
-0.0
-0.0
-0.0
-180.
print(np.sin(-0.0))
-0.0
print(np.arcsin(-0.0))
-0.0
print(np.tan(-0.0))
-0.0
print(np.arctan(-0.0))
-0.0
import numpy as np
import skimage.io
import skimage.util
img = skimage.io.imread('data/src/lena_square.png')
print(img.shape)
def split_image_cut(img, div_v, div_h):
h, w = img.shape[:2]
block_h, out_h = divmod(h, div_v)
block_w, out_w = divmod(w, div_h)
block_shape = (block_h, block_w, 3) if len(img.shape) == 3 else (block_h, block_w)
return skimage.util.view_as_blocks(img[:h - out_h, :w - out_w], block_shape)
blocks = split_image_cut(img, 2, 3)
print(blocks.shape)
def split_image_unequal(img, div_v, div_h):
l_v = np.array_split(img, div_v)
return [np.array_split(img_v, div_h, 1) for img_v in l_v]
l = split_image_unequal(img, 2, 3)
print(type(l))
print(len(l))
print(type(l[0]))
print(len(l[0]))
print(type(l[0][0]))
numpy.ndarray
print(l[0][0].shape)
print(l[0][1].shape)
print(l[0][2].shape)
print(l[1][0].shape)
print(l[1][1].shape)
print(l[1][2].shape)
l = split_image_unequal(img, 2, 3)
print(np.shares_memory(img, l[0][0]))
l_copy = split_image_unequal(img.copy(), 2, 3)
print(np.shares_memory(img, l_copy[0][0]))
import numpy as np
l = list(range(10))
print(l)
print(l[4:8])
print(l[-5:-2])
print(l[::-1])
a = np.arange(10)
print(a)
print(a[4:8])
print(a[-5:-2])
print(a[::-1])
a[3:6] = 100
print(a)
a[3:6] = [100, 200, 300]
print(a)
a[3:6] = [100, 200, 300, 400]
ValueError: cannot
a = np.arange(10)
print(a)
print(a[2:8:2])
a[2:8:2] = 100
print(a)
a[2:8:2] = [100, 200, 300]
print(a)
a = np.arange(12).reshape((3, 4))
print(a)
print(a[1:, 1:3])
print(a[1:, :])
print(a[1:])
print(a[1])
print(a[1].shape)
print(a[1:2])
print(a[1:2].shape)
print(a[:, 1:3])
print(a[:, 1])
print(a[:, 1].shape)
print(a[:, 1:2])
print(a[:, 1:2].shape)
a = np.arange(12).reshape((3, 4))
print(a)
print(a[1:, 1:3])
a[1:, 1:3] = 100
print(a)
a[1:, 1:3] = [100, 200]
print(a)
a[1:, 1:3] = [[100, 200], [300, 400]]
print(a)
a = np.arange(12).reshape((3, 4))
print(a)
print(a[1:, ::2])
a[1:, ::2] = 100
print(a)
a[1:, ::2] = [100, 200]
print(a)
a[1:, ::2] = [[100, 200], [300, 400]]
print(a)
a = np.arange(12).reshape((3, 4))
print(a)
a_slice = a[1:, 1:3]
print(a_slice)
a_slice[0, 0] = 100
print(a_slice)
print(a)
a = np.arange(12).reshape((3, 4))
print(a)
a_slice_copy = a[1:, 1:3].copy()
print(a_slice_copy)
a_slice_copy[0, 0] = 100
print(a_slice_copy)
print(a)
a = np.arange(12).reshape((3, 4))
print(a)
print(a[[0, 2], 1:3])
a[[0, 2], 1:3] = 100
print(a)
a[[0, 2], 1:3] = [100, 200]
print(a)
a[[0, 2], 1:3] = [[100, 200], [300, 400]]
print(a)
a_subset = a[[0, 2], 1:3]
print(a_subset)
a_subset[0, 0] = -1
print(a_subset)
-1
print(a)
import numpy as np
a = np.array([3, 4, 2, 0, 1])
print(a)
a_sort = np.sort(a)
print(a_sort)
print(a)
a_sort_reverse = np.sort(a)[::-1]
print(a_sort_reverse)
a_2d = np.array([[20, 3, 100], [1, 200, 30], [300, 10, 2]])
print(a_2d)
a_2d_sort_col = np.sort(a_2d, axis=0)
print(a_2d_sort_col)
a_2d_sort_row = np.sort(a_2d, axis=1)
print(a_2d_sort_row)
a_2d_sort_row = np.sort(a_2d)
print(a_2d_sort_row)
a_2d_sort_row = np.sort(a_2d, axis=-1)
print(a_2d_sort_row)
a_2d_sort_col_reverse = np.sort(a_2d, axis=0)[::-1]
print(a_2d_sort_col_reverse)
a_2d_sort_row_reverse = np.sort(a_2d, axis=1)[:, ::-1]
print(a_2d_sort_row_reverse)
print(a_2d)
a_2d.sort()
print(a_2d)
a_2d.sort(axis=0)
print(a_2d[::-1])
a_2d = np.array([[20, 3, 100], [1, 200, 30], [300, 10, 2]])
print(a_2d)
a_2d_sort_col_index = np.argsort(a_2d, axis=0)
print(a_2d_sort_col_index)
a_2d_sort_row_index = np.argsort(a_2d)
print(a_2d_sort_row_index)
print(a_2d)
col_num = 1
print(a_2d[:, col_num])
print(np.argsort(a_2d[:, col_num]))
a_2d_sort_col_num = a_2d[np.argsort(a_2d[:, col_num])]
print(a_2d_sort_col_num)
print(np.argsort(a_2d[:, col_num])[::-1])
a_2d_sort_col_num_reverse = a_2d[np.argsort(a_2d[:, col_num])[::-1]]
print(a_2d_sort_col_num_reverse)
print(a_2d)
row_num = 1
print(a_2d[row_num])
print(np.argsort(a_2d[row_num]))
a_2d_sort_row_num = a_2d[:, np.argsort(a_2d[row_num])]
print(a_2d_sort_row_num)
print(np.argsort(a_2d[row_num])[::-1])
a_2d_sort_row_num_inverse = a_2d[:, np.argsort(a_2d[row_num])[::-1]]
print(a_2d_sort_row_num_inverse)
import numpy as np
a = np.arange(16).reshape(4, 4)
print(a)
a_split = np.split(a, 2)
print(type(a_split))
print(len(a_split))
print(a_split[0])
print(a_split[1])
print(type(a_split[0]))
numpy.ndarray
print(a)
a0, a1 = np.split(a, 2)
print(a0)
print(a1)
np.split(a, 3)
ValueError: array
not result in an 
a0, a1, a2 = np.split(a, [1, 3])
print(a0)
print(a1)
print(a2)
a0, a1 = np.split(a, [1])
print(a0)
print(a1)
a0, a1 = np.split(a, 2, 0)
print(a0)
print(a1)
a0, a1 = np.split(a, 2, 1)
print(a0)
print(a1)
np.split(a, 2, 2)
IndexError: tuple
a_3d = np.arange(24).reshape(2, 3, 4)
print(a_3d)
print(a_3d.shape)
a0, a1 = np.split(a_3d, 2, 0)
print(a0)
print(a1)
a0, a1 = np.split(a_3d, [1], 2)
print(a0)
print(a1)
import numpy as np
a = np.arange(16).reshape(4, 4)
print(a)
l = np.split(a, 2)
print(l[0])
print(np.shares_memory(a, l[0]))
a[0, 0] = 100
print(a)
print(l[0])
print(np.shares_memory(a, np.vsplit(a, 2)[0]))
print(np.shares_memory(a, np.hsplit(a, 2)[0]))
print(np.shares_memory(a, np.array_split(a, 3)[0]))
a_3d = np.arange(24).reshape(2, 3, 4)
print(np.shares_memory(a_3d, np.dsplit(a_3d, 2)[0]))
a = np.arange(16).reshape(4, 4)
l_copy = np.split(a.copy(), 2)
print(np.shares_memory(a, l_copy[0]))
a[0, 0] = 100
print(a)
print(l_copy[0])
import numpy as np
a = np.arange(6).reshape(1, 2, 1, 3, 1)
print(a)
print(a.shape)
a_s = np.squeeze(a)
print(a_s)
print(a_s.shape)
print(np.shares_memory(a, a_s))
a_s_copy = np.squeeze(a).copy()
print(np.shares_memory(a, a_s_copy))
print(np.squeeze(a, 0))
print(np.squeeze(a, 0).shape)
print(np.squeeze(a, 1))
ValueError: cannot
not equal 
print(np.squeeze(a, 5))
AxisError: axis
print(np.squeeze(a, -1))
print(np.squeeze(a, -1).shape)
print(np.squeeze(a, -3))
print(np.squeeze(a, -3).shape)
print(np.squeeze(a, (0, -1)))
print(np.squeeze(a, (0, -1)).shape)
print(np.squeeze(a, (0, 1)))
ValueError: cannot
not equal 
print(a.squeeze())
print(a.squeeze().shape)
print(a.squeeze((0, -1)))
print(a.squeeze((0, -1)).shape)
a_s = a.squeeze()
print(a_s)
print(np.shares_memory(a, a_s))
print(a)
import numpy as np
a1 = np.ones(3, int)
print(a1)
print(a1.shape)
a2 = np.full(3, 2)
print(a2)
print(a2.shape)
print(np.stack([a1, a2]))
print(np.stack([a1, a2], 0))
print(np.stack([a1, a2], 0).shape)
print(np.stack([a1, a2], 1))
print(np.stack([a1, a2], 1).shape)
print(np.stack([a1, a2], 2))
AxisError: axis
print(np.stack([a1, a2], -1))
print(np.stack([a1, a2], -1).shape)
a2_ = np.full(4, 2)
print(a2_)
print(a2_.shape)
print(np.stack([a1, a2]))
ValueError: all
a1 = np.ones((3, 4), int)
print(a1)
print(a1.shape)
a2 = np.full((3, 4), 2)
print(a2)
print(a2.shape)
print(np.stack([a1, a2]))
print(np.stack([a1, a2], 0))
print(np.stack([a1, a2], 0).shape)
print(np.stack([a1, a2], 1))
print(np.stack([a1, a2], 1).shape)
np.stack([a1, a2], 1)
np.stack([a1, a2], 1)
print(np.stack([a1, a2], 2))
print(np.stack([a1, a2], 2).shape)
np.stack([a1, a2], 2)
np.stack([a1, a2], 2)
print(np.stack([a1, a2], 3))
AxisError: axis
print(np.stack([a1, a2], -1))
print(np.stack([a1, a2], -1).shape)
a2_ = np.full((2, 3), 2)
print(a2_)
print(np.stack([a1, a2_]))
ValueError: all
import numpy as np
a = np.arange(12).reshape(3, 4)
print(a.shape)
print(a)
print(np.sum(a))
print(np.sum(a, axis=0))
print(np.sum(a, axis=1))
print(a.sum())
print(a.sum(axis=0))
print(a.sum(axis=1))
print(np.mean(a))
print(np.mean(a, axis=0))
print(np.mean(a, axis=1))
print(a.mean())
print(a.mean(axis=0))
print(a.mean(axis=1))
print(np.min(a))
print(np.min(a, axis=0))
print(a.max())
print(a.max(axis=1))
b = np.arange(24).reshape(2, 3, 4)
print(b.shape)
print(b)
print(b.sum(axis=0))
+ b
print(b.sum(axis=1))
+ b
+ b
print(b.sum(axis=2))
+ b
+ b
+ b
print(b.sum(axis=(0, 1)))
print(b.sum(axis=(0, 2)))
print(b.sum(axis=(1, 2)))
import numpy as np
a = np.arange(10, 35).reshape(5, 5)
print(a)
col_swap = a[:, [3, 2, 4, 0, 1]]
print(col_swap)
col_inverse = a[:, ::-1]
print(col_inverse)
col_select = a[:, [2, 4, 0]]
print(col_select)
col_select2 = a[:, [2, 2, 2]]
print(col_select2)
row_swap = a[[3, 2, 4, 0, 1], :]
print(row_swap)
row_swap = a[[3, 2, 4, 0, 1]]
print(row_swap)
row_inverse = a[::-1]
print(row_inverse)
row_select = a[[2, 4, 0]]
print(row_select)
row_select2 = a[[2, 2, 2]]
print(row_select2)
import numpy as np
a = np.arange(16).reshape(4, 4)
print(a)
print(np.tril(a))
print(np.triu(a))
print(np.tril(a).T)
print(np.tril(a) + np.tril(a).T)
print(a.diagonal())
print(np.diag(a.diagonal()))
print(np.tril(a) + np.tril(a).T - np.diag(a.diagonal()))
def get_symmetric(a, use_tril=True):
if use_tril:
a = np.tril(a)
a = np.triu(a)
return a + a.T - np.diag(a.diagonal())
print(get_symmetric(a))
print(get_symmetric(a, False))
def is_symmetric(a):
return np.array_equal(a, a.T)
a_sym = get_symmetric(a)
print(a_sym)
print(is_symmetric(a_sym))
print(is_symmetric(a))
def get_skew_symmetric(a, use_tril=True):
if use_tril:
a = np.tril(a)
a = np.triu(a)
return a - a.T
print(get_skew_symmetric(a))
print(get_skew_symmetric(a, False))
-1
def is_skew_symmetric(a):
return np.array_equal(a, -a.T)
a_sk_sym = get_skew_symmetric(a)
print(a_sk_sym)
print(is_skew_symmetric(a_sk_sym))
print(is_skew_symmetric(a))
import numpy as np
a = np.array([0, 1, 2, 3])
print(np.tile(a, 2))
print(np.tile(a, (3, 2)))
print(np.tile(a, (2, 1)))
a = np.array([[11, 12], [21, 22]])
print(np.tile(a, 2))
print(np.tile(a, (3, 2)))
print(np.tile(a, (2, 1)))
import numpy as np
a = np.arange(3)
print(a)
a_tile = np.tile(a, 3)
print(a_tile)
print(a)
print(np.tile(a, (2, 3)))
print(np.tile(a, (2, 3)).shape)
print(np.tile(a, (2, 3, 4)))
print(np.tile(a, (2, 3, 4)).shape)
a_2d = np.arange(6).reshape(2, 3)
print(a_2d)
print(np.tile(a_2d, 2))
print(np.tile(a_2d, (2, )))
print(np.tile(a_2d, (1, 2)))
print(np.tile(a_2d, (2, 1)))
print(np.tile(a_2d, (2, 2, 2)))
print(np.tile(a_2d, (2, 2, 2)).shape)
import numpy as np
from PIL import Image
img = np.array(Image.open('data/src/lena_square.png').resize((128, 128)))
print(img.shape)
img_tile = np.tile(img, (2, 3, 1))
print(img_tile.shape)
Image.fromarray(img_tile).save('data/dst/lena_numpy_tile.jpg')
import numpy as np
import pandas as pd
a = np.arange(4)
print(a)
s = pd.Series(a)
print(s)
dtype: int64
index = ['A', 'B', 'C', 'D']
name = 'sample'
s = pd.Series(data=a, index=index, name=name, dtype='float')
print(s)
Name: sample
dtype: float64
a = np.arange(12).reshape((4, 3))
print(a)
s = pd.Series(a)
print(s)
Exception: Data
s = pd.Series(a[2])
print(s)
dtype: int64
s = pd.Series(a.T[2])
print(s)
dtype: int64
a = np.arange(12).reshape((4, 3))
print(a)
df = pd.DataFrame(a)
print(df)
index = ['A', 'B', 'C', 'D']
columns = ['a', 'b', 'c']
df = pd.DataFrame(data=a, index=index, columns=columns, dtype='float')
print(df)
import numpy as np
a_2d = np.arange(6).reshape(2, 3)
print(a_2d)
a_2d_T = a_2d.T
print(a_2d_T)
print(np.shares_memory(a_2d, a_2d_T))
a_2d_T[0, 1] = 100
print(a_2d_T)
print(a_2d)
a_2d[1, 0] = 3
print(a_2d)
print(a_2d_T)
a_2d_T_copy = a_2d.T.copy()
print(a_2d_T_copy)
print(np.shares_memory(a_2d, a_2d_T_copy))
a_2d_T_copy[0, 1] = 100
print(a_2d_T_copy)
print(a_2d)
a_2d = np.arange(6).reshape(2, 3)
print(a_2d)
print(a_2d.transpose())
print(np.shares_memory(a_2d, a_2d.transpose()))
print(np.transpose(a_2d))
print(np.shares_memory(a_2d, np.transpose(a_2d)))
a_1d = np.arange(3)
print(a_1d)
print(a_1d.T)
print(a_1d.transpose())
print(np.transpose(a_1d))
a_row = a_1d.reshape(1, -1)
print(a_row)
print(a_row.shape)
print(a_row.ndim)
a_col = a_1d.reshape(-1, 1)
print(a_col)
print(a_col.shape)
print(a_col.ndim)
print(a_row.T)
print(a_col.T)
a_3d = np.arange(24).reshape(2, 3, 4)
print(a_3d)
print(a_3d.T)
print(a_3d.T.shape)
print(a_3d.transpose())
print(a_3d.transpose().shape)
print(a_3d.transpose(2, 1, 0))
print(a_3d.transpose(2, 1, 0).shape)
print(a_3d.transpose((2, 1, 0)).shape)
print(np.transpose(a_3d, (2, 1, 0)))
print(np.transpose(a_3d, (2, 1, 0)).shape)
print(np.transpose(a_3d, 2, 1, 0))
TypeError: transpose
print(a_3d.transpose(0, 1))
ValueError: axes
print(a_3d.transpose(0, 1, 2, 3))
ValueError: axes
print(a_3d.transpose(0, 1, 3))
AxisError: axis
print(a_3d)
print(a_3d.shape)
print(a_3d.transpose(0, 2, 1))
print(a_3d.transpose(0, 2, 1).shape)
print(a_3d.transpose(1, 0, 2))
print(a_3d.transpose(1, 0, 2).shape)
a_3d.transpose(1, 0, 2)
import numpy as np
print(np.tri(4))
print(type(np.tri(4)))
numpy.ndarray
print(np.tri(4).dtype)
print(np.tri(4, dtype=int))
print(np.tri(4, dtype=int).dtype)
print(np.tri(4, k=1))
print(np.tri(4, k=-1))
print(np.tri(3, 4))
print(np.tri(3, 4, k=-1))
print(np.tri(4).T)
print(np.tri(3, 4).T)
print(np.tri(4, k=1).T)
print(np.tri(4, k=-1).T)
import numpy as np
a = np.arange(16).reshape(4, 4)
print(a)
print(np.tril(a))
print(np.tril(a, k=2))
print(np.tril(a, k=-1))
a = np.arange(12).reshape(3, 4)
print(a)
print(np.tril(a))
print(np.tril(a, k=-1))
print(np.tril(np.arange(32).reshape(2, 4, 4)))
print(np.tril(np.arange(16).reshape(1, 1, 4, 4)))
a_tril = np.tril(np.arange(16).reshape(4, 4))
print(a_tril)
print(a_tril.T)
print(a_tril.T.T)
import numpy as np
a = np.arange(16).reshape(4, 4)
print(a)
print(np.triu(a))
print(np.triu(a, k=2))
print(np.triu(a, k=-1))
a = np.arange(12).reshape(3, 4)
print(a)
print(np.triu(a))
print(np.triu(a, k=-1))
print(np.triu(np.arange(16).reshape(1, 1, 4, 4)))
import numpy as np
import pandas as pd
with open('data/src/sample_header_index.csv') as f:
print(f.read())
df = pd.read_csv('data/src/sample_header_index.csv', index_col=0)
print(df)
a = df.values
print(a)
print(type(a))
numpy.ndarray
a = np.arange(6).reshape(2, 3)
print(a)
df = pd.DataFrame(a, index=['ONE', 'TWO'], columns=['a', 'b', 'c'])
print(df)
df.to_csv('data/temp/sample_pd.csv')
with open('data/temp/sample_pd.csv') as f:
print(f.read())
with open('data/src/sample_nan.csv') as f:
print(f.read())
df = pd.read_csv('data/src/sample_nan.csv', header=None)
print(df)
with open('data/src/sample_pandas_normal.csv') as f:
print(f.read())
df = pd.read_csv('data/src/sample_pandas_normal.csv')
print(df)
print(df.dtypes)
dtype: object
print(df.select_dtypes('int'))
a = pd.read_csv('data/src/sample_pandas_normal.csv').select_dtypes('int').values
print(a)
print(type(a))
numpy.ndarray
print(a.dtype)
import numpy as np
a = np.array([0, 0, 30, 10, 10, 20])
print(a)
print(np.unique(a))
print(type(np.unique(a)))
numpy.ndarray
l = [0, 0, 30, 10, 10, 20]
print(l)
print(np.unique(l))
print(type(np.unique(l)))
numpy.ndarray
print(np.unique(a).size)
print(len(np.unique(a)))
u, counts = np.unique(a, return_counts=True)
print(u)
print(counts)
print(u[counts == 1])
print(u[counts != 1])
print(np.unique(a, return_counts=True))
array([ 0, 10, 20, 30])
array([2, 2, 1, 1])
print(type(np.unique(a, return_counts=True)))
u, indices = np.unique(a, return_index=True)
print(u)
print(indices)
print(a)
print(a[indices])
u, inverse = np.unique(a, return_inverse=True)
print(u)
print(inverse)
print(a)
print(u[inverse])
u, indices, inverse, counts = np.unique(a, return_index=True, return_inverse=True, return_counts=True)
print(u)
print(indices)
print(inverse)
print(counts)
print(np.unique(a, return_counts=True, return_index=True, return_inverse=True))
array([ 0, 10, 20, 30])
array([0, 3, 5, 2])
array([0, 0, 3, 1, 1, 2])
array([2, 2, 1, 1])
a_2d = np.array([[20, 20, 10, 10], [0, 0, 10, 30], [20, 20, 10, 10]])
print(a_2d)
print(np.unique(a_2d))
print(np.unique(a_2d, axis=0))
print(np.unique(a_2d, axis=1))
print(a_2d[0])
print(np.unique(a_2d[0]))
print(a_2d[:, 2])
print(np.unique(a_2d[:, 2]))
print([np.unique(row) for row in a_2d])
array([10, 20])
array([ 0, 10, 30])
array([10, 20])
print([np.unique(row).tolist() for row in a_2d])
print([np.unique(row).size for row in a_2d])
print(a_2d.T)
print([np.unique(row) for row in a_2d.T])
array([ 0, 20])
array([ 0, 20])
array([10])
array([10, 30])
print(a_2d.shape)
print([np.unique(a_2d[:, i]) for i in range(a_2d.shape[1])])
array([ 0, 20])
array([ 0, 20])
array([10])
array([10, 30])
u, indices, inverse, counts = np.unique(a_2d, return_index=True, return_inverse=True, return_counts=True)
print(u)
print(indices)
print(a_2d.flatten())
print(a_2d.flatten()[indices])
print(inverse)
print(u[inverse])
print(u[inverse].reshape(a_2d.shape))
print(counts)
u, indices, inverse, counts = np.unique(a_2d, axis=0, return_index=True, return_inverse=True, return_counts=True)
print(u)
print(indices)
print(a_2d[indices])
print(inverse)
print(u[inverse])
print(counts)
print(a_2d)
u, indices = np.unique(a_2d, return_index=True)
print(u)
print(a_2d.flatten())
print(indices)
print(list(zip(*np.where(a_2d == 0))))
d = {u: list(zip(*np.where(a_2d == u))) for u in np.unique(a_2d)}
print(d)
print(d[0])
print(d[10])
print(d[20])
print(d[30])
u: list
zip(*np.where(a_2d == u))
c == 1
print(d)
u: list
zip(*np.where(a_2d == u))
print(d)
import numpy as np
v = np.array([0, 1, 2])
print(v)
print(v.shape)
print(np.inner(v, v))
print(type(np.inner(v, v)))
numpy.int64
print(np.dot(v, v))
print(np.matmul(v, v))
arr_row = np.arange(3).reshape(1, 3)
print(arr_row)
print(arr_row.shape)
arr_col = np.arange(3).reshape(3, 1)
print(arr_col)
print(arr_col.shape)
arr = np.arange(9).reshape(3, 3)
print(arr)
import numpy as np
print(np.__version__)
print(np.version.version)
print(np.version.short_version)
print(np.version.full_version)
print(np.version.git_revision)
print(np.version.release)
import numpy as np
a = np.arange(16).reshape(4, 4)
print(a)
a0, a1 = np.vsplit(a, 2)
print(a0)
print(a1)
a0, a1 = np.split(a, [1])
print(a0)
print(a1)
import numpy as np
a1 = np.ones((2, 3), int)
print(a1)
a2 = np.full((2, 3), 2)
print(a2)
print(np.vstack([a1, a2]))
print(np.concatenate([a1, a2], 0))
a1 = np.ones(3, int)
print(a1)
a2 = np.full(3, 2)
print(a2)
print(np.vstack([a1, a2]))
print(np.stack([a1, a2], 0))
a1 = np.ones((2, 3), int)
print(a1)
a2 = np.full(3, 2)
print(a2)
print(np.vstack([a1, a2]))
import numpy as np
a = np.arange(9).reshape((3, 3))
print(a)
print(np.where(a < 4, -1, 100))
-1
print(np.where(a < 4, True, False))
print(a < 4)
print(np.where((a > 2) & (a < 6), -1, 100))
print(np.where((a > 2) & (a < 6) | (a == 7), -1, 100))
print((a > 2) & (a < 6))
print((a > 2) & (a < 6) | (a == 7))
print(np.where(a < 4, -1, a))
-1
print(np.where(a < 4, a, 100))
a_org = np.arange(9).reshape((3, 3))
print(a_org)
a_new = np.where(a_org < 4, -1, a_org)
print(a_new)
-1
print(a_org)
a_org[a_org < 4] = -1
print(a_org)
-1
print(np.where(a < 4, a * 10, a))
print(np.where(a < 4))
array([0, 0, 0, 1])
array([0, 1, 2, 0])
print(type(np.where(a < 4)))
print(list(zip(*np.where(a < 4))))
a_3d = np.arange(24).reshape(2, 3, 4)
print(a_3d)
print(np.where(a_3d < 5))
array([0, 0, 0, 0, 0])
array([0, 0, 0, 0, 1])
array([0, 1, 2, 3, 0])
print(list(zip(*np.where(a_3d < 5))))
a_1d = np.arange(6)
print(a_1d)
print(np.where(a_1d < 3))
array([0, 1, 2])
print(list(zip(*np.where(a_1d < 3))))
print(np.where(a_1d < 3)[0])
print(np.where(a_1d < 3)[0].tolist())
import numpy as np
a_2d = np.array([[10, 20, 30], [30, 20, 10]])
print(a_2d)
print(np.max(a_2d))
print(np.min(a_2d))
print(np.unravel_index(np.argmax(a_2d), a_2d.shape))
print(np.unravel_index(np.argmin(a_2d), a_2d.shape))
print(np.where(a_2d == np.max(a_2d)))
array([0, 1])
array([2, 0])
print(list(zip(*np.where(a_2d == np.max(a_2d)))))
print(np.where(a_2d == np.min(a_2d)))
array([0, 1])
array([0, 2])
print(list(zip(*np.where(a_2d == np.min(a_2d)))))
import numpy as np
print(np.zeros(3))
print(np.zeros((2, 3)))
print(np.zeros(3).dtype)
print(np.zeros(3, dtype=np.int))
print(np.zeros(3, dtype=np.int).dtype)
print(np.ones(3))
print(np.ones((2, 3)))
print(np.ones(3).dtype)
print(np.ones(3, dtype=np.int))
print(np.ones(3, dtype=np.int).dtype)
print(np.full(3, 100))
print(np.full(3, np.pi))
print(np.full((2, 3), 100))
print(np.full((2, 3), np.pi))
print(np.full(3, 100).dtype)
print(np.full(3, 100.0).dtype)
print(np.full(3, np.pi).dtype)
print(np.full(3, 100, dtype=float))
print(np.full(3, np.pi, dtype=int))
import numpy as np
a_int = np.arange(6).reshape((2,3))
print(a_int)
a_float = np.arange(6).reshape((2,3)) / 10
print(a_float)
print(np.zeros_like(a_int))
print(np.zeros_like(a_float))
print(np.zeros_like(a_int, dtype=np.float))
print(np.ones_like(a_int))
print(np.ones_like(a_float))
print(np.ones_like(a_int, dtype=np.float))
print(np.full_like(a_int, 100))
print(np.full_like(a_float, 100))
print(np.full_like(a_int, 0.123))
print(np.full_like(a_float, 0.123))
print(np.full_like(a_int, 0.123, dtype=np.float))
import cv2
face_cascade_path = '/usr/local/opt/opencv3/share/'
eye_cascade_path = '/usr/local/opt/opencv3/share/'
face_cascade = cv2.CascadeClassifier(face_cascade_path)
eye_cascade = cv2.CascadeClassifier(eye_cascade_path)
cap = cv2.VideoCapture(0)
while True:
ret, img = cap.read()
gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
faces = face_cascade.detectMultiScale(gray, 1.3, 5)
for (x, y, w, h) in faces:
cv2.rectangle(img, (x, y), (x + w, y + h), (255, 0, 0), 2)
roi_gray = gray[y: y + h, x: x + w]
roi_color = img[y: y + h, x: x + w]
eyes = eye_cascade.detectMultiScale(roi_gray)
for (ex, ey, ew, eh) in eyes:
cv2.rectangle(roi_color, (ex, ey), (ex + ew, ey + eh), (0, 255, 0), 2)
cv2.imshow('video image', img)
key = cv2.waitKey(10)
if key == 27:  
break
cap.release()
cv2.destroyAllWindows()
import cv2
face_cascade_path = '/usr/local/opt/opencv3/share/'
eye_cascade_path = '/usr/local/opt/opencv3/share/'
face_cascade = cv2.CascadeClassifier(face_cascade_path)
eye_cascade = cv2.CascadeClassifier(eye_cascade_path)
img = cv2.imread('data/src/lena_square.png')
gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
faces = face_cascade.detectMultiScale(gray, scaleFactor=1.3, minNeighbors=5)
for (x, y, w, h) in faces:
cv2.rectangle(img, (x, y), (x + w, y + h), (255, 0, 0), 2)
roi_gray = gray[y: y + h, x: x + w]
roi_color = img[y: y + h, x: x + w]
eyes = eye_cascade.detectMultiScale(roi_gray)
for (ex, ey, ew, eh) in eyes:
cv2.rectangle(roi_color, (ex, ey), (ex + ew, ey + eh), (0, 255, 0), 2)
cv2.imwrite('data/dst/lena_opencv_face_detect.jpg', img)
import cv2
src1 = cv2.imread('data/src/lena.jpg')
src2 = cv2.imread('data/src/rocket.jpg')
src2 = cv2.resize(src2, src1.shape[1::-1])
dst = cv2.addWeighted(src1, 0.5, src2, 0.5, 0)
cv2.imwrite('data/dst/opencv_add_weighted.jpg', dst)
dst = cv2.addWeighted(src1, 0.5, src2, 0.2, 128)
cv2.imwrite('data/dst/opencv_add_weighted_gamma.jpg', dst)
import cv2
src1 = cv2.imread('data/src/lena.jpg')
src2 = cv2.imread('data/src/horse_r.png')
src2 = cv2.resize(src2, src1.shape[1::-1])
print(src2.shape)
print(src2.dtype)
dst = cv2.bitwise_and(src1, src2)
cv2.imwrite('data/dst/opencv_bitwise_and.jpg', dst)
import cv2
import os
import datetime
def save_frame_camera_cycle(device_num, dir_path, basename, cycle, ext='jpg', delay=1, window_name='frame'):
cap = cv2.VideoCapture(device_num)
if not cap.isOpened():
return
os.makedirs(dir_path, exist_ok=True)
base_path = os.path.join(dir_path, basename)
n = 0
while True:
ret, frame = cap.read()
cv2.imshow(window_name, frame)
if cv2.waitKey(delay) & 0xFF == ord('q'):
break
if n == cycle:
n = 0
cv2.imwrite('{}_{}.{}'.format(base_path, datetime.datetime.now().strftime('%Y%m%d%H%M%S%f'), ext), frame)
n += 1
cv2.destroyWindow(window_name)
save_frame_camera_cycle(0, 'data/temp', 'camera_capture_cycle', 300)
import cv2
import os
def save_frame_camera_key(device_num, dir_path, basename, ext='jpg', delay=1, window_name='frame'):
cap = cv2.VideoCapture(device_num)
if not cap.isOpened():
return
os.makedirs(dir_path, exist_ok=True)
base_path = os.path.join(dir_path, basename)
n = 0
while True:
ret, frame = cap.read()
cv2.imshow(window_name, frame)
key = cv2.waitKey(delay) & 0xFF
if key == ord('c'):
cv2.imwrite('{}_{}.{}'.format(base_path, n, ext), frame)
n += 1
elif key == ord('q'):
break
cv2.destroyWindow(window_name)
save_frame_camera_key(0, 'data/temp', 'camera_capture')
import cv2
import numpy as np
from PIL import Image
im_cv = cv2.imread('data/src/lena.jpg')
cv2.imwrite('data/dst/lena_bgr_cv.jpg', im_cv)
pil_img = Image.fromarray(im_cv)
pil_img.save('data/dst/lena_bgr_pillow.jpg')
im_rgb = cv2.cvtColor(im_cv, cv2.COLOR_BGR2RGB)
Image.fromarray(im_rgb).save('data/dst/lena_rgb_pillow.jpg')
cv2.imwrite('data/dst/lena_rgb_cv.jpg', im_rgb)
im_pillow = np.array(Image.open('data/src/lena.jpg'))
im_bgr = cv2.cvtColor(im_pillow, cv2.COLOR_RGB2BGR)
cv2.imwrite('data/dst/lena_bgr_cv_2.jpg', im_bgr)
im_bgr = cv2.imread('data/src/lena.jpg')
im_rgb = im_bgr
Image.fromarray(im_rgb).save('data/dst/lena_swap.jpg')
im_rgb = im_bgr
-1
Image.fromarray(im_rgb).save('data/dst/lena_swap_2.jpg')
import cv2
import numpy as np
import pprint
w = h = 360
n = 6
np.random.seed(0)
pts = np.random.randint(0, w, (n, 2))
print(pts)
print(type(pts))
numpy.ndarray
print(pts.shape)
img = np.zeros((w, h, 3), np.uint8)
for p in pts:
cv2.drawMarker(img, tuple(p), (255, 255, 255), thickness=2)
cv2.imwrite('data/dst/opencv_random_pts.png', img)
rect = (0, 0, w, h)
subdiv = cv2.Subdiv2D(rect)
for p in pts:
subdiv.insert((p[0], p[1]))
triangles = subdiv.getTriangleList()
print(triangles)
pols = triangles.reshape(-1, 3, 2)
print(pols)
img_draw = img.copy()
cv2.polylines(img_draw, pols.astype(int), True, (0, 0, 255), thickness=2)
cv2.imwrite('data/dst/opencv_delaunay.png', img_draw)
np.all
w, axis=1
pols_inner = pols
np.all
w, axis=1
np.all
axis=1
np.all
h, axis=1
np.all
axis=1
print(pols_inner)
img_draw = img.copy()
cv2.polylines(img_draw, pols_inner.astype(int), True, (0, 0, 255), thickness=2)
cv2.imwrite('data/dst/opencv_delaunay_inner.png', img_draw)
facets, centers = subdiv.getVoronoiFacetList([])
pprint.pprint(facets)
-503.5626
-461.44025
-1621.6836
-125.45706
dtype=float32
dtype=float32
-125.45706
dtype=float32
-183.30182
dtype=float32
-183.30182
-1793.752
-503.5626
-461.44025
dtype=float32
dtype=float32
print(centers)
img_draw = img.copy()
cv2.polylines(img_draw, [f.astype(int) for f in facets], True, (255, 255, 255), thickness=2)
cv2.imwrite('data/dst/opencv_voronoi.png', img_draw)
img_draw = img.copy()
step = int(255 / len(facets))
for i, p in enumerate(f.astype(int) for f in facets):
cv2.fillPoly(img_draw, [p], (step * i, step * i, step * i))
cv2.imwrite('data/dst/opencv_voronoi_fill.png', img_draw)
img_draw = img.copy()
step = int(255 / len(facets))
for i, p in enumerate(f.astype(int) for f in facets):
cv2.fillPoly(img_draw, [p], (step * i, step * i, step * i))
cv2.polylines(img_draw, pols_inner.astype(int), True, (0, 0, 255), thickness=2)
for c in centers:
cv2.drawMarker(img_draw, tuple(c), (255, 255, 255), thickness=2)
cv2.imwrite('data/dst/opencv_delaunay_voronoi.png', img_draw)
import cv2
import numpy as np
src = cv2.imread('data/src/lena.jpg')
mask = np.zeros_like(src)
print(mask.shape)
print(mask.dtype)
cv2.rectangle(mask, (50, 50), (100, 200), (255, 255, 255), thickness=-1)
cv2.circle(mask, (200, 100), 50, (255, 255, 255), thickness=-1)
cv2.fillConvexPoly(mask, np.array([[330, 50], [300, 200], [360, 150]]), (255, 255, 255))
cv2.imwrite('data/dst/opencv_draw_mask.jpg', mask)
mask_blur = cv2.GaussianBlur(mask, (51, 51), 0)
cv2.imwrite('data/dst/opencv_draw_mask_blur.jpg', mask_blur)
dst = src * (mask_blur / 255)
cv2.imwrite('data/dst/opencv_draw_mask_blur_result.jpg', dst)
import cv2
import numpy as np
print(cv2.__version__)
img = np.full((210, 425, 3), 128, dtype=np.uint8)
cv2.rectangle(img, (50, 10), (125, 60), (255, 0, 0))
cv2.rectangle(img, (50, 80), (125, 130), (0, 255, 0), thickness=-1)
cv2.rectangle(img, (50, 150), (125, 200), (0, 0, 255), thickness=-1)
cv2.rectangle(img, (50, 150), (125, 200), (255, 255, 0))
cv2.rectangle(img, (175, 10), (250, 60), (255, 255, 255), thickness=8, lineType=cv2.LINE_4)
cv2.line(img, (175, 10), (250, 60), (0, 0, 0), thickness=1, lineType=cv2.LINE_4)
cv2.rectangle(img, (175, 80), (250, 130), (255, 255, 255), thickness=8, lineType=cv2.LINE_8)
cv2.line(img, (175, 80), (250, 130), (0, 0, 0), thickness=1, lineType=cv2.LINE_8)
cv2.rectangle(img, (175, 150), (250, 200), (255, 255, 255), thickness=8, lineType=cv2.LINE_AA)
cv2.line(img, (175, 150), (250, 200), (0, 0, 0), thickness=1, lineType=cv2.LINE_AA)
cv2.rectangle(img, (600, 20), (750, 120), (0, 0, 0), lineType=cv2.LINE_AA, shift=1)
cv2.rectangle(img, (601, 160), (751, 260), (0, 0, 0), lineType=cv2.LINE_AA, shift=1)
cv2.rectangle(img, (602, 300), (752, 400), (0, 0, 0), lineType=cv2.LINE_AA, shift=1)
cv2.imwrite('data/dst/opencv_draw_argument.png', img)
img_rect = cv2.rectangle(img, (10, 10), (110, 60), (255, 0, 0))
print(img is img_rect)
img = np.full((210, 425, 3), 128, dtype=np.uint8)
cv2.line(img, (50, 10), (125, 60), (255, 0, 0))
cv2.line(img, (50, 60), (125, 10), (0, 255, 255), thickness=4, lineType=cv2.LINE_AA)
cv2.arrowedLine(img, (50, 80), (125, 130), (0, 255, 0), thickness=4)
cv2.arrowedLine(img, (50, 130), (125, 80), (255, 0, 255), tipLength=0.3)
cv2.rectangle(img, (50, 150), (125, 200), (255, 255, 0))
cv2.circle(img, (190, 35), 15, (255, 255, 255), thickness=-1)
cv2.circle(img, (240, 35), 20, (0, 0, 0), thickness=3, lineType=cv2.LINE_AA)
cv2.ellipse(img, ((190, 105), (20, 50), 0), (255, 255, 255))
cv2.ellipse(img, ((240, 105), (20, 50), 30), (0, 0, 0), thickness=-1)
cv2.ellipse(img, (190, 175), (10, 25), 0, 0, 270, (255, 255, 255))
cv2.ellipse(img, (240, 175), (10, 25), 30, 0, 270, (0, 0, 0), thickness=-1)
cv2.drawMarker(img, (300, 20), (255, 0, 0))
cv2.drawMarker(img, (337, 20), (0, 255, 0), markerType=cv2.MARKER_TILTED_CROSS, markerSize=15)
cv2.drawMarker(img, (375, 20), (0, 0, 255), markerType=cv2.MARKER_STAR, markerSize=10)
cv2.drawMarker(img, (300, 50), (0, 255, 255), markerType=cv2.MARKER_DIAMOND)
cv2.drawMarker(img, (337, 50), (255, 0, 255), markerType=cv2.MARKER_SQUARE, markerSize=15)
cv2.drawMarker(img, (375, 50), (255, 255, 0), markerType=cv2.MARKER_TRIANGLE_UP, markerSize=10)
pts = np.array(((300, 80), (300, 130), (335, 130)))
cv2.polylines(img, [pts], True, (255, 255, 255), thickness=2)
pts = np.array(((335, 80), (375, 80), (375, 130)))
cv2.fillPoly(img, [pts], (0, 0, 0))
cv2.putText(img, 'nkmk', (300, 170), cv2.FONT_HERSHEY_SIMPLEX, 1.0, (255, 255, 255), thickness=2)
cv2.putText(img, 'nkmk', (300, 195), cv2.FONT_HERSHEY_COMPLEX, 0.8, (0, 0, 0), lineType=cv2.LINE_AA)
cv2.imwrite('data/dst/opencv_draw_etc.png', img)
import cv2
img = cv2.imread('data/src/lena_square.png')
img_gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
print(type(img), img.shape)
print(type(img_gray), img_gray.shape)
numpy.ndarray
numpy.ndarray
cascade_path = '/usr/local/Cellar/opencv3/3.2.0/share/OpenCV/haarcascades/haarcascade_frontalface_alt.xml'
cascade = cv2.CascadeClassifier(cascade_path)
faces = cascade.detectMultiScale(img_gray)
print(faces)
for x, y, w, h in faces:
cv2.rectangle(img, (x, y), (x + w, y + h), (255, 0, 0), 2)
cv2.imwrite('data/dst/lena_square_face.jpg', img)
import cv2
face_cascade_path = '/usr/local/opt/opencv/share/'
eye_cascade_path = '/usr/local/opt/opencv/share/'
face_cascade = cv2.CascadeClassifier(face_cascade_path)
eye_cascade = cv2.CascadeClassifier(eye_cascade_path)
src = cv2.imread('data/src/lena_square.png')
src_gray = cv2.cvtColor(src, cv2.COLOR_BGR2GRAY)
faces = face_cascade.detectMultiScale(src_gray)
for x, y, w, h in faces:
cv2.rectangle(src, (x, y), (x + w, y + h), (255, 0, 0), 2)
face = src[y: y + h, x: x + w]
face_gray = src_gray[y: y + h, x: x + w]
eyes = eye_cascade.detectMultiScale(face_gray)
for (ex, ey, ew, eh) in eyes:
cv2.rectangle(face, (ex, ey), (ex + ew, ey + eh), (0, 255, 0), 2)
cv2.imwrite('data/dst/opencv_face_detect_rectangle.jpg', src)
src = cv2.imread('data/src/lena_square.png')
src_gray = cv2.cvtColor(src, cv2.COLOR_BGR2GRAY)
faces = face_cascade.detectMultiScale(src_gray)
for x, y, w, h in faces:
src[y: y + h, x: x + w] = [0, 128, 255]
cv2.imwrite('data/dst/opencv_face_detect_fill.jpg', src)
src = cv2.imread('data/src/lena_square.png')
src_gray = cv2.cvtColor(src, cv2.COLOR_BGR2GRAY)
faces = face_cascade.detectMultiScale(src_gray)
ratio = 0.05
for x, y, w, h in faces:
small = cv2.resize(src[y: y + h, x: x + w], None, fx=ratio, fy=ratio, interpolation=cv2.INTER_NEAREST)
src[y: y + h, x: x + w] = cv2.resize(small, (w, h), interpolation=cv2.INTER_NEAREST)
cv2.imwrite('data/dst/opencv_face_detect_mosaic.jpg', src)
import cv2
face_cascade_path = '/usr/local/opt/opencv/share/'
eye_cascade_path = '/usr/local/opt/opencv/share/'
face_cascade = cv2.CascadeClassifier(face_cascade_path)
eye_cascade = cv2.CascadeClassifier(eye_cascade_path)
cap = cv2.VideoCapture(0)
while True:
ret, img = cap.read()
gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
faces = face_cascade.detectMultiScale(gray, scaleFactor=1.3, minNeighbors=5)
for x, y, w, h in faces:
cv2.rectangle(img, (x, y), (x + w, y + h), (255, 0, 0), 2)
face = img[y: y + h, x: x + w]
face_gray = gray[y: y + h, x: x + w]
eyes = eye_cascade.detectMultiScale(face_gray)
for (ex, ey, ew, eh) in eyes:
cv2.rectangle(face, (ex, ey), (ex + ew, ey + eh), (0, 255, 0), 2)
cv2.imshow('video image', img)
key = cv2.waitKey(10)
if key == 27:  
break
cap.release()
cv2.destroyAllWindows()
import cv2
img = cv2.imread('data/src/lena.jpg')
print(type(img))
numpy.ndarray
print(img.shape)
img_flip_ud = cv2.flip(img, 0)
cv2.imwrite('data/dst/lena_cv_flip_ud.jpg', img_flip_ud)
img_flip_lr = cv2.flip(img, 1)
cv2.imwrite('data/dst/lena_cv_flip_lr.jpg', img_flip_lr)
img_flip_ud_lr = cv2.flip(img, -1)
cv2.imwrite('data/dst/lena_cv_flip_ud_lr.jpg', img_flip_ud_lr)
import cv2
print(cv2.getBuildInformation())
control:               unknown
Location (extra)
control (extra)
Host:                        Darwin
generator:             Unix
Configuration:               Release
Baseline:                    SSE
requested:                 DETECT
disabled:                  SSE4_1
generation:  SSE4_1
requested:                 SSE4_1
+ FP16
+ AVX
+ FP16
+ FP16
flags (Release)
flags (Debug)
flags (Release)
flags (Debug)
ccache:                      NO
headers:         NO
built:                 aruco
Disabled:                    text
Unavailable:                 cnn_3dobj
Applications:                apps
Documentation:               NO
algorithms:         YES
Cocoa:                       YES
JPEG:                        build
WEBP:                        build
HDR:                         YES
SUNRASTER:                   YES
PXM:                         YES
PFM:                         YES
FFMPEG:                      YES
avcodec:                   YES
avformat:                  YES
avutil:                    YES
swscale:                   YES
avresample:                YES
AVFoundation:                YES
framework:            TBB
Trace:                         YES
IW:                sources
Lapack:                      YES
Eigen:                       YES
HAL:                  NO
Protobuf:                    build
OpenCL:                        YES
path:                NO
-framework
path:                lib
path:                lib
ant:                         NO
wrappers:               NO
tests:                  NO
print(type(cv2.getBuildInformation()))
import cv2
import numpy as np
im1 = cv2.imread('data/src/lena.jpg')
im2 = cv2.imread('data/src/rocket.jpg')
im_v = cv2.vconcat([im1, im1])
cv2.imwrite('data/dst/opencv_vconcat.jpg', im_v)
im_v_np = np.tile(im1, (2, 1, 1))
cv2.imwrite('data/dst/opencv_vconcat_np.jpg', im_v_np)
def vconcat_resize_min(im_list, interpolation=cv2.INTER_CUBIC):
w_min = min(im.shape[1] for im in im_list)
cv2.resize(im, (w_min, int(im.shape[0] * w_min / im.shape[1])), interpolation=interpolation)
return cv2.vconcat(im_list_resize)
im_v_resize = vconcat_resize_min([im1, im2, im1])
cv2.imwrite('data/dst/opencv_vconcat_resize.jpg', im_v_resize)
im_h = cv2.hconcat([im1, im1])
cv2.imwrite('data/dst/opencv_hconcat.jpg', im_h)
im_h_np = np.tile(im1, (1, 2, 1))
cv2.imwrite('data/dst/opencv_hconcat_np.jpg', im_h_np)
def hconcat_resize_min(im_list, interpolation=cv2.INTER_CUBIC):
h_min = min(im.shape[0] for im in im_list)
cv2.resize(im, (int(im.shape[1] * h_min / im.shape[0]), h_min), interpolation=interpolation)
return cv2.hconcat(im_list_resize)
im_h_resize = hconcat_resize_min([im1, im2, im1])
cv2.imwrite('data/dst/opencv_hconcat_resize.jpg', im_h_resize)
def concat_tile(im_list_2d):
return cv2.vconcat([cv2.hconcat(im_list_h) for im_list_h in im_list_2d])
im1_s = cv2.resize(im1, dsize=(0, 0), fx=0.5, fy=0.5)
im_tile = concat_tile
cv2.imwrite('data/dst/opencv_concat_tile.jpg', im_tile)
im_tile_np = np.tile(im1_s, (3, 4, 1))
cv2.imwrite('data/dst/opencv_concat_tile_np.jpg', im_tile_np)
def concat_tile_resize(im_list_2d, interpolation=cv2.INTER_CUBIC):
im_list_v = [hconcat_resize_min(im_list_h, interpolation=cv2.INTER_CUBIC) for im_list_h in im_list_2d]
return vconcat_resize_min(im_list_v, interpolation=cv2.INTER_CUBIC)
im_tile_resize = concat_tile_resize
cv2.imwrite('data/dst/opencv_concat_tile_resize.jpg', im_tile_resize)
import cv2
src = cv2.imread('data/src/lena.jpg')
def mosaic(src, ratio=0.1):
small = cv2.resize(src, None, fx=ratio, fy=ratio, interpolation=cv2.INTER_NEAREST)
return cv2.resize(small, src.shape[:2][::-1], interpolation=cv2.INTER_NEAREST)
dst_01 = mosaic(src)
cv2.imwrite('data/dst/opencv_mosaic_01.jpg', dst_01)
dst_005 = mosaic(src, ratio=0.05)
cv2.imwrite('data/dst/opencv_mosaic_005.jpg', dst_005)
def mosaic_area(src, x, y, width, height, ratio=0.1):
dst = src.copy()
dst[y:y + height, x:x + width] = mosaic(dst[y:y + height, x:x + width], ratio)
return dst
dst_area = mosaic_area(src, 100, 50, 100, 150)
cv2.imwrite('data/dst/opencv_mosaic_area.jpg', dst_area)
face_cascade_path = '/usr/local/opt/opencv/share/'
face_cascade = cv2.CascadeClassifier(face_cascade_path)
src_gray = cv2.cvtColor(src, cv2.COLOR_BGR2GRAY)
faces = face_cascade.detectMultiScale(src_gray)
for x, y, w, h in faces:
dst_face = mosaic_area(src, x, y, w, h)
cv2.imwrite('data/dst/opencv_mosaic_face.jpg', dst_face)
import cv2
from PIL import Image
def mosaic(src, ratio=0.1):
small = cv2.resize(src, None, fx=ratio, fy=ratio, interpolation=cv2.INTER_NEAREST)
return cv2.resize(small, src.shape[:2][::-1], interpolation=cv2.INTER_NEAREST)
src = cv2.cvtColor(cv2.imread('data/src/lena.jpg'), cv2.COLOR_BGR2RGB)
imgs = [Image.fromarray(mosaic(src, 1 / i)) for i in range(1, 25)]
imgs += imgs[-2::-1] + [Image.fromarray(src)] * 5
imgs[0].save
save_all=True
append_images=imgs[1:]
optimize=False
duration=50
loop=0
import cv2
import numpy as np
im = cv2.imread('data/src/lena.jpg')
print(im.shape)
print(im.dtype)
im_gray = cv2.cvtColor(im, cv2.COLOR_BGR2GRAY)
print(im_gray.shape)
print(im_gray.dtype)
cv2.imwrite('data/dst/opencv_gray_cvtcolr.jpg', im_gray)
im_gray_read = cv2.imread('data/src/lena.jpg', cv2.IMREAD_GRAYSCALE)
print(im_gray_read.shape)
print(im_gray_read.dtype)
cv2.imwrite('data/dst/opencv_gray_imread.jpg', im_gray_read)
im_gray_diff = im_gray.astype(int) - im_gray_read.astype(int)
print(im_gray_diff.max())
print(im_gray_diff.min())
-10
im_gray_calc = 0.299 * im
print(im_gray_calc.shape)
print(im_gray_calc.dtype)
cv2.imwrite('data/dst/numpy_gray_calc.jpg', im_gray_calc)
im_gray_diff = im_gray_calc - im_gray
print(im_gray_diff.max())
print(im_gray_diff.min())
-0.4980000000000473
a = np.array([0.0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0])
print(a.round())
print(a.astype('uint8'))
print(np.array_equal(im_gray_calc.round(), im_gray))
print(np.array_equal(im_gray_calc.astype('uint8'), im_gray))
import cv2
img_org = cv2.imread('data/src/lena.jpg')
img_q95 = cv2.imread('data/src/lena_q95.jpg')
img_q50 = cv2.imread('data/src/lena_q50.jpg')
print(cv2.PSNR(img_org, img_q95))
print(cv2.PSNR(img_org, img_q50))
print(cv2.PSNR(img_org, img_org))
print(cv2.PSNR(img_org, img_q95, R=255))
import cv2
im = cv2.imread('data/src/lena.jpg')
print(type(im))
numpy.ndarray
print(im.shape)
print(im.dtype)
cv2.imwrite('data/dst/lena_opencv_red.jpg', im)
cv2.imwrite('data/dst/lena_opencv_red_high.jpg', im, [cv2.IMWRITE_JPEG_QUALITY, 100])
cv2.imwrite('data/dst/lena_opencv_red_low.jpg', im, [cv2.IMWRITE_JPEG_QUALITY, 50])
im_gray = cv2.imread('data/src/lena.jpg', cv2.IMREAD_GRAYSCALE)
im_gray = cv2.imread('data/src/lena.jpg', 0)
print(type(im_gray))
numpy.ndarray
print(im_gray.shape)
print(im_gray.dtype)
cv2.imwrite('data/dst/lena_opencv_gray.jpg', im_gray)
im_gray_read = cv2.imread('data/dst/lena_opencv_gray.jpg')
print(im_gray_read.shape)
import numpy as np
np.array_equal
np.array_equal
im = cv2.imread('xxxxxxx')
print(im)
print(im.shape)
im = cv2.imread('data/src/sample.csv')
print(im)
im = cv2.imread('xxxxxxx')
if im:
print('Image is read.')
print('Image is not read.')
im = cv2.imread('xxxxxxx')
if not im:
print('Image is not read.')
print('Image is read.')
import cv2
img = cv2.imread('data/src/lena.jpg')
print(type(img))
numpy.ndarray
print(img.shape)
img_rotate_90_clockwise = cv2.rotate(img, cv2.ROTATE_90_CLOCKWISE)
cv2.imwrite('data/dst/lena_cv_rotate_90_clockwise.jpg', img_rotate_90_clockwise)
img_rotate_90_counterclockwise = cv2.rotate(img, cv2.ROTATE_90_COUNTERCLOCKWISE)
cv2.imwrite('data/dst/lena_cv_rotate_90_counterclockwise.jpg', img_rotate_90_counterclockwise)
img_rotate_180 = cv2.rotate(img, cv2.ROTATE_180)
cv2.imwrite('data/dst/lena_cv_rotate_180.jpg', img_rotate_180)
import cv2
im = cv2.imread('data/src/lena.jpg')
print(type(im))
numpy.ndarray
print(im.shape)
print(type(im.shape))
h, w, c = im.shape
print('width:  ', w)
print('height: ', h)
print('channel:', c)
h, w, _ = im.shape
print('width: ', w)
print('height:', h)
print('width: ', im.shape[1])
print('height:', im.shape[0])
print(im.shape[1::-1])
im_gray = cv2.imread('data/src/lena.jpg', cv2.IMREAD_GRAYSCALE)
print(im_gray.shape)
print(type(im_gray.shape))
h, w = im_gray.shape
print('width: ', w)
print('height:', h)
print('width: ', im_gray.shape[1])
print('height:', im_gray.shape[0])
h, w = im.shape[0], im.shape[1]
print('width: ', w)
print('height:', h)
print(im_gray.shape[::-1])
print(im_gray.shape[1::-1])
import cv2
im = cv2.imread('data/src/lena_square_half.png')
th, im_th = cv2.threshold(im, 128, 255, cv2.THRESH_BINARY)
print(th)
cv2.imwrite('data/dst/opencv_th.jpg', im_th)
th, im_th_tz = cv2.threshold(im, 128, 255, cv2.THRESH_TOZERO)
print(th)
cv2.imwrite('data/dst/opencv_th_tz.jpg', im_th_tz)
th, im_th_otsu = cv2.threshold(im, 128, 192, cv2.THRESH_OTSU)
error: OpenCV
-215
im_gray = cv2.cvtColor(im, cv2.COLOR_BGR2GRAY)
th, im_gray_th_otsu = cv2.threshold(im_gray, 128, 192, cv2.THRESH_OTSU)
print(th)
cv2.imwrite('data/dst/opencv_th_otsu.jpg', im_gray_th_otsu)
import cv2
import os
def save_all_frames(video_path, dir_path, basename, ext='jpg'):
cap = cv2.VideoCapture(video_path)
if not cap.isOpened():
return
os.makedirs(dir_path, exist_ok=True)
base_path = os.path.join(dir_path, basename)
digit = len(str(int(cap.get(cv2.CAP_PROP_FRAME_COUNT))))
n = 0
while True:
ret, frame = cap.read()
if ret:
cv2.imwrite('{}_{}.{}'.format(base_path, str(n).zfill(digit), ext), frame)
n += 1
return
save_all_frames('data/temp/sample_video.mp4', 'data/temp/result', 'sample_video_img')
save_all_frames('data/temp/sample_video.mp4', 'data/temp/result_png', 'sample_video_img', 'png')
def save_frame(video_path, frame_num, result_path):
cap = cv2.VideoCapture(video_path)
if not cap.isOpened():
return
os.makedirs(os.path.dirname(result_path), exist_ok=True)
cap.set(cv2.CAP_PROP_POS_FRAMES, frame_num)
ret, frame = cap.read()
if ret:
cv2.imwrite(result_path, frame)
save_frame('data/temp/sample_video.mp4', 100, 'data/temp/result_single/sample_100.jpg')
dir_path, basename, ext='jpg'
cap = cv2.VideoCapture(video_path)
if not cap.isOpened():
return
os.makedirs(dir_path, exist_ok=True)
base_path = os.path.join(dir_path, basename)
digit = len(str(int(cap.get(cv2.CAP_PROP_FRAME_COUNT))))
for n in range(start_frame, stop_frame, step_frame):
cap.set(cv2.CAP_PROP_POS_FRAMES, n)
ret, frame = cap.read()
if ret:
cv2.imwrite('{}_{}.{}'.format(base_path, str(n).zfill(digit), ext), frame)
return
def save_frame_sec(video_path, sec, result_path):
cap = cv2.VideoCapture(video_path)
if not cap.isOpened():
return
os.makedirs(os.path.dirname(result_path), exist_ok=True)
fps = cap.get(cv2.CAP_PROP_FPS)
cap.set(cv2.CAP_PROP_POS_FRAMES, round(fps * sec))
ret, frame = cap.read()
if ret:
cv2.imwrite(result_path, frame)
save_frame_sec('data/temp/sample_video.mp4', 10, 'data/temp/result_10sec.jpg')
dir_path, basename, ext='jpg'
cap = cv2.VideoCapture(video_path)
if not cap.isOpened():
return
os.makedirs(dir_path, exist_ok=True)
base_path = os.path.join(dir_path, basename)
digit = len(str(int(cap.get(cv2.CAP_PROP_FRAME_COUNT))))
fps = cap.get(cv2.CAP_PROP_FPS)
fps_inv = 1 / fps
sec = start_sec
while sec < stop_sec:
n = round(fps * sec)
cap.set(cv2.CAP_PROP_POS_FRAMES, n)
ret, frame = cap.read()
if ret:
cv2.imwrite
str(n).zfill(digit)
return
sec += step_sec
import cv2
import os
def save_frame_play(video_path, dir_path, basename, ext='jpg', delay=1, window_name='frame'):
cap = cv2.VideoCapture(video_path)
if not cap.isOpened():
return
os.makedirs(dir_path, exist_ok=True)
base_path = os.path.join(dir_path, basename)
digit = len(str(int(cap.get(cv2.CAP_PROP_FRAME_COUNT))))
n = 0
while True:
ret, frame = cap.read()
if ret:
cv2.imshow(window_name, frame)
key = cv2.waitKey(delay) & 0xFF
if key == ord('c'):
cv2.imwrite('{}_{}.{}'.format(base_path, str(n).zfill(digit), ext), frame)
elif key == ord('q'):
break
n += 1
cap.set(cv2.CAP_PROP_POS_FRAMES, 0)
n = 0
cv2.destroyWindow(window_name)
save_frame_play('data/temp/sample_video.mp4', 'data/temp', 'sample_video_cap', delay=0)
import cv2
cap_cam = cv2.VideoCapture(0)
print(type(cap_cam))
cv2.VideoCapture
print(cap_cam.isOpened())
cap_cam_wrong = cv2.VideoCapture(1)
print(type(cap_cam_wrong))
cv2.VideoCapture
print(cap_cam_wrong.isOpened())
cap_cam.release()
import cv2
cap_file = cv2.VideoCapture('data/temp/sample_video.mp4')
print(type(cap_file))
cv2.VideoCapture
print(cap_file.isOpened())
cap_file_wrong = cv2.VideoCapture('wrong_path')
print(type(cap_file_wrong))
cv2.VideoCapture
print(cap_file_wrong.isOpened())
import cv2
import sys
camera_id = 0
delay
window_name = 'frame'
cap = cv2.VideoCapture(camera_id)
if not cap.isOpened():
sys.exit()
while True:
ret, frame = cap.read()
cv2.imshow(window_name, frame)
if cv2.waitKey(delay) & 0xFF == ord('q'):
break
cv2.destroyWindow(window_name)
import cv2
import sys
file_path = 'data/temp/sample_video.mp4'
delay
window_name = 'frame'
cap = cv2.VideoCapture(file_path)
if not cap.isOpened():
sys.exit()
while True:
ret, frame = cap.read()
if ret:
cv2.imshow(window_name, frame)
if cv2.waitKey(delay) & 0xFF == ord('q'):
break
cap.set(cv2.CAP_PROP_POS_FRAMES, 0)
cv2.destroyWindow(window_name)
import cv2
import sys
camera_id = 0
delay
window_name = 'frame'
cap = cv2.VideoCapture(camera_id)
if not cap.isOpened():
sys.exit()
tm = cv2.TickMeter()
tm.start()
count = 0
max_count = 10
fps = 0
while cap.isOpened():
ret, frame = cap.read()
if count == max_count:
tm.stop()
fps = max_count / tm.getTimeSec()
tm.reset()
tm.start()
count = 0
cv2.putText
format(fps)
cv2.FONT_HERSHEY_SIMPLEX
thickness=2
cv2.imshow(window_name, frame)
count += 1
if cv2.waitKey(delay) & 0xFF == ord('q'):
break
cv2.destroyWindow(window_name)
import cv2
cap = cv2.VideoCapture('data/temp/sample_video.mp4')
print(type(cap))
cv2.VideoCapture
print(cap.isOpened())
print(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
print(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
print(cap.get(cv2.CAP_PROP_FPS))
print(cap.get(cv2.CAP_PROP_FRAME_COUNT))
print(cap.get(cv2.CAP_PROP_FRAME_COUNT) / cap.get(cv2.CAP_PROP_FPS))
print(cap.set(cv2.CAP_PROP_FRAME_WIDTH, 320))
print(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
print(cap.get(cv2.CAP_PROP_POS_FRAMES))
print(cap.set(cv2.CAP_PROP_POS_FRAMES, 100))
print(cap.get(cv2.CAP_PROP_POS_FRAMES))
import cv2
cap = cv2.VideoCapture(0)
print(cap.get(cv2.CAP_PROP_FPS))
print(cap.set(cv2.CAP_PROP_FPS, 10))
print(cap.get(cv2.CAP_PROP_FPS))
print(cap.set(cv2.CAP_PROP_FPS, 120))
print(cap.get(cv2.CAP_PROP_FPS))
cap.release()
import cv2
cap_cam = cv2.VideoCapture(0)
print(type(cap_cam))
cv2.VideoCapture
print(cap_cam.isOpened())
print(cap_cam.get(cv2.CAP_PROP_POS_FRAMES))
ret, frame = cap_cam.read()
print(ret)
print(type(frame))
numpy.ndarray
print(frame.shape)
print(cap_cam.get(cv2.CAP_PROP_POS_FRAMES))
cap_cam.release()
import cv2
cap = cv2.VideoCapture('data/temp/sample_video.mp4')
print(type(cap))
cv2.VideoCapture
print(cap.isOpened())
print(cap.get(cv2.CAP_PROP_POS_FRAMES))
print(cap.get(cv2.CAP_PROP_POS_MSEC))
ret, frame = cap.read()
print(ret)
print(type(frame))
numpy.ndarray
print(frame.shape)
print(cap.get(cv2.CAP_PROP_POS_FRAMES))
print(cap.get(cv2.CAP_PROP_POS_MSEC))
print(1 / cap.get(cv2.CAP_PROP_FPS) * 1000)
cap.set(cv2.CAP_PROP_POS_FRAMES, 100)
print(cap.get(cv2.CAP_PROP_POS_FRAMES))
ret, frame = cap.read()
print(ret)
print(cap.get(cv2.CAP_PROP_POS_FRAMES))
print(cap.get(cv2.CAP_PROP_POS_MSEC))
cap.set(cv2.CAP_PROP_POS_FRAMES, cap.get(cv2.CAP_PROP_FRAME_COUNT))
print(cap.get(cv2.CAP_PROP_POS_FRAMES))
ret, frame = cap.read()
print(ret)
print(frame)
print(cap.get(cv2.CAP_PROP_POS_FRAMES))
cap.set(cv2.CAP_PROP_POS_FRAMES, 1000)
print(cap.get(cv2.CAP_PROP_POS_FRAMES))
import cv2
import sys
camera_id = 0
delay
window_name = 'frame'
cap = cv2.VideoCapture(camera_id)
if not cap.isOpened():
sys.exit()
while True:
ret, frame = cap.read()
gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
blur = cv2.GaussianBlur(gray, (0, 0), 5)
cv2.imshow(window_name, blur)
if cv2.waitKey(delay) & 0xFF == ord('q'):
break
cv2.destroyWindow(window_name)
import cv2
import numpy as np
src = cv2.imread('data/src/lena.jpg')
dst = cv2.imread('data/src/rocket.jpg')
src_pts = [[100, 80], [150, 200], [300, 20]]
dst_pts = [[280, 120], [320, 300], [400, 150]]
src_mark = src.copy()
for pt in src_pts:
cv2.drawMarker(src_mark, tuple(pt), (0, 255, 0), thickness=4)
cv2.imwrite('data/dst/opencv_warp_src_mark.jpg', src_mark)
dst_mark = dst.copy()
for pt in dst_pts:
cv2.drawMarker(dst_mark, tuple(pt), (0, 255, 0), thickness=4)
cv2.imwrite('data/dst/opencv_warp_dst_mark.jpg', dst_mark)
src_pts_arr = np.array(src_pts, dtype=np.float32)
dst_pts_arr = np.array(dst_pts, dtype=np.float32)
src_rect = cv2.boundingRect(src_pts_arr)
dst_rect = cv2.boundingRect(dst_pts_arr)
print(src_rect)
print(dst_rect)
src_crop = src[src_rect[1]:src_rect[1] + src_rect[3], src_rect[0]:src_rect[0] + src_rect[2]]
dst_crop = dst[dst_rect[1]:dst_rect[1] + dst_rect[3], dst_rect[0]:dst_rect[0] + dst_rect[2]]
src_pts_crop = src_pts_arr - src_rect[:2]
dst_pts_crop = dst_pts_arr - dst_rect[:2]
print(src_pts_crop)
print(dst_pts_crop)
src_crop_mark = src_crop.copy()
for pt in src_pts_crop.astype(np.int):
cv2.drawMarker(src_crop_mark, tuple(pt), (0, 255, 0), thickness=4)
cv2.imwrite('data/dst/opencv_warp_src_crop_mark.jpg', src_crop_mark)
dst_crop_mark = dst_crop.copy()
for pt in dst_pts_crop.astype(np.int):
cv2.drawMarker(dst_crop_mark, tuple(pt), (0, 255, 0), thickness=4)
cv2.imwrite('data/dst/opencv_warp_dst_crop_mark.jpg', dst_crop_mark)
mat = cv2.getAffineTransform(src_pts_crop.astype(np.float32), dst_pts_crop.astype(np.float32))
affine_img = cv2.warpAffine(src_crop, mat, tuple(dst_rect[2:]))
cv2.imwrite('data/dst/opencv_warp_affine_crop.jpg', affine_img)
mask = np.zeros_like(dst_crop, dtype=np.float32)
cv2.fillConvexPoly(mask, dst_pts_crop.astype(np.int), (1.0, 1.0, 1.0), cv2.LINE_AA)
dst_crop_merge = affine_img * mask + dst_crop * (1 - mask)
cv2.imwrite('data/dst/opencv_warp_affine_crop_merge.jpg', dst_crop_merge)
dst[dst_rect[1]:dst_rect[1] + dst_rect[3], dst_rect[0]:dst_rect[0] + dst_rect[2]] = dst_crop_merge
cv2.imwrite('data/dst/opencv_warp_dst_result.jpg', dst)
import cv2
import numpy as np
import math
img = cv2.imread('data/src/lena.jpg')
h, w, c = img.shape
print(h, w, c)
mat = cv2.getRotationMatrix2D((w / 2, h / 2), 45, 0.5)
print(mat)
-0.35355339
affine_img = cv2.warpAffine(img, mat, (w, h))
cv2.imwrite('data/dst/opencv_affine.jpg', affine_img)
affine_img_half = cv2.warpAffine(img, mat, (w, h // 2))
cv2.imwrite('data/dst/opencv_affine_half.jpg', affine_img_half)
affine_img_flags = cv2.warpAffine(img, mat, (w, h), flags=cv2.INTER_CUBIC)
cv2.imwrite('data/dst/opencv_affine_flags.jpg', affine_img_flags)
affine_img_bv = cv2.warpAffine(img, mat, (w, h), borderValue=(0, 128, 255))
cv2.imwrite('data/dst/opencv_affine_border_value.jpg', affine_img_bv)
dst = img // 4
affine_img_bm_bt = cv2.warpAffine(img, mat, (w, h), borderMode=cv2.BORDER_TRANSPARENT, dst=dst)
cv2.imwrite('data/dst/opencv_affine_border_transparent.jpg', affine_img_bm_bt)
affine_img_bm_br = cv2.warpAffine(img, mat, (w, h), borderMode=cv2.BORDER_REPLICATE)
cv2.imwrite('data/dst/opencv_affine_border_replicate.jpg', affine_img_bm_br)
affine_img_bm_bw = cv2.warpAffine(img, mat, (w, h), borderMode=cv2.BORDER_WRAP)
cv2.imwrite('data/dst/opencv_affine_border_wrap.jpg', affine_img_bm_bw)
mat = np.array([[1, 0, 50], [0, 1, 20]], dtype=np.float32)
print(mat)
affine_img_translation = cv2.warpAffine(img, mat, (w, h))
cv2.imwrite('data/dst/opencv_affine_translation.jpg', affine_img_translation)
a = math.tan(math.radians(15))
mat = np.array([[1, a, 0], [0, 1, 0]], dtype=np.float32)
print(mat)
affine_img_skew_x = cv2.warpAffine(img, mat, (int(w + h * a), h))
cv2.imwrite('data/dst/opencv_affine_skew_x.jpg', affine_img_skew_x)
mat = np.array([[1, 0, 0], [a, 1, 0]], dtype=np.float32)
print(mat)
affine_img_skew_y = cv2.warpAffine(img, mat, (w, int(h + w * a)))
cv2.imwrite('data/dst/opencv_affine_skew_y.jpg', affine_img_skew_y)
import cv2
import numpy as np
img = cv2.imread('data/src/lena.jpg')
h, w, c = img.shape
src_pts = np.array([[30, 30], [50, 200], [350, 50]], dtype=np.float32)
dst_pts = np.array([[90, 20], [140, 170], [280, 80]], dtype=np.float32)
mat = cv2.getAffineTransform(src_pts, dst_pts)
print(mat)
img_mark = img.copy()
for pt in src_pts:
cv2.drawMarker(img_mark, tuple(pt), (0, 255, 0), thickness=4)
cv2.imwrite('data/dst/opencv_affine_mark_src.jpg', img_mark)
affine_img = cv2.warpAffine(img_mark, mat, (w, h))
cv2.imwrite('data/dst/opencv_affine_mark_dst.jpg', affine_img)
affine_img_mark = affine_img.copy()
for pt in dst_pts:
cv2.drawMarker(affine_img_mark, tuple(pt), (255, 0, 0), markerType=cv2.MARKER_SQUARE, thickness=4)
cv2.imwrite('data/dst/opencv_affine_mark_dst_mark.jpg', affine_img_mark)
src_pts = np.array([[30, 30], [50, 200], [350, 50]], dtype=np.float32)
dst_pts = np.array([[140, 170], [280, 80], [90, 20]], dtype=np.float32)
mat = cv2.getAffineTransform(src_pts, dst_pts)
print(mat)
-0.20925926
affine_img = cv2.warpAffine(img_mark, mat, (w, h))
cv2.imwrite('data/dst/opencv_affine_mark_dst_another.jpg', affine_img)
import cv2
import numpy as np
def warp(src, dst, src_pts, dst_pts, transform_func, warp_func, **kwargs):
src_pts_arr = np.array(src_pts, dtype=np.float32)
dst_pts_arr = np.array(dst_pts, dtype=np.float32)
src_rect = cv2.boundingRect(src_pts_arr)
dst_rect = cv2.boundingRect(dst_pts_arr)
src_crop = src[src_rect[1]:src_rect[1] + src_rect[3], src_rect[0]:src_rect[0] + src_rect[2]]
dst_crop = dst[dst_rect[1]:dst_rect[1] + dst_rect[3], dst_rect[0]:dst_rect[0] + dst_rect[2]]
src_pts_crop = src_pts_arr - src_rect[:2]
dst_pts_crop = dst_pts_arr - dst_rect[:2]
mat = transform_func(src_pts_crop.astype(np.float32), dst_pts_crop.astype(np.float32))
warp_img = warp_func(src_crop, mat, tuple(dst_rect[2:]), **kwargs)
mask = np.zeros_like(dst_crop, dtype=np.float32)
cv2.fillConvexPoly(mask, dst_pts_crop.astype(np.int), (1.0, 1.0, 1.0), cv2.LINE_AA)
dst_crop_merge = warp_img * mask + dst_crop * (1 - mask)
dst[dst_rect[1]:dst_rect[1] + dst_rect[3], dst_rect[0]:dst_rect[0] + dst_rect[2]] = dst_crop_merge
def warp_triangle(src, dst, src_pts, dst_pts, **kwargs):
cv2.getAffineTransform
cv2.warpAffine
def warp_rectangle(src, dst, src_pts, dst_pts, **kwargs):
cv2.getPerspectiveTransform
cv2.warpPerspective
src = cv2.imread('data/src/lena.jpg')
dst = cv2.imread('data/src/rocket.jpg')
src_pts = [[100, 80], [150, 200], [300, 20]]
dst_pts = [[280, 120], [320, 300], [400, 150]]
warp_triangle(src, dst, src_pts, dst_pts)
cv2.imwrite('data/dst/opencv_warp_triangle.jpg', dst)
src = cv2.imread('data/src/lena.jpg')
dst = cv2.imread('data/src/rocket.jpg')
src_pts = [[100, 80], [150, 200], [350, 160], [300, 20]]
dst_pts = [[280, 120], [200, 280], [500, 300], [400, 150]]
warp_rectangle(src, dst, src_pts, dst_pts, flags=cv2.INTER_CUBIC)
cv2.imwrite('data/dst/opencv_warp_rectangle.jpg', dst)
import cv2
import numpy as np
img = cv2.imread('data/src/lena.jpg')
h, w, c = img.shape
src_pts = np.array([[0, 0], [0, h], [w, h], [w, 0]], dtype=np.float32)
dst_pts = np.array([[20, 50], [50, 175], [300, 205], [380, 20]], dtype=np.float32)
mat = cv2.getPerspectiveTransform(src_pts, dst_pts)
print(mat)
-9.38078109e-02
-9.40390545e-04
perspective_img = cv2.warpPerspective(img, mat, (w, h))
cv2.imwrite('data/dst/opencv_perspective_dst.jpg', perspective_img)
mat_i = cv2.getPerspectiveTransform(dst_pts, src_pts)
print(mat_i)
perspective_img_i = cv2.warpPerspective(perspective_img, mat_i, (w, h))
cv2.imwrite('data/dst/opencv_perspective_dst_inverse.jpg', perspective_img_i)
import openpyxl
import pprint
wb = openpyxl.load_workbook('data/src/sample.xlsx')
print(type(wb))
openpyxl.workbook.workbook.Workbook
print(wb.sheetnames)
sheet = wb['sheet1']
print(type(sheet))
openpyxl.worksheet.worksheet.Worksheet
cell = sheet['A2']
print(type(cell))
openpyxl.cell.cell.Cell
print(cell.value)
cell = sheet.cell(row=2, column=1)
print(type(cell))
openpyxl.cell.cell.Cell
print(cell.value)
pprint.pprint(sheet['A2:C4'])
g = sheet.iter_rows(min_row=2, max_row=4, min_col=1, max_col=3)
print(type(g))
pprint.pprint(list(g))
def get_value_list(t_2d):
return([[cell.value for cell in row] for row in t_2d])
l_2d = get_value_list(sheet['A2:C4'])
pprint.pprint(l_2d, width=40)
def get_list_2d(sheet, start_row, end_row, start_col, end_col):
return get_value_list
sheet.iter_rows
min_row=start_row
max_row=end_row
min_col=start_col
max_col=end_col
l_2d = get_list_2d(sheet, 2, 4, 1, 3)
pprint.pprint(l_2d, width=40)
g_all = sheet.values
print(type(g_all))
pprint.pprint(list(g_all), width=40)
sheet['C1'] = 'XXX'
sheet['E1'] = 'new'
pprint.pprint(list(sheet.values), width=40)
sheet.cell(row=2, column=5, value=14)
pprint.pprint(list(sheet.values), width=40)
def write_list_2d(sheet, l_2d, start_row, start_col):
for y, row in enumerate(l_2d):
for x, cell in enumerate(row):
sheet.cell
row=start_row + y
column=start_col + x
value=l_2d[y][x]
l_2d = [['four', 41, 42, 43], ['five', 51, 52, 53]]
write_list_2d(sheet, l_2d, 5, 1)
pprint.pprint(list(sheet.values), width=40)
sheet_new = wb.create_sheet('sheet_new')
print(wb.worksheets)
sheet_new['A1'] = 'new sheet!'
print(list(sheet_new.values))
sheet_copy = wb.copy_worksheet(wb['sheet1'])
print(wb.worksheets)
pprint.pprint(list(sheet_copy.values))
wb.remove(wb['sheet1 Copy'])
print(wb.worksheets)
wb.save('data/dst/openpyxl_sample.xlsx')
import operator
l = [{'k1': i} for i in range(10000)]
print(len(l))
print(l[:5])
print(l[-5:])
l, key=lambda x: x['k1']
std. dev. of
sorted(l, key=operator.itemgetter('k1'))
std. dev. of
l, key=lambda x: x['k1'], reverse=True
std. dev. of
sorted(l, key=operator.itemgetter('k1'), reverse=True)
std. dev. of
l, key=lambda x: x['k1']
std. dev. of
max(l, key=operator.itemgetter('k1'))
std. dev. of
l, key=lambda x: x['k1']
std. dev. of
min(l, key=operator.itemgetter('k1'))
std. dev. of
import operator
print(2 + 3)
print(operator.add(2, 3))
print(operator.sub(2, 3))  
-1
print(operator.mul(2, 3))  
print(operator.truediv(2, 3))  
print(operator.floordiv(2, 3))  
print(operator.lt(2, 3))  
print(operator.le(2, 3))  
print(operator.gt(2, 3))  
print(operator.ge(2, 3))  
print(operator.eq(2, 3))  
print(operator.ne(2, 3))  
print(bin(0b1100 & 0b1010))
print(bin(operator.and_(0b1100, 0b1010)))
print(bin(0b1100 | 0b1010))
print(bin(operator.or_(0b1100, 0b1010)))
print(bin(0b1100 ^ 0b1010))
print(bin(operator.xor(0b1100, 0b1010)))
print(bin(0b1100 and 0b1010))
print(bin(0b1100 or 0b1010))
l = [0, 10, 20, 30, 40, 50]
print(l[3])
f = operator.itemgetter(3)
print(f(l))
print(operator.itemgetter(3)(l))
print(l[-1])
print(operator.itemgetter(-1)(l))
print(l[1:4])
print(operator.itemgetter(slice(1, 4))(l))
print(l[1::2])
print(operator.itemgetter(slice(1, None, 2))(l))
print(operator.itemgetter(0, slice(1, 4), -1)(l))
print(type(operator.itemgetter(0, slice(1, 4), -1)(l)))
d = {'k1': 0, 'k2': 10, 'k3': 20}
print(d['k2'])
print(operator.itemgetter('k2')(d))
print(operator.itemgetter('k2', 'k1')(d))
import datetime
dt = datetime.date(2001, 10, 20)
print(dt)
print(type(dt))
datetime.date
print(dt.year)
f = operator.attrgetter('year')
print(f(dt))
print(operator.attrgetter('year', 'month', 'day')(dt))
s = 'xxxAyyyAzzz'
print(s.upper())
f = operator.methodcaller('upper')
print(f(s))
print(s.split('A', maxsplit=1))
print(operator.methodcaller('split', 'A', maxsplit=1)(s))
import collections
od = collections.OrderedDict()
od['k1'] = 1
od['k2'] = 2
od['k3'] = 3
print(od)
OrderedDict([('k1', 1), ('k2', 2), ('k3', 3)])
print(collections.OrderedDict(k1=1, k2=2, k3=3))
print(collections.OrderedDict([('k1', 1), ('k2', 2), ('k3', 3)]))
print(collections.OrderedDict((['k1', 1], ['k2', 2], ['k3', 3])))
OrderedDict([('k1', 1), ('k2', 2), ('k3', 3)])
OrderedDict([('k1', 1), ('k2', 2), ('k3', 3)])
OrderedDict([('k1', 1), ('k2', 2), ('k3', 3)])
print(collections.OrderedDict({'k1': 1, 'k2': 2, 'k3': 3}))
OrderedDict([('k1', 1), ('k2', 2), ('k3', 3)])
print(issubclass(collections.OrderedDict, dict))
print(od['k1'])
od['k2'] = 200
print(od)
OrderedDict([('k1', 1), ('k2', 200), ('k3', 3)])
od.update(k4=4, k5=5)
print(od)
OrderedDict([('k1', 1), ('k2', 200), ('k3', 3), ('k4', 4), ('k5', 5)])
del od['k4'
od['k5']
print(od)
OrderedDict([('k1', 1), ('k2', 200), ('k3', 3)])
od.move_to_end('k1')
print(od)
OrderedDict([('k2', 200), ('k3', 3), ('k1', 1)])
od.move_to_end('k1', False)
print(od)
OrderedDict([('k1', 1), ('k2', 200), ('k3', 3)])
l = list(od.items())
print(l)
l.insert(1, ('kx', -1))
print(l)
-1
od = collections.OrderedDict(l)
print(od)
OrderedDict([('k1', 1), ('kx', -1), ('k2', 200), ('k3', 3)])
l = list(od.items())
print(l)
-1
l[0], l[2] = l[2], l[0]
print(l)
-1
od = collections.OrderedDict(l)
print(od)
OrderedDict([('k2', 200), ('kx', -1), ('k1', 1), ('k3', 3)])
l = list(od.items())
k = list(od.keys())
print(k)
print(k.index('kx'))
l[k.index('kx')], l[k.index('k3')] = l[k.index('k3')], l[k.index('kx')]
print(l)
-1
od = collections.OrderedDict(l)
print(od)
OrderedDict([('k2', 200), ('k3', 3), ('k1', 1), ('kx', -1)])
od_sorted_key = collections.OrderedDict
od.items()
key=lambda x: x[0]
print(od_sorted_key)
OrderedDict([('k1', 1), ('k2', 200), ('k3', 3), ('kx', -1)])
od_sorted_value = collections.OrderedDict
od.items()
key=lambda x: x[1]
reverse=True
print(od_sorted_value)
OrderedDict([('k2', 200), ('k3', 3), ('k1', 1), ('kx', -1)])
import os
import subprocess
print(type(os.environ))
os._Environ
print(os.environ)
print(os.environ['LANG'])
print(os.environ['NEW_KEY'])
print(os.environ.get('LANG'))
print(os.environ.get('NEW_KEY'))
print(os.environ.get('NEW_KEY', 'default'))
print(os.getenv('LANG'))
print(os.getenv('NEW_KEY'))
print(os.getenv('NEW_KEY', 'default'))
os.environ['NEW_KEY'] = 'test'
print(os.environ['NEW_KEY'])
os.environ['NEW_KEY'] = 'test2'
print(os.environ['NEW_KEY'])
os.environ['NEW_KEY'] = 100
TypeError: str
not int
os.environ['NEW_KEY'] = '100'
print(os.environ.pop('NEW_KEY'))
print(os.environ.pop('NEW_KEY'))
print(os.environ.pop('NEW_KEY', None))
os.environ['NEW_KEY'] = '100'
print(os.getenv('NEW_KEY'))
del os.environ['NEW_KEY'
print(os.getenv('NEW_KEY'))
del os.environ['NEW_KEY'
print(os.getenv('LANG'))
print(subprocess.check_output('date', encoding='utf-8'))
os.environ['LANG'] = 'en_US'
print(subprocess.check_output('date', encoding='utf-8'))
print(os.getenv('LANG'))
if os.getenv('LANG').startswith('ja'):
print('こんにちは')
print('Hello')
os.environ['LANG'] = 'ja_JP'
if os.getenv('LANG').startswith('ja'):
print('こんにちは')
print('Hello')
import os
path = os.getcwd()
print(path)
print(type(path))
os.chdir('../')
print(os.getcwd())
import os
new_dir_path = 'data/temp/new-dir'
os.mkdir(new_dir_path)
os.mkdir(new_dir_path)
new_dir_path_recursive = 'data/temp/new-dir2/new-sub-dir'
os.mkdir(new_dir_path_recursive)
file or directory
new_dir_path_recursive = 'data/temp/new-dir2/new-sub-dir'
os.makedirs(new_dir_path_recursive)
os.makedirs(new_dir_path_recursive)
os.makedirs(new_dir_path_recursive, exist_ok=True)
os.makedirs(new_dir_path_recursive)
except FileExistsError:
pass
def my_makedirs(path):
if not os.path.isdir(path):
os.makedirs(path)
my_makedirs(new_dir_path_recursive)
import ntpath
print(ntpath.sep)
print('\\')
print(ntpath.sep is '\\')
file_path = 'c:\\dir\\subdir\\filename.ext'
file_path_raw = r
filename.ext
print(file_path == file_path_raw)
print(ntpath.basename(file_path))
filename.ext
print(ntpath.dirname(file_path))
print(ntpath.split(file_path))
filename.ext
print(ntpath.splitdrive(file_path))
filename.ext
drive_letter = ntpath.splitdrive(file_path)[0][0]
print(drive_letter)
print(ntpath.join('c:', 'dir', 'subdir', 'filename.ext'))
c:dir
filename.ext
print(ntpath.join('c:', ntpath.sep, 'dir', 'subdir', 'filename.ext'))
filename.ext
print(ntpath.join('c:\\', 'dir', 'subdir', 'filename.ext'))
filename.ext
import os
filepath = './dir/subdir/filename.ext'
print(os.sep)
print(os.sep is os.path.sep)
basename = os.path.basename(filepath)
print(basename)
filename.ext
print(type(basename))
basename_without_ext = os.path.splitext(os.path.basename(filepath))[0]
print(basename_without_ext)
filepath_tar_gz = './dir/subdir/filename.tar.gz'
print(os.path.splitext(os.path.basename(filepath_tar_gz))[0])
filename.tar
print(os.path.basename(filepath_tar_gz).split('.', 1)[0])
dirname = os.path.dirname(filepath)
print(dirname)
print(type(dirname))
subdirname = os.path.basename(os.path.dirname(filepath))
print(subdirname)
base_dir_pair = os.path.split(filepath)
print(base_dir_pair)
filename.ext
print(type(base_dir_pair))
print(os.path.split(filepath)[0] == os.path.dirname(filepath))
print(os.path.split(filepath)[1] == os.path.basename(filepath))
dirname, basename = os.path.split(filepath)
print(dirname)
print(basename)
filename.ext
dirpath_without_sep = './dir/subdir'
print(os.path.split(dirpath_without_sep))
print(os.path.basename(dirpath_without_sep))
dirpath_with_sep = './dir/subdir/'
print(os.path.split(dirpath_with_sep))
print(os.path.basename(os.path.dirname(dirpath_with_sep)))
root_ext_pair = os.path.splitext(filepath)
print(root_ext_pair)
print(type(root_ext_pair))
root, ext = os.path.splitext(filepath)
print(root)
print(ext)
path = root + ext
print(path)
other_ext_filepath = os.path.splitext(filepath)[0] + '.jpg'
print(other_ext_filepath)
ext_without_dot = os.path.splitext(filepath)[1][1:]
print(ext_without_dot)
print(os.path.splitext(filepath_tar_gz))
print(filepath_tar_gz.split('.', 1))
dirname, basename = os.path.split(filepath_tar_gz)
basename_without_ext, ext = basename.split('.', 1)
path_without_ext = os.path.join(dirname, basename_without_ext)
print(path_without_ext)
print(ext)
tar.gz
ext_with_dot = '.' + ext
print(ext_with_dot)
tar.gz
path = os.path.join('dir', 'subdir', 'filename.ext')
print(path)
other_filepath = os.path.join(os.path.dirname(filepath), 'other_file.ext')
print(other_filepath)
import os
filepath = './data/temp/dir/file.txt'
dirpath = './data/temp/dir'
print(os.path.exists(filepath))
print(os.path.exists(dirpath))
print(os.path.isfile(filepath))
print(os.path.isfile(dirpath))
print(os.path.isdir(filepath))
print(os.path.isdir(dirpath))
dirpath_with_sep = './data/temp/dir/'
print(os.path.isfile(dirpath_with_sep))
print(os.path.isdir(dirpath_with_sep))
import os
print(os.path.getsize('data/src/lena_square.png'))
def get_dir_size(path='.'):
total = 0
with os.scandir(path) as it:
for entry in it:
if entry.is_file():
total += entry.stat().st_size
elif entry.is_dir():
total += get_dir_size(entry.path)
return total
print(get_dir_size('data/src'))
def get_size(path='.'):
if os.path.isfile(path):
return os.path.getsize(path)
elif os.path.isdir(path):
return get_dir_size(path)
print(get_size('data/src'))
print(get_size('data/src/lena_square.png'))
def get_dir_size_old(path='.'):
total = 0
for p in os.listdir(path):
full_path = os.path.join(path, p)
if os.path.isfile(full_path):
total += os.path.getsize(full_path)
elif os.path.isdir(full_path):
total += get_dir_size_old(full_path)
return total
print(get_dir_size_old('data/src'))
def get_size_old(path='.'):
if os.path.isfile(path):
return os.path.getsize(path)
elif os.path.isdir(path):
return get_dir_size_old(path)
print(get_size_old('data/src'))
print(get_size_old('data/src/lena_square.png'))
import os
os.makedirs('temp/', exist_ok=True)
with open('temp/file.txt', 'w') as f:
f.write('')
print(os.listdir('temp'))
file.txt
os.remove('temp/file.txt')
print(os.listdir('temp'))
os.remove('temp/')
not permitted
import os
import glob
def remove_glob(pathname, recursive=True):
for p in glob.glob(pathname, recursive=recursive):
if os.path.isfile(p):
os.remove(p)
remove_glob('temp/**/*.txt')
import os
import glob
import re
def remove_glob_re(pattern, pathname, recursive=True):
for p in glob.glob(pathname, recursive=recursive):
if os.path.isfile(p) and re.search(pattern, p):
os.remove(p)
import os
os.makedirs('temp/dir1/dir2/dir3/', exist_ok=True)
with open('temp/file.txt', 'w') as f:
f.write('')
print(os.listdir('temp/'))
file.txt
os.removedirs('temp/dir1/dir2/dir3/')
print(os.listdir('temp/'))
file.txt
os.remove('temp/file.txt')
import os
os.makedirs('temp/dir_empty/', exist_ok=True)
print(os.listdir('temp/'))
os.rmdir('temp/dir_empty/')
print(os.listdir('temp/'))
os.makedirs('temp/dir_not_empty/', exist_ok=True)
with open('temp/dir_not_empty/file.txt', 'w') as f:
f.write('')
os.rmdir('temp/dir_not_empty/')
not empty
os.rmdir('temp/dir_not_empty/')
except OSError as e:
pass
import shutil
shutil.rmtree('temp/dir_not_empty/')
print(os.listdir('temp/'))
import os
import pathlib
import datetime
import time
import platform
p = pathlib.Path('data/temp/test.txt')
p.write_text('test')
time.sleep(10)
p.write_text('update')
print(p.stat())
os.stat_result(st_mode=33188, st_ino=8728494137, st_dev=16777220, st_nlink=1, st_uid=501, st_gid=20, st_size=6, st_atime=1549094615, st_mtime=1549094615, st_ctime=1549094615)
print(type(p.stat()))
os.stat_result
print(os.stat('data/temp/test.txt'))
os.stat_result(st_mode=33188, st_ino=8728494137, st_dev=16777220, st_nlink=1, st_uid=501, st_gid=20, st_size=6, st_atime=1549094615, st_mtime=1549094615, st_ctime=1549094615)
print(type(os.stat('data/temp/test.txt')))
os.stat_result
print(os.stat(p))
os.stat_result(st_mode=33188, st_ino=8728494137, st_dev=16777220, st_nlink=1, st_uid=501, st_gid=20, st_size=6, st_atime=1549094615, st_mtime=1549094615, st_ctime=1549094615)
print(type(os.stat(p)))
os.stat_result
print(p.stat() == os.stat('data/temp/test.txt') == os.stat(p))
st = p.stat()
print(st.st_atime)
print(st.st_mtime)
print(st.st_ctime)
print(st.st_birthtime)
print(type(st.st_ctime))
print(st.st_ctime_ns)
print(type(st.st_ctime_ns))
print(os.path.getatime('data/temp/test.txt'))
print(os.path.getmtime('data/temp/test.txt'))
print(os.path.getctime('data/temp/test.txt'))
print(os.path.getctime(p))
print(os.path.getctime(p) == p.stat().st_ctime)
dt = datetime.datetime.fromtimestamp(p.stat().st_ctime)
print(dt)
print(type(dt))
datetime.datetime
print(dt.strftime('%Y年%m月%d日 %H:%M:%S'))
print(dt.isoformat())
print(os.path.getmtime('data/temp/test.txt'))
print(p.stat().st_mtime)
print(datetime.datetime.fromtimestamp(p.stat().st_mtime))
def creation_date(path_to_file):
if platform.system() == 'Windows':
return os.path.getctime(path_to_file)
stat = os.stat(path_to_file)
return stat.st_birthtime
except AttributeError:
Linux. No
return stat.st_mtime
print(creation_date(p))
print(datetime.datetime.fromtimestamp(creation_date(p)))
import pandas as pd
df = pd.DataFrame
index=['ONE', 'TWO', 'THREE']
print(df)
df['A'] = 0
print(df)
df['D'] = 0
print(df)
df['E'] = [0, 1, 2]
print(df)
df['F'] = [0, 1, 2, 3]
ValueError: Length
not match 
df['F'] = df['B'] + df['C']
df['G'] = df['B'].str.lower()
print(df)
s = pd.Series(['X2', 'X3', 'X4'], index=['TWO', 'THREE', 'FOUR'], name='X')
print(s)
Name: X
dtype: object
df['H'] = s
print(df)
print(s.values)
df['I'] = s.values
print(df)
df = pd.DataFrame
index=['ONE', 'TWO', 'THREE']
print(df.assign(A=0))
print(df.assign(D=0))
print(df)
s = pd.Series(['X2', 'X3', 'X4'], index=['TWO', 'THREE', 'FOUR'], name='X')
print(s)
Name: X
dtype: object
df_new = df.assign
C='XXX'
D=0
E=[0, 1, 2]
F=s
G=s.values
H=df['A'] + df['B']
print(df_new)
df = pd.DataFrame
index=['ONE', 'TWO', 'THREE']
s = pd.Series(['X2', 'X3', 'X4'], index=['TWO', 'THREE', 'FOUR'], name='X')
df.insert(0, 'D', 0)
print(df)
df.insert(len(df.columns), 'E', s)
print(df)
df.insert(10, 'F', 10)
ValueError: cannot
df.insert(-1, 'F', 10)
ValueError: unbounded
df.insert(0, 'D', 10)
ValueError: cannot
df.insert(0, 'D', 10, allow_duplicates=True)
print(df)
df = pd.DataFrame
index=['ONE', 'TWO', 'THREE']
s = pd.Series(['X2', 'X3', 'X4'], index=['TWO', 'THREE', 'FOUR'], name='X')
print(pd.concat([df, s], axis=1))
print(pd.concat([df, s], axis=1, join='inner'))
s1 = pd.Series(['X1', 'X2', 'X3'], index=df.index, name='X')
s2 = pd.Series(['Y1', 'Y2', 'Y3'], index=df.index, name='Y')
print(pd.concat([df, s1, s2], axis=1))
df2 = pd.DataFrame({'df_col1': 0, 'df_col2': range(3)}, index=df.index)
print(df2)
print(pd.concat([df, df2], axis=1))
import pandas as pd
df = pd.DataFrame
index=['ONE', 'TWO', 'THREE']
print(df)
df.loc['ONE'] = 0
print(df)
df.loc['FOUR'] = 0
df.loc['FIVE'] = ['A5', 'B5', 'C5']
print(df)
df.loc['SIX'] = ['A6', 'B6']
ValueError: cannot
s = pd.Series(['B6', 'C6', 'D6'], index=['B', 'C', 'D'], name='SIX')
print(s)
Name: SIX
dtype: object
df.loc['XXX'] = df.loc['TWO'] + df.loc['THREE']
df.loc['YYY'] = s
df.loc['ZZZ'] = s.values
print(df)
df = pd.DataFrame
index=['ONE', 'TWO', 'THREE']
print(df.append(0))
TypeError: cannot
Series and DataFrame
print(df.append([0, 1, 2]))
print(df.append({'A': 0, 'B': 1, 'C': 2}, ignore_index=True))
s = pd.Series(['A4', 'B4', 'C4'], index=df.columns, name='FOUR')
print(df.append(s))
s_mismatch = pd.Series(['B5', 'C5', 'D5'], index=['B', 'C', 'D'], name='FIVE')
print(df.append(s_mismatch))
print(df)
s_no_name = pd.Series(['B4', 'C4', 'D4'], index=['B', 'C', 'D'])
print(df.append(s_no_name))
TypeError: Can
ignore_index=True or if
print(df.append(s_no_name, ignore_index=True))
print(df.append([s, s_mismatch]))
df2 = pd.DataFrame
index=['SIX', 'SEVEN']
columns=['B', 'C', 'D']
print(df2)
print(df.append(df2))
print(df.append([s, df2]))
ValueError: all
5 and the
print(df.append(s).append(df2))
print(pd.concat([df, df2]))
print(pd.concat([df, df2], join='inner'))
print(df.append(s))
print(pd.concat([df, s]))
print(pd.DataFrame(s).T)
print(pd.concat([df, pd.DataFrame(s).T]))
print(df.T.assign(FOUR=0, FIVE=['A5', 'B5', 'C5']).T)
df_insert = df.T.insert(0, 'FOUR', 0).T
df_T = df.T
df_T.insert(0, 'FOUR', 0)
print(df_T.T)
import pandas as pd
import numpy as np
print(pd.__version__)
print(pd.DataFrame.agg is pd.DataFrame.aggregate)
df = pd.DataFrame({'A': [0, 1, 2], 'B': [3, 4, 5]})
print(df)
print(df.agg(['sum', 'mean', 'min', 'max']))
print(type(df.agg(['sum', 'mean', 'min', 'max'])))
pandas.core.frame.DataFrame
print(df.agg(['sum']))
print(type(df.agg(['sum'])))
pandas.core.frame.DataFrame
print(df.agg('sum'))
dtype: int64
print(type(df.agg('sum')))
pandas.core.series.Series
df.agg
print(df.agg({'A': 'sum', 'B': 'mean'}))
dtype: float64
print(df.agg({'A': ['sum'], 'B': ['mean']}))
print(df.agg({'A': ['min', 'max'], 'B': 'mean'}))
print(df.agg(['sum', 'mean', 'min', 'max'], axis=1))
s = df['A']
print(s)
Name: A
dtype: int64
print(s.agg(['sum', 'mean', 'min', 'max']))
Name: A
dtype: float64
print(type(s.agg(['sum', 'mean', 'min', 'max'])))
pandas.core.series.Series
print(s.agg(['sum']))
Name: A
dtype: int64
print(type(s.agg(['sum'])))
pandas.core.series.Series
print(s.agg('sum'))
print(type(s.agg('sum')))
numpy.int64
print(s.agg({'Total': 'sum', 'Average': 'mean', 'Min': 'min', 'Max': 'max'}))
Name: A
dtype: float64
print(s.agg({'NewLabel_1': ['sum', 'max'], 'NewLabel_2': ['mean', 'min']}))
SpecificationError: nested
print(df.agg(['mad', 'amax', 'dtype']))
print(df['A'].mad())
print(np.amax(df['A']))
print(df['A'].dtype)
print(df.agg(['xxx']))
not a 
print(df.agg('xxx'))
not a 
print(hasattr(pd.DataFrame, '__array__'))
print(hasattr(pd.core.groupby.GroupBy, '__array__'))
print(df.agg([np.sum, max]))
print(np.sum(df['A']))
print(max(df['A']))
print(np.abs(df['A']))
Name: A
dtype: int64
print(df.agg([np.abs]))
print(df.agg([np.abs, max]))
ValueError: cannot
transform and aggregation
def my_func(x):
return min(x) / max(x)
print(df.agg([my_func, lambda x: min(x) / max(x)]))
print(df['A'].std())
print(df['A'].std(ddof=0))
print(df.agg(['std', lambda x: x.std(ddof=0)]))
print(df.agg('std', ddof=0))
dtype: float64
print(df.agg(['std'], ddof=0))
df_str = df.assign(C=['X', 'Y', 'Z'])
print(df_str)
df_str['C'].mean()
TypeError: Could
not convert 
print(df_str.agg(['sum', 'mean']))
print(df_str.agg(['mean', 'std']))
print(df_str.agg(['sum', 'min', 'max']))
print(df_str.select_dtypes(include='number').agg(['sum', 'mean']))
import pandas as pd
df = pd.read_csv('data/src/sample_header.csv')
print(df)
s = df['c']
print(s)
Name: c
dtype: int64
s_f = s.astype('float64')
print(s_f)
Name: c
dtype: float64
print(s)
Name: c
dtype: int64
s_f = s.astype('float')
print(s_f.dtype)
s_f = s.astype(float)
print(s_f.dtype)
s_f = s.astype('f8')
print(s_f.dtype)
s_s = s.astype(str)
print(s_s)
Name: c
dtype: object
print(s_s.map(type))
Name: c
dtype: object
s_o = s.astype('object')
print(s_o)
Name: c
dtype: object
print(s_o.map(type))
Name: c
dtype: object
print(df)
print(df.dtypes)
dtype: object
df_f = df.astype('float64')
print(df_f)
print(df_f.dtypes)
dtype: object
df_fcol = df.astype({'a': float})
print(df_fcol)
print(df_fcol.dtypes)
dtype: object
df_fcol2 = df.astype({'a': 'float32', 'c': 'int8'})
print(df_fcol2)
print(df_fcol2.dtypes)
dtype: object
import pandas as pd
import datetime
print(pd.__version__)
df = pd.DataFrame
range(6)
columns=['A']
index=pd.date_range('2021-01-01', periods=6, freq='8H')
print(df)
print(type(df.index))
pandas.core.indexes.datetimes.DatetimeIndex
print(datetime.time(8, 0, 0))
print(df.at_time(datetime.time(8, 0, 0)))
print(df.at_time(datetime.time(12, 0, 0)))
print(df.at_time('16:00'))
print(df.at_time('4PM'))
df_sec = pd.DataFrame
range(6)
columns=['A']
index=pd.date_range('2021-01-01', periods=6, freq='8H5s')
print(df_sec)
print(df_sec.at_time('8:00'))
print(df_sec.at_time('8:00:00'))
print(df_sec.at_time('8:00:05'))
print(df_sec.between_time('8:00:00', '8:00:30'))
print(df_sec.between_time(datetime.time(8, 0, 0), datetime.time(8, 0, 30)))
print(df_sec.between_time('8:00:30', '8:00:00'))
print(df_sec.between_time('8:00:05', '8:00:30'))
print(df_sec.between_time('8:00:05', '8:00:30', include_start=False))
import pandas as pd
df1 = pd.DataFrame
index=['ONE', 'TWO', 'THREE']
print(df1)
df2 = pd.DataFrame
index=['TWO', 'THREE', 'FOUR']
print(df2)
s1 = pd.Series(['X1', 'X2', 'X3'], index=['ONE', 'TWO', 'THREE'], name='X')
print(s1)
Name: X
dtype: object
s2 = pd.Series(['Y2', 'Y3', 'Y4'], index=['TWO', 'THREE', 'FOUR'], name='Y')
print(s2)
Name: Y
dtype: object
df_concat = pd.concat([df1, df2])
print(df_concat)
df_concat_multi = pd.concat([df1, df2, df1])
print(df_concat_multi)
df_v = pd.concat([df1, df2], axis=0)
print(df_v)
df_h = pd.concat([df1, df2], axis=1)
print(df_h)
df_v_out = pd.concat([df1, df2], join='outer')
print(df_v_out)
df_v_in = pd.concat([df1, df2], join='inner')
print(df_v_in)
df_h_out = pd.concat([df1, df2], axis=1, join='outer')
print(df_h_out)
df_h_in = pd.concat([df1, df2], axis=1, join='inner')
print(df_h_in)
df_concat = pd.concat([df1, df2])
print(df_concat)
print(type(df_concat))
pandas.core.frame.DataFrame
s_v = pd.concat([s1, s2])
print(s_v)
dtype: object
print(type(s_v))
pandas.core.series.Series
s_h = pd.concat([s1, s2], axis=1)
print(s_h)
print(type(s_h))
pandas.core.frame.DataFrame
s_h_in = pd.concat([s1, s2], axis=1, join='inner')
print(s_h_in)
df_s_h = pd.concat([df1, s2], axis=1)
print(df_s_h)
df_s_h_in = pd.concat([df1, s2], axis=1, join='inner')
print(df_s_h_in)
df_s_v = pd.concat([df1, s2])
print(df_s_v)
df1.loc['FOUR'] = ['A4', 'B4', 'C4']
print(df1)
s = pd.Series(['A5', 'B5', 'C5'], index=df1.columns, name='FIVE')
print(s)
Name: FIVE
dtype: object
df_append = df1.append(s)
print(df_append)
import pandas as pd
import numpy as np
import seaborn as sns
import matplotlib.pyplot as plt
df = pd.DataFrame
range(5)
x**2 for x in range(5)
x**3 for x in range(5)
print(df)
df_corr = df.corr()
print(df_corr)
print(type(df_corr))
pandas.core.frame.DataFrame
df['D'] = list('abcde')
df['E'] = [True, False, True, True, False]
print(df)
print(df.dtypes)
dtype: object
df_corr = df.corr()
print(df_corr)
df_nan = df.copy()
df_nan.iloc[[2, 3, 4], 1] = np.nan
print(df_nan)
df_nan_corr = df_nan.corr()
print(df_nan_corr)
sns.heatmap(df_corr, vmax=1, vmin=-1, center=0)
plt.savefig('data/dst/seaborn_heatmap_corr_example.png')
df_house = pd.read_csv('data/src/house_prices_train.csv', index_col=0)
print(df_house.shape)
print(df_house.dtypes.value_counts())
dtype: int64
df_house_corr = df_house.corr()
print(df_house_corr.shape)
fig, ax = plt.subplots(figsize=(12, 9)) 
sns.heatmap(df_house_corr, square=True, vmax=1, vmin=-1, center=0)
plt.savefig('data/dst/seaborn_heatmap_house_price.png')
import pandas as pd
df = pd.read_csv('data/src/sample_pandas_normal.csv')
print(df)
df_bool = (df == 'CA')
print(df_bool)
print(df_bool.sum())
dtype: int64
print(df_bool.sum(axis=1))
dtype: int64
print(df_bool.values)
print(type(df_bool.values))
numpy.ndarray
print(df_bool.values.sum())
s_bool = df['age'] < 25
print(s_bool)
Name: age
dtype: bool
print(s_bool.sum())
df_bool_multi = ((df == 'CA') | (df == 70))
print(df_bool_multi)
print(df_bool_multi.sum())
dtype: int64
print(df_bool_multi.sum(axis=1))
dtype: int64
print(df_bool_multi.values.sum())
df_bool_multi_and = ((df['state'] == 'CA') & (df['age'] < 30))
print(df_bool_multi_and)
dtype: bool
print(df_bool_multi_and.sum())
df_bool_multi_or = ((df['state'] == 'CA') | (df['age'] < 30))
print(df_bool_multi_or)
dtype: bool
print(df_bool_multi_or.sum())
df == 'CA'
print(df_bool_not)
print(df_bool_not.sum())
dtype: int64
print(df_bool_not.sum(axis=1))
dtype: int64
print(df_bool_not.values.sum())
df_num = df[['age', 'point']]
print(df_num)
print((df_num <= 70).sum())
dtype: int64
print(((df['age'] > 20) & (df['age'] < 40)).sum())
print((df_num % 2 == 1).sum())
dtype: int64
df_str = df[['name', 'state']]
print(df_str)
print((df_str == 'NY').sum())
dtype: int64
print(df_str['name'].str.endswith('e'))
Name: name
dtype: bool
print(df_str['name'].str.endswith('e').sum())
df = pd.read_csv('data/src/titanic_train.csv')
print(df.head())
Mr. Owen
Mrs. John
Miss. Laina
Mrs. Jacques
Mr. William
df.info()
pandas.core.frame.DataFrame
dtypes: float64
int64(5)
object(5)
print(df.isnull().head())
print(df.isnull().sum())
dtype: int64
print(df.isnull().sum(axis=1).head())
dtype: int64
print(df.isnull().values.sum())
print(df.count())
dtype: int64
print(df.count(axis=1).head())
dtype: int64
print(df.count().sum())
import pandas as pd
df = pd.read_csv('data/src/titanic_train.csv', index_col=0).drop(['Name', 'Ticket', 'SibSp', 'Parch'], axis=1)
print(df.head())
print(pd.crosstab(df['Sex'], df['Pclass']))
print(type(pd.crosstab(df['Sex'], df['Pclass'])))
pandas.core.frame.DataFrame
print(pd.crosstab([df['Sex'], df['Survived']], [df['Pclass'], df['Embarked']]))
pd.crosstab
df['Sex']
df['Survived']
df['Pclass']
df['Embarked']
margins=True
pd.crosstab
df['Sex']
df['Survived']
df['Pclass']
df['Embarked']
margins=True
margins_name='Total'
print(pd.crosstab(df['Sex'], df['Pclass'], margins=True, normalize=True))
print(pd.crosstab(df['Sex'], df['Pclass'], margins=True, normalize='index'))
print(pd.crosstab(df['Sex'], df['Pclass'], margins=True, normalize='columns'))
pd.crosstab
df['Sex']
df['Pclass']
df['Embarked']
margins=True
normalize=True
TypeError: Expected
pd.crosstab
df['Sex']
df['Pclass']
df['Embarked']
margins=True
normalize='index'
pd.crosstab
df['Sex']
df['Pclass']
df['Embarked']
margins=True
normalize='columns'
ValueError: Length
print(pd.crosstab(df['Sex'], [df['Pclass'], df['Embarked']], normalize=True))
print(pd.crosstab(df['Sex'], [df['Pclass'], df['Embarked']], normalize='index'))
print(pd.crosstab(df['Sex'], [df['Pclass'], df['Embarked']], normalize='columns'))
import pandas as pd
df = pd.read_csv('data/src/sample.csv')
print(df)
print(df.columns)
Index(['11', '12', '13', '14'], dtype='object')
df_none = pd.read_csv('data/src/sample.csv', header=None)
print(df_none)
df_names = pd.read_csv('data/src/sample.csv', names=('A', 'B', 'C', 'D'))
print(df_names)
df_header = pd.read_csv('data/src/sample_header.csv')
print(df_header)
df_header_0 = pd.read_csv('data/src/sample_header.csv', header=0)
print(df_header_0)
df_header_2 = pd.read_csv('data/src/sample_header.csv', header=2)
print(df_header_2)
df_header_index = pd.read_csv('data/src/sample_header_index.csv')
print(df_header_index)
print(df_header_index.index)
RangeIndex(start=0, stop=3, step=1)
df_header_index_col = pd.read_csv('data/src/sample_header_index.csv', index_col=0)
print(df_header_index_col)
print(df_header_index_col.index)
Index(['ONE', 'TWO', 'THREE'], dtype='object')
df_none_usecols = pd.read_csv('data/src/sample.csv', header=None, usecols=[1, 3])
print(df_none_usecols)
df_none_usecols = pd.read_csv('data/src/sample.csv', header=None, usecols=[2])
print(df_none_usecols)
df_header_usecols = pd.read_csv('data/src/sample_header.csv', usecols=['a', 'c'])
print(df_header_usecols)
df_header_usecols = pd.read_csv
usecols=lambda x: x is not 'b'
print(df_header_usecols)
df_header_usecols = pd.read_csv
usecols=lambda x: x not in ['a', 'c']
print(df_header_usecols)
df_index_usecols = pd.read_csv
index_col=0
usecols=[0, 1, 3]
print(df_index_usecols)
df_none = pd.read_csv('data/src/sample.csv', header=None)
print(df_none)
df_none = pd.read_csv('data/src/sample.csv', header=None, skiprows=2)
print(df_none)
df_none_skiprows = pd.read_csv('data/src/sample.csv', header=None, skiprows=[0, 2])
print(df_none_skiprows)
df_none_skiprows = pd.read_csv('data/src/sample.csv', header=None, skiprows=[1])
print(df_none_skiprows)
df_none_skiprows = pd.read_csv
header=None
skiprows=lambda x: x not in [0, 2]
print(df_none_skiprows)
df_header_skiprows = pd.read_csv('data/src/sample_header.csv', skiprows=[1])
print(df_header_skiprows)
df_header_skiprows = pd.read_csv('data/src/sample_header.csv', skiprows=[0, 3])
print(df_header_skiprows)
df_none_skipfooter = pd.read_csv
header=None
skipfooter=1
engine='python'
print(df_none_skipfooter)
df_none_nrows = pd.read_csv('data/src/sample.csv', header=None, nrows=2)
print(df_none_nrows)
df_default = pd.read_csv('data/src/sample_header_index_dtype.csv', index_col=0)
print(df_default)
print(df_default.dtypes)
dtype: object
print(df_default.applymap(type))
df_str = pd.read_csv
index_col=0
dtype=str
print(df_str)
print(df_str.dtypes)
dtype: object
print(df_str.applymap(type))
df_object = pd.read_csv
index_col=0
dtype=object
print(df_object)
print(df_object.dtypes)
dtype: object
print(df_object.applymap(type))
df_int = pd.read_csv
index_col=0
dtype=int
ValueError: invalid
int() 
df_str_cast = df_str.astype({'a': int})
print(df_str_cast)
print(df_str_cast.dtypes)
dtype: object
df_str_col = pd.read_csv
index_col=0
dtype={'b': str, 'c': str}
print(df_str_col)
print(df_str_col.dtypes)
dtype: object
df_str_col_num = pd.read_csv
index_col=0
dtype={2: str, 3: str}
print(df_str_col_num)
print(df_str_col_num.dtypes)
dtype: object
df_nan = pd.read_csv('data/src/sample_header_index_nan.csv', index_col=0)
print(df_nan)
print(df_nan.isnull())
df_nan_set_na = pd.read_csv
index_col=0
na_values='-'
print(df_nan_set_na)
print(df_nan_set_na.isnull())
df_nan_set_na_no_keep = pd.read_csv
index_col=0
na_values=['-', 'NaN', 'null']
keep_default_na=False
print(df_nan_set_na_no_keep)
print(df_nan_set_na_no_keep.isnull())
df_nan_no_filter = pd.read_csv
index_col=0
na_filter=False
print(df_nan_no_filter)
print(df_nan_no_filter.isnull())
df_tsv = pd.read_table('data/src/sample_header_index.tsv', index_col=0)
print(df_tsv)
df_tsv_sep = pd.read_csv('data/src/sample_header_index.tsv', index_col=0, sep='\t')
print(df_tsv_sep)
import pandas as pd
print(pd.__version__)
df = pd.DataFrame({'A': [1, 2, 3], 'B': [4, 5, 6]}, index=['X', 'Y', 'Z'])
print(df)
print(df.cumsum())
print(df.cumsum(axis=1))
print(df.cumprod())
print(df.cumprod(axis=1))
print(df['B'])
Name: B
dtype: int64
print(type(df['B']))
pandas.core.series.Series
print(df['B'].cumsum())
Name: B
dtype: int64
print(df['B'].cumprod())
Name: B
dtype: int64
df_nan = pd.DataFrame({'A': [1, 2, 3], 'B': [4, float('nan'), 6]}, index=['X', 'Y', 'Z'])
print(df_nan)
print(df_nan.cumsum())
print(float('nan') + 4)
print(df_nan.cumsum(skipna=False))
print(df_nan.cumprod())
print(df_nan.cumprod(skipna=False))
df2 = pd.DataFrame({'A': [1, 4, 2], 'B': [6, 3, 5]}, index=['X', 'Y', 'Z'])
print(df2)
print(df2.cummax())
print(df2.cummax(axis=1))
print(df2.cummin())
print(df2.cummin(axis=1))
df2_nan = pd.DataFrame({'A': [1, 4, 2], 'B': [6, float('nan'), 5]}, index=['X', 'Y', 'Z'])
print(df2_nan)
print(df2_nan.cummax())
print(df2_nan.cummax(skipna=False))
print(df2_nan.cummin())
print(df2_nan.cummin(skipna=False))
import pandas as pd
s = pd.Series
data=[x**2 for x in range(11)]
index=list('abcdefghijk')
print(s)
dtype: int64
s_cut = pd.cut(s, 4)
print(s_cut)
-0.1
-0.1
-0.1
-0.1
-0.1
-0.1
dtype: category
Categories (4, interval[float64])
-0.1
print(type(s_cut))
pandas.core.series.Series
print(pd.cut(s, [0, 10, 50, 100]))
dtype: category
Categories (3, interval[int64])
s_cut, bins = pd.cut(s, 4, retbins=True)
print(s_cut)
-0.1
-0.1
-0.1
-0.1
-0.1
-0.1
dtype: category
Categories (4, interval[float64])
-0.1
print(bins)
print(type(bins))
-0.1
numpy.ndarray
print(pd.cut(s, 4, right=False))
dtype: category
Categories (4, interval[float64])
print(pd.cut(s, 4, labels=False))
dtype: int64
print(pd.cut(s, 4, labels=['small', 'medium', 'large', 'x-large']))
dtype: category
Categories (4, object)
print(pd.cut(s, 3))
-0.1
-0.1
-0.1
-0.1
-0.1
-0.1
dtype: category
Categories (3, interval[float64])
-0.1
print(pd.cut(s, 3, precision=1))
-0.1
-0.1
-0.1
-0.1
-0.1
-0.1
dtype: category
Categories (3, interval[float64])
-0.1
print(pd.qcut(s, 2))
-0.001
-0.001
-0.001
-0.001
-0.001
-0.001
dtype: category
Categories (2, interval[float64])
-0.001
s_qcut, bins = pd.qcut(s, 4, labels=['Q1', 'Q2', 'Q3', 'Q4'], retbins=True)
print(s_qcut)
dtype: category
Categories (4, object)
print(bins)
s_duplicate = pd.Series
data=[0, 0, 0, 0, 0, 1, 2, 3, 4, 5, 6]
index=list('abcdefghijk')
print(s_duplicate)
dtype: int64
print(pd.qcut(s_duplicate, 2))
-0.001
-0.001
-0.001
-0.001
-0.001
-0.001
dtype: category
Categories (2, interval[float64])
-0.001
print(pd.qcut(s_duplicate, 4))
ValueError: Bin
unique: array
print(pd.qcut(s_duplicate, 4, duplicates='drop'))
-0.001
-0.001
-0.001
-0.001
-0.001
-0.001
dtype: category
Categories (3, interval[float64])
-0.001
counts = pd.cut(s, 3, labels=['S', 'M', 'L']).value_counts()
print(counts)
dtype: int64
print(type(counts))
pandas.core.series.Series
print(counts['M'])
print(pd.value_counts(pd.cut(s, 3, labels=['S', 'M', 'L'])))
dtype: int64
l = [x**2 for x in range(11)]
print(l)
l_cut = pd.cut(l, 3, labels=['S', 'M', 'L'])
print(l_cut)
Categories (3, object)
print(type(l_cut))
pandas.core.categorical.Categorical
print(l_cut[0])
print(list(l_cut))
print(pd.value_counts(l_cut))
dtype: int64
df_titanic = pd.read_csv('data/src/titanic_train.csv').drop(['Name', 'Ticket', 'Cabin', 'Embarked'], axis=1)
print(df_titanic.head())
print(df_titanic['Age'].describe())
Name: Age
dtype: float64
print(pd.cut(df_titanic['Age'], 5, precision=0).value_counts(sort=False, dropna=False))
Name: Age
dtype: int64
df_titanic['Age_bin'] = pd.cut(df_titanic['Age'], 5, labels=False)
print(df_titanic.head())
import pandas as pd
import numpy as np
df_simple = pd.DataFrame(np.arange(12).reshape(3, 4))
print(df_simple)
print(df_simple.values)
print(type(df_simple.values))
numpy.ndarray
print(df_simple.columns)
RangeIndex(start=0, stop=4, step=1)
print(type(df_simple.columns))
pandas.core.indexes.range.RangeIndex
print(df_simple.index)
RangeIndex(start=0, stop=3, step=1)
print(type(df_simple.index))
pandas.core.indexes.range.RangeIndex
print(list(df_simple.columns))
print(type(list(df_simple.columns)))
print(df_simple.columns.tolist())
print(type(df_simple.columns.tolist()))
df = pd.DataFrame
np.arange(12).reshape(3, 4)
columns=['col_0', 'col_1', 'col_2', 'col_3']
index=['row_0', 'row_1', 'row_2']
print(df)
print(df.columns)
Index(['col_0', 'col_1', 'col_2', 'col_3'], dtype='object')
print(type(df.columns))
pandas.core.indexes.base.Index
print(df.index)
Index(['row_0', 'row_1', 'row_2'], dtype='object')
print(type(df.index))
pandas.core.indexes.base.Index
print(df.columns.tolist())
print(type(df.columns.tolist()))
print(df.columns[0])
df.columns[0] = 'Col_0'
TypeError: Index
not support 
df.columns = ['Col_0', 'Col_1', 'Col_2', 'Col_3']
df.index = ['Row_0', 'Row_1', 'Row_2']
print(df)
df.values = np.arange(12).reshape(3, 4) * 10
AttributeError: can
print(df['Col_1'])
Name: Col_1
dtype: int64
print(type(df['Col_1']))
pandas.core.series.Series
print(df.Col_1)
Name: Col_1
dtype: int64
print(df.loc[:, 'Col_1'])
Name: Col_1
dtype: int64
print(df.loc['Row_1', :])
Name: Row_1
dtype: int64
print(type(df.loc['Row_1', :]))
pandas.core.series.Series
print(df.loc['Row_1'])
Name: Row_1
dtype: int64
print(df.loc[['Row_0', 'Row_2'], ['Col_1', 'Col_3']])
print(type(df.loc[['Row_0', 'Row_2'], ['Col_1', 'Col_3']]))
pandas.core.frame.DataFrame
print(df.loc['Row_0', 'Col_1'])
print(type(df.loc['Row_0', 'Col_1']))
numpy.int64
print(df.at['Row_0', 'Col_1'])
print(df.at[:, 'Col_1'])
slice(None, None, None)
print(df.iloc[[0, 2], [1, 3]])
print(df.iat[0, 1])
print(df.query('Col_0 > 2'))
df.loc[:, 'Col_1'] = [10, 50, 90]
print(df)
df.loc[:] = np.arange(12).reshape(3, 4) * 100
print(df)
df.loc[:, 'Col_1'] = [10, 50, 90, 130]
ValueError: Length
not match 
df.loc[:] = np.arange(16).reshape(4, 4) * 100
ValueError: cannot
df_multi = pd.read_csv('data/src/sample_pandas_normal.csv', index_col=0)
print(df_multi)
print(df_multi.dtypes)
dtype: object
import pandas as pd
import numpy as np
print(np.arange(9).reshape(3, 3))
print(pd.DataFrame(np.arange(9).reshape(3, 3)))
pd.DataFrame
np.arange(9).reshape(3, 3)
columns=['col_0', 'col_1', 'col_2']
index=['row_0', 'row_1', 'row_2']
print(pd.DataFrame([[0, 1, 2], [3, 4, 5], [6, 7, 8]]))
print(pd.DataFrame([[0, 1, 2], [3, 4, 5], [6, 7, 8, 9, 10]]))
pd.DataFrame
np.arange(3, 6)
pd.DataFrame
np.arange(3, 6)
index=['row_0', 'row_1', 'row_2']
pd.DataFrame
np.arange(3, 6)
ValueError: arrays
pd.DataFrame
np.arange(3, 6)
index=['row_0', 'row_1', 'row_2']
pd.DataFrame
np.arange(3, 6)
pd.DataFrame.from_dict
orient='index'
pd.DataFrame.from_dict
np.array([0, 1, 2])
np.array([3, 4, 5])
np.array([6, 7, 8])
orient='index'
pd.DataFrame.from_dict
np.array([3, 4, 5])
orient='index'
TypeError: Expected
numpy.ndarray
pd.DataFrame
pd.DataFrame
index=['row_0', 'row_1', 'row_2']
pd.DataFrame
pd.DataFrame
import pandas as pd
df = pd.read_csv('data/src/sample_pandas_normal.csv', index_col=0)
df['sex'] = ['Female', 'Male', 'Male', 'Male', 'Female', 'Male']
print(df)
print(df.mean())
dtype: float64
print(df.pivot_table(index='state', columns='sex', aggfunc='mean'))
import pandas as pd
df = pd.DataFrame(pd.np.arange(1000).reshape(100, 10))
print(df.shape)
print(df.head())
print(df.tail())
for i, row in df.iterrows():
pass
std. dev. of
for t in df.itertuples():
pass
std. dev. of
for t in df.itertuples(name=None):
pass
std. dev. of
for i in df[0]:
pass
std. dev. of
for i, j, k in zip(df[0], df[4], df[9]):
pass
std. dev. of
for t in zip(df[0], df[1], df[2], df[3], df[4], df[5], df[6], df[7], df[8], df[9]):
pass
std. dev. of
import pandas as pd
df = pd.DataFrame
index=['ONE', 'TWO', 'THREE']
print(df)
df_new = df.rename(columns={'A': 'Col_1'}, index={'ONE': 'Row_1'})
print(df_new)
print(df)
print(df.rename(columns={'A': 'Col_1', 'C': 'Col_3'}))
df_copy = df.copy()
df_copy.rename(columns={'A': 'Col_1'}, index={'ONE': 'Row_1'}, inplace=True)
print(df_copy)
print(df.rename(columns=str.lower, index=str.title))
df.rename
columns=lambda s: s*3
index=lambda s: s + '!!'
print(df.add_prefix('X_'))
print(df.add_suffix('_X'))
print(df.set_axis(['Row_1', 'Row_2', 'Row_3'], axis=0))
print(df.set_axis(['Row_1', 'Row_2', 'Row_3'], axis='index'))
print(df.set_axis(['Col_1', 'Col_2', 'Col_3'], axis=1))
print(df.set_axis(['Col_1', 'Col_2', 'Col_3'], axis='columns'))
print(df.set_axis(['Row_1', 'Row_2', 'Row_3']))
print(df.set_axis(['Row_1', 'Row_2', 'Row_3', 'Row_4']))
ValueError: Length
mismatch: Expected
df_copy = df.copy()
df_copy.set_axis(['Row_1', 'Row_2', 'Row_3'], inplace=True)
print(df_copy)
df.index = ['Row_1', 'Row_2', 'Row_3']
df.columns = ['Col_1', 'Col_2', 'Col_3']
print(df)
df.index = ['Row_1', 'Row_2', 'Row_3', 'Row_4']
ValueError: Length
mismatch: Expected
import pandas as pd
df = pd.DataFrame
index=['row0', 'row1', 'row2']
print(df)
print(df.dtypes)
dtype: object
s = df['col1']
print(s)
Name: col1
dtype: float64
print(type(s))
pandas.core.series.Series
print(s.dtype)
print(s.index)
Index(['row0', 'row1', 'row2'], dtype='object')
print(s.name)
print(df.loc[:, 'col1'])
Name: col1
dtype: float64
print(df.iloc[:, 2])
Name: col2
dtype: object
print(df.iloc[[0, 2], 2])
Name: col2
dtype: object
print(df.iloc[:2, 2])
Name: col2
dtype: object
df_only = df[['col1']]
print(df_only)
print(type(df_only))
pandas.core.frame.DataFrame
df_only2 = df.iloc[:, 1:2]
print(df_only2)
print(type(df_only2))
pandas.core.frame.DataFrame
s_r = df.loc['row1', :]
print(s_r)
Name: row1
dtype: object
print(type(s_r))
pandas.core.series.Series
print(s_r.dtype)
print(s_r.index)
Index(['col0', 'col1', 'col2'], dtype='object')
print(s_r.name)
print(df.loc['row1'])
Name: row1
dtype: object
print(df.iloc[2, [0, 2]])
Name: row2
dtype: object
df_only_r = df.iloc[[1]]
print(df_only_r)
print(type(df_only_r))
pandas.core.frame.DataFrame
df_only_r2 = df[1:2]
print(df_only_r2)
print(type(df_only_r2))
pandas.core.frame.DataFrame
print(s_r[0])
print(type(s_r[0]))
numpy.int64
print(s_r[1])
print(type(s_r[1]))
numpy.float64
print(s_r[2])
print(type(s_r[2]))
df_n = df[['col0', 'col1']]
print(df_n)
print(df_n.dtypes)
dtype: object
s_n_r = df_n.iloc[1]
print(s_n_r)
Name: row1
dtype: float64
print(s_n_r[0])
print(type(s_n_r[0]))
numpy.float64
print(s_n_r[1])
print(type(s_n_r[1]))
numpy.float64
print(df)
s = df['col0']
print(s)
Name: col0
dtype: int64
df.iat[0, 0] = 100
print(df)
print(s)
Name: col0
dtype: int64
s_copy = df['col1'].copy()
print(s_copy)
Name: col1
dtype: float64
df.iat[0, 1] = 100
print(df)
print(s_copy)
Name: col1
dtype: float64
s_l = df.loc[['row0', 'row2'], 'col2']
print(s_l)
Name: col2
dtype: object
df.iat[0, 2] = 'XXX'
print(df)
print(s_l)
Name: col2
dtype: object
import pandas_datareader.data as web
import datetime
import matplotlib.pyplot as plt
start = datetime.datetime(2012, 1, 1)
end = datetime.datetime(2017, 12, 31)
f = web.DataReader('SNE', 'morningstar', start, end)
print(f.head())
f2 = web.DataReader(['SNE', 'AAPL'], 'morningstar', start, end)
print(type(f2.index))
print(f2.head())
print(f2.tail())
pandas.core.indexes.multi.MultiIndex
f2_u = f2.unstack(0)
print(f2_u.head())
print(f2_u['Close'].head())
f2_u['Close'].plot(title='SNE vs AAPL', grid=True)
plt.show()
plt.savefig('data/dst/pandas_datareader_morningstar.png')
f2_u['Close', 'AAPL'] /= f2_u['Close'].loc[f2_u.index[0], 'AAPL']
f2_u['Close', 'SNE'] /= f2_u['Close'].loc[f2_u.index[0], 'SNE']
f2_u['Close'].plot(title='SNE vs AAPL', grid=True)
plt.show()
plt.savefig('data/dst/pandas_datareader_morningstar_normalize.png')
import warnings
warnings.simplefilter('ignore', FutureWarning)
import datetime
import pandas as pd
import pandas_datareader.data as web
import matplotlib.pyplot as plt
with open('data/temp/alpha_vantage_api_key.txt') as f:
api_key = f.read()
start = datetime.datetime(2015, 1, 1)
end = datetime.datetime(2019, 12, 31)
df_sne = web.DataReader('SNE', 'av-daily', start, end, api_key=api_key)
print(df_sne)
df_aapl = web.DataReader('AAPL', 'av-daily', start, end, api_key=api_key)
print(df_aapl)
df_sne_aapl = pd.DataFrame({'SNE': df_sne['close'], 'AAPL': df_aapl['close']})
print(df_sne_aapl)
df_sne.to_csv('data/src/sne_2015_2019.csv')
df_aapl.to_csv('data/src/aapl_2015_2019.csv')
df_sne_aapl.to_csv('data/src/sne_aapl_2015_2019.csv')
print(type(df_sne_aapl.index))
pandas.core.indexes.base.Index
df_sne_aapl.index = pd.to_datetime(df_sne_aapl.index)
print(type(df_sne_aapl.index))
pandas.core.indexes.datetimes.DatetimeIndex
df_sne_aapl.plot(title='SNE vs. AAPL', grid=True)
plt.show()
plt.savefig('data/dst/pandas_datareader_stock.png')
plt.close()
vs. AAPL
df_sne_aapl['SNE'] /= df_sne_aapl['SNE'][0]
df_sne_aapl['AAPL'] /= df_sne_aapl['AAPL'][0]
df_sne_aapl.plot(title='SNE vs. AAPL', grid=True)
plt.savefig('data/dst/pandas_datareader_stock_normalize.png')
plt.close()
vs. AAPL
import pandas_datareader.data as web
import datetime
start = datetime.datetime(2010, 1, 1)
end = datetime.datetime(2015, 12, 31)
f = web.DataReader('SNE', 'yahoo', start, end)
print(f)
f = web.DataReader(['SNE', 'AAPL'], 'yahoo', start, end)
print(f['Adj Close'])
print(f)
pandas.core.panel.Panel
6 (items) 
1511 (major_axis) 
2 (minor_axis)
axis: Adj
axis: AAPL
import matplotlib.pyplot as plt
f['Adj Close']['SNE'] /= f['Adj Close']['SNE'][-1]
f['Adj Close']['AAPL'] /= f['Adj Close']['AAPL'][-1]
f['Adj Close'].plot(title='SNE vs AAPL', grid=True)
plt.show()
plt.savefig('data/dst/pandas_datareader_yahoo.png')
pandas_datareader_yahoo.png
from pandas_datareader import wb
f = wb.download
indicator='SP.POP.TOTL'
country=['JP', 'US']
start=1960
end=2014
print(f)
SP.POP.TOTL
print(wb.search('gdp.*capita.*const').iloc[:, :2])
6.0.GDPpc_constant
NY.GDP.PCAP.KD
NY.GDP.PCAP.KN
NY.GDP.PCAP.PP.KD
NY.GDP.PCAP.PP.KD
f = wb.download
indicator='SP.POP.TOTL'
country=['JP', 'US']
start=1960
end=2014
f2 = f.unstack(level=0)
print(f2)
SP.POP.TOTL
f2.columns = ['Japan', 'United States']
f2.plot(grid=True)
plt.show()
plt.savefig('data/dst/pandas_datareader_wb.png')
pandas_datareader_wb.png
import warnings
warnings.simplefilter('ignore', FutureWarning)
from pandas_datareader import wb
import matplotlib.pyplot as plt
df = wb.download
indicator='SP.POP.TOTL'
country=['JP', 'US']
start=1960
end=2014
print(df)
SP.POP.TOTL
df2 = df.unstack(level=0)
print(df2.head())
SP.POP.TOTL
print(df2.tail())
SP.POP.TOTL
print(df2.columns)
SP.POP.TOTL
SP.POP.TOTL
names=[None, 'country']
df2.columns = ['Japan', 'United States']
print(df2.head())
df2.plot(grid=True)
plt.savefig('data/dst/pandas_datareader_wb.png')
plt.close()
pandas_datareader_wb.png
import pandas as pd
import datetime
df = pd.read_csv('data/src/sample_datetime_multi.csv')
print(df)
print(df.dtypes)
dtype: object
print(type(df['A'][0]))
print(pd.to_datetime(df['A']))
Name: A
dtype: datetime64[ns]
print(pd.to_datetime(df['B'], format='%Y年%m月%d日 %H時%M分'))
Name: B
dtype: datetime64[ns]
print(pd.to_datetime(df['A']) == pd.to_datetime(df['B'], format='%Y年%m月%d日 %H時%M分'))
dtype: bool
df['X'] = pd.to_datetime(df['A'])
print(df)
print(df.dtypes)
datetime64[ns]
dtype: object
print(df['X'][0])
print(type(df['X'][0]))
pandas._libs.tslib.Timestamp
print(issubclass(pd.Timestamp, datetime.datetime))
print(df['X'][0].year)
print(df['X'][0].weekday_name)
py_dt = df['X'][0].to_pydatetime()
print(type(py_dt))
datetime.datetime
dt64 = df['X'][0].to_datetime64()
print(type(dt64))
numpy.datetime64
print(df['X'][0].timestamp())
print(pd.to_datetime('1970-01-01 00:00:00').timestamp())
print(int(df['X'][0].timestamp()))
print(df['X'][0].strftime('%Y/%m/%d'))
print(df['X'].dt.year)
Name: X
dtype: int64
print(df['X'].dt.hour)
Name: X
dtype: int64
print(df['X'].dt.dayofweek)
Name: X
dtype: int64
print(df[df['X'].dt.dayofweek == 4])
print(df['X'].astype(str))
Name: X
dtype: object
print(df['X'].dt.strftime('%A, %B %d, %Y'))
Name: X
dtype: object
print(df['X'].dt.strftime('%Y年%m月%d日'))
Name: X
dtype: object
df['en'] = df['X'].dt.strftime('%A, %B %d, %Y')
df['jp'] = df['X'].dt.strftime('%Y年%m月%d日')
print(df)
print(df['X'].dt.to_pydatetime())
datetime.datetime(2017, 11, 1, 12, 24)
datetime.datetime(2017, 11, 18, 23, 0)
datetime.datetime(2017, 12, 5, 5, 5)
datetime.datetime(2017, 12, 22, 8, 54)
datetime.datetime(2018, 1, 8, 14, 20)
datetime.datetime(2018, 1, 19, 20, 1)
print(type(df['X'].dt.to_pydatetime()))
print(type(df['X'].dt.to_pydatetime()[0]))
numpy.ndarray
datetime.datetime
print(df['X'].values)
print(type(df['X'].values))
print(type(df['X'].values[0]))
numpy.ndarray
numpy.datetime64
print(df['X'].map(pd.Timestamp.timestamp))
Name: X
dtype: float64
print(df['X'].map(pd.Timestamp.timestamp).astype(int))
Name: X
dtype: int64
df_i = df.set_index('X').drop(['en', 'jp'], axis=1)
print(df_i)
print(df_i.index)
dtype='datetime64[ns]'
name='X'
freq=None
print(df_i.index.minute)
Int64Index([24, 0, 5, 54, 20, 1], dtype='int64', name='X')
print(df_i.index.strftime('%y/%m/%d'))
df_i['min'] = df_i.index.minute
df_i['str'] = df_i.index.strftime('%y/%m/%d')
print(df_i)
df_csv = pd.read_csv('data/src/sample_datetime_multi.csv', parse_dates=[0])
print(df_csv)
print(df_csv.dtypes)
datetime64[ns]
dtype: object
df_csv_jp = pd.read_csv
parse_dates=[1]
date_parser=lambda date: pd.to_datetime(date, format='%Y年%m月%d日 %H時%M分')
print(df_csv_jp)
print(df_csv_jp.dtypes)
datetime64[ns]
dtype: object
df_csv_jp_i = pd.read_csv
index_col=1
parse_dates=True
date_parser=lambda date: pd.to_datetime(date, format='%Y年%m月%d日 %H時%M分')
print(df_csv_jp_i)
print(df_csv_jp_i.index)
dtype='datetime64[ns]'
name='B'
freq=None
import pandas as pd
df = pd.DataFrame
print(df)
print(df.dtypes)
dtype: object
print(df.describe())
print(type(df.describe()))
pandas.core.frame.DataFrame
print(df.describe().loc['std'])
Name: std
dtype: float64
print(df.describe().at['std', 'b'])
print(df.describe(exclude='number'))
df_notnum = df[['c', 'd', 'e']]
print(df_notnum)
print(df_notnum.dtypes)
dtype: object
print(df_notnum.describe())
print(df.describe(include='all'))
print(df.describe(include=int))
print(type(df.describe(include=int)))
pandas.core.frame.DataFrame
print(df.describe(include=[object, bool]))
print(df.describe(exclude=[float, object]))
print(df.count())
dtype: int64
print(df.nunique())
dtype: int64
print(df.mode())
print(df.mode().count())
dtype: int64
print(df.mode().iloc[0])
dtype: object
print(df['c'].value_counts().iat[0])
print(df.apply(lambda x: x.value_counts().iat[0]))
dtype: int64
print(df.mean(numeric_only=True))
dtype: float64
print(df.std(numeric_only=True))
dtype: float64
print(df.min(numeric_only=True))
dtype: float64
print(df.max(numeric_only=True))
dtype: float64
print(df.median(numeric_only=True))
dtype: float64
print(df.quantile(q=[0.25, 0.75], numeric_only=True))
print(df.quantile(q=[0, 0.25, 0.5, 0.75, 1], numeric_only=True))
print(df.describe(percentiles=[0.2, 0.4, 0.6, 0.8]))
print(df.astype('str').describe())
print(df.astype({'a': str}).describe(exclude='number'))
print(df.astype({'d': int, 'e': int}).describe())
s_int = df['a']
print(s_int)
Name: a
dtype: int64
print(s_int.describe())
Name: a
dtype: float64
print(type(s_int.describe()))
pandas.core.series.Series
s_str = df['d']
print(s_str.describe())
Name: d
dtype: object
print(s_str.astype('int').describe())
Name: d
dtype: float64
df['dt'] = pd.to_datetime(['2018-01-01', '2018-03-15', '2018-02-20', '2018-03-15'])
print(df.dtypes)
datetime64[ns]
dtype: object
print(df.describe(include='datetime'))
print(df['dt'].min())
print(df['dt'].max())
print(df.T.describe())
import pandas as pd
df = pd.DataFrame
range(1, 6)
x**2 for x in range(1, 6)
x**3 for x in range(1, 6)
print(df)
print(df.diff())
print(df.diff(1))
print(df.diff(2))
print(df.diff(-1))
print(df.diff(axis=1))
print(df.diff(-1, axis=1))
print(df.diff(2).dropna())
print(df.diff(2).fillna(0))
print(df.diff(2).fillna(method='bfill'))
df['b_diff'] = df['b'].diff(-1)
print(df)
import pandas as pd
df = pd.read_csv('data/src/sample_pandas_normal.csv', index_col=0)
print(df)
print(df.drop('Charlie', axis=0))
print(df.drop('Charlie'))
print(df.drop(index='Charlie'))
print(df.drop(['Bob', 'Dave', 'Frank']))
print(df.drop(index=['Bob', 'Dave', 'Frank']))
df_org = df.copy()
df_org.drop(index=['Bob', 'Dave', 'Frank'], inplace=True)
print(df_org)
print(df.index[[1, 3, 5]])
Index(['Bob', 'Dave', 'Frank'], dtype='object', name='name')
print(df.drop(df.index[[1, 3, 5]]))
print(df.drop(index=df.index[[1, 3, 5]]))
df_noindex = pd.read_csv('data/src/sample_pandas_normal.csv')
print(df_noindex)
print(df_noindex.index)
RangeIndex(start=0, stop=6, step=1)
print(df_noindex.drop([1, 3, 5]))
print(df_noindex.drop(df_noindex.index[[1, 3, 5]]))
df_noindex_sort = df_noindex.sort_values('state')
print(df_noindex_sort)
print(df_noindex_sort.index)
Int64Index([1, 2, 4, 0, 5, 3], dtype='int64')
print(df_noindex_sort.drop([1, 3, 5]))
print(df_noindex_sort.drop(df_noindex_sort.index[[1, 3, 5]]))
print(df.drop('state', axis=1))
print(df.drop(columns='state'))
print(df.drop(['state', 'point'], axis=1))
print(df.drop(columns=['state', 'point']))
df_org = df.copy()
df_org.drop(columns=['state', 'point'], inplace=True)
print(df_org)
print(df.columns[[1, 2]])
Index(['state', 'point'], dtype='object')
print(df.drop(df.columns[[1, 2]], axis=1))
print(df.drop(columns=df.columns[[1, 2]]))
df.drop
index=['Bob', 'Dave', 'Frank']
columns=['state', 'point']
df.drop
index=df.index[[1, 3, 5]]
columns=df.columns[[1, 2]]
df_org = df.copy()
df_org.drop
index=['Bob', 'Dave', 'Frank']
columns=['state', 'point']
inplace=True
print(df_org)
import pandas as pd
import numpy as np
s = pd.Series([0, 1, 2], dtype=np.float64)
print(s.dtype)
s = pd.Series([0, 1, 2], dtype='float64')
print(s.dtype)
s = pd.Series([0, 1, 2], dtype='f8')
print(s.dtype)
s = pd.Series([0, 1, 2], dtype='float')
print(s.dtype)
s = pd.Series([0, 1, 2], dtype=float)
print(s.dtype)
s = pd.Series([0, 1, 2], dtype='uint')
print(s.dtype)
s_object = pd.Series([0, 0.1, 'abc', pd.np.nan])
print(s_object)
dtype: object
print(s_object.map(type))
dtype: object
s_str_constructor = pd.Series([0, 0.1, 'abc', pd.np.nan], dtype=str)
print(s_str_constructor)
dtype: object
print(s_str_constructor.map(type))
dtype: object
s_str_astype = s_object.astype(str)
print(s_str_astype)
dtype: object
print(s_str_astype.map(type))
dtype: object
print(s_str_astype.str.len())
dtype: int64
print(s_object.str.len())
dtype: float64
print(s_object.astype(str).str.len())
dtype: int64
print(s_object)
dtype: object
print(s_object.map(type))
dtype: object
print(s_object.isnull())
dtype: bool
print(s_object.dropna())
dtype: object
print(s_str_astype)
dtype: object
print(s_str_astype.map(type))
dtype: object
print(s_str_astype.isnull())
dtype: bool
print(s_str_astype.dropna())
dtype: object
s_str_astype_nan = s_str_astype.replace('nan', pd.np.nan)
print(s_str_astype_nan)
dtype: object
print(s_str_astype_nan.map(type))
dtype: object
print(s_str_astype_nan.isnull())
dtype: bool
import pandas as pd
df = pd.read_csv('data/src/sample_pandas_normal.csv', index_col=0)
df['f_data'] = [0.1, 0.2, 0.3, 0.4, 0.5, 0.6]
print(df)
print(df.dtypes)
dtype: object
print(type(df.dtypes))
print(type(df.dtypes[0]))
pandas.core.series.Series
numpy.dtype
print(df.dtypes == 'int64')
dtype: bool
print(df.loc[:, df.dtypes == 'int64'])
print(df.dtypes != 'object')
dtype: bool
print(df.loc[:, df.dtypes != 'object'])
print(df.dtypes == 'int')
dtype: bool
print(df.loc[:, df.dtypes == 'int'])
df = df.astype({'age': 'int8'})
print(df.dtypes)
dtype: object
print(df.dtypes == 'int')
dtype: bool
print(df.dtypes.isin(['int8', 'int64']))
dtype: bool
print((df.dtypes == 'int8') | (df.dtypes == 'int64'))
dtype: bool
print(df.loc[:, (df.dtypes == 'int8') | (df.dtypes == 'int64')])
print(df.dtypes.astype('str').str.contains('int'))
dtype: bool
print(df.loc[:, df.dtypes.astype('str').str.contains('int')])
print(df.loc[:, df.dtypes != 'object'])
import pandas as pd
df = pd.read_csv('data/src/sample_pandas_normal.csv')
print(df)
df = df.append({'name': 'Dave', 'age': 68, 'state': 'TX', 'point': 70}, ignore_index=True)
print(df)
print(df.duplicated())
dtype: bool
print(df[df.duplicated()])
print(df.duplicated(keep='last'))
dtype: bool
print(df.duplicated(keep=False))
dtype: bool
print(df.duplicated(subset='state'))
dtype: bool
print(df.duplicated(subset=['state', 'point']))
dtype: bool
print(df.duplicated().sum())
df.duplicated()
dtype: bool
df.duplicated()
sum()
print(df.duplicated().value_counts())
dtype: int64
print(df.duplicated(keep=False).value_counts())
dtype: int64
df.duplicated()
print(df.drop_duplicates())
print(df.drop_duplicates(keep=False))
print(df.drop_duplicates(subset='state'))
df.drop_duplicates(subset='state', keep='last', inplace=True)
print(df)
df = pd.read_csv('data/src/sample_pandas_normal.csv')
print(df)
print(df.groupby('state').mean())
df.groupby('state').agg
lambda x: ','.join(x)
df.groupby('state').agg
NY            [Alice, Frank]   
TX                    [Dave]   
import pandas as pd
df = pd.DataFrame(pd.np.arange(12).reshape(3, 4), columns=['a', 'b', 'c', 'd'], index=['x', 'y', 'z'])
print(df)
print((df > 3) & (df % 2 == 0))
print((df > 3) and (df % 2 == 0))
ValueError: The
a.empty
a.bool()
a.item()
a.any() or a.all()
print(df > 3 & df % 2 == 0)
ValueError: The
a.empty
a.bool()
a.item()
a.any() or a.all()
print(df & 7)
print(df | 1)
print(df << 1)
TypeError: unsupported
type(s) 
print(df << df)
TypeError: unsupported
type(s) 
print(df > 3)
print((df > 3).all())
dtype: bool
print((df > 3).all(axis=1))
dtype: bool
print((df > 3).all(axis=None))
print(df.empty)
df_empty = pd.DataFrame()
print(df_empty.empty)
print(df.size)
print(df_empty.size)
import pandas as pd
print(pd.__version__)
df = pd.DataFrame
index=['apple', 'banana', 'pineapple']
columns=['A', 'B', 'C']
print(df)
print(df.filter(items=['A', 'C']))
print(df.filter(items=['A', 'C'], like='A'))
TypeError: Keyword
print(df.filter(items=['X']))
print(df.filter(items=['apple', 'pineapple'], axis=0))
print(df.filter(items=['apple', 'pineapple'], axis='index'))
print(df.filter(items=['A', 'C'], axis=1))
print(df.filter(items=['A', 'C'], axis='columns'))
print(df.filter(items=['A', 'C']).filter(items=['apple', 'pineapple'], axis=0))
print(df.filter(items=['A', 'C']))
print(df.filter(items=['C', 'A']))
print(df[['C', 'A']])
print(df.loc[:, ['C', 'A']])
print(df.loc[['pineapple', 'apple']])
print(df.filter(like='apple', axis=0))
print(df.filter(regex='e$', axis=0))
print(df.filter(regex='^(a|b)', axis=0))
print(df.filter(regex='(na|ne)', axis=0))
s = pd.Series([0, 1, 2], index=['apple', 'banana', 'pineapple'])
print(s)
dtype: int64
print(s.filter(items=['pineapple', 'banana']))
dtype: int64
print(s.filter(like='apple'))
dtype: int64
print(s.filter(regex='^(a|b)'))
dtype: int64
import pandas as pd
df = pd.DataFrame
index=['Alice', 'Bob']
print(df)
for column_name in df:
print(type(column_name))
print(column_name)
print('======\n')
for column_name in df.__iter__():
print(type(column_name))
print(column_name)
print('======\n')
for column_name, item in df.iteritems():
print(type(column_name))
print(column_name)
print('~~~~~~')
print(type(item))
print(item)
print('------')
print(item['Alice'])
print(item[0])
print(item.Alice)
print('======\n')
pandas.core.series.Series
Name: age
dtype: int64
pandas.core.series.Series
Name: state
dtype: object
pandas.core.series.Series
Name: point
dtype: int64
for index, row in df.iterrows():
print(type(index))
print(index)
print('~~~~~~')
print(type(row))
print(row)
print('------')
print(row['point'])
print(row[2])
print(row.point)
print('======\n')
pandas.core.series.Series
Name: Alice
dtype: object
pandas.core.series.Series
Name: Bob
dtype: object
for row in df.itertuples():
print(type(row))
print(row)
print('------')
print(row[3])
print(row.point)
print('======\n')
pandas.core.frame.Pandas
Pandas(Index='Alice', age=24, state='NY', point=64)
pandas.core.frame.Pandas
Pandas(Index='Bob', age=42, state='CA', point=92)
for row in df.itertuples(name=None):
print(type(row))
print(row)
print('------')
print(row[3])
print('======\n')
print(df['age'])
Name: age
dtype: int64
print(type(df['age']))
pandas.core.series.Series
for age in df['age']:
print(age)
for age, point in zip(df['age'], df['point']):
print(age, point)
print(df.index)
Index(['Alice', 'Bob'], dtype='object')
print(type(df.index))
pandas.core.indexes.base.Index
for index in df.index:
print(index)
for index, state in zip(df.index, df['state']):
print(index, state)
for index, row in df.iterrows():
row['point'] += row['age']
print(df)
for index, row in df.iterrows():
df.at[index, 'point'] += row['age']
print(df)
df['point'] += df['age']
print(df)
df['new'] = df['point'] + df['age'] * 2
print(df)
df['age_sqrt'] = pd.np.sqrt(df['age'])
print(df)
df['state_0'] = df['state'].str.lower().str[0]
print(df)
import pandas as pd
import numpy as np
df = pd.read_csv('data/src/sample_pandas_normal.csv', index_col=0)
df['sex'] = ['female', np.nan, 'male', 'male', 'female', 'male']
df['rank'] = [2, 1, 1, 0, 2, 0]
print(df)
print(pd.get_dummies(df['sex']))
print(pd.get_dummies(['male', 1, 1, 2]))
print(pd.get_dummies(np.arange(6)))
print(pd.get_dummies(np.arange(6).reshape((2, 3))))
Exception: Data
print(pd.get_dummies(df))
print(pd.get_dummies(df, drop_first=True))
print(pd.get_dummies(df, drop_first=True, dummy_na=True))
print(pd.get_dummies(df, drop_first=True, prefix='', prefix_sep=''))
print(pd.get_dummies(df, drop_first=True, prefix=['ST', 'sex'], prefix_sep='-'))
print(pd.get_dummies(df, drop_first=True, prefix={'state': 'ST', 'sex': 'sex'}, prefix_sep='-'))
print(pd.get_dummies(df, drop_first=True, columns=['sex', 'rank']))
df['rank'] = df['rank'].astype(object)
print(pd.get_dummies(df, drop_first=True))
print(df['state'].map({'CA': 0, 'NY': 1, 'TX': 2}))
Name: state
dtype: int64
df['state'] = df['state'].map({'CA': 0, 'NY': 1, 'TX': 2})
print(df)
import pandas as pd
df = pd.read_csv('data/src/sample_pandas_normal.csv', index_col=0)
df_A, df_B = df[:3].copy(), df[3:].copy()
print(df_A)
print(df_B)
print(pd.get_dummies(df_A))
print(pd.get_dummies(df_B))
categories = set(df_A['state'].unique().tolist() + df_B['state'].unique().tolist())
print(categories)
df_A['state'] = pd.Categorical(df_A['state'], categories=categories)
df_B['state'] = pd.Categorical(df_B['state'], categories=categories)
print(df_A['state'].dtypes)
print(pd.get_dummies(df_A))
print(pd.get_dummies(df_B))
df = pd.read_csv('data/src/sample_pandas_normal.csv', index_col=0)
df_train, df_test = df[:3].copy(), df[3:].copy()
categories = df_train['state'].unique()
df_train['state'] = pd.Categorical(df_train['state'], categories=categories)
df_test['state'] = pd.Categorical(df_test['state'], categories=categories)
print(df_test)
print(pd.get_dummies(df_train))
print(pd.get_dummies(df_test))
df = pd.read_csv('data/src/sample_pandas_normal.csv', index_col=0)
df_train, df_test = df[:3].copy(), df[3:].copy()
cols = df_train.select_dtypes('object').columns
for col in cols:
categories = df_train[col].unique()
df_train[col] = pd.Categorical(df_train[col], categories=categories)
df_test[col] = pd.Categorical(df_test[col], categories=categories)
df_train = pd.get_dummies(df_train)
df_test = pd.get_dummies(df_test)
print(df_train)
print(df_test)
import pandas as pd
df = pd.read_csv('data/src/sample_pandas_normal.csv', index_col=0)
print(df)
print(df.index.values)
print(df.columns.values)
print(df.at['Bob', 'age'])
print(df.at['Dave', 'state'])
df.at['Bob', 'age'] = 60
print(df.at['Bob', 'age'])
print(df.iat[1, 0])
print(df.iat[3, 1])
df.iat[1, 0] = 42
print(df.iat[1, 0])
print(df.loc['Bob', 'age'])
print(df.iloc[3, 1])
df.loc['Bob', 'age'] = 60
print(df.loc['Bob', 'age'])
df.iloc[1, 0] = 42
print(df.iloc[1, 0])
print(df.loc['Bob':'Dave', 'age'])
print(type(df.loc['Bob':'Dave', 'age']))
Name: age
dtype: int64
pandas.core.series.Series
print(df.loc[:'Dave', ['age', 'point']])
print(type(df.loc[:'Dave', 'age':'point']))
pandas.core.frame.DataFrame
print(df.iloc[:3, [0, 2]])
print(type(df.iloc[:3, [0, 2]]))
pandas.core.frame.DataFrame
print(df.iloc[::2, 0])
print(type(df.iloc[::2, 0]))
Name: age
dtype: int64
pandas.core.series.Series
print(df.iloc[1::2, 0])
print(type(df.iloc[1::2, 0]))
Name: age
dtype: int64
pandas.core.series.Series
df.loc['Bob':'Dave', 'age'] = [20, 30, 40]
print(df.loc['Bob':'Dave', 'age'])
Name: age
dtype: int64
print(df['Bob':'Ellen'])
print(df[:3])
print(df['age'])
Name: age
dtype: int64
print(df[['age', 'point']])
print(df.loc['Bob'])
print(type(df.loc['Bob']))
Name: Bob
dtype: object
pandas.core.series.Series
print(df.iloc[[1, 4]])
print(type(df.iloc[[1, 4]]))
pandas.core.frame.DataFrame
print(df.loc[:, 'age':'point'])
print(type(df.loc[:, 'age':'point']))
pandas.core.frame.DataFrame
print(df.iloc[:, [0, 2]])
print(type(df.iloc[:, [0, 2]]))
pandas.core.frame.DataFrame
print(df.loc['Bob'])
print(type(df.loc['Bob']))
Name: Bob
dtype: object
pandas.core.series.Series
print(df.loc['Bob':'Bob'])
print(type(df.loc['Bob':'Bob']))
pandas.core.frame.DataFrame
print(df.loc[['Bob']])
print(type(df.loc[['Bob']]))
pandas.core.frame.DataFrame
print(df.iloc[:, 1])
print(type(df.iloc[:, 1]))
Name: state
dtype: object
pandas.core.series.Series
print(df.iloc[:, 1:2])
print(type(df.iloc[:, 1:2]))
pandas.core.frame.DataFrame
print(df.iloc[:, [1]])
print(type(df.iloc[:, [1]]))
pandas.core.frame.DataFrame
df_state = pd.read_csv('data/src/sample_pandas_normal.csv', index_col=2)
print(df_state)
print(df_state.index.values)
print(df_state.at['NY', 'age'])
print(type(df_state.at['NY', 'age']))
numpy.ndarray
print(df_state.loc['NY', 'age'])
print(type(df_state.loc['NY', 'age']))
Name: age
dtype: int64
pandas.core.series.Series
print(df_state.loc['NY', ['age', 'point']])
print(type(df_state.loc['NY', ['age', 'point']]))
pandas.core.frame.DataFrame
print(df_state.iat[0, 1])
print(df_state.index.is_unique)
print(df_state.columns.is_unique)
print(df)
print(df.index[2])
print(df.columns[1])
print(df.at[df.index[2], 'age'])
print(df.loc[['Alice', 'Dave'], df.columns[1]])
Name: state
dtype: object
print(df['age'][2])
print(df.age[2])
print(df.loc[['Alice', 'Dave']].iloc[:, 1])
Name: state
dtype: object
import pandas as pd
import seaborn as sns
import numpy as np
df = sns.load_dataset('iris')
print(df.shape)
print(df.head(5))
df.columns = ['sl', 'sw', 'pl', 'pw', 'species']
print(df.head(5))
grouped = df.groupby('species')
print(grouped)
pandas.core.groupby.generic.DataFrameGroupBy
print(type(grouped))
pandas.core.groupby.generic.DataFrameGroupBy
print(grouped.size())
dtype: int64
print(grouped.mean())
print(grouped.min())
print(grouped.max())
print(grouped.sum())
print(type(grouped.mean()))
pandas.core.frame.DataFrame
print(grouped.agg('mean'))
print(grouped.agg(max))
print(grouped.agg(np.min))
print(grouped.agg(lambda x: max(x) - min(x)))
print(grouped.agg(lambda x: type(x))['sl'])
pandas.core.series.Series
pandas.core.series.Series
pandas.core.series.Series
Name: sl
dtype: object
print(grouped.agg(lambda x: x + 1))
Exception: Must
def my_func(x):
return max(x) - min(x)
print(grouped.agg(my_func))
print(grouped.agg(['mean', max, np.min]))
print(grouped.agg({'sl': 'mean', 'sw': max, 'pl': np.min, 'pw': my_func}))
print(grouped.describe()['sl'])
print(type(grouped.max()))
pandas.core.frame.DataFrame
ax = grouped.max().plot.bar(rot=0)
fig = ax.get_figure()
fig.savefig('data/dst/iris_pandas_groupby_max.jpg')
import pandas as pd
import seaborn as sns
df = sns.load_dataset("iris")
print(df.shape)
print(df.head())
print(df.head(3))
print(df.tail())
print(df.tail(3))
print(df[50:55])
print(df[:5])
print(df[-5:])
print(df.head(1))
print(type(df.head(1)))
pandas.core.frame.DataFrame
print(df.iloc[0])
dtype: object
print(type(df.iloc[0]))
pandas.core.series.Series
print(df.iloc[0]['sepal_length'])
print(df.iloc[-1])
dtype: object
print(type(df.iloc[-1]))
pandas.core.series.Series
print(df.iloc[-1]['sepal_length'])
print(df[0])
print(df[-1])
-1
import pandas as pd
df = pd.DataFrame
index=['a', 'b', 'c', 'd']
print(df)
print(df['col1'])
Name: col1
dtype: int64
print(type(df['col1']))
pandas.core.series.Series
print(df['col1'].max())
print(df['col1'].min())
print(df.max())
dtype: int64
print(df.min())
dtype: int64
print(df.max(axis=1))
dtype: int64
print(df.min(axis=1))
dtype: int64
print(type(df.max()))
pandas.core.series.Series
print(df['col1'].idxmax())
print(df['col1'].idxmin())
print(df['col1'] == df['col1'].max())
Name: col1
dtype: bool
print(df['col1'][df['col1'] == df['col1'].max()])
Name: col1
dtype: int64
print(df['col1'][df['col1'] == df['col1'].max()].index)
Index(['b', 'd'], dtype='object')
print(df['col1'][df['col1'] == df['col1'].max()].index.values)
print(type(df['col1'][df['col1'] == df['col1'].max()].index.values))
numpy.ndarray
print(list(df['col1'][df['col1'] == df['col1'].max()].index))
print(type(list(df['col1'][df['col1'] == df['col1'].max()].index)))
print(df['col1'][df['col1'] == df['col1'].min()].index.values)
print(df.loc['a'])
Name: a
dtype: int64
print(df.loc['a'].idxmax())
print(df.loc['a'].idxmin())
print(df.idxmax())
dtype: object
print(df.idxmin())
dtype: object
print(df.apply(lambda x: list(x[x == x.max()].index)))
col1    [b, d]
col2       [a]
dtype: object
print(df.apply(lambda x: list(x[x == x.min()].index)))
col1    [a]
col2    [b]
dtype: object
print(df.idxmax(axis=1))
dtype: object
print(df.idxmin(axis=1))
dtype: object
print(df.apply(lambda x: list(x[x == x.max()].index), axis=1))
a          [col2]
b          [col1]
c    [col1, col2]
d          [col1]
dtype: object
print(df.apply(lambda x: list(x[x == x.min()].index), axis=1))
a          [col1]
b          [col2]
c    [col1, col2]
d          [col2]
dtype: object
df_nan = df.copy()
df_nan.at['b'] = pd.np.nan
print(df_nan)
print(df_nan.idxmax())
dtype: object
print(df_nan.idxmin())
dtype: object
print(df_nan.idxmax(axis=1))
dtype: object
print(df_nan.idxmin(axis=1))
dtype: object
print(df_nan.idxmax(skipna=False))
dtype: float64
print(df_nan.idxmin(skipna=False))
dtype: float64
print(df_nan.idxmax(axis=1, skipna=False))
dtype: object
print(df_nan.idxmin(axis=1, skipna=False))
dtype: object
print(df_nan['col1'].idxmax())
print(df_nan['col1'].idxmin())
print(df_nan['col1'].idxmax(skipna=False))
print(df_nan['col1'].idxmin(skipna=False))
import pandas as pd
df_mix = pd.DataFrame({'col_int': [0, 1, 2], 'col_float': [0.1, 0.2, 0.3]}, index=['A', 'B', 'C'])
print(df_mix)
print(df_mix.dtypes)
dtype: object
print(df_mix['col_int'] + df_mix['col_float'])
dtype: float64
print(df_mix / 1)
print((df_mix / 1).dtypes)
dtype: object
print(df_mix * 1)
print((df_mix * 1).dtypes)
dtype: object
print(df_mix * 1.0)
print((df_mix * 1.0).dtypes)
dtype: object
print(df_mix.loc['A'])
Name: A
dtype: float64
print(df_mix.T)
print(df_mix.T.dtypes)
dtype: object
df_mix.at['A', 'col_int'] = 10.9
df_mix.at['A', 'col_float'] = 10
print(df_mix)
print(df_mix.dtypes)
dtype: object
df_mix.at['A', 'col_int'] = 'abc'
ValueError: invalid
int() 
import pandas as pd
df_mix = pd.DataFrame({'col_int': [0, 1, 2], 'col_float': [0.1, 0.2, 0.3]}, index=['A', 'B', 'C'])
print(df_mix)
print(df_mix.dtypes)
dtype: object
print(df_mix.loc['B'])
Name: B
dtype: float64
print(type(df_mix.loc['B']))
pandas.core.series.Series
print(df_mix.loc['B']['col_int'])
print(type(df_mix.loc['B']['col_int']))
numpy.float64
print(df_mix.at['B', 'col_int'])
print(type(df_mix.at['B', 'col_int']))
numpy.int64
print(df_mix.loc[['B']])
print(type(df_mix.loc[['B']]))
pandas.core.frame.DataFrame
print(df_mix.loc[['B']].dtypes)
dtype: object
import pandas as pd
df = pd.read_csv('data/src/sample_pandas_normal.csv', index_col=0)
print(df)
print(df['age'])
print(type(df['age']))
Name: age
dtype: int64
pandas.core.series.Series
print(df.age)
print(type(df.age))
Name: age
dtype: int64
pandas.core.series.Series
print(df[['age', 'point']])
print(type(df[['age', 'point']]))
pandas.core.frame.DataFrame
print(df[['age']])
print(type(df[['age']]))
pandas.core.frame.DataFrame
print(df['age':'point'])
print(df.loc[:, 'age':'point'])
print(type(df.loc[:, 'age':'point']))
pandas.core.frame.DataFrame
print(df.iloc[:, [0, 2]])
print(type(df.iloc[:, [0, 2]]))
pandas.core.frame.DataFrame
print(df[1:4])
print(type(df[1:4]))
pandas.core.frame.DataFrame
print(df[:-3])
print(type(df[:-3]))
pandas.core.frame.DataFrame
print(df[::2])
print(type(df[::2]))
pandas.core.frame.DataFrame
print(df[1::2])
print(type(df[1::2]))
pandas.core.frame.DataFrame
print(df[1])
print(df[1:2])
print(type(df[1:2]))
pandas.core.frame.DataFrame
print(df['Bob':'Ellen'])
print(type(df['Bob':'Ellen']))
pandas.core.frame.DataFrame
print(df.loc['Bob'])
print(type(df.loc['Bob']))
Name: Bob
dtype: object
pandas.core.series.Series
print(df.loc[['Bob', 'Ellen']])
print(type(df.loc[['Bob', 'Ellen']]))
pandas.core.frame.DataFrame
print(df.iloc[[1, 4]])
print(type(df.iloc[[1, 4]]))
pandas.core.frame.DataFrame
print(df['age']['Alice'])
print(df['Bob':'Dave'][['age', 'point']])
print(df.at['Alice', 'age'])
print(df.loc['Bob':'Dave', ['age', 'point']])
import pandas as pd
df = pd.read_csv('data/src/sample_pandas_normal.csv', index_col=0)
print(df)
print(df.index)
Index(['Alice', 'Bob', 'Charlie', 'Dave', 'Ellen', 'Frank'], dtype='object', name='name')
print(df.index.str.contains('li'))
print(df[df.index.str.contains('li')])
print(df.index.str.endswith('e'))
print(df[df.index.str.endswith('e')])
print(df.columns)
Index(['age', 'state', 'point'], dtype='object')
print(df.columns.str.endswith('e'))
print(df.loc[:, df.columns.str.endswith('e')])
print(df.iloc[:, df.columns.str.endswith('e')])
print(df.loc[df.index.str.contains('li'), df.columns.str.endswith('e')])
import pandas as pd
df = pd.DataFrame
index=[2, 0, 1]
columns=[1, 2, 0]
print(df)
print(df[0])
dtype: int64
print(df[[0, 2]])
print(df[:2])
print(df[-2:])
print(df.loc[:2])
print(df.iloc[:2])
s = df[2]
print(s)
dtype: int64
print(s[0])
print(s.at[0])
print(s.iat[0])
print(s[-1])
-1
print(s.iat[-1])
import pandas as pd
df = pd.read_csv('data/src/sample_pandas_normal.csv', index_col=0)
print(df)
print(df.query('index.str.contains("li")', engine='python'))
print(df.query('name.str.endswith("e")', engine='python'))
import pandas as pd
df = pd.read_csv('data/src/sample_pandas_normal.csv', index_col=0)
print(df)
s = df['age']
print(s)
Name: age
dtype: int64
print(s[3])
print(type(s[3]))
numpy.int64
print(s['Dave'])
print(type(s['Dave']))
numpy.int64
print(s[-1])
print(type(s[-1]))
numpy.int64
print(s.Dave)
print(type(s.Dave))
numpy.int64
print(s[[1, 3]])
print(type(s[[1, 3]]))
Name: age
dtype: int64
pandas.core.series.Series
print(s[['Bob', 'Dave']])
print(type(s[['Bob', 'Dave']]))
Name: age
dtype: int64
pandas.core.series.Series
print(s[[1]])
print(type(s[[1]]))
Name: age
dtype: int64
pandas.core.series.Series
print(s[['Bob']])
print(type(s[['Bob']]))
Name: age
dtype: int64
pandas.core.series.Series
print(s[1:3])
print(type(s[1:3]))
Name: age
dtype: int64
pandas.core.series.Series
print(s['Bob':'Dave'])
print(type(s['Bob':'Dave']))
Name: age
dtype: int64
pandas.core.series.Series
print(s[1:2])
print(type(s[1:2]))
Name: age
dtype: int64
pandas.core.series.Series
print(s['Bob':'Bob'])
print(type(s['Bob':'Bob']))
Name: age
dtype: int64
pandas.core.series.Series
import pandas as pd
df = pd.DataFrame
pd.np.nan
pd.np.nan
pd.np.nan
pd.np.nan
pd.np.nan
pd.np.nan
pd.np.nan
print(df)
print(df.interpolate())
print(df.interpolate(axis=1))
print(df.interpolate(limit=1))
print(df.interpolate(limit=1, limit_direction='forward'))
print(df.interpolate(limit=1, limit_direction='backward'))
print(df.interpolate(limit=1, limit_direction='both'))
print(df.interpolate(limit_direction='both'))
print(df.interpolate(limit_area='inside'))
print(df.interpolate(limit_area='outside'))
print(df.interpolate(limit_area='outside', limit_direction='both'))
df_copy = df.copy()
df_copy.interpolate(inplace=True)
print(df_copy)
s = pd.Series
pd.np.nan
pd.np.nan
pd.np.nan
pd.np.nan
pd.np.nan
index=[0, 2, 5, 6, 8, 10, 14]
print(s)
dtype: float64
print(s.interpolate())
dtype: float64
print(s.interpolate('index'))
dtype: float64
print(s.interpolate('values'))
dtype: float64
s.index = list('abcdefg')
print(s)
dtype: float64
print(s.interpolate())
dtype: float64
print(s.interpolate('values'))
TypeError: Cannot
dtype('O') 
dtype('float64') 
s = pd.Series
pd.np.nan
pd.np.nan
pd.np.nan
pd.np.nan
index=[0, 2, 5, 6, 8, 10, 14]
print(s.interpolate('spline', order=2))
dtype: float64
s.index = range(7)
print(s.interpolate('spline', order=2))
dtype: float64
s.index = list('abcdefg')
print(s.interpolate('spline', order=2))
TypeError: unsupported
type(s) 
import pandas as pd
print(pd.DataFrame(l_simple))
print(pd.json_normalize(l_simple))
print(pd.DataFrame(l_nested))
print(pd.json_normalize(l_nested))
id.x
id.y
print(pd.json_normalize(l_nested, sep='_'))
print(pd.json_normalize(l_complex))
info.n
info.m
X      [{'a': 1, 'b': 2}, {'a': 3, 'b': 4}]     
Y  [{'a': 10, 'b': 20}, {'a': 30, 'b': 40}]     
print(pd.json_normalize(l_complex, record_path='data'))
print(pd.json_normalize(l_complex, record_path='data', record_prefix='data_'))
pd.json_normalize
l_complex, record_path='data'
meta='label'
pd.json_normalize
l_complex, record_path='data'
meta='label'
meta_prefix='meta_'
pd.json_normalize
l_complex, record_path='data'
meta='info'
pd.json_normalize
l_complex, record_path='data'
meta=[['info', 'n'], ['info', 'm']]
info.n
info.m
pd.json_normalize
l_complex, record_path='data'
meta=[['info', 'n'], ['info', 'm']]
sep='_'
pd.json_normalize
l_complex, record_path='data'
meta=['label', ['info', 'n'], ['info', 'm']]
sep='_'
pd.json_normalize
l_complex, record_path='data'
meta=[['info', 'n']]
info.n
pd.json_normalize
l_complex, record_path='data'
meta=['info', 'n']
errors='ignore'
not always 
import pandas as pd
import json
s = '{"OTHER": "x", "DATA": [{"name":"Alice","age":25},{"name":"Bob","age":42}]}'
print(pd.read_json(s))
d = json.loads(s)
print(d)
print(type(d))
print(pd.json_normalize(d))
print(pd.json_normalize(d, record_path='DATA'))
print(pd.json_normalize(d, record_path='DATA', meta='OTHER'))
print(d['DATA'])
print(type(d['DATA']))
print(pd.DataFrame(d['DATA']))
print(pd.json_normalize(d['DATA']))
import pandas as pd
df = pd.read_csv('data/src/titanic_train.csv')
print(df.head())
Mr. Owen
Mrs. John
Miss. Laina
Mrs. Jacques
Mr. William
df.info()
pandas.core.frame.DataFrame
dtypes: float64
int64(5)
object(5)
print(len(df))
print(len(df.columns))
print(df.shape)
print(df.shape[0])
print(df.shape[1])
row, col = df.shape
print(row)
print(col)
print(df.size)
print(df.shape[0] * df.shape[1])
s = df['PassengerId']
print(s.head())
Name: PassengerId
dtype: int64
print(len(s))
print(s.size)
print(s.shape)
df_multiindex = df.set_index(['Sex', 'Pclass', 'Embarked', 'PassengerId'])
print(len(df_multiindex))
print(len(df_multiindex.columns))
print(df_multiindex.shape)
print(df_multiindex.size)
import pandas as pd
l_1d = [0, 1, 2]
s = pd.Series(l_1d)
print(s)
dtype: int64
s = pd.Series(l_1d, index=['row1', 'row2', 'row3'])
print(s)
dtype: int64
l_2d = [[0, 1, 2], [3, 4, 5]]
df = pd.DataFrame(l_2d)
print(df)
df = pd.DataFrame
index=['row1', 'row2']
columns=['col1', 'col2', 'col3']
print(df)
l_1d_index = [['Alice', 0], ['Bob', 1], ['Charlie', 2]]
index, value = zip(*l_1d_index)
print(index)
print(value)
s_index = pd.Series(value, index=index)
print(s_index)
dtype: int64
l_2d_index = [['Alice', 0, 0.0], ['Bob', 1, 0.1], ['Charlie', 2, 0.2]]
df_index = pd.DataFrame(l_2d_index, columns=['name', 'val1', 'val2'])
print(df_index)
df_index_set = df_index.set_index('name')
print(df_index_set)
print(df_index_set.dtypes)
dtype: object
l_2d_index_columns = [['name', 'val1', 'val2'], ['Alice', 0, 0.0], ['Bob', 1, 0.1], ['Charlie', 2, 0.2]]
df_index_columns = pd.DataFrame(l_2d_index_columns[1:], columns=l_2d_index_columns[0])
print(df_index_columns)
df_index_columns_set = df_index_columns.set_index('name')
print(df_index_columns_set)
s = pd.Series([0, 1, 2])
print(s)
dtype: int64
l_1d = s.values.tolist()
print(l_1d)
df = pd.DataFrame([[0, 1, 2], [3, 4, 5]])
print(df)
l_2d = df.values.tolist()
print(l_2d)
s_index = pd.Series([0, 1, 2], index=['row1', 'row2', 'row3'])
print(s_index)
dtype: int64
l_1d = s_index.values.tolist()
print(l_1d)
df_index = pd.DataFrame
index=['row1', 'row2']
columns=['col1', 'col2', 'col3']
print(df_index)
l_2d = df_index.values.tolist()
print(l_2d)
l_1d_index = s_index.reset_index().values.tolist()
print(l_1d_index)
l_2d_index = df_index.reset_index().values.tolist()
print(l_2d_index)
l_2d_index_columns = df_index.reset_index().T.reset_index().T.values.tolist()
print(l_2d_index_columns)
print(s_index)
dtype: int64
print(s_index.index)
Index(['row1', 'row2', 'row3'], dtype='object')
print(type(s_index.index))
pandas.core.indexes.base.Index
print(s_index.index.tolist())
print(type(s_index.index.tolist()))
for i in s_index.index:
print(i, type(i))
print(s_index.index[0])
print(s_index.index[:2])
Index(['row1', 'row2'], dtype='object')
s_index.index[0] = 'ROW1'
TypeError: Index
not support 
print(df_index)
print(df_index.index)
Index(['row1', 'row2'], dtype='object')
print(df_index.index.tolist())
print(df_index.columns)
Index(['col1', 'col2', 'col3'], dtype='object')
print(df_index.columns.tolist())
import pandas as pd
df = pd.DataFrame
index=['a', 'b', 'c', 'd']
print(df)
df['l'] = df['s'].str.split(',')
print(df)
X         [X]
Y     [XY, Y]
print(df.dtypes)
dtype: object
print(type(df.at['a', 's']))
print(type(df.at['a', 'l']))
print(df['s'].apply(lambda x: [s.strip() for s in x.split(',')]))
b           [X]
c       [XY, Y]
Name: s
dtype: object
print(df['l'].apply(len))
Name: l
dtype: int64
print(df['l'].apply(sorted))
b           [X]
c       [XY, Y]
Name: l
dtype: object
print(df['l'].apply(lambda x: ','.join(x)))
Name: l
dtype: object
print(df['l'].apply(','.join))
Name: l
dtype: object
print(df['l'].apply(lambda x: ','.join(sorted(x))))
Name: l
dtype: object
df['l'].apply(lambda x: x.append('A'))
print(df)
X         [X, A]
df['l'].apply(lambda x: x.remove('Z') if 'Z' in x else x)
print(df)
X      [X, A]
print(df['l'].apply(lambda x: 'X' in x))
Name: l
dtype: bool
print(df[df['l'].apply(lambda x: 'X' in x)])
X     [X, A]
print(df['s'].str.contains('Z'))
Name: s
dtype: bool
print(df[df['s'].str.contains('Z')])
import pandas as pd
import numpy as np
df = pd.read_csv('data/src/sample_header.csv')
print(df)
print(np.sqrt(df))
print(np.amax(df))
dtype: int64
print(np.mean(df, axis=1))
dtype: float64
print(df.max())
dtype: int64
print(df.max(axis=1))
dtype: int64
s = df['a']
print(s)
Name: a
dtype: int64
f_brackets = lambda x: '[{}]'.format(x)
print(s.map(f_brackets))
0    [11]
1    [21]
2    [31]
Name: a
dtype: object
def f_str(x):
return str(x).replace('1', 'One').replace('2', 'Two').replace('3', 'Three').replace('4', 'Four')
print(s.map(f_str))
Name: a
dtype: object
f_oddeven = lambda x: 'odd' if x % 2 == 1 else 'even'
print(df.applymap(f_oddeven))
f_maxmin = lambda x: max(x) - min(x)
print(df.apply(f_maxmin))
dtype: int64
print(df.apply(f_maxmin, axis=1))
dtype: int64
df['b'] = df['b'].map(f_str)
print(df)
df.iloc[2] = df.iloc[2].map(f_str)
print(df)
import pandas as pd
df = pd.read_csv('data/src/sample_pandas_normal.csv')
print(df)
s = df['state']
print(s)
Name: state
dtype: object
s_map_all = s.map({'NY': 'NewYork', 'CA': 'California', 'TX': 'Texas'})
print(s_map_all)
Name: state
dtype: object
s_replace_all = s.replace({'NY': 'NewYork', 'CA': 'California', 'TX': 'Texas'})
print(s_replace_all)
Name: state
dtype: object
s_map = s.map({'NY': 'NewYork'})
print(s_map)
Name: state
dtype: object
s_replace = s.replace({'NY': 'NewYork'})
print(s_replace)
Name: state
dtype: object
s_copy = s.copy()
s_copy.update(s_copy.map({'NY': 'NewYork'}))
print(s_copy)
Name: state
dtype: object
s_copy = s.copy()
s_copy.replace({'NY': 'NewYork'}, inplace=True)
print(s_copy)
Name: state
dtype: object
s_map_num = s.map({'NY': 0, 'CA': 1, 'TX': 2})
print(s_map_num)
Name: state
dtype: int64
df['state'] = df['state'].map({'NY': 0, 'CA': 1, 'TX': 2})
print(df)
print(df['state'].dtype)
import pandas as pd
df = pd.read_csv('data/src/sample_pandas_normal.csv')
s = df['state']
s.map({'NY': 'NewYork', 'CA': 'California', 'TX': 'Texas'})
std. dev. of
s.replace({'NY': 'NewYork', 'CA': 'California', 'TX': 'Texas'})
std. dev. of
s_copy = s.copy()
s_copy.update(s_copy.map({'NY': 'NewYork'}))
std. dev. of
s_copy = s.copy()
s_copy.replace({'NY': 'NewYork'}, inplace=True)
std. dev. of
s_copy = s.copy()
s_copy.update(s_copy.map({'NY': 'NewYork', 'CA': 'California', 'TX': 'Texas'}))
std. dev. of
s_copy = s.copy()
s_copy.replace({'NY': 'NewYork', 'CA': 'California', 'TX': 'Texas'}, inplace=True)
std. dev. of
s.map({'NY': 'NewYork'})
std. dev. of
s.replace({'NY': 'NewYork'})
std. dev. of
df = pd.read_csv('data/src/titanic_train.csv')
s = df['Sex']
s.map({'male': 0, 'female': 1})
std. dev. of
s.replace({'male': 0, 'female': 1})
std. dev. of
import pandas as pd
df = pd.DataFrame
range(5)
i ** 2 for i in range(5)
list('abcde')
print(df)
print(df.median())
dtype: float64
print(type(df.median()))
pandas.core.series.Series
print(df.mean())
dtype: float64
print(df['col_1'].median())
print(type(df['col_1'].median()))
numpy.float64
df_even = pd.DataFrame
range(6)
i ** 2 for i in range(6)
list('abcdef')
print(df_even)
print(df_even.median())
dtype: float64
print(df.median(axis=1))
dtype: float64
import pandas as pd
df_ab = pd.DataFrame({'a': ['a_1', 'a_2', 'a_3'], 'b': ['b_1', 'b_2', 'b_3']})
df_ac = pd.DataFrame({'a': ['a_1', 'a_2', 'a_4'], 'c': ['c_1', 'c_2', 'c_4']})
print(df_ab)
print(df_ac)
print(pd.merge(df_ab, df_ac))
print(df_ab.merge(df_ac))
print(pd.merge(df_ab, df_ac, on='a'))
df_ac_ = df_ac.rename(columns={'a': 'a_'})
print(df_ac_)
print(pd.merge(df_ab, df_ac_, left_on='a', right_on='a_'))
print(pd.merge(df_ab, df_ac_, left_on='a', right_on='a_').drop(columns='a_'))
print(pd.merge(df_ab, df_ac, on='a', how='inner'))
print(pd.merge(df_ab, df_ac, on='a', how='left'))
print(pd.merge(df_ab, df_ac, on='a', how='right'))
print(pd.merge(df_ab, df_ac, on='a', how='outer'))
print(pd.merge(df_ab, df_ac, on='a', how='inner', indicator=True))
print(pd.merge(df_ab, df_ac, on='a', how='outer', indicator=True))
print(pd.merge(df_ab, df_ac, on='a', how='outer', indicator='indicator'))
df_ac_b = df_ac.rename(columns={'c': 'b'})
print(df_ac_b)
print(pd.merge(df_ab, df_ac_b, on='a'))
print(pd.merge(df_ab, df_ac_b, on='a', suffixes=['_left', '_right']))
df_abx = df_ab.assign(x=['x_2', 'x_2', 'x_3'])
df_acx = df_ac.assign(x=['x_1', 'x_2', 'x_2'])
print(df_abx)
print(df_acx)
print(pd.merge(df_abx, df_acx))
print(pd.merge(df_abx, df_acx, on=['a', 'x']))
print(pd.merge(df_abx, df_acx, on='a'))
df_acx_ = df_acx.rename(columns={'x': 'x_'})
print(df_acx_)
print(pd.merge(df_abx, df_acx_, left_on=['a', 'x'], right_on=['a', 'x_']))
print(pd.merge(df_abx, df_acx, on=['a', 'x'], how='inner'))
print(pd.merge(df_abx, df_acx, on=['a', 'x'], how='left'))
print(pd.merge(df_abx, df_acx, on=['a', 'x'], how='right'))
print(pd.merge(df_abx, df_acx, on=['a', 'x'], how='outer'))
print(pd.merge(df_abx, df_acx, on=['a', 'x'], how='outer', sort=True))
df_ac_i = df_ac.set_index('a')
print(df_ac_i)
print(pd.merge(df_ab, df_ac_i, left_on='a', right_index=True))
df_ab_i = df_ab.set_index('a')
print(df_ab_i)
print(pd.merge(df_ab_i, df_ac_i, left_index=True, right_index=True))
print(df_ab_i)
print(df_ac_i)
print(df_ab_i.join(df_ac_i))
print(df_ab_i.join(df_ac_i, how='inner'))
print(df_ab)
print(df_ab.join(df_ac_i, on='a'))
df_ad_i = pd.DataFrame({'a': ['a_1', 'a_4', 'a_5'], 'd': ['d_1', 'd_4', 'd_5']}).set_index('a')
print(df_ad_i)
print(df_ab_i.join([df_ac_i, df_ad_i]))
print(df_ac_i.join([df_ad_i, df_ab_i]))
import pandas as pd
df = pd.read_csv('data/src/sample_pandas_normal.csv', index_col=0)
print(df)
df = df.assign(point_ratio=df['point'] / 100)
df = df.drop(columns='state')
df = df.sort_values('age')
df = df.head(3)
print(df)
df_mc = pd.read_csv('data/src/sample_pandas_normal.csv', index_col=0).assign(point_ratio=df['point'] / 100).drop(columns='state').sort_values('age').head(3)
print(df_mc)
df_mc_break = pd.read_csv
index_col=0
point_ratio=df['point'] / 100
columns='state'
print(df_mc_break)
df_mc_break = pd.read_csv
pandas_normal.csv
index_col=0
point_ratio=df['point'] / 100
columns='state'
SyntaxError: EOL
dfdf_mc_break_mc = pd.read_csv
index_col=0
point_ratio=df['point'] / 100
drop(columns='state').sort_values('age').head(3)
print(df_mc_break)
df_mc_break_backslash = pd.read_csv('data/src/sample_pandas_normal.csv', index_col=0) \

assign(point_ratio=df['point'] / 100) \

drop(columns='state') \

sort_values('age') \

head(3)
print(df_mc_break_backslash)
pd.read_csv('data/src/sample_pandas_normal.csv', index_col=0)
assign(point_ratio=df['point'] / 100)
drop(columns='state')
sort_values('age')
head(3)
print(df_mc_break_parens)
pd.read_csv('data/src/sample_pandas_normal.csv', index_col=0)
assign(point_ratio=df['point'] / 100)
drop(columns='state')
sort_values('age')
head(3)
print(df_mc_break_parens)
pd.read_csv('data/src/sample_pandas_normal.csv', index_col=0)
assign(point_ratio=df['point'] / 100)
drop(columns='state')
sort_values('age')
head(3)
print(df_mc_break_parens)
import pandas as pd
s = pd.Series(['X', 'X', 'Y', 'X'])
print(s)
dtype: object
print(s.mode())
dtype: object
print(type(s.mode()))
pandas.core.series.Series
mode_value = s.mode()[0]
print(mode_value)
print(type(mode_value))
s_same = pd.Series(['X', 'Y', 'Y', 'X'])
print(s_same)
dtype: object
print(s_same.mode())
dtype: object
print(s_same.mode()[0])
l_modes = s_same.mode().tolist()
print(l_modes)
print(type(l_modes))
df = pd.DataFrame
index=['row1', 'row2', 'row3', 'row4']
print(df)
print(df.mode())
print(type(df.mode()))
pandas.core.frame.DataFrame
print(df.mode().count())
dtype: int64
print(df.mode().iloc[0])
dtype: object
print(df.mode()['col1'])
Name: col1
dtype: object
print(df['col1'].mode())
dtype: object
l_mode = df.apply(lambda x: x.mode().tolist())
print(l_mode)
col1       [X]
col2    [X, Y]
dtype: object
print(l_mode.iat[1])
print(type(l_mode.iat[1]))
print(l_mode.iat[1][1])
print(type(l_mode.iat[1][1]))
print(df.mode(axis=1))
print(df.mode(axis=1).count(axis=1))
dtype: int64
df_t = df.T
print(df_t)
print(df_t.mode())
print(df['col1'].value_counts())
Name: col1
dtype: int64
print(df['col1'].value_counts().iat[0])
print(df.apply(lambda x: x.value_counts().iat[0]))
dtype: int64
print(df.describe())
print(df.describe().loc['freq'])
Name: freq
dtype: object
print(df.describe().at['freq', 'col2'])
print(df.T.describe())
import pandas as pd
df = pd.read_csv('data/src/sample_multi.csv', index_col=[0, 1, 2])
print(df)
print(df.index)
levels=[['A0', 'A1', 'A2', 'A3'], ['B0', 'B1', 'B2', 'B3'], ['C0', 'C1', 'C2', 'C3']]
labels=[[0, 0, 0, 0, 1, 1, 1, 1, 2, 2, 2, 2, 3, 3, 3, 3], [0, 0, 1, 1, 2, 2, 3, 3, 0, 0, 1, 1, 2, 2, 3, 3], [0, 1, 2, 3, 0, 1, 2, 3, 0, 1, 2, 3, 0, 1, 2, 3]]
names=['level_1', 'level_2', 'level_3']
print(df.loc['A0', 'val_1'])
Name: val_1
dtype: int64
print(df.loc['A0', :])
print(df.loc['A0'])
print(df.loc['A0':'A2', :])
print(df.loc[['A0', 'A2'], :])
print(df.loc[('A0', 'B1'), :])
print(df.loc[('A0', 'B1', 'C2'), :])
Name: (A0, B1, C2)
dtype: int64
print(df.loc[(['A0', 'A1'], ['B0', 'B3']), :])
df.loc
SyntaxError: invalid
df.loc
SyntaxError: invalid
print(df.loc[(slice(None), 'B1'), :])
print(df.loc[(slice('A1', 'A3'), 'B2'), :])
print(df.loc[(slice('A1', 'A3'), ['B0', 'B2'], 'C1'), :])
print(df.loc[pd.IndexSlice[:, 'B1'], :])
print(df.loc[pd.IndexSlice['A1':'A3', 'B2'], :])
df.loc
pd.IndexSlice
print(df.xs('B1', level='level_2'))
print(df.xs('C1', level=2))
print(df.xs(['B1', 'C2'], level=['level_2', 'level_3']))
print(df.xs(pd.IndexSlice['A1':'A3'], level='level_1'))
print(df.xs(slice('A1', 'A3'), level='level_1'))
print(df.xs(['B1', 'B2'], level='level_2'))
print(df.loc[pd.IndexSlice[:, ['B1', 'B2']], :])
df.loc[(['A0', 'A1'], ['B0', 'B3']), :] = -100
print(df)
df.loc[(['A0', 'A1'], ['B0', 'B3']), :] = [-200, -300]
print(df)
df.loc[(['A0', 'A1'], ['B0', 'B3']), :] = [[-1, -2], [-3, -4], [-5, -6], [-7, -8]]
print(df)
SyntaxError: can
import pandas as pd
df = pd.read_csv('data/src/titanic_train.csv')
df.drop(labels=['Name', 'Ticket', 'Cabin'], axis=1, inplace=True)
df_single = df.set_index('PassengerId')
print(df_single.head())
df_multi = df.set_index(['Sex', 'Pclass', 'Embarked', 'PassengerId']).sort_index()
print(df_multi.head())
print(df_multi.tail())
print(df_multi.mean())
dtype: float64
print(df_single.mean())
dtype: float64
print(df_multi.max())
dtype: float64
print(df_single.max())
dtype: object
print(df_multi.mean(level='Sex'))
print(df_multi.mean(level=0))
print(df_multi.mean(level=1))
print(df_multi.mean(level=2))
print(df_multi.mean(level=['Sex', 'Pclass']))
print(df_multi.mean(level=[0, 1, 2]))
print(df_single.groupby(by='Sex').mean())
print(df_single.groupby(by=['Sex', 'Pclass', 'Embarked']).mean())
print(df_multi.groupby(level='Sex').size())
dtype: int64
print(df_multi.groupby(level=2).size())
dtype: int64
print(df_multi.groupby(level=[0, 1, 2]).size())
dtype: int64
print(df_single.groupby(by=['Sex', 'Pclass', 'Embarked']).size())
dtype: int64
import pandas as pd
df = pd.read_csv('data/src/sample_multi.csv')
print(df)
df_m_csv = pd.read_csv('data/src/sample_multi.csv', index_col=['level_1', 'level_2', 'level_3'])
print(df_m_csv)
df_m = df.set_index(['level_1', 'level_2', 'level_3'])
print(df_m)
df_m_1 = df.set_index('level_1')
print(df_m_1)
df_m_2 = df_m_1.set_index('level_2', append=True)
print(df_m_2)
df_r_all = df_m.reset_index()
print(df_r_all)
df_r_1 = df_m.reset_index(level='level_1')
print(df_r_1)
df_r_1 = df_m.reset_index(level=0)
print(df_r_1)
df_r_2 = df_m.reset_index(level=['level_1', 'level_2'])
print(df_r_2)
df_r_drop = df_m.reset_index(level='level_1', drop=True)
print(df_r_drop)
df_r_drop_sort = df_r_drop.sort_index()
print(df_r_drop_sort)
df_r_drop_sort_2 = df_r_drop.sort_index(level='level_3')
print(df_r_drop_sort_2)
print(df_m)
df_m_swap = df_m.swaplevel('level_1', 'level_3')
print(df_m_swap)
df_m_swap = df_m.swaplevel(0, 2)
print(df_m_swap)
df_m_swap_sort = df_m.swaplevel('level_1', 'level_3').sort_index()
print(df_m_swap_sort)
df_m_change = df_m.reset_index().set_index(['level_2', 'level_3', 'level_1']).sort_index()
print(df_m_change)
import pandas as pd
df = pd.read_csv('data/src/sample_pandas_normal.csv')
print(df)
mask = [True, False, True, False, True, False]
df_mask = df[mask]
print(df_mask)
print(df['age'] < 35)
Name: age
dtype: bool
df['state'] == 'NY'
Name: state
dtype: bool
df['state'] == 'NY'
dtype: bool
df_and = df
df['state'] == 'NY'
print(df_and)
print((df['age'] < 20) | (df['point'] > 90))
dtype: bool
df_or = df[(df['age'] < 20) | (df['point'] > 90)]
print(df_or)
df_multi_1 = df
df['state'] == 'NY'
print(df_multi_1)
df_multi_2 = df
df['state'] == 'NY'
print(df_multi_2)
df_multi_3 = df
df['state'] == 'NY'
print(df_multi_3)
import pandas as pd
df = pd.read_csv('data/src/sample_pandas_normal_nan.csv')
print(df)
print(df.dropna(how='all'))
print(df.dropna(how='all', axis=1))
print(df.dropna(how='all').dropna(how='all', axis=1))
df2 = df.dropna(how='all').dropna(how='all', axis=1)
print(df2)
print(df2.dropna(how='any'))
print(df2.dropna())
print(df2.dropna(how='any', axis=1))
print(df2.dropna(axis=1))
print(df.dropna(thresh=3))
print(df.dropna(thresh=3, axis=1))
print(df.dropna(subset=['age']))
print(df.dropna(subset=['age', 'state']))
print(df.dropna(subset=['age', 'state'], how='all'))
print(df.dropna(subset=[0, 4], axis=1))
print(df.dropna(subset=[0, 4], axis=1, how='all'))
print(df.dropna(subset=['age', 'state', 'xxx']))
print(df.dropna(subset=[0, 4, 10], axis=1))
s = df['age']
print(s)
Name: age
dtype: float64
print(s.dropna())
Name: age
dtype: float64
print(df.fillna(0))
print(df.fillna({'name': 'XXX', 'age': 20, 'ZZZ': 100}))
s_for_fill = pd.Series(['XXX', 20, 100], index=['name', 'age', 'ZZZ'])
print(s_for_fill)
dtype: object
print(df.fillna(s_for_fill))
print(df.mean())
dtype: float64
print(df.fillna(df.mean()))
print(df.fillna(df.median()))
RuntimeWarning: Mean
return np.nanmean(a, axis, out=out, keepdims=keepdims)
print(df.fillna(df.mode().iloc[0]))
print(df.fillna(method='ffill'))
print(df.fillna(method='bfill'))
print(df.fillna(method='bfill', limit=1))
s = df['age']
print(s)
Name: age
dtype: float64
print(s.fillna(100))
Name: age
dtype: float64
print(s.fillna({1: 100, 4: 0}))
Name: age
dtype: float64
print(s.fillna(method='bfill', limit=1))
Name: age
dtype: float64
print(df)
print(df['point'].isnull())
Name: point
dtype: bool
print(df[df['point'].isnull()])
print(df.iloc[2].isnull())
dtype: bool
print(df.loc[:, df.iloc[2].isnull()])
df2 = df.dropna(how='all').dropna(how='all', axis=1)
print(df2)
print(df2.isnull())
print(df2.isnull().any(axis=1))
dtype: bool
print(df2[df2.isnull().any(axis=1)])
print(df2.isnull().any())
dtype: bool
print(df2.loc[:, df2.isnull().any()])
import pandas as pd
df = pd.read_csv('data/src/sample_pandas_normal_nan.csv')
print(df)
print(df.isnull())
print(df.isnull().all())
dtype: bool
print(df.isnull().all(axis=1))
dtype: bool
print(df.isnull().any())
dtype: bool
print(df.isnull().any(axis=1))
dtype: bool
print(df.isnull().sum())
dtype: int64
print(df.isnull().sum(axis=1))
dtype: int64
print(df.count())
dtype: int64
print(df.count(axis=1))
dtype: int64
print(df.isnull().values)
print(type(df.isnull().values))
numpy.ndarray
print(df.isnull().values.sum())
print(df.count().sum())
print(df.isnull().values.sum() != 0)
print(df.size)
print(df.isnull().values.sum() == df.size)
s = df['state']
print(s)
Name: state
dtype: object
print(s.isnull())
Name: state
dtype: bool
print(s.isnull().any())
print(s.isnull().all())
print(s.isnull().sum())
print(s.count())
import pandas as pd
df = pd.read_csv('data/src/sample_pandas_normal_nan.csv')
print(df)
print(df.dtypes)
dtype: object
print(df.at[1, 'name'])
print(type(df.at[1, 'name']))
print(df.at[0, 'point'])
print(type(df.at[0, 'point']))
numpy.float64
print(df.isnull())
import numpy as np
import math
s_nan = pd.Series([None, np.nan, math.nan, float('nan')])
print(s_nan)
dtype: float64
print(s_nan[0])
print(type(s_nan[0]))
numpy.float64
print(s_nan.isnull())
dtype: bool
s_nan_str = pd.Series([None, float('nan'), 'NaN', 'nan'])
print(s_nan_str)
dtype: object
print(s_nan_str[0])
print(type(s_nan_str[0]))
print(s_nan_str.isnull())
dtype: bool
print(s_nan_str.dropna())
dtype: object
print(s_nan_str.fillna(100))
dtype: object
s_nan_str_replace = s_nan_str.replace({'NaN': float('nan'), 'nan': float('nan')})
print(s_nan_str_replace)
dtype: float64
print(s_nan_str_replace.isnull())
dtype: bool
import pandas as pd
import numpy as np
df = pd.DataFrame
index=['A', 'B', 'C', 'D', 'E', 'F']
print(df)
s = df['col1']
print(s)
Name: col1
dtype: int64
print(df.max())
dtype: int64
print(type(df.max()))
pandas.core.series.Series
print(df.min())
dtype: int64
print(type(df.min()))
pandas.core.series.Series
print(df.max(axis=1))
dtype: int64
print(df.min(axis=1))
dtype: int64
print(s.max())
print(type(s.max()))
numpy.int64
print(s.min())
print(type(s.min()))
numpy.int64
print(s.nlargest(4))
Name: col1
dtype: int64
print(type(s.nlargest(4)))
pandas.core.series.Series
print(s.nsmallest(4))
Name: col1
dtype: int64
print(type(s.nsmallest(4)))
pandas.core.series.Series
print(s.nlargest(1))
Name: col1
dtype: int64
print(type(s.nlargest(1)))
pandas.core.series.Series
print(df.nlargest(4, 'col1'))
print(type(df.nlargest(4, 'col1')))
pandas.core.frame.DataFrame
print(df.nsmallest(4, 'col1'))
print(type(df.nsmallest(4, 'col1')))
pandas.core.frame.DataFrame
print(df.nlargest(1, 'col1'))
print(type(df.nlargest(1, 'col1')))
pandas.core.frame.DataFrame
print(df.nlargest(4, ['col1', 'col2']))
print(df.nlargest(4, ['col2', 'col1']))
print(df.nsmallest(4, 'col1'))
print(df.nsmallest(4, 'col1', keep='first'))
print(df.nsmallest(4, 'col1', keep='last'))
print(df.nsmallest(4, 'col1', keep='all'))
print(df.nsmallest(3, ['col1', 'col2'], keep='all'))
print(df.nsmallest(4, ['col1', 'col2'], keep='all'))
print(df['col1'].nsmallest(4).tolist())
print(type(df['col1'].nsmallest(4).tolist()))
print(df['col1'].nsmallest(4).to_numpy())
print(type(df['col1'].nsmallest(4).to_numpy()))
numpy.ndarray
print(df['col1'].nsmallest(4).values)
print(type(df['col1'].nsmallest(4).values))
numpy.ndarray
print(df['col1'].nsmallest(3))
Name: col1
dtype: int64
print(df['col2'].nsmallest(3))
Name: col2
dtype: int64
print([df[col_name].nsmallest(3).tolist() for col_name in df])
print({col_name: col.nsmallest(3).tolist() for col_name, col in df.iteritems()})
print(np.array([df[col_name].nsmallest(3).tolist() for col_name in df]))
print([df[col_name].nsmallest(3, keep='all').tolist() for col_name in df])
print({col_name: col.nsmallest(3, keep='all').tolist() for col_name, col in df.iteritems()})
print(np.array([df[col_name].nsmallest(3, keep='all').tolist() for col_name in df]))
list([1, 2, 3, 3, 3]) 
list([10, 10, 20])
import pandas as pd
import scipy.stats
from sklearn import preprocessing
df = pd.DataFrame
columns=['col1', 'col2', 'col3']
index=['a', 'b', 'c']
print(df)
print((df - df.min()) / (df.max() - df.min()))
print(((df.T - df.T.min()) / (df.T.max() - df.T.min())).T)
print((df - df.values.min()) / (df.values.max() - df.values.min()))
print((df - df.mean()) / df.std())
print((df - df.mean()) / df.std(ddof=0))
print(((df.T - df.T.mean()) / df.T.std()).T)
print(((df.T - df.T.mean()) / df.T.std(ddof=0)).T)
print((df - df.values.mean()) / df.values.std())
print((df - df.values.mean()) / df.values.std(ddof=1))
df_ = df.copy()
s = df_['col1']
df_['col1_min_max'] = (s - s.min()) / (s.max() - s.min())
df_['col1_standardization'] = (s - s.mean()) / s.std()
print(df_)
print(scipy.stats.zscore(df))
print(type(scipy.stats.zscore(df)))
numpy.ndarray
print(scipy.stats.zscore(df, axis=None, ddof=1))
-0.36514837
df_standardization = pd.DataFrame
scipy.stats.zscore(df)
index=df.index
columns=df.columns
print(df_standardization)
df_ = df.copy()
df_['col1_standardization'] = scipy.stats.zscore(df_['col1'])
print(df_)
mm = preprocessing.MinMaxScaler()
print(mm.fit_transform(df))
print(type(mm.fit_transform(df)))
numpy.ndarray
print(preprocessing.minmax_scale(df))
print(type(preprocessing.minmax_scale(df)))
numpy.ndarray
df_min_max = pd.DataFrame
mm.fit_transform(df)
index=df.index
columns=df.columns
print(df_min_max)
df_ = df.copy()
s = df_['col1'].astype(float)
df_['col1_min_max'] = preprocessing.minmax_scale(s)
df_['col1_standardization'] = preprocessing.scale(s)
print(df_)
import pandas as pd
dates = pd.date_range('2018-08-01', '2018-08-31', freq='B')
df = pd.DataFrame({'price': dates.day, 'volume': dates.day * 10}, index=dates)
print(df)
print(df.index)
dtype='datetime64[ns]'
freq='B'
print(df['price'].resample('W').ohlc())
print(df['price'].resample('W-MON', label='left', closed='left').ohlc())
print(df['volume'].resample('W').sum())
Freq: W
-SUN
Name: volume
dtype: int64
print(df['volume'].resample('W-MON', label='left', closed='left').sum())
Freq: W
-MON
Name: volume
dtype: int64
pd.concat
df['price'].resample('W-MON', label='left', closed='left').ohlc()
df['volume'].resample('W-MON', label='left', closed='left').sum()
axis=1
df['price'].resample('W-MON', label='left', closed='left').ohlc()
assign(volume=df['volume'].resample('W-MON', label='left', closed='left').sum())
pd.concat
df['price'].resample('W-MON', label='left', closed='left').ohlc()
df['volume'].resample('W').sum()
axis=1
df['price'].resample('W-MON', label='left', closed='left').ohlc()
assign(volume=df['volume'].resample('W').sum().values)
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import mpl_finance
df_org = pd.read_csv('data/src/aapl_2015_2019.csv', index_col=0, parse_dates=True)['2017']
print(df_org)
df = df_org.copy()
df.index = mdates.date2num(df.index)
data = df.reset_index().values
print(type(data))
numpy.ndarray
print(data.shape)
print(data)
fig = plt.figure(figsize=(12, 4))
ax = fig.add_subplot(1, 1, 1)
mpl_finance.candlestick_ohlc(ax, data, width=2, alpha=0.5, colorup='r', colordown='b')
ax.grid()
locator = mdates.AutoDateLocator()
ax.xaxis.set_major_locator(locator)
ax.xaxis.set_major_formatter(mdates.AutoDateFormatter(locator))
plt.savefig('data/dst/candlestick_day.png')
plt.close()
fig = plt.figure(figsize=(12, 4))
ax = fig.add_subplot(1, 1, 1)
mpl_finance.candlestick_ohlc(ax, data, width=2, alpha=0.5, colorup='r', colordown='b')
ax.grid()
ax.xaxis.set_major_locator(mdates.MonthLocator())
ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y/%m'))
plt.savefig('data/dst/candlestick_day_format.png')
plt.close()
df_w = df_org.resample('W-MON', closed='left', label='left').agg(d_ohlcv)
df_w.index = mdates.date2num(df_w.index)
data_w = df_w.reset_index().values
fig = plt.figure(figsize=(12, 4))
ax = fig.add_subplot(1, 1, 1)
mpl_finance.candlestick_ohlc(ax, data_w, width=4, alpha=0.75, colorup='r', colordown='b')
ax.grid()
locator = mdates.AutoDateLocator()
ax.xaxis.set_major_locator(locator)
ax.xaxis.set_major_formatter(mdates.AutoDateFormatter(locator))
plt.savefig('data/dst/candlestick_week.png')
plt.close()
fig = plt.figure(figsize=(12, 4))
ax = fig.add_subplot(1, 1, 1)
mpl_finance.candlestick_ohlc(ax, data_w, width=4, alpha=0.75, colorup='r', colordown='b')
ax.plot(df_w.index, df_w['close'].rolling(4).mean())
ax.plot(df_w.index, df_w['close'].rolling(13).mean())
ax.plot(df_w.index, df_w['close'].rolling(26).mean())
ax.grid()
locator = mdates.AutoDateLocator()
ax.xaxis.set_major_locator(locator)
ax.xaxis.set_major_formatter(mdates.AutoDateFormatter(locator))
plt.savefig('data/dst/candlestick_week_sma.png')
plt.close()
fig, axes = plt.subplots
nrows=2
ncols=1
figsize=(12, 4)
sharex=True
gridspec_kw={'height_ratios': [3, 1]}
mpl_finance.candlestick_ohlc(axes[0], data_w, width=4, alpha=0.75, colorup='r', colordown='b')
axes[1].bar(df_w.index, df_w['volume'], width=4, color='navy')
axes[0].grid()
axes[1].grid()
locator = mdates.AutoDateLocator()
axes[0].xaxis.set_major_locator(locator)
axes[0].xaxis.set_major_formatter(mdates.AutoDateFormatter(locator))
plt.savefig('data/dst/candlestick_week_v.png')
plt.close()
fig, axes = plt.subplots
nrows=2
ncols=1
figsize=(12, 4)
sharex=True
gridspec_kw={'height_ratios': [3, 1]}
mpl_finance.candlestick_ohlc(axes[0], data_w, width=4, alpha=0.75, colorup='r', colordown='b')
axes[0].plot(df_w.index, df_w['close'].rolling(4).mean())
axes[0].plot(df_w.index, df_w['close'].rolling(13).mean())
axes[0].plot(df_w.index, df_w['close'].rolling(26).mean())
axes[1].bar(df_w.index, df_w['volume'], width=4, color='navy')
axes[0].grid()
axes[1].grid()
locator = mdates.AutoDateLocator()
axes[0].xaxis.set_major_locator(locator)
axes[0].xaxis.set_major_formatter(mdates.AutoDateFormatter(locator))
plt.savefig('data/dst/candlestick_week_sma_v.png')
plt.close()
import pandas as pd
import mplfinance as mpf
df = pd.read_csv('data/src/aapl_2015_2019.csv', index_col=0, parse_dates=True)['2017']
print(df)
df.columns = ['Open', 'High', 'Low', 'Close', 'Volume']
mpf.plot(df, savefig='data/dst/candlestick_mpf.png')
mpf.plot
df[50:100], figratio=(12,4)
savefig='data/dst/candlestick_mpf_figratio.png'
mpf.plot
df[50:100], type='candle', figratio=(12,4)
savefig='data/dst/candlestick_mpf_candle.png'
mpf.plot
df[50:100], type='candle', volume=True, figratio=(12,4)
savefig='data/dst/candlestick_mpf_volume.png'
mpf.plot
df[50:100], type='candle', volume=True, mav=(5, 25), figratio=(12,4)
savefig='data/dst/candlestick_mpf_mav.png'
mpf.plot
df[50:100], type='candle', figratio=(12,4)
volume=True
mav=(5, 25)
style='yahoo'
savefig='data/dst/candlestick_mpf_style_yahoo.png'
df_w = df.resample('W-MON', closed='left', label='left').agg(d_ohlcv)
print(df_w.head())
mpf.plot
df_w, type='candle', figratio=(12,4)
volume=True
mav=(5, 25)
savefig='data/dst/candlestick_mpf_week.png'
import pandas as pd
df = pd.read_csv('data/src/aapl_2015_2019.csv', index_col=0, parse_dates=True)['2017']
print(df)
print(df.resample('MS').agg(d_ohlc))
print(df.resample('QS').agg(d_ohlc))
print(df.resample('2W-MON', closed='left', label='left').agg(d_ohlc))
print(df.resample('MS').agg(d_ohlcv))
import pandas as pd
import numpy as np
print(pd.__version__)
print(pd.options.display.precision)
s_decimal = pd.Series([123.456, 12.3456, 1.23456, 0.123456, 0.0123456, 0.00123456])
print(s_decimal)
dtype: float64
print(s_decimal[5])
pd.options.display.precision = 4
print(s_decimal)
dtype: float64
pd.options.display.precision = 2
print(s_decimal)
dtype: float64
print(pd.options.display.float_format)
pd.options.display.float_format = '{:.2f}'.format
print(s_decimal)
dtype: float64
pd.options.display.float_format = '{:.4g}'.format
print(s_decimal)
dtype: float64
pd.options.display.float_format = '{:.4e}'.format
print(s_decimal)
dtype: float64
pd.options.display.float_format = '{: <10.2%}'.format
print(s_decimal)
dtype: float64
df_decimal = pd.DataFrame
pd.options.display.float_format = '{:.0f}'.format
print(df_decimal)
df_decimal2 = pd.DataFrame
pd.options.display.float_format = '{:.1f}'.format
print(df_decimal2)
print(pd.options.display.max_rows)
df_tall = pd.DataFrame(np.arange(300).reshape((100, 3)))
pd.options.display.max_rows = 10
print(df_tall)
print(df_tall.head(10))
print(df_tall.head(20))
pd.options.display.max_rows = None
print(pd.options.display.max_columns)
df_wide = pd.DataFrame(np.arange(90).reshape((3, 30)))
print(df_wide)
pd.options.display.max_columns = 10
print(df_wide)
pd.options.display.max_columns = None
print(df_wide)
print(pd.options.display.show_dimensions)
pd.options.display.max_columns = 10
print(df_wide)
df = pd.DataFrame(np.arange(12).reshape((3, 4)))
print(df)
pd.options.display.show_dimensions = True
print(df_wide)
print(df)
pd.options.display.show_dimensions = False
print(df_wide)
print(df)
print(pd.options.display.width)
pd.options.display.max_columns = None
print(df_wide)
pd.options.display.width = 60
print(df_wide)
pd.options.display.width = None
print(df_wide)
print(pd.options.display.max_colwidth)
df_long_col = pd.DataFrame({'col': ['a' * 10, 'a' * 30, 'a' * 60]})
print(df_long_col)
pd.options.display.max_colwidth = 80
print(df_long_col)
df_long_col2 = pd.DataFrame
pd.options.display.max_colwidth = 20
print(df_long_col2)
df_long_col_header = pd.DataFrame
pd.options.display.max_colwidth = 40
print(df_long_col_header)
print(pd.options.display.colheader_justify)
print(df_long_col)
pd.options.display.colheader_justify = 'left'
print(df_long_col)
import pandas as pd
import pprint
print(pd.__version__)
print(pd.options.display.max_rows)
pd.options.display.max_rows = 100
print(pd.options.display.max_rows)
print(dir(pd.options))
pprint.pprint(dir(pd.options.display))
pd.describe_option()
compute.use_bottleneck
values: False
default: True
currently: True
compute.use_numexpr
values: False
default: True
currently: True
display.chop_threshold
float or None
repr and friends
default: None
currently: None
display.colheader_justify
headers. used
default: right
currently: right
display.column_space
display.date_dayfirst
prints and parses
default: False
currently: False
display.date_yearfirst
prints and parses
default: False
currently: False
display.encoding
returned
default: UTF
-8
currently: UTF
-8
display.expand_frame_repr
display.width
default: True
currently: True
display.float_format
number and return
formats.format.EngFormatter
default: None
currently: None
display.html.border
border=value
display.html.table_schema
default: False
default: False
currently: False
display.html.use_mathjax
default: True
default: True
currently: True
display.large_repr
df.info() 
default: truncate
currently: truncate
display.latex.escape
values: False
default: True
currently: True
display.latex.longtable
values: False
default: False
currently: False
display.latex.multicolumn
values: False
default: True
currently: True
display.latex.multicolumn_format
values: False
default: l
currently: l
display.latex.multirow
values: False
default: False
currently: False
display.latex.repr
default: False
default: False
currently: False
display.max_categories
display.max_columns
view. Depending
truncated or printed
0 and pandas
terminal and print
width. The
not run in a 
terminal and hence
display.max_colwidth
structure. When
display.max_info_columns
display.max_info_rows
int or None
df.info() 
slow. max_info_rows and max_info_cols
display.max_rows
view. Depending
truncated or printed
0 and pandas
terminal and print
height. The
not run in a 
terminal and hence
display.max_seq_items
int or None
printed. If
display.memory_usage
string or None
default: True
currently: True
display.multi_sparse
default: True
currently: True
display.notebook_repr_html
default: True
currently: True
display.pprint_nest_depth
display.precision
display.show_dimensions
boolean or 'truncate'
e.g. not
default: truncate
currently: truncate
display.unicode.ambiguous_as_wide
default: False
default: False
currently: False
display.unicode.east_asian_width
default: False
default: False
currently: False
display.width
None and pandas
not run in a
terminal and hence
html.border
border=value
display.html.border
io.excel.xls.writer
files. Available
default: auto
currently: auto
io.excel.xlsm.writer
files. Available
default: auto
currently: auto
io.excel.xlsx.writer
files. Available
default: auto
currently: auto
io.hdf.default_format
default: None
currently: None
io.hdf.dropna_table
default: False
currently: False
io.parquet.engine
engine. Available
default: auto
currently: auto
mode.chained_assignment
default: warn
currently: warn
mode.sim_interactive
default: False
currently: False
mode.use_inf_as_na
-INF
None and NaN
-INF
not NA
default: False
currently: False
mode.use_inf_as_null
deprecated and will
version. Use
default: False
currently: False
mode.use_inf_as_na
plotting.matplotlib.register_converters
Periods. Toggling
default: True
currently: True
pd.describe_option('compute')
compute.use_bottleneck
values: False
default: True
currently: True
compute.use_numexpr
values: False
default: True
currently: True
pd.describe_option('max_col')
display.max_columns
view. Depending
truncated or printed
0 and pandas
terminal and print
width. The
not run in a 
terminal and hence
display.max_colwidth
structure. When
pd.describe_option('max.*col')
display.max_columns
view. Depending
truncated or printed
0 and pandas
terminal and print
width. The
not run in a 
terminal and hence
display.max_colwidth
structure. When
display.max_info_columns
print(pd.get_option('display.max_rows'))
pd.set_option('display.max_rows', 60)
print(pd.get_option('max_r'))
pd.set_option('max_r', 100)
pd.get_option('max')
pd.set_option('max', 60)
l = ['display.max_rows', 'display.max_columns', 'display.max_colwidth']
print([pd.get_option(i) for i in l])
print({i: pd.get_option(i) for i in l})
display.max_rows
display.max_columns
display.max_colwidth
display.max_rows
display.max_columns
display.max_colwidth
pd.set_option(k, v) for k, v in d.items()
print({i: pd.get_option(i) for i in d.keys()})
display.max_rows
display.max_columns
display.max_colwidth
print(pd.options.display.max_rows)
pd.reset_option('display.max_rows')
print(pd.options.display.max_rows)
print(pd.options.display.max_columns)
print(pd.options.display.max_colwidth)
pd.reset_option('max_col')
print(pd.options.display.max_columns)
print(pd.options.display.max_colwidth)
pd.options.display.max_rows = 100
pd.options.display.max_columns = 100
pd.options.display.max_colwidth = 100
pd.reset_option('^display', silent=True)
print(pd.options.display.max_rows)
print(pd.options.display.max_columns)
print(pd.options.display.max_colwidth)
pd.reset_option('all')
html.border
display.html.border
deprecated and will
version. Use
FutureWarning: html
display.html.border
warnings.warn(d.msg, FutureWarning)
deprecated and will
version. Use
warnings.warn(d.msg, FutureWarning)
pd.reset_option('all', silent=True)
pd.option_context('display.max_rows', 100)
print(pd.options.display.max_rows)
print(pd.options.display.max_rows)
pd.options.display.max_rows = 80
pd.option_context('display.max_rows', 100)
print(pd.options.display.max_rows)
print(pd.options.display.max_rows)
pd.option_context('display.max_rows', 100, 'display.max_columns', 100)
print(pd.options.display.max_rows)
print(pd.options.display.max_columns)
print(pd.options.display.max_rows)
print(pd.options.display.max_columns)
pd.option_context('display.max_rows', 100)
print(pd.options.display.max_rows)
import pandas as pd
df = pd.DataFrame
range(1, 6)
x**2 for x in range(1, 6)
x**3 for x in range(1, 6)
print(df)
print(df.pct_change())
print(df.pct_change(2))
print(df.pct_change(-1))
print(df.pct_change(axis=1))
print(df.pct_change(-1, axis=1))
print(df.pct_change(2).dropna())
print(df.pct_change(2).fillna(0))
print(df.pct_change(2).fillna(method='bfill'))
df['b_pct_change'] = df['b'].pct_change(-1)
print(df)
import pandas as pd
print(pd)
print(pd.np)
import numpy as np
print(pd.np is np)
print(pd.np.mean is np.mean)
a = pd.np.arange(12).reshape(3, 4)
print(a)
print(type(a))
numpy.ndarray
print(pd.np.pi)
print(pd.np.e)
import numpy as np
import pandas as pd
df = pd.read_csv('data/src/titanic_train.csv', index_col=0).drop(['Name', 'Ticket', 'SibSp', 'Parch'], axis=1)
print(df.head())
print(pd.pivot_table(df, index='Pclass', columns='Sex'))
print(type(pd.pivot_table(df, index='Pclass', columns='Sex')))
pandas.core.frame.DataFrame
print(pd.pivot_table(df, index='Pclass', columns='Sex', values='Age'))
print(pd.pivot_table(df, index=['Sex', 'Pclass'], columns='Survived', values=['Age', 'Fare']))
print(pd.pivot_table(df, index='Sex', columns='Pclass', values='Age', margins=True))
pd.pivot_table
df, index='Sex', columns='Pclass', values='Age'
margins=True
margins_name='Total'
pd.pivot_table
df, index='Sex', columns='Pclass', values='Age'
margins=True
aggfunc=np.min
pd.pivot_table
df, index='Sex', columns='Pclass', values='Age'
margins=True
aggfunc=[np.min, np.max]
pd.pivot_table
df, index='Sex', columns='Pclass', values='Age'
margins=True
aggfunc=len
print(len(df))
print(df.isnull().sum())
dtype: int64
pd.pivot_table
df, index='Sex', columns='Pclass', values='Age'
margins=True
aggfunc=len
dropna=False
import pandas as pd
import matplotlib as mpl
import matplotlib.pyplot as plt
df = pd.read_csv('data/src/iris.csv', index_col=0)
print(df.head())
plt.figure()
df.plot()
plt.savefig('data/dst/pandas_iris_line.png')
plt.close('all')
print(type(df.plot()))
matplotlib.axes._subplots.AxesSubplot
fig = plt.figure()
ax = fig.add_subplot(1, 1, 1)
df.plot(ax=ax)
fig.savefig('data/dst/pandas_iris_line.png')
fig, axes = plt.subplots(nrows=2, ncols=3, figsize=(9, 6))
df.plot(ax=axes[0, 0], legend=False)
df.plot(ax=axes[1, 2], legend=False, kind='hist')
fig.savefig('data/dst/pandas_iris_line_axes.png')
current_figsize = mpl.rcParams['figure.figsize']
print(current_figsize)
plt.figure()
df.plot(figsize=(9, 6))
plt.savefig('data/dst/pandas_iris_line_figsize.png')
plt.close('all')
current_dpi = mpl.rcParams['figure.dpi']
print(current_dpi)
plt.figure()
df.plot()
plt.savefig('data/dst/pandas_iris_line_dpi.png', dpi=current_dpi * 1.5)
plt.close('all')
plt.figure()
df.plot(subplots=True)
plt.savefig('data/dst/pandas_iris_line_subplots.png')
plt.close('all')
print(type(df.plot(subplots=True)))
print(type(df.plot(subplots=True)[0]))
numpy.ndarray
matplotlib.axes._subplots.AxesSubplot
plt.figure()
df.plot(subplots=True, layout=(2, 2))
plt.savefig('data/dst/pandas_iris_line_subplots_layout.png')
plt.close('all')
plt.figure()
df.plot
subplots=True
layout=(2, 2)
sharex=True
sharey=True
plt.savefig('data/dst/pandas_iris_line_subplots_share.png')
plt.close('all')
plt.figure()
df.plot(x='sepal_length', y='sepal_width')
plt.savefig('data/dst/pandas_iris_line_xy.png')
plt.close('all')
plt.figure()
df.plot(x='sepal_length')
plt.savefig('data/dst/pandas_iris_line_x.png')
plt.close('all')
plt.figure()
df.plot(y='sepal_length')
plt.savefig('data/dst/pandas_iris_line_y.png')
plt.close('all')
df.plot(y=['sepal_length', 'sepal_width'])
UserWarning: Pandas
plt.figure()
ax = df.plot(y='sepal_length')
df.plot(y='sepal_width', ax=ax)
plt.savefig('data/dst/pandas_iris_line_multi.png')
plt.close('all')
plt.figure()
df[['sepal_length', 'sepal_width']].plot()
plt.savefig('data/dst/pandas_iris_line_multi.png')
plt.close('all')
plt.figure()
df.plot
title='Iris Data Set'
grid=True
colormap='Accent'
legend=False
alpha=0.5
plt.savefig('data/dst/pandas_iris_line_etc.png')
plt.close('all')
plt.figure()
df.plot(kind='line')
plt.savefig('data/dst/pandas_iris_line.png')
plt.close('all')
plt.figure()
df.plot.line()
plt.savefig('data/dst/pandas_iris_line.png')
plt.close('all')
plt.figure()
df.plot.line(subplots=True, layout=(2, 2))
plt.savefig('data/dst/pandas_iris_line_subplots_layout.png')
plt.close('all')
plt.figure()
df.plot.line(style=['r--', 'b.-', 'g+', 'k:'])
plt.savefig('data/dst/pandas_iris_line_style.png')
plt.close('all')
plt.figure()
df[:5].plot.bar()
plt.savefig('data/dst/pandas_iris_bar.png')
plt.close('all')
plt.figure()
df[:5].plot.barh()
plt.savefig('data/dst/pandas_iris_barh.png')
plt.close('all')
plt.figure()
df[:5].plot.bar(stacked=True)
plt.savefig('data/dst/pandas_iris_bar_stack.png')
plt.close('all')
plt.figure()
df.plot.box()
plt.savefig('data/dst/pandas_iris_box.png')
plt.close('all')
plt.figure()
df.plot.hist()
plt.savefig('data/dst/pandas_iris_hist.png')
plt.close('all')
plt.figure()
df.plot.hist(alpha=0.5)
plt.savefig('data/dst/pandas_iris_hist_alpha.png')
plt.close('all')
plt.figure()
df.plot.hist(stacked=True)
plt.savefig('data/dst/pandas_iris_stacked.png')
plt.close('all')
plt.figure()
df.plot.hist(bins=20, histtype='step', orientation='horizontal')
plt.savefig('data/dst/pandas_iris_hist_h_step.png')
plt.close('all')
plt.figure()
df.plot.kde(style=['r-', 'g--', 'b-.', 'c:'])
plt.savefig('data/dst/pandas_iris_kde.png')
plt.close('all')
plt.figure()
df.plot.area()
plt.savefig('data/dst/pandas_iris_area.png')
plt.close('all')
plt.figure()
df.plot.scatter(x='sepal_length', y='petal_length', alpha=0.5)
plt.savefig('data/dst/pandas_iris_scatter.png')
plt.close('all')
plt.figure()
ax = df.plot.scatter(x='sepal_length', y='petal_length', alpha=0.5)
df.plot.scatter
x='sepal_length'
y='petal_width'
marker='s'
c='r'
s=50
alpha=0.5
ax=ax
plt.savefig('data/dst/pandas_iris_scatter_multi.png')
plt.close('all')
plt.figure()
df.plot.line
x='sepal_length'
style=['ro', 'g+', 'bs']
alpha=0.5
plt.savefig('data/dst/pandas_iris_scatter_line.png')
plt.close('all')
plt.figure()
df.plot.hexbin
x='sepal_length'
y='petal_length'
gridsize=15
sharex=False
plt.savefig('data/dst/pandas_iris_hexbin.png')
plt.close('all')
df_pie = pd.DataFrame
index=['a', 'b', 'c']
columns=['ONE', 'TWO']
print(df_pie)
plt.figure()
df_pie.plot.pie(subplots=True)
plt.savefig('data/dst/pandas_pie.png')
plt.close('all')
print(type(df_pie.plot.pie(subplots=True)))
print(type(df_pie.plot.pie(subplots=True)[0]))
numpy.ndarray
matplotlib.axes._subplots.AxesSubplot
plt.figure()
df_pie['ONE'].plot.pie()
plt.savefig('data/dst/pandas_pie_single.png')
plt.close('all')
import pandas as pd
df = pd.DataFrame
range(11)
i ** 2 for i in range(11)
list('abcdefghijk')
print(df)
print(df.quantile())
dtype: float64
print(type(df.quantile()))
pandas.core.series.Series
print(df['col_1'].quantile())
print(type(df['col_1'].quantile()))
print(df.quantile(0.2))
dtype: float64
print(df.quantile([0, 0.25, 0.5, 0.75, 1.0]))
print(type(df.quantile([0, 0.25, 0.5, 0.75, 1.0])))
pandas.core.frame.DataFrame
print(df['col_1'].quantile([0, 0.25, 0.5, 0.75, 1.0]))
Name: col_1
dtype: float64
print(type(df['col_1'].quantile([0, 0.25, 0.5, 0.75, 1.0])))
pandas.core.series.Series
print(df.quantile(0.21))
dtype: float64
print(df.quantile(0.21, interpolation='linear'))
dtype: float64
print(df.quantile(0.21, interpolation='lower'))
dtype: int64
print(df.quantile(0.21, interpolation='higher'))
dtype: int64
print(df.quantile(0.21, interpolation='nearest'))
dtype: int64
print(df.quantile(0.21, interpolation='midpoint'))
dtype: float64
print(df.quantile(0.2))
dtype: float64
print(df.quantile(0.2, interpolation='lower'))
dtype: int64
print(df.quantile(axis=1))
dtype: float64
print(df.quantile(numeric_only=False))
TypeError: can
print(df.quantile(numeric_only=False, interpolation='lower'))
dtype: object
print(df.quantile(0.25, numeric_only=False, interpolation='lower'))
dtype: object
print(df.quantile(0.25, numeric_only=False, interpolation='higher'))
dtype: object
df['col_3'] = pd.date_range('2019-01-01', '2019-01-11')
print(df)
print(df.dtypes)
datetime64[ns]
dtype: object
print(df.quantile())
dtype: float64
print(df.quantile(numeric_only=False))
dtype: object
print(df.quantile(0.25, numeric_only=False))
dtype: object
print(df.quantile(0.25, numeric_only=False, interpolation='lower'))
dtype: object
import pandas as pd
df = pd.read_csv('data/src/sample_pandas_normal.csv')
print(df)
print(df[df['age'] < 25])
print(df.query('age < 25'))
print(df.query('not age < 25'))
print(df.query('24 <= age < 50'))
print(df.query('age < point / 3'))
print(df.query('state == "CA"'))
print(df.query('state != "CA"'))
print(df[df['state'].isin(['NY', 'TX'])])
print(df.query('state in ["NY", "TX"]'))
print(df.query('state == ["NY", "TX"]'))
print(df.query('name.str.endswith("e")', engine='python'))
print(df.query('name.str.contains("li")', engine='python'))
print(df.query('name.str.match(".*i.*e")', engine='python'))
print(df.query('age.astype("str").str.endswith("8")', engine='python'))
df.at[0, 'name'] = None
print(df)
print(df.query('name.str.endswith("e")', engine='python'))
ValueError: cannot
print(df[df['name'].str.endswith('e', na=False)])
print(df.query('name.str.endswith("e", na=False)', engine='python'))
df['name'].fillna('Alice', inplace=True)
print(df)
print(df.query('index % 2 == 0'))
df_name = df.set_index('name')
print(df_name)
print(df_name.query('name.str.endswith("e")', engine='python'))
print(df_name.query('index.str.endswith("e")', engine='python'))
val = 80
print(df.query('point > @val'))
print(df[(df['age'] < 25) & (df['point'] > 65)])
print(df.query('age < 25 & point > 65'))
print(df.query('age < 25 and point > 65'))
print(df.query('age < 25 | point > 65'))
print(df.query('age < 25 or point > 65'))
print(df.query('not age < 25 and not point > 65'))
print(df.query('age == 24 | point > 80 & state == "CA"'))
print(df.query('(age == 24 | point > 80) & state == "CA"'))
df.columns = ['名前', 'age.year', 'state name', 3]
print(df)
age.year
print(df.query('名前 == ["Alice", "Dave"]'))
age.year
print(df.query('名前.str.contains("li")', engine='python'))
age.year
print(df.query('age.year < 25'))
UndefinedVariableError: name
not defined
print(df.query('state name == "CA"'))
SyntaxError: invalid
print(df.query('3 > 75'))
KeyError: False
print(df[df['age.year'] < 25])
age.year
print(df[df['state name'] == 'CA'])
age.year
print(df[df[3] > 75])
age.year
df.rename(columns={3: 'point'}, inplace=True)
print(df)
age.year
df.columns = [str(s).replace(' ', '_').replace('.', '_') for s in df.columns]
print(df)
df.query('age_year > 25', inplace=True)
print(df)
import pandas as pd
df = pd.read_csv('data/src/sample_pandas_normal.csv')
print(df)
df_s = df.sample(frac=1)
print(df_s)
df_s = df.sample(frac=1, random_state=0)
print(df_s)
df_s = df.sample(frac=1).reset_index(drop=True)
print(df_s)
df = df.sample(frac=1).reset_index(drop=True)
print(df)
import pandas as pd
df = pd.DataFrame
pd.np.nan
pd.np.nan
index=['a', 'b', 'c', 'd']
print(df)
print(df.rank())
print(df.rank(axis=1))
print(df.rank(numeric_only=True))
print(df.rank(axis=1, numeric_only=False))
not supported 
print(df.rank(ascending=False))
print(df.rank(method='average'))
print(df.rank(method='min'))
print(df.rank(method='max'))
print(df.rank(method='first'))
ValueError: first
not supported 
print(df.rank(method='first', numeric_only=True))
print(df.rank(method='dense'))
print(df.rank(na_option='keep'))
print(df.rank(na_option='top'))
print(df.rank(na_option='top', method='min'))
print(df.rank(na_option='bottom'))
print(df.rank(na_option='bottom', method='min'))
print(df.rank(pct=True))
print(df.rank(pct=True, method='min', ascending=False, na_option='bottom'))
print(df['col1'].rank(method='min', ascending=False))
Name: col1
dtype: float64
import pandas as pd
df = pd.read_clipboard()
print(df)
df.to_csv('data/dst/test.csv')
import pandas as pd
df = pd.read_csv('data/src/sample_header_index_dtype.csv', index_col=0)
print(df)
print(df.dtypes)
dtype: object
print(df.applymap(type))
df_str = pd.read_csv
index_col=0
dtype=str
print(df_str)
print(df_str.dtypes)
dtype: object
print(df_str.applymap(type))
df_object = pd.read_csv
index_col=0
dtype=object
print(df_object)
print(df_object.dtypes)
dtype: object
print(df_object.applymap(type))
print(df.astype(str).applymap(type))
df_col = pd.read_csv
index_col=0
dtype={'a': float, 'b': str}
print(df_col)
print(df_col.dtypes)
dtype: object
df_col = pd.read_csv
index_col=0
dtype={1: float, 2: str}
print(df_col)
print(df_col.dtypes)
dtype: object
import pandas as pd
df_sjis = pd.read_csv('data/src/sample_header_shift_jis.csv')
df_sjis = pd.read_csv
encoding='shift_jis'
print(df_sjis)
import pandas as pd
df = pd.read_csv
tokyo.zip
header=None
encoding='shift_jis'
print(df.shape)
print(df.head())
print(df.tail())
import pandas as pd
df_zip = pd.read_csv('data/src/sample_header.csv.zip')
print(df_zip)
import pandas as pd
print(pd.__version__)
df = pd.read_excel('data/src/sample.xlsx', index_col=0)
print(df)
print(type(df))
pandas.core.frame.DataFrame
df_sheet_index = pd.read_excel('data/src/sample.xlsx', sheet_name=0, index_col=0)
print(df_sheet_index)
df_sheet_name = pd.read_excel('data/src/sample.xlsx', sheet_name='sheet2', index_col=0)
print(df_sheet_name)
df_sheet_multi = pd.read_excel('data/src/sample.xlsx', sheet_name=[0, 'sheet2'], index_col=0)
print(type(df_sheet_multi))
print(len(df_sheet_multi))
print(df_sheet_multi.keys())
dict_keys([0, 'sheet2'])
print(df_sheet_multi[0])
print(type(df_sheet_multi[0]))
pandas.core.frame.DataFrame
print(df_sheet_multi['sheet2'])
print(type(df_sheet_multi['sheet2']))
pandas.core.frame.DataFrame
df_sheet_all = pd.read_excel('data/src/sample.xlsx', sheet_name=None, index_col=0)
print(type(df_sheet_all))
print(df_sheet_all.keys())
dict_keys(['sheet1', 'sheet2'])
df_header_index = pd.read_excel('data/src/sample.xlsx', header=None, index_col=None)
print(df_header_index)
print(df_header_index.columns)
Int64Index([0, 1, 2, 3], dtype='int64')
print(df_header_index.index)
RangeIndex(start=0, stop=4, step=1)
df_default = pd.read_excel('data/src/sample.xlsx')
print(df_default)
print(df_default.columns)
Index(['Unnamed: 0', 'A', 'B', 'C'], dtype='object')
print(df_default.index)
RangeIndex(start=0, stop=3, step=1)
print(pd.read_excel('data/src/sample.xlsx', index_col=0))
df_use_skip = pd.read_excel
index_col=0
usecols=[0, 1, 3]
skiprows=[1]
skipfooter=1
print(df_use_skip)
import pandas as pd
url = 'https://info.finance.yahoo.co.jp/ranking/?kd=4'
dfs = pd.read_html(url)
print(len(dfs))
print(dfs[0].head())
print(dfs[0][['名称', '時価総額（百万円）']].head())
dfs[0].columns = ['順位', 'コード', '市場', '名称', '時刻', '取引値', '発行済み株式数', '時価総額（百万円）', '単元株数', '掲示板']
print(dfs[0][['名称', '時価総額（百万円）']].head())
url = 'https://ja.wikipedia.org/wiki/Python'
dfs = pd.read_html(url)
print(len(dfs))
dfs = pd.read_html(url, match='リリース日')
print(len(dfs))
print(dfs[0])
print(dfs[1])
dfs = pd.read_html(url, match='リリース日', header=0)
print(len(dfs))
print(dfs[0])
print(dfs[1])
df = pd.concat([dfs[0], dfs[1]], ignore_index=True).sort_values('リリース日[17]')
print(df)
df.to_csv('data/dst/pandas_read_html_sample.csv')
import pandas as pd
import json
s = '{"col1":{"row1":1,"row2":2,"row3":3},"col2":{"row1":"a","row2":"x","row3":"\u3042"}}'
df_s = pd.read_json(s)
print(df_s)
s_single_quote = "{'col1':{'row1':1,'row2':2,'row3':3},'col2':{'row1':'a','row2':'x','row3':'\u3042'}}"
df_s_single_quote = pd.read_json(s_single_quote)
ValueError: Expected
object or value
print(pd.read_json(s_single_quote.replace("'", '"')))
df_f = pd.read_json('data/src/sample_from_pandas_columns.json')
print(df_f)
df_gzip = pd.read_json('data/src/sample_from_pandas_columns.gz', compression='infer')
print(df_gzip)
df_s_index = pd.read_json(s, orient='index')
print(df_s_index)
df_s_split = pd.read_json(s, orient='split')
ValueError: JSON
key(s)
s_jsonl = ''
print(s_jsonl)
df_s_jsonl = pd.read_json(s_jsonl, orient='records', lines=True)
print(df_s_jsonl)
s_nested = '{"OTHER": "x", "DATA": {"col1":{"row1":1,"row2":2},"col2":{"row1":"a","row2":"x"}}}'
print(pd.read_json(s_nested))
d = json.loads(s_nested)
print(d)
print(type(d))
d_target = d['DATA']
print(d_target)
print(type(d_target))
s_target = json.dumps(d_target)
print(s_target)
print(type(s_target))
df_target = pd.read_json(s_target)
print(df_target)
df_target2 = pd.read_json(json.dumps(json.loads(s_nested)['DATA']))
print(df_target2)
import pandas as pd
df = pd.DataFrame
index=['One', 'Two', 'Three']
print(df)
print(df.reindex(index=['Two', 'Three', 'One']))
print(df.reindex(columns=['B', 'C', 'A']))
print(df.reindex(index=['Two', 'Three', 'One'], columns=['B', 'C', 'A']))
print(df.reindex(columns=['B', 'A'], index=['Three', 'One']))
print(df.reindex(['Two', 'Three', 'One'], axis=0))
print(df.reindex(['B', 'C', 'A'], axis='columns'))
print(df[['B', 'C', 'A']])
print(df.reindex(columns=['B', 'X', 'C'], index=['Two', 'One', 'Four']))
df.reindex
columns=['B', 'X', 'C']
index=['Two', 'One', 'Four']
fill_value=0
df = pd.DataFrame
index=[10, 20]
print(df)
print(df.reindex(index=[5, 10, 15, 20, 25]))
print(df.reindex(index=[5, 10, 15, 20, 25], method='bfill'))
print(df.reindex(index=[5, 10, 15, 20, 25], method='ffill'))
print(df.reindex(index=[5, 10, 15, 20, 25], method='nearest'))
print(df.reindex(index=[10, 12, 14, 16, 18, 20]))
print(df.reindex(index=[10, 12, 14, 16, 18, 20], method='bfill', limit=2))
print(df.reindex(index=[25, 5, 15], method='bfill'))
print(df.reindex(index=[5, 15, 25], method='bfill'))
print(df.reindex(index=[5, 10, 15, 20, 25]).fillna(method='bfill'))
print(df.reindex(index=[5, 10, 15, 20, 25]).interpolate())
print(df.reindex(columns=['A', 'X', 'C'], method='bfill'))
print(df.reindex(columns=['A', 'X', 'C']).fillna(method='bfill', axis=1))
print(df.reindex(columns=['A', 'X', 'C']).interpolate(axis=1))
df = pd.DataFrame
index=[20, 10, 30]
print(df)
print(df.reindex(index=[10, 15, 20], method='ffill'))
ValueError: index
increasing or decreasing
print(df.reindex(index=[10, 15, 20]))
print(df.reindex(index=[10, 15, 20]).fillna(method='bfill'))
import pandas as pd
df1 = pd.DataFrame
index=[10, 20, 30]
print(df1)
df2 = pd.DataFrame
index=[10, 15, 20]
print(df2)
print(df1.reindex(columns=df2.columns, index=df2.index))
print(df1.reindex_like(df2))
print(df1.reindex_like(df2, method='bfill'))
print(df1.reindex_like(df2, fill_value=0))
TypeError: reindex_like
print(df1.reindex_like(df2).fillna(0))
import pandas as pd
df = pd.DataFrame
index=pd.date_range('2020-06-01', '2020-06-05', freq='2D')
print(df)
new_index = pd.date_range('2020-06-01', '2020-06-05', freq='1D')
print(new_index)
dtype='datetime64[ns]'
freq='D'
print(df.reindex(index=new_index))
print(df.reindex(index=new_index, method='bfill'))
print(df.reindex(index=new_index).interpolate(method='time'))
print(df.reindex(index=['2020-06-01', '2020-06-02', '2020-06-03']))
df2 = pd.DataFrame
index=pd.date_range('2020-06-01', '2020-06-05', freq='D')
print(df2)
print(df.reindex_like(df2))
print(df.reindex_like(df2).interpolate(method='time'))
import pandas as pd
print(pd.__version__)
df = pd.read_csv('data/src/sample_pandas_normal.csv')
df.iloc[1, 3] = 24
print(df)
print(df.replace('CA', 'California'))
print(df.replace(24, 100))
print(df.replace({'CA': 'California', 24: 100}))
print(df.replace(['CA', 24], ['California', 100]))
print(df.replace(['CA', 24, 'NY'], ['California', 100]))
ValueError: Replacement
print(df.replace(['CA', 24], 'XXX'))
print(df.replace({'CA': 'NY', 'NY': 'XXX'}))
print(df.replace({'NY': 'XXX', 'CA': 'NY'}))
print(df.replace(['CA', 'NY'], ['NY', 'XXX']))
print(df.replace(['NY', 'CA'], ['XXX', 'NY']))
print(df.replace('CA', 'NY').replace('NY', 'XXX'))
print(df.replace('NY', 'XXX').replace('CA', 'NY'))
print(df.replace({'age': {24: 100}}))
print(df.replace({'age': {24: 100, 18: 0}, 'point': {24: 50}}))
print(df.replace({'age': [[24, 18], [100, 0]], 'point': {24: 50}}))
TypeError: If
print(df.replace({'age': 24, 'point': 70}, 100))
print(df.replace({'age': [24, 18], 'point': 70}, 100))
print(df.replace('li', 'LI'))
df.replace
regex=True
df['name'] = df['name'].str.replace('li', 'LI')
print(df)
df = pd.read_csv('data/src/sample_pandas_normal.csv')
print(df)
df.replace('CA', 'California', inplace=True)
print(df)
import pandas as pd
df = pd.read_csv('data/src/sample_pandas_normal.csv')
print(df)
df.sort_values('state', inplace=True)
print(df)
df_r = df.reset_index()
print(df_r)
df_r = df.reset_index(drop=True)
print(df_r)
df.reset_index(inplace=True, drop=True)
print(df)
df = pd.read_csv('data/src/sample_pandas_normal.csv', index_col=0)
print(df)
df_r = df.reset_index()
print(df_r)
df_s = df.set_index('state')
print(df_s)
df_rs = df.reset_index().set_index('state')
print(df_rs)
import pandas as pd
s = pd.Series(range(10))
print(s)
dtype: int64
print(s.rolling(3))
window=3
center=False
axis=0
print(type(s.rolling(3)))
pandas.core.window.rolling.Rolling
print(s.rolling(3).sum())
dtype: float64
print(s.rolling(2).sum())
dtype: float64
print(s.rolling(4).sum())
dtype: float64
print(s.rolling(3, center=True).sum())
dtype: float64
print(s.rolling(4, center=True).sum())
dtype: float64
print(s.rolling(3, min_periods=2).sum())
dtype: float64
print(s.rolling(3, min_periods=1).sum())
dtype: float64
df = pd.DataFrame
range(10)
range(10, 0, -1)
range(10, 20)
range(20, 10, -1)
print(df.rolling(2).sum())
print(df.rolling(2, axis=1).sum())
print(s.rolling(3).mean())
dtype: float64
s.rolling(3).agg
lambda x: max(x) - min(x)
import pandas as pd
import numpy as np
from decimal import Decimal, ROUND_HALF_UP, ROUND_HALF_EVEN
import platform
print(platform.python_version())
print(pd.__version__)
print(np.__version__)
s_f = pd.Series([123.456, 654.123])
print(s_f)
dtype: float64
print(s_f.round())
dtype: float64
print(s_f.round().astype(int))
dtype: int64
print(s_f.round(2))
dtype: float64
print(s_f.round(-2))
dtype: float64
s_i = pd.Series([123, 654])
print(s_i)
dtype: int64
print(s_i.round())
dtype: int64
print(s_i.round(2))
dtype: int64
print(s_i.round(-2))
dtype: int64
s_i_round = s_i.round(-2)
print(s_i_round)
dtype: int64
print(s_i)
dtype: int64
df = pd.DataFrame({'f': [123.456, 654.321], 'i': [123, 654], 's': ['abc', 'xyz']})
print(df)
print(df.dtypes)
dtype: object
print(df.round())
print(df.round(2))
print(df.round(-2))
print(df.round({'f': 2, 'i': -1}))
print(df.round({'i': -2}))
s = pd.Series([0.5, 1.5, 2.5, 3.5, 4.5])
print(s)
dtype: float64
print(s.round())
dtype: float64
s = pd.Series([5, 15, 25, 5.1, 15.1, 25.1])
print(s)
dtype: float64
print(s.round(-1))
dtype: float64
print('Python round')
print('0.005 => ', round(0.005, 2))
print('0.015 => ', round(0.015, 2))
print('0.025 => ', round(0.025, 2))
print('0.035 => ', round(0.035, 2))
print('0.045 => ', round(0.045, 2))
print('2.675 => ', round(2.675, 2))
print('NumPy np.around (np.round)')
print('0.005 => ', np.around(0.005, 2))
print('0.015 => ', np.around(0.015, 2))
print('0.025 => ', np.around(0.025, 2))
print('0.035 => ', np.around(0.035, 2))
print('0.045 => ', np.around(0.045, 2))
print('2.675 => ', np.around(2.675, 2))
np.around (np.round)
s = pd.Series([0.005, 0.015, 0.025, 0.035, 0.045, 2.675])
print(s)
dtype: float64
print(s.round(2))
dtype: float64
print(Decimal(2.675))
print(Decimal('2.675'))
print(Decimal('2.675').quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
print(Decimal('2.675').quantize(Decimal('0.01'), rounding=ROUND_HALF_EVEN))
print(Decimal('0.5').quantize(Decimal('0'), rounding=ROUND_HALF_UP))
print(Decimal('0.5').quantize(Decimal('0'), rounding=ROUND_HALF_EVEN))
s = pd.Series([0.5, 1.5, 2.5, 3.5, 4.5])
s.map
lambda x: float
Decimal(str(x))
quantize(Decimal('0'), rounding=ROUND_HALF_UP)
dtype: float64
s = pd.Series([0.005, 0.015, 0.025, 0.035, 0.045, 2.675])
s.map
lambda x: float
Decimal(str(x))
quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
dtype: float64
s = pd.Series([5, 15, 25, 5.1, 15.1, 25.1])
s.map
lambda x: int
Decimal(str(x))
quantize(Decimal('1E1'), rounding=ROUND_HALF_UP)
dtype: int64
s = pd.Series([0.005, 0.015, 0.025, 0.035, 0.045, 2.675])
s.map
lambda x: float
Decimal(str(x))
quantize(Decimal('0.01'), rounding=ROUND_HALF_EVEN)
dtype: float64
import pandas as pd
df = pd.read_csv('data/src/sample_pandas_normal.csv', index_col=0)
print(df)
print(df.index.get_loc('Alice'))
print(df.index.get_loc('Ellen'))
print(df.columns.get_loc('age'))
print(df.columns.get_loc('point'))
print(df.index.get_loc('XXX'))
print(df.columns.get_loc('XXX'))
df_dup = df.rename(index={'Charlie': 'Bob'})
print(df_dup)
print(df_dup.index.get_loc('Bob'))
slice(1, 3, None)
print(type(df_dup.index.get_loc('Bob')))
df_dup.rename(index={'Ellen': 'Bob'}, inplace=True)
print(df_dup)
print(df_dup.index.get_loc('Bob'))
print(type(df_dup.index.get_loc('Bob')))
numpy.ndarray
print(df_dup[df_dup.index.get_loc('Bob')])
print(df_dup.iloc[df_dup.index.get_loc('Bob'), 0])
Name: age
dtype: int64
print(df_dup.query('index == "Bob"'))
l_index = list(df.index)
print(l_index)
print(type(l_index))
l_columns = list(df.columns)
print(l_columns)
print(type(l_columns))
print(l_index.index('Bob'))
l_index_dup = list(df_dup.index)
print(l_index_dup)
print([i for i, x in enumerate(l_index_dup) if x == 'Bob'])
print(df.query('state == "CA"'))
print(list(df.query('state == "CA"').index))
print(df.query('state == "TX"'))
print(list(df.query('state == "TX"').index))
print(df.query('state == "TX"').index[0])
print(df.reset_index())
print(list(df.reset_index().query('state == "CA"').index))
print(list(df.reset_index().query('state == "TX"').index))
print(df.reset_index().query('state == "TX"').index[0])
import pandas as pd
import seaborn as sns
df = sns.load_dataset("iris")
print(df.shape)
print(df.sample())
print(df.sample(n=3))
print(df.sample(n=3, random_state=0))
print(df.sample(frac=0.04))
print(df.head(3).sample(n=3, replace=True))
print(df.head(3).sample(n=5, replace=True))
print(df.head().sample(n=2, axis=1))
import pandas as pd
df = pd.DataFrame
df['f'] = pd.to_datetime(['2018-01-01', '2018-03-15', '2018-02-20', '2018-03-15'])
print(df)
X  [0, 0]   
Y  [0, 1]   
X  [1, 0]  
Z  [1, 1]   
print(df.dtypes)
datetime64[ns]
dtype: object
print(df.select_dtypes(include=int))
print(df.select_dtypes(include='int'))
print(df.select_dtypes(include='int64'))
print(df.select_dtypes(include='int32'))
print(df.select_dtypes(include=[int, float, 'datetime']))
print(df.select_dtypes(include='number'))
print(df.select_dtypes(include=object))
X  [0, 0]
Y  [0, 1]
X  [1, 0]
Z  [1, 1]
print(type(df.at[0, 'c']))
print(type(df.at[0, 'd']))
print(df.select_dtypes(exclude='number'))
X  [0, 0]   
Y  [0, 1]   
X  [1, 0]  
Z  [1, 1]   
print(df.select_dtypes(exclude=[bool, 'datetime']))
X  [0, 0]
Y  [0, 1]
X  [1, 0]
Z  [1, 1]
print(df.select_dtypes(include='number', exclude=int))
print(df.select_dtypes(include=[int, bool], exclude=int))
ValueError: include
numpy.int64
import pandas as pd
s = pd.Series([1, 2, 3], index=['ONE', 'TWO', 'THREE'])
print(s)
dtype: int64
print(s.rename({'ONE': 'a', 'THREE': 'c'}))
dtype: int64
print(s.rename(str.lower))
dtype: int64
print(s.add_prefix('X_'))
dtype: int64
print(s.add_suffix('_X'))
dtype: int64
print(s.set_axis(['a', 'b', 'c']))
dtype: int64
s.index = ['a', 'b', 'c']
print(s)
dtype: int64
import pandas as pd
import timeit
s = pd.Series(['a', 'b', 'c', 'd', 'e'])
print(s)
dtype: object
s_swap = pd.Series(s.index.values, s.values)
print(s_swap)
dtype: int64
print(s.values)
print(type(s.values))
numpy.ndarray
print(s.index.values)
print(type(s.index.values))
numpy.ndarray
s_swap = pd.Series(s.index, s)
print(s_swap)
dtype: int64
loop = 10000
result = timeit.timeit(lambda: pd.Series(s.index.values, s.values), number=loop)
print(result / loop)
result = timeit.timeit(lambda: pd.Series(s.index, s), number=loop)
print(result / loop)
s_large = pd.concat([s] * 100000)
print(len(s_large))
loop = 100
result = timeit.timeit(lambda: pd.Series(s_large.index.values, s_large.values), number=loop)
print(result / loop)
result = timeit.timeit(lambda: pd.Series(s_large.index, s_large), number=loop)
print(result / loop)
import pandas as pd
s = pd.Series([0, 1, 2], index=['a', 'b', 'c'])
print(s)
dtype: int64
df = pd.DataFrame(s)
print(df)
print(type(df))
pandas.core.frame.DataFrame
df_ = pd.DataFrame([s])
print(df_)
print(type(df_))
pandas.core.frame.DataFrame
s_name = pd.Series([0, 1, 2], index=['a', 'b', 'c'], name='X')
print(s_name)
Name: X
dtype: int64
print(pd.DataFrame(s_name))
print(pd.DataFrame([s_name]))
s1 = pd.Series([0, 1, 2], index=['a', 'b', 'c'])
print(s1)
dtype: int64
s2 = pd.Series([0.0, 0.1, 0.2], index=['a', 'b', 'c'])
print(s2)
dtype: float64
print(pd.DataFrame({'col0': s1, 'col1': s2}))
print(pd.DataFrame([s1, s2]))
print(pd.concat([s1, s2], axis=1))
s1_name = pd.Series([0, 1, 2], index=['a', 'b', 'c'], name='X')
print(s1_name)
Name: X
dtype: int64
s2_name = pd.Series([0.0, 0.1, 0.2], index=['a', 'b', 'c'], name='Y')
print(s2_name)
Name: Y
dtype: float64
pd.DataFrame
s1_name.name
s2_name.name
print(pd.DataFrame([s1_name, s2_name]))
print(pd.concat([s1_name, s2_name], axis=1))
s3 = pd.Series([0.1, 0.2, 0.3], index=['b', 'c', 'd'])
print(s3)
dtype: float64
print(pd.DataFrame({'col0': s1, 'col1': s3}))
print(pd.DataFrame([s1, s3]))
print(pd.concat([s1, s3], axis=1))
FutureWarning: Sorting
not sort 
pass 
sort=False
behavior and silence
pass 
sort=True
print(pd.concat([s1, s3], axis=1, join='inner'))
print(s1.values)
print(type(s1.values))
numpy.ndarray
print(pd.DataFrame({'col0': s1.values, 'col1': s3.values}))
print(pd.DataFrame([s1.values, s3.values]))
print(pd.concat([s1.values, s3.values], axis=1))
TypeError: cannot
numpy.ndarray
Series and DataFrame
print(pd.DataFrame({'col0': s1, 'col1': s3.values}))
print(pd.DataFrame([s1, s3.values]))
print(pd.DataFrame({'col0': s1.values, 'col1': s3.values}, index=s1.index))
print(pd.DataFrame([s1.values, s3.values], columns=s1.index))
s4 = pd.Series([0.1, 0.2], index=['b', 'd'])
print(s4)
dtype: float64
print(pd.DataFrame({'col0': s1, 'col1': s4}))
print(pd.DataFrame([s1, s4]))
print(pd.concat([s1, s4], axis=1, join='inner'))
print(pd.DataFrame({'col0': s1.values, 'col1': s4.values}))
ValueError: arrays
print(pd.DataFrame([s1.values, s4.values]))
s4.index = ['a', 'b']
print(s4)
dtype: float64
print(pd.DataFrame({'col0': s1, 'col1': s4}))
print(pd.DataFrame({'col0': s1, 'col1': s4}).fillna(100))
print(s)
dtype: int64
df = pd.DataFrame(s)
print(df)
s[0] = 100
print(s)
dtype: int64
print(df)
df_copy = pd.DataFrame(s, copy=True)
print(df_copy)
s[1] = 100
print(s)
dtype: int64
print(df_copy)
df_c = pd.concat([s1, s2], axis=1)
print(df_c)
s1[0] = 100
print(s1)
dtype: int64
print(df_c)
df_c_false = pd.concat([s1, s2], axis=1, copy=False)
print(df_c_false)
s1[1] = 100
print(s1)
dtype: int64
print(df_c_false)
import pandas as pd
df = pd.read_csv('data/src/sample_pandas_normal.csv')
print(df)
df_i = df.set_index('name')
print(df_i)
df_id = df.set_index('name', drop=False)
print(df_id)
df_mi = df.set_index(['state', 'name'])
print(df_mi)
df_mi.sort_index(inplace=True)
print(df_mi)
print(df_i)
df_ii = df_i.set_index('state')
print(df_ii)
df_mi = df_i.set_index('state', append=True)
print(df_mi)
print(df_mi.swaplevel(0, 1))
print(df_i)
df_ri = df_i.reset_index()
print(df_ri)
df_change = df_i.reset_index().set_index('state')
print(df_change)
df.set_index('name', inplace=True)
print(df)
df = pd.read_csv('data/src/sample_pandas_normal.csv', index_col=0)
print(df)
print(df.loc['Bob'])
Name: Bob
dtype: object
print(df.at['Bob', 'age'])
import pandas as pd
df = pd.DataFrame({'a': [0, 1, 2], 'b': [3, 4, 5]}, index=['x', 'y', 'z'])
print(df)
print(df.loc['x':'y']['a'])
Name: a
dtype: int64
df.loc['x':'y']['a'] = 100
using .loc[row_indexer,col_indexer] = value
documentation: http
returning-a-view-versus-a-copy
print(df)
print(df.loc[['x', 'y']]['a'])
Name: a
dtype: int64
df.loc[['x', 'y']]['a'] = 0
print(df)
df.loc['x':'y', 'a'] = 10
print(df)
df.loc[['x', 'y'], 'a'] = 0
print(df)
print(df.iloc[[0, 1]]['a'])
Name: a
dtype: int64
df.loc[[0, 1], 'a']
of [Int64Index([0, 1], dtype='int64')] 
df.iloc[[0, 1], 'a']
ValueError: Location
print(df.index[0])
print(df.index[1])
print(df.columns[0])
print(df.columns[1])
print(df.loc[[df.index[0], df.index[1]], 'a'])
Name: a
dtype: int64
print(df.iloc[:2]['a'])
Name: a
dtype: int64
print(df.loc[:df.index[2], 'a'])
Name: a
dtype: int64
print(df.loc[:df.index[2 - 1], 'a'])
Name: a
dtype: int64
df = pd.DataFrame({'a': [0, 1, 2], 'b': [3, 4, 5]}, index=['x', 'y', 'z'])
print(df)
df_slice = df.loc['x':'y']
print(df_slice)
print(df_slice['a'])
Name: a
dtype: int64
df_slice['a'] = 100
using .loc[row_indexer,col_indexer] = value
documentation: http
returning-a-view-versus-a-copy
print(df_slice)
print(df)
df_list = df.loc[['x', 'y']]
print(df_list)
print(df_list['a'])
Name: a
dtype: int64
df_list['a'] = 0
print(df_list)
print(df)
df_slice['c'] = 0
using .loc[row_indexer,col_indexer] = value
documentation: http
returning-a-view-versus-a-copy
print(df_slice)
df.at['x', 'a'] = 0
print(df)
print(df_slice)
df_slice_copy = df.loc['x':'y'].copy()
print(df_slice_copy)
df.at['x', 'a'] = 100
print(df)
print(df_slice_copy)
import pandas as pd
import numpy as np
a = np.array([[0, 1], [2, 3], [4, 5]])
print(a)
df = pd.DataFrame(a)
print(df)
print(np.shares_memory(a, df))
print(df._is_view)
a[0, 0] = 100
print(a)
print(df)
a_str = np.array([['a', 'x'], ['b', 'y'], ['c', 'z']])
print(a_str)
df_str = pd.DataFrame(a_str)
print(df_str)
print(np.shares_memory(a_str, df_str))
print(df_str._is_view)
a_str[0, 0] = 'n'
print(a_str)
print(df_str)
df_homo = pd.DataFrame({'a': [0, 1, 2], 'b': [3, 4, 5]})
print(df_homo)
print(df_homo.dtypes)
dtype: object
a_homo = df_homo.values
print(a_homo)
print(np.shares_memory(a_homo, df_homo))
df_homo.iat[0, 0] = 100
print(df_homo)
print(a_homo)
df_hetero = pd.DataFrame({'a': [0, 1, 2], 'b': ['x', 'y', 'z']})
print(df_hetero)
print(df_hetero.dtypes)
dtype: object
a_hetero = df_hetero.values
print(a_hetero)
print(np.shares_memory(a_hetero, df_hetero))
df_hetero.iat[0, 0] = 100
print(df_hetero)
print(a_hetero)
import pandas as pd
df = pd.DataFrame
range(1, 6)
x**2 for x in range(1, 6)
x**3 for x in range(1, 6)
print(df)
print(df.shift())
print(df.shift(1))
print(df.shift(2))
print(df.shift(-1))
print(df.shift(axis=1))
print(df.shift(-1, axis=1))
import pandas as pd
df = pd.read_csv('data/src/sample_pandas_normal.csv')
print(df)
print(df.sort_values('name'))
print(df.sort_values('name', ascending=False))
l_order = ['Charlie', 'Alice', 'Dave', 'Bob']
df['order'] = df['name'].apply(lambda x: l_order.index(x) if x in l_order else -1)
print(df)
print(df.sort_values('order'))
print(df.sort_values('order').reset_index(drop=True).drop(columns='order'))
d_order = {'Charlie': 0, 'Alice': 1, 'Dave': 2, 'Bob': 3}
df['order'] = df['name'].map(d_order)
print(df)
print(df.sort_values('order'))
print(df.sort_values('order', na_position='first'))
print(df.sort_values('order', na_position='first').reset_index(drop=True).drop(columns='order'))
import pandas as pd
df = pd.read_csv('data/src/sample_pandas_normal.csv')
print(df)
df_s = df.sort_values('state')
print(df_s)
df_s = df.sort_values('state', ascending=False)
print(df_s)
df_s = df.sort_values(['state', 'age'])
print(df_s)
df_s = df.sort_values(['age', 'state'])
print(df_s)
df_s = df.sort_values(['age', 'state'], ascending=[True, False])
print(df_s)
df_nan = df.copy()
df_nan.iloc[:2, 1] = pd.np.nan
print(df_nan)
df_nan_s = df_nan.sort_values('age')
print(df_nan_s)
df_nan_s = df_nan.sort_values('age', na_position='first')
print(df_nan_s)
df.sort_values('state', inplace=True)
print(df)
df_d = df.drop(['name', 'state'], axis=1)
print(df_d)
df_d .sort_values(by=1, axis=1, ascending=False, inplace=True)
print(df_d)
print(df)
df_s = df.sort_index()
print(df_s)
df_s = df.sort_index(ascending=False)
print(df_s)
df.sort_index(inplace=True)
print(df)
df_s = df.sort_index(axis=1)
print(df_s)
df.sort_index(axis=1, ascending=False, inplace=True)
print(df)
import pandas as pd
df = pd.DataFrame
print(df)
s = df.stack()
print(s)
dtype: object
print(type(s))
pandas.core.series.Series
print(s.index)
levels=[[0, 1, 2, 3, 4, 5], ['A', 'B', 'C', 'D']]
labels=[[0, 0, 0, 0, 1, 1, 1, 1, 2, 2, 2, 2, 3, 3, 3, 3, 4, 4, 4, 4, 5, 5, 5, 5], [0, 1, 2, 3, 0, 1, 2, 3, 0, 1, 2, 3, 0, 1, 2, 3, 0, 1, 2, 3, 0, 1, 2, 3]]
print(s.unstack())
print(s.unstack(level=0))
print(df.T)
df_m = df.set_index(['A', 'B'])
print(df_m)
print(df_m.index)
levels=[['a', 'b'], ['x', 'y', 'z']]
labels=[[0, 0, 0, 1, 1, 1], [0, 1, 2, 0, 1, 2]]
names=['A', 'B']
df_mu = df_m.unstack()
print(df_mu)
print(df_mu.columns)
levels=[['C', 'D'], ['x', 'y', 'z']]
labels=[[0, 0, 0, 1, 1, 1], [0, 1, 2, 0, 1, 2]]
names=[None, 'B']
print(df_m.unstack(level=0))
print(df_m.unstack(level='A'))
print(df_mu)
print(df_mu.swaplevel(axis=1))
print(df_mu.swaplevel(axis=1).sort_index(axis=1))
print(df_mu.stack())
print(df_mu.stack(level=0))
print(df)
print(df.pivot(index='A', columns='B', values='C'))
print(df.pivot(index='A', columns='B', values=['C', 'D']))
print(df.pivot(index='A', columns='B'))
print(df.set_index(['A', 'B']).unstack('B'))
df.loc[1, 'B'] = 'x'
print(df)
print(df.pivot(index='A', columns='B'))
ValueError: Index
print(df.pivot_table(index='A', columns='B'))
print(df.pivot_table(index='A', columns='B', aggfunc=sum))
import pandas as pd
df = pd.read_csv('data/src/estat_0003215840.csv')
df = df[['男女別・性比', '人口', '年齢各歳', 'value']]
print(df.head(10))
print(df['男女別・性比'].unique())
print(df['人口'].unique())
print(df[['男女別・性比', '人口']].drop_duplicates())
df.pivot(index='年齢各歳', columns='男女別・性比', values='value')
ValueError: Index
df_jp = df.query('人口 == "日本人人口"')
print(df_jp.pivot(index='年齢各歳', columns='男女別・性比', values='value').head(10))
print(df.set_index(['年齢各歳', '人口', '男女別・性比']).unstack(['人口', '男女別・性比']).sort_index(axis=1).head(10))
df_pt = df.pivot_table(index='年齢各歳', columns=['人口', '男女別・性比'], values='value')
print(df_pt.head(10))
print(df_pt.columns)
levels=[['日本人人口', '総人口'], ['人口性比', '女', '男', '男女計']]
labels=[[0, 0, 0, 0, 1, 1, 1, 1], [0, 1, 2, 3, 0, 1, 2, 3]]
names=['人口', '男女別・性比']
print(df_pt.loc[:, ('日本人人口', '男')].head(10))
dtype: float64
print(df_pt.loc[:, ('日本人人口', ['男', '女'])].head(10))
import pandas as pd
df = pd.read_csv('data/src/sample_pandas_normal.csv').head(3)
print(df)
print(df['name'].str.cat(df['state']))
Name: name
dtype: object
print(df['name'].str.cat(df['state'], sep=' in '))
Name: name
dtype: object
print(df['name'].str.cat(['X', 'Y', 'Z'], sep=' in '))
Name: name
dtype: object
print(df['name'].str.cat([df['state'], ['X', 'Y', 'Z']], sep='-'))
Name: name
dtype: object
print(df['name'].str.cat('X', sep='-'))
ValueError: Did
print(df['name'] + df['state'])
dtype: object
print(df['name'] + ' in ' + df['state'])
dtype: object
print(df['name'] + ' in ' + df['state'] + ' - ' + ['X', 'Y', 'Z'])
dtype: object
df['col_NaN'] = ['X', pd.np.nan, 'Z']
print(df)
print(df['name'].str.cat(df['col_NaN'], sep='-'))
Name: name
dtype: object
print(df['name'].str.cat(df['col_NaN'], sep='-', na_rep='No Data'))
Name: name
dtype: object
print(df['name'] + '-' + df['col_NaN'])
dtype: object
print(df['name'] + '-' + df['col_NaN'].fillna('No Data'))
dtype: object
print(df['name'].str.cat(df['age'], sep='-'))
TypeError: sequence
print(df['name'].str.cat(df['age'].astype(str), sep='-'))
Name: name
dtype: object
print(df['name'] + '-' + df['age'])
TypeError: can
str (not "int") 
print(df['name'] + '-' + df['age'].astype(str))
dtype: object
df['name_state'] = df['name'].str.cat(df['state'], sep=' in ')
print(df)
print(df.drop(columns=['name', 'state']))
df = pd.read_csv('data/src/sample_pandas_normal.csv').head(3)
print(df)
print(df.assign(name_state=df['name'] + ' in ' + df['state']))
print(df.assign(name_state=df['name'] + ' in ' + df['state']).drop(columns=['name', 'state']))
import pandas as pd
s_org = pd.Series(['aaa@xxx.com', 'bbb@yyy.com', 'ccc@zzz.com', 'ddd'], index=['A', 'B', 'C', 'D'])
print(s_org)
xxx.com
yyy.com
zzz.com
dtype: object
df = s_org.str.extract
expand=True
print(df)
df = s_org.str.extract
expand=False
print(df)
df_single = s_org.str.extract
expand=True
print(df_single)
print(type(df_single))
pandas.core.frame.DataFrame
s = s_org.str.extract
expand=False
print(s)
print(type(s))
dtype: object
pandas.core.series.Series
df_name = s_org.str.extract
expand=True
print(df_name)
import pandas as pd
s_org = pd.Series(['aaa@xxx.com', 'bbb@yyy.net', 'ccc@zzz.co.jp'], index=['A', 'B', 'C'])
print(s_org)
xxx.com
yyy.net
zzz.co.jp
dtype: object
df_single = s_org.str.extract('(.+)@', expand=True)
print(df_single)
print(type(df_single))
pandas.core.frame.DataFrame
s = s_org.str.extract('(.+)@', expand=False)
print(s)
print(type(s))
dtype: object
pandas.core.series.Series
df_name = s_org.str.extract('(?P<local>.+)@', expand=True)
print(df_name)
print(type(df_name))
pandas.core.frame.DataFrame
print(s_org.str.extract('(.+)@(.+)'))
xxx.com
yyy.net
zzz.co.jp
print(s_org.str.extract('(?P<local>.+)@(?P<domain>.+)'))
xxx.com
yyy.net
zzz.co.jp
print(s_org.str.extract('(a+)', expand=True))
s_org2 = pd.Series
xxx.com
xxx.com
yyy.net
yyy.net
zzz.co.jp
index=['A', 'B', 'C']
print(s_org2)
xxx.com
xxx.com
yyy.net
yyy.net
zzz.co.jp
dtype: object
print(s_org2.str.extract('([a-z]+)@([a-z.]+)', expand=True))
xxx.com
yyy.net
zzz.co.jp
df_all = s_org2.str.extractall('([a-z]+)@([a-z.]+)')
print(df_all)
xxx.com
xxx.com
yyy.net
yyy.net
zzz.co.jp
print(df_all.index)
levels=[['A', 'B', 'C'], [0, 1]]
labels=[[0, 0, 1, 1, 2], [0, 1, 0, 1, 0]]
names=[None, 'match']
print(s_org.str.extractall('([a-z]+)@([a-z.]+)'))
xxx.com
yyy.net
zzz.co.jp
import pandas as pd
print(pd.__version__)
df = pd.DataFrame
print(df)
print(df.dtypes)
dtype: object
print(df['i'].astype(str))
Name: i
dtype: object
print(df['f'].astype(str))
Name: f
dtype: object
print(df.astype(str))
print(df.astype(str).dtypes)
dtype: object
print(df['i'].astype(float))
Name: i
dtype: float64
print(df['f'].astype(int))
Name: f
dtype: int64
print(df['s_i'].astype(int))
Name: s_i
dtype: int64
print(df['s_i'].astype(float))
Name: s_i
dtype: float64
print(df['s_f'].astype(float))
Name: s_f
dtype: float64
print(df['s_f'].astype(int))
ValueError: invalid
int() 
print(df['s_f'].astype(float).astype(int))
Name: s_f
dtype: int64
df['i'] = df['i'].astype(str)
print(df)
df['f_s'] = df['f'].astype(str)
print(df)
print(df.dtypes)
dtype: object
s_int = pd.Series
print(s_int)
dtype: int64
print(s_int.map(bin))
dtype: object
print(s_int.map(oct))
dtype: object
print(s_int.map(hex))
dtype: object
print(s_int.map('{:b}'.format))
dtype: object
print(s_int.map('{:#b}'.format))
dtype: object
print(s_int.map('{:#010b}'.format))
dtype: object
df_str = pd.DataFrame
print(df_str)
print(df_str.dtypes)
dtype: object
print(df_str['bin'].astype(int))
ValueError: invalid
int() 
print(df_str['bin'].map(lambda x: int(x, 2)))
Name: bin
dtype: int64
print(df_str['oct'].map(lambda x: int(x, 8)))
Name: oct
dtype: int64
print(df_str['hex'].map(lambda x: int(x, 16)))
Name: hex
dtype: int64
print(df_str.applymap(lambda x: int(x, 0)))
print(df_str['dec'].map(lambda x: int(x, 2)))
Name: dec
dtype: int64
s_str_dec = pd.Series(['01', '10', '11'])
print(s_str_dec)
dtype: object
print(s_str_dec.astype(int))
dtype: int64
print(s_str_dec.map(lambda x: int(x, 0)))
ValueError: invalid
int() 
print(df_str['oct'].map(lambda x: int(x, 8)).map(hex))
Name: oct
dtype: object
s_str = pd.Series(['0', '10', 'xxx'])
print(s_str)
dtype: object
print(s_str.str.zfill(8))
dtype: object
print(s_str.str.rjust(8))
dtype: object
print(s_str.str.rjust(8, '_'))
dtype: object
print(s_str.str.center(8))
dtype: object
print(s_str.str.center(8, '_'))
dtype: object
print(s_str.str.ljust(8))
dtype: object
print(s_str.str.ljust(8, '_'))
dtype: object
s_num = pd.Series([0, 10, 100])
print(s_num.str.rjust(8, '_'))
AttributeError: Can
use .str
np.object_
print(s_num.astype(str).str.rjust(8, '_'))
dtype: object
df = pd.DataFrame
print(df)
print(df.dtypes)
dtype: object
print(df['i'].map('{:08}'.format))
Name: i
dtype: object
print(df['i'].map('{:_<8}'.format))
Name: i
dtype: object
print(df['i'].map('{:x}'.format))
Name: i
dtype: object
print(df['i'].map('{:#010b}'.format))
Name: i
dtype: object
print(df['f'].map('{:.2f}'.format))
Name: f
dtype: object
print(df['f'].map('{:.2g}'.format))
Name: f
dtype: object
print(df['f'].map('{:.2e}'.format))
Name: f
dtype: object
print(df['f'].map('{:.2%}'.format))
Name: f
dtype: object
print(df['round'].map('{:.0f}'.format))
Name: round
dtype: object
import pandas as pd
s_sep = pd.Series(['1,000,000', '1,000', '1'])
print(s_sep)
dtype: object
print(s_sep.astype(int))
ValueError: invalid
int() 
print(s_sep.str.replace(',', '').astype(int))
dtype: int64
print(s_sep.str.replace(',', '').astype(float))
dtype: float64
import pandas as pd
s = pd.Series([' a-a-x ', ' b-x-b ', ' x-c-c '])
print(s)
dtype: object
s_new = s.str.replace('x', 'z')
print(s_new)
dtype: object
df = pd.DataFrame
columns=['col1', 'col2']
print(df)
df['col1'] = df['col1'].str.replace('x', 'z')
print(df)
s_new = s.str.strip()
print(s_new)
dtype: object
s_new = s.str.strip(' x')
print(s_new)
dtype: object
df['col1'] = df['col1'].str.strip()
print(df)
s_new = s.str.lstrip()
print(s_new)
dtype: object
s_new = s.str.rstrip()
print(s_new)
dtype: object
s = pd.Series(['Hello World', 'hello world', 'HELLO WORLD'])
print(s)
dtype: object
s_new = s.str.lower()
print(s_new)
dtype: object
s_new = s.str.upper()
print(s_new)
dtype: object
s_new = s.str.capitalize()
print(s_new)
dtype: object
s_new = s.str.title()
print(s_new)
dtype: object
import pandas as pd
df = pd.read_csv('data/src/sample_pandas_normal.csv').head(3)
print(df)
mask = [True, False, True]
df_mask = df[mask]
print(df_mask)
print(df['state'] == 'CA')
Name: state
dtype: bool
print(df[df['state'] == 'CA'])
print(df['name'].str.contains('li'))
Name: name
dtype: bool
print(df[df['name'].str.contains('li')])
df_nan = df.copy()
df_nan.iloc[2, 0] = float('nan')
print(df_nan)
print(df_nan['name'].str.contains('li'))
Name: name
dtype: object
print(df_nan[df_nan['name'].str.contains('li')])
ValueError: cannot
print(df_nan['name'].str.contains('li', na=False))
Name: name
dtype: bool
print(df_nan['name'].str.contains('li', na=True))
Name: name
dtype: bool
print(df['name'].str.contains('LI'))
Name: name
dtype: bool
print(df['name'].str.contains('LI', case=False))
Name: name
dtype: bool
print(df['name'].str.contains('i.*e'))
Name: name
dtype: bool
print(df['name'].str.contains('i.*e', regex=False))
Name: name
dtype: bool
df_q = df.copy()
df_q.iloc[2, 0] += '?'
print(df_q)
print(df_q['name'].str.contains('?'))
error: nothing
print(df_q['name'].str.contains('?', regex=False))
Name: name
dtype: bool
df_q['name'].str.contains
Name: name
dtype: bool
print(df['name'].str.endswith('e'))
Name: name
dtype: bool
print(df[df['name'].str.endswith('e')])
print(df['name'].str.startswith('B'))
Name: name
dtype: bool
print(df[df['name'].str.startswith('B')])
print(df['name'].str.match('.*i.*e'))
Name: name
dtype: bool
print(df[df['name'].str.match('.*i.*e')])
print(df['name'].str.match('.*i'))
Name: name
dtype: bool
print(df['name'].str.match('i.*e'))
Name: name
dtype: bool
import pandas as pd
df = pd.DataFrame
print(df)
print(df.dtypes)
dtype: object
print(df['a'].str[:2])
Name: a
dtype: object
print(df['a'].str[-2:])
Name: a
dtype: object
print(df['a'].str[::2])
Name: a
dtype: object
print(df['a'].str[2])
Name: a
dtype: object
print(df['a'].str[0])
Name: a
dtype: object
print(df['a'].str[-1])
Name: a
dtype: object
df['a_head'] = df['a'].str[:2]
print(df)
print(df['b'].str[:2])
AttributeError: Can
use .str
np.object_
print(df['b'].astype(str).str[:2])
Name: b
dtype: object
print(df['b'].astype(str).str[:2].astype(int))
Name: b
dtype: int64
print((df['b'] / 10).astype(int))
Name: b
dtype: int64
import pandas as pd
s_org = pd.Series(['aaa@xxx.com', 'bbb@yyy.com', 'ccc@zzz.com', 'ddd'], index=['A', 'B', 'C', 'D'])
print(s_org)
print(type(s_org))
xxx.com
yyy.com
zzz.com
dtype: object
pandas.core.series.Series
s = s_org.str.split('@')
print(s)
print(type(s))
A    [aaa, xxx.com]
B    [bbb, yyy.com]
C    [ccc, zzz.com]
D             [ddd]
dtype: object
pandas.core.series.Series
df = s_org.str.split('@', expand=True)
print(df)
print(type(df))
xxx.com
yyy.com
zzz.com
pandas.core.frame.DataFrame
df.columns = ['local', 'domain']
print(df)
xxx.com
yyy.com
zzz.com
print(df['domain'].str.split('.', expand=True))
df2 = pd.concat([df, df['domain'].str.split('.', expand=True)], axis=1).drop('domain', axis=1)
print(df2)
df3 = pd.concat([df['local'], df['domain'].str.split('.', expand=True)], axis=1)
print(df3)
df3.rename(columns={0: 'second_LD', 1: 'TLD'}, inplace=True)
print(df3)
import pandas as pd
df = pd.DataFrame
range(1, 32, 2)
index=pd.date_range('2018-08-01', '2018-08-31', freq='2D')
print(df)
print(df.asfreq('10D'))
print(df.asfreq('5D'))
print(df.asfreq('W'))
print(df.asfreq('W-WED'))
print(df.asfreq('W', fill_value=0))
print(df.asfreq('W', method='pad'))
print(df.asfreq('W', method='ffill'))
print(df.asfreq('W', method='backfill'))
print(df.asfreq('W', method='bfill'))
df_3D = pd.DataFrame
range(1, 32, 3)
index=pd.date_range('2018-08-01', '2018-08-31', freq='3D')
print(df_3D.asfreq('D', method='bfill'))
df_h = pd.DataFrame
range(9)
index=pd.date_range('2018-08-01', '2018-08-05', freq='12H')
print(df_h)
print(df_h.asfreq('D'))
print(df_h.resample('D').mean())
import pandas as pd
print(pd.__version__)
df = pd.read_csv
index_col='date'
parse_dates=True
head(3)
print(df)
print(type(df.index))
pandas.core.indexes.datetimes.DatetimeIndex
print(df.index[0])
print(type(df.index[0]))
pandas._libs.tslibs.timestamps.Timestamp
df_default = pd.read_csv('data/src/sample_date.csv', index_col='date').head(3)
print(df_default)
print(type(df_default.index))
pandas.core.indexes.base.Index
print(df_default.index[0])
print(type(df_default.index[0]))
df_jp_ng = pd.read_csv
index_col='date'
parse_dates=True
head(3)
print(df_jp_ng)
print(type(df_jp_ng.index))
pandas.core.indexes.base.Index
df_jp = pd.read_csv
index_col='date'
parse_dates=True
date_parser=lambda x: pd.to_datetime(x, format='%Y年%m月%d日')
head(3)
print(df_jp)
print(type(df_jp.index))
pandas.core.indexes.datetimes.DatetimeIndex
df = pd.read_csv('data/src/sample_date.csv').head(3)
print(df)
print(type(df.index))
pandas.core.indexes.range.RangeIndex
df['date'] = pd.to_datetime(df['date'])
print(df['date'].dtype)
datetime64[ns]
df.set_index('date', inplace=True)
print(df)
print(type(df.index))
pandas.core.indexes.datetimes.DatetimeIndex
df = pd.read_csv('data/src/sample_date.csv').head(3)
df.set_axis(pd.to_datetime(df['date']), axis='index', inplace=True)
print(df)
print(type(df.index))
pandas.core.indexes.datetimes.DatetimeIndex
df.drop(['date'], axis='columns', inplace=True)
print(df)
df = pd.read_csv('data/src/sample_date.csv').head(3)
df.index = pd.to_datetime(df['date'])
df.drop(['date'], axis='columns', inplace=True)
print(df)
print(type(df.index))
pandas.core.indexes.datetimes.DatetimeIndex
df = pd.read_csv('data/src/sample_date.csv', index_col='date', parse_dates=True)
print(df)
print(type(df.index))
pandas.core.indexes.datetimes.DatetimeIndex
print(df.loc['2018'])
print(df.loc['2017-11'])
print(df.loc['2017-12-15':'2018-01-15'])
print(df.at['01/19/2018', 'val_1'])
print(df.at['20180103', 'val_2'])
import pandas as pd
df = pd.DataFrame
range(1, 16, 2)
index=pd.date_range('2018-01-01', '2018-01-15', freq='2D')
print(df)
print(df.diff())
print(df.diff(2))
print(df.shift(freq='2D'))
print(df - df.shift(freq='2D'))
print(df - df.shift(freq='D'))
import pandas as pd
print(pd.date_range('2018-01-01', '2018-12-31', freq='M'))
dtype='datetime64[ns]'
freq='M'
print(pd.date_range('2018-01-01', '2018-12-31', freq='MS'))
dtype='datetime64[ns]'
freq='MS'
print(pd.date_range('2018-01-01', '2018-12-31', freq='BMS'))
dtype='datetime64[ns]'
freq='BMS'
print(pd.date_range('2018-01-01', '2018-12-31', freq='SM'))
dtype='datetime64[ns]'
freq='SM-15'
print(pd.date_range('2018-01-01', '2018-12-31', freq='SMS'))
dtype='datetime64[ns]'
freq='SMS-15'
print(pd.date_range('2018-01-01', '2018-12-31', freq='W'))
dtype='datetime64[ns]'
freq='W-SUN'
print(pd.date_range('2018-01-01', '2018-12-31', freq='W-WED'))
dtype='datetime64[ns]'
freq='W-WED'
print(pd.date_range('2018-01-01', '2018-12-31', freq='QS'))
DatetimeIndex(['2018-01-01', '2018-04-01', '2018-07-01', '2018-10-01'], dtype='datetime64[ns]', freq='QS-JAN')
print(pd.date_range('2018-01-01', '2018-12-31', freq='QS-FEB'))
DatetimeIndex(['2018-02-01', '2018-05-01', '2018-08-01', '2018-11-01'], dtype='datetime64[ns]', freq='QS-FEB')
print(pd.date_range('2015-01-01', '2018-12-31', freq='A'))
DatetimeIndex(['2015-12-31', '2016-12-31', '2017-12-31', '2018-12-31'], dtype='datetime64[ns]', freq='A-DEC')
print(pd.date_range('2015-01-01', '2018-12-31', freq='A-JUL'))
DatetimeIndex(['2015-07-31', '2016-07-31', '2017-07-31', '2018-07-31'], dtype='datetime64[ns]', freq='A-JUL')
print(pd.date_range('2018-01-01', '2018-12-31', freq='WOM-4FRI'))
dtype='datetime64[ns]'
freq='WOM-4FRI'
print(pd.date_range('2018-01-01', '2018-12-31', freq='WOM-2MON'))
dtype='datetime64[ns]'
freq='WOM-2MON'
print(pd.date_range('2018-01-01', '2018-01-02', freq='H'))
dtype='datetime64[ns]'
freq='H'
print(pd.date_range('2018-01-01', '2018-12-31', freq='100D'))
DatetimeIndex(['2018-01-01', '2018-04-11', '2018-07-20', '2018-10-28'], dtype='datetime64[ns]', freq='100D')
print(pd.date_range('2018-01-01', '2018-12-31', freq='100B'))
DatetimeIndex(['2018-01-01', '2018-05-21', '2018-10-08'], dtype='datetime64[ns]', freq='100B')
print(pd.date_range('2018-01-01', '2018-12-31', freq='10W'))
dtype='datetime64[ns]'
freq='10W-SUN'
print(pd.date_range('2018-01-01', '2018-12-31', freq='10W-WED'))
dtype='datetime64[ns]'
freq='10W-WED'
print(pd.date_range('2018-01-01', '2018-12-31', freq='2M'))
dtype='datetime64[ns]'
freq='2M'
print(pd.date_range('2018-01-01', '2018-01-02', freq='90T'))
dtype='datetime64[ns]'
freq='90T'
print(pd.date_range('2018-01-01', '2018-01-10', freq='36H'))
dtype='datetime64[ns]'
freq='36H'
print(pd.date_range('2018-01-01', '2018-01-10', freq='1D12H'))
dtype='datetime64[ns]'
freq='36H'
print(pd.date_range('2018-01-01', '2018-01-2', freq='30min30S100ms100us'))
dtype='datetime64[ns]'
freq='1830100100U'
import pandas as pd
df = pd.DataFrame
index=pd.to_datetime(['2018-01-01', '2018-01-15', '2018-01-31'])
print(df)
print(df.asfreq('5D'))
print(df.asfreq('5D', fill_value=15))
print(df.asfreq('5D', method='pad'))
print(df.asfreq('5D', method='bfill'))
print(df.resample('5D').ffill())
print(df.resample('5D').bfill())
print(df.resample('5D').nearest())
print(df.resample('5D').interpolate())
df.loc['2018-01-15', 'value'] = 100
print(df)
print(df.resample('5D').interpolate())
print(df.resample('D').interpolate())
print(df.resample('D').interpolate().asfreq('5D'))
print(df.resample('D').interpolate('spline', order=2))
print(df.resample('D').interpolate('spline', order=2).asfreq('5D'))
df_nan = pd.DataFrame
pd.np.nan
pd.np.nan
pd.np.nan
index=pd.to_datetime(['2018-01-01', '2018-01-02', '2018-01-15', '2018-01-20', '2018-01-31'])
print(df_nan)
print(df_nan.interpolate())
print(df_nan.interpolate('time'))
import pandas as pd
df = pd.read_csv('data/src/sample_date.csv', index_col=0, parse_dates=True)
print(df)
print(type(df.index))
pandas.core.indexes.datetimes.DatetimeIndex
print(df.index.weekday)
Int64Index([2, 1, 5, 0, 1, 1, 4, 4, 2, 0, 4, 1], dtype='int64', name='date')
print(df.index.dayofweek)
Int64Index([2, 1, 5, 0, 1, 1, 4, 4, 2, 0, 4, 1], dtype='int64', name='date')
print(df.index.day_name())
dtype='object'
name='date'
print(df[df.index.weekday == 0])
print(df[df.index.weekday == 0].sum())
dtype: int64
print(df[df.index.weekday == 0].mean())
dtype: float64
print(df[df.index.weekday == 0].agg(['sum', 'mean']))
df_w = df.set_index([df.index.weekday, df.index])
print(df_w)
df_w.index.names = ['weekday', 'date']
print(df_w)
df_w.sort_index(inplace=True)
print(df_w)
print(df_w.sum())
dtype: int64
print(df_w.sum(level='weekday'))
print(df_w.mean(level='weekday'))
print(df_w.groupby(level='weekday').agg(['sum', 'mean']))
print(df.index.year)
dtype='int64'
name='date'
df_y = df.set_index([df.index.year, df.index])
df_y.index.names = ['year', 'date']
print(df_y)
print(df_y.sum(level='year'))
df_q = df.set_index([df.index.quarter, df.index])
df_q.index.names = ['quarter', 'date']
print(df_q)
print(df_q.sum(level='quarter'))
df_m = df.set_index([df.index.month, df.index])
df_m.index.names = ['month', 'date']
print(df_m)
print(df_m.sum(level='month'))
print(df.index.month_name())
dtype='object'
name='date'
df_w2 = df.set_index([df.index.week, df.index])
df_w2.index.names = ['week', 'date']
print(df_w2)
print(df_w2.sum(level='week'))
print(pd.date_range('2017-01-01', '2017-01-07').week)
Int64Index([52, 1, 1, 1, 1, 1, 1], dtype='int64')
print(pd.date_range('2017-01-01', '2017-01-07').day_name())
dtype='object'
df_yw = df.set_index([df.index.year, df.index.weekday, df.index])
df_yw.index.names = ['year', 'weekday', 'date']
df_yw.sort_index(inplace=True)
print(df_yw)
print(df_yw.sum(level='weekday'))
print(df_yw.sum(level=['year', 'weekday']))
print(df_yw.loc[(2017, 1), :])
print(df_yw.xs(1, level='weekday'))
print(df_yw.loc[(2017, [0, 4]), :])
print(df_yw.loc[pd.IndexSlice[:, [0, 4]], :])
df_yqmw = df.set_index([df.index.year, df.index.quarter, df.index.month, df.index.weekday, df.index])
df_yqmw.index.names = ['year', 'quarter', 'month', 'weekday', 'date']
df_yqmw.sort_index(inplace=True)
print(df_yqmw)
print(df_yqmw.sum(level='month'))
print(df_yqmw.sum(level='weekday'))
print(df_yqmw.sum(level=['quarter', 'weekday']))
print(df_yqmw.xs(1, level='weekday'))
print(df_yqmw.xs((1, 2017), level=('weekday', 'year')))
df_yqmw.loc
pd.IndexSlice
import pandas as pd
df = pd.DataFrame
range(1, 16, 2)
index=pd.date_range('2018-01-01', '2018-01-15', freq='2D')
print(df)
print(df.pct_change())
print(df.pct_change(2))
print(df.pct_change(freq='2D'))
print(df.pct_change(freq='D'))
import pandas as pd
df = pd.DataFrame
range(1, 32, 2)
index=pd.date_range('2018-08-01', '2018-08-31', freq='2D')
print(df)
print(df.resample('W'))
Week: weekday=6
axis=0
closed=right
label=right
convention=start
base=0
print(type(df.resample('W')))
pandas.core.resample.DatetimeIndexResampler
print(df.resample('W').mean())
print(df.resample('W').median())
print(df.resample('W').sum())
print(df.resample('W').first())
print(df.resample('W').last())
print(df.resample('W').count())
print(df.resample('W').ohlc())
print(df.resample('W').apply(list))
print(df.resample('W').agg(['min', 'max', 'sum']))
print(df.resample('W').apply(list))
print(df.resample('W', label='left').apply(list))
print(df.resample('W', label='left', closed='left').apply(list))
df_reset = df.reset_index()
print(df_reset)
print(df_reset.resample('W', on='index').sum())
import pandas as pd
df = pd.read_csv('data/src/sample_date.csv', index_col=0, parse_dates=True)
print(df)
print(type(df.index))
pandas.core.indexes.datetimes.DatetimeIndex
print(df.resample('W').sum())
print(df.resample('M').sum())
print(df.resample('Q').sum())
print(df.resample('Y').sum())
print(df.resample('10D').sum())
print(df.resample('3W').sum())
print(df.resample('Y').mean())
print(df.resample('Y').agg(['sum', 'mean', 'max', 'min']))
import pandas as pd
df = pd.DataFrame
range(1, 32, 2)
index=pd.date_range('2018-01-01', '2018-01-31', freq='2D')
print(df)
print(df.rolling(5).mean())
print(df.rolling('5D').mean())
print(df.resample('5D').mean())
import pandas as pd
df = pd.DataFrame
range(1, 16, 2)
index=pd.date_range('2018-01-01', '2018-01-15', freq='2D')
print(df)
print(df.shift())
print(df.shift(2))
print(df.shift(freq='D'))
print(df.shift(freq='3D'))
print(df.shift(3, freq='D'))
print(df.shift(freq='W'))
print(df.shift(freq='M'))
print(df.shift(freq='7D'))
print(df.shift(freq='31D'))
df_m = pd.DataFrame
range(1, 13)
index=pd.date_range('2018-01-01', '2018-12-31', freq='M')
print(df_m)
print(df_m.shift(freq='M'))
import pandas as pd
df = pd.DataFrame({'a': [0, 1, 2], 'b': [3, 4, 5]})
print(df)
df.to_clipboard()
df.to_clipboard(excel=False)
df.to_clipboard(sep=',')
import pandas as pd
df = pd.read_csv('data/src/sample_pandas_normal.csv', index_col=0)
print(df)
df.to_csv('data/dst/to_csv_out.csv')
df.to_csv('data/dst/to_csv_out_columns.csv', columns=['age'])
df.to_csv('data/dst/to_csv_out_header_index.csv', header=False, index=False)
df.to_csv('data/dst/to_csv_out.tsv', sep='\t')
df.to_csv('data/dst/to_csv_out_a.csv')
df.to_csv('data/dst/to_csv_out_a.csv', mode='a', header=False)
df.to_csv('data/dst/to_csv_out_a_new_column.csv')
df = pd.read_csv('data/dst/to_csv_out_a_new_column.csv', index_col=0)
print(df)
df['new_col'] = 'new data'
print(df)
df.to_csv('data/dst/to_csv_out_a_new_column.csv')
import pandas as pd
df = pd.DataFrame
print(df)
print(df.dtypes)
dtype: object
print(df.iat[0, 0])
print(df.iat[1, 0])
df.to_csv('data/dst/to_csv_out_float_default.csv')
print('%.3f' % 0.123456789)
print('%.3f' % 123456789)
df.to_csv('data/dst/to_csv_out_float_format_3f.csv', float_format='%.3f')
print('%.3e' % 0.123456789)
print('%.3e' % 123456789)
df.to_csv('data/dst/to_csv_out_float_format_3e.csv', float_format='%.3e')
df['col1'] = df['col1'].map('{:.3f}'.format)
df['col2'] = df['col2'].map('{:.3e}'.format)
df['col3'] = df['col3'].map('{:#010x}'.format)
print(df)
print(df.dtypes)
dtype: object
df.to_csv('data/dst/to_csv_out_float_format_str.csv')
df = pd.read_csv('data/dst/to_csv_out_float_format_str.csv', index_col=0)
print(df)
print(df.dtypes)
dtype: object
df['col3'] = df['col3'].map(lambda x: int(x, 16))
print(df)
print(df.dtypes)
dtype: object
import pandas as pd
import pprint
from collections import OrderedDict
df = pd.DataFrame
index=['row1', 'row2', 'row3']
print(df)
d = df.to_dict()
pprint.pprint(d)
print(type(d))
d_dict = df.to_dict(orient='dict')
pprint.pprint(d_dict)
print(d_dict['col1'])
print(type(d_dict['col1']))
d_list = df.to_dict(orient='list')
pprint.pprint(d_list)
print(d_list['col1'])
print(type(d_list['col1']))
d_series = df.to_dict(orient='series')
pprint.pprint(d_series)
Name: col1
dtype: int64
Name: col2
dtype: object
print(d_series['col1'])
Name: col1
dtype: int64
print(type(d_series['col1']))
pandas.core.series.Series
d_split = df.to_dict(orient='split')
pprint.pprint(d_split)
print(d_split['columns'])
print(type(d_split['columns']))
l_records = df.to_dict(orient='records')
pprint.pprint(l_records)
print(type(l_records))
print(l_records[0])
print(type(l_records[0]))
d_index = df.to_dict(orient='index')
pprint.pprint(d_index)
print(d_index['row1'])
print(type(d_index['row1']))
od = df.to_dict(into=OrderedDict)
pprint.pprint(od)
OrderedDict([('row1', 1), ('row2', 2), ('row3', 3)])
OrderedDict([('row1', 'a'), ('row2', 'x'), ('row3', 'あ')])
print(type(od))
collections.OrderedDict
print(od['col1'])
OrderedDict([('row1', 1), ('row2', 2), ('row3', 3)])
print(type(od['col1']))
collections.OrderedDict
print(df.index)
Index(['row1', 'row2', 'row3'], dtype='object')
print(df['col1'])
Name: col1
dtype: int64
d_col = dict(zip(df.index, df['col1']))
print(d_col)
d_row = dict(zip(df.columns, df.loc['row1']))
print(d_row)
import pandas as pd
from collections import OrderedDict
df = pd.DataFrame
index=['row1', 'row2', 'row3']
print(df)
s = df['col1']
print(s)
Name: col1
dtype: int64
print(type(s))
pandas.core.series.Series
d = s.to_dict()
print(d)
print(type(d))
s = df.set_index('col2')['col1']
print(s)
Name: col1
dtype: int64
d = s.to_dict()
print(d)
od = df['col1'].to_dict(OrderedDict)
print(od)
OrderedDict([('row1', 1), ('row2', 2), ('row3', 3)])
print(type(od))
collections.OrderedDict
import pandas as pd
print(pd.__version__)
df = pd.DataFrame
index=['one', 'two', 'three']
columns=['a', 'b', 'c']
print(df)
df.to_excel('data/dst/pandas_to_excel.xlsx', sheet_name='new_sheet_name')
df.to_excel
index=False
header=False
df2 = df[['a', 'c']]
print(df2)
with pd.ExcelWriter('data/dst/pandas_to_excel_multi.xlsx') as writer:
df.to_excel(writer, sheet_name='sheet1')
df2.to_excel(writer, sheet_name='sheet2')
with pd.ExcelWriter('data/dst/pandas_to_excel.xlsx', mode='a') as writer:
df.to_excel(writer, sheet_name='new_sheet1')
df2.to_excel(writer, sheet_name='new_sheet2')
import pandas as pd
import pprint
import json
df = pd.DataFrame
index=['row1', 'row2', 'row3']
print(df)
print(df.to_json())
print(type(df.to_json()))
path = 'data/src/sample_from_pandas_columns.json'
df.to_json(path)
with open(path) as f:
s = f.read()
print(s)
print(type(s))
with open(path, encoding='unicode-escape') as f:
s = f.read()
print(s)
print(type(s))
with open(path) as f:
d = json.load(f)
print(d)
print(type(d))
df_read = pd.read_json(path)
print(df_read)
df.to_json('data/src/sample_from_pandas_columns.gz', compression='gzip')
print(df.to_json(orient='split'))
pprint.pprint(json.loads(df.to_json(orient='split')))
print(df.to_json(orient='records'))
pprint.pprint(json.loads(df.to_json(orient='records')), width=40)
print(df.to_json(orient='records', lines=True))
print(df.to_json(orient='index'))
pprint.pprint(json.loads(df.to_json(orient='index')))
print(df.to_json(orient='columns'))
pprint.pprint(json.loads(df.to_json(orient='columns')))
print(df.to_json(orient='values'))
pprint.pprint(json.loads(df.to_json(orient='values')))
print(df.to_json(orient='table'))
pprint.pprint(json.loads(df.to_json(orient='table')))
print(df.to_json(force_ascii=False))
import numpy as np
import pandas as pd
df = pd.DataFrame(data=[[1, 2, 3], [4, 5, 6]], columns=['a', 'b', 'c'])
print(df)
a_df = df.values
print(a_df)
print(type(a_df))
numpy.ndarray
print(a_df.dtype)
s = df['a']
print(s)
Name: a
dtype: int64
a_s = s.values
print(a_s)
print(type(a_s))
numpy.ndarray
print(a_s.dtype)
df_f = pd.DataFrame([[0.1, 0.2, 0.3], [0.4, 0.5, 0.6]])
print(df_f)
a_df_f = df_f.values
print(a_df_f)
print(type(a_df_f))
numpy.ndarray
print(a_df_f.dtype)
df_multi = pd.read_csv('data/src/sample_pandas_normal.csv')
print(df_multi)
a_df_multi = df_multi.values
print(a_df_multi)
print(type(a_df_multi))
numpy.ndarray
print(a_df_multi.dtype)
a_df_int = df_multi[['age', 'point']].values
print(a_df_int)
print(type(a_df_int))
numpy.ndarray
print(a_df_int.dtype)
print(a_df_int.T)
a_df_int = df_multi.select_dtypes(include=int).values
print(a_df_int)
print(type(a_df_int))
numpy.ndarray
print(a_df_int.dtype)
import numpy as np
import pandas as pd
df = pd.DataFrame(data=[[1, 2, 3], [4, 5, 6]], columns=['a', 'b', 'c'])
print(df)
a_values = df.values
print(a_values)
print(np.shares_memory(a_values, df))
a_values[0, 0] = 100
print(a_values)
print(df)
df_if = pd.DataFrame(data=[[1, 0.1], [2, 0.2]], columns=['int', 'float'])
print(df_if)
print(df_if.dtypes)
dtype: object
a_values_if = df_if.values
print(a_values_if)
print(np.shares_memory(a_values_if, df_if))
a_values_if[0, 0] = 100
print(a_values_if)
print(df_if)
print(df[['a', 'c']].values)
print(np.shares_memory(df[['a', 'c']].values, df))
print(df.iloc[:, ::2].values)
print(np.shares_memory(df.iloc[:, ::2].values, df))
a_values_copy = df.values.copy()
print(a_values_copy)
print(np.shares_memory(a_values_copy, df))
a_values_copy[0, 0] = 10
print(a_values_copy)
print(df)
a = np.array([[1, 2, 3], [4, 5, 6]])
print(a)
df_a = pd.DataFrame(a, columns=['a', 'b', 'c'])
print(df_a)
print(np.shares_memory(a, df_a))
a[0, 0] = 100
print(a)
print(df_a)
df_a.iat[1, 0] = 10
print(df_a)
print(a)
df_a_copy = pd.DataFrame(a.copy(), columns=['a', 'b', 'c'])
print(df_a_copy)
a[0, 0] = 1
print(a)
print(df_a_copy)
import numpy as np
import pandas as pd
df = pd.DataFrame(data=[[1, 2, 3], [4, 5, 6]], columns=['a', 'b', 'c'])
print(df)
a = df.to_numpy()
print(a)
print(type(a))
numpy.ndarray
print(np.shares_memory(df, a))
a[0, 0] = 100
print(a)
print(df)
a_copy = df.to_numpy(copy=True)
print(a_copy)
print(np.shares_memory(df, a_copy))
a_copy[0, 0] = 10
print(a_copy)
print(df)
a_cols = df[['a', 'c']].to_numpy()
print(a_cols)
print(np.shares_memory(df, a_cols))
a_f = df.to_numpy(dtype=float)
print(a_f)
print(np.shares_memory(df, a_f))
import pandas as pd
df = pd.DataFrame
index=pd.date_range('2018-01-01', '2018-01-04', freq='D')
print(df)
print(df.index)
DatetimeIndex(['2018-01-01', '2018-01-02', '2018-01-03', '2018-01-04'], dtype='datetime64[ns]', freq='D')
print(type(df['list'][0]))
df.to_csv('data/dst/pandas_obj.csv')
df_from_csv = pd.read_csv('data/dst/pandas_obj.csv', index_col=0, parse_dates=True)
print(df_from_csv)
print(df_from_csv.index)
DatetimeIndex(['2018-01-01', '2018-01-02', '2018-01-03', '2018-01-04'], dtype='datetime64[ns]', freq=None)
print(type(df_from_csv['list'][0]))
df_from_csv['list'] = df_from_csv['list'].apply(eval)
print(df_from_csv)
print(type(df_from_csv['list'][0]))
df.to_pickle('data/dst/pandas_obj.pkl')
df_from_pkl = pd.read_pickle('data/dst/pandas_obj.pkl')
print(df_from_pkl)
print(df_from_pkl.index)
DatetimeIndex(['2018-01-01', '2018-01-02', '2018-01-03', '2018-01-04'], dtype='datetime64[ns]', freq='D')
print(type(df_from_pkl['list'][0]))
df.to_pickle('data/dst/pandas_obj.zip')
df_from_pkl_zip = pd.read_pickle('data/dst/pandas_obj.zip')
print(df_from_pkl_zip)
import pandas as pd
df = pd.DataFrame({'X': [0, 1, 2], 'Y': [3, 4, 5]}, index=['A', 'B', 'C'])
print(df)
print(df.T)
print(df.transpose())
df = df.T
print(df)
df = pd.DataFrame({'X': [0, 1, 2], 'Y': [3, 4, 5]}, index=['A', 'B', 'C'])
print(df)
print(df.dtypes)
dtype: object
print(df.T)
print(df.T.dtypes)
dtype: object
df_mix = pd.DataFrame({'col_int': [0, 1, 2], 'col_float': [0.1, 0.2, 0.3]}, index=['A', 'B', 'C'])
print(df_mix)
print(df_mix.dtypes)
dtype: object
print(df_mix.T)
print(df_mix.T.dtypes)
dtype: object
print(df_mix.T.T)
print(df_mix.T.T.dtypes)
dtype: object
df_mix2 = pd.DataFrame
index=['A', 'B', 'C']
print(df_mix2)
print(df_mix2.dtypes)
dtype: object
print(df_mix2.T)
print(df_mix2.T.dtypes)
dtype: object
print(df_mix2.T.T)
print(df_mix2.T.T.dtypes)
dtype: object
df = pd.DataFrame({'X': [0, 1, 2], 'Y': [3, 4, 5]}, index=['A', 'B', 'C'])
print(df)
df_T = df.T
print(df_T)
df_transpose = df.transpose()
print(df_transpose)
df.at['A', 'X'] = 100
print(df)
print(df_T)
print(df_transpose)
df_mix = pd.DataFrame({'col_int': [0, 1, 2], 'col_float': [0.1, 0.2, 0.3]}, index=['A', 'B', 'C'])
print(df_mix)
df_mix_T = df_mix.T
print(df_mix_T)
df_mix_transpose = df_mix.transpose()
print(df_mix_transpose)
df_mix.at['A', 'col_int'] = 100
print(df_mix)
print(df_mix_T)
print(df_mix_transpose)
df = pd.DataFrame({'X': [0, 1, 2], 'Y': [3, 4, 5]}, index=['A', 'B', 'C'])
print(df)
df_T_copy = df.T.copy()
print(df_T_copy)
df_transpose_copy = df.transpose(copy=True)
print(df_transpose_copy)
df.at['A', 'X'] = 100
print(df)
print(df_T_copy)
print(df_transpose_copy)
import pandas as pd
s = '2018-01-01T12:00+09:00'
print(s)
print(type(s))
ts = pd.to_datetime(s)
print(ts)
print(type(ts))
pandas._libs.tslibs.timestamps.Timestamp
print(ts.tz)
pytz.FixedOffset(540)
ts_utc = pd.to_datetime(s, utc=True)
print(ts_utc)
print(ts_utc.tz)
s_without_tz = '2018-01-01T12:00'
ts_naive = pd.to_datetime(s_without_tz)
print(ts_naive)
print(ts_naive.tz)
ts_set_utc = pd.to_datetime(s_without_tz, utc=True)
print(ts_set_utc)
print(ts_set_utc.tz)
print(ts_utc)
print(ts_utc.tz)
ts_jst = ts_utc.tz_convert('Asia/Tokyo')
print(ts_jst)
print(ts_jst.tz)
print(ts_utc.value)
print(ts_jst.value)
print(ts_utc == ts_jst)
ts_pst = ts_utc.tz_convert('US/Pacific')
print(ts_pst)
print(ts_pst.tz)
print(ts_utc.tz_convert('America/Los_Angeles'))
print(ts_utc.tz_convert('America/Vancouver'))
print(ts_naive)
print(ts_naive.tz)
print(ts_naive.tz_convert('Asia/Tokyo'))
TypeError: Cannot
ts_jst_localize = ts_naive.tz_localize('Asia/Tokyo')
print(ts_jst_localize)
print(ts_jst_localize.tz)
print(ts_naive.tz_localize('US/Pacific'))
print(ts_naive.tz_localize('Asia/Tokyo') == ts_naive.tz_localize('US/Pacific'))
print(ts_jst)
print(ts_jst.tz)
print(ts_jst.tz_localize('US/Pacific'))
TypeError: Cannot
print(ts_jst)
print(ts_jst.tz)
print(ts_jst.tz_convert(None))
print(ts_jst.tz_localize(None))
df = pd.DataFrame
print(df)
s_naive = pd.to_datetime(df['date'])
print(s_naive)
Name: date
dtype: datetime64[ns]
print(s_naive[0])
print(type(s_naive[0]))
pandas._libs.tslibs.timestamps.Timestamp
print(s_naive[0].tz)
s_utc = pd.to_datetime(df['date'], utc=True)
print(s_utc)
Name: date
dtype: datetime64[ns, UTC]
print(s_utc[0].tz)
print(s_naive.tz_localize('Asia/Tokyo'))
TypeError: index
not a 
DatetimeIndex or PeriodIndex
print(s_utc.tz_convert('Asia/Tokyo'))
TypeError: index
not a 
DatetimeIndex or PeriodIndex
print(s_naive.dt.tz_localize('Asia/Tokyo'))
Name: date
dtype: datetime64
print(s_utc.dt.tz_convert('Asia/Tokyo'))
Name: date
dtype: datetime64
print(s_naive.dt.tz_convert('Asia/Tokyo'))
TypeError: Cannot
print(s_utc.dt.tz_localize('Asia/Tokyo'))
TypeError: Already
print(df['date'].dt.tz_localize('Asia/Tokyo'))
AttributeError: Can
use .dt
df['date'] = pd.to_datetime(df['date'])
df_ts = df.set_index('date')
print(df_ts)
print(df_ts.index)
dtype='datetime64[ns]'
name='date'
freq=None
print(type(df_ts.index))
pandas.core.indexes.datetimes.DatetimeIndex
print(df_ts['2018-01-03'])
print(df_ts.index.tz_localize('Asia/Tokyo'))
dtype='datetime64[ns, Asia/Tokyo]'
name='date'
freq=None
print(df_ts.tz_localize('Asia/Tokyo'))
s_ts = df_ts['value']
print(s_ts)
Name: value
dtype: object
print(s_ts.tz_localize('Asia/Tokyo'))
Name: value
dtype: object
df = pd.DataFrame
print(df)
print(pd.to_datetime(df['date']))
Name: date
dtype: datetime64
pytz.FixedOffset(540)
print(pd.to_datetime(df['date'], utc=True))
Name: date
dtype: datetime64[ns, UTC]
print(pd.to_datetime(df['date']).dt.tz_convert('US/Pacific'))
Name: date
dtype: datetime64
df['date'] = pd.to_datetime(df['date'])
df_ts = df.set_index('date')
print(df_ts)
print(df_ts.index)
dtype='datetime64[ns, pytz.FixedOffset(540)]'
name='date'
freq=None
print(df_ts.tz_convert('US/Pacific'))
df = pd.DataFrame
print(df)
print(pd.to_datetime(df['date']))
Name: date
dtype: object
print(type(pd.to_datetime(df['date'])[0]))
datetime.datetime
print(pd.to_datetime(df['date'])[0].tzinfo)
tzoffset(None, 32400)
print(pd.to_datetime(df['date'])[2].tzinfo)
tzoffset(None, -18000)
print(pd.to_datetime(df['date'], utc=True))
Name: date
dtype: datetime64[ns, UTC]
print(type(pd.to_datetime(df['date'], utc=True)[0]))
pandas._libs.tslibs.timestamps.Timestamp
print(pd.to_datetime(df['date']).dt.tz_convert('Asia/Tokyo'))
ValueError: Tz
-aware
datetime.datetime
utc=True
df['date'] = pd.to_datetime(df['date'])
df_dt = df.set_index('date')
print(df_dt)
print(df_dt.index)
dtype='object'
name='date'
print(df_dt.tz_convert('Asia/Tokyo'))
TypeError: index
not a 
DatetimeIndex or PeriodIndex
print(df_dt.tz_localize('Asia/Tokyo'))
TypeError: index
not a 
DatetimeIndex or PeriodIndex
import pandas as pd
import numpy as np
df = pd.read_csv('data/src/sample_pandas_normal.csv')
df.iloc[1] = np.nan
print(df)
u = df['state'].unique()
print(u)
print(type(u))
numpy.ndarray
vc = df['state'].value_counts()
print(vc)
print(type(vc))
Name: state
dtype: int64
pandas.core.series.Series
print(df['state'].value_counts(ascending=True))
Name: state
dtype: int64
print(df['state'].value_counts(sort=False))
Name: state
dtype: int64
print(df['state'].value_counts(dropna=False))
Name: state
dtype: int64
print(df['state'].value_counts(dropna=False, normalize=True))
Name: state
dtype: float64
nu = df['state'].nunique()
print(nu)
print(type(nu))
print(df['state'].nunique(dropna=False))
nu_col = df.nunique()
print(nu_col)
print(type(nu_col))
dtype: int64
pandas.core.series.Series
print(df.nunique(dropna=False))
dtype: int64
print(df.nunique(dropna=False, axis='columns'))
dtype: int64
print(df['state'].nunique())
print(df.nunique())
dtype: int64
print(df['state'].unique().tolist())
print(type(df['state'].unique().tolist()))
print(df['state'].value_counts().index.tolist())
print(type(df['state'].value_counts().index.tolist()))
print(df['state'].value_counts(dropna=False).index.values)
print(type(df['state'].value_counts().index.values))
numpy.ndarray
print(df['state'].value_counts()['NY'])
print(df['state'].value_counts().NY)
for index, value in df['state'].value_counts().iteritems():
print(index, ': ', value)
d = df['state'].value_counts().to_dict()
print(d)
print(type(d))
print(d['NY'])
for key, value in d.items():
print(key, ': ', value)
print(df['state'].value_counts())
Name: state
dtype: int64
print(df['state'].value_counts().index[0])
print(df['state'].value_counts().iat[0])
print(df.apply(lambda x: x.value_counts().index[0]))
dtype: object
print(df.apply(lambda x: x.value_counts().iat[0]))
dtype: int64
print(df['state'].mode())
dtype: object
print(df['state'].mode().tolist())
print(df['age'].mode().tolist())
s_mode = df.apply(lambda x: x.mode().tolist())
print(s_mode)
age                                    [24.0]
state                                [CA, NY]
point                                  [70.0]
dtype: object
print(type(s_mode))
pandas.core.series.Series
print(s_mode['name'])
print(type(s_mode['name']))
print(df.mode())
print(df.mode().count())
dtype: int64
print(df.astype('str').describe())
print(df.astype('str').describe().loc['top'])
Name: top
dtype: object
import pandas as pd
print(pd.__version__)
pd.show_versions()
commit: None
OS: Darwin
machine: x86_64
processor: i386
byteorder: little
LC_ALL: None
LANG: ja_JP
LOCALE: ja_JP
pytest: None
pyarrow: None
xarray: None
sphinx: None
patsy: None
blosc: None
bottleneck: None
tables: None
numexpr: None
feather: None
xlwt: None
xlsxwriter: None
pymysql: None
psycopg2: None
s3fs: None
fastparquet: None
pandas_gbq: None
import pandas as pd
import numpy as np
df_homo = pd.DataFrame({'a': [0, 1, 2], 'b': [3, 4, 5]})
print(df_homo)
print(df_homo.dtypes)
dtype: object
df_homo_slice = df_homo.iloc[:2]
print(df_homo_slice)
print(np.shares_memory(df_homo, df_homo_slice))
print(df_homo_slice._is_view)
df_homo_list = df_homo.iloc[[0, 1]]
print(df_homo_list)
print(np.shares_memory(df_homo, df_homo_list))
print(df_homo_list._is_view)
print(df_homo['a'] < 2)
Name: a
dtype: bool
df_homo_bool = df_homo.loc[df_homo['a'] < 2]
print(df_homo_bool)
print(np.shares_memory(df_homo, df_homo_bool))
print(df_homo_bool._is_view)
s_homo_scalar = df_homo.iloc[0]
print(s_homo_scalar)
dtype: int64
print(np.shares_memory(df_homo, s_homo_scalar))
print(s_homo_scalar._is_view)
s_homo_col = df_homo['a']
print(s_homo_col)
Name: a
dtype: int64
print(np.shares_memory(df_homo, s_homo_col))
print(s_homo_col._is_view)
df_homo_col_multi = df_homo[['a', 'b']]
print(df_homo_col_multi)
print(np.shares_memory(df_homo, df_homo_col_multi))
print(df_homo_col_multi._is_view)
df_homo.iat[0, 0] = 100
print(df_homo)
print(df_homo_slice)
print(df_homo_list)
print(df_homo_bool)
print(s_homo_scalar)
dtype: int64
print(s_homo_col)
Name: a
dtype: int64
print(df_homo_col_multi)
df_hetero = pd.DataFrame({'a': [0, 1, 2], 'b': ['x', 'y', 'z']})
print(df_hetero)
print(df_hetero.dtypes)
dtype: object
df_hetero_slice = df_hetero.iloc[:2]
print(df_hetero_slice)
print(np.shares_memory(df_hetero, df_hetero_slice))
print(df_hetero_slice._is_view)
df_hetero_slice2 = df_hetero.iloc[:2, 0:]
print(df_hetero_slice2)
print(np.shares_memory(df_hetero, df_hetero_slice2))
print(df_hetero_slice2._is_view)
df_hetero_list = df_hetero.iloc[[0, 1]]
print(df_hetero_list)
print(np.shares_memory(df_hetero, df_hetero_list))
print(df_hetero_list._is_view)
df_hetero_bool = df_hetero.loc[df_hetero['a'] < 2]
print(df_hetero_bool)
print(df_hetero_bool._is_view)
print(df_hetero_bool._is_view)
s_hetero_scalar = df_hetero.iloc[0]
print(s_hetero_scalar)
dtype: object
print(np.shares_memory(df_hetero, s_hetero_scalar))
print(s_hetero_scalar._is_view)
s_hetero_col = df_hetero['a']
print(s_hetero_col)
Name: a
dtype: int64
print(np.shares_memory(df_hetero, s_hetero_col))
print(s_hetero_col._is_view)
df_hetero_col_multi = df_hetero[['a', 'b']]
print(df_hetero_col_multi)
print(np.shares_memory(df_hetero, df_hetero_col_multi))
print(df_hetero_col_multi._is_view)
df_hetero.iat[0, 0] = 100
print(df_hetero)
print(df_hetero_slice)
print(df_hetero_slice2)
print(df_hetero_list)
print(df_hetero_bool)
print(s_hetero_scalar)
dtype: object
print(s_hetero_col)
Name: a
dtype: int64
print(df_hetero_col_multi)
import pandas as pd
import numpy as np
df = pd.DataFrame
-20
-10
print(df)
df.loc[df['A'] < 0, 'A'] = -100
df.loc
print(df)
print(df['A'] < 0)
Name: A
dtype: bool
Name: A
dtype: bool
print(df.loc[df['A'] < 0, 'A'])
Name: A
dtype: int64
df.loc[df['A'] < 0, 'A'] = -10
print(df)
df.loc
df['B']
print(df)
df.loc[df['B'] % 2 == 0, 'D'] = 'even'
df.loc[df['B'] % 2 != 0, 'D'] = 'odd'
print(df)
df.loc
df['C'] == 'b'
print(df)
df.loc
print(df)
df.loc[df['C'] == 'a', 'F'] = df['A']
df.loc[df['C'] == 'b', 'F'] = df['B']
print(df)
df['F'] = df['F'].astype(int)
print(df)
df.loc[df['C'] == 'a', ['E', 'F']] = 100
print(df)
print(df < 0)
print(df[df < 0])
df[df < 0] = 0
TypeError: Cannot
np.nan
df = pd.DataFrame
-20
-10
print(df)
print(df['A'].where(df['C'] == 'a'))
Name: A
dtype: float64
print(df['A'].where(df['C'] == 'a', 100))
Name: A
dtype: int64
print(df['A'].where(df['C'] == 'a', df['B']))
Name: A
dtype: int64
df['D'] = df['A'].where(df['C'] == 'a', df['B'])
print(df)
df['D'].where((df['D'] % 2 == 0) & (df['A'] < 0), df['D'] * 100, inplace=True)
print(df)
print(df < 0)
print(df.where(df < 0))
print(df.where(df < 0, df * 2))
print(df.where(df < 0, 100))
df = pd.DataFrame
-20
-10
print(df)
print(df['C'].mask(df['C'] == 'a'))
Name: C
dtype: object
print(df['C'].mask(df['C'] == 'a', 100))
Name: C
dtype: object
df['D'] = df['A'].mask(df['C'] == 'a', df['B'])
print(df)
df['D'].mask(df['D'] % 2 != 0, df['D'] * 10, inplace=True)
print(df)
print(df.mask(df < 0, -100))
print(df.select_dtypes(include='number').mask(df < 0, -100))
df_mask = df.select_dtypes(include='number').mask(df < 0, -100)
df_mask = pd.concat([df_mask, df.select_dtypes(exclude='number')], axis=1)
print(df_mask.sort_index(axis=1))
df = pd.DataFrame
-20
-10
print(df)
print(np.where(df['B'] % 2 == 0, 'even', 'odd'))
print(np.where(df['C'] == 'a', df['A'], df['B']))
-20
df['D'] = np.where(df['B'] % 2 == 0, 'even', 'odd')
print(df)
df['E'] = np.where(df['C'] == 'a', df['A'], df['B'])
print(df)
print(np.where(df < 0, df, 100))
-20
-20
-10
df_np_where = pd.DataFrame
np.where(df < 0, df, 100)
index=df.index
columns=df.columns
print(df_np_where)
for i in range(3):
print(i)
if i == 1:
continue
print('CONTINUE')
for i in range(3):
print(i)
if i == 1:
pass
print('PASS')
def empty_func():
SyntaxError: unexpected
def empty_func():
pass
EmptyClass()
SyntaxError: unexpected
EmptyClass()
pass
def empty_func_one_line(): 
pass
EmptyClassOneLine()
pass
def divide(a, b):
print(a / b)
divide(1, 0)
ZeroDivisionError: division
def divide_exception(a, b):
print(a / b)
except ZeroDivisionError as e:
print('ZeroDivisionError: ', e)
divide_exception(1, 0)
ZeroDivisionError:  division
def divide_exception_pass(a, b):
print(a / b)
except ZeroDivisionError as e:
pass
divide_exception_pass(1, 0)
a = 3
if a % 2 == 0:
print('Even')
pass
open('temp/empty.txt', 'w')
pass
open('temp/empty.txt', 'w')
pass
import util_make_files
util_make_files.pathlib_basic()
import pathlib
import os
p = pathlib.Path('temp/file.txt')
print(p)
print(type(p))
pathlib.PosixPath
print(p.cwd())
print(type(p.cwd()))
pathlib.PosixPath
print(pathlib.Path.cwd())
print(type(pathlib.Path.cwd()))
pathlib.PosixPath
print(os.getcwd())
print(type(os.getcwd()))
print(p.resolve())
p_rel = pathlib.Path('temp/dir/../file.txt')
print(p_rel)
file.txt
print(p_rel.resolve())
p_abs = pathlib.Path('/Users/mbp/Documents/my-project/python-snippets/notebook/temp/file.txt')
print(p_abs)
print(p_abs.relative_to(p.cwd()))
print(p_abs.relative_to('/Users/mbp/Documents/my-project'))
print(p_abs.relative_to('/usr/'))
not start 
p_rel = pathlib.Path('temp/dir/sub_dir/file2.txt')
print(p_rel.relative_to('temp/dir'))
print(p_abs)
print(p_abs.is_absolute())
print(p_rel)
print(p_rel.is_absolute())
import shutil
shutil.rmtree('temp')
import util_make_files
util_make_files.pathlib_basic()
import pathlib
import os
import pprint
p_file = pathlib.Path('temp/file.txt')
print(p_file)
print(type(p_file))
pathlib.PosixPath
p_dir = pathlib.Path('temp/dir')
print(p_dir)
print(type(p_dir))
pathlib.PosixPath
print(p_file.is_file())
print(p_dir.is_file())
p_new_file = pathlib.Path('temp/new_file.txt')
print(p_new_file.exists())
p_new_file.touch()
print(p_new_file.exists())
pathlib.Path('temp/new_file2.txt').touch()
pprint.pprint(list(pathlib.Path('temp').iterdir()))
PosixPath('temp/file.txt')
PosixPath('temp/new_file.txt')
PosixPath('temp/new_file2.txt')
PosixPath('temp/dir')
p_sub_dir_file = p_dir / 'sub_dir' / 'file2.txt'
print(p_sub_dir_file)
print(p_sub_dir_file.is_file())
p_sub_dir_file = p_dir.joinpath('sub_dir', 'file2.txt')
print(p_sub_dir_file)
print(p_sub_dir_file.is_file())
p_file_join = p_dir.joinpath('..', 'file.txt')
print(p_file_join)
file.txt
print(p_file)
print(p_file.samefile(p_file_join))
print(p_file == p_file_join)
print(p_file_join.resolve())
print(p_file.resolve())
print(p_file_join.resolve() == p_file.resolve())
print(type(p_file.resolve()))
pathlib.PosixPath
print(p_file_join.resolve().relative_to(pathlib.Path.cwd()))
print(p_dir.parent)
print(p_dir.parent.joinpath('file.txt'))
print(p_file_join.parent)
print(p_file.parent)
s = str(p_file)
print(s)
print(type(s))
print(os.path.isfile('temp/file.txt'))
print(os.path.isfile(p_file))
import shutil
shutil.rmtree('temp')
import pathlib
p = pathlib.Path('temp')
print(p)
print(type(p))
pathlib.PosixPath
print(p.exists())
p.mkdir()
print(p.exists())
print(p.is_dir())
pathlib.Path('temp/dir').mkdir()
print(pathlib.Path('temp/dir').is_dir())
pathlib.Path('temp/dir/sub_dir/sub_dir2').mkdir()
file or directory
pathlib.Path('temp/dir/sub_dir/sub_dir2').mkdir(parents=True)
print(pathlib.Path('temp/dir/sub_dir/sub_dir2').is_dir())
pathlib.Path('temp/dir').mkdir()
pathlib.Path('temp/dir').mkdir(exist_ok=True)
pathlib.Path('temp/dir/file').touch()
print(pathlib.Path('temp/dir/file').is_file())
pathlib.Path('temp/dir/file').mkdir(exist_ok=True)
p_sub_dir = pathlib.Path('temp/dir/sub_dir/sub_dir2')
print(p_sub_dir.is_dir())
p_sub_dir.rmdir()
print(p_sub_dir.exists())
p = pathlib.Path('temp')
p.rmdir()
not empty
import shutil
shutil.rmtree(p)
print(p.exists())
import pathlib
import os
os.makedirs('temp', exist_ok=True)
p_empty = pathlib.Path('temp/empty_file.txt')
print(p_empty)
print(type(p_empty))
pathlib.PosixPath
print(p_empty.exists())
p_empty.touch()
print(p_empty.exists())
p_empty.touch()
p_empty.touch(exist_ok=False)
pathlib.Path('temp/new_dir/empty_file.txt').touch()
file or directory
p_empty_new = pathlib.Path('temp/new_dir/empty_file.txt')
p_empty_new.parent.mkdir(parents=True, exist_ok=True)
p_empty_new.touch()
pathlib.Path('temp/empty_file2.txt').touch()
p_new = pathlib.Path('temp/new_file.txt')
print(p_new.exists())
with p_new.open(mode='w') as f:
f.write('line 1\nline 2\nline 3')
with p_new.open() as f:
print(f.read())
s = p_new.read_text()
print(s)
print(type(s))
i = p_new.write_text('new text')
print(i)
print(p_new.read_text())
p_new2 = pathlib.Path('temp/new_file2.txt')
print(p_new2.exists())
print(p_new2.read_text())
file or directory
print(p_new2.write_text('new text2'))
print(p_new2.read_text())
print(pathlib.Path('temp/new_dir2/new_file.txt').write_text('new_text'))
file or directory
p_text_new = pathlib.Path('temp/new_dir2/new_file.txt')
p_text_new.parent.mkdir(parents=True, exist_ok=True)
print(p_text_new.write_text('new_text'))
print(p_text_new.read_text())
print(pathlib.Path('temp/new_file3.txt').write_text('new_text3'))
print(pathlib.Path('temp/new_file3.txt').read_text())
p_empty = pathlib.Path('temp/empty_file.txt')
print(p_empty.exists())
p_empty.unlink()
print(p_empty.exists())
p_empty.unlink()
file or directory
p_dir = pathlib.Path('temp/')
p_dir.unlink()
not permitted
for p in p_dir.iterdir():
if p.is_file():
p.unlink()
p.unlink() for p in p_dir.iterdir() if p.is_file()
import shutil
shutil.rmtree('temp')
import util_make_files
util_make_files.glob_example_detail()
import pathlib
import glob
import re
import pprint
p_temp = pathlib.Path('temp')
print(p_temp)
print(type(p_temp))
pathlib.PosixPath
print(type(p_temp.iterdir()))
pprint.pprint(list(p_temp.iterdir()))
PosixPath('temp/[x].txt')
PosixPath('temp/aaa.text')
PosixPath('temp/dir')
PosixPath('temp/1.txt')
PosixPath('temp/12.text')
PosixPath('temp/123.txt')
print(list(pathlib.Path('temp/1.txt').iterdir()))
p_temp = pathlib.Path('temp')
print(type(p_temp.glob('**/*.txt')))
pprint.pprint(list(p_temp.glob('**/*.txt')))
PosixPath('temp/[x].txt')
PosixPath('temp/1.txt')
PosixPath('temp/123.txt')
PosixPath('temp/dir/bbb.txt')
PosixPath('temp/dir/sub_dir1/98.txt')
pprint.pprint(list(p_temp.glob('*')))
PosixPath('temp/[x].txt')
PosixPath('temp/aaa.text')
PosixPath('temp/dir')
PosixPath('temp/1.txt')
PosixPath('temp/12.text')
PosixPath('temp/123.txt')
pprint.pprint(list(p_temp.glob('dir/*/*.text')))
PosixPath('temp/dir/sub_dir1/ccc.text')
PosixPath('temp/dir/sub_dir2/ddd.text')
pprint.pprint(list(p_temp.glob('???.*')))
PosixPath('temp/[x].txt')
PosixPath('temp/aaa.text')
PosixPath('temp/123.txt')
pprint.pprint(list(p_temp.glob('[a-z][a-z][a-z].*')))
PosixPath('temp/aaa.text')
pprint.pprint(glob.glob('temp/**', recursive=True))
pprint.pprint(list(p_temp.glob('**')))
PosixPath('temp')
PosixPath('temp/dir')
PosixPath('temp/dir/sub_dir1')
PosixPath('temp/dir/sub_dir2')
pprint.pprint(list(p_temp.glob('**/*')))
PosixPath('temp/[x].txt')
PosixPath('temp/aaa.text')
PosixPath('temp/dir')
PosixPath('temp/1.txt')
PosixPath('temp/12.text')
PosixPath('temp/123.txt')
PosixPath('temp/dir/sub_dir1')
PosixPath('temp/dir/987.text')
PosixPath('temp/dir/bbb.txt')
PosixPath('temp/dir/sub_dir2')
PosixPath('temp/dir/sub_dir1/98.txt')
PosixPath('temp/dir/sub_dir1/ccc.text')
PosixPath('temp/dir/sub_dir2/ddd.text')
pprint.pprint
p for p in p_temp.glob('**/*')
re.search
str(p)
PosixPath('temp/1.txt')
PosixPath('temp/123.txt')
PosixPath('temp/dir/sub_dir1/98.txt')
pprint.pprint
p for p in p_temp.glob('**/*')
re.search
str(p)
PosixPath('temp/[x].txt')
PosixPath('temp/aaa.text')
PosixPath('temp/dir/bbb.txt')
PosixPath('temp/dir/sub_dir1/ccc.text')
PosixPath('temp/dir/sub_dir2/ddd.text')
pprint.pprint([p.resolve() for p in p_temp.iterdir()])
PosixPath('/Users/mbp/Documents/my-project/python-snippets/notebook/temp/[x].txt')
PosixPath('/Users/mbp/Documents/my-project/python-snippets/notebook/temp/aaa.text')
PosixPath('/Users/mbp/Documents/my-project/python-snippets/notebook/temp/dir')
PosixPath('/Users/mbp/Documents/my-project/python-snippets/notebook/temp/1.txt')
PosixPath('/Users/mbp/Documents/my-project/python-snippets/notebook/temp/12.text')
PosixPath('/Users/mbp/Documents/my-project/python-snippets/notebook/temp/123.txt')
pprint.pprint([str(p) for p in p_temp.iterdir()])
pprint.pprint([p for p in p_temp.iterdir() if p.is_file()])
PosixPath('temp/[x].txt')
PosixPath('temp/aaa.text')
PosixPath('temp/1.txt')
PosixPath('temp/12.text')
PosixPath('temp/123.txt')
pprint.pprint([p for p in p_temp.iterdir() if p.is_dir()])
PosixPath('temp/dir')
pprint.pprint([p.name for p in p_temp.iterdir() if p.is_file()])
aaa.text
pprint.pprint
p for p in p_temp.glob('**/*')
re.search
str(p)
PosixPath('temp/1.txt')
PosixPath('temp/123.txt')
PosixPath('temp/dir/sub_dir1/98.txt')
for p in p_temp.glob('**/*'):
re.search
str(p)
p.is_file()
p.unlink()
pprint.pprint
p for p in p_temp.glob('**/*')
re.search
str(p)
p.unlink() for p in p_temp.glob('**/*') if re.search
str(p)
p.is_file()
import shutil
shutil.rmtree('temp')
import util_make_files
util_make_files.pathlib_basic()
import pathlib
p_file = pathlib.Path('temp/file.txt')
print(p_file)
print(type(p_file))
pathlib.PosixPath
print(str(p_file))
print(type(str(p_file)))
print(p_file.name)
file.txt
print(type(p_file.name))
print(p_file.stem)
print(type(p_file.stem))
p_dir = pathlib.Path('temp/dir/')
print(p_dir)
print(type(p_dir))
pathlib.PosixPath
print(p_dir.name)
print(p_dir.stem)
print(p_file.suffix)
print(type(p_file.suffix))
print(p_dir.suffix)
print(p_file.suffix.lstrip('.'))
print(p_file.suffix[1:])
print(p_dir.suffix.lstrip('.'))
print(p_dir.suffix[1:])
p_sub = pathlib.Path('temp/dir/sub_dir/file2.txt')
print(p_sub)
print(p_sub.parent)
print(type(p_sub.parent))
pathlib.PosixPath
print(p_sub.parents[0])
print(p_sub.parents[1])
print(p_sub.parents[2])
print(p_sub.parents[3])
print(p_sub.parents[4])
p_abs = p_sub.resolve()
print(p_abs)
print(p_abs.parents[4])
print(p_abs.parents[10])
p_file = pathlib.Path('temp/file.txt')
print(p_file)
p_file_rel = pathlib.Path('temp/dir/sub_dir/../../file.txt')
print(p_file_rel)
file.txt
print(p_file.samefile(p_file_rel))
print(p_file.parents[0])
print(p_file.parents[1])
print(p_file_rel.parents[0])
print(p_file_rel.parents[1])
print(p_file_rel.parents[2])
print(p_file_rel.parents[3])
print(p_file_rel.resolve())
print(p_file_rel.resolve().relative_to(p_file_rel.cwd()))
print(p_file.with_name('file_new.txt'))
print(type(p_file.with_name('file_new.txt')))
pathlib.PosixPath
print(p_dir.with_name('dir_new'))
print(p_dir.with_name('file_new.txt'))
p_file.with_name('file_new.txt').touch()
print(p_file.with_name('file_new.txt').exists())
print(p_file.with_suffix('.text'))
print(type(p_file.with_suffix('.text')))
pathlib.PosixPath
print(p_file.with_suffix('text'))
ValueError: Invalid
import shutil
shutil.rmtree('temp')
from my_lib.imagelib import add_margin
from PIL import Image
im = Image.open('data/src/astronaut_rect.bmp')
im_new = add_margin(im, 50, 10, 0, 100, (128, 0, 64))
im_new.save('data/dst/astronaut_add_margin.jpg', quality=95)
from PIL import Image, ImageDraw, ImageFilter
im1 = Image.open('data/src/lena.jpg')
im2 = Image.open('data/src/rocket.jpg').resize(im1.size)
im2.save('data/src/rocket_resize.jpg')
mask = Image.new("L", im1.size, 128)
im = Image.composite(im1, im2, mask)
im = Image.blend(im1, im2, 0.5)
im.save('data/dst/pillow_composite_solid.jpg', quality=95)
mask = Image.new("L", im1.size, 0)
draw = ImageDraw.Draw(mask)
draw.ellipse((140, 50, 260, 170), fill=255)
im = Image.composite(im1, im2, mask)
im.save('data/dst/pillow_composite_circle.jpg', quality=95)
mask_blur = mask.filter(ImageFilter.GaussianBlur(10))
im = Image.composite(im1, im2, mask_blur)
im.save('data/dst/pillow_composite_circle_blur.jpg', quality=95)
mask = Image.open('data/src/horse.png').convert('L').resize(im1.size)
im = Image.composite(im1, im2, mask)
im.save('data/dst/pillow_composite_horse.jpg', quality=95)
mask = Image.open('data/src/gradation_h.jpg').convert('L').resize(im1.size)
im = Image.composite(im1, im2, mask)
im.save('data/dst/pillow_composite_gradation.jpg', quality=95)
from PIL import Image
im1 = Image.open('data/src/lena.jpg')
im2 = Image.open('data/src/rocket.jpg')
def get_concat_h(im1, im2):
dst = Image.new('RGB', (im1.width + im2.width, im1.height))
dst.paste(im1, (0, 0))
dst.paste(im2, (im1.width, 0))
return dst
def get_concat_v(im1, im2):
dst = Image.new('RGB', (im1.width, im1.height + im2.height))
dst.paste(im1, (0, 0))
dst.paste(im2, (0, im1.height))
return dst
get_concat_h(im1, im1).save('data/dst/pillow_concat_h.jpg')
get_concat_v(im1, im1).save('data/dst/pillow_concat_v.jpg')
def get_concat_h_cut(im1, im2):
dst = Image.new('RGB', (im1.width + im2.width, min(im1.height, im2.height)))
dst.paste(im1, (0, 0))
dst.paste(im2, (im1.width, 0))
return dst
def get_concat_v_cut(im1, im2):
dst = Image.new('RGB', (min(im1.width, im2.width), im1.height + im2.height))
dst.paste(im1, (0, 0))
dst.paste(im2, (0, im1.height))
return dst
get_concat_h_cut(im1, im2).save('data/dst/pillow_concat_h_cut.jpg')
get_concat_v_cut(im1, im2).save('data/dst/pillow_concat_v_cut.jpg')
def get_concat_h_cut_center(im1, im2):
dst = Image.new('RGB', (im1.width + im2.width, min(im1.height, im2.height)))
dst.paste(im1, (0, 0))
dst.paste(im2, (im1.width, (im1.height - im2.height) // 2))
return dst
def get_concat_v_cut_center(im1, im2):
dst = Image.new('RGB', (min(im1.width, im2.width), im1.height + im2.height))
dst.paste(im1, (0, 0))
dst.paste(im2, ((im1.width - im2.width) // 2, im1.height))
return dst
get_concat_h_cut_center(im1, im2).save('data/dst/pillow_concat_h_cut_center.jpg')
get_concat_v_cut_center(im1, im2).save('data/dst/pillow_concat_v_cut_center.jpg')
def get_concat_h_blank(im1, im2, color=(0, 0, 0)):
dst = Image.new('RGB', (im1.width + im2.width, max(im1.height, im2.height)), color)
dst.paste(im1, (0, 0))
dst.paste(im2, (im1.width, 0))
return dst
def get_concat_v_blank(im1, im2, color=(0, 0, 0)):
dst = Image.new('RGB', (max(im1.width, im2.width), im1.height + im2.height), color)
dst.paste(im1, (0, 0))
dst.paste(im2, (0, im1.height))
return dst
get_concat_h_blank(im1, im2).save('data/dst/pillow_concat_h_blank.jpg')
get_concat_v_blank(im1, im2, (0, 64, 128)).save('data/dst/pillow_concat_v_blank.jpg')
def get_concat_h_resize(im1, im2, resample=Image.BICUBIC, resize_big_image=True):
if im1.height == im2.height:
_im1 = im1
_im2 = im2
not resize_big_image
_im1 = im1.resize((int(im1.width * im2.height / im1.height), im2.height), resample=resample)
_im2 = im2
_im1 = im1
_im2 = im2.resize((int(im2.width * im1.height / im2.height), im1.height), resample=resample)
dst = Image.new('RGB', (_im1.width + _im2.width, _im1.height))
dst.paste(_im1, (0, 0))
dst.paste(_im2, (_im1.width, 0))
return dst
def get_concat_v_resize(im1, im2, resample=Image.BICUBIC, resize_big_image=True):
if im1.width == im2.width:
_im1 = im1
_im2 = im2
not resize_big_image
_im1 = im1.resize((im2.width, int(im1.height * im2.width / im1.width)), resample=resample)
_im2 = im2
_im1 = im1
_im2 = im2.resize((im1.width, int(im2.height * im1.width / im2.width)), resample=resample)
dst = Image.new('RGB', (_im1.width, _im1.height + _im2.height))
dst.paste(_im1, (0, 0))
dst.paste(_im2, (0, _im1.height))
return dst
get_concat_h_resize(im1, im2).save('data/dst/pillow_concat_h_resize.jpg')
get_concat_v_resize(im1, im2, resize_big_image=False).save('data/dst/pillow_concat_v_resize.jpg')
def get_concat_h_multi_blank(im_list):
_im = im_list.pop(0)
for im in im_list:
_im = get_concat_h_blank(_im, im)
return _im
get_concat_h_multi_blank([im1, im2, im1]).save('data/dst/pillow_concat_h_multi_blank.jpg')
def get_concat_h_multi_resize(im_list, resample=Image.BICUBIC):
min_height = min(im.height for im in im_list)
im.resize((int(im.width * min_height / im.height), min_height),resample=resample)
total_width = sum(im.width for im in im_list_resize)
dst = Image.new('RGB', (total_width, min_height))
pos_x = 0
for im in im_list_resize:
dst.paste(im, (pos_x, 0))
pos_x += im.width
return dst
def get_concat_v_multi_resize(im_list, resample=Image.BICUBIC):
min_width = min(im.width for im in im_list)
im.resize((min_width, int(im.height * min_width / im.width)),resample=resample)
total_height = sum(im.height for im in im_list_resize)
dst = Image.new('RGB', (min_width, total_height))
pos_y = 0
for im in im_list_resize:
dst.paste(im, (0, pos_y))
pos_y += im.height
return dst
get_concat_h_multi_resize([im1, im2, im1]).save('data/dst/pillow_concat_h_multi_resize.jpg')
get_concat_v_multi_resize([im1, im2, im1]).save('data/dst/pillow_concat_v_multi_resize.jpg')
def get_concat_tile_resize(im_list_2d, resample=Image.BICUBIC):
im_list_v = [get_concat_h_multi_resize(im_list_h, resample=resample) for im_list_h in im_list_2d]
return get_concat_v_multi_resize(im_list_v, resample=resample)
save('data/dst/pillow_concat_tile_resize.jpg')
def get_concat_h_repeat(im, column):
dst = Image.new('RGB', (im.width * column, im.height))
for x in range(column):
dst.paste(im, (x * im.width, 0))
return dst
def get_concat_v_repeat(im, row):
dst = Image.new('RGB', (im.width, im.height * row))
for y in range(row):
dst.paste(im, (0, y * im.height))
return dst
def get_concat_tile_repeat(im, row, column):
dst_h = get_concat_h_repeat(im, column)
return get_concat_v_repeat(dst_h, row)
im_s = im1.resize((im1.width // 2, im1.height // 2))
get_concat_tile_repeat(im_s, 3, 4).save('data/dst/pillow_concat_tile_repeat.jpg')
from my_lib.imagelib import crop_center, crop_max_square
from PIL import Image
im = Image.open('data/src/astronaut_rect.bmp')
im_crop = im.crop((60, 20, 400, 200))
im_crop.save('data/dst/astronaut_pillow_crop.jpg', quality=95)
im.crop((60, 20, 400, 200)).save('data/dst/astronaut_pillow_crop.jpg', quality=95)
im_crop_outside = im.crop((200, 150, 600, 360))
im_crop_outside.save('data/dst/astronaut_pillow_crop_outside.jpg', quality=95)
im_new = crop_center(im, 300, 150)
im_new.save('data/dst/astronaut_pillow_crop_center.jpg', quality=95)
im_new = crop_max_square(im)
im_new.save('data/dst/astronaut_pillow_crop_max_square.jpg', quality=95)
from my_lib.imagelib import expand2square
from PIL import Image
im = Image.open('data/src/astronaut_rect.bmp')
im_new = expand2square(im, (0, 0, 0))
im_new.save('data/dst/astronaut_expand_square.jpg', quality=95)
im_new = expand2square(im, (0, 0, 0)).resize((150, 150))
from PIL import Image, ImageOps
im = Image.open('data/src/lena.jpg')
im_flip = ImageOps.flip(im)
im_flip.save('data/dst/lena_flip.jpg', quality=95)
im_mirror = ImageOps.mirror(im)
im_mirror.save('data/dst/lena_mirror.jpg', quality=95)
im = Image.open('data/src/horse.png')
im_flip = ImageOps.flip(im)
im_flip.save('data/dst/horse_flip.png', quality=95)
im_mirror = ImageOps.mirror(im)
im_mirror.save('data/dst/horse_mirror.png', quality=95)
from PIL import Image, ImageDraw
images = []
width = 200
center = width // 2
color_1 = (0, 0, 0)
color_2 = (255, 255, 255)
max_radius = int(center * 1.5)
step = 8
for i in range(0, max_radius, step):
im = Image.new('RGB', (width, width), color_1)
draw = ImageDraw.Draw(im)
draw.ellipse((center - i, center - i, center + i, center + i), fill=color_2)
images.append(im)
for i in range(0, max_radius, step):
im = Image.new('RGB', (width, width), color_2)
draw = ImageDraw.Draw(im)
draw.ellipse((center - i, center - i, center + i, center + i), fill=color_1)
images.append(im)
images[0].save
save_all=True
append_images=images[1:]
optimize=False
duration=40
loop=0
from PIL import Image, ImageFilter
im = Image.open('data/src/lena_square.png')
print(im.format, im.size, im.mode)
PNG (512, 512) 
print(im.getextrema()) 
print(im.getpixel((256, 256)))
new_im = im.convert('L').rotate(90).filter(ImageFilter.GaussianBlur())
new_im.show()
new_im.save('data/dst/lena_square_pillow.jpg', quality=95)
from PIL import Image, ImageDraw, ImageFont
im = Image.new("RGB", (512, 512), (128, 128, 128))
draw = ImageDraw.Draw(im)
draw.line((0, im.height, im.width, 0), fill=(255, 0, 0), width=8)
draw.rectangle((100, 100, 200, 200), fill=(0, 255, 0))
draw.ellipse((250, 300, 450, 400), fill=(0, 0, 255))
font = ImageFont.truetype('/Library/Fonts/Arial Bold.ttf', 48)
draw.multiline_text((0, 0), 'Pillow sample', fill=(0, 0, 0), font=font)
im.save('data/dst/pillow_iamge_draw.jpg', quality=95)
from PIL import Image
img = Image.open('data/src/lena_square.png')
img_resize = img.resize((256, 256))
img_resize.save('data/dst/lena_pillow_resize_nearest.jpg')
img_resize_lanczos = img.resize((256, 256), Image.LANCZOS)
img_resize_lanczos.save('data/dst/lena_pillow_resize_lanczos.jpg')
img_resize = img.resize((img.width // 2, img.height // 2))
img_resize_lanczos.save('data/dst/lena_pillow_resize_half.jpg')
import os
import glob
from PIL import Image
dst_dir = 'data/temp/images_half'
os.makedirs(dst_dir, exist_ok=True)
files = glob.glob('./data/temp/images/*.jpg')
for f in files:
img = Image.open(f)
img_resize = img.resize((img.width // 2, img.height // 2))
root, ext = os.path.splitext(f)
basename = os.path.basename(root)
img_resize.save(os.path.join(dst_dir, basename + '_half' + ext))
files = glob.glob('./data/temp/images/*')
for f in files:
root, ext = os.path.splitext(f)
if ext in ['.jpg', '.png']:
img = Image.open(f)
img_resize = img.resize((img.width // 2, img.height // 2))
basename = os.path.basename(root)
img_resize.save(os.path.join(dst_dir, basename + '_half' + ext))
files = glob.glob('./data/temp/images/*')
for f in files:
img = Image.open(f)
img_resize = img.resize((img.width // 2, img.height // 2))
root, ext = os.path.splitext(f)
basename = os.path.basename(root)
img_resize.save(os.path.join(dst_dir, basename + '_half' + ext))
except OSError as e:
pass
from PIL import Image, ImageDraw
im = Image.new('RGB', (500, 300), (128, 128, 128))
draw = ImageDraw.Draw(im)
draw.ellipse((100, 100, 150, 200), fill=(255, 0, 0), outline=(0, 0, 0))
draw.rectangle((200, 100, 300, 200), fill=(0, 192, 192), outline=(255, 255, 255))
draw.line((350, 200, 450, 100), fill=(255, 255, 0), width=10)
im.save('data/dst/pillow_imagedraw.jpg', quality=95)
im = Image.new('RGB', (500, 250), (128, 128, 128))
draw = ImageDraw.Draw(im)
draw.line(((30, 200), (130, 100), (80, 50)), fill=(255, 255, 0))
draw.line(((80, 200), (180, 100), (130, 50)), fill=(255, 255, 0), width=10)
draw.polygon(((200, 200), (300, 100), (250, 50)), fill=(255, 255, 0), outline=(0, 0, 0))
draw.point(((350, 200), (450, 100), (400, 50)), fill=(255, 255, 0))
im.save('data/dst/pillow_imagedraw2.jpg', quality=95)
im = Image.new('RGB', (600, 250), (128, 128, 128))
draw = ImageDraw.Draw(im)
draw.arc((25, 50, 175, 200), start=30, end=270, fill=(255, 255, 0))
draw.chord((225, 50, 375, 200), start=30, end=270, fill=(255, 255, 0), outline=(0, 0, 0))
draw.pieslice((425, 50, 575, 200), start=30, end=270, fill=(255, 255, 0), outline=(0, 0, 0))
im.save('data/dst/pillow_imagedraw3.jpg', quality=95)
im = Image.open('data/src/lena.jpg')
draw = ImageDraw.Draw(im)
draw.pieslice((15, 50, 140, 175), start=30, end=330, fill=(255, 255, 0))
im.save('data/dst/pillow_imagedraw_lena.jpg', quality=95)
from PIL import Image, ImageOps
im = Image.open('data/src/lena.jpg')
im_invert = ImageOps.invert(im)
im_invert.save('data/dst/lena_invert.jpg', quality=95)
im = Image.open('data/src/horse.png').convert('RGB')
im_invert = ImageOps.invert(im)
im_invert.save('data/dst/horse_invert.png')
from PIL import Image, ImageFilter
im = Image.open('data/src/lena_square.png')
im = im.convert('L')
im = im.rotate(90)
im = im.filter(ImageFilter.GaussianBlur())
im.save('data/temp/lena_square_pillow.jpg', quality=95)
Image.open('data/src/lena_square.png').convert('L').rotate(90).filter(ImageFilter.GaussianBlur()).save('data/temp/lena_square_pillow.jpg', quality=95)
Image.open('data/src/lena_square.png')
convert('L')
rotate(90)
filter(ImageFilter.GaussianBlur())
save('data/temp/lena_square_pillow.jpg', quality=95)
from PIL import Image
import numpy as np
im = np.array(Image.open('data/src/lena.jpg'))
print(type(im))
numpy.ndarray
print(im.dtype)
print(im.shape)
im_gray = np.array(Image.open('data/src/lena.jpg').convert('L'))
print(im_gray.shape)
print(im.flags.writeable)
im_as = np.asarray(Image.open('data/src/lena.jpg'))
print(type(im_as))
numpy.ndarray
print(im_as.flags.writeable)
ValueError: assignment
im_f = im.astype(np.float64)
print(im_f.dtype)
im_f = np.array(Image.open('data/src/lena.jpg'), np.float64)
print(im_f.dtype)
print(im_gray.dtype)
pil_img = Image.fromarray(im)
print(pil_img.mode)
pil_img.save('data/temp/lena_save_pillow.jpg')
pil_img_gray = Image.fromarray(im_gray)
print(pil_img_gray.mode)
pil_img_gray.save('data/temp/lena_save_pillow_gray.jpg')
Image.fromarray(im).save('data/temp/lena_save_pillow.jpg')
Image.fromarray(im_gray).save('data/temp/lena_save_pillow_gray.jpg')
pil_img = Image.fromarray(im_f)
TypeError: Cannot
pil_img = Image.fromarray(im_f.astype(np.uint8))
pil_img.save('data/temp/lena_save_pillow.jpg')
from PIL import Image, ImageDraw, ImageFilter
im1 = Image.open('data/src/rocket.jpg')
im2 = Image.open('data/src/lena.jpg')
im1.paste(im2)
im1.save('data/dst/rocket_pillow_paste.jpg', quality=95)
im1 = Image.open('data/src/rocket.jpg')
im2 = Image.open('data/src/lena.jpg')
back_im = im1.copy()
back_im.paste(im2)
back_im.save('data/dst/rocket_pillow_paste.jpg', quality=95)
back_im = im1.copy()
back_im.paste(im2, (100, 50))
back_im.save('data/dst/rocket_pillow_paste_pos.jpg', quality=95)
back_im = im1.copy()
back_im.paste(im2, (400, 100))
back_im.save('data/dst/rocket_pillow_paste_out.jpg', quality=95)
mask_im = Image.new("L", im2.size, 0)
draw = ImageDraw.Draw(mask_im)
draw.ellipse((140, 50, 260, 170), fill=255)
mask_im.save('data/dst/mask_circle.jpg', quality=95)
back_im = im1.copy()
back_im.paste(im2, (0, 0), mask_im)
back_im.save('data/dst/rocket_pillow_paste_mask_circle.jpg', quality=95)
mask_im_blur = mask_im.filter(ImageFilter.GaussianBlur(10))
mask_im_blur.save('data/dst/mask_circle_blur.jpg', quality=95)
back_im = im1.copy()
back_im.paste(im2, (0, 0), mask_im_blur)
back_im.save('data/dst/rocket_pillow_paste_mask_circle_blur.jpg', quality=95)
mask_im = Image.open('data/src/horse.png').resize(im2.size).convert('L')
back_im = im1.copy()
back_im.paste(im2, (100, 50), mask_im)
back_im.save('data/dst/rocket_pillow_paste_mask_horse.jpg', quality=95)
from PIL import Image, ImageDraw, ImageFilter
im_rgb = Image.open('data/src/lena.jpg')
im_rgba = im_rgb.copy()
im_rgba.putalpha(128)
im_rgba.save('data/dst/pillow_putalpha_solid.png')
im_a = Image.new("L", im_rgb.size, 0)
draw = ImageDraw.Draw(im_a)
draw.ellipse((140, 50, 260, 170), fill=255)
im_rgba = im_rgb.copy()
im_rgba.putalpha(im_a)
im_rgba_crop = im_rgba.crop((140, 50, 260, 170))
im_rgba_crop.save('data/dst/pillow_putalpha_circle.png')
im_a_blur = im_a.filter(ImageFilter.GaussianBlur(4))
im_rgba = im_rgb.copy()
im_rgba.putalpha(im_a_blur)
im_rgba_crop = im_rgba.crop((135, 45, 265, 175))
im_rgba_crop.save('data/dst/pillow_putalpha_circle_blur.png')
im_a = Image.open('data/src/horse_r.png').convert('L').resize(im_rgb.size)
im_rgba = im_rgb.copy()
im_rgba.putalpha(im_a)
im_rgba.save('data/dst/pillow_putalpha_horse.png')
from PIL import Image
im = Image.open('data/src/lena.jpg')
im_rotate = im.rotate(90)
im_rotate.save('data/dst/lena_rotate_90.jpg', quality=95)
im_rotate = im.rotate(45)
im_rotate.save('data/dst/lena_rotate_45.jpg', quality=95)
im_rotate = im.rotate(45, resample=Image.BICUBIC)
im_rotate.save('data/dst/lena_rotate_45_bicubic.jpg', quality=95)
im_rotate = im.rotate(90, expand=True)
im_rotate.save('data/dst/lena_rotate_90_expand.jpg', quality=95)
im_rotate = im.rotate(45, expand=True)
im_rotate.save('data/dst/lena_rotate_45_expand.jpg', quality=95)
im_rotate = im.rotate(45, center=(0, 60))
im_rotate.save('data/dst/lena_rotate_45_change_center.jpg', quality=95)
im_rotate = im.rotate(45, center=(0, 60), expand=True)
im_rotate.save('data/dst/lena_rotate_45_change_center_expand.jpg', quality=95)
im_rotate = im.rotate(0, translate=(100, 50))
im_rotate.save('data/dst/lena_rotate_0_translate.jpg', quality=95)
im_rotate = im.rotate(45, translate=(100, 50))
im_rotate.save('data/dst/lena_rotate_45_translate.jpg', quality=95)
im_rotate = im.rotate(45, translate=(100, 50), expand=True)
im_rotate.save('data/dst/lena_rotate_45_translate_expand.jpg', quality=95)
im_rotate = im.rotate(45, fillcolor=(255, 128, 0), expand=True)
im_rotate.save('data/dst/lena_rotate_45_fillcolor_expand.jpg', quality=95)
from PIL import Image
img = Image.open('data/src/lena.jpg')
img.save('data/src/lena_q95.jpg', quality=95)
img.save('data/src/lena_q50.jpg', quality=50)
from PIL import Image
im = Image.open('data/src/lena.jpg')
print(im.size)
print(type(im.size))
w, h = im.size
print('width: ', w)
print('height:', h)
print('width: ', im.width)
print('height:', im.height)
im_gray = Image.open('data/src/lena.jpg').convert('L')
print(im.size)
print('width: ', im.width)
print('height:', im.height)
from my_lib.imagelib import crop_center, crop_max_square, expand2square, mask_circle_solid, mask_circle_transparent
import os
import glob
from PIL import Image, ImageDraw, ImageFilter
im = Image.open('data/src/astronaut_rect.bmp')
thumb_width = 200
im_thumb = crop_center(im, thumb_width, thumb_width)
im_thumb.save('data/dst/astronaut_thumbnail_center_square.jpg', quality=95)
im_thumb = crop_max_square(im).resize((thumb_width, thumb_width), Image.LANCZOS)
im_thumb.save('data/dst/astronaut_thumbnail_max_square.jpg', quality=95)
im_thumb = expand2square(im, (0, 0, 0)).resize((thumb_width, thumb_width), Image.LANCZOS)
im_thumb.save('data/dst/astronaut_thumbnail_expand.jpg', quality=95)
im_square = crop_max_square(im).resize((thumb_width, thumb_width), Image.LANCZOS)
im_thumb = mask_circle_solid(im_square, (0, 0, 0), 4)
im_thumb.save('data/dst/astronaut_thumbnail_mask_circle_solid.jpg', quality=95)
im_square = crop_max_square(im).resize((thumb_width, thumb_width), Image.LANCZOS)
im_thumb = mask_circle_transparent(im_square, 4)
im_thumb.save('data/dst/astronaut_thumbnail_mask_circle_transparent.png')
src_dir = 'data/temp/src'
dst_dir = 'data/temp/dst'
files = glob.glob(os.path.join(src_dir, '*.jpg'))
for f in files:
im = Image.open(f)
im_thumb = crop_max_square(im).resize((thumb_width, thumb_width), Image.LANCZOS)
ftitle, fext = os.path.splitext(os.path.basename(f))
im_thumb.save(os.path.join(dst_dir, ftitle + '_thumbnail' + fext), quality=95)
import platform
print(platform.python_version())
print(type(platform.python_version()))
print(platform.python_version_tuple())
print(type(platform.python_version_tuple()))
import platform
print(platform.system())
print(platform.release())
print(platform.version())
root:xnu
-4903.231
print(platform.platform())
print(platform.platform(terse=True))
print(platform.platform(aliased=True))
print(platform.mac_ver())
import platform
print(platform.system())
print(platform.release())
print(platform.version())
print(platform.platform())
print(platform.linux_distribution())
import platform
print(platform.system())
print(platform.release())
print(platform.version())
print(platform.platform())
print(platform.win32_ver())
print(0)
print(-0)
print(0 == -0)
print(0 is -0)
print(0.0)
print(-0.0)
-0.0
print(0.0 == -0.0)
print(0.0 is -0.0)
import pprint
print(l)
pprint.pprint(l)
pprint.pprint(l, width=40)
pprint.pprint(l, width=200)
pprint.pprint(l, width=1)
pprint.pprint(l, depth=1)
pprint.pprint(l, depth=2)
pprint.pprint(l, depth=2, width=40)
pprint.pprint(l, indent=4, width=4)
l_long = [list(range(10)), list(range(100, 110))]
print(l_long)
pprint.pprint(l_long, width=40)
pprint.pprint(l_long, width=40, compact=True)
s_normal = str(l)
print(s_normal)
print(type(s_normal))
s_pp = pprint.pformat(l)
print(s_pp)
print(type(s_pp))
s_pp = pprint.pformat(l, depth=2, width=40, indent=2)
print(s_pp)
l_2d = [list(range(10)), list(range(10)), list(range(10))]
print(l_2d)
pprint.pprint(l_2d)
l_2d = [[0, 1, 2], [3, 4, 5], [6, 7, 8]]
print(l_2d)
pprint.pprint(l_2d)
pprint.pprint(l_2d, width=20)
s = pprint.pformat(l_2d, width=20)
print(s)
print(type(s))
import collections
def prime_factorize(n):
a = []
while n % 2 == 0:
a.append(2)
n //= 2
f = 3
while f * f <= n:
if n % f == 0:
a.append(f)
n //= f
f += 2
if n != 1:
a.append(n)
return a
print(prime_factorize(1))
print(prime_factorize(36))
print(prime_factorize(840))
c = collections.Counter(prime_factorize(840))
print(c)
Counter({2: 3, 3: 1, 5: 1, 7: 1})
print(c.keys())
dict_keys([2, 3, 5, 7])
print(c.values())
dict_values([3, 1, 1, 1])
print(c.items())
dict_items([(2, 3), (3, 1), (5, 1), (7, 1)])
print(list(c.keys()))
def prime_factorize(n):
a = []
while n % 2 == 0:
a.append(2)
n //= 2
f = 3
while f * f <= n:
if n % f == 0:
a.append(f)
n //= f
f += 2
if n != 1:
a.append(n)
return a
print(prime_factorize(67280421310721))
prime_factorize(67280421310721)
std. dev. of
print(prime_factorize(67280421310722))
prime_factorize(67280421310722)
std. dev. of
print(prime_factorize(2147483647))
prime_factorize(2147483647)
std. dev. of
SyntaxError: Missing
print('this is a pen')
print('this is a pen')
print(100)
print([0, 1, 2])
print({'a': 0, 'b': 1, 'c': 2})
print('1.00000')
print(1.00000)
s = 'this is a pen'
print(s)
l = [0, 1, 2]
print(l)
print(l[0])
d = {'a': 0, 'b': 1, 'c': 2}
print(d)
print(d['b'])
f = 1.00000
print(f)
i = 100
print('apple', i, 0.123)
print('apple', i, 0.123, sep='----')
print('apple', i, 0.123, sep='\n')
s = 'Alice'
i = 25
print('Alice is %d years old' % i)
print('%s is %d years old' % (s, i))
print('Alice is {} years old'.format(i))
print('{} is {} years old'.format(s, i))
print('{0} is {1} years old / {0}{0}{0}'.format(s, i))
print('{name} is {age} years old'.format(name=s, age=i))
print('{} is {} years old / {{xxx}}'.format(s, i))
s = 'Alice'
i = 25
print(f'{s} is {i} years old')
number = 0.45
print('{0:.4f} is {0:.2%}'.format(number))
print(f'{number:.4f} is {number:.2%}')
i = 255
print('left   : {:<8}'.format(i))
print('center : {:^8}'.format(i))
print('right  : {:>8}'.format(i))
print('zero   : {:08}'.format(i))
print('bin    : {:b}'.format(i))
print('oct    : {:o}'.format(i))
print('hex    : {:x}'.format(i))
hex    : ff
f = 0.1234
print('digit   : {:.2}'.format(f))
print('digit   : {:.6f}'.format(f))
print('exp     : {:.4e}'.format(f))
print('percent : {:.0%}'.format(f))
print('abc')
print('xyz')
print('abc', end='---')
print('xyz')
print('abc', end='')
print('xyz')
import collections
def print_len_abc_sized(x):
if isinstance(x, collections.abc.Sized):
print(len(x))
print('x is not Sized')
print_len_abc_sized([0, 1, 2])
print_len_abc_sized('abc')
print_len_abc_sized({0, 1, 2})
print_len_abc_sized(100)
def print_len_abc_sequence(x):
if isinstance(x, collections.abc.Sequence):
print(len(x))
print('x is not Sequence')
print_len_abc_sequence([0, 1, 2])
print_len_abc_sequence('abc')
print_len_abc_sequence({0, 1, 2})
print_len_abc_sequence({'k1': 1, 'k2': 2, 'k3': 3})
def print_len_abc_mutablesequence(x):
if isinstance(x, collections.abc.MutableSequence):
print(len(x))
print('x is not MutableSequence')
print_len_abc_mutablesequence([0, 1, 2])
print_len_abc_mutablesequence('abc')
print_len_abc_mutablesequence((0, 1, 2))
MySequence(collections.abc.Sequence)
def __len__(self):
return 10
ms = MySequence()
TypeError: Can
MySequence(collections.abc.Sequence)
def __len__(self):
return 10
def __getitem__(self, i):
return i
ms = MySequence()
print(len(ms))
print(ms[3])
print(ms.index(5))
print(list(reversed(ms)))
print(isinstance(ms, collections.abc.Sequence))
print(hasattr(ms, '__len__'))
MySequence_bare()
def __len__(self):
return 10
def __getitem__(self, i):
return i
msb = MySequence_bare()
print(len(msb))
print(msb[3])
print(msb.index(5))
print(isinstance(msb, collections.abc.Sequence))
print(hasattr(msb, '__len__'))
import numpy as np
def print_len_eafp(x):
print(len(x))
except TypeError as e:
print(e)
print_len_eafp([0, 1, 2])
print_len_eafp(100)
len()
print_len_eafp((0, 1, 2))
print_len_eafp('abc')
a = np.arange(3)
print(a)
print_len_eafp(a)
import numpy as np
l = [0, 1, 2]
print(type(l))
print(hasattr(l, 'append'))
print(hasattr(l, 'xxx'))
print(len(l))
print(l.__len__())
def print_len_hasattr(x):
if hasattr(x, '__len__'):
print(len(x))
print('x has no __len__')
print_len_hasattr([0, 1, 2])
print_len_hasattr('abc')
print_len_hasattr(100)
a = np.arange(3)
print(a)
print_len_hasattr(a)
def print_len_lbyl_list(x):
if isinstance(x, list):
print(len(x))
print('x is not list')
print_len_lbyl_list([0, 1, 2])
print_len_lbyl_list(100)
print_len_lbyl_list((0, 1, 2))
print_len_lbyl_list('abc')
def print_len_lbyl_list_tuple(x):
if isinstance(x, (list, tuple)):
print(len(x))
print('x is not list or tuple')
print_len_lbyl_list_tuple([0, 1, 2])
print_len_lbyl_list_tuple(100)
x is not list or tuple
print_len_lbyl_list_tuple((0, 1, 2))
print_len_lbyl_list_tuple('abc')
x is not list or tuple
l = [0, 1, 2]
print(l)
print(*l)  
print(0, 1, 2)
print(*l, sep='')
print(*l, sep='-')
s = '-'.join([str(i) for i in l])
print(s)
import sys
import pprint
pprint.pprint(sys.path)
import PyPDF2
import glob
import os
def merge_pdf_in_dir(dir_path, dst_path):
l = glob.glob(os.path.join(dir_path, '*.pdf'))
l.sort()
merger = PyPDF2.PdfFileMerger()
for p in l:
if not PyPDF2.PdfFileReader(p).isEncrypted:
merger.append(p)
merger.write(dst_path)
merger.close()
merge_pdf_in_dir('data/src/pdf', 'data/temp/sample_dir.pdf')
import PyPDF2
merger = PyPDF2.PdfFileMerger()
merger.append('data/src/pdf/sample1.pdf')
merger.append('data/src/pdf/sample2.pdf')
merger.append('data/src/pdf/sample3.pdf')
merger.write('data/temp/sample_merge.pdf')
merger.close()
merger = PyPDF2.PdfFileMerger()
merger.append('data/src/pdf/sample1.pdf')
merger.merge(2, 'data/src/pdf/sample2.pdf')
merger.merge(4, 'data/src/pdf/sample3.pdf')
merger.write('data/temp/sample_insert.pdf')
merger.close()
merger = PyPDF2.PdfFileMerger()
merger.append('data/src/pdf/sample1.pdf')
merger.append('data/src/pdf/sample2.pdf')
d = PyPDF2.PdfFileReader('data/src/pdf/sample1.pdf').documentInfo
d = {k: d[k] for k in d.keys()}
d['/Title'] = 'merged file'
merger.addMetadata(d)
merger.write('data/temp/sample_merge_meta.pdf')
merger.close()
import PyPDF2
merger = PyPDF2.PdfFileMerger()
merger.append('data/src/pdf/sample1.pdf', pages=(0, 1))
merger.append('data/src/pdf/sample2.pdf', pages=(2, 4))
merger.merge(2, 'data/src/pdf/sample3.pdf', pages=(0, 3, 2))
merger.write('data/temp/sample_merge_page.pdf')
merger.close()
merger = PyPDF2.PdfFileMerger()
merger.append('data/src/pdf/sample1.pdf', pages=PyPDF2.pagerange.PageRange('-1'))
merger.append('data/src/pdf/sample2.pdf', pages=PyPDF2.pagerange.PageRange('2:'))
merger.merge(2, 'data/src/pdf/sample3.pdf', pages=PyPDF2.pagerange.PageRange('::-1'))
merger.write('data/temp/sample_merge_pagerange.pdf')
merger.close()
reader1 = PyPDF2.PdfFileReader('data/src/pdf/sample1.pdf')
reader2 = PyPDF2.PdfFileReader('data/src/pdf/sample2.pdf')
writer = PyPDF2.PdfFileWriter()
writer.addPage(reader1.getPage(0))
writer.addPage(reader2.getPage(2))
with open('data/temp/sample_merge_wr.pdf', 'wb') as f:
writer.write(f)
import PyPDF2
pdf = PyPDF2.PdfFileReader('data/src/pdf/sample1.pdf')
print(type(pdf.documentInfo))
PyPDF2.pdf.DocumentInformation
print(isinstance(pdf.documentInfo, dict))
print(pdf.documentInfo)
IndirectObject(33, 0)
IndirectObject(34, 0)
IndirectObject(35, 0)
IndirectObject(36, 0)
IndirectObject(36, 0)
print(pdf.documentInfo['/Title'])
for k in pdf.documentInfo.keys():
print(k, ':', pdf.documentInfo[k])
Title : sample1
Producer : macOS
Creator : Keynote
CreationDate : D
ModDate : D
import PyPDF2
src_pdf = PyPDF2.PdfFileReader('data/src/pdf/sample1.pdf')
dst_pdf = PyPDF2.PdfFileWriter()
dst_pdf.cloneReaderDocumentRoot(src_pdf)
with open('data/temp/sample1_no_meta.pdf', 'wb') as f:
dst_pdf.write(f)
print(PyPDF2.PdfFileReader('data/temp/sample1_no_meta.pdf').documentInfo)
dst_pdf.addMetadata({'/Producer': ''})
with open('data/temp/sample1_no_meta.pdf', 'wb') as f:
dst_pdf.write(f)
print(PyPDF2.PdfFileReader('data/temp/sample1_no_meta.pdf').documentInfo)
def remove_all_metadata(src_path, dst_path, producer=''):
src_pdf = PyPDF2.PdfFileReader(src_path)
dst_pdf = PyPDF2.PdfFileWriter()
dst_pdf.cloneReaderDocumentRoot(src_pdf)
dst_pdf.addMetadata({'/Producer': producer})
with open(dst_path, 'wb') as f:
dst_pdf.write(f)
remove_all_metadata('data/src/pdf/sample1.pdf', 'data/temp/sample1_no_meta.pdf')
print(PyPDF2.PdfFileReader('data/temp/sample1_no_meta.pdf').documentInfo)
src_pdf = PyPDF2.PdfFileReader('data/src/pdf/sample1.pdf')
dst_pdf = PyPDF2.PdfFileWriter()
d = {key: src_pdf.documentInfo[key] for key in src_pdf.documentInfo.keys()}
print(d)
d.pop('/Creator')
d.pop('/Producer')
print(d)
dst_pdf.addMetadata(d)
with open('data/temp/sample1_remove_meta.pdf', 'wb') as f:
dst_pdf.write(f)
print(PyPDF2.PdfFileReader('data/temp/sample1_remove_meta.pdf').documentInfo)
def remove_metadata(src_path, dst_path, *args, producer=''):
src_pdf = PyPDF2.PdfFileReader(src_path)
dst_pdf = PyPDF2.PdfFileWriter()
dst_pdf.cloneReaderDocumentRoot(src_pdf)
key: src_pdf
documentInfo[key] for key in src_pdf.documentInfo.keys()
d.setdefault('/Producer', producer)
dst_pdf.addMetadata(d)
with open(dst_path, 'wb') as f:
dst_pdf.write(f)
print(PyPDF2.PdfFileReader('data/temp/sample1_no_meta.pdf').documentInfo)
print(PyPDF2.PdfFileReader('data/temp/sample1_no_meta.pdf').documentInfo)
producer='XXX'
print(PyPDF2.PdfFileReader('data/temp/sample1_no_meta.pdf').documentInfo)
def select_metadata(src_path, dst_path, *args, producer=''):
src_pdf = PyPDF2.PdfFileReader(src_path)
dst_pdf = PyPDF2.PdfFileWriter()
dst_pdf.cloneReaderDocumentRoot(src_pdf)
key: src_pdf
documentInfo[key] for key in src_pdf.documentInfo.keys()
d.setdefault('/Producer', producer)
dst_pdf.addMetadata(d)
with open(dst_path, 'wb') as f:
dst_pdf.write(f)
print(PyPDF2.PdfFileReader('data/temp/sample1_no_meta.pdf').documentInfo)
print(PyPDF2.PdfFileReader('data/temp/sample1_no_meta.pdf').documentInfo)
import PyPDF2
src_pdf = PyPDF2.PdfFileReader('data/src/pdf/sample1.pdf')
dst_pdf = PyPDF2.PdfFileWriter()
dst_pdf.cloneReaderDocumentRoot(src_pdf)
d = {key: src_pdf.documentInfo[key] for key in src_pdf.documentInfo.keys()}
d['/Title'] = 'new title'
d['/Author'] = 'new author'
d['/XXX'] = 'special data'
dst_pdf.addMetadata(d)
with open('data/temp/sample1_new_meta.pdf', 'wb') as f:
dst_pdf.write(f)
print(PyPDF2.PdfFileReader('data/temp/sample1_new_meta.pdf').documentInfo)
def update_metadata(src_path, dst_path, metadata):
src_pdf = PyPDF2.PdfFileReader(src_path)
dst_pdf = PyPDF2.PdfFileWriter()
dst_pdf.cloneReaderDocumentRoot(src_pdf)
d = {key: src_pdf.documentInfo[key] for key in src_pdf.documentInfo.keys()}
d.update(metadata)
dst_pdf.addMetadata(d)
with open(dst_path, 'wb') as f:
dst_pdf.write(f)
print(PyPDF2.PdfFileReader('data/temp/sample1_new_meta.pdf').documentInfo)
def set_metadata(src_path, dst_path, metadata):
src_pdf = PyPDF2.PdfFileReader(src_path)
dst_pdf = PyPDF2.PdfFileWriter()
dst_pdf.cloneReaderDocumentRoot(src_pdf)
dst_pdf.addMetadata(metadata)
with open(dst_path, 'wb') as f:
dst_pdf.write(f)
print(PyPDF2.PdfFileReader('data/temp/sample1_new_meta.pdf').documentInfo)
import PyPDF2
pdf = PyPDF2.PdfFileReader('data/temp/Simple PDF 2.0 file.pdf')
print(pdf.documentInfo)
print(type(pdf.xmpMetadata))
PyPDF2.xmp.XmpInformation
print(pdf.xmpMetadata.dc_title)
print(pdf.xmpMetadata.pdf_keywords)
print(pdf.xmpMetadata.xmp_metadataDate)
import PyPDF2
pdf = PyPDF2.PdfFileReader('data/src/pdf/sample1.pdf')
print(pdf.isEncrypted)
src_pdf = PyPDF2.PdfFileReader('data/src/pdf/sample1.pdf')
dst_pdf = PyPDF2.PdfFileWriter()
dst_pdf.cloneReaderDocumentRoot(src_pdf)
print(src_pdf.documentInfo)
IndirectObject(33, 0)
IndirectObject(34, 0)
IndirectObject(35, 0)
IndirectObject(36, 0)
IndirectObject(36, 0)
dst_pdf.addMetadata(src_pdf.documentInfo)
TypeError: createStringObject
str or unicode
d = {key: src_pdf.documentInfo[key] for key in src_pdf.documentInfo.keys()}
print(d)
dst_pdf.addMetadata(d)
dst_pdf.encrypt('password')
with open('data/temp/sample1_pass.pdf', 'wb') as f:
dst_pdf.write(f)
def set_password(src_path, dst_path, password):
src_pdf = PyPDF2.PdfFileReader(src_path)
dst_pdf = PyPDF2.PdfFileWriter()
dst_pdf.cloneReaderDocumentRoot(src_pdf)
d = {key: src_pdf.documentInfo[key] for key in src_pdf.documentInfo.keys()}
dst_pdf.addMetadata(d)
dst_pdf.encrypt(password)
with open(dst_path, 'wb') as f:
dst_pdf.write(f)
set_password('data/src/pdf/sample1.pdf', 'data/temp/sample1_pass.pdf', 'password')
pdf_rc4 = PyPDF2.PdfFileReader('data/src/pdf/sample1_pass_rc4.pdf')
print(pdf_rc4.isEncrypted)
print(pdf_rc4.documentInfo)
PdfReadError: file
not been 
print(pdf_rc4.decrypt('wrong-password'))
print(pdf_rc4.decrypt('password'))
print(pdf_rc4.documentInfo)
pdf_aes = PyPDF2.PdfFileReader('data/src/pdf/sample1_pass_aes.pdf')
print(pdf_aes.decrypt('password'))
NotImplementedError: only
1 and 2
def change_password(src_path, dst_path, src_password, dst_password=None):
src_pdf = PyPDF2.PdfFileReader(src_path)
src_pdf.decrypt(src_password)
dst_pdf = PyPDF2.PdfFileWriter()
dst_pdf.cloneReaderDocumentRoot(src_pdf)
d = {key: src_pdf.documentInfo[key] for key in src_pdf.documentInfo.keys()}
dst_pdf.addMetadata(d)
if dst_password:
dst_pdf.encrypt(dst_password)
with open(dst_path, 'wb') as f:
dst_pdf.write(f)
change_password('data/src/pdf/sample1_pass_rc4.pdf', 'data/temp/sample1_no_pass.pdf', 'password')
pass
import PyPDF2
merger = PyPDF2.PdfFileMerger()
merger.append('data/src/pdf/sample1.pdf', pages=PyPDF2.pagerange.PageRange(':2'))
merger.write('data/temp/sample_split1.pdf')
merger.close()
merger = PyPDF2.PdfFileMerger()
merger.append('data/src/pdf/sample1.pdf', pages=PyPDF2.pagerange.PageRange('2:'))
merger.write('data/temp/sample_split2.pdf')
merger.close()
import PyPDF2
def split_pdf_pages(src_path, dst_basepath):
src_pdf = PyPDF2.PdfFileReader(src_path)
for i in range(src_pdf.numPages):
dst_pdf = PyPDF2.PdfFileWriter()
dst_pdf.addPage(src_pdf.getPage(i))
with open('{}_{}.pdf'.format(dst_basepath, i), 'wb') as f:
dst_pdf.write(f)
split_pdf_pages('data/src/pdf/sample1.pdf', 'data/temp/sample1')
import requests
import pprint
with open('data/temp/qiita_access_token.txt') as f:
qiita_access_token = f.read().strip()
headers = {'Authorization': 'Bearer {}'.format(qiita_access_token)}
url_items = 'https://qiita.com/api/v2/items'
r_post = requests.post(url_items, headers=headers, json=item_data)
print(r_post.status_code)
pprint.pprint(r_post.json())
item_id = r_post.json()['id']
print(item_id)
r_get = requests.get(url_items + '/' + item_id, headers=headers)
print(r_get.status_code)
pprint.pprint(r_get.json())
item_data_updated = item_data.copy()
item_data_updated['title'] = 'テスト記事更新'
print(item_data_updated)
r_patch = requests.patch(url_items + '/' + item_id, headers=headers, json=item_data_updated)
print(r_patch.status_code)
pprint.pprint(r_patch.json())
r_delete = requests.delete(url_items + '/' + item_id, headers=headers)
print(r_delete.status_code)
url_tag = 'https://qiita.com/api/v2/tags/{}/following'
tag = 'python'
r_put = requests.put(url_tag.format(tag), headers=headers)
print(r_put.status_code)
r_delete_tag = requests.delete(url_tag.format(tag), headers=headers)
print(r_delete_tag.status_code)
import qrcode
from PIL import Image
img_bg = Image.open('data/src/lena.jpg')
qr = qrcode.QRCode(box_size=2)
qr.add_data('I am Lena')
qr.make()
img_qr = qr.make_image()
pos = (img_bg.size[0] - img_qr.size[0], img_bg.size[1] - img_qr.size[1])
img_bg.paste(img_qr, pos)
img_bg.save('data/dst/qr_lena.png')
face = Image.open('data/src/lena.jpg').crop((175, 90, 235, 150))
qr_big = qrcode.QRCode
error_correction=qrcode.constants.ERROR_CORRECT_H
qr_big.add_data('I am Lena')
qr_big.make()
img_qr_big = qr_big.make_image().convert('RGB')
pos = ((img_qr_big.size[0] - face.size[0]) // 2, (img_qr_big.size[1] - face.size[1]) // 2)
img_qr_big.paste(face, pos)
img_qr_big.save('data/dst/qr_lena2.png')
import qrcode
img = qrcode.make('test text')
print(type(img))
print(img.size)
qrcode.image.pil.PilImage
img.save('data/dst/qrcode_test.png')
qr = qrcode.QRCode()
qr.add_data('test text')
qr.make()
img = qr.make_image()
img.save('data/dst/qrcode_test2.png')
qr = qrcode.QRCode
version=12
error_correction=qrcode.constants.ERROR_CORRECT_H
box_size=2
border=8
qr.add_data('test text')
qr.make()
img = qr.make_image(fill_color="red", back_color="#23dda0")
img.save('data/dst/qrcode_test2_2.png')
import random
l = [0, 1, 2, 3, 4]
print(random.choice(l))
print(random.choice(('xxx', 'yyy', 'zzz')))
print(random.choice('abcde'))
print(random.choice([]))
IndexError: Cannot
random.seed(0)
print(random.choice(l))
import random
l = [0, 1, 2, 3, 4]
print(random.choices(l, k=3))
print(random.choices(l, k=10))
print(random.choices(l))
print(random.choices(l, k=3, weights=[1, 1, 1, 10, 1]))
print(random.choices(l, k=3, weights=[1, 1, 0, 0, 0]))
print(random.choices(l, k=3, cum_weights=[1, 2, 3, 13, 14]))
print(random.choices(l, k=3, weights=[1, 1, 1, 10, 1, 1, 1]))
ValueError: The
not match 
print(random.choices(l, k=3, weights=[1, 1, 1, 10, 1], cum_weights=[1, 2, 3, 13, 14]))
TypeError: Cannot
weights and cumulative
import random
print([random.random() for i in range(5)])
print([random.randint(0, 10) for i in range(5)])
print(random.sample(range(10), k=5))
print(random.sample(range(100, 200, 10), k=5))
import random
print(random.random())
random.seed(0)
print(random.random())
import random
print(list(range(10)))
print(random.randrange(10))
print(list(range(10, 20, 2)))
print(random.randrange(10, 20, 2))
print(random.randint(50, 100))
print(random.randrange(50, 101))
random.seed(0)
print(random.randint(50, 100))
import random
l = [0, 1, 2, 3, 4]
print(random.sample(l, 3))
print(type(random.sample(l, 3)))
print(random.sample(l, 1))
print(random.sample(l, 0))
print(random.sample(l, 10))
ValueError: Sample
population or is
print(random.sample(('xxx', 'yyy', 'zzz'), 2))
print(random.sample('abcde', 2))
print(tuple(random.sample(('xxx', 'yyy', 'zzz'), 2)))
print(''.join(random.sample('abcde', 2)))
l_dup = [0, 0, 0, 1, 1, 1, 2, 2, 2, 3, 3, 3]
print(random.sample(l_dup, 3))
print(set(l_dup))
print(random.sample(set(l_dup), 3))
import random
l = list(range(5))
print(l)
random.shuffle(l)
print(l)
l = list(range(5))
print(l)
lr = random.sample(l, len(l))
print(lr)
print(l)
s = 'abcde'
random.shuffle(s)
not support 
t = tuple(range(5))
print(t)
random.shuffle(t)
not support 
sr = ''.join(random.sample(s, len(s)))
print(sr)
tr = tuple(random.sample(t, len(l)))
print(tr)
random.seed(0)
l = list(range(5))
random.shuffle(l)
print(l)
import random
print(random.uniform(100, 200))
print(random.uniform(100, -100))
-27.82338731501028
print(random.uniform(100, 100))
print(random.uniform(1.234, 5.637))
print(xrange(3))
NameError: name
not defined
print(range(3))
range(0, 3)
print(type(range(3)))
for i in range(3):
print(i)
print(list(range(3)))
print(list(range(10)))
print(list(range(-3)))
print(list(range(3, 10)))
print(list(range(10, 3)))
print(list(range(-3, 3)))
-3
-2
-1
print(list(range(3, -3)))
print(range(0, 3) == range(3))
print(list(range(3, 10, 2)))
print(list(range(10, 3, 2)))
print(list(range(10, 3, -2)))
print(list(range(3, 10, -2)))
print(range(3, 10, 1) == range(3, 10))
print(range(0, 10, 1) == range(0, 10) == range(10))
print(list(range(3, 10, 2)))
print(list(range(9, 2, -2)))
print(list(reversed(range(3, 10, 2))))
for i in reversed(range(3, 10, 2)):
print(i)
print(list(range(0.3, 1.0, 0.2)))
print([i / 10 for i in range(3, 10, 2)])
print([i * 0.1 for i in range(3, 10, 2)])
print([round(i * 0.1, 1) for i in range(3, 10, 2)])
s = 'a\tb\nA\tB'
print(s)
rs = r
print(rs)
print(type(rs))
print(rs == 'a\\tb\\nA\\tB')
print(len(s))
print(list(s))
print(len(rs))
print(list(rs))
path = 'C:\\Windows\\system32\\cmd.exe'
rpath = r
cmd.exe
print(path == rpath)
path2 = 'C:\\Windows\\system32\\'
rpath2 = r
SyntaxError: EOL
rpath2 = r
+ '\\'
print(path2 == rpath2)
s_r = repr(s)
print(s_r)
print(list(s_r))
s_r2 = repr(s)[1:-1]
print(s_r2)
print(s_r2 == rs)
repr('\t')[1:-1]
SyntaxError: EOL
SyntaxError: EOL
import re
p = re.compile('[0123456789]+')
print(p.fullmatch('123'))
re.Match
span=(0, 3)
match='123'
p = re.compile('[^0123456789]+')
print(p.fullmatch('123'))
print(p.fullmatch('abc'))
re.Match
span=(0, 3)
match='abc'
p = re.compile('[0-9]+')
print(p.fullmatch('123'))
re.Match
span=(0, 3)
match='123'
p = re.compile('[9-0]+')
error: bad
p = re.compile('[\u3041-\u309F]+')
print(p.fullmatch('あいうえおぁぃぅぇぉわをん'))
re.Match
span=(0, 13)
match='あいうえおぁぃぅぇぉわをん'
p = re.compile('[ぁ-ゟ]+')
print(p.fullmatch('あいうえおぁぃぅぇぉわをん'))
re.Match
span=(0, 13)
match='あいうえおぁぃぅぇぉわをん'
print('[\u3041-\u309F]+')
-{}
+'.format('
p = re.compile('[a-zA-Z]+')
print(p.fullmatch('abcABC'))
re.Match
span=(0, 6)
match='abcABC'
p = re.compile
print(p.fullmatch('abc-[ABC]'))
re.Match
span=(0, 9)
match='abc-[ABC]'
p = re.compile('[a-zA-Z]+')
print(p.findall('abcABCxyzXYZ'))
p = re.compile('[a-z]+|[A-Z]+')
print(p.findall('abcABCxyzXYZ'))
import re
import regex
p = re.compile('[a-z]+')
print(p.fullmatch('abc'))
re.Match
span=(0, 3)
match='abc'
p = re.compile('[A-Z]+')
print(p.fullmatch('ABC'))
re.Match
span=(0, 3)
match='ABC'
p = re.compile('[ａ-ｚ]+')
print(p.fullmatch('ａｂｃ'))
re.Match
span=(0, 3)
match='ａｂｃ'
p = re.compile('[Ａ-Ｚ]+')
print(p.fullmatch('ＡＢＣ'))
re.Match
span=(0, 3)
match='ＡＢＣ'
p = re.compile('[a-zA-Zａ-ｚＡ-Ｚ]+')
print(p.fullmatch('abcABCａｂｃＡＢＣ'))
re.Match
span=(0, 12)
match='abcABCａｂｃＡＢＣ'
p = regex.compile
Script=Latin
print(p.fullmatch('AÁÀÂÄÆ'))
regex.Match
span=(0, 6)
match='AÁÀÂÄÆ'
p = re.compile('[0-9]+')
print(p.fullmatch('123'))
re.Match
span=(0, 3)
match='123'
p = re.compile('[０-９]+')
print(p.fullmatch('１２３'))
re.Match
span=(0, 3)
match='１２３'
p = regex.compile
Numeric_Type=Numeric
print(p.fullmatch('一二三ⅠⅡⅢ百万億⑩⑽'))
regex.Match
span=(0, 11)
match='一二三ⅠⅡⅢ百万億⑩⑽'
print(p.fullmatch('123'))
p = re.compile('[\u2160-\u217F]+')
print(p.fullmatch('ⅠⅡⅢ'))
re.Match
span=(0, 3)
match='ⅠⅡⅢ'
p = re.compile('[〇一二三四五六七八九十百千万億兆]+')
print(p.fullmatch('三十五億'))
re.Match
span=(0, 4)
match='三十五億'
p = re.compile('[\u0000-\u007F]+')
print(p.fullmatch('(abc)!_(123)?'))
re.Match
span=(0, 13)
match='(abc)!_(123)?'
p = re.compile('[\u0020-\u002F\u003A-\u0040\u005B-\u0060\u007B-\u007E]+')
print(p.fullmatch('!_? ()[]'))
re.Match
span=(0, 8)
match='!_? ()[]'
p = re.compile
print(p.fullmatch(',.!?[]()'))
re.Match
span=(0, 8)
match=',.!?[]()'
p = re.compile('[\uFF01-\uFF0F\uFF1A-\uFF20\uFF3B-\uFF40\uFF5B-\uFF65]+')
print(p.fullmatch('！？（）［］｢｣'))
re.Match
span=(0, 8)
match='！？（）［］｢｣'
p = re.compile('[\u3000-\u303F]+')
print(p.fullmatch('、。「」【】'))
re.Match
span=(0, 6)
match='、。「」【】'
p = re.compile('[\uFF01-\uFF0F\uFF1A-\uFF20\uFF3B-\uFF40\uFF5B-\uFF65\u3000-\u303F]+')
print(p.fullmatch('！？（）［］｢｣、。「」【】'))
re.Match
span=(0, 14)
match='！？（）［］｢｣、。「」【】'
p = re.compile('[\u3041-\u309F]+')
print(p.fullmatch('あいうえおぁぃぅぇぉ'))
re.Match
span=(0, 10)
match='あいうえおぁぃぅぇぉ'
p = re.compile('[ぁ-ゟ]+')
print(p.fullmatch('あいうえおぁぃぅぇぉ'))
re.Match
span=(0, 10)
match='あいうえおぁぃぅぇぉ'
p = re.compile('[\u30A1-\u30FF]+')
print(p.fullmatch('アイウエオァィゥェォ'))
re.Match
span=(0, 10)
match='アイウエオァィゥェォ'
p = re.compile
print(p.fullmatch('アイウエオァィゥェォ'))
re.Match
span=(0, 10)
match='アイウエオァィゥェォ'
p = re.compile('[\uFF66-\uFF9F]+')
print(p.fullmatch('ｱｲｳｴｵｧｨｩｪｫ'))
re.Match
span=(0, 10)
match='ｱｲｳｴｵｧｨｩｪｫ'
p = re.compile('[ｦ-ﾟ]+')
print(p.fullmatch('ｱｲｳｴｵｧｨｩｪｫ'))
re.Match
span=(0, 10)
match='ｱｲｳｴｵｧｨｩｪｫ'
p = regex.compile
Script=Han
print(p.fullmatch('漢字'))
regex.Match
span=(0, 2)
match='漢字'
p = regex.compile
Script_Extensions=Han
print(p.fullmatch('漢字〆㈠㈱㊊㏩'))
regex.Match
span=(0, 7)
match='漢字〆㈠㈱㊊㏩'
p = re.compile('[\u2E80-\u2FDF\u3005-\u3007\u3400-\u4DBF\u4E00-\u9FFF\uF900-\uFAFF\U00020000-\U0002EBEF]+')
print(p.fullmatch('漢字'))
re.Match
span=(0, 2)
match='漢字'
p = regex.compile
Emoji=Yes
print(p.fullmatch('💯123'))
regex.Match
span=(0, 4)
match='💯123'
p = regex.compile
Emoji_Presentation=Yes
print(p.fullmatch('💯'))
regex.Match
span=(0, 1)
match='💯'
p = regex.compile
Basic_Emoji=Yes
error: unknown
p = re.compile('[\U0001F300-\U0001F5FF]+')
print(p.fullmatch('💯'))
re.Match
span=(0, 1)
match='💯'
p = re.compile('[\U0001F600-\U0001F64F]+')
print(p.fullmatch('😀'))
re.Match
span=(0, 1)
match='😀'
import re
p = re.compile('[a-z]+')
print(p.fullmatch('abc'))
re.Match
span=(0, 3)
match='abc'
print(p.fullmatch('abc123'))
s = 'abc'
if p.fullmatch(s):
print('match')
print('no match')
s = 'abc123'
if p.fullmatch(s):
print('match')
print('no match')
p = re.compile('[a-z]+$')
print(p.match('abc'))
re.Match
span=(0, 3)
match='abc'
print(p.match('abc123'))
p = re.compile('[a-z]+')
print(p.search('123abcABC'))
re.Match
span=(3, 6)
match='abc'
print(p.search('123ABC'))
p = re.compile('[a-z]+')
result = p.findall('123abcABCxyz')
print(result)
print(type(result))
print(type(result[0]))
print(p.findall('123ABC'))
s_result = ''.join(p.findall('123abcABCxyz'))
print(s_result)
print(len(s_result))
print(len('🇯🇵'))
import re
s = 'aaa@xxx.com, bbb@yyy.com, ccc@zzz.net'
m = re.match
print(m)
re.Match
span=(0, 11)
match='aaa@xxx.com'
result = re.sub
print(result)
zzz.net
p = re.compile
print(p)
re.compile('([a-z]+)@([a-z]+)\\.com')
print(type(p))
re.Pattern
m = p.match(s)
print(m)
re.Match
span=(0, 11)
match='aaa@xxx.com'
result = p.sub('new-address', s)
print(result)
zzz.net
import re
s = 'aaa@xxx.com, bbb@yyy.com, ccc@zzz.net'
result = re.findall
print(result)
xxx.com
yyy.com
zzz.net
print(len(result))
result = re.findall
print(result)
result = re.findall
print(result)
xxx.com
yyy.com
zzz.net
result = re.findall('[0-9]+', s)
print(result)
s = 'aaa@xxx.com, bbb@yyy.com, ccc@zzz.net'
result = re.finditer
print(result)
print(type(result))
for m in result:
print(m)
re.Match
span=(0, 11)
match='aaa@xxx.com'
re.Match
span=(13, 24)
match='bbb@yyy.com'
re.Match
span=(26, 37)
match='ccc@zzz.net'
l = list
re.finditer
print(l)
re.Match
span=(0, 11)
match='aaa@xxx.com'
re.Match
span=(13, 24)
match='bbb@yyy.com'
re.Match
span=(26, 37)
match='ccc@zzz.net'
print(l[0])
re.Match
span=(0, 11)
match='aaa@xxx.com'
print(type(l[0]))
re.Match
print(l[0].span())
m.span() for m in re.finditer
result = re.finditer
for m in result:
print(m)
re.Match
span=(0, 11)
match='aaa@xxx.com'
re.Match
span=(13, 24)
match='bbb@yyy.com'
re.Match
span=(26, 37)
match='ccc@zzz.net'
print(list(result))
import re
m = re.match
print(m)
re.Match
span=(0, 11)
match='あいう漢字ＡＢＣ１２３'
m = re.match('[a-zA-Z0-9_]+', 'あいう漢字ＡＢＣ１２３')
print(m)
m = re.match
flags=re.ASCII
print(m)
m = re.match
print(m)
p = re.compile
flags=re.ASCII
print(p)
re.compile('\\w+', re.ASCII)
print(p.match('あいう漢字ＡＢＣ１２３'))
p = re.compile
print(p)
re.compile('(?a)\\w+', re.ASCII)
print(p.match('あいう漢字ＡＢＣ１２３'))
print(re.ASCII is re.A)
m = re.match
print(m)
m = re.match
flags=re.ASCII
print(m)
re.Match
span=(0, 11)
match='あいう漢字ＡＢＣ１２３'
m = re.match
print(m)
re.Match
span=(0, 3)
match='123'
m = re.match
print(m)
re.Match
span=(0, 3)
match='１２３'
m = re.match
flags=re.ASCII
print(m)
re.Match
span=(0, 3)
match='123'
m = re.match
flags=re.ASCII
print(m)
m = re.match
print(m)
re.Match
span=(0, 1)
match='\u3000'
m = re.match
flags=re.ASCII
print(m)
m = re.match('[a-zA-Z]+', 'abcABC')
print(m)
re.Match
span=(0, 6)
match='abcABC'
m = re.match('[a-z]+', 'abcABC', flags=re.IGNORECASE)
print(m)
re.Match
span=(0, 6)
match='abcABC'
m = re.match('[A-Z]+', 'abcABC', flags=re.IGNORECASE)
print(m)
re.Match
span=(0, 6)
match='abcABC'
s = ''
print(s)
result = re.findall('[a-z]+', s)
print(result)
result = re.findall('^[a-z]+', s)
print(result)
result = re.findall('^[a-z]+', s, flags=re.MULTILINE)
print(result)
result = re.findall('[a-z]+$', s)
print(result)
result = re.findall('[a-z]+$', s, flags=re.MULTILINE)
print(result)
s = ''
print(s)
result = re.findall
s, flags=re.M
print(result)
result = re.findall
s, flags=re.M | re.A
print(result)
result = re.findall
print(result)
import re
s = 'aaa@xxx.com, bbb@yyy.com'
m = re.match
+com
print(m)
re.Match
span=(0, 24)
match='aaa@xxx.com, bbb@yyy.com'
print(m.group())
xxx.com
yyy.com
m = re.match
print(m)
re.Match
span=(0, 11)
match='aaa@xxx.com'
print(m.group())
xxx.com
import re
import pandas as pd
url = 'https://raw.githubusercontent.com/cjkvi/cjkvi-tables/15569eaae99daef9f99f0383e9d8efbec64a7c5a/joyo2010.txt'
df = pd.read_csv(url, header=None, skiprows=1, delimiter='\t')
print(df.shape)
print(df.head())
print(df.tail())
kanji = ''.join(df.iloc[:, 0])
print(kanji)
print(len(kanji))
print(kanji[:30])
print(kanji[-30:])
p = re.compile('[{}]+'.format(kanji))
print(p.fullmatch('常用漢字'))
re.Match
span=(0, 4)
match='常用漢字'
kanji_ex = ''.join(df.iloc[:, 0]) + ''.join(df.iloc[:, 1].dropna())
print(len(kanji_ex))
print(kanji_ex[-30:])
p_ex = re.compile('[{}]+'.format(kanji_ex))
print(p_ex.fullmatch('常用漢字'))
re.Match
span=(0, 4)
match='常用漢字'
print(p.fullmatch('野村證券'))
print(p_ex.fullmatch('野村證券'))
re.Match
span=(0, 4)
match='野村證券'
print('叱' in kanji_ex)
print('𠮟' in kanji_ex)
kanji_ex2 = kanji_ex + '叱'
p_ex2 = re.compile('[{}]+'.format(kanji_ex2))
print(p_ex2.fullmatch('常用漢字野村證券叱𠮟'))
re.Match
span=(0, 10)
match='常用漢字野村證券叱𠮟'
import re
s = 'aaa@xxx.com'
m = re.match
print(m)
re.Match
span=(0, 11)
match='aaa@xxx.com'
print(type(m))
re.Match
print(m.start())
print(m.end())
print(m.span())
print(m.group())
xxx.com
print(type(m.group()))
m = re.match
print(m)
re.Match
span=(0, 11)
match='aaa@xxx.com'
print(m.groups())
print(m.group())
xxx.com
print(m.group(0))
xxx.com
print(m.group(1))
print(m.group(2))
print(m.group(3))
print(m.group(4))
IndexError: no
print(m.group(0, 1, 3))
xxx.com
print(m.span())
print(m.span(3))
print(m.span(4))
IndexError: no
print(m.span(0, 1))
TypeError: span
m = re.match
print(m)
re.Match
span=(0, 11)
match='aaa@xxx.com'
print(m.groups())
xxx.com
m = re.match
print(m)
re.Match
span=(0, 11)
match='aaa@xxx.com'
print(m.group('local'))
print(m.group('SLD'))
print(m.group('TLD'))
print(m.group(0))
xxx.com
print(m.group(3))
print(m.group(0, 2, 'TLD'))
xxx.com
print(m.groups())
print(m.groupdict())
print(type(m.groupdict()))
re.match
re.Match
span=(0, 11)
match='aaa@xxx.com'
re.match
print(re.match('[0-9]+', s))
print(bool(re.match('[0-9]+', s)))
re.match
print('match')
print('no match')
if re.match('[0-9]+', s):
print('match')
print('no match')
m = re.match('[0-9]*', s)
print(m)
re.Match
span=(0, 0)
match=''
print(m.group() == '')
print(bool(m))
if re.match('[0-9]*', s):
print('match')
print('no match')
import re
s = 'aaa@xxx.com, bbb@yyy.com, ccc@zzz.net'
m = re.match
print(m)
re.Match
span=(0, 11)
match='aaa@xxx.com'
m = re.match
print(m)
s = 'aaa@xxx.com, bbb@yyy.com, ccc@zzz.net'
m = re.search
print(m)
re.Match
span=(26, 37)
match='ccc@zzz.net'
m = re.search
print(m)
re.Match
span=(0, 11)
match='aaa@xxx.com'
s = 'aaa@xxx.com'
m = re.fullmatch
print(m)
re.Match
span=(0, 11)
match='aaa@xxx.com'
s = '!!!aaa@xxx.com!!!'
m = re.fullmatch
print(m)
s = '!!!aaa@xxx.com!!!'
m = re.match
print(m)
import re
s = '111aaa222bbb333'
result = re.split('[a-z]+', s)
print(result)
result = re.split('[0-9]+', s)
print(result)
result = re.split('[a-z]+', s, 1)
print(result)
import re
s = 'aaa@xxx.com, bbb@yyy.com, ccc@zzz.net'
result = re.sub
print(result)
zzz.net
print(type(result))
result = re.sub
print(result)
xxx.net
yyy.net
zzz.net
result = re.sub
print(result)
xxx.net
yyy.net
zzz.net
result = re.sub
s, count=1
print(result)
yyy.com
zzz.net
result = re.subn
print(result)
zzz.net
print(result[0])
zzz.net
print(result[1])
result = re.subn
print(result)
xxx.net
yyy.net
zzz.net
result = re.subn
s, count=1
print(result)
yyy.com
zzz.net
import re
s = 'one two one two'
m = re.match('one', s)
print(m)
_sre.SRE_Match
span=(0, 3)
match='one'
print(m.group())
print(m.start())
print(m.end())
print(m.span())
m = re.match('one two', s)
print(m)
print(m.group())
_sre.SRE_Match
span=(0, 7)
match='one two'
m = re.match('(one) (two)', s)
print(m)
print(m.group())
print(m.groups())
_sre.SRE_Match
span=(0, 7)
match='one two'
m = re.match('two', s)
print(m)
m = re.search('one', s)
print(m)
_sre.SRE_Match
span=(0, 3)
match='one'
m = re.search('two', s)
print(m)
_sre.SRE_Match
span=(4, 7)
match='two'
m = re.findall('one', s)
print(m)
m = re.findall('one two', s)
print(m)
m = re.findall('(one) (two)', s)
print(m)
m = re.finditer('one', s)
print(m)
for match in m:
print(match)
_sre.SRE_Match
span=(0, 3)
match='one'
_sre.SRE_Match
span=(8, 11)
match='one'
m = re.sub('one', 'ONE', s)
print(m)
m = re.sub('one two', 'xxx', s)
print(m)
m = re.sub('(one) (two)', '\\1X\\2', s)
print(m)
m = re.sub
print(m)
m = re.subn('one', 'ONE', s)
print(m)
m = re.split(' ', s)
print(m)
p = re.compile('one')
m = p.match(s)
print(m)
_sre.SRE_Match
span=(0, 3)
match='one'
m = p.findall(s)
print(m)
m = p.sub('ONE', s)
print(m)
import regex
p = regex.compile
Block=Hiragana
print(p.fullmatch('あいうえおぁぃぅぇぉわをんゟ'))
regex.Match
span=(0, 14)
match='あいうえおぁぃぅぇぉわをんゟ'
p = regex.compile
Script=Hiragana
print(p.fullmatch('あいうえおぁぃぅぇぉわをんゟ🈀'))
regex.Match
span=(0, 15)
match='あいうえおぁぃぅぇぉわをんゟ🈀'
p = regex.compile
print(p.fullmatch('あいうえおぁぃぅぇぉわをんゟ🈀'))
regex.Match
span=(0, 15)
match='あいうえおぁぃぅぇぉわをんゟ🈀'
p = regex.compile
subhead=Hiragana_letters
error: unknown
p = regex.compile
Script=Hiragana
Script=Katakana
print(p.fullmatch('あーいアイウabc🈀'))
regex.Match
span=(0, 10)
match='あーいアイウabc🈀'
import requests
import os
url_image = 'https://www.python.org/static/community_logos/python-logo.png'
r_image = requests.get(url_image)
print(r_image.headers['Content-Type'])
filename_image = os.path.basename(url_image)
print(filename_image)
with open('data/temp/' + filename_image, 'wb') as f:
f.write(r_image.content)
url_zip = 'http://www.post.japanpost.jp/zipcode/dl/oogaki/zip/13tokyo.zip'
r_zip = requests.get(url_zip)
print(r_zip.headers['Content-Type'])
filename_zip = os.path.basename(url_zip)
print(filename_zip)
tokyo.zip
with open('data/temp/' + filename_zip, 'wb') as f:
f.write(r_zip.content)
import requests
import pprint
import json
url = 'http://weather.livedoor.com/forecast/webservice/json/v1'
params = {'city': 130010}
r = requests.get(url, params=params)
print(r.headers['Content-Type'])
charset=utf-8
json_data = r.json()
print(type(json_data))
pprint.pprint(json_data, depth=2, compact=True)
weather.livedoor.com
print(json_data['description']['text'])
pprint.pprint(json_data['forecasts'][0])
with open('data/temp/download.json', 'w') as f:
json.dump(json_data, f, ensure_ascii=False, indent=4)
import requests
url = 'https://en.wikipedia.org'
r = requests.get(url)
print(r.url)
print(r.status_code)
print(r.history)
Response [301]
print(len(r.history))
print(type(r.history[0]))
requests.models.Response
print(r.history[0].url)
en.wikipedia.org
print(r.history[0].status_code)
print([response.url for response in r.history])
en.wikipedia.org
r_not_redirect = requests.get(url, allow_redirects=False)
print(r_not_redirect.url)
en.wikipedia.org
print(r_not_redirect.status_code)
import requests
url = 'https://www.yahoo.co.jp/'
ua = 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_13_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/67.0.3396.99 Safari/537.36'
headers = {'User-Agent': ua}
r_ua = requests.get(url, headers=headers)
import requests
url = 'https://example.com/'
response = requests.get(url)
print(response)
Response [200]
print(type(response))
requests.models.Response
print(response.url)
example.com
print(response.status_code)
print(response.headers)
ECS (oxr/8313)
print(type(response.headers))
requests.structures.CaseInsensitiveDict
print(response.headers['Content-Type'])
print(response.headers['content-type'])
print(response.headers['cOntEnt-typE'])
print(response.headers['xxxxx'])
print(response.headers.get('xxxxx'))
print(response.encoding)
print(response.text)
charset="utf-8"
content="text/html; charset=utf-8"
name="viewport"
content="width=device-width, initial-scale=1"
type="text/css"
a:link
a:visited
width: auto
coordination or asking
href="http://www.iana.org/domains/example">More
print(type(response.text))
print(response.content)
charset="utf-8"
content="text/html; charset=utf-8"
name="viewport"
content="width=device-width, initial-scale=1"
type="text/css"
a:link
a:visited
width: auto
coordination or asking
href="http://www.iana.org/domains/example">More
print(type(response.content))
print(response.content.decode(response.encoding) == response.text)
import requests
url = 'https://www.google.co.jp/search'
params = {'q': '日本代表', 'tbm': 'nws'}
r = requests.get(url, params=params)
print(r.url)
def test():
return 'abc', 100
result = test()
print(result)
print(type(result))
print(result[0])
print(type(result[0]))
print(result[1])
print(type(result[1]))
print(result[2])
IndexError: tuple
a, b = test()
print(a)
print(b)
def test2():
return 'abc', 100, [0, 1, 2]
a, b, c = test2()
print(a)
print(b)
print(c)
def test_list():
return ['abc', 100]
result = test_list()
print(result)
print(type(result))
org_list = [1, 2, 3, 4, 5]
org_list.reverse()
print(org_list)
print(org_list.reverse())
org_list = [1, 2, 3, 4, 5]
reverse_iterator = reversed(org_list)
print(org_list)
print(type(reverse_iterator))
for i in reversed(org_list):
print(i)
new_list = list(reversed(org_list))
print(org_list)
print(new_list)
org_list = [1, 2, 3, 4, 5]
new_list = org_list[::-1]
print(org_list)
print(new_list)
org_str = 'abcde'
new_str_list = list(reversed(org_str))
print(new_str_list)
new_str = ''.join(list(reversed(org_str)))
print(new_str)
new_str = org_str[::-1]
print(new_str)
org_tuple = (1, 2, 3, 4, 5)
new_tuple = tuple(reversed(org_tuple))
print(new_tuple)
new_tuple = org_tuple[::-1]
print(new_tuple)
s = 'abc'
s_rjust = s.rjust(8)
print(s_rjust)
print(type(s_rjust))
print(s.rjust(2))
print(s.rjust(8, '+'))
+++++abc
print(s.rjust(8, '漢'))
print(s.rjust(8, 'xyz'))
TypeError: The
s_n = '-123'
print(s_n.rjust(8, '0'))
print(s_n.zfill(8))
-0000123
s = 'abc'
print(s.center(8))
print(s.center(8, '+'))
++abc
print(s.center(9, '+'))
+++abc
print(s.center(10, '+'))
+++abc
s = 'abc'
print(s.ljust(8))
print(s.ljust(8, '+'))
n = 123
print(type(n))
print(str(n).rjust(8, '@'))
print(str(n).center(8, '@'))
print(str(n).ljust(8, '@'))
s = 'abc'
print('right : {:@>8}'.format(s))
print('center: {:@^8}'.format(s))
print('left  : {:@<8}'.format(s))
left  : abc
n = 255
print('right : {:08}'.format(n))
print('right : {:08x}'.format(n))
print(f'right : {n:08}')
print(f'right : {n:08x}')
f = 123.456
print(round(f))
print(type(round(f)))
print(round(f, 1))
print(round(f, 2))
print(round(f, -1))
print(round(f, -2))
print(round(f, 0))
print(type(round(f, 0)))
i = 99518
print(round(i))
print(round(i, 2))
print(round(i, -1))
print(round(i, -2))
print(round(i, -3))
print('0.4 =>', round(0.4))
print('0.5 =>', round(0.5))
print('0.6 =>', round(0.6))
print('4 =>', round(4, -1))
print('5 =>', round(5, -1))
print('6 =>', round(6, -1))
print('0.5 =>', round(0.5))
print('1.5 =>', round(1.5))
print('2.5 =>', round(2.5))
print('3.5 =>', round(3.5))
print('4.5 =>', round(4.5))
print('0.05 =>', round(0.05, 1))
print('0.15 =>', round(0.15, 1))
print('0.25 =>', round(0.25, 1))
print('0.35 =>', round(0.35, 1))
print('0.45 =>', round(0.45, 1))
import math
def round_towards_infinity(x):
return int(math.copysign(math.ceil(abs(x)), x))
print(round_towards_infinity(10.123))
print(round_towards_infinity(-10.123))
-11
print(math.floor(10.123))
print(math.floor(-10.123))
-11
print(math.ceil(10.123))
print(math.ceil(-10.123))
-10
print(int(10.123))
print(int(-10.123))
-10
import os
def save_file_at_dir(dir_path, filename, file_content, mode='w'):
os.makedirs(dir_path, exist_ok=True)
with open(os.path.join(dir_path, filename), mode) as f:
f.write(file_content)
import numpy as np
from scipy.sparse.csgraph import connected_components
from scipy.sparse import csr_matrix
n, labels = connected_components(l)
a = np.array(l)
print(type(a))
numpy.ndarray
n, labels = connected_components(a)
print(n)
print(labels)
csr = csr_matrix(l)
print(csr)
print(type(csr))
scipy.sparse.csr.csr_matrix
n, labels = connected_components(csr)
print(n)
print(labels)
n = 5
d = [1, 1, 1, 1]
i = [0, 0, 1, 3]
j = [1, 2, 2, 4]
csr = csr_matrix((d, (i, j)), (n, n))
print(csr)
print(connected_components(csr, return_labels=False))
n, labels = connected_components(csr_matrix(ld))
print(n)
print(labels)
n, labels = connected_components(csr_matrix(ld), connection='strong')
print(n)
print(labels)
n, labels = connected_components(csr_matrix(ld2), connection='strong')
print(n)
print(labels)
n, labels = connected_components(csr_matrix(ld), directed=False, connection='strong')
print(n)
print(labels)
-2
n, labels = connected_components(csr_matrix(lw))
print(n)
print(labels)
from scipy.spatial import Delaunay, delaunay_plot_2d, Voronoi, voronoi_plot_2d
import matplotlib.pyplot as plt
import numpy as np
w = h = 360
n = 6
np.random.seed(0)
pts = np.random.randint(0, w, (n, 2))
print(pts)
print(type(pts))
numpy.ndarray
print(pts.shape)
tri = Delaunay(pts)
print(type(tri))
scipy.spatial.qhull.Delaunay
fig = delaunay_plot_2d(tri)
fig.savefig('data/dst/scipy_matplotlib_delaunay.png')
plt.close()
print(tri.points)
print(tri.points == pts)
print(tri.simplices)
print(pts[tri.simplices])
vor = Voronoi(pts)
print(type(vor))
scipy.spatial.qhull.Voronoi
fig = voronoi_plot_2d(vor)
fig.savefig('data/dst/scipy_matplotlib_voronoi.png')
plt.close()
print(vor.vertices)
print(vor.regions)
-1
-1
-1
-1
print([r for r in vor.regions if -1 not in r and r])
for region in [r for r in vor.regions if -1 not in r and r]:
print(vor.vertices[region])
fig, ax = plt.subplots()
voronoi_plot_2d(vor, ax)
for region, c in zip([r for r in vor.regions if -1 not in r and r], ['yellow', 'pink']):
ax.fill
vor.vertices[region][:, 0]
vor.vertices[region][:, 1]
color=c
fig.savefig('data/dst/scipy_matplotlib_voronoi_fill.png')
plt.close()
MatplotlibDeprecationWarning: The
was_held = ax.ishold()
fig, ax = plt.subplots(figsize=(4, 4))
delaunay_plot_2d
voronoi_plot_2d(vor, ax, show_vertices=False)
ax.set_xlim(0, w)
ax.set_ylim(0, h)
ax.grid(linestyle='--')
fig.savefig('data/dst/scipy_matplotlib_delaunay_voronoi.png')
plt.close()
MatplotlibDeprecationWarning: The
was_held = ax.ishold()
import numpy as np
from scipy.sparse.csgraph import minimum_spanning_tree
from scipy.sparse import csr_matrix, coo_matrix, lil_matrix
mst = minimum_spanning_tree(l)
print(mst)
print(type(mst))
scipy.sparse.csr.csr_matrix
print(mst.toarray())
print(mst.toarray().astype(int))
print(mst.astype(int).toarray())
a = np.array(l)
print(a)
print(type(a))
numpy.ndarray
print(minimum_spanning_tree(a))
csr = csr_matrix(l)
print(csr)
print(type(csr))
scipy.sparse.csr.csr_matrix
print(minimum_spanning_tree(csr))
print(minimum_spanning_tree(coo_matrix(l)))
print(minimum_spanning_tree(lil_matrix(l)))
n = 4
d = [8, 3, 2, 5, 6]
i = [0, 0, 1, 1, 2]
j = [1, 3, 2, 3, 3]
csr_ = csr_matrix((d, (i, j)), shape=(n, n))
print(csr_)
print(csr_.toarray())
print(minimum_spanning_tree(csr))
print(mst.toarray().astype(int).tolist())
print(type(mst.toarray().astype(int).tolist()))
print(mst.sum())
print(int(mst.sum()))
r, c = mst.nonzero()
print(r, c)
print(list(zip(*mst.nonzero())))
print(mst.data)
print(list(zip(*mst.nonzero(), mst.data.astype(int))))
print(minimum_spanning_tree(l).toarray().astype(int))
print(minimum_spanning_tree(l).toarray().astype(int))
print(minimum_spanning_tree(l).toarray().astype(int))
from scipy.sparse.csgraph import minimum_spanning_tree
from scipy.sparse import csr_matrix, coo_matrix
minimum_spanning_tree(l)
std. dev. of
minimum_spanning_tree(l, True)
std. dev. of
csr = csr_matrix(l)
minimum_spanning_tree(csr)
std. dev. of
minimum_spanning_tree(csr, True)
std. dev. of
coo = coo_matrix(l)
minimum_spanning_tree(coo)
std. dev. of
minimum_spanning_tree(coo, True)
std. dev. of
minimum_spanning_tree(csr_matrix(l))
std. dev. of
minimum_spanning_tree(csr_matrix(l), True)
std. dev. of
mst = minimum_spanning_tree(csr, True)
print(csr)
print(mst)
import numpy as np
from scipy.sparse.csgraph import shortest_path, floyd_warshall, dijkstra, bellman_ford, johnson
from scipy.sparse import csr_matrix
print(shortest_path(l))
a = np.array(l)
print(type(a))
numpy.ndarray
print(shortest_path(a))
print(type(shortest_path(a)))
numpy.ndarray
csr = csr_matrix(l)
print(csr)
print(type(csr))
scipy.sparse.csr.csr_matrix
print(shortest_path(csr))
print(shortest_path(csr, indices=0))
print(shortest_path(csr, indices=[0, 3]))
print(shortest_path(csr, directed=False))
print(shortest_path(csr_matrix(l_ud)))
print(shortest_path(csr_matrix(l2), directed=False))
print(shortest_path(csr, unweighted=True))
print(shortest_path(csr_matrix(l_uw)))
d, p = shortest_path(csr, return_predecessors=True)
print(d)
print(p)
-9999
def get_path(start, goal, pred):
return get_path_row(start, goal, pred[start])
def get_path_row(start, goal, pred_row):
path = []
i = goal
while i != start and i >= 0:
path.append(i)
i = pred_row[i]
if i < 0:
return []
path.append(i)
return path[::-1]
print(get_path(0, 3, p))
print(get_path(2, 1, p))
print(get_path(3, 3, p))
print(get_path(3, 1, p))
d2, p2 = shortest_path(csr, indices=2, return_predecessors=True)
print(d2)
print(p2)
print(get_path_row(2, 1, p2))
print(get_path_row(2, 0, p2))
print(floyd_warshall(csr, indices=0))
TypeError: floyd_warshall
print(shortest_path(csr, indices=0, method='FW'))
ValueError: Cannot
method == 'FW'
-3
print(dijkstra(csr_matrix(l_n)))
UserWarning: Graph
weights: dijkstra
cycles. Consider
johnson or bellman_ford
print(dijkstra(csr))
print(dijkstra(csr, limit=2))
-10
print(bellman_ford(csr_matrix(l_nc)))
NegativeCycleError: Negative
print(floyd_warshall(csr_matrix(l_nc)))
NegativeCycleError: Negative
print(johnson(csr_matrix(l_nc)))
NegativeCycleError: Negative
print(dijkstra(csr_matrix(l_nc)))
-7.   0
UserWarning: Graph
weights: dijkstra
cycles. Consider
johnson or bellman_ford
import numpy as np
from scipy.sparse.csgraph import shortest_path, dijkstra, floyd_warshall, bellman_ford, johnson
from scipy.sparse import csr_matrix
n = 100
c = n * 2
np.random.seed(1)
d = np.random.randint(0, n, c)
i = np.random.randint(0, n, c)
j = np.random.randint(0, n, c)
csr = csr_matrix((d, (i, j)), shape=(n, n))
a = csr.toarray()
print(a.shape)
print((a > 0).sum())
shortest_path(a)
std. dev. of
shortest_path(csr)
std. dev. of
shortest_path(a, method='D')
std. dev. of
shortest_path(csr, method='D')
std. dev. of
dijkstra(a)
std. dev. of
dijkstra(csr)
std. dev. of
shortest_path(a, method='FW')
std. dev. of
shortest_path(csr, method='FW')
std. dev. of
floyd_warshall(a)
std. dev. of
floyd_warshall(csr)
std. dev. of
dijkstra(a, indices=0)
std. dev. of
dijkstra(csr, indices=0)
std. dev. of
a_n = a.copy()
a_n[0, 1] = -10
csr_n = csr_matrix(a_n)
johnson(csr_n)
std. dev. of
bellman_ford(csr_n)
std. dev. of
floyd_warshall(csr_n)
std. dev. of
johnson(csr_n, indices=0)
std. dev. of
import numpy as np
from scipy.sparse.csgraph import shortest_path, dijkstra, floyd_warshall, bellman_ford, johnson
from scipy.sparse import csr_matrix
n = 100
c = n * 50
np.random.seed(1)
d = np.random.randint(0, n, c)
i = np.random.randint(0, n, c)
j = np.random.randint(0, n, c)
csr = csr_matrix((d, (i, j)), shape=(n, n))
a = csr.toarray()
print(a.shape)
print((a > 0).sum())
shortest_path(a)
std. dev. of
shortest_path(csr)
std. dev. of
shortest_path(a, method='FW')
std. dev. of
shortest_path(csr, method='FW')
std. dev. of
floyd_warshall(a)
std. dev. of
floyd_warshall(csr)
std. dev. of
dijkstra(a, indices=0)
std. dev. of
dijkstra(csr, indices=0)
std. dev. of
a_n = a.copy()
a_n[0, 1] = -10
csr_n = csr_matrix(a_n)
johnson(csr_n)
std. dev. of
bellman_ford(csr_n)
std. dev. of
floyd_warshall(csr_n)
std. dev. of
johnson(csr_n, indices=0)
std. dev. of
import numpy as np
from scipy.sparse import csr_matrix, csc_matrix, coo_matrix, lil_matrix
csr = csr_matrix(l)
print((csr + csr).toarray())
print(type((csr + csr)))
scipy.sparse.csr.csr_matrix
print((csr - csr).toarray())
print((csr * csr).toarray())
print((csr.multiply(csr)).toarray())
print((csr.dot(csr)).toarray())
print(csr / csr)
print(type(csr / csr))
numpy.matrix
print(np.array(csr / csr))
print(type(np.array(csr / csr)))
numpy.ndarray
csr_full = csr_matrix([[1, 2, 3], [4, 5, 6], [7, 8, 9]])
print(csr_full / csr_full)
print(type(csr_full / csr_full))
numpy.matrix
print(csr + 10)
NotImplementedError: adding
print(csr - 10)
NotImplementedError: subtracting
print(csr_full + 10)
NotImplementedError: adding
print(csr_full - 10)
NotImplementedError: subtracting
print((csr * 10).toarray())
print((csr / 10).toarray())
print((csr ** 2).toarray())
print((csr ** 3).toarray())
print((csr * csr * csr).toarray())
print((csr ** -1).toarray())
alueError: exponent
print((csr ** 0.5).toarray())
ValueError: exponent
csc = csc_matrix(l)
coo = coo_matrix(l)
lil = lil_matrix(l)
print(type(csc + csc))
scipy.sparse.csc.csc_matrix
print(type(csr + csc))
scipy.sparse.csr.csr_matrix
print(type(csc + csr))
scipy.sparse.csc.csc_matrix
print(type(coo + coo))
scipy.sparse.csr.csr_matrix
print(type(lil + lil))
scipy.sparse.csr.csr_matrix
a = np.array(l)
print(a)
print(type(a))
numpy.ndarray
print(a + csr)
print(type(a + csr))
numpy.matrix
print(type(csr - a))
numpy.matrix
print(type(a / csr))
TypeError: unsupported
type(s) 
numpy.ndarray
print(csr / a)
RuntimeWarning: invalid
return np.true_divide(self.todense(), other)
print(type(csr / a))
numpy.matrix
RuntimeWarning: invalid
return np.true_divide(self.todense(), other)
print(csr * a)
print(type(csr * a))
numpy.ndarray
print(type(a * csr))
numpy.ndarray
print(type(csr.dot(a)))
numpy.ndarray
print(a.dot(csr))
numpy.int64
numpy.int64
numpy.int64
numpy.int64
numpy.int64
numpy.int64
numpy.int64
numpy.int64
numpy.int64
print(csr.multiply(a))
print(type(csr.multiply(a)))
scipy.sparse.coo.coo_matrix
print(csr.multiply(a).toarray())
print(np.multiply(a, csr))
numpy.int64
numpy.int64
numpy.int64
numpy.int64
numpy.int64
numpy.int64
numpy.int64
numpy.int64
numpy.int64
from scipy.sparse import csr_matrix, csc_matrix, coo_matrix, lil_matrix
data = [10, 20, 30, 40, 50]
row = [0, 0, 1, 1, 2]
col = [1, 2, 0, 2, 2]
csr = csr_matrix(l)
print(csr)
print(type(csr))
scipy.sparse.csr.csr_matrix
print(csr.data)
print(csr.indices)
print(csr.indptr)
csc = csc_matrix(l)
print(csc)
print(type(csc))
scipy.sparse.csc.csc_matrix
print(csc.data)
print(csc.indices)
print(csc.indptr)
coo = coo_matrix(l)
print(coo)
print(type(coo))
scipy.sparse.coo.coo_matrix
print(coo.data)
print(coo.row)
print(coo.col)
lil = lil_matrix(l)
print(lil)
print(type(lil))
scipy.sparse.lil.lil_matrix
print(lil.data)
list([10, 20]) 
list([30, 40]) 
list([50])
print(lil.rows)
list([1, 2]) 
list([0, 2]) 
list([2])
import numpy as np
def calc_indptr(n, data, row, column, tocsr=True):
if tocsr:
Ai, Aj = row, column
Ai, Aj = column, row
indptr = np.zeros(n + 1, dtype=int)
indices = np.zeros_like(Aj)
data_ = np.zeros_like(data)
for a in Ai:
indptr[a + 1] += 1
indptr = indptr.cumsum()
for i, j, d in zip(Ai, Aj, data):
dest = indptr[i]
indices[dest] = j
data_[dest] = d
indptr[i] += 1
indptr[-1] = 0
return data_, indices, np.roll(indptr, 1)
print(calc_indptr(3, [10, 20, 30, 40], [0, 0, 1, 1], [1, 2, 0, 2]))
array([10, 20, 30, 40])
array([1, 2, 0, 2])
array([0, 2, 4, 4])
print(calc_indptr(3, [10, 20, 30, 40], [0, 0, 1, 1], [1, 2, 0, 2], False))
array([30, 10, 20, 40])
array([1, 0, 0, 1])
array([0, 1, 2, 4])
from scipy.sparse import csr_matrix, csc_matrix, coo_matrix
data = [10, 20, 30, 40]
row = [0, 0, 1, 1]
col = [1, 2, 0, 2]
print(csr_matrix((data, (row, col))).toarray())
print(csr_matrix((data, (row, col)), shape=(3, 3)).toarray())
print(csr_matrix((data, (row, col)), shape=(4, 4)).toarray())
data = [10, 20, 30, 40]
row = [0, 0, 1, 1]
col = [1, 2, 2, 2]
print(csr_matrix((data, (row, col))))
print(csc_matrix((data, (row, col))))
print(coo_matrix((data, (row, col))))
print(coo_matrix((data, (row, col))).tocsr())
print(coo_matrix((data, (row, col))).toarray())
print(csr_matrix(([10, 20, 30, 40], [1, 2, 0, 2], [0, 2, 4, 4]), shape=(3, 3)).toarray())
print(csc_matrix(([30, 10, 20, 40], [1, 0, 0, 1], [0, 1, 2, 4]), shape=(3, 3)).toarray())
from scipy.sparse import csr_matrix, coo_matrix, lil_matrix
lil = lil_matrix((3, 3), dtype=int)
print(lil.toarray())
lil[1, 0] = 10
lil[2, 2] = 30
print(lil)
print(lil.toarray())
lil[2, 2] = 0
print(lil)
print(lil.toarray())
csr = csr_matrix((3, 3), dtype=int)
print(csr.toarray())
csr[1, 0] = 10
SparseEfficiencyWarning: Changing
coo = coo_matrix((3, 3), dtype=int)
coo[1, 0] = 10
not support 
import numpy as np
from scipy.sparse import csr_matrix
csr = csr_matrix(l)
print(csr)
print(type(csr))
scipy.sparse.csr.csr_matrix
print(csr.shape)
a = np.array(l)
print(a)
print(type(a))
numpy.ndarray
csr = csr_matrix(a)
print(csr)
print(type(csr))
scipy.sparse.csr.csr_matrix
print(csr_matrix([0, 1, 2]))
print(csr_matrix([0, 1, 2]).shape)
print(csr_matrix([[[0, 1, 2]]]))
TypeError: expected
array or matrix
print(csr.toarray())
print(type(csr.toarray()))
numpy.ndarray
print(csr.toarray().tolist())
print(type(csr.toarray().tolist()))
print(csr.todense())
print(type(csr.todense()))
numpy.matrix
import numpy as np
from scipy.sparse import csr_matrix, csc_matrix, coo_matrix, lil_matrix
n = 1000
np.random.seed(0)
d = np.ones(n, dtype=int)
i = np.random.randint(0, n, n)
j = np.random.randint(0, n, n)
print(d[:10])
print(i[:10])
print(j[:10])
csr = csr_matrix((d, (i, j)), (n, n))
std. dev. of
csc = csc_matrix((d, (i, j)), (n, n))
std. dev. of
coo = coo_matrix((d, (i, j)), (n, n))
std. dev. of
csr = coo_matrix((d, (i, j)), (n, n)).tocsr()
std. dev. of
csc = coo_matrix((d, (i, j)), (n, n)).tocsc()
std. dev. of
lil = lil_matrix((n, n))
for d_, i_, j_ in zip(d, i, j):
lil[i_, j_] = d_
std. dev. of
lil = lil_matrix((n, n))
std. dev. of
for d_, i_, j_ in zip(d, i, j):
pass
std. dev. of
dij = list(zip(d, i, j))
print(dij[:5])
lil = lil_matrix((n, n))
for d, i, j in dij:
lil[i, j] = d
std. dev. of
d_, i_, j_ = zip(*dij)
coo = coo_matrix((d_, (i_, j_)), (n, n))
std. dev. of
dij_ = []
for t in dij:
dij_.append(t)
d_, i_, j_ = zip(*dij_)
coo = coo_matrix((d_, (i_, j_)), (n, n))
std. dev. of
d_ = []
i_ = []
j_ = []
for d, i, j in dij:
d_.append(d)
i_.append(i)
j_.append(j)
coo = coo_matrix((d_, (i_, j_)), (n, n))
std. dev. of
a = coo_matrix((d, (i, j)), (n, n)).toarray()
print(a.shape)
csr = csr_matrix(a)
std. dev. of
csc = csc_matrix(a)
std. dev. of
coo = coo_matrix(a)
std. dev. of
lil = lil_matrix(a)
std. dev. of
lil = coo_matrix(a).tolil()
std. dev. of
from scipy.sparse import csr_matrix, lil_matrix
csr = csr_matrix(l)
print(csr)
print(type(csr))
scipy.sparse.csr.csr_matrix
lil = csr.tolil()
print(lil)
print(type(lil))
scipy.sparse.lil.lil_matrix
lil = lil_matrix(csr)
print(lil)
print(type(lil))
scipy.sparse.lil.lil_matrix
lil[0, 0] = 100
print(lil.toarray())
print(csr.toarray())
lil2 = lil_matrix(lil)
print(lil2.toarray())
lil[0, 0] = 0
print(lil2.toarray())
lil2_copy = lil_matrix(lil, copy=True)
print(lil2_copy.toarray())
lil[0, 0] = 100
print(lil2_copy.toarray())
from scipy.sparse import csr_matrix, csc_matrix, lil_matrix, dok_matrix, eye
n = 1000
csr = eye(n, format='csr')
csc = eye(n, format='csc')
lil = eye(n, format='lil')
dok = eye(n, format='dok')
csr.getrow(0)
std. dev. of
csc.getrow(0)
std. dev. of
lil.getrow(0)
std. dev. of
dok.getrow(0)
std. dev. of
csr[0]
std. dev. of
csc[0]
std. dev. of
lil[0]
std. dev. of
dok[0]
std. dev. of
csr.getcol(0)
std. dev. of
csc.getcol(0)
std. dev. of
lil.getcol(0)
std. dev. of
dok.getcol(0)
std. dev. of
csr[:, 0]
std. dev. of
csc[:, 0]
std. dev. of
lil[:, 0]
std. dev. of
dok[:, 0]
std. dev. of
csr[0, 0]
std. dev. of
csc[0, 0]
std. dev. of
lil[0, 0]
std. dev. of
dok[0, 0]
std. dev. of
csr[:10]
std. dev. of
csc[:10]
std. dev. of
lil[:10]
std. dev. of
dok[:10]
std. dev. of
csr[:, :10]
std. dev. of
csc[:, :10]
std. dev. of
lil[:, :10]
std. dev. of
dok[:, :10]
std. dev. of
csr[:10, :10]
std. dev. of
csc[:10, :10]
std. dev. of
lil[:10, :10]
std. dev. of
dok[:10, :10]
std. dev. of
import numpy as np
from scipy.sparse import csr_matrix, csc_matrix, lil_matrix, dok_matrix
n = 1000
np.random.seed(0)
d = np.random.randint(1, n, n*10)
i = np.random.randint(0, n, n*10)
j = np.random.randint(0, n, n*10)
csr = csr_matrix((d, (i, j)), (n, n))
csc = csr.tocsc()
lil = csr.tolil()
dok = csr.todok()
csr.getrow(0)
std. dev. of
csc.getrow(0)
std. dev. of
lil.getrow(0)
std. dev. of
dok.getrow(0)
std. dev. of
csr[0]
std. dev. of
csc[0]
std. dev. of
lil[0]
std. dev. of
dok[0]
std. dev. of
csr.getcol(0)
std. dev. of
csc.getcol(0)
std. dev. of
lil.getcol(0)
std. dev. of
dok.getcol(0)
std. dev. of
csr[:, 0]
std. dev. of
csc[:, 0]
std. dev. of
lil[:, 0]
std. dev. of
dok[:, 0]
std. dev. of
csr[0, 0]
std. dev. of
csc[0, 0]
std. dev. of
lil[0, 0]
std. dev. of
dok[0, 0]
std. dev. of
csr[:10]
std. dev. of
csc[:10]
std. dev. of
lil[:10]
std. dev. of
dok[:10]
std. dev. of
csr[:, :10]
std. dev. of
csc[:, :10]
std. dev. of
lil[:, :10]
std. dev. of
dok[:, :10]
std. dev. of
csr[:10, :10]
std. dev. of
csc[:10, :10]
std. dev. of
lil[:10, :10]
std. dev. of
dok[:10, :10]
std. dev. of
from scipy.sparse import csr_matrix, csc_matrix, coo_matrix, lil_matrix
csr = csr_matrix(l)
csc = csc_matrix(l)
coo = coo_matrix(l)
lil = lil_matrix(l)
print(csr.getrow(0))
print(type(csr.getrow(0)))
scipy.sparse.csr.csr_matrix
print(csr.getrow(0).shape)
print(csr.getrow(0).toarray())
print(type(csc.getrow(0)))
scipy.sparse.csr.csr_matrix
print(type(coo.getrow(0)))
scipy.sparse.csr.csr_matrix
print(type(lil.getrow(0)))
scipy.sparse.lil.lil_matrix
print(csr.getcol(0))
print(type(csr.getcol(0)))
scipy.sparse.csr.csr_matrix
print(csr.getcol(0).shape)
print(csr.getcol(0).toarray())
print(type(csc.getcol(0)))
scipy.sparse.csc.csc_matrix
print(type(coo.getcol(0)))
scipy.sparse.csr.csr_matrix
print(type(lil.getcol(0)))
scipy.sparse.csr.csr_matrix
lil_row = lil.getrow(0)
lil_row[0, 0] = 100
print(lil.toarray())
print(lil_row.toarray())
from scipy.sparse import csr_matrix, lil_matrix, hstack, vstack
csr = csr_matrix(l)
lil = lil_matrix(l)
print(hstack([csr, lil]).toarray())
print(type(hstack([csr, lil])))
scipy.sparse.coo.coo_matrix
print(type(hstack([csr, lil], format='csr')))
scipy.sparse.csr.csr_matrix
print(vstack([csr, lil]).toarray())
print(type(vstack([csr, lil])))
scipy.sparse.coo.coo_matrix
print(type(vstack([csr, lil], format='csr')))
scipy.sparse.csr.csr_matrix
print(vstack([csr, lil[:2]]).toarray())
print(hstack([csr, lil[:2]]))
ValueError: blocks
dimensions. Got
blocks[0,1].shape[0] == 2
from scipy.sparse import csr_matrix, csc_matrix, coo_matrix, lil_matrix
csr = csr_matrix(l)
csc = csc_matrix(l)
coo = coo_matrix(l)
lil = lil_matrix(l)
print(csr[1, 1])
print(csc[1, 1])
print(lil[1, 1])
print(coo[1, 1])
lil[0, 0] = 10
lil[0, 1] = 100
print(lil)
print(lil.toarray())
lil[1, 1] = 0
print(lil)
print(lil.toarray())
csr[0, 0] = 10
csr[0, 1] = 100
SparseEfficiencyWarning: Changing
csr[1, 1] = 0
print(csr)
print(csr.toarray())
import numpy as np
from scipy.sparse import csr_matrix, csc_matrix
from scipy.sparse.linalg import inv, eigs
csr = csr_matrix(l)
csc = csc_matrix(l)
csr_inv = linalg.inv(csr)
SparseEfficiencyWarning: splu
SparseEfficiencyWarning: spsolve
csc_inv = inv(csc)
print(csc_inv)
-1.0
-5.0
print(type(csc_inv))
scipy.sparse.csc.csc_matrix
print(csc_inv.toarray())
-1.  2
-1
w, v = np.linalg.eig(l)
print(w)
print(v)
csr_f = csr_matrix(l, dtype=float)
csr_i = csc_matrix(l, dtype=int)
w, v = eigs(csr_f, k=1)
print(w)
print(v)
w, v = eigs(csr_f, k=2)
TypeError: Cannot
scipy.linalg.eig
scipy.linalg.eig(A.toarray()) or reduce
w, v = eigs(csr_i, k=1)
ValueError: matrix
import numpy as np
from scipy.sparse import csr_matrix, lil_matrix
csr = csr_matrix(l)
lil = lil_matrix(l)
print(csr.sum())
print(csr.mean())
print(csr.max())
print(csr.min())
print(lil.max())
AttributeError: max
not found
print(csr.sqrt().toarray())
print(csr.sin().toarray())
print(csr.tan().toarray())
-0.14254654
print(lil.sqrt())
AttributeError: sqrt
not found
csr_ = csr.copy()
print(csr_.data)
print(type(csr_.data))
numpy.ndarray
csr_.data = np.cos(csr_.data)
print(csr_.toarray())
-0.9899925
csr_ = csr.copy()
csr_.data = csr_.data ** 2
print(csr_.toarray())
print(lil.data)
list([1, 2]) 
list([3, 4]) 
list([])
print(csr)
r, c = csr.nonzero()
print(r, c)
print(csr.count_nonzero())
print(csr.nnz)
csr[0, 1] = 0
print(csr)
print(csr.count_nonzero())
print(csr.nnz)
r, c = csr.nonzero()
print(r, c)
print(lil)
print(lil.nnz)
lil[0, 1] = 0
print(lil)
print(lil.nnz)
import numpy as np
from scipy.sparse import csr_matrix
a = np.eye(1000, dtype=np.int64)
print(type(a))
numpy.ndarray
print(a[:10, :10])
print(a.shape)
csr = csr_matrix(a)
print(type(csr))
scipy.sparse.csr.csr_matrix
def get_size_of_csr(csr):
return csr.data.nbytes + csr.indices.nbytes + csr.indptr.nbytes
print(a.nbytes)
print(get_size_of_csr(csr))
std. dev. of
std. dev. of
a_dense = np.random.randint(1, 100, (1000, 1000))
csr_dense = csr_matrix(a_dense)
print(a_dense[:10, :10])
print(a_dense.dtype)
print(a_dense.shape)
print(a_dense.nbytes)
print(get_size_of_csr(csr_dense))
std. dev. of
std. dev. of
a_10_10 = np.eye(10, dtype=np.int64)
csr_10_10 = csr_matrix(a_10_10)
print(a_10_10.nbytes)
print(get_size_of_csr(csr_10_10))
std. dev. of
std. dev. of
a_100_100 = np.eye(100, dtype=np.int64)
csr_100_100 = csr_matrix(a_100_100)
print(a_100_100.nbytes)
print(get_size_of_csr(csr_100_100))
std. dev. of
std. dev. of
a_200_200 = np.eye(200, dtype=np.int64)
csr_200_200 = csr_matrix(a_200_200)
print(a_200_200.nbytes)
print(get_size_of_csr(csr_200_200))
std. dev. of
std. dev. of
import numpy as np
from scipy.sparse import csr_matrix, csc_matrix, lil_matrix, save_npz, load_npz
csr = csr_matrix(l)
csc = csc_matrix(l)
save_npz('data/temp/csr.npz', csr)
save_npz('data/temp/csc.npz', csc)
csr_ = load_npz('data/temp/csr.npz')
print(csr_.toarray())
print(type(csr))
scipy.sparse.csr.csr_matrix
csc_ = load_npz('data/temp/csc.npz')
print(csc_.toarray())
print(type(csc))
scipy.sparse.csc.csc_matrix
lil = lil_matrix(l)
save_npz('data/temp/lil.npz', lil)
NotImplementedError: Save
not implemented 
npz = np.load('data/temp/csr.npz')
print(npz.files)
print(npz['data'])
print(npz['indices'])
print(npz['indptr'])
print(npz['format'])
print(npz['shape'])
import numpy as np
from scipy.sparse import csr_matrix, csc_matrix, lil_matrix
csr = csr_matrix(l)
csc = csc_matrix(l)
lil = lil_matrix(l)
print(csr[0, :])
print(csr[0, :].toarray())
print(csr[0].toarray())
print(csr[:, 0])
print(csr[:, 0].toarray())
print(csr[1:3, 1:3])
print(csr[1:3, 1:3].toarray())
print(csr[:, ::2])
print(csr[:, ::2].toarray())
print(type(csr[0]))
scipy.sparse.csr.csr_matrix
print(type(csc[0]))
scipy.sparse.csc.csc_matrix
print(type(lil[0]))
scipy.sparse.lil.lil_matrix
print(type(csr[:, 0]))
scipy.sparse.csr.csr_matrix
print(type(csc[:, 0]))
scipy.sparse.csc.csc_matrix
print(type(lil[:, 0]))
scipy.sparse.lil.lil_matrix
print(type(csr[1:3, 1:3]))
scipy.sparse.csr.csr_matrix
print(type(csc[1:3, 1:3]))
scipy.sparse.csc.csc_matrix
print(type(lil[1:3, 1:3]))
scipy.sparse.lil.lil_matrix
csr_slice = csr[1:3, 1:3]
csr_slice[0, 0] = 100
print(csr.toarray())
print(csr_slice.toarray())
csc_slice = csc[1:3, 1:3]
csc_slice[0, 0] = 100
print(csc.toarray())
print(csc_slice.toarray())
lil_slice = lil[1:3, 1:3]
lil_slice[0, 0] = 100
print(lil.toarray())
print(lil_slice.toarray())
lil[0] = [10, 20, 30, 40]
print(lil)
print(lil.toarray())
lil[1:3, 1:3] = np.arange(4).reshape(2, 2) * 100
print(lil)
print(lil.toarray())
lil[:, 0] = csr[:, 3]
print(lil)
print(lil.toarray())
lil[1:3, 1:3] = [10, 20, 30, 40]
ValueError: shape
mismatch: objects
csr[0] = [0, 0, 0, 100]
SparseEfficiencyWarning: Changing
self._set_arrayXarray(i, j, x)
print(csr)
print(csr.toarray())
from scipy.sparse import csr_matrix, triu, tril
csr = csr_matrix(l)
print(triu(csr).toarray())
print(type(triu(csr)))
scipy.sparse.coo.coo_matrix
print(type(triu(csr, format='csr')))
scipy.sparse.csr.csr_matrix
print(triu(csr, k=1).toarray())
print(triu(csr, k=-1).toarray())
print(tril(csr).toarray())
print(type(tril(csr)))
scipy.sparse.coo.coo_matrix
print(type(tril(csr, format='csr')))
scipy.sparse.csr.csr_matrix
print(tril(csr, k=1).toarray())
print(tril(csr, k=-1).toarray())
from scipy.special import comb
print(comb(4, 2))
print(comb(4, 2, exact=True))
print(comb(4, 0, exact=True))
print(comb(4, 2, exact=True, repetition=True))
from scipy.special import perm
print(perm(4, 2))
print(perm(4, 2, exact=True))
print(perm(4, 4, exact=True))
import scipy.stats
l = [0, 1, 2, 3, 4]
print(l)
print(scipy.stats.zscore(l))
print(type(scipy.stats.zscore(l)))
numpy.ndarray
print(scipy.stats.zscore(l, ddof=1))
l_2d = [[0, 1, 2], [3, 4, 5], [6, 7, 8]]
print(l_2d)
print(scipy.stats.zscore(l_2d))
print(scipy.stats.zscore(l_2d, ddof=1))
print(scipy.stats.zscore(l_2d, axis=1))
-1.22474487
-1.22474487
-1.22474487
print(scipy.stats.zscore(l_2d, axis=1, ddof=1))
-1.  0
-1.  0
-1.  0
print(scipy.stats.zscore(l_2d, axis=None))
-0.38729833
print(scipy.stats.zscore(l_2d, axis=None, ddof=1))
-0.36514837
import seaborn as sns
import pandas as pd
import numpy as np
import matplotlib as mpl
import matplotlib.pyplot as plt
list_2d = [[0, 1, 2], [3, 4, 5]]
plt.figure()
sns.heatmap(list_2d)
plt.savefig('data/dst/seaborn_heatmap_list.png')
plt.close('all')
arr_2d = np.arange(-8, 8).reshape((4, 4))
print(arr_2d)
plt.figure()
sns.heatmap(arr_2d)
plt.savefig('data/dst/seaborn_heatmap_ndarray.png')
df = pd.DataFrame(data=arr_2d, index=['a', 'b', 'c', 'd'], columns=['A', 'B', 'C', 'D'])
print(df)
plt.figure()
sns.heatmap(df)
plt.savefig('data/dst/seaborn_heatmap_dataframe.png')
print(type(sns.heatmap(list_2d)))
matplotlib.axes._subplots.AxesSubplot
fig = plt.figure()
ax = fig.add_subplot(1, 1, 1)
sns.heatmap(list_2d, ax=ax)
fig.savefig('data/dst/seaborn_heatmap_list.png')
fig, axes = plt.subplots(nrows=2, ncols=3, figsize=(8, 6))
sns.heatmap(list_2d, ax=axes[0, 0])
sns.heatmap(arr_2d, ax=axes[1, 2])
fig.savefig('data/dst/seaborn_heatmap_list_sub.png')
plt.figure()
sns.heatmap(df, annot=True)
plt.savefig('data/dst/seaborn_heatmap_annot.png')
plt.figure()
sns.heatmap(df, cbar=False)
plt.savefig('data/dst/seaborn_heatmap_no_cbar.png')
plt.figure()
sns.heatmap(df, square=True)
plt.savefig('data/dst/seaborn_heatmap_square.png')
plt.figure()
sns.heatmap(df, vmax=10, vmin=-10, center=0)
plt.savefig('data/dst/seaborn_heatmap_vmax_vmin_center.png')
plt.figure()
sns.heatmap(df, cmap='hot')
plt.savefig('data/dst/seaborn_heatmap_hot.png')
plt.figure()
sns.heatmap(df, cmap='Blues')
plt.savefig('data/dst/seaborn_heatmap_blues.png')
plt.figure()
sns.heatmap(df, cmap='Blues_r')
plt.savefig('data/dst/seaborn_heatmap_blues_r.png')
current_figsize = mpl.rcParams['figure.figsize']
print(current_figsize)
plt.figure(figsize=(9, 6)) 
sns.heatmap(df, square=True)
plt.savefig('data/dst/seaborn_heatmap_big.png')
current_dpi = mpl.rcParams['figure.dpi']
print(current_dpi)
plt.figure()
sns.heatmap(df, square=True)
plt.savefig('data/dst/seaborn_heatmap_big_2.png', dpi=current_dpi * 1.5)
df_house = pd.read_csv('data/src/house_prices_train.csv', index_col=0)
df_house_corr = df_house.corr()
print(df_house_corr.shape)
print(df_house_corr.head())
plt.figure(figsize=(12, 9)) 
sns.heatmap(df_house_corr, square=True, vmax=1, vmin=-1, center=0)
plt.savefig('data/dst/seaborn_heatmap_house_price.png')
import seaborn as sns
sns.set(style="ticks")
df = sns.load_dataset("iris")
sns.pairplot(df, hue='species', markers=["o", "s", "+"]).savefig('data/dst/seaborn_iris.png')
import pandas as pd
import seaborn as sns
df = pd.read_csv('data/src/iris.csv', index_col=0)
df = sns.load_dataset("iris")
print(df.head())
print(df.dtypes)
dtype: object
print(df['species'].value_counts())
Name: species
dtype: int64
pg = sns.pairplot(df)
print(type(pg))
seaborn.axisgrid.PairGrid
pg.savefig('data/dst/seaborn_pairplot_default.png')
sns.pairplot(df).savefig('data/dst/seaborn_pairplot_default.png')
sns.pairplot(df, hue='species').savefig('data/dst/seaborn_pairplot_hue.png')
sns.pairplot
df, hue='species'
hue_order=['virginica', 'versicolor', 'setosa']
savefig('data/dst/seaborn_pairplot_hue_order.png')
sns.pairplot(df, hue='species', palette='Blues').savefig('data/dst/seaborn_pairplot_palette.png')
sns.pairplot
df, hue='species'
savefig('data/dst/seaborn_pairplot_palette_dict.png')
sns.pairplot
df, hue='species'
vars=['sepal_length', 'sepal_width']
savefig('data/dst/seaborn_pairplot_vars.png')
sns.pairplot
df, hue='species'
x_vars=['sepal_length', 'sepal_width']
y_vars=['petal_length', 'petal_width']
savefig('data/dst/seaborn_pairplot_xy_vars.png')
sns.pairplot(df, hue='species', markers='+').savefig('data/dst/seaborn_pairplot_markers.png')
sns.pairplot(df, hue='species', markers=['+', 's', 'd']).savefig('data/dst/seaborn_pairplot_markers_multi.png')
sns.pairplot(df, hue='species', kind='reg').savefig('data/dst/seaborn_pairplot_kind_reg.png')
sns.pairplot(df, hue='species', diag_kind='kde').savefig('data/dst/seaborn_pairplot_diag_kind_kde.png')
sns.pairplot(df, hue='species', size=2).savefig('data/dst/seaborn_pairplot_size.png')
sns.pairplot
df, hue='species'
plot_kws={'alpha': 0.2}
diag_kws={'bins': 20, 'histtype': 'step'}
savefig('data/dst/seaborn_pairplot_kws.png')
sns.pairplot
df, hue='species', kind='reg'
--'}}).savefig('
s = {1, 2, 2, 3, 1, 4}
print(s)
print(type(s))
s = {1.23, '百', (0, 1, 2), '百'}
print(s)
s = {[0, 1, 2]}
TypeError: unhashable
s = {100, 100.0}
print(s)
s = {}
print(s)
print(type(s))
l = [1, 2, 2, 3, 1, 4]
print(l)
print(type(l))
s_l = set(l)
print(s_l)
print(type(s_l))
fs_l = frozenset(l)
print(fs_l)
print(type(fs_l))
frozenset({1, 2, 3, 4})
s = set()
print(s)
print(type(s))
set()
l = [2, 2, 3, 1, 3, 4]
l_unique = list(set(l))
print(l_unique)
s = {i**2 for i in range(5)}
print(s)
s = {1, 2, 2, 3, 1, 4}
print(s)
print(len(s))
s = {0, 1, 2}
s.add(3)
print(s)
s = {0, 1, 2}
s.discard(1)
print(s)
s = {0, 1, 2}
s.discard(10)
print(s)
s = {0, 1, 2}
s.remove(1)
print(s)
s = {0, 1, 2}
s.remove(10)
s = {2, 1, 0}
v = s.pop()
print(s)
print(v)
s = {2, 1, 0}
print(s.pop())
print(s.pop())
print(s.pop())
print(s.pop())
s = {0, 1, 2}
s.clear()
print(s)
set()
s1 = {0, 1, 2}
s2 = {1, 2, 3}
s3 = {2, 3, 4}
s_union = s1 | s2
print(s_union)
s_union = s1.union(s2)
print(s_union)
s_union = s1.union(s2, s3)
print(s_union)
s_union = s1.union(s2, [5, 6, 5, 7, 5])
print(s_union)
s_intersection = s1 & s2
print(s_intersection)
s_intersection = s1.intersection(s2)
print(s_intersection)
s_intersection = s1.intersection(s2, s3)
print(s_intersection)
s_difference = s1 - s2
print(s_difference)
s_difference = s1.difference(s2)
print(s_difference)
s_difference = s1.difference(s2, s3)
print(s_difference)
s_symmetric_difference = s1 ^ s2
print(s_symmetric_difference)
s_symmetric_difference = s1.symmetric_difference(s2)
print(s_symmetric_difference)
s1 = {0, 1}
s2 = {0, 1, 2, 3}
print(s1 <= s2)
print(s1.issubset(s2))
print(s1 <= s1)
print(s1.issubset(s1))
print(s1 < s1)
s1 = {0, 1}
s2 = {0, 1, 2, 3}
print(s2 >= s1)
print(s2.issuperset(s1))
print(s1 >= s1)
print(s1.issuperset(s1))
print(s1 > s1)
s1 = {0, 1}
s2 = {1, 2}
s3 = {2, 3}
print(s1.isdisjoint(s2))
print(s1.isdisjoint(s3))
def test():
print('function is called')
return True
print(True and test())
print(False and test())
print(True or test())
print(False or test())
import shutil
shutil.make_archive('data/temp/new_shutil', 'zip', root_dir='data/temp/dir')
shutil.make_archive('data/temp/new_shutil_sub', 'zip', root_dir='data/temp/dir', base_dir='dir_sub')
import shutil
import os
os.makedirs('temp/dir1/dir', exist_ok=True)
os.makedirs('temp/dir2', exist_ok=True)
with open('temp/dir1/file.txt', 'w') as f:
f.write('original')
print(os.listdir('temp/dir1/'))
file.txt
print(os.listdir('temp/dir2/'))
new_path = shutil.move('temp/dir1/file.txt', 'temp/dir2/')
print(new_path)
print(os.listdir('temp/dir1/'))
print(os.listdir('temp/dir2/'))
file.txt
new_path = shutil.move('temp/dir2/file.txt', 'temp/dir1/new_dir/')
file or directory
new_path = shutil.move('temp/dir1/dir/', 'temp/dir2/')
print(new_path)
print(os.listdir('temp/dir1/'))
print(os.listdir('temp/dir2/'))
file.txt
new_path = shutil.move('temp/dir2/file.txt', 'temp/dir1/file_new.txt')
print(new_path)
print(os.listdir('temp/dir1/'))
file_new.txt
print(os.listdir('temp/dir2/'))
new_path = shutil.move('temp/dir1/file_new.txt', 'temp/dir2/dir_new/file_new.txt')
file or directory
new_path = shutil.move('temp/dir2/dir/', 'temp/dir1/dir_new/')
print(new_path)
print(os.listdir('temp/dir1/'))
file_new.txt
print(os.listdir('temp/dir2/'))
new_path = shutil.move('temp/dir1/dir_new', 'temp/dir2/dir_new/dir_new2/')
print(new_path)
print(os.listdir('temp/dir1/'))
file_new.txt
print(os.listdir('temp/dir2/'))
print(os.listdir('temp/dir2/dir_new/'))
with open('temp/dir2/file_other.txt', 'w') as f:
f.write('other')
new_path = shutil.move('temp/dir1/file_new.txt', 'temp/dir2/file_other.txt')
print(new_path)
print(os.listdir('temp/dir1/'))
print(os.listdir('temp/dir2/'))
file_other.txt
with open('temp/dir2/file_other.txt') as f:
print(f.read())
shutil.rmtree('temp/dir1/')
shutil.rmtree('temp/dir2/')
import shutil
import os
os.makedirs('temp/dir1/dir', exist_ok=True)
os.makedirs('temp/dir2', exist_ok=True)
with open('temp/dir1/file.txt', 'w') as f:
f.write('original')
print(os.listdir('temp/dir1/'))
file.txt
print(os.listdir('temp/dir2/'))
src_dir = 'temp/dir1/'
dst_dir = 'temp/dir2/'
for p in os.listdir(src_dir):
shutil.move(os.path.join(src_dir, p), dst_dir)
print(os.listdir(src_dir))
print(os.listdir(dst_dir))
file.txt
shutil.rmtree('temp/dir1/')
shutil.rmtree('temp/dir2/')
import shutil
import glob
import os
def move_glob(dst_path, pathname, recursive=True):
for p in glob.glob(pathname, recursive=recursive):
shutil.move(p, dst_path)
os.mkdir('temp/dir2')
move_glob('temp/dir2', 'temp/**/*.txt')
import shutil
import glob
import re
import os
def move_glob_re(dst_path, pattern, pathname, recursive=True):
for p in glob.glob(pathname, recursive=recursive):
if re.search(pattern, p):
shutil.move(p, dst_path)
os.mkdir('temp/dir2')
import shutil
import os
os.makedirs('temp/dir/sub_dir/', exist_ok=True)
with open('temp/dir/file.txt', 'w') as f:
f.write('')
print(os.listdir('temp/'))
print(os.listdir('temp/dir/'))
file.txt
shutil.rmtree('temp/dir/file.txt')
shutil.rmtree('temp/dir/')
print(os.listdir('temp/'))
import skimage.io
import skimage.metrics
img_org = skimage.io.imread('data/src/lena.jpg')
img_q95 = skimage.io.imread('data/src/lena_q95.jpg')
img_q50 = skimage.io.imread('data/src/lena_q50.jpg')
print(skimage.metrics.peak_signal_noise_ratio(img_org, img_q95))
print(skimage.metrics.peak_signal_noise_ratio(img_org, img_q50))
print(skimage.metrics.peak_signal_noise_ratio(img_org, img_org))
RuntimeWarning: divide
return 10 * np.log10((data_range ** 2) / err)
print(skimage.metrics.peak_signal_noise_ratio(img_org, img_q95, data_range=255))
import numpy as np
import skimage.util
a = np.arange(1, 7).reshape(2, 3)
print(a)
b = a * 10
print(b)
c = a * 100
print(c)
m = skimage.util.montage([a, b, c])
print(m)
print(m.shape)
abc = np.array([a, b, c])
print(abc)
print(abc.shape)
print(skimage.util.montage(abc))
d = a[:, :2]
print(d)
skimage.util.montage([a, b, c, d])
ValueError: could
not broadcast 
shape (2,3) 
shape (2)
print(np.mean(np.array([a, b, c])))
print(skimage.util.montage([a, b, c], fill=0))
print(skimage.util.montage([a, b, c], grid_shape=(1, 3)))
print(skimage.util.montage([a, b, c], grid_shape=(3, 1)))
print(skimage.util.montage([a, b, c], grid_shape=(1, 2)))
IndexError: list
print(skimage.util.montage([a, b, c], grid_shape=(2, 3)))
print(skimage.util.montage([a, b, c], grid_shape=(2, 3), fill=0))
print(skimage.util.montage([a, b, c], padding_width=1))
print(skimage.util.montage([a, b, c], padding_width=1, fill=0))
import skimage.io
import skimage.util
a = skimage.io.imread('data/src/lena.jpg')
print(a.shape)
b = a // 2
c = a // 3
m = skimage.util.montage([a, b, c], multichannel=True)
print(m.shape)
skimage.io.imsave('data/dst/skimage_montage_default.jpg', m)
skimage.util.montage([a, b, c])
ValueError: Input
m_fill = skimage.util.montage([a, b, c], fill=(255, 128, 0), multichannel=True)
skimage.io.imsave('data/dst/skimage_montage_fill.jpg', m_fill)
m_1_3_pad = skimage.util.montage
fill=(0, 0, 0)
grid_shape=(1, 3)
padding_width=10
multichannel=True
print(m_1_3_pad.shape)
skimage.io.imsave('data/dst/skimage_montage_1_3_pad.jpg', m_1_3_pad)
import numpy as np
import skimage.util
a = np.arange(24).reshape(4, 6)
print(a)
blocks = skimage.util.view_as_blocks(a, (2, 3))
print(blocks)
print(type(blocks))
numpy.ndarray
print(blocks.shape)
print(blocks[0, 0])
print(blocks[0, 1])
print(blocks[1, 0])
print(blocks[1, 1])
blocks = skimage.util.view_as_blocks(a, (2, 4))
not compatible 
print(np.shares_memory(a, blocks))
a[0, 0] = 100
print(blocks[0, 0])
a = np.arange(24).reshape(4, 6)
blocks_copy = skimage.util.view_as_blocks(a, (2, 3)).copy()
print(np.shares_memory(a, blocks_copy))
a[0, 0] = 100
print(blocks_copy[0, 0])
blocks_copy2 = skimage.util.view_as_blocks(a.copy(), (2, 3))
print(np.shares_memory(a, blocks_copy2))
import numpy as np
import skimage.io
import skimage.util
img = skimage.io.imread('data/src/lena_square.png')
print(img.shape)
blocks = skimage.util.view_as_blocks(img, (256, 256, 3))
print(blocks.shape)
blocks = skimage.util.view_as_blocks(img, (256, 256))
arr_in.shape
skimage.io.imsave
skimage.io.imsave
skimage.io.imsave
skimage.io.imsave
print(np.shares_memory(img, blocks))
skimage.io.imsave('data/dst/skimage_block_change.jpg', img)
blocks_s = skimage.util.view_as_blocks(img, (256, 256, 3)).squeeze()
print(blocks_s.shape)
print(blocks_s[0, 0].shape)
from sklearn.metrics import accuracy_score
y_true = [0, 0, 0, 0, 0, 1, 1, 1, 1, 1]
y_pred = [0, 1, 1, 1, 1, 0, 0, 0, 1, 1]
print(accuracy_score(y_true, y_pred))
from sklearn.metrics import classification_report
import pandas as pd
import pprint
y_true = [0, 0, 0, 0, 0, 1, 1, 1, 1, 1]
y_pred = [0, 1, 1, 1, 1, 0, 0, 0, 1, 1]
print(classification_report(y_true, y_pred))
print(type(classification_report(y_true, y_pred)))
target_names=['class_0', 'class_1']
d = classification_report(y_true, y_pred, output_dict=True)
pprint.pprint(d)
print(d['0'])
print(d['0']['precision'])
print(type(d['0']['precision']))
df = pd.DataFrame(d)
print(df)
print(df.iloc[:, :-3])
print(df.iloc[:, -3:])
from sklearn.metrics import confusion_matrix
from sklearn.metrics import precision_score
from sklearn.metrics import classification_report
y_true_multi = [0, 0, 0, 1, 1, 1, 2, 2, 2]
y_pred_multi = [0, 1, 1, 1, 1, 2, 2, 2, 2]
print(confusion_matrix(y_true_multi, y_pred_multi))
print(precision_score(y_true_multi, y_pred_multi))
ValueError: Target
average='binary'. Please
print(precision_score(y_true_multi, y_pred_multi, average=None))
print(precision_score(y_true_multi, y_pred_multi, average='macro'))
print(precision_score(y_true_multi, y_pred_multi, average='micro'))
print(precision_score(y_true_multi, y_pred_multi, average='weighted'))
print(classification_report(y_true_multi, y_pred_multi))
from sklearn.metrics import confusion_matrix
y_true = [0, 0, 0, 0, 0, 1, 1, 1, 1, 1]
y_pred = [0, 1, 1, 1, 1, 0, 0, 0, 1, 1]
cm = confusion_matrix(y_true, y_pred)
print(cm)
print(type(cm))
numpy.ndarray
print(cm.flatten())
tn, fp, fn, tp = cm.flatten()
print(tn)
print(fp)
print(fn)
print(tp)
y_true_ab = ['A', 'A', 'A', 'A', 'A', 'B', 'B', 'B', 'B', 'B']
y_pred_ab = ['A', 'B', 'B', 'B', 'B', 'A', 'A', 'A', 'B', 'B']
print(confusion_matrix(y_true_ab, y_pred_ab))
print(confusion_matrix(y_true_ab, y_pred_ab, labels=['B', 'A']))
y_true_multi = [0, 0, 0, 1, 1, 1, 2, 2, 2]
y_pred_multi = [0, 1, 1, 1, 1, 2, 2, 2, 2]
print(confusion_matrix(y_true_multi, y_pred_multi))
print(confusion_matrix(y_true_multi, y_pred_multi, labels=[2, 1]))
from sklearn.metrics import confusion_matrix
import seaborn as sns
import matplotlib.pyplot as plt
y_true = [0, 0, 0, 0, 0, 1, 1, 1, 1, 1]
y_pred = [0, 1, 1, 1, 1, 0, 0, 0, 1, 1]
cm = confusion_matrix(y_true, y_pred)
print(cm)
sns.heatmap(cm)
plt.savefig('data/dst/sklearn_confusion_matrix.png')
plt.close()
sns.heatmap(cm, annot=True, cmap='Blues')
plt.savefig('data/dst/sklearn_confusion_matrix_annot_blues.png')
plt.close()
from sklearn.metrics import f1_score
y_true = [0, 0, 0, 0, 0, 1, 1, 1, 1, 1]
y_pred = [0, 1, 1, 1, 1, 0, 0, 0, 1, 1]
print(f1_score(y_true, y_pred))
from sklearn.datasets import fetch_olivetti_faces
data = fetch_olivetti_faces()
print(type(data))
sklearn.utils.Bunch
print(data.keys())
dict_keys(['data', 'images', 'target', 'DESCR'])
print(data.data.shape)
print(data.target.shape)
print(data.images.shape)
import pandas as pd
from sklearn import datasets, model_selection, svm, metrics
iris = datasets.load_iris()
print(type(iris))
print(iris.keys())
sklearn.datasets.base.Bunch
dict_keys(['data', 'target', 'target_names', 'DESCR', 'feature_names'])
iris_data = pd.DataFrame(data=iris.data, columns=iris.feature_names)
print(iris_data.head())
length (cm)  
width (cm)  
length (cm)  
width (cm)
iris_label = pd.Series(data=iris.target)
print(iris_label.head())
dtype: int64
print(len(iris_data))
data_train, data_test, label_train, label_test = model_selection.train_test_split(iris_data, iris_label)
print(data_train.head())
length (cm)  
width (cm)  
length (cm)  
width (cm)
print(label_train.head())
dtype: int64
test_size = 0.25
print(len(data_train), len(data_test))
clf = svm.SVC()
clf.fit(data_train, label_train)
pre = clf.predict(data_test)
print(type(pre))
print(pre)
numpy.ndarray
ac_score = metrics.accuracy_score(label_test, pre)
print(ac_score)
scores = model_selection.cross_val_score(clf, iris_data, iris_label, cv=3)
print(scores)
print(scores.mean())
import sklearn.datasets
data_iris = sklearn.datasets.load_iris()
print(type(data_iris))
sklearn.utils.Bunch
data_boston = sklearn.datasets.load_boston()
print(type(data_boston))
sklearn.utils.Bunch
import pandas as pd
from sklearn.datasets import load_iris
data = load_iris()
print(type(data))
sklearn.utils.Bunch
print(issubclass(type(data), dict))
print(data.keys())
dict_keys(['data', 'target', 'target_names', 'DESCR', 'feature_names', 'filename'])
print(data['feature_names'])
length (cm)
width (cm)
length (cm)
width (cm)
print(data.feature_names)
length (cm)
width (cm)
length (cm)
width (cm)
print(data.filename)
print(data.DESCR)
attributes and the
- sepal
- sepal
- petal
- petal
- class
Values: None
Creator: R
A. Fisher
Donor: Michael
io.arc.nasa.gov
Date: July
R.A. Fisher. The
paper. Note
not as in the 
literature.  Fisher
plant.  One
- Fisher
R.A
- Duda
R.O
P.E
Classification and Scene
Q327.D83
- Dasarathy
B.V
Neighborhood: A
Structure and Classification
Rule for Recognition in Partially
Analysis and Machine
- Gates
G.W
- See
- Many
X = data.data
y = data.target
print(type(X))
numpy.ndarray
print(X.shape)
print(X[:5])
print(type(y))
numpy.ndarray
print(y.shape)
print(y)
df_X = pd.DataFrame(data.data, columns=data.feature_names)
print(df_X.head())
length (cm)  
width (cm)  
length (cm)  
width (cm)
s_y = pd.Series(data.target)
print(s_y.head())
dtype: int64
print(df_X.describe())
length (cm)  
width (cm)  
length (cm)  \

width (cm)  
print(s_y.value_counts())
dtype: int64
X, y = load_iris(return_X_y=True)
print(type(X))
numpy.ndarray
print(X.shape)
print(type(y))
numpy.ndarray
print(y.shape)
from sklearn import datasets, model_selection, svm, metrics
mnist = datasets.fetch_mldata('MNIST original', data_home='data/src/download/')
print(type(mnist))
print(mnist.keys())
sklearn.datasets.base.Bunch
dict_keys(['DESCR', 'COL_NAMES', 'target', 'data'])
mnist_data = mnist.data / 255
mnist_label = mnist.target
print(mnist_data.shape)
print(mnist_label.shape)
train_size = 500
test_size = 100
data_train, data_test, label_train, label_test = model_selection.train_test_split(mnist_data, mnist_label, test_size=test_size, train_size=train_size)
clf = svm.SVC()
clf.fit(data_train, label_train)
pre = clf.predict(data_test)
ac_score = metrics.accuracy_score(label_test, pre)
print(ac_score)
import timeit
num = 10
train_size = 500
test_size = 100
data_train, data_test, label_train, label_test = model_selection.train_test_split(mnist_data, mnist_label, test_size=test_size, train_size=train_size)
clf = svm.SVC()
print(timeit.timeit(lambda: clf.fit(data_train, label_train), number=num) / num)
pre = clf.predict(data_test)
ac_score = metrics.accuracy_score(label_test, pre)
print(ac_score)
clf = svm.LinearSVC()
print(timeit.timeit(lambda: clf.fit(data_train, label_train), number=num) / num)
pre = clf.predict(data_test)
ac_score = metrics.accuracy_score(label_test, pre)
print(ac_score)
train_size = 10000
test_size = 2000
data_train, data_test, label_train, label_test = model_selection.train_test_split(mnist_data, mnist_label, test_size=test_size, train_size=train_size)
clf = svm.LinearSVC()
print(timeit.timeit(lambda: clf.fit(data_train, label_train), number=num) / num)
pre = clf.predict(data_test)
ac_score = metrics.accuracy_score(label_test, pre)
print(ac_score)
co_mat = metrics.confusion_matrix(label_test, pre)
print(co_mat)
from sklearn.metrics import precision_score
y_true = [0, 0, 0, 0, 0, 1, 1, 1, 1, 1]
y_pred = [0, 1, 1, 1, 1, 0, 0, 0, 1, 1]
print(precision_score(y_true, y_pred))
from sklearn.metrics import precision_score
from sklearn.metrics import confusion_matrix
y_true = [0, 0, 0, 0, 0, 1, 1, 1, 1, 1]
y_pred = [0, 1, 1, 1, 1, 0, 0, 0, 1, 1]
print(precision_score(y_true, y_pred))
print(precision_score(y_true, y_pred, pos_label=0))
print(precision_score(y_true, y_pred, average=None))
print(precision_score(y_true, y_pred, average='macro'))
print(precision_score(y_true, y_pred, average='micro'))
print(confusion_matrix(y_true, y_pred))
print(confusion_matrix(y_true, y_pred, labels=[1, 0]))
print(precision_score(y_true, y_pred, average='weighted'))
y_true_2 = [0, 1, 1, 1, 1]
y_pred_2 = [0, 0, 0, 0, 1]
print(confusion_matrix(y_true_2, y_pred_2))
print(confusion_matrix(y_true_2, y_pred_2, labels=[1, 0]))
print(precision_score(y_true_2, y_pred_2))
print(precision_score(y_true_2, y_pred_2, pos_label=0))
print(precision_score(y_true_2, y_pred_2, average='macro'))
print(precision_score(y_true_2, y_pred_2, average='micro'))
print(precision_score(y_true_2, y_pred_2, average='weighted'))
y_true_ab = ['A', 'A', 'A', 'A', 'A', 'B', 'B', 'B', 'B', 'B']
y_pred_ab = ['A', 'B', 'B', 'B', 'B', 'A', 'A', 'A', 'B', 'B']
print(precision_score(y_true_ab, y_pred_ab))
ValueError: pos_label=1 is not a
label: array
dtype='<U1'
print(precision_score(y_true_ab, y_pred_ab, pos_label='A'))
from sklearn import preprocessing
l = [0, 1, 2, 3, 4]
print(l)
l_2d = [[0, 1, 2], [3, 4, 5], [6, 7, 8]]
print(l_2d)
mm = preprocessing.MinMaxScaler()
mm.fit_transform(l)
ValueError: Expected
l_2d_min_max = mm.fit_transform(l_2d)
print(l_2d_min_max)
print(type(l_2d_min_max))
numpy.ndarray
print(preprocessing.minmax_scale(l))
print(preprocessing.minmax_scale(l_2d))
print(preprocessing.minmax_scale(l_2d, axis=1))
ss = preprocessing.StandardScaler()
print(ss.fit_transform(l))
ValueError: Expected
l_2d_standardization = ss.fit_transform(l_2d)
print(l_2d_standardization)
print(type(l_2d_standardization))
numpy.ndarray
print(preprocessing.scale(l))
print(preprocessing.scale(l_2d))
print(preprocessing.scale(l_2d, axis=1))
-1.22474487
-1.22474487
-1.22474487
from sklearn.metrics import recall_score
y_true = [0, 0, 0, 0, 0, 1, 1, 1, 1, 1]
y_pred = [0, 1, 1, 1, 1, 0, 0, 0, 1, 1]
print(recall_score(y_true, y_pred))
from sklearn.metrics import roc_auc_score
import numpy as np
y_true = np.array([0, 0, 0, 0, 1, 1, 1, 1])
y_score = np.array([0.2, 0.3, 0.6, 0.8, 0.4, 0.5, 0.7, 0.9])
print(roc_auc_score(y_true, y_score))
y_true_perfect = np.array([0, 0, 0, 0, 1, 1, 1, 1])
y_score_perfect = np.array([0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8])
print(roc_auc_score(y_true_perfect, y_score_perfect))
np.random.seed(0)
y_true_random = np.array([0] * 5000 + [1] * 5000)
y_score_random = np.random.rand(10000)
print(roc_auc_score(y_true_random, y_score_random))
y_score_inv = 1 - y_score
print(y_score_inv)
print(roc_auc_score(y_true, y_score_inv))
y_score_perfect_inv = 1 - y_score_perfect
print(y_score_perfect_inv)
print(roc_auc_score(y_true_perfect, y_score_perfect_inv))
from sklearn.metrics import roc_curve
import matplotlib.pyplot as plt
y_true = [0, 0, 0, 0, 1, 1, 1, 1]
y_score = [0.2, 0.3, 0.6, 0.8, 0.4, 0.5, 0.7, 0.9]
roc = roc_curve(y_true, y_score)
print(type(roc))
print(len(roc))
fpr, tpr, thresholds = roc_curve(y_true, y_score)
print(fpr)
print(tpr)
print(thresholds)
plt.plot(fpr, tpr, marker='o')
plt.xlabel('FPR: False positive rate')
plt.ylabel('TPR: True positive rate')
plt.grid()
plt.savefig('data/dst/sklearn_roc_curve.png')
plt.close()
fpr_all, tpr_all, thresholds_all = roc_curve
drop_intermediate=False
print(fpr_all)
print(tpr_all)
print(thresholds_all)
plt.plot(fpr_all, tpr_all, marker='o')
plt.xlabel('FPR: False positive rate')
plt.ylabel('TPR: True positive rate')
plt.grid()
plt.savefig('data/dst/sklearn_roc_curve_all.png')
plt.close()
from sklearn.metrics import roc_curve, recall_score, confusion_matrix
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
y_true = np.array([0, 0, 0, 0, 1, 1, 1, 1])
y_score = np.array([0.2, 0.3, 0.6, 0.8, 0.4, 0.5, 0.7, 0.9])
print(y_score >= 0.5)
print((y_score >= 0.5).astype(int))
def fpr_score(y_true, y_pred):
tn, fp, fn, tp = confusion_matrix(y_true, y_pred).flatten()
return fp / (tn + fp)
print(fpr_score(y_true, y_score >= 0.5))
print(recall_score(y_true, y_score >= 0.5))
th_min = min(y_score)
print(th_min)
print((y_score >= th_min).astype(int))
print(fpr_score(y_true, y_score >= th_min))
print(recall_score(y_true, y_score >= th_min))
th_max = max(y_score) + 1
print(th_max)
print((y_score >= th_max).astype(int))
print(fpr_score(y_true, y_score >= th_max))
print(recall_score(y_true, y_score >= th_max))
df = pd.DataFrame({'true': y_true, 'score': y_score})
df['TPR'] = df.apply(lambda row: recall_score(y_true, y_score >= row['score']), axis=1)
df['FPR'] = df.apply(lambda row: fpr_score(y_true, y_score >= row['score']), axis=1)
print(df)
print(df.sort_values('score', ascending=False))
fpr_all, tpr_all, th_all = roc_curve
drop_intermediate=False
df_roc = pd.DataFrame({'th_all': th_all, 'tpr_all': tpr_all, 'fpr_all': fpr_all})
print(df_roc)
y_true_perfect = np.array([0, 0, 0, 0, 1, 1, 1, 1])
y_score_perfect = np.array([0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8])
print(y_true_perfect)
print((y_score_perfect >= 0.5).astype(int))
print(fpr_score(y_true_perfect, y_score_perfect >= 0.5))
print(recall_score(y_true_perfect, y_score_perfect >= 0.5))
roc_p = roc_curve(y_true_perfect, y_score_perfect, drop_intermediate=False)
plt.plot(roc_p[0], roc_p[1], marker='o')
plt.xlabel('FPR: False positive rate')
plt.ylabel('TPR: True positive rate')
plt.grid()
plt.savefig('data/dst/sklearn_roc_curve_perfect.png')
plt.close()
y_true_1 = np.array([0, 0, 0, 1, 0, 1, 1, 1])
y_score_1 = y_score_perfect
roc_1 = roc_curve(y_true_1, y_score_1, drop_intermediate=False)
y_true_2 = np.array([0, 0, 1, 1, 0, 0, 1, 1])
y_score_2 = y_score_perfect
roc_2 = roc_curve(y_true_2, y_score_2, drop_intermediate=False)
plt.plot(roc_p[0], roc_p[1], marker='s')
plt.plot(roc_1[0], roc_1[1], marker='o')
plt.plot(roc_2[0], roc_2[1], marker='x')
plt.xlabel('FPR: False positive rate')
plt.ylabel('TPR: True positive rate')
plt.grid()
plt.savefig('data/dst/sklearn_roc_curve_compare.png')
plt.close()
y_true_org = np.array([0, 0, 1, 1, 0, 0, 1, 1])
y_score_org = np.array([0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8])
roc_org = roc_curve(y_true_org, y_score_org, drop_intermediate=False)
y_score_scale = y_score_org / 2
print(y_score_scale)
roc_scale = roc_curve(y_true_org, y_score_scale, drop_intermediate=False)
y_score_interval = np.array([0.01, 0.02, 0.91, 0.92, 0.93, 0.94, 0.95, 0.96])
roc_interval = roc_curve(y_true_org, y_score_interval, drop_intermediate=False)
plt.plot(roc_org[0], roc_org[1], marker='s')
plt.plot(roc_scale[0], roc_scale[1], marker='o', linestyle='-.')
plt.plot(roc_interval[0], roc_interval[1], marker='x', linestyle=':')
plt.xlabel('FPR: False positive rate')
plt.ylabel('TPR: True positive rate')
plt.grid()
plt.savefig('data/dst/sklearn_roc_curve_same.png')
plt.close()
s = pd.Series(y_score_interval)
print(s)
dtype: float64
print(s.rank())
dtype: float64
np.random.seed(0)
y_true_random = np.array([0] * 5000 + [1] * 5000)
y_score_random = np.random.rand(10000)
roc_random = roc_curve(y_true_random, y_score_random)
plt.plot(roc_random[0], roc_random[1])
plt.xlabel('FPR: False positive rate')
plt.ylabel('TPR: True positive rate')
plt.grid()
plt.savefig('data/dst/sklearn_roc_curve_random.png')
plt.close()
import pandas as pd
from sklearn.model_selection import train_test_split
from sklearn.datasets import load_iris
data = load_iris()
X_df = pd.DataFrame(data['data'], columns=data['feature_names'])
y_s = pd.Series(data['target'])
print(X_df)
length (cm)  
width (cm)  
length (cm)  
width (cm)
print(type(X_df))
pandas.core.frame.DataFrame
print(X_df.shape)
print(y_s)
dtype: int64
print(type(y_s))
pandas.core.series.Series
print(y_s.shape)
X_train_df, X_test_df, y_train_s, y_test_s = train_test_split
X_df, y_s, test_size=0.2, random_state=0, stratify=y_s
print(type(X_train_df))
pandas.core.frame.DataFrame
print(X_train_df.shape)
print(type(X_test_df))
pandas.core.frame.DataFrame
print(X_test_df.shape)
print(type(y_train_s))
pandas.core.series.Series
print(y_train_s.shape)
print(type(y_test_s))
pandas.core.series.Series
print(y_test_s.shape)
print(y_train_s.value_counts())
dtype: int64
print(y_test_s.value_counts())
dtype: int64
from sklearn.model_selection import train_test_split
from sklearn.datasets import load_iris
data = load_iris()
X = data['data']
y = data['target']
print(X.shape)
print(X[:5])
print(y.shape)
print(y)
X_train, X_test, y_train, y_test = train_test_split(X, y, random_state=0)
print(X_train.shape)
print(X_test.shape)
print(y_train.shape)
print(y_test.shape)
print(y_test)
print((y_test == 0).sum())
print((y_test == 1).sum())
print((y_test == 2).sum())
X_train, X_test, y_train, y_test = train_test_split(X, y, random_state=0, stratify=y)
print(y_test)
print((y_test == 0).sum())
print((y_test == 1).sum())
print((y_test == 2).sum())
import numpy as np
from sklearn.model_selection import train_test_split
a = np.arange(10)
print(a)
print(train_test_split(a))
array([3, 9, 6, 1, 5, 0, 7])
array([2, 8, 4])
print(type(train_test_split(a)))
print(len(train_test_split(a)))
a_train, a_test = train_test_split(a)
print(a_train)
print(a_test)
a_train, a_test = train_test_split(a, test_size=0.6)
print(a_train)
print(a_test)
a_train, a_test = train_test_split(a, test_size=6)
print(a_train)
print(a_test)
a_train, a_test = train_test_split(a, train_size=0.6)
print(a_train)
print(a_test)
a_train, a_test = train_test_split(a, train_size=6)
print(a_train)
print(a_test)
a_train, a_test = train_test_split(a, train_size=0.25)
print(a_train)
print(a_test)
a_train, a_test = train_test_split(a, test_size=0.3, train_size=0.4)
print(a_train)
print(a_test)
a_train, a_test = train_test_split(a, test_size=3, train_size=4)
print(a_train)
print(a_test)
a_train, a_test = train_test_split(a, test_size=0.8, train_size=0.7)
ValueError: The
test_size and train_size = 1.500000
1.0. Reduce
a_train, a_test = train_test_split(a, test_size=8, train_size=7)
ValueError: The
train_size and test_size = 15
a_train, a_test = train_test_split(a, shuffle=False)
print(a_train)
print(a_test)
a_train, a_test = train_test_split(a, random_state=0)
print(a_train)
print(a_test)
X = np.arange(20).reshape(2, 10).T
print(X)
y = np.arange(10)
print(y)
X_train, X_test, y_train, y_test = train_test_split(X, y)
print(X_train)
print(X_test)
print(y_train)
print(y_test)
z = np.arange(10) * 10
print(z)
X_train, X_test, y_train, y_test, z_train, z_test = train_test_split(X, y, z)
print(X_train)
print(X_test)
print(y_train)
print(y_test)
print(z_train)
print(z_test)
y_mismatch = np.arange(8)
print(y_mismatch)
X_train, X_test, y_train, y_test = train_test_split(X, y_mismatch)
ValueError: Found
y = np.array([0] * 5 + [1] * 5)
print(y)
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=100)
print(y_train)
print(y_test)
X_train, X_test, y_train, y_test = train_test_split
X, y, test_size=0.2, random_state=100
stratify=y
print(y_train)
print(y_test)
l = [0, 10, 20, 30, 40, 50, 60]
print(l)
print(l[2:5])
print(l[:3])
print(l[3:])
print(l[:])
print(l[2:10])
print(l[5:2])
print(l[2:2])
print(l[10:20])
print(l[::2])
print(l[1::2])
print(l[::3])
print(l[2:5:2])
print(l[3:-1])
print(l[-2:])
print(l[-5:-2])
print(l[5:2:-1])
print(l[2:5:-1])
print(l[-2:-5:-1])
print(l[-2:2:-1])
print(l[5:2:-2])
print(l[::-1])
sl = slice(2, 5, 2)
print(sl)
slice(2, 5, 2)
print(type(sl))
print(l[sl])
sl = slice(2, 5)
print(sl)
slice(2, 5, None)
print(l[sl])
sl = slice(2)
print(sl)
slice(None, 2, None)
print(l[sl])
sl = slice()
TypeError: slice
sl = slice(None)
print(sl)
slice(None, None, None)
print(l[sl])
print(l)
l[2:5] = [200, 300, 400]
print(l)
l[2:5] = [-2, -3]
print(l)
-2
-3
l[2:4] = [2000, 3000, 4000, 5000]
print(l)
l[2:6] = [20000]
print(l)
l[2:3] = 200
TypeError: can
l[1:4] = []
print(l)
l[20:60] = [-1, -2, -3]
print(l)
-1
-2
-3
l[2:2] = [-100]
print(l)
-100
-1
-2
-3
print(l[:5:2])
-100
-2
l[:5:2] = [100, 200, 300]
print(l)
-1
-3
l[:5:2] = [100, 200]
ValueError: attempt
l_2d = [[0, 1, 2], [3, 4, 5], [6, 7, 8], [9, 10, 11]]
print(l_2d)
print(l_2d[1:3])
print([l[:2] for l in l_2d[1:3]])
l_2d_t = [list(x) for x in zip(*l_2d)]
print(l_2d_t)
print(l_2d_t[1])
l = [0, 10, 20, 30, 40, 50, 60]
print(l)
l_slice = l[2:5]
print(l_slice)
l_slice[1] = 300
print(l_slice)
print(l)
l_2d = [[0, 1, 2], [3, 4, 5], [6, 7, 8], [9, 10, 11]]
print(l_2d)
l_2d_slice = l_2d[1:3]
print(l_2d_slice)
l_2d_slice[0][1] = 400
print(l_2d_slice)
print(l_2d)
import copy
l_2d = [[0, 1, 2], [3, 4, 5], [6, 7, 8], [9, 10, 11]]
print(l_2d)
l_2d_slice_deepcopy = copy.deepcopy(l_2d[1:3])
print(l_2d_slice_deepcopy)
l_2d_slice_deepcopy[0][1] = 400
print(l_2d_slice_deepcopy)
print(l_2d)
s = 'abcdefg'
print(s)
print(s[2:5])
print(s[::-1])
s[2:5] = 'CDE'
not support 
t = (0, 10, 20, 30, 40, 50, 60)
print(t)
print(t[2:5])
t[2:5] = (200, 300, 400)
not support 
l = ['Banana', 'Alice', 'Apple', 'Bob']
print(sorted(l))
print(sorted(l, reverse=True))
l_order = ['Alice', 'Bob', 'Apple', 'Banana']
print(sorted(l, key=l_order.index))
print(l)
print([l_order.index(s) for s in l])
print(sorted(l, key=l_order.index))
d_order = {'Alice': 0, 'Bob': 1, 'Apple': 2, 'Banana': 3}
l, key=lambda x: d_order[x]
l = ['Banana', 'Alice', 'Apple', 'Bob', 'xxx']
l_order = ['Alice', 'Bob', 'Apple', 'Banana']
print(sorted(l, key=l_order.index))
l, key=lambda x: l_order.index(x) if x in l_order else -1
l, key=lambda x: l_order.index(x) if x in l_order else float('inf')
def my_index(x):
l_order = ['Alice', 'Bob', 'Apple', 'Banana']
return l_order.index(x) if x in l_order else -1
print(sorted(l, key=my_index))
l = ['Banana', 'Alice', 'Apple', 'Bob']
l_order = ['Alice', 'Bob', 'Apple', 'Banana', 'xxx']
print(sorted(l, key=l_order.index))
l = ['Banana', 'Alice', 'Apple', 'Bob', 'xxx']
d_order = {'Alice': 0, 'Bob': 1, 'Apple': 2, 'Banana': 3}
l, key=lambda x: d_order[x]
l, key=lambda x: d_order.get(x, -1)
l, key=lambda x: d_order.get(x, float('inf'))
l = ['Banana', 'Alice', 'Apple', 'Bob']
d_order = {'Alice': 0, 'Bob': 1, 'Apple': 2, 'Banana': 3, 'xxx': 4}
l, key=lambda x: d_order[x]
import pprint
l = ['沖縄県', '東京都', '北海道', '京都府']
print(sorted(l))
print(sorted(l, key=tdfk.index))
d_tdfk = dict(zip(tdfk, range(len(tdfk))))
print(d_tdfk)
l, key=lambda x: d_tdfk[x]
l = ['沖縄県', '東京都', '北海道', '京都府', 'xxx']
print(sorted(l, key=tdfk.index))
l, key=lambda x: tdfk.index(x) if x in tdfk else -1
l, key=lambda x: tdfk.index(x) if x in tdfk else float('inf')
l, key=lambda x: d_tdfk.get(x, -1)
l, key=lambda x: d_tdfk.get(x, float('inf'))
l = ['沖縄県', '東京', '北海道', '京都府']
print(sorted(l, key=tdfk.index))
tdfk_head2 = [s[:2] for s in tdfk]
pprint.pprint(tdfk_head2, compact=True)
l, key=lambda x: tdfk_head2.index(x[:2])
l = ['沖縄県', '東京', '北海道', '京都府', 'xxx']
l, key=lambda x: tdfk_head2.index(x[:2]) if x[:2] in tdfk_head2 else -1
l = ['沖縄県', '東京', '東京都', '北海道', '京都府']
l, key=lambda x: tdfk_head2.index(x[:2])
l = ['沖縄県', '東京都', '東京', '北海道', '京都府']
l, key=lambda x: tdfk_head2.index(x[:2])
tdfk_ex = sorted
key=lambda x: tdfk_head2.index(x[:2])
tdfk_ex.remove('北海')
pprint.pprint(tdfk_ex, compact=True)
l = ['沖縄県', '東京', '東京都', '北海道', '京都府']
print(sorted(l, key=tdfk_ex.index))
l = ['沖縄県', '東京都', '東京', '北海道', '京都府']
print(sorted(l, key=tdfk_ex.index))
l = [10, 1, 5]
l.sort()
print(l)
l = [10, 1, 5]
print(sorted(l))
print(l)
print(sorted(l, reverse=True))
l = ['10', '01', '05']
print(sorted(l))
l = ['10', '1', '5']
print(sorted(l))
l = ['10', '1', '5']
print(sorted(l, key=int))
print(sorted(l, key=float))
l = ['10.0', '1.0', '5.0']
print(sorted(l, key=float))
l = ['10', '1', '5']
l.sort(key=int)
print(l)
l = ['10', '1', '5']
print([int(s) for s in l])
print(sorted([int(s) for s in l]))
import re
s = 'file5.txt'
re.search
group()
re.search
group()
re.search
group()
l = ['file10.txt', 'file1.txt', 'file5.txt']
print(sorted(l))
file1.txt
file10.txt
file5.txt
l, key=lambda s: int
re.search
group()
file1.txt
file5.txt
file10.txt
p = re.compile
l, key=lambda s: int(p.search(s).group())
file1.txt
file5.txt
file10.txt
s = '100file5.txt'
re.search
group()
re.findall
re.findall
re.search
groups()
re.search
groups()[0]
re.search
groups()[0]
l = ['100file10.txt', '100file1.txt', '100file5.txt']
l, key=lambda s: int
re.findall
file1.txt
file5.txt
file10.txt
l, key=lambda s: int
re.search
groups()[0]
file1.txt
file5.txt
file10.txt
l, key=lambda s: int
re.search
groups()[0]
file1.txt
file5.txt
file10.txt
p = re.compile
l, key=lambda s: int(p.search(s).groups()[0])
file1.txt
file5.txt
file10.txt
l = ['file10.txt', 'file1.txt', 'file5.txt', 'file.txt']
l, key=lambda s:int
re.search
group()
def extract_num(s, p, ret=0):
search = p.search(s)
if search:
return int(search.groups()[0])
return ret
p = re.compile
print(extract_num('file10.txt', p))
print(extract_num('file.txt', p))
print(extract_num('file.txt', p, 100))
l, key=lambda s: extract_num(s, p)
file.txt
file1.txt
file5.txt
file10.txt
l, key=lambda s: extract_num(s, p, float('inf'))
file1.txt
file5.txt
file10.txt
file.txt
l = ['100file10.txt', '100file1.txt', '100file5.txt', '100file.txt']
p = re.compile
l, key=lambda s: extract_num(s, p)
file.txt
file1.txt
file5.txt
file10.txt
l, key=lambda s: extract_num(s, p, float('inf'))
file1.txt
file5.txt
file10.txt
file.txt
org_list = [3, 1, 4, 5, 2]
org_list.sort()
print(org_list)
print(org_list.sort())
org_list.sort(reverse=True)
print(org_list)
org_list = [3, 1, 4, 5, 2]
new_list = sorted(org_list)
print(org_list)
print(new_list)
new_list_reverse = sorted(org_list, reverse=True)
print(org_list)
print(new_list_reverse)
org_str = 'cebad'
new_str_list = sorted(org_str)
print(org_str)
print(new_str_list)
new_str = ''.join(new_str_list)
print(new_str)
new_str = ''.join(sorted(org_str))
print(new_str)
new_str_reverse = ''.join(sorted(org_str, reverse=True))
print(new_str_reverse)
org_tuple = (3, 1, 4, 5, 2)
new_tuple_list = sorted(org_tuple)
print(org_tuple)
print(new_tuple_list)
new_tuple = tuple(new_tuple_list)
print(new_tuple)
new_tuple = tuple(sorted(new_tuple_list))
print(new_tuple)
new_tuple_reverse = tuple(sorted(new_tuple_list, reverse=True))
print(new_tuple_reverse)
s = 'one two three'
l = s.split()
print(l)
s = 'one two        three'
l = s.split()
print(l)
s = 'one\ttwo\tthree'
l = s.split()
print(l)
s = 'one::two::three'
l = s.split('::')
print(l)
s = 'one,two,three'
l = s.split(',')
print(l)
s = 'one, two, three'
l = s.split(',')
print(l)
s = 'one, two, three'
l = s.split(', ')
print(l)
s = 'one, two,  three'
l = s.split(', ')
print(l)
s = '  one  '
print(s.strip())
print(s)
s = '-+-one-+-'
print(s.strip('-+'))
s = '-+- one -+-'
print(s.strip('-+'))
s = '-+- one -+-'
print(s.strip('-+ '))
s = 'one, two, three'
l = [x.strip() for x in s.split(',')]
print(l)
s = ''
l = [x.strip() for x in s.split(',')]
print(l)
print(len(l))
s = ''
l = [x.strip() for x in s.split(',') if not s == '']
print(l)
print(len(l))
s = 'one, , three'
l = [x.strip() for x in s.split(',')]
print(l)
print(len(l))
s = 'one, ,three'
l = [x.strip() for x in s.split(',') if not x.strip() == '']
print(l)
print(len(l))
s = '1, 2, 3, 4'
l = [x.strip() for x in s.split(',')]
print(l)
print(type(l[0]))
s = '1, 2, 3, 4'
l = [int(x.strip()) for x in s.split(',')]
print(l)
print(type(l[0]))
s = 'one, two,  three'
l = [x.strip() for x in s.split(',')]
print(l)
print(','.join(l))
print('::'.join(l))
s = 'one, two,  three'
s_new = '-'.join([x.strip() for x in s.split(',')])
print(s_new)
s = 'one,two,three'
s_new = s.replace(',', '+')
print(s_new)
import sqlite3
path = ':memory:'
conn = sqlite3.connect(path)
c = conn.cursor()
c.execute("DROP TABLE IF EXISTS items;")
c.execute
c.execute("INSERT INTO items (name, value) VALUES (?, ?)", ('one', 100))
c.execute("INSERT INTO items (name, value) VALUES (?, ?)", ('two', 200))
c.execute("INSERT INTO items (name, value) VALUES (?, ?)", ('three', 300))
data = [('four', 400), ('five', 500), ('six', 600)]
c.executemany("INSERT INTO items (name, value) VALUES (?, ?)", data)
conn.commit()
c.execute("SELECT * FROM items")
print(c.fetchall())
c.execute("SELECT * FROM items WHERE id >= ?", (4, ))
print(c.fetchall())
conn.close()
import math
print(math.cos(0))
from math import sin
print(sin(0))
import statistics
import math
l = [10, 1, 3, 7, 1]
mean = statistics.mean(l)
print(mean)
my_mean = sum(l) / len(l)
print(my_mean)
harmonic_mean = statistics.harmonic_mean(l)
print(harmonic_mean)
my_harmonic_mean = len(l) / sum(1 / x for x in l)
print(my_harmonic_mean)
median = statistics.median(l)
print(median)
l_even = [10, 1, 3, 7, 1, 6]
median = statistics.median(l_even)
print(median)
median_low = statistics.median_low(l_even)
print(median_low)
median_high = statistics.median_high(l_even)
print(median_high)
print(statistics.median_high(l) == statistics.median_low(l) == statistics.median(l))
mode = statistics.mode(l)
print(mode)
l_mode_error = [1, 2, 3, 4, 5]
mode = statistics.mode(l_mode_error)
StatisticsError: no
l_mode_error = [1, 1, 1, 2, 2, 2, 3]
mode = statistics.mode(l_mode_error)
StatisticsError: no
pvariance = statistics.pvariance(l)
print(pvariance)
my_pvariance = sum((x - sum(l) / len(l))**2 for x in l) / len(l)
print(my_pvariance)
pstdev = statistics.pstdev(l)
print(pstdev)
print(math.sqrt(pvariance))
variance = statistics.variance(l)
print(variance)
my_variance = sum((x - sum(l) / len(l))**2 for x in l) / (len(l) - 1)
print(my_variance)
stdev = statistics.stdev(l)
print(stdev)
print(math.sqrt(variance))
print('abc' == 'abc')
print('abc' == 'xyz')
print('abc' == 'ABC')
print('abc' != 'xyz')
print('abc' != 'abc')
print('bbb' in 'aaa-bbb-ccc')
print('xxx' in 'aaa-bbb-ccc')
print('abc' in 'aaa-bbb-ccc')
print('xxx' not in 'aaa-bbb-ccc')
print('bbb' not in 'aaa-bbb-ccc')
s = 'aaa-bbb-ccc'
print(s.startswith('aaa'))
print(s.startswith('bbb'))
print(s.startswith(('aaa', 'bbb', 'ccc')))
print(s.startswith(('xxx', 'yyy', 'zzz')))
print(s.startswith(['a', 'b', 'c']))
TypeError: startswith
str or a
not list
print(s.endswith('ccc'))
print(s.endswith('bbb'))
print(s.endswith(('aaa', 'bbb', 'ccc')))
print('a' < 'b')
print('aa' < 'ab')
print('abc' < 'abcd')
print(ord('a'))
print(ord('b'))
print('Z' < 'a')
print(ord('Z'))
print('あ' < 'い')
print(ord('あ'))
print(ord('い'))
print('ん' < 'ア')
print(ord('ん'))
print(ord('ア'))
print('乙' < '亜')
print(ord('乙'))
print(ord('亜'))
print('七' < '三')
print(ord('七'))
print(ord('三'))
print(sorted(['aaa', 'abc', 'Abc', 'ABC']))
print(sorted('一二三四五六七八九十'))
s1 = 'abc'
s2 = 'ABC'
print(s1 == s2)
print(s1.lower() == s2.lower())
s = 'Abcあいうえお'
print(s.lower())
print(s.upper())
import re
s = 'aaa-AAA-123'
print(re.search('aaa', s))
re.Match
span=(0, 3)
match='aaa'
print(re.search('xxx', s))
print(re.search('^aaa', s))
re.Match
span=(0, 3)
match='aaa'
print(re.search('^123', s))
print(re.search('aaa$', s))
print(re.search('123$', s))
re.Match
span=(8, 11)
match='123'
print(re.search('[A-Z]+', s))
re.Match
span=(4, 7)
match='AAA'
s = '012-3456-7890'
re.fullmatch
re.Match
span=(0, 13)
match='012-3456-7890'
s = 'tel: 012-3456-7890'
re.fullmatch
s = '012-3456-7890'
re.search
re.Match
span=(0, 13)
match='012-3456-7890'
s = 'tel: 012-3456-7890'
re.search
s = 'ABC'
print(re.search('abc', s))
print(re.search('abc', s, re.IGNORECASE))
re.Match
span=(0, 3)
match='ABC'
i = 100
s_i = str(i)
print(s_i)
print(type(s_i))
f = 0.123
s_f = str(f)
print(s_f)
print(type(s_f))
i = 0xFF
print(i)
s_i = str(i)
print(s_i)
f = 1.23e+10
print(f)
s_f = str(f)
print(s_f)
s_i_format = format(i, '#X')
print(s_i_format)
s_f_format = format(f, '.2e')
print(s_f_format)
l = [0, 1, 2]
s_l = str(l)
print(s_l)
print(type(s_l))
s_d = str(d)
print(s_d)
print(type(s_d))
import pprint
dl = {'a': 1, 'b': 2, 'c': [100, 200, 300]}
s_dl = str(dl)
print(s_dl)
p_dl = pprint.pformat(dl, width=10)
print(p_dl)
print(type(p_dl))
import re
s = '012-3456-7890'
re.search
re.Match
span=(0, 3)
match='012'
m = re.search
print(m.group())
print(type(m.group()))
re.findall
print(re.findall('a.*b', 'axyzb'))
print(re.findall('a.*b', 'a---b'))
print(re.findall('a.*b', 'aあいうえおb'))
print(re.findall('a.*b', 'ab'))
print(re.findall('a.+b', 'ab'))
print(re.findall('a.+b', 'axb'))
print(re.findall('a.+b', 'axxxxxxb'))
print(re.findall('a.?b', 'ab'))
print(re.findall('a.?b', 'axb'))
print(re.findall('a.?b', 'axxb'))
print(re.findall('a(.*)b', 'axyzb'))
re.findall
abc(def)
re.findall
abc(def)
s = 'axb-axxxxxxb'
print(re.findall('a.*b', s))
print(re.findall('a.*?b', s))
print(re.findall('a.+b', s))
print(re.findall('a.+?b', s))
print(re.findall('[abc]x', 'ax-bx-cx'))
print(re.findall('[abc]+', 'abc-aaa-cba'))
print(re.findall('[a-z]+', 'abc-xyz'))
s = 'abc-012-あいうえお'
print(re.findall('[0-9]+', s))
print(re.findall('[a-z]+', s))
print(re.findall('[ぁ-ゟ]+', s))
s = 'abc-def-ghi'
print(re.findall('[a-z]+', s))
print(re.findall('^[a-z]+', s))
print(re.findall('[a-z]+$', s))
s = 'axxxb-012'
print(re.findall('a.*b', s))
re.findall
re.findall
s = 'abc-Abc-ABC'
print(re.findall('[a-z]+', s))
print(re.findall('[A-Z]+', s))
print(re.findall('[a-z]+', s, flags=re.IGNORECASE))
s = 'I am Sam'
print('Sam' in s)
print('sam' in s)
print('I' in s and 'Sam' in s)
s = 'I am Sam'
print(s.find('Sam'))
print(s.find('XXX'))
-1
print(s.find('am'))
print(s.find('am', 3))
print(s.find('am', 3, 5))
-1
print(s.rfind('am'))
print(s.rfind('XXX'))
-1
print(s.rfind('am', 2))
print(s.rfind('am', 2, 5))
print(s.index('am'))
print(s.index('XXX'))
ValueError: substring
not found
print(s.rindex('am'))
print(s.rindex('XXX'))
ValueError: substring
not found
s = 'I am Sam'
print(s.count('am'))
print(s.count('XXX'))
print(s.count('am', 2, 4))
s = 'aaaa'
print(s.count('aa'))
s = 'I am Sam'
l = s.split()
print(l)
print(l.count('am'))
s = 'I am Sam'
print(s.upper())
print(s.lower())
print('sam' in s)
print('sam' in s.lower())
print(s.find('sam'))
-1
print(s.lower().find('sam'))
print(s.count('sam'))
print(s.lower().count('sam'))
s = '私はSam'
print(s.lower())
print(s.upper())
s = 'abcde'
print(s[0])
print(s[4])
print(s[-1])
print(s[-5])
print(s[5])
IndexError: string
print(s[-6])
IndexError: string
s = 'abcde'
print(s[1:3])
print(s[:3])
print(s[1:])
print(s[-4:-2])
print(s[:-2])
print(s[-4:])
print(s[3:1])
print(s[3:1] == '')
print(s[-100:100])
print(s[1:4:2])
print(s[::2])
print(s[::3])
print(s[::-1])
print(s[::-2])
s = 'abcdefghi'
print(len(s))
print(s[len(s) / 2])
TypeError: string
print(s[len(s) // 2])
print(s[:len(s) // 2])
print(s[len(s) // 2:])
s = 'abcあいう'
print(len(s))
print(s[4])
print(s[1:5])
s = 'abcde'
print(len(s))
s_length = len(s)
print(s_length)
print(type(s_length))
s = 'あいうえお'
print(len(s))
s = 'abcdeあいうえお'
print(len(s))
s = 'a\tb\\c'
print(s)
print(len(s))
s = r
print(s)
print(len(s))
s = '\u3042\u3044\u3046'
print(s)
print(len(s))
s = r
print(s)
print(len(s))
s = 'a\nb'
print(s)
print(len(s))
s = 'a\r\nb'
print(s)
print(len(s))
s = 'abc\nabcd\r\nab'
print(s)
print(len(s))
print(s.splitlines())
print(len(s.splitlines()))
print([len(line) for line in s.splitlines()])
print(sum(len(line) for line in s.splitlines()))
s = 'abc'
print(s)
print(type(s))
s = "abc"
print(s)
print(type(s))
s_sq = 'abc'
s_dq = "abc"
print(s_sq == s_dq)
s_sq = 'a\'b"c'
print(s_sq)
s_sq = 'a\'b\"c'
print(s_sq)
s_dq = "a'b\"c"
print(s_dq)
s_dq = "a\'b\"c"
print(s_dq)
s_sq = 'a\'b"c'
s_dq = "a'b\"c"
print(s_sq == s_dq)
SyntaxError: EOL
s = 'abc\nxyz'
print(s)
s_tq = ''
print(s_tq)
print(type(s_tq))
s_tq = '''abc'
print(s_tq)
s_tq_sq = ''
print(s_tq_sq)
s_tq_dq = ""
print(s_tq_dq)
print(s_tq_sq == s_tq_dq)
s_tq = ''
print(s_tq)
print(s_multi)
print(int('100'))
print(type(int('100')))
print(int('1.23'))
ValueError: invalid
int() 
print(int('10,000'))
ValueError: invalid
int() 
print(int('10,000'.replace(',', '')))
print(float('1.23'))
print(type(float('1.23')))
print(float('.23'))
print(float('100'))
print(type(float('100')))
print(int('100', 2))
print(int('100', 8))
print(int('100', 16))
print(int('100', 10))
print(int('100'))
print(int('0b100', 0))
print(int('0o100', 0))
print(int('0x100', 0))
print(int('FF', 16))
print(int('ff', 16))
print(int('0xFF', 0))
print(int('0XFF', 0))
print(int('0xff', 0))
print(int('0Xff', 0))
print(float('1.23e-4'))
print(type(float('1.23e-4')))
print(float('1.23e4'))
print(type(float('1.23e4')))
print(float('1.23E-4'))
print(int('１００'))
print(type(int('１００')))
print(float('１００'))
print(type(float('１００')))
print(float('ー１．２３'))
ValueError: could
not convert 
print(float('-１.２３'))
-1.23
print(float('ー１．２３'.replace('ー', '-').replace('．', '.')))
-1.23
import unicodedata
print(unicodedata.numeric('五'))
print(type(unicodedata.numeric('五')))
print(unicodedata.numeric('十'))
print(unicodedata.numeric('参'))
print(unicodedata.numeric('億'))
print(unicodedata.numeric('五十'))
TypeError: numeric
not str
print(unicodedata.numeric('漢'))
ValueError: not
s = '1234567890'
print('s =', s)
print('isdecimal:', s.isdecimal())
print('isdigit:', s.isdigit())
print('isnumeric:', s.isnumeric())
s = 1234567890
isdecimal: True
isdigit: True
isnumeric: True
s = '１２３４５６７８９０'
print('s =', s)
print('isdecimal:', s.isdecimal())
print('isdigit:', s.isdigit())
print('isnumeric:', s.isnumeric())
s = １２３４５６７８９０
isdecimal: True
isdigit: True
isnumeric: True
s = '-1.23'
print('s =', s)
print('isdecimal:', s.isdecimal())
print('isdigit:', s.isdigit())
print('isnumeric:', s.isnumeric())
s = -1.23
isdecimal: False
isdigit: False
isnumeric: False
s = '10\u00B2'
print('s =', s)
print('isdecimal:', s.isdecimal())
print('isdigit:', s.isdigit())
print('isnumeric:', s.isnumeric())
s = 10
isdecimal: False
isdigit: True
isnumeric: True
s = '\u00BD'
print('s =', s)
print('isdecimal:', s.isdecimal())
print('isdigit:', s.isdigit())
print('isnumeric:', s.isnumeric())
isdecimal: False
isdigit: False
isnumeric: True
s = '\u2166'
print('s =', s)
print('isdecimal:', s.isdecimal())
print('isdigit:', s.isdigit())
print('isnumeric:', s.isnumeric())
isdecimal: False
isdigit: False
isnumeric: True
s = '一二三四五六七八九〇'
print('s =', s)
print('isdecimal:', s.isdecimal())
print('isdigit:', s.isdigit())
print('isnumeric:', s.isnumeric())
isdecimal: False
isdigit: False
isnumeric: True
s = '壱億参阡萬'
print('s =', s)
print('isdecimal:', s.isdecimal())
print('isdigit:', s.isdigit())
print('isnumeric:', s.isnumeric())
isdecimal: False
isdigit: False
isnumeric: True
s = 'abc'
print('s =', s)
print('isalpha:', s.isalpha())
s = abc
isalpha: True
s = 'あいうえお'
print('s =', s)
print('isalpha:', s.isalpha())
isalpha: True
s = 'アイウエオ'
print('s =', s)
print('isalpha:', s.isalpha())
isalpha: True
s = '漢字'
print('s =', s)
print('isalpha:', s.isalpha())
isalpha: True
s = '1234567890'
print('s =', s)
print('isalpha:', s.isalpha())
s = 1234567890
isalpha: False
s = '１２３４５６７８９０'
print('s =', s)
print('isalpha:', s.isalpha())
s = １２３４５６７８９０
isalpha: False
s = '一二三四五六七八九'
print('s =', s)
print('isalpha:', s.isalpha())
isalpha: True
s = '壱億参阡萬'
print('s =', s)
print('isalpha:', s.isalpha())
isalpha: True
s = '〇'
print('s =', s)
print('isalpha:', s.isalpha())
isalpha: False
s = '\u2166'
print('s =', s)
print('isalpha:', s.isalpha())
isalpha: False
s = 'abc123'
print('s =', s)
print('isalnum:', s.isalnum())
print('isalpha:', s.isalpha())
print('isdecimal:', s.isdecimal())
print('isdigit:', s.isdigit())
print('isnumeric:', s.isnumeric())
s = abc123
isalnum: True
isalpha: False
isdecimal: False
isdigit: False
isnumeric: False
s = 'abc123+-,.&'
print('s =', s)
print('isascii:', s.isascii())
print('isalnum:', s.isalnum())
s = abc123
isascii: True
isalnum: False
s = 'あいうえお'
print('s =', s)
print('isascii:', s.isascii())
print('isalnum:', s.isalnum())
isascii: False
isalnum: True
s = ''
print('s =', s)
print('isalnum:', s.isalnum())
print('isalpha:', s.isalpha())
print('isdecimal:', s.isdecimal())
print('isdigit:', s.isdigit())
print('isnumeric:', s.isnumeric())
print('isascii:', s.isascii())
isalnum: False
isalpha: False
isdecimal: False
isdigit: False
isnumeric: False
isascii: True
print(bool(''))
print(bool('abc123'))
s = '-1.23'
print('s =', s)
print('isalnum:', s.isalnum())
print('isalpha:', s.isalpha())
print('isdecimal:', s.isdecimal())
print('isdigit:', s.isdigit())
print('isnumeric:', s.isnumeric())
print('isascii:', s.isascii())
s = -1.23
isalnum: False
isalpha: False
isdecimal: False
isdigit: False
isnumeric: False
isascii: True
print(float('-1.23'))
-1.23
print(type(float('-1.23')))
print(float('abc'))
ValueError: could
not convert 
def is_num(s):
float(s)
except ValueError:
return False
return True
print(is_num('123'))
print(is_num('-1.23'))
print(is_num('+1.23e10'))
print(is_num('abc'))
print(is_num('10,000,000'))
def is_num_delimiter(s):
float(s.replace(',', ''))
except ValueError:
return False
return True
print(is_num_delimiter('10,000,000'))
def is_num_delimiter2(s):
float(s.replace(',', '').replace(' ', ''))
except ValueError:
return False
return True
print(is_num_delimiter2('10,000,000'))
print(is_num_delimiter2('10 000 000'))
import re
s_nums = 'one1two22three333four'
re.split
re.split
s_marks = 'one-two+three#four'
print(re.split('[-+#]', s_marks))
s_strs = 'oneXXXtwoYYYthreeZZZfour'
print(re.split('XXX|YYY|ZZZ', s_strs))
import re
s = 'aaa@xxx.com bbb@yyy.com ccc@zzz.com'
print(re.sub('[a-z]*@', 'ABC@', s))
xxx.com
yyy.com
zzz.com
print(re.sub('[a-z]*@', 'ABC@', s, 2))
xxx.com
yyy.com
zzz.com
print(re.sub('[xyz]', '1', s))
print(re.sub('aaa|bbb|ccc', 'ABC', s))
xxx.com
yyy.com
zzz.com
print(re.sub('([a-z]*)@', '\\1-123@', s))
xxx.com
yyy.com
zzz.com
re.sub
xxx.com
yyy.com
zzz.com
t = re.subn('[a-z]*@', 'ABC@', s)
print(t)
xxx.com
yyy.com
zzz.com
print(type(t))
print(t[0])
xxx.com
yyy.com
zzz.com
print(t[1])
l = ['Alice', 'Bob', 'Charlie']
print([s.strip('bce') for s in l])
print([s[:2] for s in l])
s = 'Alice\nBob\nCharlie'
print(s)
print(s.replace('li', ''))
print(s.strip('bce'))
print(s[2:-2])
l_s = s.splitlines()
print(l_s)
l_s_strip = [line.strip('bce') for line in l_s]
print(l_s_strip)
s_line_strip = '\n'.join(l_s_strip)
print(s_line_strip)
print('\n'.join([line[:2] for line in s.splitlines()]))
l_remove = [line for line in s.splitlines() if not line.startswith('B')]
print(l_remove)
s_line_remove = '\n'.join(l_remove)
print(s_line_remove)
print('\n'.join([line for line in s.splitlines() if 'li' in line]))
s = '0123456789'
print(s[3:7])
print(s[3:-3])
print(s[:5])
print(s[5:])
print(s[:3] + s[6:])
def remove_str_start_end(s, start, end):
return s[:start] + s[end + 1:]
print(remove_str_start_end(s, 3, 5))
def remove_str_start_length(s, start, length):
return s[:start] + s[start + length:]
print(remove_str_start_length(s, 3, 5))
s = 'abc-abcxyz'
print(s.removeprefix('abc-'))
print(s.removeprefix('aabc-'))
print(s.lstrip('abc-'))
def my_removeprefix(s, prefix):
if s.startswith(prefix):
return s[len(prefix):]
return s
print(my_removeprefix(s, 'abc-'))
s = 'abcxyz-xyz'
print(s.removesuffix('-xyz'))
print(s.removesuffix('-xyzz'))
def my_removesuffix(s, suffix):
return s[:-len(suffix)] if s.endswith(suffix) else s
print(my_removesuffix(s, '-xyz'))
s = 'abc-abcxyz-xyz'
print(s.removeprefix('abc-').removesuffix('-xyz'))
print(my_removeprefix(my_removesuffix(s, '-xyz'), 'abc-'))
s = 'one two one two one'
print(s.replace(' ', '-'))
print(s.replace(' ', ''))
print(s.replace('one', 'XXX'))
print(s.replace('one', 'XXX', 2))
print(s.replace('one', 'XXX').replace('two', 'YYY'))
print(s.replace('one', 'XtwoX').replace('two', 'YYY'))
print(s.replace('two', 'YYY').replace('one', 'XtwoX'))
s_lines = 'one\ntwo\nthree'
print(s_lines)
print(s_lines.replace('\n', '-'))
s_lines_multi = 'one\ntwo\r\nthree'
print(s_lines_multi)
print(repr(s_lines_multi))
print(s_lines_multi.replace('\r\n', '-').replace('\n', '-'))
print(repr(s_lines_multi.replace('\r\n', '-').replace('\n', '-')))
print(s_lines_multi.replace('\n', '-').replace('\r\n', '-'))
-threeo
print(repr(s_lines_multi.replace('\n', '-').replace('\r\n', '-')))
print(s_lines_multi.splitlines())
print('-'.join(s_lines_multi.splitlines()))
s = 'one two one two one'
print(s.translate(str.maketrans({'o': 'O', 't': 'T'})))
print(s.translate(str.maketrans({'o': 'XXX', 't': None})))
print(s.translate(str.maketrans('ow', 'XY', 'n')))
print(s.translate(str.maketrans('ow', 'XXY', 'n')))
ValueError: the
s = 'abcdefghij'
print(s[:4] + 'XXX' + s[7:])
s_replace = 'XXX'
i = 4
print(s[:i] + s_replace + s[i + len(s_replace):])
print(s[:4] + '-' + s[7:])
print(s[:4] + '+++++' + s[4:])
s = 'abc-xyz-123-789-ABC-XYZ'
print(s.replace('xyz', ''))
import re
s = 'abc-xyz-123-789-ABC-XYZ'
re.sub
import re
s = 'I am Sam'
print(re.search('Sam', s))
re.Match
span=(5, 8)
match='Sam'
print(re.search('XXX', s))
m = re.search('Sam', s)
print(m.group())
print(m.start())
print(m.end())
print(m.span())
s = 'I am Sam'
print(re.search('am', s))
re.Match
span=(2, 4)
match='am'
print(re.findall('am', s))
print(len(re.findall('am', s)))
print([m.span() for m in re.finditer('am', s)])
s = 'I am Sam Adams'
print(re.findall('Sam|Adams', s))
print([m.span() for m in re.finditer('Sam|Adams', s)])
s = 'I am Sam Adams'
print(re.findall('am', s))
print(re.findall('[a-zA-Z]+am[a-z]*', s))
s = 'I am Sam'
print(re.search('sam', s))
print(re.search('sam', s, flags=re.IGNORECASE))
re.Match
span=(5, 8)
match='Sam'
s_blank = 'one two     three\nfour\tfive'
print(s_blank)
print(s_blank.split())
print(type(s_blank.split()))
s_comma = 'one,two,three,four,five'
print(s_comma.split(','))
print(s_comma.split('three'))
print(s_comma.split(',', 2))
s_lines = 'one\ntwo\nthree\nfour'
print(s_lines)
print(s_lines.split('\n', 1))
print(s_lines.split('\n', 1)[0])
print(s_lines.split('\n', 1)[1])
print(s_lines.split('\n', 1)[-1])
print(s_lines.split('\n', 2)[-1])
print(s_lines.rsplit('\n', 1))
print(s_lines.rsplit('\n', 1)[0])
print(s_lines.rsplit('\n', 1)[1])
print(s_lines.rsplit('\n', 2)[0])
s_lines_multi = '1 one\n2 two\r\n3 three\n'
print(s_lines_multi)
print(s_lines_multi.split())
print(s_lines_multi.split('\n'))
print(s_lines_multi.splitlines())
print(s_lines_multi.splitlines(True))
l = ['one', 'two', 'three']
print(','.join(l))
print('\n'.join(l))
print(''.join(l))
s = 'abcdefghij'
print(s[:5])
print(s[5:])
s_tuple = s[:5]
s[5:]
print(s_tuple)
print(type(s_tuple))
s_first, s_last = s[:5], s[5:]
print(s_first)
print(s_last)
s_first, s_second, s_last = s[:3], s[3:6], s[6:]
print(s_first)
print(s_second)
print(s_last)
half = len(s) // 2
print(half)
s_first, s_last = s[:half], s[half:]
print(s_first)
print(s_last)
print(s_first + s_last)
s = ' \n a b c　\t'
print(s)
print(repr(s))
print(s.strip())
print(repr(s.strip()))
s_strip = s.strip()
print(repr(s_strip))
print(repr(s))
s = s.strip()
print(repr(s))
s = 'aabbcc-abc-aabbcc'
print(s.strip('abc'))
-abc
print(s.strip('cba'))
-abc
print(s.strip('ab'))
s = ' \n aabbcc-abc-aabbcc　\t'
print(repr(s))
print(repr(s.strip('abc')))
print(repr(s.strip('abc \n　\t')))
-abc
print(repr(s.strip().strip('abc')))
-abc
s = ' \n a b c 　\t'
print(repr(s.lstrip()))
s = 'aabbcc-abc-aabbcc'
print(s.lstrip('abc'))
s = ' \n a b c 　\t'
print(repr(s.rstrip()))
s = 'aabbcc-abc-aabbcc'
print(s.rstrip('abc'))
s = 'one two one two one'
print(s.replace('one', 'two').replace('two', 'one'))
print(s.replace('one', 'X').replace('two', 'one').replace('X', 'two'))
def swap_str(s_org, s1, s2, temp='*q@w-e~r^'):
return s_org.replace(s1, temp).replace(s2, s1).replace(temp, s2)
print(swap_str(s, 'one', 'two'))
print(s.replace('o', 't').replace('t', 'o'))
print(s.translate(str.maketrans({'o': 't', 't': 'o'})))
print(s.translate(str.maketrans('ot', 'to')))
s_org = 'pYThon proGramminG laNguAge'
print(s_org.upper())
print(s_org)
s_new = s_org.upper()
print(s_new)
s_org = s_org.upper()
print(s_org)
s_org = 'Pyhon Ｐｙｔｈｏｎ パイソン 123'
print(s_org.upper())
s_org = 'pYThon proGramminG laNguAge'
print(s_org.upper())
s_org = 'pYThon proGramminG laNguAge'
print(s_org.lower())
s_org = 'pYThon proGramminG laNguAge'
print(s_org.capitalize())
s_org = 'pYThon proGramminG laNguAge'
print(s_org.title())
s_org = 'pYThon proGramminG laNguAge'
print(s_org.swapcase())
print('PYTHON'.isupper())
print('Python'.isupper())
print('ＰＹＴＨＯＮ'.isupper())
print('PYTHON パイソン 123'.isupper())
print('パイソン 123'.isupper())
print('python'.islower())
print('Python'.islower())
print('ｐｙｔｈｏｎ'.islower())
print('python パイソン 123'.islower())
print('パイソン 123'.islower())
print('Python Programming Language'.istitle())
print('PYTHON Programming Language'.istitle())
print('★Python Programming Language'.istitle())
print('Python★ Programming Language'.istitle())
print('Py★thon Programming Language'.istitle())
print('The 1st Team'.istitle())
print('The 1St Team'.istitle())
print('パイソン 123'.istitle())
s1 = 'python'
s2 = 'PYTHON'
print(s1 == s2)
print(s1.upper() == s2.upper())
print(s1.lower() == s2.lower())
print(s1.capitalize() == s2.capitalize())
print(s1.title() == s2.title())
s = 'aaa' + 'bbb' + 'ccc'
print(s)
s1 = 'aaa'
s2 = 'bbb'
s3 = 'ccc'
s = s1 + s2 + s3
print(s)
s = s1 + s2 + s3 + 'ddd'
print(s)
s1 += s2
print(s1)
s = 'aaa'
s += 'xxx'
print(s)
s = 'aaa'
print(s)
s = 'aaa'
print(s)
s = 'aaa'
print(s)
s = s1
SyntaxError: invalid
s1 = 'aaa'
s2 = 'bbb'
i = 100
f = 0.25
s = s1 + i
TypeError: must
not int
s = s1 + '_' + str(i) + '_' + s2 + '_' + str(f)
print(s)
s = s1 + '_' + format(i, '05') + '_' + s2 + '_' + format(f, '.5f')
print(s)
s = '{}_{:05}_{}_{:.5f}'.format(s1, i, s2, f)
print(s)
s = '{}_{}_{}_{}'.format(s1, i, s2, f)
print(s)
s = f'{s1}_{i:05}_{s2}_{f:.5f}
print(s)
s = f'{s1}_{i}_{s2}_{f}
print(s)
l = ['aaa', 'bbb', 'ccc']
s = ''.join(l)
print(s)
s = ','.join(l)
print(s)
s = '-'.join(l)
print(s)
s = '\n'.join(l)
print(s)
l = [0, 1, 2]
s = '-'.join(l)
TypeError: sequence
s = '-'.join([str(n) for n in l])
print(s)
s = '-'.join((str(n) for n in l))
print(s)
s = '-'.join(str(n) for n in l)
print(s)
s = 'Line1\nLine2\nLine3'
print(s)
s = 'Line1\r\nLine2\r\nLine3'
print(s)
s = ''
print(s)
s = ''
print(s)
s = 'Line1\n'
print(s)
s = 'Line1\n'
print(s)
print(s)
print(s)
s = ''
print(s)
s = ''
print(s)
l = ['Line1', 'Line2', 'Line3']
s_n = '\n'.join(l)
print(s_n)
print(repr(s_n))
s_rn = '\r\n'.join(l)
print(s_rn)
print(repr(s_rn))
s = 'Line1\nLine2\r\nLine3'
print(s.splitlines())
s = 'Line1\nLine2\r\nLine3'
print(''.join(s.splitlines()))
print(' '.join(s.splitlines()))
print(','.join(s.splitlines()))
s_n = '\n'.join(s.splitlines())
print(s_n)
print(repr(s_n))
s = 'Line1\nLine2\nLine3'
print(s.replace('\n', ''))
print(s.replace('\n', ','))
s = 'Line1\nLine2\r\nLine3'
s_error = s.replace('\n', ',')
print(s_error)
print(repr(s_error))
s_error = s.replace('\r\n', ',')
print(s_error)
print(repr(s_error))
s = 'Line1\nLine2\r\nLine3'
print(s.replace('\r\n', ',').replace('\n', ','))
s_error = s.replace('\n', ',').replace('\r\n', ',')
print(s_error)
print(repr(s_error))
print(','.join(s.splitlines()))
s = 'aaa\n'
print(s + 'bbb')
print(s.rstrip() + 'bbb')
print('a')
print('b')
print('c')
print('a', end='')
print('b', end='')
print('c', end='')
print('a', end='-')
print('b', end='-')
print('c')
from distutils.util import strtobool
print(strtobool('true'))
print(strtobool('True'))
print(strtobool('TRUE'))
print(strtobool('t'))
print(strtobool('yes'))
print(strtobool('y'))
print(strtobool('on'))
print(strtobool('1'))
print(strtobool('false'))
print(strtobool('False'))
print(strtobool('FALSE'))
print(strtobool('f'))
print(strtobool('no'))
print(strtobool('n'))
print(strtobool('off'))
print(strtobool('0'))
print(strtobool('abc'))
ValueError: invalid
strtobool('abc')
except ValueError as e:
print('other value')
print(type(strtobool('true')))
if strtobool('yes'):
print('True!')
import pprint
pprint.pprint(ArithmeticError.__subclasses__())
decimal.DecimalException
print(type(ArithmeticError.__subclasses__()))
print(type(ArithmeticError.__subclasses__()[0]))
Base()
pass
class SubA(Base
pass
class SubB(Base
pass
class SubSubA(SubA
pass
class SUbSubB(SubB
pass
print(Base.__subclasses__())
__main__.SubA
__main__.SubB
def all_subclasses(cls):
return set(cls.__subclasses__()).union
s for c in cls.__subclasses__() for s in all_subclasses(c)
pprint.pprint(all_subclasses(Base))
__main__.SubA
__main__.SubB
__main__.SubSubA
__main__.SUbSubB
print(len(dict.__subclasses__()))
import pandas as pd
print(len(dict.__subclasses__()))
a = 1
b = 2
a, b = b, a
print('a = ', a)
print('b = ', b)
a =  2
b =  1
a, b = 100, 200
print('a = ', a)
print('b = ', b)
a =  100
b =  200
a, b, c, d = 0, 1, 2, 3
a, b, c, d = c, d, a, b
print('a = ', a)
print('b = ', b)
print('c = ', c)
print('d = ', d)
a =  2
b =  3
c =  0
d =  1
l = [0, 1, 2, 3, 4]
l[0], l[3] = l[3], l[0]
print(l)
print(sorted(l))
print(sorted(l, reverse=True))
import sympy
x = sympy.Symbol('x')
y = sympy.Symbol('y')
print(type(x))
sympy.core.symbol.Symbol
expr = x**2 + y + 1
print(expr)
z = sympy.Symbol('ZZZZ')
expr_z = z**2 + 3 * z
print(expr_z)
print(expr)
print(expr.subs(x, 1))
print(expr.subs(x, y))
print(expr.subs([(x, 1), (y, 2)]))
expr = (x + 1)**2
print(expr)
expr_ex = sympy.expand(expr)
print(expr_ex)
expr_factor = sympy.factor(expr_ex)
print(expr_factor)
print(sympy.factor(x**3 - x**2 - 3 * x + 3))
print(sympy.factor(x * y + x + y + 1))
print(sympy.solve(x**2 - 3 * x + 2))
print(sympy.solve(x**2 + x + 1))
expr = x + y**2 - 4
print(sympy.solve(expr, x))
print(sympy.solve(expr, y))
-sqrt(-x + 4)
sqrt(-x + 4)
expr1 = 3 * x + 5 * y - 29
expr2 = x + y - 7
print(sympy.solve([expr1, expr2]))
print(sympy.diff(x**3 + 2 * x**2 + x))
expr = x**3 + y**2 - y
print(sympy.diff(expr, x))
print(sympy.diff(expr, y))
print(sympy.integrate(3 * x**2 + 4 * x + 1))
print(sympy.diff(sympy.cos(x)))
-sin(x)
print(sympy.diff(sympy.exp(x)))
exp(x)
print(sympy.diff(sympy.log(x)))
print(sympy.integrate(sympy.cos(x)))
sin(x)
print(sympy.integrate(sympy.exp(x)))
exp(x)
print(sympy.integrate(sympy.log(x)))
import sympy
from sympy import sin, exp
x = sympy.Symbol('x')
print(sympy.diff(sin(x)))
cos(x)
print(sympy.diff(exp(x)))
exp(x)
import sys
print('sys.argv         : ', sys.argv)
print('type(sys.argv)   : ', type(sys.argv))
print('len(sys.argv)    : ', len(sys.argv))
print()
print('sys.argv[0]      : ', sys.argv[0])
print('sys.argv[1]      : ', sys.argv[1])
print('sys.argv[2]      : ', sys.argv[2])
print('type(sys.argv[0]): ', type(sys.argv[0]))
print('type(sys.argv[1]): ', type(sys.argv[1]))
print('type(sys.argv[2]): ', type(sys.argv[2]))
import sys
print(sys.float_info)
sys.float_info(max=1.7976931348623157e+308, max_exp=1024, max_10_exp=308, min=2.2250738585072014e-308, min_exp=-1021, min_10_exp=-307, dig=15, mant_dig=53, epsilon=2.220446049250313e-16, radix=2, rounds=1)
print(type(sys.float_info))
sys.float_info
print(sys.float_info.max)
print(1.8e+308)
print(type(1.8e+308))
print(sys.float_info.max.hex())
print(-sys.float_info.max)
-1.7976931348623157e+308
print(-1.8e+308)
-inf
print(type(-1.8e+308))
print(sys.float_info.min)
print(sys.float_info.min.hex())
print(float.fromhex('0x0.0000000000001p-1022'))
print(format(float.fromhex('0x0.0000000000001p-1022'), '.17'))
print(1e-323)
print(1e-324)
import sys
print(sys.maxsize)
print(type(sys.maxsize))
print(sys.maxsize == 2**63 - 1)
print(bin(sys.maxsize))
print(hex(sys.maxsize))
i = 10**100
print(i)
print(i > sys.maxsize)
print(sys.float_info.max)
i_e309 = 10**309
print(type(i_e309))
print(i_e309 > sys.float_info.max)
print(float('inf'))
print(float('inf') > sys.float_info.max)
print(float('inf') > i_e309)
int(float('inf'))
OverflowError: cannot
import sys
import resource
print(sys.getrecursionlimit())
def recu_test(n):
if n == 1:
print('Finish')
return
recu_test(n - 1)
recu_test(950)
recu_test(1500)
RecursionError: maximum
recu_test(995)
RecursionError: maximum
sys.setrecursionlimit(2000)
print(sys.getrecursionlimit())
recu_test(1500)
sys.setrecursionlimit(4)
print(sys.getrecursionlimit())
sys.setrecursionlimit(3)
RecursionError: cannot
sys.setrecursionlimit(10 ** 9)
print(sys.getrecursionlimit())
sys.setrecursionlimit(10 ** 10)
OverflowError: signed
recu_test(10 ** 4)
recu_test(10 ** 5)
print(resource.getrlimit(resource.RLIMIT_STACK))
-1
resource.setrlimit(resource.RLIMIT_STACK, (-1, -1))
print(resource.getrlimit(resource.RLIMIT_STACK))
-1
-1
recu_test(10 ** 5)
import sys
print(sys.version)
print(type(sys.version))
print(sys.version_info)
sys.version_info(major=3, minor=7, micro=0, releaselevel='final', serial=0)
print(type(sys.version_info))
sys.version_info
print(sys.version_info[0])
print(sys.version_info.major)
if sys.version_info.major == 3:
print('Python3')
print('Python2')
import tensorflow as tf
const1 = tf.constant(5)
const2 = tf.constant(10)
add_op = tf.add(const1, const2)
mul_op = tf.multiply(add_op, const2)
print(add_op)
print(mul_op)
dtype=int32
dtype=int32
with tf.Session() as sess:
add_result, mul_result = sess.run([add_op, mul_op])
print(add_result)
print(mul_result)
add_op_2 = const1 + const2
mul_op_2 = add_op_2 * const2
with tf.Session() as sess:
add_op_2_result, mul_op_2_result = sess.run([add_op_2, mul_op_2])
print(add_op_2_result)
print(mul_op_2_result)
import tensorflow as tf
const1 = tf.constant(5)
const2 = tf.constant(10)
print(const1)
print(const2)
dtype=int32
dtype=int32
with tf.Session() as sess:
const1_result, const2_result = sess.run([const1, const2])
print(const1_result)
print(const2_result)
import tensorflow as tf
import numpy as np
from sklearn import datasets, model_selection, utils
def conv2d(x, W):
return tf.nn.conv2d(x, W, strides=[1, 1, 1, 1], padding='SAME')
def max_pool_2x2(x):
return tf.nn.max_pool(x, ksize=[1, 2, 2, 1], strides=[1, 2, 2, 1], padding='SAME')
def weight_variable(shape):
initial = tf.truncated_normal(shape, stddev=0.1)
return tf.Variable(initial)
def bias_variable(shape):
initial = tf.constant(0.1, shape=shape)
return tf.Variable(initial)
width = 28
height = 28
n_in = width * height
n_out = 10
x = tf.placeholder(tf.float32, [None, n_in])
y_ = tf.placeholder(tf.float32, [None, n_out])
tf.name_scope('reshape')
x_image = tf.reshape(x, [-1, height, width, 1])
tf.name_scope('conv1')
W_conv1 = weight_variable([5, 5, 1, 32])
b_conv1 = bias_variable([32])
h_conv1 = tf.nn.relu(conv2d(x_image, W_conv1) + b_conv1)
tf.name_scope('pool1')
h_pool1 = max_pool_2x2(h_conv1)
tf.name_scope('conv2')
W_conv2 = weight_variable([5, 5, 32, 64])
b_conv2 = bias_variable([64])
h_conv2 = tf.nn.relu(conv2d(h_pool1, W_conv2) + b_conv2)
tf.name_scope('pool2')
h_pool2 = max_pool_2x2(h_conv2)
tf.name_scope('fc1')
W_fc1 = weight_variable([7 * 7 * 64, 1024])
b_fc1 = bias_variable([1024])
h_pool2_flat = tf.reshape(h_pool2, [-1, 7*7*64])
h_fc1 = tf.nn.relu(tf.matmul(h_pool2_flat, W_fc1) + b_fc1)
tf.name_scope('dropout')
keep_prob = tf.placeholder(tf.float32)
h_fc1_drop = tf.nn.dropout(h_fc1, keep_prob)
tf.name_scope('fc2')
W_fc2 = weight_variable([1024, 10])
b_fc2 = bias_variable([10])
y_conv = tf.matmul(h_fc1_drop, W_fc2) + b_fc2
tf.name_scope('loss')
cross_entropy = tf.nn.softmax_cross_entropy_with_logits
labels=y_
logits=y_conv
cross_entropy = tf.reduce_mean(cross_entropy)
tf.name_scope('training')
train_step = tf.train.AdamOptimizer(1e-4).minimize(cross_entropy)
tf.name_scope('accuracy')
correct_prediction = tf.equal(tf.argmax(y_conv, 1), tf.argmax(y_, 1))
correct_prediction = tf.cast(correct_prediction, tf.float32)
accuracy = tf.reduce_mean(correct_prediction)
mnist = datasets.fetch_mldata('MNIST original', data_home='data/src/download')
X = mnist.data / 255
y = mnist.target
Y = np.identity(10)[y.astype(int)]
train_size = 60000
test_size = 10000
X_train, X_test, Y_train, Y_test = model_selection.train_test_split
X, Y, test_size=test_size, train_size=train_size
batch_size = 100
batch_num = (int)(train_size // batch_size)
epochs = 10
with tf.Session() as sess:
sess.run(tf.global_variables_initializer())
for epoch in range(epochs):
X_, Y_ = utils.shuffle(X_train, Y_train)
for i in range(batch_num):
batch_X = X_train[i * batch_size: (i+1) * batch_size]
batch_Y = Y_train[i * batch_size: (i+1) * batch_size]
sess.run(train_step, feed_dict={x: batch_X, y_: batch_Y, keep_prob: 0.5})
if i % 10 == 0:
loss = sess.run(cross_entropy, feed_dict={x: X_test, y_: Y_test, keep_prob: 1})
acc = sess.run(accuracy, feed_dict={x: X_test, y_: Y_test, keep_prob: 1})
print('step: {}, loss: {:.5f}, train_accuracy: {:.5f}'.format(i, loss, acc))
loss = sess.run(cross_entropy, feed_dict={x: X_test, y_: Y_test, keep_prob: 1})
acc = sess.run(accuracy, feed_dict={x: X_test, y_: Y_test, keep_prob: 1})
print('epoch: {}, loss: {:.5f}, train_accuracy: {:.5f}'.format(epoch, loss, acc))
acc = sess.run(accuracy, feed_dict={x: X_test, y_: Y_test, keep_prob: 1})
print('test_accuracy: {:.5f}'.format(acc))
import tensorflow as tf
import numpy as np
from sklearn import datasets, model_selection, utils
mnist = datasets.fetch_mldata('MNIST original', data_home='data/src/download')
X = mnist.data / 255
y = mnist.target
Y = np.identity(10)[y.astype(int)]
train_size = 60000
test_size = 10000
X_train, X_test, Y_train, Y_test = model_selection.train_test_split
X, Y, test_size=test_size, train_size=train_size
n_in = 784
n_out = 10
x = tf.placeholder(tf.float32, [None, n_in])
y_ = tf.placeholder(tf.float32, [None, n_out])
W = tf.Variable(tf.zeros([n_in, n_out]))
b = tf.Variable(tf.zeros([n_out]))
y = tf.matmul(x, W) + b
cross_entropy = tf.reduce_mean
tf.nn.softmax_cross_entropy_with_logits(labels=y_, logits=y)
train_step = tf.train.GradientDescentOptimizer(0.5).minimize(cross_entropy)
correct_prediction = tf.equal(tf.argmax(y, 1), tf.argmax(y_, 1))
accuracy = tf.reduce_mean(tf.cast(correct_prediction, tf.float32))
batch_size = 100
batch_num = (int)(train_size // batch_size)
epochs = 20
with tf.Session() as sess:
sess.run(tf.global_variables_initializer())
for epoch in range(epochs):
X_, Y_ = utils.shuffle(X_train, Y_train)
for i in range(batch_num):
batch_X = X_train[i * batch_size: (i+1) * batch_size]
batch_Y = Y_train[i * batch_size: (i+1) * batch_size]
sess.run(train_step, feed_dict={x: batch_X, y_: batch_Y})
loss, acc = sess.run([cross_entropy, accuracy], feed_dict={x: X_test, y_: Y_test})
print('epoch: {:2}, loss: {:.5f}, accuracy: {:.5f}'.format(epoch, loss, acc))
acc = sess.run(accuracy, feed_dict={x: X_test, y_: Y_test})
print('Final accuracy: {:.5f}'.format(acc))
import tensorflow as tf
const = tf.constant(1)
holder = tf.placeholder(tf.int32)
add_op = const + holder
with tf.Session() as sess:
result = sess.run(add_op, feed_dict={holder: 5})
print(result)
result = sess.run(add_op, feed_dict={holder: 10})
print(result)
holder1 = tf.placeholder(tf.int32)
holder2 = tf.placeholder(tf.int32, [3])
mul_op = holder1 * holder2
with tf.Session() as sess:
result = sess.run(mul_op, feed_dict={holder1: 2, holder2: [0, 1, 2]})
print(result)
result = sess.run(mul_op, feed_dict={holder1: 5, holder2: [0, 10, 20]})
print(result)
holder1 = tf.placeholder(tf.int32)
holder2 = tf.placeholder(tf.int32, [None])
mul_op = holder1 * holder2
with tf.Session() as sess:
result = sess.run(mul_op, feed_dict={holder1: 2, holder2: [0, 1]})
print(result)
result = sess.run(mul_op, feed_dict={holder1: 5, holder2: [0, 1, 2, 3, 4]})
print(result)
import tensorflow as tf
var = tf.Variable(10)
const = tf.constant(5)
calc_op = var * const
assign_op = tf.assign(var, calc_op)
with tf.Session() as sess:
sess.run(tf.global_variables_initializer())
print(sess.run(var))
sess.run(assign_op)
print(sess.run(var))
sess.run(assign_op)
print(sess.run(var))
with tf.Session() as sess:
sess.run(tf.global_variables_initializer())
print(sess.run(var))
import test_module
print('This is test_main.py')
print('test_module.__name__ is', test_module.__name__)
print('---')
print('call test_module.func()')
test_module.func()
def func():
print('    This is func() in test_module.py')
print('    __name__ is', __name__)
if __name__ == '__main__':
print("Start if __name__ == '__main__'")
print('call func()')
func()
import textwrap
s = "Python can be easy to pick up whether you're a first time programmer or you're experienced with other languages"
s_wrap_list = textwrap.wrap(s, 40)
print(s_wrap_list)
programmer or you
print('\n'.join(s_wrap_list))
programmer or you
print(textwrap.fill(s, 40))
programmer or you
print(textwrap.wrap(s, 40, max_lines=2))
print(textwrap.fill(s, 40, max_lines=2))
print(textwrap.fill(s, 40, max_lines=2, placeholder=' ~'))
print(textwrap.fill(s, 40, max_lines=2, placeholder=' ~', initial_indent='  '))
s = 'あいうえお、かきくけこ、12345,67890, さしすせそ、abcde'
print(textwrap.fill(s, 12))
s = 'Python is powerful'
print(textwrap.shorten(s, 12))
print(textwrap.shorten(s, 12, placeholder=' ~'))
s = 'Pythonについて。Pythonは汎用のプログラミング言語である。'
print(textwrap.shorten(s, 20))
s_short = s[:12] + '...'
print(s_short)
wrapper = textwrap.TextWrapper(width=30, max_lines=3, placeholder=' ~', initial_indent='  ')
s = "Python can be easy to pick up whether you're a first time programmer or you're experienced with other languages"
print(wrapper.wrap(s))
programmer or you
print(wrapper.fill(s))
programmer or you
import numpy as np
a = np.array([0, 1, 2])
print(a)
print(type(a))
numpy.ndarray
import time
ut = time.time()
print(ut)
print(type(ut))
start = time.time()
time.sleep(3)
t = time.time() - start
print(t)
import timeit
def test(n):
return sum(range(n))
n = 10000
loop = 1000
result = timeit.timeit('test(n)', globals=globals(), number=loop)
print(result / loop)
result = timeit.timeit(lambda: test(n), number=loop)
print(result / loop)
print(timeit.timeit(lambda: test(n), number=1))
print(timeit.timeit(lambda: test(n), number=10))
print(timeit.timeit(lambda: test(n), number=100))
repeat = 5
print(timeit.repeat(lambda: test(n), repeat=repeat, number=100))
test(n)
std. dev. of
test(n)
std. dev. of
import numpy as np
a = np.arange(n)
np.sum(a)
std. dev. of
t = (0, 1, 2)
a, b, c = t
print(a)
print(b)
print(c)
l = [0, 1, 2]
a, b, c = l
print(a)
print(b)
print(c)
a, b = 0, 1
print(a)
print(b)
a, b = t
ValueError: too
a, b, c, d = t
ValueError: not
t = (0, 1, (2, 3, 4))
a, b, c = t
print(a)
print(b)
print(c)
print(type(c))
print(a)
print(b)
print(c)
print(d)
print(e)
t = (0, 1, 2)
a, b, _ = t
print(a)
print(b)
print(_)
t = (0, 1, 2, 3, 4)
c = t
print(a)
print(b)
print(c)
print(type(c))
b, c = t
print(a)
print(b)
print(c)
a, b, c = t
print(a)
print(b)
print(c)
_ = t
print(a)
print(b)
print(_)
a, b = t[0], t[1]
print(a)
print(b)
c = t
SyntaxError: two
t = (0, 1, 2)
c = t
print(a)
print(b)
print(c)
print(type(c))
d = t
print(a)
print(b)
print(c)
print(d)
t = (0, 1, 2)
print(t)
print(type(t))
print(t[0])
print(t[:2])
t[0] = 100
not support 
t.append(100)
t_add = t + (3, 4, 5)
print(t_add)
print(t)
print(t + [3, 4, 5])
TypeError: can
tuple (not "list") 
print(t + tuple([3, 4, 5]))
print(t + tuple(3))
print(t + (3,))
l = list(t)
print(l)
print(type(l))
l.insert(2, 100)
print(l)
t_insert = tuple(l)
print(t_insert)
print(type(t_insert))
l = list(t)
l[1] = 100
t_change = tuple(l)
print(t_change)
l = list(t)
l.remove(1)
t_remove = tuple(l)
print(t_remove)
single_tuple_error = (0)
print(single_tuple_error)
print(type(single_tuple_error))
single_tuple = (0, )
print(single_tuple)
print(type(single_tuple))
print((0, 1, 2) + (3))
TypeError: can
tuple (not "int") 
print((0, 1, 2) + (3, ))
t = 0
print(t)
print(type(t))
t_ = 0
print(t_)
print(type(t_))
print(empty_tuple)
print(type(empty_tuple))
SyntaxError: invalid
SyntaxError: invalid
SyntaxError: invalid
empty_tuple = tuple()
print(empty_tuple)
print(type(empty_tuple))
def example(a, b):
print(a, type(a))
print(b, type(b))
example(0, 1)
example((0, 1))
TypeError: example
example((0, 1), 2)
example(*(0, 1))
print(type('string'))
print(type(100))
print(type([0, 1, 2]))
print(type(type('string')))
print(type(str))
print(type('string') is str)
print(type('string') is int)
def is_str(v):
return type(v) is str
print(is_str('string'))
print(is_str(100))
print(is_str([0, 1, 2]))
def is_str_or_int(v):
return type(v) in (str, int)
print(is_str_or_int('string'))
print(is_str_or_int(100))
print(is_str_or_int([0, 1, 2]))
def type_condition(v):
if type(v) is str:
print('type is str')
elif type(v) is int:
print('type is int')
print('type is not str or int')
type_condition('string')
type_condition(100)
type_condition([0, 1, 2])
type is not str or int
from typing import Union, List
def func_u(x: List[Union[int, float]]) -> float:
return sum(x) ** 0.5
print(func_u([0.5, 9.5, 90]))
s = 'あいうえお'
b = s.encode('unicode-escape')
print(b)
print(type(b))
s_from_b = b.decode('unicode-escape')
print(s_from_b)
print(type(s_from_b))
s_from_b_error = b.decode('utf-8')
print(s_from_b_error)
print(type(s_from_b_error))
s_from_s = s_from_b_error.encode().decode('unicode-escape')
print(s_from_s)
print(type(s_from_s))
import codecs
s_from_s_codecs = codecs.decode(s_from_b_error, 'unicode-escape')
print(s_from_s_codecs)
print(type(s_from_s_codecs))
s_ascii = ascii('あ')
print(s_ascii)
print(type(s_ascii))
print(s_ascii[0])
print(s_ascii[-1])
print(len(s_ascii))
print(ascii('あ') == "'\\u3042'")
s_unicode_escape = ascii('あ')[1:-1]
print(s_unicode_escape)
print(type(s_unicode_escape))
print(s_unicode_escape == '\\u3042')
print('\u3042')
print(len('\u3042'))
print('\u3042' == 'あ')
print('\\u3042')
with open('data/src/unicode_escape.txt') as f:
s = f.read()
print(s)
print(type(s))
print(len(s))
with open('data/src/unicode_escape.txt', encoding='unicode-escape') as f:
s = f.read()
print(s)
print(type(s))
print(len(s))
b_json = b
print(b_json)
print(b_json.decode())
print(b_json.decode('unicode-escape'))
import json
print(json.loads(b_json.decode()))
print(type(json.loads(b_json.decode())))
print(json.loads(b_json))
import unicodedata
print(unicodedata.east_asian_width('あ'))  
print(type(unicodedata.east_asian_width('あ')))
print(unicodedata.east_asian_width('a'))  
print(unicodedata.east_asian_width('Ａ'))  
print(unicodedata.east_asian_width('ｱ'))  
print(unicodedata.east_asian_width('Å'))  
def get_east_asian_width_count(text):
count = 0
for c in text:
if unicodedata.east_asian_width(c) in 'FWA':
count += 2
count += 1
return count
print(get_east_asian_width_count('あいうえお'))
print(get_east_asian_width_count('abcde'))
print(get_east_asian_width_count('ｱｲｳｴｵ'))
print(get_east_asian_width_count('ａｂｃｄｅ'))
print(get_east_asian_width_count('あｱaａ'))  
www.unicode.org
import unicodedata
s = 'あ'
print(unicodedata.name(s))
name = 'grinning face'
value = unicodedata.lookup(name)
print(value)
print(type(value))
u8 = value.encode('utf-8')
print(u8)
print(type(u8))
ue = value.encode('unicode-escape')
print(ue)
print(type(ue))
emoji = '\U0001f600'
print(emoji)
emoji_2 = '\U0001f601'
print(emoji_2)
print(unicodedata.name(emoji_2))
u8_2 = emoji_2.encode('utf-8')
print(u8_2)
print(type(u8_2))
from collections import defaultdict
UnionFind()
def __init__(self, n):
self.n = n
self.parents = [-1] * n
def find(self, x):
if self.parents[x] < 0:
return x
self.parents[x] = self.find(self.parents[x])
return self.parents[x]
def union(self, x, y):
x = self.find(x)
y = self.find(y)
if x == y:
return
if self.parents[x] > self.parents[y]:
x, y = y, x
self.parents[x] += self.parents[y]
self.parents[y] = x
def size(self, x):
return -self.parents[self.find(x)]
def same(self, x, y):
return self.find(x) == self.find(y)
def members(self, x):
root = self.find(x)
return [i for i in range(self.n) if self.find(i) == root]
def roots(self):
return [i for i, x in enumerate(self.parents) if x < 0]
def group_count(self):
return len(self.roots())
def all_group_members(self):
group_members = defaultdict(list)
for member in range(self.n):
group_members[self.find(member)].append(member)
return group_members
def __str__(self):
return '\n'.join(f'{r}: {m}' for r, m in self.all_group_members().items())
class UnionFindLabel(UnionFind
def __init__(self, labels):
assert len(labels) == len(set(labels))
self.n = len(labels)
self.parents = [-1] * self.n
self.d = {x: i for i, x in enumerate(labels)}
self.d_inv = {i: x for i, x in enumerate(labels)}
def find_label(self, x):
return self.d_inv[super().find(self.d[x])]
def union(self, x, y):
super().union(self.d[x], self.d[y])
def size(self, x):
return super().size(self.d[x])
def same(self, x, y):
return super().same(self.d[x], self.d[y])
def members(self, x):
root = self.find(self.d[x])
return [self.d_inv[i] for i in range(self.n) if self.find(i) == root]
def roots(self):
return [self.d_inv[i] for i, x in enumerate(self.parents) if x < 0]
def all_group_members(self):
group_members = defaultdict(list)
for member in range(self.n):
group_members[self.d_inv[self.find(member)]].append(self.d_inv[member])
return group_members
UnionFindBasic()
def __init__(self, n):
self.parents = list(range(n))
def find(self, x):
if self.parents[x] == x:
return x
return self.find(self.parents[x])
def union(self, x, y):
x = self.find(x)
y = self.find(y)
if x == y:
return
self.parents[y] = x
UnionFindPathCompression()
def __init__(self, n):
self.parents = list(range(n))
def find(self, x):
if self.parents[x] == x:
return x
self.parents[x] = self.find(self.parents[x])
return self.parents[x]
def union(self, x, y):
x = self.find(x)
y = self.find(y)
if x == y:
return
self.parents[y] = x
UnionFindByRank()
def __init__(self, n):
self.parents = list(range(n))
self.rank = [0] * n
def find(self, x):
if self.parents[x] == x:
return x
self.parents[x] = self.find(self.parents[x])
return self.parents[x]
def union(self, x, y):
x = self.find(x)
y = self.find(y)
if x == y:
return
if self.rank[x] < self.rank[y]:
self.parents[x] = y
self.parents[y] = x
if self.rank[x] == self.rank[y]:
self.rank[x] += 1
UnionFindBySize()
def __init__(self, n):
self.parents = list(range(n))
self.size = [1] * n
def find(self, x):
if self.parents[x] == x:
return x
self.parents[x] = self.find(self.parents[x])
return self.parents[x]
def union(self, x, y):
x = self.find(x)
y = self.find(y)
if x == y:
return
if self.size[x] < self.size[y]:
self.size[y] += self.size[x]
self.parents[x] = y
self.size[x] += self.size[y]
self.parents[y] = x
if self.size[x] < self.size[y]:
x, y = y, x
self.size[x] += self.size[y]
self.parents[y] = x
UnionFind()
def __init__(self, n):
self.parents = [-1] * n
def find(self, x):
if self.parents[x] < 0:
return x
self.parents[x] = self.find(self.parents[x])
return self.parents[x]
def union(self, x, y):
x = self.find(x)
y = self.find(y)
if x == y:
return
if self.parents[x] > self.parents[y]:
x, y = y, x
self.parents[x] += self.parents[y]
self.parents[y] = x
from union_find_basic import UnionFindBasic, UnionFindPathCompression, UnionFindByRank, UnionFindBySize, UnionFind
ufb = UnionFindBasic(5)
print(ufb.parents)
ufb.union(3, 4)
print(ufb.parents)
ufb.union(2, 3)
print(ufb.parents)
ufb.union(1, 2)
print(ufb.parents)
ufb.union(0, 4)
print(ufb.parents)
print([ufb.find(i) for i in range(5)])
ufpc = UnionFindPathCompression(5)
print(ufpc.parents)
ufpc.union(3, 4)
print(ufpc.parents)
ufpc.union(2, 3)
print(ufpc.parents)
ufpc.union(1, 2)
print(ufpc.parents)
ufpc.union(0, 4)
print(ufpc.parents)
print([ufpc.find(i) for i in range(5)])
ufbr = UnionFindByRank(5)
print(ufbr.parents)
ufbr.union(3, 4)
print(ufbr.parents)
ufbr.union(2, 3)
print(ufbr.parents)
ufbr.union(1, 2)
print(ufbr.parents)
ufbr.union(0, 4)
print(ufbr.parents)
ufbs = UnionFindBySize(5)
print(ufbs.parents)
ufbs.union(3, 4)
print(ufbs.parents)
ufbs.union(2, 3)
print(ufbs.parents)
ufbs.union(1, 2)
print(ufbs.parents)
ufbs.union(0, 4)
print(ufbs.parents)
print(ufbs.size)
print(ufbs.size[ufbs.find(0)])
uf = UnionFind(5)
print(uf.parents)
-1
-1
-1
-1
-1
uf.union(3, 4)
print(uf.parents)
uf.union(2, 3)
print(uf.parents)
uf.union(1, 2)
print(uf.parents)
uf.union(0, 4)
print(uf.parents)
-1
-1
-1
-2
-1
-1
-3
-1
-4
-5
from union_find import UnionFind, UnionFindLabel
uf = UnionFind(6)
print(uf.parents)
-1
-1
-1
-1
-1
-1
print(uf)
uf.union(0, 2)
print(uf.parents)
-2
-1
-1
-1
-1
print(uf)
uf.union(1, 3)
print(uf.parents)
uf.union(4, 5)
print(uf.parents)
uf.union(1, 4)
print(uf.parents)
-2
-2
-1
-1
-2
-2
-2
-2
-4
print(uf)
print(uf.parents)
-2
-4
print(uf.find(0))
print(uf.find(5))
print(uf.size(0))
print(uf.size(5))
print(uf.same(0, 2))
print(uf.same(0, 5))
print(uf.members(0))
print(uf.members(5))
print(uf.roots())
print(uf.group_count())
print(uf.all_group_members())
print(list(uf.all_group_members().values()))
l = ['A', 'B', 'C', 'D', 'E']
d = {x: i for i, x in enumerate(l)}
print(d)
d_inv = {i: x for i, x in enumerate(l)}
print(d_inv)
uf_s = UnionFind(len(l))
print(uf_s)
uf_s.union(d['A'], d['D'])
uf_s.union(d['D'], d['C'])
uf_s.union(d['E'], d['B'])
print(uf_s)
print(d_inv[uf_s.find(d['D'])])
print(uf_s.size(d['D']))
print(uf_s.same(d['A'], d['D']))
print([d_inv[i] for i in uf_s.members(d['D'])])
print([d_inv[i] for i in uf_s.roots()])
print(uf_s.group_count())
l = ['A', 'B', 'C', 'D', 'E']
ufl = UnionFindLabel(l)
print(ufl)
ufl.union('A', 'D')
ufl.union('D', 'C')
ufl.union('E', 'B')
print(ufl)
print(ufl.find_label('D'))
print(ufl.size('D'))
print(ufl.same('A', 'D'))
print(ufl.members('D'))
print(ufl.roots())
print(ufl.group_count())
print(ufl.all_group_members())
ufl_n = UnionFindLabel([1, 2, 3, 4, 5])
print(ufl_n)
ufl_n.union(1, 4)
ufl_n.union(4, 3)
ufl_n.union(5, 2)
print(ufl_n)
ufl_n2 = UnionFind(6)
print(ufl_n2)
ufl_n2.union(1, 4)
ufl_n2.union(4, 3)
ufl_n2.union(5, 2)
print(ufl_n2)
ufl_t = UnionFindLabel([(0, 0), (0, 1), (1, 0), (1, 1)])
print(ufl_t)
ufl_t.union((0, 1), (1, 0))
ufl_t.union((0, 0), (1, 0))
print(ufl_t)
print(ufl_t.same((0, 1), (0, 0)))
import urllib.parse
values = {'key1': 'value1', 'key2': 'バリュー2'}
print(urllib.parse.urlencode(values))
key1=value1&key2
base = 'http://example.com/sub1/index.html'
print(urllib.parse.urljoin(base, 'index2.html'))
print(urllib.parse.urljoin(base, '../../sub2/index.html'))
print(urllib.parse.urljoin(base, 'https://google.com'))
print(urllib.parse.urljoin(base, '//google.com'))
google.com
google.com
url = 'http://example.com/sub1/index.html?key=value'
o = urllib.parse.urlparse(url)
print(o)
print(o.scheme, o.netloc, o.path)
ParseResult(scheme='http', netloc='example.com', path='/sub1/index.html', params='', query='key=value', fragment='')
relative_url = 'sub/index.html'
o = urllib.parse.urlparse(relative_url)
print(o)
ParseResult(scheme='', netloc='', path='sub/index.html', params='', query='', fragment='')
import urllib.parse
url = 'https://www.google.co.jp/search?q=%E6%A1%9C&tbm=isch'
print(urllib.parse.urlparse(url))
ParseResult(scheme='https', netloc='www.google.co.jp', path='/search', params='', query='q=%E6%A1%9C&tbm=isch', fragment='')
qs = urllib.parse.urlparse(url).query
print(qs)
print(type(qs))
qs_d = urllib.parse.parse_qs(qs)
print(qs_d)
print(type(qs_d))
print(qs_d['q'])
print(type(qs_d['q']))
print(qs_d['q'][0])
print(type(qs_d['q'][0]))
qs_l = urllib.parse.parse_qsl(qs)
print(qs_l)
print(type(qs_l))
print(qs_l[0])
print(type(qs_l[0]))
print(qs_l[0][1])
print(type(qs_l[0][1]))
d = {'key1': 'value / one', 'key2': 'バリュー2'}
d_qs = urllib.parse.urlencode(d)
print(d_qs)
key1=value
print(type(d_qs))
l = [('key1', 'value / one'), ('key2', 'バリュー2')]
l_qs = urllib.parse.urlencode(l)
print(l_qs)
key1=value
print(type(l_qs))
print(urllib.parse.urlencode(d))
key1=value
print(urllib.parse.urlencode(d, quote_via=urllib.parse.quote))
key1=value%20%2
print(urllib.parse.urlencode(d, safe='/'))
key1=value
print(urllib.parse.urlencode(d, safe='/', quote_via=urllib.parse.quote))
key1=value%20
print(qs_d)
print(urllib.parse.urlencode(qs_d))
print(urllib.parse.urlencode(qs_d, doseq=True))
print(url)
print(url.replace('isch', 'vid'))
def update_query(url, key, org_val, new_val):
pr = urllib.parse.urlparse(url)
d = urllib.parse.parse_qs(pr.query)
l = d.get(key)
if l:
d[key] = [new_val if v == org_val else v for v in l]
d[key] = new_val
return urllib.parse.urlunparse(pr._replace(query=urllib.parse.urlencode(d, doseq=True)))
print(update_query(url, 'tbm', 'isch', 'vid'))
print(update_query(url, 'q', '桜', '梅'))
print(update_query(url, 'new-key', 'xxx', 'yyy'))
def remove_query(url, key):
pr = urllib.parse.urlparse(url)
d = urllib.parse.parse_qs(pr.query)
d.pop(key, None)
return urllib.parse.urlunparse(pr._replace(query=urllib.parse.urlencode(d, doseq=True)))
print(remove_query(url, 'tbm'))
print(remove_query(url, 'new-key'))
def remove_all_query(url):
return urllib.parse.urlunparse(urllib.parse.urlparse(url)._replace(query=None))
print(remove_all_query(url))
import urllib.parse
s = '日本語'
s_quote = urllib.parse.quote(s)
print(s_quote)
print(type(s_quote))
b = s.encode()
print(b)
print(type(b))
print(urllib.parse.quote(b))
s_quote_sj = urllib.parse.quote(s, encoding='shift-jis')
print(s_quote_sj)
b_sj_quote = urllib.parse.quote(s.encode('shift-jis'))
print(b_sj_quote)
print(s_quote_sj == b_sj_quote)
print(urllib.parse.quote('http://x-y_z.com'))
print(urllib.parse.quote('http://x-y_z.com', safe=''))
print(urllib.parse.quote('http://x-y_z.com', safe='/:'))
print(urllib.parse.quote('+ /'))
print(urllib.parse.quote_plus('+ /'))
print(urllib.parse.quote_plus('+ /', safe='+/'))
page_title = '日本語'
base_ja = 'https://ja.wikipedia.org/wiki/'
print(base_ja + urllib.parse.quote(page_title))
full_url = 'https://ja.wikipedia.org/wiki/日本語'
print(urllib.parse.quote(full_url, safe=':/'))
print(base_ja + urllib.parse.quote('OK コンピューター'))
print(base_ja + urllib.parse.quote('OK コンピューター'.replace(' ', '_')))
print(s_quote)
print(urllib.parse.unquote(s_quote))
print(s_quote_sj)
print(urllib.parse.unquote(s_quote_sj))
print(urllib.parse.unquote(s_quote_sj, 'shift-jis'))
print(urllib.parse.unquote('a+b'))
print(urllib.parse.unquote_plus('a+b'))
b_unquote = urllib.parse.unquote_to_bytes(s_quote)
print(b_unquote)
print(b_unquote.decode())
b_unquote_sj = urllib.parse.unquote_to_bytes(s_quote_sj)
print(b_unquote_sj)
print(b_unquote_sj.decode('shift-jis'))
import urllib.request
url = "https://upload.wikimedia.org/wikipedia/en/2/24/Lenna.png"
path = "data/src/lena_square.png"
result = urllib.request.urlretrieve(url, path)
print(result)
http.client.HTTPMessage
data = urllib.request.urlopen(url).read()
with open(path, mode="wb") as f:
f.write(data)
import os
import pprint
import time
import urllib.error
import urllib.request
def download_file(url, dst_path):
with urllib.request.urlopen(url) as web_file:
data = web_file.read()
with open(dst_path, mode='wb') as local_file:
local_file.write(data)
urllib.error.URLError
print(e)
def download_file(url, dst_path):
urllib.request.urlopen(url) 
open(dst_path, 'wb') 
local_file.write(web_file.read())
urllib.error.URLError
print(e)
url = 'https://www.python.org/static/img/python-logo.png'
dst_path = 'data/temp/py-logo.png'
download_file(url, dst_path)
def download_file_to_dir(url, dst_dir):
download_file(url, os.path.join(dst_dir, os.path.basename(url)))
dst_dir = 'data/temp'
download_file_to_dir(url, dst_dir)
url_error = 'https://www.python.org/static/img/python-logo_xxx.png'
download_file_to_dir(url_error, dst_dir)
url_zip = 'https://github.com/nkmk/python-snippets/raw/master/notebook/data/src/sample_header.csv.zip'
download_file_to_dir(url_zip, dst_dir)
url_xlsx = 'https://github.com/nkmk/python-snippets/raw/master/notebook/data/src/sample.xlsx'
download_file_to_dir(url_xlsx, dst_dir)
url_pdf = 'https://github.com/nkmk/python-snippets/raw/master/notebook/data/src/pdf/sample1.pdf'
download_file_to_dir(url_pdf, dst_dir)
url_list = ['https://example.com/basedir/base_{:03}.jpg'.format(i) for i in range(5)]
pprint.pprint(url_list)
download_dir = 'data/temp'
sleep_time_sec = 1
for url in url_list:
print(url)
download_file_dir(url, download_dir)
time.sleep(sleep_time_sec)
import os
import shutil
def glob_example_detail():
if os.path.exists('temp'):
shutil.rmtree('temp')
os.makedirs('temp/dir/', exist_ok=True)
os.makedirs('temp/dir/sub_dir1/', exist_ok=True)
os.makedirs('temp/dir/sub_dir2/', exist_ok=True)
open('temp/aaa.text', 'w')
pass
open('temp/1.txt', 'w')
pass
open('temp/12.text', 'w')
pass
open('temp/123.txt', 'w')
pass
open('temp/[x].txt', 'w')
pass
open('temp/dir/bbb.txt', 'w')
pass
open('temp/dir/987.text', 'w')
pass
open('temp/dir/sub_dir1/98.txt', 'w')
pass
open('temp/dir/sub_dir1/ccc.text', 'w')
pass
open('temp/dir/sub_dir2/ddd.text', 'w')
pass
def pathlib_basic():
if os.path.exists('temp'):
shutil.rmtree('temp')
os.makedirs('temp/dir/sub_dir/', exist_ok=True)
open('temp/file.txt', 'w')
pass
open('temp/dir/sub_dir/file2.txt', 'w')
pass
import sys
print(sys.version)
print(type(sys.version))
print(sys.version_info)
sys.version_info(major=3, minor=6, micro=2, releaselevel='final', serial=0)
print(type(sys.version_info))
sys.version_info
print(sys.version_info.major)
print(type(sys.version_info.major))
import platform
print(platform.python_version())
print(type(platform.python_version()))
print(platform.python_version_tuple())
print(type(platform.python_version_tuple()))
print(platform.python_version_tuple()[0])
print(type(platform.python_version_tuple()[0]))
import warnings
import pandas as pd
df = pd.DataFrame([[0, 1, 2], [3, 4, 5]])
df.ix[0, 0] = 0
iloc for positional indexing
df.iloc[:1][0] = 0
using .loc[row_indexer,col_indexer] = value
documentation: http
returning-a-view-versus-a-copy
warnings.simplefilter('ignore')
df.ix[0, 0] = 0
df.iloc[:1][0] = 0
warnings.resetwarnings()
warnings.simplefilter('ignore', FutureWarning)
df.ix[0, 0] = 0
df.iloc[:1][0] = 0
using .loc[row_indexer,col_indexer] = value
documentation: http
returning-a-view-versus-a-copy
warnings.resetwarnings()
warnings.simplefilter('ignore', SettingWithCopyWarning)
NameError: name
not defined
warnings.simplefilter('ignore', pd.core.common.SettingWithCopyWarning)
df.ix[0, 0] = 0
iloc for positional indexing
df.iloc[:1][0] = 0
warnings.resetwarnings()
warnings.simplefilter('error')
df.ix[0, 0] = 0
warnings.resetwarnings()
warnings.simplefilter('ignore', FutureWarning)
warnings.simplefilter('error', pd.core.common.SettingWithCopyWarning)
df.ix[0, 0] = 0
df.iloc[:1][0] = 0
warnings.resetwarnings()
warnings.catch_warnings()
warnings.simplefilter('ignore')
df.ix[0, 0] = 0
df.ix[0, 0] = 0
iloc for positional indexing
import time
while True:
time.sleep(1)
print('processing...')
except KeyboardInterrupt:
print('!!FINISH!!')
i = 0
while i < 3:
print(i)
i += 1
i = 0
while i < 3:
print(i)
if i == 1:
print('!!BREAK!!')
break
i += 1
i = 0
while i < 3:
if i == 1:
print('!!CONTINUE!!')
i += 1
continue
print(i)
i += 1
i = 0
while i < 3:
print(i)
i += 1
print('!!FINISH!!')
i = 0
while i < 3:
print(i)
if i == 1:
print('!!BREAK!!')
break
i += 1
print('!!FINISH!!')
i = 0
while i < 3:
if i == 1:
print('!!SKIP!!')
i += 1
continue
print(i)
i += 1
print('!!FINISH!!')
import time
start = time.time()
while True:
time.sleep(1)
print('processing...')
if time.time() - start > 5:
print('!!BREAK!!')
break
start = time.time()
while 1:
time.sleep(1)
print('processing...')
if time.time() - start > 5:
print('!!BREAK!!')
break
start = time.time()
while time.time() - start <= 5:
time.sleep(1)
print('processing...')
print('!!FINISH!!')
import xlrd
import pprint
wb = xlrd.open_workbook('data/src/sample.xlsx')
print(type(wb))
xlrd.book.Book
print(wb.sheet_names())
sheets = wb.sheets()
print(type(sheets))
print(type(sheets[0]))
xlrd.sheet.Sheet
sheet = wb.sheet_by_name('sheet1')
print(type(sheet))
xlrd.sheet.Sheet
cell = sheet.cell(1, 2)
print(cell)
print(type(cell))
xlrd.sheet.Cell
print(cell.value)
print(sheet.cell_value(1, 2))
col = sheet.col(1)
print(col)
print(type(col[0]))
xlrd.sheet.Cell
col_values = sheet.col_values(1)
print(col_values)
print(sheet.row_values(1))
pprint.pprint([sheet.row_values(x) for x in range(4)])
def get_list_2d(sheet, start_row, end_row, start_col, end_col):
return [sheet.row_values(row, start_col, end_col + 1) for row in range(start_row, end_row + 1)]
l_2d = get_list_2d(sheet, 2, 3, 1, 2)
print(l_2d)
print(sheet.nrows)
def get_list_2d_all(sheet):
return [sheet.row_values(row) for row in range(sheet.nrows)]
l_2d_all = get_list_2d_all(sheet)
pprint.pprint(l_2d_all)
print(l_2d_all[1][0])
import xlwt
wb = xlwt.Workbook()
print(type(wb))
xlwt.Workbook.Workbook
sheet = wb.add_sheet('sheet1')
print(type(sheet))
xlwt.Worksheet.Worksheet
sheet.write(0, 0, 'A')
sheet.write(0, 1, 'B')
sheet.write(1, 0, 10)
sheet.write(1, 1, 20)
sheet.write(0, 0, 'A')
Exception: Attempt
cell: sheetname='sheet1'
rowx=0
colx=0
wb.save('data/dst/xlwt_sample.xls')
sheet2 = wb.add_sheet('sheet2')
def write_list_1d(sheet, l, start_row, start_col):
for i, val in enumerate(l):
sheet.write(start_row, start_col + i, val)
def write_list_2d(sheet, l_2d, start_row, start_col):
for i, l in enumerate(l_2d):
write_list_1d(sheet, l, start_row + i, start_col)
l_2d = [['A', 'B', 'C'], [1, 2, 3]]
write_list_2d(sheet2, l_2d, 1, 2)
wb.save('data/dst/xlwt_sample.xls')
s = '1234'
s_zero = s.zfill(8)
print(s_zero)
print(type(s_zero))
print(s.zfill(3))
s = '-1234'
print(s.zfill(8))
-0001234
s = '+1234'
print(s.zfill(8))
+0001234
s = 'abcd'
print(s.zfill(8))
n = 1234
print(type(n))
print(n.zfill(8))
print(str(n).zfill(8))
s = '1234'
print(s.rjust(8, '0'))
print(s.ljust(8, '0'))
print(s.center(8, '0'))
s = '-1234'
print(s.rjust(8, '0'))
print(s.ljust(8, '0'))
-1234000
print(s.center(8, '0'))
s = '1234'
print(format(s, '0>8'))
print('Zero Padding: {:0>8}'.format(s))
s = '-1234'
print(format(s, '0>8'))
print('Zero Padding: {:0>8}'.format(s))
print(format(int(s), '08'))
-0001234
print('Zero Padding: {:0>8}'.format(int(s)))
n = 1234
print(format(n, '08'))
print('Zero Padding: {:08}'.format(n))
print(format(n, '08x'))
print('Zero Padding: {:08x}'.format(n))
n = -1234
print(format(n, '08'))
-0001234
print('Zero Padding: {:08}'.format(n))
-0001234
print(f'Zero Padding: {n:08}')
-0001234
n = 1234
print('Zero Padding: %08d' % n)
n = -1234
print('Zero Padding: %08d' % n)
-0001234
s = '1234'
print('Zero Padding: %08s' % s)
print('Zero Padding: %08d' % int(s))
names = ['Alice', 'Bob', 'Charlie']
ages = [24, 50, 18]
for name, age in zip(names, ages):
print(name, age)
points = [100, 85, 90]
for name, age, point in zip(names, ages, points):
print(name, age, point)
names = ['Alice', 'Bob', 'Charlie', 'Dave']
ages = [24, 50, 18]
for name, age in zip(names, ages):
print(name, age)
names = ['Alice', 'Bob', 'Charlie']
ages = (24, 50, 18)
z = zip(names, ages)
print(z)
print(type(z))
l = list(zip(names, ages))
print(l)
print(type(l))
print(type(l[0]))
from itertools import zip_longest
names = ['Alice', 'Bob', 'Charlie', 'Dave']
ages = [24, 50, 18]
for name, age in zip_longest(names, ages):
print(name, age)
for name, age in zip_longest(names, ages, fillvalue=20):
print(name, age)
points = [100, 85]
for name, age, point in zip_longest(names, ages, points, fillvalue=20):
print(name, age, point)
fill_name = 'XXX'
fill_age = 20
fill_point = 50
len_names = len(names)
len_ages = len(ages)
len_points = len(points)
max_len = max(len_names, len_ages, len_points)
names = names + [fill_name] * (max_len - len_names)
ages = ages + [fill_age] * (max_len - len_ages)
points = points + [fill_point] * (max_len - len_points)
print(names)
print(ages)
print(points)
for name, age, point in zip(names, ages, points):
print(name, age, point)
def my_zip_longest(iterables, fillvalues):
max_len = max(len(i) for i in iterables)
return zip(*[list(i) + [v] * (max_len - len(i)) for i, v in zip(iterables, fillvalues)])
for name, age, point in my_zip_longest((names, ages, points), ('XXX', 20, 50)):
print(name, age, point)
import zipfile
with zipfile.ZipFile('data/temp/new_comp.zip', 'w', compression=zipfile.ZIP_DEFLATED) as new_zip:
new_zip.write('data/temp/test1.txt', arcname='test1.txt')
new_zip.write('data/temp/test2.txt', arcname='zipdir/test2.txt')
new_zip.write('data/temp/test3.txt', arcname='zipdir/sub_dir/test3.txt')
with zipfile.ZipFile('data/temp/new_comp_single.zip', 'w') as new_zip:
new_zip.write('data/temp/test1.txt', arcname='test1.txt', compress_type=zipfile.ZIP_DEFLATED)
new_zip.write('data/temp/test2.txt', arcname='zipdir/test2.txt')
new_zip.write('data/temp/test3.txt', arcname='zipdir/sub_dir/test3.txt')
with zipfile.ZipFile('data/temp/new_comp.zip', 'a') as existing_zip:
existing_zip.write('data/temp/test4.txt', arcname='test4.txt')
with zipfile.ZipFile('data/temp/new_comp.zip') as existing_zip:
print(existing_zip.namelist())
test1.txt
test4.txt
with zipfile.ZipFile('data/temp/new_comp.zip') as existing_zip:
existing_zip.extractall('data/temp/ext')
with zipfile.ZipFile('data/temp/new_comp_with_pass.zip') as pass_zip:
pass
_zip.extractall('data/temp/ext_pass', pwd='password')
with zipfile.ZipFile('data/temp/new_comp.zip') as existing_zip:
existing_zip.extract('test1.txt', 'data/temp/ext2')
with zipfile.ZipFile('data/temp/new_comp_with_pass.zip') as pass_zip:
pass
_zip.extract('test1.txt', 'data/temp/ext_pass2', pwd='password')
